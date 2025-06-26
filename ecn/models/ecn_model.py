from odoo import fields, models, api, _
from odoo.exceptions import UserError, ValidationError
from datetime import datetime
import json
import base64
import io
from io import BytesIO
import logging
from openpyxl import Workbook
from odoo.modules.module import get_module_resource
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, PatternFill
from openpyxl.styles import Font, Border, Side
from PIL import Image as PILImage
from odoo.tools import html2plaintext


class EngineeringChangeNote(models.Model):
    _name = 'asd.ecn'
    _description = 'Engineering Change Note'
    _rec_name = 'ecn_number'
    _inherit = ['iatf.sign.off.members','mail.thread', 'mail.activity.mixin']

    # Basic information fields
    project = fields.Many2one('xf.doc.approval.document.package', string='Project')
    part_id = fields.Many2one('product.template', related='project.part_id', string='Part No.', required=True)
    part_name = fields.Char(string='Part Name', related='part_id.name')
    part_number = fields.Char(string='Part Number', related='part_id.default_code')
    drg_no = fields.Char('Drg. No.', related='part_id.drg_no')
    related_parts = fields.Many2many('product.template', string='Related Parts')
    partner_id = fields.Many2one('res.partner', string='Customer Name', related='project.partner_id', required=True)

    ecn_number = fields.Char(string='ECN/ECR Number',
                             default=lambda self: self.env['ir.sequence'].next_by_code('asd.ecn'),
                             readonly=True)
    ecn_date = fields.Date(string='ECN/ECR Date', default=fields.Date.today)
    start_date = fields.Date(string='Start Date')
    target_completion_date = fields.Date(string='Target Completion Date')
    end_date = fields.Date(string='Actual Completion Date', compute='_compute_end_date', inverse='_inverse_end_date',
                           store=True)
    completion_status = fields.Char(
        string='Completion Status',
        compute='_compute_completion_status',
        store=True,
        help='Shows whether the task was completed early, on time, or late.'
    )

    @api.depends('target_completion_date', 'end_date')
    def _compute_completion_status(self):
        for record in self:
            if record.target_completion_date and record.end_date:
                # Calculate the difference in days
                delta = record.end_date - record.target_completion_date
                days_diff = delta.days

                if days_diff < 0:
                    # Early completion
                    record.completion_status = f"{abs(days_diff)} days early"
                elif days_diff > 0:
                    # Late completion
                    record.completion_status = f"{days_diff} days late"
                else:
                    # On time
                    record.completion_status = "On time"
            else:
                record.completion_status = False
    # People related fields
    emp_id = fields.Many2one('hr.employee', string='Proposer Name')
    co_ord = fields.Many2one('hr.employee', string='Co-Ordinator')
    department = fields.Many2one('hr.department', related='emp_id.department_id', string='Department', store=True)

    # Team approval fields
    team_approval_ids = fields.One2many('ecn.team.approval', 'ecn_id', string='CFT Team Approvals')
    all_approved = fields.Boolean(compute='_compute_approval_status', store=True)
    any_rejected = fields.Boolean(compute='_compute_approval_status', store=True)
    any_review = fields.Boolean(compute='_compute_approval_status', store=True)
    user_is_team_member = fields.Boolean(compute='_compute_user_is_team_member')

    # Workflow state
    state = fields.Selection([
        ('draft', 'Draft'),
        ('submitted', 'Submitted'),
        ('under_review', 'Under Review'),
        ('feasibility_check', 'Feasibility Check'),
        ('cft_approval', 'CFT Approval'),
        ('implementation', 'Implementation Planning'),
        ('completed', 'Completed'),
        ('rejected', 'Rejected')
    ], string='Status', default='draft', tracking=True)

    # ECN origin and description
    change_person = fields.Many2one('hr.employee', string='Change')
    change_received_from = fields.Selection([('customer', 'Customer'),
                                             ('internal', 'Internal'),
                                             ('supplier', 'Supplier')], string='Change Received From')
    description = fields.Html('Change Description', placeholder="Description of the change")
    existing = fields.Text('Existing')
    proposed = fields.Text('Proposed')

    # Material disposition fields
    scraped = fields.Boolean('Scraped')
    quantity = fields.Integer('Quantity')
    userd_in = fields.Boolean('To be used as it is')
    specification = fields.Boolean('To be converted as per new specification')

    # Purpose of change fields
    specific_change = fields.Boolean('Design / Specification Change')
    process_change = fields.Boolean('Process Change')
    quality_imp = fields.Boolean('Improvement In Performance / Quality')
    cost_reduce = fields.Boolean('Cost Reduction')
    standardisation = fields.Boolean('Standardisation')
    sub_suppiler = fields.Boolean('Sub- Supplier Requirement')
    improve_machine = fields.Boolean('Improvement In Machining')
    customer_requirement = fields.Boolean('Customer Requirement')
    remark = fields.Text('Remarks')

    # Impact assessment fields
    first_impact = fields.Boolean('1.Can engineering performance specifications be met as written?')
    secound_impact = fields.Boolean('2.Is there adequate capacity to produce product?')
    third_impact = fields.Boolean("3.Can product be manufactured with Cpk's that meet requirements?")
    forth_impact = fields.Boolean('4.Can statistical process control required on product?')
    fifth_impact = fields.Boolean('5.Does the design allow the use of efficient material handling techniques?')
    sixth_impact = fields.Boolean('6.Costs for tooling?')
    seven_impact = fields.Boolean('7.Costs for capital equipment?')
    eight_impact = fields.Boolean('8.Alternative manufacturing methods?')

    # Feasibility fields
    feasible = fields.Boolean('Feasible')
    not_feasible = fields.Boolean('Not Feasible')

    # Implementation planning fields
    ppap_approval_required = fields.Boolean('PPAP Approval Required')
    date_of_ppap_submission = fields.Date('Date of PPAP Submission')
    date_of_approval_received = fields.Date('Date of Approval Submission')
    date_of_pilot_lot_submission = fields.Date('Date of Pilot Lot Submission')
    date_of_regular_submission = fields.Date('Date of Regular Production')

    completion_notes = fields.Text('Notes')
    

    # Related fields for changes and formats
    change_required_ids = fields.One2many(
        comodel_name='asd.ecn.line',
        inverse_name='change_required_id',
        string='Change Required'
    )
    new_format_ids = fields.One2many(
        comodel_name='ecn.new.format',
        inverse_name='ecn_id',
        string='New Format'
    )
    # 4M Type One2many relationship
    four_m_type_ids = fields.One2many(
        'ecn.four.m.type',
        'ecn_id',
        string='4M Type Impacts'
    )

    # Approval history tracking
    approval_history_ids = fields.One2many(
        comodel_name='ecn.approval.history',
        inverse_name='ecn_id',
        string='Approval History'
    )
    final_status = fields.Selection([
        ('approved', 'Approved'),
        ('rejected', 'Rejected')
    ], string='Final Status', compute='_compute_final_status', store=True)
    generate_xlsx_file = fields.Binary(string="Generate XLSX File", attachment=True)
    def generate_xlsx_report(self):
        output = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = 'ECN\u2044ECR Report'

        border = Border(top=Side(style='thin'),left=Side(style='thin'),right=Side(style='thin'),bottom=Side(style='thin'))
        align_center = Alignment(vertical='center', horizontal='center', wrapText=True)
        align_left = Alignment(vertical='center', horizontal='left')
        font_header = Font(name='Arial', size=12, bold=True)
        font_all = Font(name='Times New Roman', size=11, bold=False)
        # Company Logo
        if self.env.user.company_id.logo:
            max_width, max_height = 150, 60
            image_data = base64.b64decode(self.env.user.company_id.logo)
            image = PILImage.open(io.BytesIO(image_data))
            image.thumbnail((max_width, max_height), PILImage.LANCZOS)
            img_bytes = io.BytesIO()
            image.save(img_bytes, format='PNG')
            logo_image = Image(img_bytes)
            ws.add_image(logo_image, 'A1')
            ws['A1'].alignment = align_center
        
        # Data

        data = {
            'B1': 'Engineering change request(ECR)/Engineering change note(ECN)',
            'A2': 'CHANGE RECEIVED FROM CUSTOMER',
            'A3': 'ECN/ECR NO',
            'A4': 'Project',
            'A5': 'Part No',
            'A6' : 'Part Name',
            'A7': 'Part Number',
            'A8': 'Related Parts',
            'A9': 'Target Completion Date',
            'E3': 'ECN/ECR Date',
            'E4': 'Proposer Name',
            'E5': 'Department',
            'E6': 'Co-Ordinator',
            'E7': 'Customer Name',
            'E8': 'Actual Completion Date',
            'E9': 'Completion Status',
            'A10': 'Initial Information',
            'A11': 'ORIGIN',
            'A12': 'Change Recieved From',
            'E12': 'Change',
            'A13': 'Change Description',
            'A14': 'Existing',
            'E14': 'Proposed',
            'A15': 'Change Details and Risk Assessment Analysis',
            'A16': 'Existing Material Dipostion',
            'A17': 'Scraped',
            'A18': 'To be used it is as',
            'A19': 'To be converted as per new specification',  
            'A20': 'Quantity',
            'A21': 'Purpose of Change',
            'A22': 'Design/Specification Change',
            'A23': 'Process Change',
            'A24': 'Improvement In Performance/Quality',
            'A25': 'Cost Reduction',
            'E22': 'Standardisation',
            'E23': 'Sub-Supplier Requirement',
            'E24': 'Improvement In Machining',
            'E25': 'Customer Requirement',
            'A26': 'Risk Assessment Analysis',
            'A27': 'Type',
            'B27': 'Man',
            'C27': 'Machine',
            'D27': 'Material',
            'E27': 'Method',
            'F27': 'Measurment',
            'G27': 'Environment',
            'H27': 'Action Plan',
            'I27': 'Responsibility',
            'J27': 'Status',
        }

        for cell, value in data.items():
            ws[cell] = value
            ws[cell].font = font_header
            ws[cell].alignment = align_center
            ws[cell].border = border
        
        for row in ws.iter_rows(min_row=1, max_row=27, min_col=1, max_col=10):
            for cell in row:
                cell.border = border
                cell.alignment = align_center
        ws['B1'].fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        ws['A10'].fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        ws['A15'].fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        ws['A21'].fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        ws['A26'].fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        
       
        
        merge_ranges = [
            'B1:J1', 'A2:E2','F2:J2', 'A3:B3', 'A4:B4', 'A5:B5', 'A6:B6', 'A7:B7', 'A8:B8', 'A9:B9', 'C3:D3', 'C4:D4', 'C5:D5', 'C6:D6', 'C7:D7', 'C8:D8', 'C9:D9',
            'E3:F3', 'E4:F4', 'E5:F5', 'E6:F6', 'E7:F7', 'E8:F8', 'E9:F9', 
            'G3:J3', 'G4:J4', 'G5:J5', 'G6:J6', 'G7:J7', 'G8:J8', 'G9:J9', 'A10:J10', 'A11:J11', 'A12:B12',
            'E12:G12', 'C12:D12', 'H12:J12', 'A13:B13', 'C13:J13', 'E13:G13', 'H13:J13', 'A14:B14', 'C14:D14',
            'E14:F14', 'G14:J14', 'A15:J15', 'A16:J16', 'A17:B17', 'C17:J17', 'A18:B18', 'C18:J18', 'A19:B19', 'C19:J19',
            'A20:B20', 'C20:J20', 'A21:J21', 'A22:B22', 'C22:D22', 'E22:G22', 'H22:J22', 'A23:B23', 'C23:D23', 'E23:G23', 'H23:J23',
            'A24:B24', 'C24:D24', 'E24:G24', 'H24:J24', 'A25:B25', 'C25:D25', 'E25:G25', 'H25:J25', 'A26:J26'
        ]

        for merge_range in merge_ranges:
            ws.merge_cells(merge_range)
        
        ws.row_dimensions[1].height = 55
        ws.row_dimensions[10].height = 25
        ws.row_dimensions[15].height = 25
        ws.row_dimensions[15].height = 25
        ws.row_dimensions[21].height = 25
        ws.row_dimensions[26].height = 25
        column_widths = { 
            'A': 25, 'B': 15, 'C': 15, 'D': 17, 'E': 15, 'F': 17, 'G': 15, 'H': 15, 'I': 17, 'J': 15}
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        for rec in self:
            ws['F2'] = rec.change_received_from
            ws['C3'] = rec.ecn_number
            ws['G3'] = rec.ecn_date
            ws['C4'] = rec.project.name
            ws['G4'] = rec.emp_id.name
            ws['C5'] = rec.part_id.name
            ws['G5'] = rec.department.name
            ws['C6'] = rec.part_name
            ws['G6'] = rec.co_ord.name
            ws['C7'] = rec.part_number
            ws['G7'] = rec.partner_id.name
            names_line = ", ".join(rec.related_parts.mapped("name"))
            ws["C8"] = names_line   
            ws['G8'] = rec.end_date
            ws['C9'] = rec.target_completion_date
            ws['G9'] = rec.completion_status
            ws['C12'] = rec.change_received_from
            ws['H12'] = rec.change_person.name
            ws['C13'] = html2plaintext(rec.description or '')
            ws['C13'].alignment = align_left
            ws['C14'] = rec.existing
            ws['G14'] = rec.proposed
            ws['C17'] = "☑" if rec.scraped else "☐"
            ws['C18'] = "☑" if rec.userd_in else "☐"
            ws['C19'] = "☑" if rec.specification else "☐"
            ws['C20'] = rec.quantity
            ws['C22'] = "☑" if rec.specific_change else "☐"
            ws['C23'] = "☑" if rec.process_change else "☐"
            ws['C24'] = "☑" if rec.quality_imp else "☐"
            ws['C25'] = "☑" if rec.cost_reduce else "☐"
            ws['H22'] = "☑" if rec.standardisation else "☐"
            ws['H23'] = "☑" if rec.sub_suppiler else "☐"
            ws['H24'] = "☑" if rec.improve_machine else "☐"
            ws['H25'] = "☑" if rec.customer_requirement else "☐"
        current_row = 28
        for rec in self.four_m_type_ids:
            ws[f'A{current_row}'] = rec.types.name if rec.types else ''
            ws[f'A{current_row}'].border = border
            ws[f'A{current_row}'].alignment = align_center
            ws[f'B{current_row}'] = "☑" if rec.man else "☐"
            ws[f'B{current_row}'].border = border
            ws[f'B{current_row}'].alignment = align_center
            ws[f'C{current_row}'] = "☑" if rec.machine else "☐"
            ws[f'C{current_row}'].border = border            
            ws[f'C{current_row}'].alignment = align_center
            ws[f'D{current_row}'] = "☑" if rec.material else "☐"
            ws[f'D{current_row}'].border = border
            ws[f'D{current_row}'].alignment = align_center
            ws[f'E{current_row}'] = "☑" if rec.method else "☐"
            ws[f'E{current_row}'].border = border
            ws[f'E{current_row}'].alignment = align_center
            ws[f'F{current_row}'] = "☑" if rec.measure else "☐"
            ws[f'F{current_row}'].border = border
            ws[f'F{current_row}'].alignment = align_center
            ws[f'G{current_row}'] = "☑" if rec.environment else "☐"
            ws[f'G{current_row}'].border = border
            ws[f'G{current_row}'].alignment = align_center
            ws[f'H{current_row}'] = rec.description if rec.description else ''
            ws[f'H{current_row}'].border = border
            ws[f'H{current_row}'].alignment = align_center
            ws[f'I{current_row}'] = rec.responsibility_id.name if rec.responsibility_id else ''
            ws[f'I{current_row}'].border = border
            ws[f'I{current_row}'].alignment = align_center
            ws[f'J{current_row}'] = rec.status if rec.status else ''
            ws[f'J{current_row}'].border = border
            ws[f'J{current_row}'].alignment = align_center
            current_row += 1
        current_row += 1
        

        data = {
            f"A{current_row+1}":  "Impact of Changes",
            # feasibility section
            f"A{current_row+6}":  "Feasibility Assesment",
            f"A{current_row+7}":  "Feasible",
            f"E{current_row+7}":  "Not Feasible",

            # team‑assignment header row
            f"A{current_row+8}":  "ASSIGN CFT MEMBERS",
            f"A{current_row+9}":  "Team Member",
            f"C{current_row+9}":  "Department",
            f"E{current_row+9}":  "Status",
            f"F{current_row+9}":  "Approval Date",
            f"H{current_row+9}":  "Comments",
        }
        for cell, value in data.items():
            ws[cell] = value
            ws[cell].font = font_header
            ws[cell].alignment = align_center
            ws[cell].border = border
        ws.row_dimensions[current_row+1].height = 30
        ws.row_dimensions[current_row+6].height = 25
        ws.row_dimensions[current_row+8].height = 30
        ws[f'A{current_row+1}'].fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        ws[f'A{current_row+6}'].fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        ws[f'A{current_row+8}'].fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")

        data1 = {
            f"A{current_row+2}":  "1.Can engineering performance specifications be met as written?",
            f"A{current_row+3}":  "2.Is there adequate capacity to produce product?",
            f"A{current_row+4}":  "3.Can product be manufactured with Cpks that meet requirements?",
            f"A{current_row+5}":  "4.Can statistical process control required on product?",

            # questions 5‑8
            f"E{current_row+2}":  "5.Does the design allow the use of efficient material handling techniques?",
            f"E{current_row+3}":  "6.Costs for tooling?",
            f"E{current_row+4}":  "7.Costs for materials?",
            f"E{current_row+5}":  "8.Alternative manufacturing methods?",
        }
        for cell, value in data1.items():
            ws[cell] = value
            ws[cell].alignment = align_left
            ws[cell].border = border

        for i in self:
            ws[f'D{current_row+2}'] = "☑" if i.first_impact else "☐"
            ws[f'D{current_row+2}'].border = border
            ws[f'D{current_row+2}'].alignment = align_center
            ws[f'D{current_row+3}'] = "☑" if i.secound_impact else "☐"
            ws[f'D{current_row+3}'].border = border
            ws[f'D{current_row+3}'].alignment = align_center
            ws[f'D{current_row+4}'] = "☑" if i.third_impact else "☐"
            ws[f'D{current_row+4}'].border = border
            ws[f'D{current_row+4}'].alignment = align_center
            ws[f'D{current_row+5}'] = "☑" if i.forth_impact else "☐"
            ws[f'D{current_row+5}'].border = border
            ws[f'D{current_row+5}'].alignment = align_center
            ws[f'H{current_row+2}'] = "☑" if i.fifth_impact else "☐"
            ws[f'H{current_row+2}'].border = border
            ws[f'H{current_row+2}'].alignment = align_center
            ws[f'H{current_row+3}'] = "☑" if i.sixth_impact else "☐"
            ws[f'H{current_row+3}'].border = border
            ws[f'H{current_row+3}'].alignment = align_center
            ws[f'H{current_row+4}'] = "☑" if i.seven_impact else "☐"
            ws[f'H{current_row+4}'].border = border
            ws[f'H{current_row+4}'].alignment = align_center
            ws[f'H{current_row+5}'] = "☑" if i.eight_impact else "☐"
            ws[f'H{current_row+5}'].border = border
            ws[f'H{current_row+5}'].alignment = align_center
            ws[f'D{current_row+7}'] = "☑" if i.feasible else "☐"
            ws[f'D{current_row+7}'].border = border
            ws[f'D{current_row+7}'].alignment = align_center
            ws[f'H{current_row+7}'] = "☑" if i.not_feasible else "☐"
            ws[f'H{current_row+7}'].border = border
            ws[f'H{current_row+7}'].alignment = align_center
        merge_ranges = [
            f"A{current_row}:J{current_row}", 
            f"A{current_row+1}:J{current_row+1}", 
            f"A{current_row+2}:C{current_row+2}", f"E{current_row+2}:G{current_row+2}", f"H{current_row+2}:J{current_row+2}",
            f"A{current_row+3}:C{current_row+3}", f"E{current_row+3}:G{current_row+3}", f"H{current_row+3}:J{current_row+3}",
            f"A{current_row+4}:C{current_row+4}", f"E{current_row+4}:G{current_row+4}", f"H{current_row+4}:J{current_row+4}",
            f"A{current_row+5}:C{current_row+5}", f"E{current_row+5}:G{current_row+5}", f"H{current_row+5}:J{current_row+5}",
            f"A{current_row+6}:J{current_row+6}",
            f"A{current_row+7}:C{current_row+7}", f"E{current_row+7}:G{current_row+7}", f"H{current_row+7}:J{current_row+7}",
            f"A{current_row+8}:J{current_row+8}", f"A{current_row+9}:B{current_row+9}", f"C{current_row+9}:D{current_row+9}",
            f"F{current_row+9}:G{current_row+9}", f"H{current_row+9}:J{current_row+9}"
        ]
        for merge_range in merge_ranges:
            ws.merge_cells(merge_range)
        
        
        
        
        
        
        current_row = current_row + 10
        for i in self:
            for rec in i.team_approval_ids:
                ws[f'A{current_row}'] = rec.employee_id.name if rec.employee_id else ''
                ws[f'A{current_row}'].border = border
                ws[f'A{current_row}'].alignment = align_center
                ws.merge_cells(f'A{current_row}:B{current_row}')
                ws[f'C{current_row}'] = rec.department_id.name if rec.department_id else ''
                ws[f'C{current_row}'].border = border
                ws[f'C{current_row}'].alignment = align_center
                ws.merge_cells(f'C{current_row}:D{current_row}')
                ws[f'E{current_row}'] = rec.approval_status if rec.approval_status else ''
                ws[f'E{current_row}'].border = border
                ws[f'E{current_row}'].alignment = align_center
                ws[f'F{current_row}'] = rec.approval_date if rec.approval_date else ''
                ws[f'F{current_row}'].border = border
                ws[f'F{current_row}'].alignment = align_center
                ws.merge_cells(f'F{current_row}:G{current_row}')
                ws[f'H{current_row}'] = rec.comments if rec.comments else ''
                ws[f'H{current_row}'].border = border
                ws[f'H{current_row}'].alignment = align_center
                ws.merge_cells(f'H{current_row}:J{current_row}')
                current_row += 1
        current_row += 2
        data = {
            f'A{current_row}': 'Implementation Planning',
            f'A{current_row+1}': 'Sl.No',
            f'B{current_row+1}': 'Format',
            f'D{current_row+1}': 'Change Required',
            f'F{current_row+1}': 'Responsibility',
            f'H{current_row+1}': 'Target Date',
            f'I{current_row+1}': 'Attachments',
        }
        for cell, value in data.items():
            ws[cell] = value
            ws[cell].font = font_header
            ws[cell].alignment = align_center
            ws[cell].border = border
        ws.row_dimensions[current_row].height = 30
        ws.row_dimensions[current_row+1].height = 20
        ws[f'A{current_row}'].fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        
        merge_ranges = [
            f"A{current_row}:J{current_row}", 
            f"B{current_row+1}:C{current_row+1}", f"D{current_row+1}:E{current_row+1}", f"F{current_row+1}:G{current_row+1}",
            f"I{current_row+1}:J{current_row+1}"
        ]
        for merge_range in merge_ranges:
            ws.merge_cells(merge_range)
        
        current_row += 2
        for i in self:
            for rec in i.change_required_ids:
                ws[f'A{current_row}'] = rec.sl_no if rec.sl_no else ''
                ws[f'A{current_row}'].border = border
                ws[f'A{current_row}'].alignment = align_center
                ws[f'B{current_row}'] = rec.format.name if rec.format else ''
                ws[f'B{current_row}'].border = border
                ws[f'B{current_row}'].alignment = align_center
                ws.merge_cells(f'B{current_row}:C{current_row}')
                ws[f'D{current_row}'] = rec.change_required if rec.change_required else ''
                ws[f'D{current_row}'].border = border
                ws[f'D{current_row}'].alignment = align_center
                ws.merge_cells(f'D{current_row}:E{current_row}')
                ws[f'F{current_row}'] = rec.responsibility_id.name if rec.responsibility_id else ''
                ws[f'F{current_row}'].border = border
                ws[f'F{current_row}'].alignment = align_center
                ws.merge_cells(f'F{current_row}:G{current_row}')
                ws[f'H{current_row}'] = rec.target_date if rec.target_date else ''
                ws[f'H{current_row}'].border = border
                ws[f'H{current_row}'].alignment = align_center
                ws[f'I{current_row}'] = rec.attachments if rec.attachments else ''
                ws[f'I{current_row}'].border = border
                ws[f'I{current_row}'].alignment = align_center
                ws.merge_cells(f'I{current_row}:J{current_row}')
                current_row += 1
        current_row += 2
        data = {
            f'A{current_row}': 'Additional Formats',
            f'A{current_row+1}': 'Format',
            f'C{current_row+1}': 'Description',
            f'E{current_row+1}': 'Change Required',
            f'G{current_row+1}': 'Responsibility',
            f'I{current_row+1}': 'Target Date',
            f'J{current_row+1}': 'Attachments',
        }
        ws[f'A{current_row}'].fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        ws[f'A{current_row+1}'].fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
        ws[f'C{current_row+1}'].fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
        ws[f'E{current_row+1}'].fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
        ws[f'G{current_row+1}'].fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
        ws[f'I{current_row+1}'].fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
        ws[f'J{current_row+1}'].fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

        for cell, value in data.items():
            ws[cell] = value
            ws[cell].font = font_header
            ws[cell].alignment = align_center
            ws[cell].border = border
        ws.row_dimensions[current_row].height = 30
        ws.row_dimensions[current_row+1].height = 20
        merge_ranges = [
            f"A{current_row}:J{current_row}",
            f"A{current_row+1}:B{current_row+1}", f"C{current_row+1}:D{current_row+1}", f"E{current_row+1}:F{current_row+1}",
            f"G{current_row+1}:H{current_row+1}"
        ]
        for merge_range in merge_ranges:
            ws.merge_cells(merge_range)
        
        current_row += 2
        for i in self:
            for rec in i.new_format_ids:
                ws[f'A{current_row}'] = rec.formate_name.name if rec.format else ''
                ws[f'A{current_row}'].border = border
                ws[f'A{current_row}'].alignment = align_center
                ws.merge_cells(f'A{current_row}:B{current_row}')
                ws[f'C{current_row}'] = rec.description if rec.description else ''
                ws[f'C{current_row}'].border = border
                ws[f'C{current_row}'].alignment = align_center
                ws.merge_cells(f'C{current_row}:D{current_row}')
                ws[f'E{current_row}'] = rec.change_required if rec.change_required else ''
                ws[f'E{current_row}'].border = border
                ws[f'E{current_row}'].alignment = align_center
                ws.merge_cells(f'E{current_row}:F{current_row}')
                ws[f'G{current_row}'] = rec.responsibility_id.name if rec.responsibility_id else ''
                ws[f'G{current_row}'].border = border
                ws[f'G{current_row}'].alignment = align_center
                ws.merge_cells(f'G{current_row}:H{current_row}')
                ws[f'I{current_row}'] = rec.target_date if rec.target_date else ''
                ws[f'I{current_row}'].border = border
                ws[f'I{current_row}'].alignment = align_center
                ws[f'J{current_row}'] = rec.attachments if rec.attachments else ''
                ws[f'J{current_row}'].border = border
                ws[f'J{current_row}'].alignment = align_center
                current_row += 1
        
        current_row += 2
        data = {
            f'A{current_row}': 'PPAP REQUIREMENTS',
            f'A{current_row+1}': 'PPAP Approval Required',
            f'D{current_row+1}': 'Date of Pilot Lot Submission',
            f'G{current_row+1}': 'Date of Regular Production',
        }
        for cell, value in data.items():
            ws[cell] = value
            ws[cell].font = font_header
            ws[cell].alignment = align_center
            ws[cell].border = border
        ws.row_dimensions[current_row].height = 30
        ws[f'A{current_row}'].fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        
        merge_ranges = [
            f"A{current_row}:J{current_row}",
            f"A{current_row+1}:B{current_row+1}", f"D{current_row+1}:E{current_row+1}",
            f"G{current_row+1}:H{current_row+1}", f"I{current_row+1}:J{current_row+1}"
        ]
        for merge_range in merge_ranges:
            ws.merge_cells(merge_range)

        for i in self:
            ws[f'C{current_row+1}'] = "☑" if i.ppap_approval_required else "☐"
            ws[f'C{current_row+1}'].border = border
            ws[f'C{current_row+1}'].alignment = align_center
            ws[f'F{current_row+1}'] = i.date_of_pilot_lot_submission if i.date_of_pilot_lot_submission else ''
            ws[f'F{current_row+1}'].border = border
            ws[f'F{current_row+1}'].alignment = align_center
            ws[f'I{current_row+1}'] = i.date_of_regular_submission if i.date_of_regular_submission else ''
            ws[f'I{current_row+1}'].border = border
            ws[f'I{current_row+1}'].alignment = align_center
            
        
        data = {
            
            f"A{current_row+2}":  "Completion / Rejection",  # Border and alignment are set here
            f"A{current_row+3}":  "Completion Notes",
            f"A{current_row+4}":  "Approval History",
            f"A{current_row+5}":  "Date",
            f"B{current_row+5}":  "User",
            f"D{current_row+5}":  "Action",
            f"G{current_row+5}":  "Comments",
        }
        ws[f'A{current_row+2}'].fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        ws[f'A{current_row+4}'].fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        ws[f'A{current_row+5}'].fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
        ws[f'B{current_row+5}'].fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
        ws[f'D{current_row+5}'].fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
        ws[f'G{current_row+5}'].fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
        
        for cell, value in data.items():
            ws[cell] = value
            ws[cell].font = font_header
            ws[cell].alignment = align_center
            ws[cell].border = border
        ws.row_dimensions[current_row+2].height = 30
        ws.row_dimensions[current_row+4].height = 20
        merge_ranges = [
            f"A{current_row+2}:J{current_row+2}",
            f"A{current_row+3}:B{current_row+3}",
            f"A{current_row+4}:J{current_row+4}",
            f"B{current_row+5}:C{current_row+5}",
            f"D{current_row+5}:F{current_row+5}",
            f"G{current_row+5}:J{current_row+5}",
        ]
        for merge_range in merge_ranges:
            ws.merge_cells(merge_range)
        for i in self:
            ws[f'C{current_row+3}'] = i.completion_notes if i.completion_notes else ''
            
            ws[f'C{current_row+3}'].border = border
            ws[f'C{current_row+3}'].alignment = align_center
            ws.merge_cells(f'C{current_row+3}:J{current_row+3}')
        current_row = current_row + 6
        for i in self:
            for rec in i.approval_history_ids:
                ws[f'A{current_row}'] = rec.date if rec.date else ''
                ws[f'A{current_row}'].border = border
                ws[f'A{current_row}'].alignment = align_center
                ws[f'B{current_row}'] = rec.user_id.name if rec.user_id else ''
                ws[f'B{current_row}'].border = border
                ws[f'B{current_row}'].alignment = align_center
                ws.merge_cells(f'B{current_row}:C{current_row}')
                ws[f'D{current_row}'] = dict(rec._fields['action'].selection).get(rec.action, '') if rec.action else ''
                ws[f'D{current_row}'].border = border
                ws[f'D{current_row}'].alignment = align_center
                ws.merge_cells(f'D{current_row}:F{current_row}')
                ws[f'G{current_row}'] = rec.comments if rec.comments else ''
                ws[f'G{current_row}'].border = border
                ws[f'G{current_row}'].alignment = align_center
                ws.merge_cells(f'G{current_row}:J{current_row}')
                
                current_row += 1






                

            



        
        wb.save(output)
        output.seek(0)

        attachment = self.env["ir.attachment"].create({
            "name": "ECN\u2044ECR Report.xlsx",
            "type": "binary",
            "datas": base64.b64encode(output.getvalue()),
            "res_model": self._name,
            "res_id": self.id,
            "mimetype": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        })










        return {"type": "ir.actions.act_url", "url": f"/web/content/{attachment.id}?download=true", "target": "self"}





    @api.depends('all_approved', 'any_rejected')
    def _compute_final_status(self):
        for record in self:
            if record.all_approved:
                record.final_status = 'approved'
            elif record.any_rejected:
                record.final_status = 'rejected'
            else:
                record.final_status = False

    def dummy_action(self):
        """
        Dummy action function for the button.
        This can be customized to perform any desired action.
        """
        # Display a notification message
        message = "button was clicked! Quantity fetched"
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': 'Information',
                'message': message,
                'sticky': False,
                'type': 'info',
            }
        }
    @api.depends('state')
    def _compute_end_date(self):
        for record in self:
            if record.state == 'completed':
                record.end_date = fields.Date.today()

    def _inverse_end_date(self):
        pass  # This method is required but can be empty

    @api.depends('team_approval_ids.approval_status')
    def _compute_approval_status(self):
        for record in self:
            record.all_approved = record.team_approval_ids and all(
                line.approval_status == 'approved' for line in record.team_approval_ids)
            record.any_rejected = any(line.approval_status == 'rejected' for line in record.team_approval_ids)
            record.any_review = any(line.approval_status == 'review' for line in record.team_approval_ids)

    @api.depends('team_approval_ids.employee_id.user_id')
    def _compute_user_is_team_member(self):
        """Check if current user is in the CFT team"""
        current_user = self.env.user
        for record in self:
            employee = self.env['hr.employee'].search([('user_id', '=', current_user.id)], limit=1)
            record.user_is_team_member = employee and employee.id in record.team_approval_ids.mapped('employee_id').ids

    @api.model
    def create(self, vals):
        # Generate ECN/ECR number if not provided
        if not vals.get('ecn_number'):
            sequence = self.env['ir.sequence'].next_by_code('asd.ecn')
            vals['ecn_number'] = sequence or _('New')
        return super(EngineeringChangeNote, self).create(vals)

    def send_notification_emails(self, template_name, recipients):
        """Send notification emails based on template name"""
        self.ensure_one()
        template = self.env.ref(template_name, raise_if_not_found=False)
        if not template:
            self.message_post(body=_("Email template %s not found") % template_name)
            return

        for recipient in recipients:
            template.with_context(email_to=recipient.work_email).send_mail(
                self.id, force_send=True,
                email_values={'email_to': recipient.work_email},


            )
            self.message_post(body=_("Notification email sent to %s") % recipient.name)

    # Workflow actions
    def action_submit(self):
        """Submit the ECN/ECR for review"""
        for record in self:
            # Validate required fields
            if not record.emp_id:
                raise ValidationError(_("Please select an Proposer Name before submitting."))

            # Validate team members are added
            if not record.team_approval_ids:
                raise ValidationError(_("Please add CFT team members before submitting."))

            record.state = 'submitted'

            # Add to approval history
            self.env['ecn.approval.history'].create({
                'ecn_id': record.id,
                'user_id': self.env.user.id,
                'action': 'submitted',
                'date': fields.Datetime.now(),
                'comments': _("ECN/ECR submitted for review")
            })

            # Send email notifications to all team members
            team_members = record.team_approval_ids.mapped('employee_id')
            record.send_notification_emails('ecn.mail_template_ecn_submitted', team_members)

    def action_start_review(self):
        """Start the review process"""
        for record in self:
            # Validate required fields
            if not record.description:
                raise ValidationError(_("Please provide a change description before submitting."))
            record.state = 'under_review'

            # Add to approval history
            self.env['ecn.approval.history'].create({
                'ecn_id': record.id,
                'user_id': self.env.user.id,
                'action': 'started_review',
                'date': fields.Datetime.now(),
                'comments': _("Review process started")
            })

    def action_reset_to_draft(self):
        """Reset to draft state"""
        for record in self:
            record.state = 'draft'

    def action_move_to_feasibility(self):
        """Move to feasibility check stage"""
        for record in self:
            record.state = 'feasibility_check'

            # Add to approval history
            self.env['ecn.approval.history'].create({
                'ecn_id': record.id,
                'user_id': self.env.user.id,
                'action': 'feasibility_check',
                'date': fields.Datetime.now(),
                'comments': _("Moved to feasibility check")
            })
            # Send email notifications to all team members
            team_members = record.team_approval_ids.mapped('employee_id')
            record.send_notification_emails('ecn.mail_template_ecn_status_update', team_members)

    def action_request_cft_approval(self):
        """Request Cross-Functional Team approval"""
        for record in self:
            # Verify all team members have approved
            missing_approvals = []
            for line in record.team_approval_ids:
                if line.approval_status != 'approved':
                    department = line.department_id.name or "Unknown Department"
                    if department not in missing_approvals:
                        missing_approvals.append(department)

            if missing_approvals:
                missing_depts = ', '.join(missing_approvals)
                raise ValidationError(
                    _("All team members must approve before moving to implementation planning. Missing approvals from: %s") % missing_depts)

            # Check if feasibility is marked
            if not (record.feasible or record.not_feasible):
                raise ValidationError(_("Please mark the feasibility status before requesting CFT approval."))

            if record.not_feasible:
                return record.action_reject()

            record.state = 'cft_approval'

            # Add to approval history
            self.env['ecn.approval.history'].create({
                'ecn_id': record.id,
                'user_id': self.env.user.id,
                'action': 'cft_approval_requested',
                'date': fields.Datetime.now(),
                'comments': _("CFT approval requested")
            })

            # Reset all approval statuses to pending if any were previously approved
            for line in record.team_approval_ids:
                line.write({
                    'approval_status': 'pending',
                    'approval_date': False,
                    'comments': False
                })

            # Send email notifications to all team members
            team_members = record.team_approval_ids.mapped('employee_id')
            record.send_notification_emails('ecn.mail_template_ecn_status_update', team_members)

    def action_plan_implementation(self):
        """Plan implementation after approval"""
        for record in self:

            # Reset all approval statuses to pending for second approval round
            record._reset_approvals_for_implementation()

            # Auto-populate formats based on project
            if record.project:
                # Find formats related to this project
                format_ids = self.env['document.formate'].search([
                    ('used_in_project_type_ids', 'in', [record.project.used_in_project_type_id.id]),
                    ('sr_no', '=', 1)
                ])

                if format_ids:
                    # Clear existing lines first to avoid duplicates
                    record.change_required_ids.unlink()

                    # Create lines for each format
                    for format_id in format_ids:
                        self.env['asd.ecn.line'].create({
                            'change_required_id': record.id,
                            'format': format_id.id,
                            'change_required': False,  # Default value
                        })
                else:
                    # Notify if no formats found
                    record.message_post(body=_("No document formats found for this project type."))
            else:
                raise ValidationError(_("Please select a project before proceeding to implementation planning."))

            record.state = 'implementation'

            # Add to approval history
            self.env['ecn.approval.history'].create({
                'ecn_id': record.id,
                'user_id': self.env.user.id,
                'action': 'implementation_planning',
                'date': fields.Datetime.now(),
                'comments': _("Moved to implementation planning")
            })

            # Send notifications to CFT team for final approval
            team_members = record.team_approval_ids.mapped('employee_id')
            record.send_notification_emails('ecn.mail_template_ecn_implementation', team_members)

    def _reset_approvals_for_implementation(self):
        """Reset all approvals to pending state for implementation review"""
        self.ensure_one()
        for line in self.team_approval_ids:
            line.write({
                'approval_status': 'pending',
                'approval_date': False,
                'comments': False
            })

        # Record in approval history
        self.env['ecn.approval.history'].create({
            'ecn_id': self.id,
            'user_id': self.env.user.id,
            'action': 'reset_approvals',
            'date': fields.Datetime.now(),
            'comments': _("CFT approvals reset for implementation review")
        })

    def action_approve(self):
        """Approve the ECN/ECR by the current user"""
        self.ensure_one()
        current_user = self.env.user
        employee = self.env['hr.employee'].search([('user_id', '=', current_user.id)], limit=1)

        if not employee:
            raise ValidationError(_("You don't have an employee record linked to your user account."))

        approval_line = self.team_approval_ids.filtered(lambda l: l.employee_id.id == employee.id)
        if not approval_line:
            raise ValidationError(_("You are not authorized to approve this ECN."))

        approval_line.write({
            'approval_status': 'approved',
            'approval_date': fields.Date.today(),
        })

        # Record in approval history
        self.env['ecn.approval.history'].create({
            'ecn_id': self.id,
            'user_id': current_user.id,
            'action': 'approved',
            'date': fields.Datetime.now(),
            'comments': approval_line.comments
        })

        # Check if all approvals are complete after implementation
        if self.state == 'implementation' and self.all_approved:
            self.action_complete()

    def action_complete(self):
        """Complete the ECN/ECR if all approvals are received"""
        for record in self:
            if not record.all_approved:
                # Get the list of members who haven't approved
                missing_approvals = []
                for line in record.team_approval_ids:
                    if line.approval_status != 'approved':
                        department = line.department_id.name or "Unknown Department"
                        if department not in missing_approvals:
                            missing_approvals.append(department)

                # Raise validation error with specific missing approvals
                if missing_approvals:
                    missing_depts = ', '.join(missing_approvals)
                    raise ValidationError(
                        _("All team members must approve implementation before completing the ECN. Missing approvals from: %s") % missing_depts)

            # Send email notifications to all team members
            team_members = record.team_approval_ids.mapped('employee_id')
            record.send_notification_emails('ecn.mail_template_ecn_status_update', team_members)

            return {
                'type': 'ir.actions.act_window',
                'res_model': 'ecn.completion.wizard',
                'view_mode': 'form',
                'target': 'new',
                'context': {
                    'default_ecn_id': record.id,
                    'default_action_type': 'complete'
                }
            }

    def action_reject(self):
        """Open wizard for rejection reason"""
        for record in self:
            # Send email notifications to all team members
            team_members = record.team_approval_ids.mapped('employee_id')
            record.send_notification_emails('ecn.mail_template_ecn_status_update', team_members)

            return {
                'type': 'ir.actions.act_window',
                'res_model': 'ecn.completion.wizard',
                'view_mode': 'form',
                'target': 'new',
                'context': {
                    'default_ecn_id': self.id,
                    'default_action_type': 'reject'
                }
            }

    def action_request_review(self):
        """Request additional review by the current user"""
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'ecn.approval.wizard',
            'view_mode': 'form',
            'target': 'new',
            'context': {
                'default_ecn_id': self.id,
                'default_action_type': 'review'
            }
        }

    def action_generate_excel_report(self):
        """Generate Excel report for completed ECN"""
        self.ensure_one()
        if self.state != 'completed':
            raise UserError(_("Report can only be generated for completed ECNs."))

        # Implementation for Excel report generation would go here
        # This is a placeholder for actual Excel generation logic

        return {
            'type': 'ir.actions.act_window',
            'res_model': 'asd.ecn',
            'view_mode': 'form',
            'res_id': self.id,
            'target': 'current',
            'context': {'show_notification': True}
        }


class ECNCompletionWizard(models.TransientModel):
    _name = 'ecn.completion.wizard'
    _description = 'ECN/ECR Completion Wizard'
    _inherit = "translation.mixin"

    ecn_id = fields.Many2one('asd.ecn', required=True)
    action_type = fields.Selection([('complete', 'Complete'), ('reject', 'Reject')], required=True)
    notes = fields.Text('Notes', required=True,translate=True)

    def confirm_action(self):
        self.ensure_one()
        ecn = self.ecn_id
        ecn.completion_notes = self.notes

        if self.action_type == 'complete':
            # Validate PPAP dates if PPAP approval is required
            # if ecn.ppap_approval_required:
            #     if not ecn.date_of_ppap_submission:
            #         raise ValidationError(_("PPAP submission date is required before completion."))
            #     if not ecn.date_of_approval_received:
            #         raise ValidationError(_("PPAP approval date is required before completion."))
            #
            # # Validate production dates
            # if not ecn.date_of_regular_submission:
            #     raise ValidationError(_("Date of regular production is required before completion."))

            # Apply draft changes to all formats
            for line in ecn.change_required_ids:
                if line.change_required and hasattr(line, 'draft_changes') and line.draft_changes:
                    line.apply_draft_changes()
                    if hasattr(line, 'original_values'):
                        line.original_values = False  # Clear original values after completion

            ecn.state = 'completed'
            comments = _("ECN/ECR process completed")
        else:  # reject
            # Restore original values for all formats
            for line in ecn.change_required_ids:
                if line.change_required and hasattr(line, 'original_values') and line.original_values:
                    line.restore_original_values()
                    if hasattr(line, 'draft_changes'):
                        line.draft_changes = False
                    line.original_values = False

            ecn.state = 'rejected'
            comments = _("ECN/ECR rejected")

        # Record in approval history
        ecn.env['ecn.approval.history'].create({
            'ecn_id': ecn.id,
            'user_id': self.env.user.id,
            'action': 'rejected' if self.action_type == 'reject' else 'completed',
            'date': fields.Datetime.now(),
            'comments': self.notes
        })

        return {
            'type': 'ir.actions.act_window',
            'res_model': 'asd.ecn',
            'view_mode': 'form',
            'res_id': ecn.id,
            'target': 'current'
        }


class ECNApprovalWizard(models.TransientModel):
    _name = 'ecn.approval.wizard'
    _description = 'ECN/ECR Approval Wizard'
    _inherit = "translation.mixin"

    ecn_id = fields.Many2one('asd.ecn', required=True)
    action_type = fields.Selection([
        ('reject', 'Reject'),
        ('review', 'Request Review')
    ], required=True)
    comments = fields.Text('Comments', required=True,translate=True)

    def confirm_action(self):
        self.ensure_one()
        ecn = self.ecn_id
        current_user = self.env.user
        employee = self.env['hr.employee'].search([('user_id', '=', current_user.id)], limit=1)

        if not employee:
            raise ValidationError(_("You don't have an employee record linked to your user account."))

        approval_line = ecn.team_approval_ids.filtered(lambda l: l.employee_id.id == employee.id)
        if not approval_line:
            raise ValidationError(_("You are not authorized to approve this ECN."))

        if self.action_type == 'reject':
            status = 'rejected'
            action = 'rejected'
        else:  # review
            status = 'review'
            action = 'requested_review'

        approval_line.write({
            'approval_status': status,
            'approval_date': fields.Date.today(),
            'comments': self.comments
        })

        # Record in approval history
        ecn.env['ecn.approval.history'].create({
            'ecn_id': ecn.id,
            'user_id': current_user.id,
            'action': action,
            'date': fields.Datetime.now(),
            'comments': self.comments
        })

        return {
            'type': 'ir.actions.act_window',
            'res_model': 'asd.ecn',
            'view_mode': 'form',
            'res_id': ecn.id,
            'target': 'current'
        }


class ECNFormat(models.Model):
    _name = 'ecn.format'
    _description = 'ECN/ECR Format'
    _inherit = "translation.mixin"

    name = fields.Char('Format Name', required=True,translate=True)


class ECNNewFormat(models.Model):
    _name = 'ecn.new.format'
    _description = 'ECN/ECR New Format'
    _inherit = "translation.mixin"

    ecn_id = fields.Many2one('asd.ecn', string='ECN', ondelete='cascade')
    formate_name = fields.Many2one('ecn.format', string='Format')
    description = fields.Text('Description',translate=True)
    change_required = fields.Boolean("Change Required (Y/N)")
    responsibility_id = fields.Many2one("res.users", string='Responsibility')
    target_date = fields.Date('Target Date')
    ecn_attachment = fields.Binary('Attachments')


class EngineeringChangeNoteLine(models.Model):
    _name = 'asd.ecn.line'
    _description = 'Engineering Change Note Line'
    _inherit = "translation.mixin"

    change_required_id = fields.Many2one('asd.ecn', string='Change Required', ondelete='cascade')
    sl_no = fields.Integer('S.No', compute='_compute_sequence_number')
    # Update format field to filter based on project_id's project type
    format = fields.Many2one(
        'document.formate',
        string='Format')
    table = fields.Char('Table', related='format.table')

    change_required = fields.Boolean("Change Required (Y/N)")
    responsibility_id = fields.Many2one("res.users", string='Responsibility')
    target_date = fields.Date('Target Date')
    ecn_attachment = fields.Binary('Attachments')

    # Store original values and draft changes (JSON)
    original_values = fields.Text('Original Values', help='Stores original values before changes',translate=True)
    draft_changes = fields.Text('Draft Changes', help='Stores draft changes as JSON',translate=True)

    @api.depends('change_required_id')
    def _compute_sequence_number(self):
        for line in self:
            if not line.change_required_id or not line.change_required_id.change_required_ids:
                line.sl_no = 1
                continue

            try:
                # Track position manually
                position = 1
                related_lines = line.change_required_id.change_required_ids
                for rel_line in related_lines:
                    if rel_line.id == line.id:
                        line.sl_no = position
                        break
                    position += 1
                else:
                    line.sl_no = 1
            except Exception:
                # Fallback for new records
                line.sl_no = 1

    def action_open_format(self):
        """Open the format in edit mode, storing original values first"""
        self.ensure_one()
        if not self.change_required:
            raise UserError(_("Change Required must be checked to open the format."))
        # Check if current user is the responsible person
        current_user = self.env.user
        if self.responsibility_id and self.responsibility_id != current_user:
            raise UserError(_("You don't have access to this record. Only the responsible person can open it."))

        if not self.format or not self.format.table:
            raise UserError(_("No valid format configured."))

        try:
            # Find the related record
            target_model = self.env[self.format.table]
            domain = [
                ('part_id', '=', self.change_required_id.part_id.id),
                ('partner_id', '=', self.change_required_id.partner_id.id),
                ('project_id', '=', self.change_required_id.project.id)
            ]
            record = target_model.search(domain, limit=1)

            if not record:
                raise UserError(_("Related record not found. Please create it first."))

            # Store original values if not already stored
            if not self.original_values:
                # Get field values in a serializable format
                field_values = {}
                for field_name, field in record._fields.items():
                    if field_name in models.MAGIC_COLUMNS:
                        continue  # Skip internal fields
                    value = record[field_name]

                    # Handle special field types
                    if isinstance(field, fields.Datetime):
                        field_values[field_name] = value and value.strftime('%Y-%m-%d %H:%M:%S') or False
                    elif isinstance(field, fields.Date):
                        field_values[field_name] = value and value.strftime('%Y-%m-%d') or False
                    elif isinstance(field, fields.Many2one):
                        field_values[field_name] = value and value.id or False
                    elif isinstance(field, (fields.One2many, fields.Many2many)):
                        field_values[field_name] = value and [(6, 0, value.ids)] or False
                    else:
                        field_values[field_name] = value

                original_data = {
                    'id': record.id,
                    'values': field_values
                }
                self.original_values = json.dumps(original_data)

            # Prepare context to pass draft changes
            context = {
                'default_draft_changes': self.draft_changes,
                'ecn_line_id': self.id,
                'create': False,
                'edit': True
            }

            return {
                'type': 'ir.actions.act_window',
                'name': self.format.name,
                'view_mode': 'form',
                'res_model': self.format.table,
                'res_id': record.id,
                'target': 'current',
                'context': context
            }
        except Exception as e:
            raise UserError(_("Error opening format: %s") % str(e))

    def save_draft_changes(self, changes):
        """Save changes as draft (JSON format)"""
        self.ensure_one()
        self.draft_changes = json.dumps(changes)
    def apply_draft_changes(self):
        """Apply draft changes to the original format"""
        self.ensure_one()
        if not self.draft_changes:
            return

        try:
            changes = json.loads(self.draft_changes)
            target_model = self.env[self.format.table]
            record = target_model.browse(changes.get('id'))
            if record.exists():
                record.write(changes.get('data', {}))
        except Exception as e:
            raise UserError(_("Error applying changes: %s") % str(e))

    def restore_original_values(self):
        """Restore original values to the format"""
        self.ensure_one()
        if not self.original_values:
            return

        try:
            original_data = json.loads(self.original_values)
            target_model = self.env[self.format.table]
            record = target_model.browse(original_data.get('id'))

            if record.exists():
                values = original_data.get('values', {})

                # Convert string dates back to datetime objects
                for field_name, field in record._fields.items():
                    if field_name in values:
                        if isinstance(field, fields.Datetime) and values[field_name]:
                            values[field_name] = datetime.strptime(values[field_name], '%Y-%m-%d %H:%M:%S')
                        elif isinstance(field, fields.Date) and values[field_name]:
                            values[field_name] = datetime.strptime(values[field_name], '%Y-%m-%d').date()

                # Filter out special fields
                values.pop('id', None)
                values.pop('create_uid', None)
                values.pop('create_date', None)
                values.pop('write_uid', None)
                values.pop('write_date', None)
                values.pop('__last_update', None)

                record.write(values)
        except Exception as e:
            raise UserError(_("Error restoring original values: %s") % str(e))

class ECNCategory(models.Model):
    _name = 'ecn.category'
    _description = 'ECN/ECR Category'
    _inherit = "translation.mixin"

    name = fields.Char('Category Name',translate=True)


class ECN4MType(models.Model):
    _name = 'ecn.four.m.type'
    _description = 'ECN/ECR 4M Type'
    _inherit = "translation.mixin"

    ecn_id = fields.Many2one('asd.ecn', string='ECN', ondelete='cascade')

    types=fields.Many2one('ecn.category', string='Type', required=True)
    description=fields.Char(string="Action Plan",translate=True)
    responsibility_id = fields.Many2one("res.users", string='Responsibility')
    status=fields.Selection([
        ('open', 'Open'), ('not_started', 'Not Started'),
        ('in_progress', 'In Progress'),
        ('closed', 'Closed')
    ], string='Status', default='not_started')

    # 4M impacts (Man, Machine, Material, Method,Measurement, Environment)
    man = fields.Boolean(string="Man")
    machine = fields.Boolean(string="Machine")
    material = fields.Boolean(string="Material")
    method = fields.Boolean(string="Method")
    environment = fields.Boolean(string="Environment")
    measure=fields.Boolean(string="Measurement")

    @api.constrains('man', 'machine', 'material', 'method', 'environment','measure')
    def _check_only_one_selected(self):
        for record in self:
            selected = sum([
                record.man,
                record.machine,
                record.material,
                record.method,
                record.environment,
                record.measure
            ])
            if selected > 1:
                raise ValidationError(
                    "Only one of the fields (Man, Machine, Material, Method, Measurement, Environment) can be selected.")


class ECNTeamApproval(models.Model):
    _name = 'ecn.team.approval'
    _description = 'ECN/ECR Team Member Approval'
    _inherit = "translation.mixin"

    ecn_id = fields.Many2one('asd.ecn', string='ECN', ondelete='cascade')
    employee_id = fields.Many2one('hr.employee', string='Team Member', required=True)
    department_id = fields.Many2one('hr.department', related='employee_id.department_id', string='Department', store=True)
    approval_status = fields.Selection([
        ('pending', 'Pending'),
        ('approved', 'Approved'),
        ('rejected', 'Rejected'),
        ('review', 'Need Review')
    ], string='Status', default='pending', tracking=True)
    approval_date = fields.Date(string='Approval Date')
    comments = fields.Text(string='Comments',translate=True)


class ECNApprovalHistory(models.Model):
    _name = 'ecn.approval.history'
    _description = 'ECN/ECR Approval History'
    _order = 'date desc'
    _inherit = "translation.mixin"

    ecn_id = fields.Many2one('asd.ecn', string='ECN', ondelete='cascade')
    user_id = fields.Many2one('res.users', string='User', required=True)
    date = fields.Datetime(string='Date', default=fields.Datetime.now)
    action = fields.Selection([
        ('submitted', 'Submitted'),
        ('started_review', 'Started Review'),
        ('feasibility_check', 'Moved to Feasibility Check'),
        ('cft_approval_requested', 'CFT Approval Requested'),
        ('approved', 'Approved'),
        ('rejected', 'Rejected'),
        ('implementation_planning', 'Implementation Planning'),
        ('reset_approvals', 'Reset Approvals for Implementation'),
        ('completed', 'Completed')
    ], string='Action', required=True)
    comments = fields.Text('Comments',translate=True)

    def name_get(self):
        result = []
        for record in self:
            name = f"{record.ecn_id.ecn_number} - {record.action} - {record.date}"
            result.append((record.id, name))
        return result