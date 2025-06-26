from odoo import api, fields, models, _
from datetime import date, datetime
from odoo.exceptions import ValidationError
import base64
from openpyxl import Workbook
from PIL import Image as PILImage, ImageOps
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
import io
from io import BytesIO
import math
import logging
_logger = logging.getLogger(__name__)
class Inprocess_IR(models.Model):
    _name = 'in.process.inspection.report'
    _description = 'Report of in process inspection'
    _inherit = ['iatf.sign.off.members', 'revision.history.mixin']
    _rec_name = 'part_name'  # Added for better record identification
    
    # customer_id=fields.Many2one('res.partner',string='Customer/Supplier Name')
    # part_name=fields.Many2one('product.template',string='Part/Assembly Name')
    # part_no = fields.Char("Part No.", related="part_name.default_code", store=True)
    # drg_no = fields.Char("Drg. No.", related="part_name.drg_no", store=True)

    batch_quantity=fields.Integer("Batch Quantity")
    page_no=fields.Char("Page No.")
    date=fields.Date("Date",default=lambda self: fields.Date.today())

    rev_no=fields.Char("Rev. No.")
    rev_date = fields.Date("Rev. Date")
    parts_to_be_checked=fields.Integer("Number of Parts to be Checked")
    line_ids = fields.One2many('in.process.operations', 'operation_id', string="Inspection Lines")
    generate_xls_file = fields.Binary(string="Generated File")

    @api.model
    def create(self, vals):
        """Ensure updates in Incoming IR reflect in Control Plan."""
        # Validate required fields before creating
        if 'project_id' not in vals or not vals.get('project_id'):
            raise ValidationError(_("Project is required to create an inspection report."))

        record = super().create(vals)
        try:

            for line in record.line_ids:
                line._update_observations()
        except Exception as e:
            _logger.error("Error syncing to control plan: %s", str(e))
            raise ValidationError(
                _("Failed to sync with control plan. Please ensure all required data is available. Details: %s") % str(
                    e))
        return record

    def write(self, vals):
        """Ensure updates in Incoming IR reflect in Control Plan."""
        res = super().write(vals)
        try:

            if 'batch_quantity' in vals:
                for line in self.line_ids:
                    line._update_observations()
        except Exception as e:
            _logger.error("Error syncing to control plan: %s", str(e))
            raise ValidationError(
                _("Failed to sync with control plan. Please ensure all required data is available. Details: %s") % str(
                    e))
        return res

    @api.onchange('batch_quantity')
    def _onchange_batch_quantity(self):
        """Update observations for all lines when batch quantity changes"""
        if self.line_ids:
            for operation in self.line_ids:
                for line in operation.op_ids:
                    line._update_observations()


    def generate_xls_report(self):
        # Create workbook and worksheet
        output = BytesIO()
        wb = Workbook()
        # Group operations by operation number
        operations_by_op_no = {}
        for rec in self:
            for operation in rec.line_ids:
                op_key = operation.op_no.id if operation.op_no else 'default'
                if op_key not in operations_by_op_no:
                    operations_by_op_no[op_key] = []
                operations_by_op_no[op_key].append(operation)
        # Remove default worksheet
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

        # Create worksheets for each operation number
        for op_key, operations in operations_by_op_no.items():
            if not operations:
                continue

            op_name = operations[0].op_name or "Operation " + str(op_key)
            op_no = operations[0].op_no.process_step if operations[0].op_no else ""

            # Create a valid worksheet name (max 31 chars, no special chars)
            sheet_name = (op_no + "-" + op_name)[:31].replace('/', '_').replace('\\', '_')
            # Ensure uniqueness
            if sheet_name in wb.sheetnames:
                sheet_name = sheet_name[:27] + "_" + str(len(wb.sheetnames))

            ws = wb.create_sheet(title=sheet_name)

        # ws = wb.active

            # Define styles
            border = Border(top=Side(style='thin'), left=Side(style='thin'),
                            right=Side(style='thin'), bottom=Side(style='thin'))
            font_header = Font(name='Times New Roman', bold=True)
            title_font=Font(size=20, bold=True)
            align_center = Alignment(vertical='center', horizontal='center', wrapText=True)

            # Define fill colors
            fill = PatternFill(start_color="e7e7e7", end_color="e7e7e7", fill_type="solid")  # Grey
            orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")


            if self.env.user.company_id and self.env.user.company_id.logo:

                max_width, max_height = 100, 200  # Set max dimensions
                image_data = base64.b64decode(self.env.user.company_id.logo)

                # Open image with PIL
                image = PILImage.open(io.BytesIO(image_data))
                width, height = image.size
                aspect_ratio = width / height

                # Resize with aspect ratio
                if width > max_width:
                    width = max_width
                    height = int(width / aspect_ratio)
                if height > max_height:
                    height = max_height
                    width = int(height * aspect_ratio)

                # Resize and add padding
                resized_image = image.resize((width, height), PILImage.LANCZOS)
                padding_top, padding_left = 5,5
                resized_image = ImageOps.expand(resized_image, border=(padding_left, padding_top, 0, 0), fill='white')

                img_bytes = io.BytesIO()
                resized_image.save(img_bytes, format='PNG')
                img_bytes.seek(0)
                logo_image = Image(img_bytes)

                ws.add_image(logo_image, 'A1')
            # Determine maximum number of observations dynamically
            max_observations = self.batch_quantity if self.batch_quantity else 5

            # Headers dictionary
            headers = {
                'C1':'In Process Inspection Report',
                'A5':'Customer/Supplier Name',
                'A7':'Part / Assembly Name',
                'A9':'Drawing No.',
                'A11':'Batch Quantity',
                # 'P1':'Page No.',
                'P1':'Date',
                'P3':'Operation No.',
                'P5':'Operation Name',
                'P7':'Part / Assembly No.',
                'P9':'Rev. No.',
                'P11':'Number of Parts to be Checked',
                'A13':'Sr. No./ Balloon No.',
                'B13':'Dimension Description',
                'C13':'Dimension Specification',
                'D13':'Inspection Method',
                'E13':'Inspection Frequency',
                'F13':'Class',
                'G13':'Shift',
                'H13':'Inspected By',
                # 'I13':'Observation',
                # 'I14':'1',
                # 'J14':'Time & Date of Inspection',
                # 'K14':'2',
                # 'L14':'Time & Date of Inspection',
                # 'M14':'3',
                # 'N14':'Time & Date of Inspection',
                # 'O14':'4',
                # 'P14':'Time & Date of Inspection',
                # 'Q14':'5',
                # 'R14':'Time & Date of Inspection',
                # 'S13':'Remarks',


            }

            # Helper function to get Excel column letter for a given index
            def get_column_letter(index):
                result = ""
                while index > 0:
                    index, remainder = divmod(index - 1, 26)
                    result = chr(65 + remainder) + result
                return result

            # Start from column I (which is index 9)
            col_index = 9  # I is the 9th column

            # Create observation headers
            for i in range(1, max_observations + 1):
                # Get column letters for observation and date columns
                obs_col = get_column_letter(col_index)
                date_col = get_column_letter(col_index + 1)

                # Add headers
                headers[f'{obs_col}13'] = f'Observation {i}'
                headers[f'{date_col}13'] = 'Time & Date of Inspection'

                # Merge cells
                ws.merge_cells(f'{obs_col}13:{obs_col}14')
                ws.merge_cells(f'{date_col}13:{date_col}14')

                # Set column widths
                ws.column_dimensions[obs_col].width = 15
                ws.column_dimensions[date_col].width = 18

                # Increment column index for next pair
                col_index += 2

            # Add remarks column
            remarks_col = get_column_letter(col_index)
            headers[f'{remarks_col}13'] = 'Remarks'
            ws.merge_cells(f'{remarks_col}13:{remarks_col}14')
            ws.column_dimensions[remarks_col].width = 20

            # Apply font and border styles
            for cell, value in headers.items():
                ws[cell] = value
                ws[cell].font = font_header
                ws[cell].alignment = align_center
                ws[cell].fill=fill
            ws['D1'].font = title_font
            ws['D1'].fill=orange_fill
            # Merge cells
            merge_cells = [
                'A1:B4','C1:O4','A5:B6','A7:B8','A9:B10','A11:B12',
                'C5:D6','C7:D8','C9:D10','C11:D12',
                'P1:Q2','P3:Q4','P5:Q6','P7:Q8','P9:Q10','P11:Q12',
                'R1:S2','R3:S4','R5:S6','R7:S8','R9:S10','R11:S12',
                'A13:A14','B13:B14','C13:C14','D13:D14','E13:E14','F13:F14','G13:G14','H13:H14',
                'E5:O12'
            ]
            for cell_range in merge_cells:
                ws.merge_cells(cell_range)

            # Set column widths
            col_widths = {
                'A': 8,'B': 20, 'C': 15, 'D': 15, 'E': 15,
                'F': 15, 'G': 15, 'H': 15,  }
            for col, width in col_widths.items():
                ws.column_dimensions[col].width = width

            # Populate data for the inspection record
            for rec in self:
                ws['C5']=rec.partner_id.name if rec.partner_id else ''
                ws['C7']=rec.part_id.name if rec.part_id else ''
                ws['C9']=rec.drawing_no if rec.drawing_no else ''
                ws['C11']=rec.batch_quantity if rec.batch_quantity else ''
                ws['R1']=rec.date if rec.date else ''
                # Use the operation's data for this specific worksheet
                ws['R3'] = op_no
                ws['R5'] = op_name
                ws['R7']=rec.part_number if rec.part_number else ''
                ws['R9']=rec.rev_no if rec.rev_no else ''
                ws['R11']=rec.parts_to_be_checked if rec.parts_to_be_checked else ''
                row = 15
                for operation in operations:

                    for ele in operation.op_ids:
                        ws[f'A{row}']=ele.sr_no if ele.sr_no else ''
                        ws[f'B{row}'] = " - ".join(
                            filter(None, [ele.dimension_des.name, ele.dimension_des.symbol])) if ele.dimension_des else ''
                        ws[f'C{row}'] = ele.dimension_spec if ele.dimension_spec else ''
                        ws[f'D{row}'] = " - ".join(filter(None, [ele.inspection_method.gauge.name,
                                                                 ele.inspection_method.gauge_no])) if ele.inspection_method else ''

                        ws[f'E{row}']=ele.inspection_frequency if ele.inspection_frequency else ''
                        ws[f'F{row}']=operation.symbol.symbol if operation.symbol else ''
                        ws[f'G{row}']=ele.shift if ele.shift else ''

                        ws[f'H{row}']=ele.inspected_by.name if ele.inspected_by else ''


                        col_index = 9  # 'G' column (starting point)
                        # Determine which observation numbers should be filled based on frequency
                        active_obs_numbers = []
                        if ele.inspection_frequency:
                            freq = str(ele.inspection_frequency).strip().lower() if ele.inspection_frequency else ""

                            # Case 1: Percentage format
                            if freq.endswith('%'):
                                try:
                                    percentage = float(freq.rstrip('%'))
                                    num_observations = math.ceil(
                                        operation.operation_id.batch_quantity * (percentage / 100))
                                    active_obs_numbers = list(range(1, num_observations + 1))
                                except ValueError:
                                    active_obs_numbers = []

                            # Case 2: Integer interval format
                            else:
                                try:
                                    interval = int(freq)
                                    if interval > 0:
                                        active_obs_numbers = list(
                                            range(1, operation.operation_id.batch_quantity + 1, interval))
                                except ValueError:
                                    active_obs_numbers = []

                        # Fill all observation columns - some will be colored, some not
                        for obs_num in range(1, operation.operation_id.batch_quantity + 1):
                            # Try to find an observation with this number
                            obs = ele.observations.filtered(lambda o: o.observation_number == obs_num)

                            # Check if this observation number should be active/fillable
                            is_active = obs_num in active_obs_numbers

                            # Add the observation value
                            cell = ws.cell(row=row, column=col_index, value=obs.observation_value if obs else '')

                            # Add the date
                            date_cell = ws.cell(row=row, column=col_index + 1,
                                                value=obs.time_date_inspection if obs else '')

                            # Apply orange fill to active observation cells
                            if is_active:
                                cell.fill = orange_fill
                                date_cell.fill = orange_fill

                            col_index += 2  # Move to the next observation column
                        # Add Remarks after the last observation column
                        ws.cell(row=row, column=col_index, value=ele.remarks if ele.remarks else '')

                        row += 1


                cur_row=30
                if cur_row<row:
                    cur_row=row
                max_border = 2 * self.batch_quantity
                for rows in ws.iter_rows(min_row=1, max_row=cur_row, min_col=1, max_col=9+max_border):
                    for cell in rows:
                        cell.alignment = align_center
                        cell.border = border

                # region SignOff Members Footer
                sign_row = cur_row
                ws.merge_cells(f'A{cur_row}:S{cur_row}')

                cur_row += 1
                ws[f'A{cur_row}'] = 'Prepared By'
                ws[f'A{cur_row}'].font = font_header
                ws.merge_cells(f'A{cur_row}:B{cur_row}')
                ws.merge_cells(f'C{cur_row}:E{cur_row}')

                ws[f'F{cur_row}'] = 'Prepared Date'
                ws[f'F{cur_row}'].font = font_header
                ws.merge_cells(f'F{cur_row}:G{cur_row}')
                ws.merge_cells(f'H{cur_row}:L{cur_row}')

                for rec in self:
                    ws[f'C{cur_row}'] = rec.create_uid.name if rec.create_uid else ''
                    ws[f'H{cur_row}'] = rec.create_date if rec.create_date else ''

                ws.row_dimensions[cur_row].height = 18
                cur_row += 1
                ws.merge_cells(f'A{cur_row}:L{cur_row}')
                cur_row += 1
                ws.merge_cells(f'A{cur_row}:L{cur_row}')
                ws[f'A{cur_row}'] = 'Sign OFF'
                ws[f'A{cur_row}'].font = Font(size=18, bold=True)
                ws.row_dimensions[cur_row].height = 25
                cur_row += 1
                ws.merge_cells(f'A{cur_row}:L{cur_row}')
                cur_row += 1
                for cell in ws[cur_row]:
                    ws[f'A{cur_row}'] = 'Member'
                    ws[f'D{cur_row}'] = 'Department'
                    ws[f'F{cur_row}'] = 'Status'
                    ws[f'H{cur_row}'] = 'Date'
                    ws[f'J{cur_row}'] = 'Comments'
                    ws.row_dimensions[cur_row].height = 25
                    ws.merge_cells(f'A{cur_row}:C{cur_row}')
                    ws.merge_cells(f'D{cur_row}:E{cur_row}')
                    ws.merge_cells(f'F{cur_row}:G{cur_row}')
                    ws.merge_cells(f'H{cur_row}:I{cur_row}')
                    ws.merge_cells(f'J{cur_row}:L{cur_row}')
                    cell.font = Font(size=12, bold=True, color='ffffff')
                    cell.fill = PatternFill(start_color='0070C0', end_color='0070C0', fill_type="solid")

                # Listing Managers Id
                for rec in self:
                    for record in rec.iatf_members_ids:
                        name_rec = record.approver_id.name
                        dept_rec = record.department_id.name if record.department_id.name else ''
                        status_rec = record.approval_status.capitalize()
                        date_rec = record.date_approved_rejected if record.date_approved_rejected else ''
                        comment_rec = record.comment if record.comment else ''

                        ws[f'A{cur_row + 1}'] = name_rec
                        ws[f'D{cur_row + 1}'] = dept_rec
                        ws[f'F{cur_row + 1}'] = status_rec
                        ws[f'H{cur_row + 1}'] = date_rec
                        ws[f'J{cur_row + 1}'] = comment_rec
                        ws.merge_cells(f'A{cur_row + 1}:C{cur_row + 1}')
                        ws.merge_cells(f'D{cur_row + 1}:E{cur_row + 1}')
                        ws.merge_cells(f'F{cur_row + 1}:G{cur_row + 1}')
                        ws.merge_cells(f'H{cur_row + 1}:I{cur_row + 1}')
                        ws.merge_cells(f'J{cur_row + 1}:L{cur_row + 1}')

                        status_dict = {'Approved': ['CFFFC3', '000000'], 'Rejected': ['FFCDCD', '000000'],
                                       'Revision': ['AFEFFF', '000000'], 'Pending': ['FDFFCD', '000000']}

                        for state, color in status_dict.items():
                            if state == status_rec:
                                ws[f'F{cur_row + 1}'].fill = PatternFill(start_color=color[0], end_color=color[0],
                                                                         fill_type="solid")
                                ws[f'F{cur_row + 1}'].font = Font(size=12, bold=False, color=color[1])

                        cur_row += 1

                    # ws.merge_cells(f'C{cur_row + 1}:D{cur_row + 1}')
                    # ws.merge_cells(f'E{cur_row + 1}:F{cur_row + 1}')
                    ws.merge_cells(f'A{cur_row + 1}:S{cur_row + 1}')
                    ws.merge_cells(f'M{sign_row + 1}:S{cur_row}')

                    for row_no in ws.iter_rows(min_row=sign_row, max_row=cur_row + 1, min_col=1, max_col=19):
                        for cell in row_no:
                            cell.border = border
                            cell.alignment = align_center
                    # endregion



        # Save workbook to BytesIO
        wb.save(output)
        output.seek(0)

        # Create attachment
        attachment = self.env['ir.attachment'].create({
            'name': 'INPROCESS_INSPECTION.xlsx',
            'type': 'binary',
            'datas': base64.b64encode(output.getvalue()),
            'res_model': 'in.process.inspection.report',
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        })

        # Return download link
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }

        
         
class Inprocess_operations(models.Model):
    _name = 'in.process.operations'
    _description = 'Operations of in process inspection'


    operation_id=fields.Many2one('in.process.inspection.report',string='Operation No.', ondelete='cascade')

    op_no = fields.Many2one(
        'control.plan.process',
        string="Operation No.",
        domain="[('process_id.project_id', '=', parent.project_id)]"
    )

    op_name = fields.Char("Operation Name", related="op_no.process_name", store=True)
    symbol = fields.Many2one('process.flow.class', string="Class",store=True)
    op_ids=fields.One2many('in.process.inspection.report.line','inspection_id',string='Operation Lines')

    @api.onchange('op_no')
    def _onchange_op_no(self):
        """Auto-populate fields based on selected Operation No."""
        if self.op_no and self.operation_id and self.operation_id.project_id:
            # Explicitly set the symbol field from the op_no's char_class
            if self.op_no.char_class:
                self.symbol = self.op_no.char_class.id
            # Search for matching control plans
            control_plans = self.env['control.plan'].search([
                ('process_line_ids.process_step', '=', self.op_no.process_step),
                ('project_id', '=', self.operation_id.project_id.id)
            ])

            if control_plans:
                # Clear existing lines
                self.op_ids = [(5, 0, 0)]
                line_commands = []

                for control_plan in control_plans:
                    matching_processes = control_plan.process_line_ids.filtered(
                        lambda p: p.process_step == self.op_no.process_step
                    )

                    for process in matching_processes:
                        # Get the symbol from the process if possible (backup option)
                        if not self.symbol and process.char_class:
                            self.symbol = process.char_class.id
                        for char in process.process_char_ids:
                            vals = {
                                'ele_no': str(char.char_no) if char.char_no else "",
                                'dimension_des': char.method_description.id if char.method_description else False,
                                 'lower_limit': char.lower_limit,
                'upper_limit': char.upper_limit,
                'uom_id': char.uom_id.id,
                                'inspection_method': char.method_evaluation.id if char.method_evaluation else False,
                                'inspected_by': char.method_inspected_by.id if char.method_inspected_by else False,
                                'inspection_frequency': char.method_sample_freq,

                            }
                            line_commands.append((0, 0, vals))

                if line_commands:
                    self.op_ids = line_commands
                # After creating lines, trigger observation generation
                if self.op_ids and self.operation_id and self.operation_id.batch_quantity:
                    for line in self.op_ids:
                        line._update_observations()

    @api.constrains('op_ids')
    def _check_required_fields(self):
        for rec in self:
            for line in rec.op_ids:
                if not all([line.ele_no, line.dimension_des, line.inspection_method, line.inspected_by]):
                    raise ValidationError(
                        _("Inspection line for element %s is incomplete. Ensure all required fields are populated.") %
                        (line.ele_no or "unknown")
                    )

    @api.model
    def create(self, vals):
        """Ensure updates in Inprocess IR reflect in Control Plan after creation."""
        record = super().create(vals)
        record._sync_to_control_plan()
        return record

    def write(self, vals):
        """Ensure updates in Inprocess IR reflect in Control Plan after writing."""
        res = super().write(vals)
        self._sync_to_control_plan()
        return res

    def _sync_to_control_plan(self):
        """Sync updates from Inprocess IR operations to Control Plan."""
        for rec in self:
            if not rec.op_no or not rec.operation_id or not rec.operation_id.project_id:
                continue

            control_plans = self.env['control.plan'].search([
                ('process_line_ids.process_step', '=', rec.op_no.process_step),
                ('project_id', '=', rec.operation_id.project_id.id)
            ])

            for control_plan in control_plans:
                for process in control_plan.process_line_ids:
                    if process.process_step == rec.op_no.process_step:
                        for char in process.process_char_ids:
                            # Find the matching inspection report line using ele_no and char_no
                            matching_line = rec.op_ids.filtered(lambda l: str(l.ele_no) == str(char.char_no))
                            if matching_line:
                                char.write({
                                    'method_description': matching_line.dimension_des.id,
                                    'lower_limit': matching_line.lower_limit,
                                    'upper_limit': matching_line.upper_limit,
                                    'uom_id': matching_line.uom_id.id,
                                    'method_evaluation': matching_line.inspection_method.id,
                                    'method_inspected_by': matching_line.inspected_by.id,
                                    'method_sample_freq': matching_line.inspection_frequency,
                                })

                        # Check for lines in rec.line_ids that don't have matching characteristics in process_char_ids
                        existing_char_nos = process.process_char_ids.mapped('char_no')
                        for line in rec.op_ids:
                            if str(line.ele_no) not in [str(char_no) for char_no in existing_char_nos]:
                                # Create new characteristic
                                self.env['control.chara.line'].create({
                                    'char_no': line.ele_no,
                                    'chara_id': process.id,
                                    'method_description': line.dimension_des.id,
                                    'lower_limit': line.lower_limit,
                                    'upper_limit': line.upper_limit,
                                    'uom_id': line.uom_id.id,
                                    'method_evaluation': line.inspection_method.id,
                                    'method_inspected_by': line.inspected_by.id,
                                    'method_sample_freq': line.inspection_frequency,
                                })

                    # Check for lines in rec.op_ids that don't have matching characteristics in process_char_ids





class Inprocess_IRLine(models.Model):
    _name = 'in.process.inspection.report.line'
    _description = 'Report of in process inspection line'

    _order = 'sequence'

    inspection_id=fields.Many2one('in.process.operations',string='Inspection Report', ondelete='cascade')
    sequence = fields.Integer(string="Sequence", index=True, default=1)
    sr_no = fields.Integer(string="Sr. No./Balloon No", related='sequence', readonly=True)

    ele_no = fields.Char(string="Element No.")
    dimension_des = fields.Many2one('gdt.symbol', string="Dimension Description", store=True)
    # dimension_spec = fields.Char(string="Dimension Specification")
    lower_limit = fields.Float(string="Lower Limit")
    upper_limit = fields.Float(string="Upper Limit")
    uom_id = fields.Many2one('uom.uom', string="Unit of Measure")

    dimension_spec = fields.Char(
        string="Method Product Spec./ Tolerance",
        compute="_compute_dimension_spec",
    )

    @api.depends('lower_limit', 'upper_limit', 'uom_id')
    def _compute_dimension_spec(self):
        for record in self:
            if record.lower_limit is not None and record.upper_limit is not None and record.uom_id:
                record.dimension_spec = f"{record.lower_limit} - {record.upper_limit} {record.uom_id.name}"
            else:
                record.dimension_spec = ''


    @api.onchange('dimension_spec')
    def _onchange_dimension_spec(self):
        for rec in self:
            if rec.dimension_spec:
                rec.dimension_des = rec.dimension_spec.dimension_des
            else:
                rec.dimension_des = False

    inspection_method = fields.Many2one('gauge.trackingsheet.line', string="Inspection Method", store=True)

    @api.onchange('lower_limit', 'upper_limit', 'uom_id')
    def _onchange_match_gauge(self):
        for record in self:
            if record.lower_limit and record.upper_limit and record.uom_id:
                domain = [
                    ('lower_limit', '<=', record.lower_limit),
                    ('upper_limit', '>=', record.upper_limit),
                    ('uom_id', '=', record.uom_id.id),
                ]

                matched_gauge = self.env['gauge.trackingsheet.line'].search(domain, limit=1)
                record.inspection_method = matched_gauge or False

    inspection_frequency=fields.Char(string="Inspection Frequency")


    shift=fields.Selection([('a', 'A'), ('b', 'B'), ('c', 'C'), ('d', 'D'), ('e', 'E')],string="Shift")
    inspected_by=fields.Many2one('res.users',string="Inspected By")


    remarks = fields.Text(string="Remarks")
    batch_quantity = fields.Integer(related='inspection_id.operation_id.batch_quantity', string="Batch Quantity")

    # Dynamic fields for observations and timestamps
    observations = fields.One2many('in.inspection.observation', 'inspection_line_id', string="Observations")
    observation_status = fields.Selection([
        ('draft', 'Draft'),
        ('auto_filled', 'Auto Filled'),
        ('manually_edited', 'Manually Edited')
    ], string="Observation Status", default='draft')

    def _generate_observations(self):
        """
        Generate observations based on batch quantity and inspection frequency
        Supports formats:
        - Integer (fixed interval): 3 means every 3rd observation
        - Percentage (100%, 70%, etc.): Percentage of total observations
        """
        self.ensure_one()

        if not self.batch_quantity or not self.inspection_frequency:
            return

        # Parse frequency - handle empty or invalid values
        freq = str(self.inspection_frequency).strip().lower() if self.inspection_frequency else ""
        if not freq:
            return

        # First determine which observation numbers should exist
        active_obs_numbers = []

        # Case 1: Percentage (e.g., "100%", "70%")
        if freq.endswith('%'):
            try:
                percentage = float(freq.rstrip('%'))
                num_observations = math.ceil(self.batch_quantity * (percentage / 100))
                active_obs_numbers = list(range(1, num_observations + 1))
            except ValueError:
                # Invalid percentage format
                return

        # Case 2: Integer interval (e.g., "3" means every 3rd observation)
        else:
            try:
                interval = int(freq)
                if interval <= 0:
                    return
                # Generate observations at specified interval
                active_obs_numbers = list(range(1, self.batch_quantity + 1, interval))
            except ValueError:
                # Invalid interval format
                return

        # Now create/update observations based on active_obs_numbers

        # Get existing observations
        existing_obs = {obs.observation_number: obs for obs in self.observations}

        # Create commands for observations
        obs_commands = []

        # Remove observations that shouldn't exist anymore
        for obs_num, obs in existing_obs.items():
            if obs_num not in active_obs_numbers:
                obs_commands.append((2, obs.id, 0))  # Delete command

        # Create or keep observations that should exist
        for obs_num in active_obs_numbers:
            if obs_num in existing_obs:
                # Keep existing observation
                continue
            else:
                # Create new observation
                obs_commands.append((0, 0, {
                    'observation_number': obs_num,
                    'observation_value': '',
                    'is_editable': True
                }))

        # Apply the commands
        if obs_commands:
            self.observations = obs_commands
    def action_button(self):
        return self._generate_observations()
    def _update_observations(self):
        """
        Update observations based on current batch quantity and inspection frequency
        This is the method called from parent models when batch_quantity changes
        """
        for line in self:
            if line.inspection_frequency and line.batch_quantity:
                line._generate_observations()

    @api.onchange('inspection_frequency', 'batch_quantity')
    def _onchange_frequency_batch(self):
        """
        Trigger observation generation when frequency or batch quantity changes
        """
        if self.inspection_frequency and self.batch_quantity:
            self._generate_observations()
            self.observation_status = 'auto_filled'



    def write(self, vals):
        res = super().write(vals)
        if 'batch_quantity' in vals or 'inspection_frequency' in vals:
            for line in self:
                line._generate_observations()
        return res

    @api.depends('inspection_method')
    def _compute_dimension_des(self):
        for record in self:
            if record.inspection_method and record.inspection_method.method_description:
                record.dimension_des = record.inspection_method.method_description
            else:
                record.dimension_des = False


    

    def _resequence_records(self, inspection_id):
        if not inspection_id:
            return
        
        lines = self.search([
            ('inspection_id', '=', inspection_id.id)
        ], order="sequence, id")
        
        for index, line in enumerate(lines, start=1):
            line.write({'sequence': index})
    # @api.constrains('time_date_inspection_1', 'time_date_inspection_2', 'time_date_inspection_3',
    #                 'time_date_inspection_4', 'time_date_inspection_5')
    # def _check_inspection_dates(self):
    #     for record in self:
    #         dates = [
    #             record.time_date_inspection_1,
    #             record.time_date_inspection_2,
    #             record.time_date_inspection_3,
    #             record.time_date_inspection_4,
    #             record.time_date_inspection_5
    #         ]
    #         dates = [d for d in dates if d]  # Remove None values
    #
    #         for i in range(len(dates)-1):
    #             if dates[i] > dates[i+1]:
    #                 raise ValidationError(_("Inspection dates must be in chronological order."))
    #

class InspectionObservation(models.Model):
    _name = 'in.inspection.observation'
    _description = 'Inspection Observations'
    _inherit = "translation.mixin"

    inspection_line_id = fields.Many2one('in.process.inspection.report.line', string="Inspection Line", ondelete="cascade")
    observation_number = fields.Integer(string="Observation No.")
    observation_value = fields.Char(string="Observation",translate=True)
    time_date_inspection = fields.Datetime(string="Time & Date of Inspection")
    # New field to control editability
    is_editable = fields.Boolean(string="Editable", default=False)

    @api.constrains('observation_number')
    def _check_observation_number(self):
        """
        Ensure observation number is within batch quantity
        """
        for rec in self:
            if rec.inspection_line_id and rec.observation_number > rec.inspection_line_id.batch_quantity:
                raise ValidationError(
                    f"Observation number {rec.observation_number} exceeds batch quantity of {rec.inspection_line_id.batch_quantity}")

class ProductTemplate(models.Model):
    _inherit = "product.template"

    drg_no = fields.Char('Drawing Number')


    