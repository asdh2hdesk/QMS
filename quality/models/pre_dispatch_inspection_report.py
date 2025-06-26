from odoo import api, fields, models, _
from datetime import date, datetime
from odoo.exceptions import ValidationError
import base64
from openpyxl import Workbook
from PIL import Image as PILImage, ImageOps
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
import io
from io import BytesIO
import logging
_logger = logging.getLogger(__name__)
class PreDispatch_IR(models.Model):
    _name = 'pre.dispatch.inspection.report'
    _description = 'Report of pre dispatch inspection'
    _inherit = ['iatf.sign.off.members', 'revision.history.mixin']
    _rec_name = 'part_name'  # Added for better record identification
    
    # customer_id=fields.Many2one('res.partner',string='Customer/Supplier Name')
    # part_name=fields.Many2one('product.template',string='Part/Assembly Name')
    # part_no = fields.Char("Part No.", related="part_name.default_code", store=True)
    # drg_no = fields.Char("Drg. No.", related="part_name.drg_no", store=True)

    batch_quantity=fields.Integer("Batch Quantity")
    page_no=fields.Char("Page No.")
    date=fields.Date("Date")

    op_no = fields.Many2one(
        'control.plan.process',
        string="Operation No.",
        domain="[('process_id.project_id', '=', project_id)]"
    )

    op_name = fields.Char("Operation Name", related="op_no.process_name", store=True)


    rev_no=fields.Char("Rev. No.")
    rev_date = fields.Date("Rev. Date")
    parts_to_be_checked=fields.Integer("Number of Parts to be Checked")
    line_ids = fields.One2many('pre.dispatch.inspection.report.line', 'inspection_id', string="Inspection Lines")
    generate_xls_file = fields.Binary(string="Generated File")

    @api.onchange('op_no')
    def _onchange_op_no(self):
        """Auto-populate fields based on selected Operation No."""
        if not self.op_no:
            self.op_name = False
            self.line_ids = [(5, 0, 0)]  # Clear lines if no operation is selected
            return

        # Set operation name
        self.op_name = self.op_no.process_name

        # Validate project_id
        if not self.project_id:
            raise ValidationError(_("Please select a project before choosing an operation."))

        # Search for a single control plan matching the project and process step
        control_plan = self.env['control.plan'].search([
            ('project_id', '=', self.project_id.id),
            ('process_line_ids.process_step', '=', self.op_no.process_step),
        ], limit=1)

        if not control_plan:
            raise ValidationError(
                _("No control plan found for the selected project and operation. Please ensure a control plan exists.")
            )

        # Clear existing lines
        self.line_ids = [(5, 0, 0)]
        line_commands = []

        # Get the matching process for the selected operation
        matching_process = control_plan.process_line_ids.filtered(
            lambda p: p.process_step == self.op_no.process_step
        )

        if not matching_process:
            raise ValidationError(
                _("No matching process found for the selected operation in the control plan. Please check the control plan configuration.")
            )

        # Validate that process has characteristics
        if not matching_process.process_char_ids:
            raise ValidationError(
                _("The selected operation has no characteristics defined in the control plan. Please define characteristics first.")
            )

        # Iterate through characteristics of the matching process
        for char in matching_process.process_char_ids:
            char_no_value = str(char.char_no) if char.char_no else ""

            # Validate required fields
            if not all([char.method_description, char.method_product_display, char.method_evaluation,
                        char.method_inspected_by]):
                _logger.warning(
                    "Incomplete data for char_no %s: Missing required fields", char_no_value)
                raise ValidationError(
                    _("Incomplete data for characteristic %s. Ensure Description, Specification, Evaluation Method, and Inspected By are defined.") % char_no_value
                )

            vals = {
                'ele_no': char_no_value,
                'dimension_des': char.method_description.id if char.method_description else False,
                'lower_limit': char.lower_limit,
                'upper_limit': char.upper_limit,
                'uom_id': char.uom_id.id,  # Now we ensure this is always set
                'inspection_method': char.method_evaluation.id if char.method_evaluation else False,
                'inspected_by': char.method_inspected_by.id if char.method_inspected_by else False,
                'symbol': matching_process.char_class.id if matching_process.char_class else False,
            }
            print(vals)

            # Debug: Log the values to check if they are being set
            _logger.info("Creating line for char_no %s: %s", char_no_value, vals)

            line_commands.append((0, 0, vals))

        # Update line_ids with the new lines
        if line_commands:
            self.line_ids = line_commands
        else:
            raise ValidationError(
                _("No valid inspection lines could be generated. Please ensure characteristics are fully defined in the control plan.")
            )

    @api.constrains('project_id', 'op_no')
    def _check_required_fields(self):
        for record in self:
            if not record.project_id:
                raise ValidationError(_("Project information is required."))

            # Check if line_ids have all required fields
            for line in record.line_ids:
                if not all([line.ele_no, line.dimension_des, line.dimension_spec,
                            line.inspection_method, line.inspected_by]):
                    raise ValidationError(
                        _("Inspection line for element %s is incomplete. Ensure all required fields are populated.") %
                        (line.ele_no or "unknown")
                    )

    @api.onchange('batch_quantity')
    def _onchange_batch_quantity(self):
        """Update observations for all lines when batch quantity changes"""
        if self.line_ids:
            for line in self.line_ids:
                line._update_observations()

    @api.model
    def create(self, vals):
        """Ensure updates in pre dispatch IR reflect in Control Plan."""
        # Validate required fields before creating
        if 'project_id' not in vals or not vals.get('project_id'):
            raise ValidationError(_("Project is required to create an inspection report."))

        record = super().create(vals)
        try:
            record._sync_to_control_plan()
            for line in record.line_ids:
                line._update_observations()
        except Exception as e:
            _logger.error("Error syncing to control plan: %s", str(e))
            raise ValidationError(
                _("Failed to sync with control plan. Please ensure all required data is available. Details: %s") % str(
                    e))
        return record

    def write(self, vals):
        """Ensure updates in pre dispatch IR reflect in Control Plan."""
        res = super().write(vals)
        try:
            self._sync_to_control_plan()
            if 'batch_quantity' in vals:
                for line in self.line_ids:
                    line._update_observations()
        except Exception as e:
            _logger.error("Error syncing to control plan: %s", str(e))
            raise ValidationError(
                _("Failed to sync with control plan. Please ensure all required data is available. Details: %s") % str(
                    e))
        return res

    def _sync_to_control_plan(self):
        """Sync updates from pre dispatch IR to Control Plan."""
        for rec in self:
            if not rec.project_id or not rec.op_no or not rec.op_no.process_step:
                # Skip syncing if critical data is missing
                _logger.warning("Skipping sync to control plan due to missing data: project_id=%s, op_no=%s",
                                rec.project_id, rec.op_no)
                continue

            control_plan = self.env['control.plan'].search([
                ('process_line_ids.process_step', '=', rec.op_no.process_step),
                ('project_id', '=', rec.project_id.id)
            ], limit=1)

            if not control_plan:
                _logger.warning("No control plan found for project_id=%s and process_step=%s",
                                rec.project_id.id, rec.op_no.process_step)
                continue

            for process in control_plan.process_line_ids:
                if process.process_step != rec.op_no.process_step:
                    continue

                # Update existing characteristics
                for char in process.process_char_ids:
                    matching_line = rec.line_ids.filtered(lambda l: str(l.ele_no) == str(char.char_no))[:1]
                    if matching_line:
                        char.write({
                            'method_description': matching_line.dimension_des.id if matching_line.dimension_des else False,
                            'lower_limit': matching_line.lower_limit,
                            'upper_limit': matching_line.upper_limit,
                            'uom_id': matching_line.uom_id.id,
                            'method_evaluation': matching_line.inspection_method.id if matching_line.inspection_method else False,
                            'method_inspected_by': matching_line.inspected_by.id if matching_line.inspected_by else False,
                        })

                # Create new characteristics for unmatched lines
                existing_char_nos = process.process_char_ids.mapped('char_no')
                for line in rec.line_ids:
                    if str(line.ele_no) not in [str(char_no) for char_no in existing_char_nos]:
                        self.env['control.chara.line'].create({
                            'char_no': line.ele_no,
                            'chara_id': process.id,
                            'method_description': line.dimension_des.id if line.dimension_des else False,
                            'lower_limit': line.lower_limit,
                            'upper_limit': line.upper_limit,
                            'uom_id': line.uom_id.id,
                            'method_evaluation': line.inspection_method.id if line.inspection_method else False,
                            'method_inspected_by': line.inspected_by.id if line.inspected_by else False,
                        })

    def generate_xls_report(self):
        # Create workbook and worksheet

        output = BytesIO()
        wb = Workbook()
        ws = wb.active
        
        # Define styles
        border = Border(top=Side(style='thin'), left=Side(style='thin'), 
                        right=Side(style='thin'), bottom=Side(style='thin'))
        font_header = Font(name='Times New Roman', bold=True)
        title_font=Font(size=20, bold=True)
        align_center = Alignment(vertical='center', horizontal='center', wrapText=True)
        
        # Define fill colors
        fill = PatternFill(start_color="e7e7e7", end_color="e7e7e7", fill_type="solid")  # Grey
        orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

        
        if self.env.user.company_id and self.env.user.company_id.logo:

            max_width, max_height = 100, 200  # Set max dimensions
            image_data = base64.b64decode(self.env.user.company_id.logo)
            
            # Open image with PIL
            image = PILImage.open(io.BytesIO(image_data))
            width, height = image.size
            aspect_ratio = width / height
            
            # Resize with aspect ratio
            if width > max_width:
                width = max_width
                height = int(width / aspect_ratio)
            if height > max_height:
                height = max_height
                width = int(height * aspect_ratio)

            # Resize and add padding
            resized_image = image.resize((width, height), PILImage.LANCZOS)
            padding_top, padding_left = 5,5
            resized_image = ImageOps.expand(resized_image, border=(padding_left, padding_top, 0, 0), fill='white')

            img_bytes = io.BytesIO()
            resized_image.save(img_bytes, format='PNG')
            img_bytes.seek(0)
            logo_image = Image(img_bytes)

            ws.add_image(logo_image, 'A1')
        # Determine maximum number of observations dynamically
        max_observations = self.batch_quantity if self.batch_quantity else 5

        # Headers dictionary
        headers = {
            'C1':'Pre - Dispatch Inspection Report',
            'A5':'Customer/Supplier Name',
            'A7':'Part / Assembly Name',
            'A9':'Drawing No.',
            'A11':'Batch Quantity',

            'N1':'Date',
            'N3':'Operation No.',
            'N5':'Operation Name',
            'N7':'Part / Assembly No.',
            'N9':'Rev. No.',
            'N11':'Number of Parts to be Checked',
            'A13':'Sr. No./ Balloon No.',
            'B13':'Dimension Description',
            'C13':'Dimension Specification',
            'D13':'Inspection Method',
            'E13':'Class',
            # 'F13':'Inspection Frequency',
            # 'G13':'Shift',
            'F13':'Inspected By',
            # 'G13':'Observation',
            # 'G14':'1',
            # 'H14':'Time & Date of Inspection',
            # 'I14':'2',
            # 'J14':'Time & Date of Inspection',
            # 'K14':'3',
            # 'L14':'Time & Date of Inspection',
            # 'M14':'4',
            # 'N14':'Time & Date of Inspection',
            # 'O14':'5',
            # 'P14':'Time & Date of Inspection',
            # 'Q13':'Remarks',
            #
            
        }

        # Helper function to get Excel column letter for a given index
        def get_column_letter(index):
            result = ""
            while index > 0:
                index, remainder = divmod(index - 1, 26)
                result = chr(65 + remainder) + result
            return result

        # Start from column I (which is index 9)
        col_index = 7  # I is the 9th column

        # Create observation headers
        for i in range(1, max_observations + 1):
            # Get column letters for observation and date columns
            obs_col = get_column_letter(col_index)
            date_col = get_column_letter(col_index + 1)

            # Add headers
            headers[f'{obs_col}13'] = f'Observation {i}'
            headers[f'{date_col}13'] = 'Time & Date of Inspection'

            # Merge cells
            ws.merge_cells(f'{obs_col}13:{obs_col}14')
            ws.merge_cells(f'{date_col}13:{date_col}14')

            # Set column widths
            ws.column_dimensions[obs_col].width = 15
            ws.column_dimensions[date_col].width = 18

            # Increment column index for next pair
            col_index += 2

        # Add remarks column
        remarks_col = get_column_letter(col_index)
        headers[f'{remarks_col}13'] = 'Remarks'
        ws.merge_cells(f'{remarks_col}13:{remarks_col}14')
        ws.column_dimensions[remarks_col].width = 20
        # Apply font and border styles
        for cell, value in headers.items():
            ws[cell] = value
            ws[cell].font = font_header
            ws[cell].alignment = align_center
            ws[cell].fill=fill
        ws['C1'].font = title_font
       
        # Merge cells
        merge_cells = [
            'A1:B4','C1:M4','A5:B6','A7:B8','A9:B10','A11:B12',
            'C5:D6','C7:D8','C9:D10','C11:D12',
            'N1:O2','N3:O4','N5:O6','N7:O8','N9:O10','N11:O12',
            'P1:Q2','P3:Q4','P5:Q6','P7:Q8','P9:Q10','P11:Q12',
            'A13:A14','B13:B14','C13:C14','D13:D14','E13:E14','F13:F14',
            'E5:M12'
        ]
        for cell_range in merge_cells:
            ws.merge_cells(cell_range)
            
        # Set column widths
        col_widths = {
            'A': 8,'B': 20, 'C': 15, 'D': 20, 'E': 15,
            'F': 15,
        }
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width
            
        # Populate data for the inspection record 
        for rec in self:
            ws['C5']=rec.partner_id.name if rec.partner_id else ''
            ws['C7']=rec.part_name if rec.part_name else ''
            ws['C9']=rec.drawing_no if rec.drawing_no else ''
            ws['C11']=rec.batch_quantity if rec.batch_quantity else ''

            ws['P1']=rec.date if rec.date else ''
            ws['P3']=rec.op_no.id if rec.op_no else ''
            ws['P5']=rec.op_name if rec.op_name else ''
            ws['P7']=rec.part_number if rec.part_number else ''
            ws['P9']=rec.rev_no if rec.rev_no else ''
            ws['P11']=rec.parts_to_be_checked if rec.parts_to_be_checked else ''
            
            row=15
            for ele in rec.line_ids:
                ws[f'A{row}']=ele.sr_no if ele.sr_no else ''
                ws[f'B{row}'] = " - ".join(filter(None, [ele.dimension_des.name, ele.dimension_des.symbol])) if ele.dimension_des else ''
                ws[f'C{row}']=ele.dimension_spec if ele.dimension_spec else ''
                ws[f'D{row}'] = " - ".join(filter(None, [ele.inspection_method.gauge.name,
                                                         ele.inspection_method.gauge_no])) if ele.inspection_method else ''

                ws[f'E{row}']=ele.symbol.symbol if ele.symbol else ''
                ws[f'F{row}']=ele.inspected_by.name if ele.inspected_by else ''

                col_index = 7  # 'G' column (starting point)
                for obs in ele.observations:
                    ws.cell(row=row, column=col_index, value=obs.observation_value)
                    ws.cell(row=row, column=col_index + 1, value=obs.time_date_inspection)
                    col_index += 2  # Move to the next observation column
                # Add Remarks after the last observation column
                ws.cell(row=row, column=col_index, value=ele.remarks if ele.remarks else '')
                row += 1
               
            
        cur_row=30
        if cur_row<row:
            cur_row=row
        max_border = 2 * max_observations
        for rows in ws.iter_rows(min_row=1, max_row=cur_row, min_col=1,  max_col=7+max_border):
            for cell in rows:
                cell.alignment = align_center
                cell.border = border     
            
        # region SignOff Members Footer
        sign_row = cur_row
        blank_rows = get_column_letter(max_border + 7)
        ws.merge_cells(f'A{cur_row}:{blank_rows}{cur_row}')

        cur_row += 1
        ws[f'A{cur_row}'] = 'Prepared By'
        ws[f'A{cur_row}'].font = font_header
        ws.merge_cells(f'A{cur_row}:B{cur_row}')
        ws.merge_cells(f'C{cur_row}:E{cur_row}')

        ws[f'F{cur_row}'] = 'Prepared Date'
        ws[f'F{cur_row}'].font = font_header
        ws.merge_cells(f'F{cur_row}:G{cur_row}')
        ws.merge_cells(f'H{cur_row}:L{cur_row}')

        for rec in self:
            ws[f'C{cur_row}'] = rec.create_uid.name if rec.create_uid else ''
            ws[f'H{cur_row}'] = rec.create_date if rec.create_date else ''

        ws.row_dimensions[cur_row].height = 18
        cur_row += 1
        ws.merge_cells(f'A{cur_row}:L{cur_row}')
        cur_row += 1
        ws.merge_cells(f'A{cur_row}:L{cur_row}')
        ws[f'A{cur_row}'] = 'Sign OFF'
        ws[f'A{cur_row}'].font = Font(size=18, bold=True)
        ws.row_dimensions[cur_row].height = 25
        cur_row += 1
        ws.merge_cells(f'A{cur_row}:L{cur_row}')
        cur_row += 1
        for cell in ws[cur_row]:
            ws[f'A{cur_row}'] = 'Member'
            ws[f'D{cur_row}'] = 'Department'
            ws[f'F{cur_row}'] = 'Status'
            ws[f'H{cur_row}'] = 'Date'
            ws[f'J{cur_row}'] = 'Comments'
            ws.row_dimensions[cur_row].height = 25
            ws.merge_cells(f'A{cur_row}:C{cur_row}')
            ws.merge_cells(f'D{cur_row}:E{cur_row}')
            ws.merge_cells(f'F{cur_row}:G{cur_row}')
            ws.merge_cells(f'H{cur_row}:I{cur_row}')
            ws.merge_cells(f'J{cur_row}:L{cur_row}')
            cell.font = Font(size=12, bold=True, color='ffffff')
            cell.fill = PatternFill(start_color='0070C0', end_color='0070C0', fill_type="solid")

        # Listing Managers Id
        for rec in self:
            for record in rec.iatf_members_ids:
                name_rec = record.approver_id.name
                dept_rec = record.department_id.name if record.department_id.name else ''
                status_rec = record.approval_status.capitalize()
                date_rec = record.date_approved_rejected if record.date_approved_rejected else ''
                comment_rec = record.comment if record.comment else ''

                ws[f'A{cur_row + 1}'] = name_rec
                ws[f'D{cur_row + 1}'] = dept_rec
                ws[f'F{cur_row + 1}'] = status_rec
                ws[f'H{cur_row + 1}'] = date_rec
                ws[f'J{cur_row + 1}'] = comment_rec
                ws.merge_cells(f'A{cur_row + 1}:C{cur_row + 1}')
                ws.merge_cells(f'D{cur_row + 1}:E{cur_row + 1}')
                ws.merge_cells(f'F{cur_row + 1}:G{cur_row + 1}')
                ws.merge_cells(f'H{cur_row + 1}:I{cur_row + 1}')
                ws.merge_cells(f'J{cur_row + 1}:L{cur_row + 1}')

                status_dict = {'Approved': ['CFFFC3', '000000'], 'Rejected': ['FFCDCD', '000000'],
                               'Revision': ['AFEFFF', '000000'], 'Pending': ['FDFFCD', '000000']}

                for state, color in status_dict.items():
                    if state == status_rec:
                        ws[f'F{cur_row + 1}'].fill = PatternFill(start_color=color[0], end_color=color[0],
                                                                 fill_type="solid")
                        ws[f'F{cur_row + 1}'].font = Font(size=12, bold=False, color=color[1])

                cur_row += 1

            # ws.merge_cells(f'C{cur_row + 1}:D{cur_row + 1}')
            # ws.merge_cells(f'E{cur_row + 1}:F{cur_row + 1}')
            ws.merge_cells(f'A{cur_row + 1}:{blank_rows}{cur_row + 1}')
            ws.merge_cells(f'M{sign_row + 1}:{blank_rows}{cur_row}')
            ws.freeze_panes = 'F1'

            for row_no in ws.iter_rows(min_row=sign_row, max_row=cur_row + 1, min_col=1, max_col=max_border+7):
                for cell in row_no:
                    cell.border = border
                    cell.alignment = align_center
            # endregion

       
                
        # Save workbook to BytesIO
        wb.save(output)
        output.seek(0)

        # Create attachment
        attachment = self.env['ir.attachment'].create({
            'name': 'PRE_DISPATCH_INSPECTION.xlsx',
            'type': 'binary',
            'datas': base64.b64encode(output.getvalue()),
            'res_model': 'pre.dispatch.inspection.report',
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        })

        # Return download link
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }

        
         





class PreDispatch_IRLine(models.Model):
    _name = 'pre.dispatch.inspection.report.line'
    _description = 'Report of pre dispatch inspection line'
    _inherit = "translation.mixin"

    _order = 'sequence'

    inspection_id=fields.Many2one('pre.dispatch.inspection.report',string='Inspection Report', ondelete='cascade')
    sequence = fields.Integer(string="Sequence", index=True, default=1)
    sr_no = fields.Integer(string="Sr. No./Balloon No", related='sequence', readonly=True)

    ele_no = fields.Char(string="Element No.",translate=True)
    dimension_des = fields.Many2one('gdt.symbol', string="Dimension Description", store=True)
    # dimension_spec = fields.Char(string="Dimension Specification")
    lower_limit = fields.Float(string="Lower Limit")
    upper_limit = fields.Float(string="Upper Limit")
    uom_id = fields.Many2one('uom.uom', string="Unit of Measure")

    dimension_spec = fields.Char(
        string="Method Product Spec./ Tolerance",
        compute="_compute_dimension_spec",translate=True
    )

    @api.depends('lower_limit', 'upper_limit', 'uom_id')
    def _compute_dimension_spec(self):
        for record in self:
            if record.lower_limit is not None and record.upper_limit is not None and record.uom_id:
                record.dimension_spec = f"{record.lower_limit} - {record.upper_limit} {record.uom_id.name}"
            else:
                record.dimension_spec = ''

    @api.onchange('dimension_spec')
    def _onchange_dimension_spec(self):
        for rec in self:
            if rec.dimension_spec:
                rec.dimension_des = rec.dimension_spec.dimension_des
            else:
                rec.dimension_des = False

    inspection_method = fields.Many2one('gauge.trackingsheet.line', string="Inspection Method", store=True)

    @api.onchange('lower_limit', 'upper_limit', 'uom_id')
    def _onchange_match_gauge(self):
        for record in self:
            if record.lower_limit and record.upper_limit and record.uom_id:
                domain = [
                    ('lower_limit', '<=', record.lower_limit),
                    ('upper_limit', '>=', record.upper_limit),
                    ('uom_id', '=', record.uom_id.id),
                ]

                matched_gauge = self.env['gauge.trackingsheet.line'].search(domain, limit=1)
                record.inspection_method = matched_gauge or False
    symbol = fields.Many2one('process.flow.class', string='Class')
    inspected_by = fields.Many2one('res.users', string="Inspected By")

    # observation_1 = fields.Char(string="Observation 1")
    # time_date_inspection_1 = fields.Datetime(string="Time & Date of Inspection 1")
    #
    # observation_2 = fields.Char(string="Observation 2")
    # time_date_inspection_2 = fields.Datetime(string="Time & Date of Inspection 2")
    #
    # observation_3 = fields.Char(string="Observation 3")
    # time_date_inspection_3 = fields.Datetime(string="Time & Date of Inspection 3")
    #
    # observation_4 = fields.Char(string="Observation 4")
    # time_date_inspection_4 = fields.Datetime(string="Time & Date of Inspection 4")
    #
    # observation_5 = fields.Char(string="Observation 5")
    # time_date_inspection_5 = fields.Datetime(string="Time & Date of Inspection 5")

    remarks = fields.Text(string="Remarks",translate=True)
    batch_quantity = fields.Integer(related='inspection_id.batch_quantity', string="Batch Quantity")

    # Dynamic fields for observations and timestamps
    observations = fields.One2many('pre.inspection.observation', 'inspection_line_id', string="Observations")

    def _update_observations(self):
        """Helper method to update observations based on batch quantity"""
        batch_qty = self.batch_quantity or 0
        _logger.info("Updating observations for line %s (element %s): batch_qty=%s",
                     self.id, self.ele_no, batch_qty)

        # Get current observations
        current_obs = self.observations
        current_count = len(current_obs)

        _logger.info("Current observation count: %s", current_count)

        # Return early if no change is needed
        if current_count == batch_qty:
            _logger.info("No change needed, returning")
            return

        # Create new observations array with commands
        new_observations = []

        # If we need to add more observations
        if batch_qty > current_count:
            _logger.info("Adding %s observations", batch_qty - current_count)
            # Keep existing observations
            for obs in current_obs:
                new_observations.append((4, obs.id))

            # Add new ones
            for i in range(current_count, batch_qty):
                new_observations.append((0, 0, {
                    'observation_number': i + 1,
                    'observation_value': False,
                    'time_date_inspection': False
                }))

        # If we need to reduce observations
        elif batch_qty < current_count and batch_qty > 0:
            _logger.info("Reducing to %s observations", batch_qty)
            # Only keep the first 'batch_qty' observations
            for i, obs in enumerate(current_obs):
                if i < batch_qty:
                    new_observations.append((4, obs.id))
                else:
                    new_observations.append((2, obs.id))  # Delete command (2)

        # If batch_qty is 0, remove all
        elif batch_qty == 0:
            _logger.info("Removing all observations")
            new_observations = [(5, 0, 0)]  # Clear all command (5)

        _logger.info("Setting %s observation commands", len(new_observations))
        self.observations = new_observations
        self.env.cr.commit()  # Force commit to ensure changes are saved

    @api.model
    def create(self, vals):
        res = super(PreDispatch_IRLine, self).create(vals)
        res._update_observations()
        return res

    def write(self, vals):
        res = super().write(vals)
        if 'batch_quantity' in vals:
            for line in self.line_ids:
                line._update_observations()
        return res

    @api.depends('inspection_method')
    def _compute_dimension_des(self):
        for record in self:
            if record.inspection_method and record.inspection_method.method_description:
                record.dimension_des = record.inspection_method.method_description
            else:
                record.dimension_des = False
    


    def _resequence_records(self, inspection_id):
        if not inspection_id:
            return
        
        lines = self.search([
            ('inspection_id', '=', inspection_id.id)
        ], order="sequence, id")
        
        for index, line in enumerate(lines, start=1):
            line.write({'sequence': index})
    # @api.constrains('time_date_inspection_1', 'time_date_inspection_2', 'time_date_inspection_3',
    #                 'time_date_inspection_4', 'time_date_inspection_5')
    # def _check_inspection_dates(self):
    #     for record in self:
    #         dates = [
    #             record.time_date_inspection_1,
    #             record.time_date_inspection_2,
    #             record.time_date_inspection_3,
    #             record.time_date_inspection_4,
    #             record.time_date_inspection_5
    #         ]
    #         dates = [d for d in dates if d]  # Remove None values
    #
    #         for i in range(len(dates)-1):
    #             if dates[i] > dates[i+1]:
    #                 raise ValidationError(_("Inspection dates must be in chronological order."))
    #

class PreInspectionObservation(models.Model):
    _name = 'pre.inspection.observation'
    _description = 'Pre Dispatch Inspection Observations'
    _inherit = "translation.mixin"

    inspection_line_id = fields.Many2one('pre.dispatch.inspection.report.line', string="Inspection Line", ondelete="cascade")
    observation_number = fields.Integer(string="Observation No.")
    observation_value = fields.Char(string="Observation",translate=True)
    time_date_inspection = fields.Datetime(string="Time & Date of Inspection")
class ProductTemplate(models.Model):
    _inherit = "product.template"

    drg_no = fields.Char('Drawing Number')


    