# -*- coding: utf-8 -*-
import base64
import copy
import io
import re
from io import BytesIO

import openpyxl
from PIL import Image as PILImage
from PIL import ImageOps
from odoo.exceptions import ValidationError
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

from odoo import api, fields, models, _


class ProcessFlow(models.Model):
    _name = 'process.flow'
    _description = 'Process Flow Diagram'
    _rec_name = "part_name"
    _inherit = ['iatf.sign.off.members', 'revision.history.mixin', "translation.mixin"]

    # part_no = fields.Char("Part No")
    part_name = fields.Char("Part Name")
    drawing_no = fields.Char("Drawing No")
    document_no = fields.Char("Document No")
    cft_team = fields.Char("CFT Team")
    # internal_part_no = fields.Char("Internal Part No")
    # customer = fields.Many2one('res.partner', "Customer")
    rev_date = fields.Date("Drawing Revision Date")
    rev_no = fields.Char("Drawing Revision No.")
    document_revision_date = fields.Date("Document Revision Date")
    effective_date = fields.Date("Effective Date")

    process_flow_line_ids = fields.One2many(
        comodel_name='process.flow.line',
        inverse_name='process_flow_id',
        string="Process Flow Lines"
    )

    verified_by = fields.Many2one('res.users', 'Verified By')
    approved_by = fields.Many2one('res.users', 'Approved By')
    hr = fields.Many2one('res.users', 'HR')  # HR APPROVED
    design_eng = fields.Many2one('res.users', 'Design Engineering')  # DESIGN
    manf_eng = fields.Many2one('res.users', 'Manufacturing Engineering')  # Engineering
    production = fields.Many2one('res.users', 'Production')  # Manufacturing
    quality = fields.Many2one('res.users', 'Quality')  # Qulity
    top_management_id = fields.Many2one('res.users', 'Top Management')  # Final Approved
    part_development_id = fields.Many2one("part.development.process")
    generate_xls_file = fields.Binary(string="Generated file")

    @api.model_create_multi
    def create(self, vals_list):
        records = super(ProcessFlow, self).create(vals_list)

        for record in records:
            pm_records = self.env["process.group"].search(
                [('project_id', '=', record.project_id.id),
                 ('final_status', '=', 'approved')
                 ])
            # if not pm_records:
            #     raise ValidationError(
            #         _("No Process Matrix is filled in this project first fill Process Matrix before continuing..."))

            for operation in pm_records.process_presentation_ids.sorted(key="operation"):
                operation_element = self.env['process.flow.line'].create({
                    'process_flow_id': record.id,
                    'step': operation.operation,
                    'desc_of_operation': operation.operation_description,
                    # 'class_f': operation.special_characteristics.id if operation.special_characteristics else False,
                    'cumulative_cy_time': operation.cumulative_cy_time,
                })
                # print(operation.operation_lines_ids.product_kpc.name)
                for i in operation.operation_lines_ids:
                    print(i.element_description)
                    self.env['process.flow.operations.line'].create({
                        'operation_element_id': operation_element.id,
                        'element_no': i.element_no,
                        'element_desc': i.element_description,
                        'product_characteristics': i.product_kpc.id,
                        'process_characteristics': i.process_kcc.id,
                        'remarks': i.remarks,
                        'cycle_time_in_min': i.cycle_time,
                        'equip_fixture_ids': [(6, 0, i.equipment_fixture_ids.ids)],
                        'tools_ids': [(6, 0, i.tool_ids.ids)],
                        'pokayoke_ids': [(6, 0, i.pokayoke_ids.ids)],
                    })
        return records

    def update_pfmea(self):
        pfmea_records = self.env["his.pfmea"].search([('project_id', '=', self.project_id.id)])

        for pfmea in pfmea_records:
            for operation in self.process_flow_line_ids.sorted(key="step"):
                # Find existing PFMEA operation
                pfmea_operation = self.env['his.pfmea.operations'].search([
                    ('pfmea_id', '=', pfmea.id),
                    ('operation', '=', operation.step),
                ], limit=1)

                if pfmea_operation:
                    # Update existing PFMEA operation
                    pfmea_operation.write({
                        'operation': operation.step,
                        'stage_pfmea': operation.stage.id,
                        'desc_of_operation_pfmea': operation.desc_of_operation,
                    })
                else:
                    # Create new PFMEA operation
                    pfmea_operation = self.env['his.pfmea.operations'].create({
                        'pfmea_id': pfmea.id,
                        'operation': operation.step,
                        'stage_pfmea': operation.stage.id,
                        'desc_of_operation_pfmea': operation.desc_of_operation,
                    })

                for i in operation.process_op_lines_ids:
                    # Find existing PFMEA operation line
                    pfmea_operation_line = self.env['his.pfmea.operations.line'].search([
                        ('pfmea_operation_id', '=', pfmea_operation.id),
                        ('process_step', '=', i.element_no)
                    ], limit=1)

                    if pfmea_operation_line:
                        # Update existing PFMEA operation line
                        pfmea_operation_line.write({
                            'process_step': i.element_no,
                            'process_desc': i.element_desc,
                        })
                    else:
                        # Create new PFMEA operation line
                        self.env['his.pfmea.operations.line'].create({
                            'pfmea_operation_id': pfmea_operation.id,
                            'process_step': i.element_no,
                            'process_desc': i.element_desc,
                        })

        return True

    def update_pfmea_2(self):
        pfmea_2_records = self.env["process.report"].search([('project_id', '=', self.project_id.id)])

        for pfmea_2 in pfmea_2_records:
            for operation in self.process_flow_line_ids.sorted(key="step"):
                for i in operation.process_op_lines_ids:

                    suggestion = self.env['pfmea.suggestions'].search(
                        [('process_step', '=', operation.desc_of_operation)], limit=1)

                    if not suggestion:
                        suggestion = self.env['pfmea.suggestions'].create({'process_step': operation.desc_of_operation})

                    # Find existing PFMEA operation
                    pfmea_2_operation = self.env['process.operations'].search([
                        ('operations_id', '=', pfmea_2.id),
                        ('issue', '=', i.element_no),
                    ], limit=1)

                    if pfmea_2_operation:
                        # Update existing PFMEA operation
                        pfmea_2_operation.write({
                            'process_step': suggestion.id,
                        })
                    else:
                        pfmea_2_operation.create({
                            'operations_id': pfmea_2.id,
                            'station_no': operation.step,
                            'issue': i.element_no,
                            'process_step': suggestion.id,
                        })

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _("Success"),
                'message': _("PFMEA updated successfully from Process Flow!"),
                'sticky': False,
                'type': 'success',
            }
        }

    def update_control_plan(self):
        control_plan_records = self.env["control.plan"].search([('project_id', '=', self.project_id.id)])

        for control_plan in control_plan_records:
            for op in self.process_flow_line_ids.sorted(key="step"):
                # Find existing PFMEA operation
                control_plan_operation = self.env['control.plan.process'].search([
                    ('process_id', '=', control_plan.id),
                    ('process_step', '=', op.step),
                ], limit=1)

                if control_plan_operation:
                    # Update existing PFMEA operation
                    control_plan_operation.write({
                        'process_step': op.step,
                        'process_name': op.desc_of_operation,
                        'char_class': op.class_f.id,
                    })
                else:
                    # Create new PFMEA operation
                    control_plan_operation = self.env['control.plan.process'].create({
                        'process_id': control_plan.id,
                        'process_step': op.step,
                        'process_name': op.desc_of_operation,
                        'char_class': op.class_f.id,
                    })

                for i in op.process_op_lines_ids:
                    # Find existing PFMEA operation line
                    cp_operation_line = self.env['control.chara.line'].search([
                        ('chara_id', '=', control_plan_operation.id),
                        ('char_no', '=', i.element_no)
                    ], limit=1)

                    if cp_operation_line:
                        # Update existing PFMEA operation line
                        cp_operation_line.write({
                            'char_no': i.element_no,
                        })
                    else:
                        # Create new PFMEA operation line
                        self.env['control.chara.line'].create({
                            'chara_id': control_plan_operation.id,
                            'char_no': i.element_no,
                            'char_product': i.product_characteristics.id,
                            'char_process': i.process_characteristics.id,
                            'method_control': i.control_method,
                            'mc_jig_tool': [(6, 0, i.equip_fixture_ids.ids)],
                        })
        return True

    def action_update_process_matrix(self):
        """ Update process matrix with modified process flow data """
        for record in self:
            process_matrix_main = self.env["process.group"].search([
                ('project_id', '=', record.project_id.id),
                ('final_status', '!=', 'approved')
            ])

            if not process_matrix_main:
                raise ValidationError(_("No In-progress Process Matrix found for this project."))

            for pf_operation in record.process_flow_line_ids:
                process_matrix_op = process_matrix_main.process_presentation_ids.filtered(
                    lambda x: x.operation == pf_operation.step
                )

                if process_matrix_op:
                    process_matrix_op.write({
                        'operation_description': pf_operation.desc_of_operation,
                    })
                else:
                    process_matrix_op = process_matrix_main.process_presentation_ids.create({
                        'process_id': process_matrix_main.id,
                        'operation': pf_operation.step,
                        'operation_description': pf_operation.desc_of_operation,
                    })

                for pf_element in pf_operation.process_op_lines_ids:
                    pm_element = process_matrix_op.operation_lines_ids.filtered(
                        lambda x: x.element_no == pf_element.element_no
                    )

                    if pm_element:
                        pm_element.write({
                            'product_kpc': pf_element.product_characteristics.id,
                            'process_kcc': pf_element.process_characteristics.id,
                            'cycle_time': pf_element.cycle_time_in_min,
                            'equipment_fixture_ids': [(6, 0, pf_element.equip_fixture_ids.ids)],
                        })
                    else:
                        self.env['operation.element.line'].create({
                            'operation_element_id': process_matrix_op.id,
                            'element_no': pf_element.element_no,
                            'product_kpc': pf_element.product_characteristics.id,
                            'process_kcc': pf_element.process_characteristics.id,
                            'cycle_time': pf_element.cycle_time_in_min,
                            'equipment_fixture_ids': [(6, 0, pf_element.equip_fixture_ids.ids)],
                        })

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _("Success"),
                'message': _("Process Matrix updated successfully!"),
                'sticky': False,
                'type': 'success',
            }
        }

    @api.depends('total_cyc_time_in_mins')
    def _compute_total_cyc_time(self):
        for rec in self:
            # list_cycle_times = [ral.cycle_time_in_min for ral in rec.process_flow_line_ids]
            list_cycle_times = [ral.cycle_time_in_min for ral in self.env['process.flow.line'].search(
                [('process_flow_id', '=', rec.id)])]
            total_cycle_time = sum(list_cycle_times)
            # print(list_cycle_times)
            # print(total_cycle_time)
            rec.total_cyc_time_in_mins = total_cycle_time

    @api.depends('total_takt_time_in_mins')
    def _compute_total_takt_time(self):
        for rec in self:
            # list_cycle_times = [ral.cycle_time_in_min for ral in rec.process_flow_line_ids]
            list_cycle_times = [ral.cycle_time_in_min for ral in self.env['process.flow.line'].search(
                [('process_flow_id', '=', rec.id)])]
            if not list_cycle_times:
                rec.total_takt_time_in_mins = 0
            else:
                total_cycle_time = max(list_cycle_times)
                # print(list_cycle_times)
                # print(total_cycle_time)
                rec.total_takt_time_in_mins = total_cycle_time

    def action_generate_excel_report(self):
        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        output = BytesIO()

        data = {
            'F1': 'PROCESS FLOW DIAGRAM',
            'A2': 'Type(SafeLaunch/ Prototype/ PreLaunch/ Production)',
            'A3': 'Part No.',
            'A4': 'Part Name',
            'A5': 'Drawing No.',
            'A6': 'Document Number',
            'K2': 'Customer',
            'K3': 'Drawing revision No.',
            'K4': 'Drawing revision Date',
            'K5': 'Document Revision/Date',
            'K6': 'Effective Date',
            'A7': 'STEPS',
            'B7': 'RECEIPT (ARRIVAL)',
            'C7': 'MANUFACTURING',
            'D7': 'HANDLING',
            'E7': 'OUTSOURCED PROCESS',
            'F7': 'STORE',
            'G7': 'CONTROL / INSPECTION',
            'H7': 'Class',
            'I7': 'DESCRIPTION OF OPERATION',
            'J7': 'PRODUCT CHARACTERISTICS',
            'K7': 'PROCESS CHARACTERISTICS',
            'L7': 'CONTROL METHOD',
            'M7': 'WORK CENTER',
            'N7': 'MACHINE NUMBER',
            'O7': 'CYCLE TIME',
            'P7': 'REMARKS',
            # 'A28': 'Prepared By',
            # 'I28': 'Verified By',
            # 'K28': 'Approved By',
            # 'A29': 'Legends',
            # 'C29': '<SC> Special Characteristics as per Drawing Given by Customer                                '
            #        '<IC>  Special characteristic defined by GTIPL                                                  '
            #        'Special Characteristics for Safe',
        }
        # Fill data into specific cells using key-value pairs
        for cell, value in data.items():
            ws[cell] = value

        thin = Side(border_style='thin', color='000000')
        font_header = Font(name='Times New Roman', size=12, bold=True)
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        align_left = Alignment(vertical='center', horizontal='left', wrapText=True)
        align_right = Alignment(vertical='center', horizontal='right', wrapText=True)
        align_center = Alignment(vertical='center', horizontal='center', wrapText=True)
        symbol_font = Font(name='Times New Roman', size=30, bold=True)

        my_grey = openpyxl.styles.colors.Color(rgb='9c9c9c')
        my_red = openpyxl.styles.colors.Color(rgb='ff9999')
        my_green = openpyxl.styles.colors.Color(rgb='c3f5a9')
        my_blue = openpyxl.styles.colors.Color(rgb='a9d8f5')
        my_yellow = openpyxl.styles.colors.Color(rgb='fff79e')
        my_orange = openpyxl.styles.colors.Color(rgb='ffb47a')
        my_purple = openpyxl.styles.colors.Color(rgb='b59eff')

        my_grey_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_grey)
        my_red_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_red)
        my_green_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_green)
        my_blue_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_blue)
        my_yellow_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_yellow)
        my_orange_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_orange)
        my_purple_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_purple)

        # region merging and formatting cells
        max_col = 16
        max_row = 29
        for row in ws.iter_rows(min_row=1, max_row=8, min_col=1, max_col=max_col):
            for cell in row:
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                alignment = copy.copy(cell.alignment)
                alignment.wrapText = True
                cell.alignment = align_center
                cell.font = font_header

        # Specific Dimension
        ws['F1'].alignment = align_center
        ws['F1'].font = Font(name='Times New Roman', size=24, bold=True)

        # Color list
        color_list = ['A2', 'A3', 'A4', 'A5', 'A6', 'K2', 'K3', 'K4', 'K5', 'K6',
                      'A7', 'B7', 'C7', 'D7', 'E7', 'F7', 'G7', 'H7', 'I7', 'J7',
                      'K7', 'L7', 'M7', 'N7', 'O7', 'P7']
        fill = PatternFill(start_color="4eafed", end_color="4eafed", fill_type="solid")

        ws['H7'].font = Font(name='Times New Roman', size=22, bold=True, color='ffffff')

        # Merging the cells as per standard sheet
        ws.merge_cells('A1:E1')
        ws.merge_cells('F1:P1')
        for mer in range(2, 6 + 1):
            ws.merge_cells(f"A{mer}:F{mer}")
            ws.merge_cells(f"G{mer}:J{mer}")
            ws.merge_cells(f"K{mer}:L{mer}")
            ws.merge_cells(f"M{mer}:P{mer}")
        ws.merge_cells('A7:A8')
        ws.merge_cells('H7:H8')
        ws.merge_cells('I7:I8')
        ws.merge_cells('J7:J8')
        ws.merge_cells('K7:K8')
        ws.merge_cells('L7:L8')
        ws.merge_cells('M7:M8')
        ws.merge_cells('N7:N8')
        ws.merge_cells('O7:O8')
        ws.merge_cells('P7:P8')

        # Dimension of Columns
        ws.column_dimensions['A'].width = 13
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 10
        ws.column_dimensions['I'].width = 40
        ws.column_dimensions['J'].width = 44
        ws.column_dimensions['K'].width = 48
        ws.column_dimensions['L'].width = 54
        ws.column_dimensions['M'].width = 25
        ws.column_dimensions['N'].width = 20
        ws.column_dimensions['O'].width = 15
        ws.column_dimensions['P'].width = 30

        ws.row_dimensions[1].height = 75
        for row_cell in range(2, 6 + 1):
            ws.row_dimensions[row_cell].height = 24
        ws.row_dimensions[7].height = 100
        ws.row_dimensions[8].height = 60

        # Adding logo
        if self.env.user.company_id.logo:
            max_width = 500  # Set your desired maximum width
            max_height = 100  # Set your desired maximum height
            image_data = base64.b64decode(self.env.user.company_id.logo)

            # Open the image using PIL
            image = PILImage.open(io.BytesIO(image_data))
            width, height = image.size
            aspect_ratio = width / height

            if width > max_width:
                width = max_width
                height = int(width / aspect_ratio)

            if height > max_height:
                height = max_height
                width = int(height * aspect_ratio)

            # Resize the image using PIL
            # Add space on the top and left side of the image
            padding_top = 10  # Adjust as needed
            padding_left = 10  # Adjust as needed

            resized_image = image.resize((width, height), PILImage.LANCZOS)
            ImageOps.expand(resized_image, border=(padding_left, padding_top, 0, 0), fill='rgba(0,0,0,0)')
            img_bytes = io.BytesIO()
            resized_image.save(img_bytes, format='PNG')
            img_bytes.seek(0)
            logo_image = Image(img_bytes)
            # logo_image = Image(self.env.user.company_id.logo)
            ws.add_image(logo_image, 'A1')

        mx_row = dflt_rows = 29
        last_row = 9

        symbol_dict = {
            'RECEIPT/ARRIVAL': ['B8', my_purple_fill],
            'MANUFACTURING': ['C8', my_blue_fill],
            'HANDLING': ['D8', my_green_fill],
            'OUTSOURCED PROCESS': ['E8', my_yellow_fill],
            'STORE': ['F8', my_orange_fill],
            'CONTROL/INSPECTION': ['G8', my_red_fill]
        }
        for symbol in self.env['process.flow.stages'].search([]):
            if symbol.name in symbol_dict:
                ws[symbol_dict.get(symbol.name)[0]] = symbol.symbol
                ws[symbol_dict.get(symbol.name)[0]].font = symbol_font
                ws[symbol_dict.get(symbol.name)[0]].fill = symbol_dict.get(symbol.name)[1]

        for rec in self:
            ws['G2'] = rec.doc_type if rec.doc_type else ''
            ws['G3'] = rec.part_number if rec.part_number else ''
            ws['G4'] = rec.part_name if rec.part_name else ''
            ws['G5'] = rec.drawing_no if rec.drawing_no else ''
            ws['G6'] = rec.document_no if rec.document_no else ''
            ws['M2'] = rec.partner_id.name if rec.partner_id else ''
            ws['M3'] = rec.drawing_rev_no if rec.drawing_rev_no else ''
            ws['M4'] = rec.drawing_rev_date.strftime('%d-%m-%Y') if rec.drawing_rev_date else ''
            ws['M5'] = rec.document_revision_date.strftime('%d-%m-%Y') if rec.document_revision_date else ''
            ws['M6'] = rec.effective_date.strftime('%d-%m-%Y') if rec.effective_date else ''
            if rec.process_flow_line_ids:
                for pfl in rec.process_flow_line_ids.sorted(key='step'):
                    ws[f'A{last_row}'] = pfl.step if pfl.step else ''
                    cell_id = None
                    c_list = ['B', 'C', 'D', 'E', 'F', 'G']

                    if pfl.stage.name == 'RECEIPT/ARRIVAL':
                        ws[f'B{last_row}'] = pfl.stage.symbol
                        ws[f'B{last_row}'].font = symbol_font
                        ws[f'B{last_row}'].fill = my_purple_fill
                        cell_id = 'B'
                    elif pfl.stage.name == 'MANUFACTURING':
                        ws[f'C{last_row}'] = pfl.stage.symbol
                        ws[f'C{last_row}'].font = symbol_font
                        ws[f'C{last_row}'].fill = my_blue_fill
                        cell_id = 'C'
                    elif pfl.stage.name == 'HANDLING':
                        ws[f'D{last_row}'] = pfl.stage.symbol
                        ws[f'D{last_row}'].font = symbol_font
                        ws[f'D{last_row}'].fill = my_green_fill
                        cell_id = 'D'
                    elif pfl.stage.name == 'OUTSOURCED PROCESS':
                        ws[f'E{last_row}'] = pfl.stage.symbol
                        ws[f'E{last_row}'].font = symbol_font
                        ws[f'E{last_row}'].fill = my_yellow_fill
                        cell_id = 'E'
                    elif pfl.stage.name == 'STORE':
                        ws[f'F{last_row}'] = pfl.stage.symbol
                        ws[f'F{last_row}'].font = symbol_font
                        ws[f'F{last_row}'].fill = my_orange_fill
                        cell_id = 'F'
                    elif pfl.stage.name == 'CONTROL/INSPECTION':
                        ws[f'G{last_row}'] = pfl.stage.symbol
                        ws[f'G{last_row}'].font = symbol_font
                        ws[f'G{last_row}'].fill = my_red_fill
                        cell_id = 'G'

                    if cell_id in c_list:
                        c_list.remove(cell_id)
                    for i in c_list:
                        ws[f"{i}{last_row}"].fill = my_grey_fill

                    ws[f'A{last_row}'].alignment = align_center
                    ws[f'B{last_row}'].alignment = align_center
                    ws[f'C{last_row}'].alignment = align_center
                    ws[f'D{last_row}'].alignment = align_center
                    ws[f'E{last_row}'].alignment = align_center
                    ws[f'F{last_row}'].alignment = align_center
                    ws[f'G{last_row}'].alignment = align_center
                    ws[f'H{last_row}'].alignment = align_center

                    ws[f'H{last_row}'] = pfl.class_f.symbol if pfl.class_f else ''
                    ws[f'H{last_row}'].font = symbol_font
                    ws[f'I{last_row}'] = pfl.desc_of_operation if pfl.desc_of_operation else ''
                    ws[f'O{last_row}'] = pfl.cumulative_cy_time if pfl.cumulative_cy_time else ''

                    ws[f'J{last_row}'] = ', '.join(
                        [f"{i.element_no}){i.product_characteristics.name}" if i.product_characteristics else '' for i
                         in pfl.process_op_lines_ids])
                    ws[f'K{last_row}'] = ', '.join(
                        [f"{i.element_no}){i.process_characteristics.name}" if i.process_characteristics else '' for i
                         in pfl.process_op_lines_ids])
                    ws[f'L{last_row}'] = ', '.join(
                        [f"{i.element_no}){i.control_method.name}" if i.control_method else '' for i in
                         pfl.process_op_lines_ids])
                    ws[f'M{last_row}'] = ', '.join(
                        [f"{i.element_no}){i.work_center_id.name}" if i.work_center_id else '' for i in
                         pfl.process_op_lines_ids])
                    ws[f'N{last_row}'] = ', '.join(
                        [f"{i.element_no}){i.machine_id.name}" if i.machine_id else '' for i in
                         pfl.process_op_lines_ids])
                    ws[f'P{last_row}'] = ', '.join(
                        [f"{i.element_no}){i.remarks}" if i.remarks else '' for i in pfl.process_op_lines_ids])

                    ws[f'I{last_row}'].font = font_header
                    ws[f'I{last_row}'].alignment = align_center
                    ws[f'J{last_row}'].font = font_header
                    ws[f'J{last_row}'].alignment = align_center
                    ws[f'K{last_row}'].font = font_header
                    ws[f'K{last_row}'].alignment = align_center
                    ws[f'L{last_row}'].font = font_header
                    ws[f'L{last_row}'].alignment = align_center
                    ws[f'M{last_row}'].font = font_header
                    ws[f'M{last_row}'].alignment = align_center
                    ws[f'N{last_row}'].font = font_header
                    ws[f'N{last_row}'].alignment = align_center
                    ws[f'O{last_row}'].font = font_header
                    ws[f'O{last_row}'].alignment = align_center
                    ws[f'P{last_row}'].font = font_header
                    ws[f'P{last_row}'].alignment = align_center

                    last_row += 1
            if last_row > dflt_rows:
                mx_row = last_row

            for i in range(last_row, dflt_rows):
                ws[f"B{i}"].fill = my_grey_fill
                ws[f"C{i}"].fill = my_grey_fill
                ws[f"D{i}"].fill = my_grey_fill
                ws[f"E{i}"].fill = my_grey_fill
                ws[f"F{i}"].fill = my_grey_fill
                ws[f"G{i}"].fill = my_grey_fill

            ws.merge_cells(f'A{mx_row}:B{mx_row}')
            ws.merge_cells(f'A{mx_row + 1}:B{mx_row + 1}')
            ws.merge_cells(f'C{mx_row + 1}:P{mx_row + 1}')
            ws[f'A{mx_row + 1}'] = 'Legends'
            ws[f'A{mx_row + 1}'].font = font_header
            class_characters = ' | '.join(
                [f"{class_.symbol}-{class_.name}" for class_ in self.env['process.flow.class'].search([])])
            ws[f'C{mx_row + 1}'] = class_characters
            # ws[f'C{mx_row + 1}'].font = font_header
            # # ws['C29'] = '<SC> Special Characteristics as per Drawing Given by Customer                                <IC>  Special characteristic defined by GTIPL                                                  Special Characteristics for Safe'
            # ws[f'C{mx_row}'] = rec.prepared_by.name if rec.prepared_by else ''
            # ws[f'C{mx_row}'].font = font_header
            # ws[f'J{mx_row}'] = rec.verified_by.name if rec.verified_by else ''
            # ws[f'J{mx_row}'].font = font_header
            # ws[f'L{mx_row}'] = rec.approved_by.name if rec.approved_by else ''
            # ws[f'L{mx_row}'].font = font_header

        # for row_cell in range(9, mx_row + 1):
        #     ws.row_dimensions[row_cell].height = 85
        for row in ws.iter_rows(min_row=9, max_row=mx_row + 1, min_col=1, max_col=max_col):
            for cell in row:
                cell.border = border
                cell.alignment = align_center

        ws.row_dimensions[mx_row].height = 15
        ws.row_dimensions[mx_row + 1].height = 34
        ws.merge_cells(f'A{mx_row}:P{mx_row}')

        ws[f'A{mx_row + 1}'].alignment = align_center
        ws[f'C{mx_row + 1}'].alignment = align_center

        cur_row = mx_row + 2
        if cur_row < max_row:
            cur_row = max_row + 1

        for color_cell in color_list:
            ws[color_cell].fill = fill
            ws[color_cell].font = Font(name='Times New Roman', size=12, bold=True, color='ffffff')
            list_90 = ['B7', 'C7', 'D7', 'E7', 'F7', 'G7', 'H7']
            if color_cell in list_90:
                ws[color_cell].alignment = Alignment(textRotation=90, horizontal='center', vertical='center',
                                                     wrap_text=True)

        # region SignOff Members Footer
        sign_row = cur_row
        ws.merge_cells(f'A{cur_row}:P{cur_row}')
        cur_row += 1
        mer_end = cur_row
        ws[f'A{cur_row}'] = 'Prepared By'
        ws[f'A{cur_row}'].font = font_header
        ws.merge_cells(f'A{cur_row}:B{cur_row}')
        ws.merge_cells(f'C{cur_row}:F{cur_row}')

        ws[f'G{cur_row}'] = 'Prepared Date'
        ws[f'G{cur_row}'].font = font_header
        ws.merge_cells(f'G{cur_row}:H{cur_row}')

        for rec in self:
            ws[f'C{cur_row}'] = rec.create_uid.name if rec.create_uid else ''
            ws[f'I{cur_row}'] = rec.create_date.strftime('%d-%m-%Y') if rec.create_date else ''

        ws.row_dimensions[cur_row].height = 18
        cur_row += 1
        ws.merge_cells(f'A{cur_row}:I{cur_row}')
        cur_row += 1
        ws.merge_cells(f'A{cur_row}:I{cur_row}')
        ws[f'A{cur_row}'] = 'Sign OFF'
        ws[f'A{cur_row}'].font = Font(size=18, bold=True)
        ws.row_dimensions[cur_row].height = 25
        cur_row += 1
        ws.merge_cells(f'A{cur_row}:I{cur_row}')
        cur_row += 1
        for cell in ws[cur_row]:
            ws[f'A{cur_row}'] = 'Member'
            ws[f'C{cur_row}'] = 'Department'
            ws[f'E{cur_row}'] = 'Status'
            ws[f'G{cur_row}'] = 'Date'
            ws[f'I{cur_row}'] = 'Comments'
            ws.row_dimensions[cur_row].height = 25
            ws.merge_cells(f'A{cur_row}:B{cur_row}')
            ws.merge_cells(f'C{cur_row}:D{cur_row}')
            ws.merge_cells(f'E{cur_row}:F{cur_row}')
            ws.merge_cells(f'G{cur_row}:H{cur_row}')
            cell.font = Font(size=12, bold=True, color='ffffff')
            cell.fill = PatternFill(start_color='0070C0', end_color='0070C0', fill_type="solid")

        # Listing Managers Id
        for rec in self:
            for record in rec.iatf_members_ids:
                name_rec = record.approver_id.name
                dept_rec = record.department_id.name if record.department_id.name else ''
                status_rec = record.approval_status.capitalize()
                date_rec = record.date_approved_rejected.strftime('%d-%m-%Y') if record.date_approved_rejected else ''
                comment_rec = record.comment if record.comment else ''

                ws[f'A{cur_row + 1}'] = name_rec
                ws[f'C{cur_row + 1}'] = dept_rec
                ws[f'E{cur_row + 1}'] = status_rec
                ws[f'G{cur_row + 1}'] = date_rec
                ws[f'I{cur_row + 1}'] = comment_rec
                ws.merge_cells(f'A{cur_row + 1}:B{cur_row + 1}')
                ws.merge_cells(f'C{cur_row + 1}:D{cur_row + 1}')
                ws.merge_cells(f'E{cur_row + 1}:F{cur_row + 1}')
                ws.merge_cells(f'G{cur_row + 1}:H{cur_row + 1}')

                status_dict = {'Approved': ['CFFFC3', '000000'], 'Rejected': ['FFCDCD', '000000'],
                               'Revision': ['AFEFFF', '000000'], 'Pending': ['FDFFCD', '000000']}

                for state, color in status_dict.items():
                    if state == status_rec:
                        ws[f'E{cur_row + 1}'].fill = PatternFill(start_color=color[0], end_color=color[0],
                                                                 fill_type="solid")
                        ws[f'E{cur_row + 1}'].font = Font(size=12, bold=False, color=color[1])

                cur_row += 1

            ws.merge_cells(f'A{cur_row + 1}:P{cur_row + 1}')
            ws.merge_cells(f'J{mer_end}:P{cur_row}')

            for row_no in ws.iter_rows(min_row=sign_row + 1, max_row=cur_row + 1, min_col=1, max_col=16):
                for cell in row_no:
                    cell.border = border
                    cell.alignment = align_center

            # endregion

        # Save the workbook
        wb.save(output)
        output.seek(0)
        self.generate_xls_file = base64.b64encode(output.getvalue()).decode('utf-8')

        return {
            "type": "ir.actions.act_url",
            "target": "self",
            "url": "/web/content?model=process.flow&download=true&field=generate_xls_file&filename={filename}.xlsx&id={pid}".format(
                filename="Process Flow Diagram", pid=self[0].id),
        }


class ProcessFlowLine(models.Model):
    _name = 'process.flow.line'
    _description = 'Process Flow Line'
    _inherit = "translation.mixin"
    _rec_name = 'step'

    process_flow_id = fields.Many2one(
        comodel_name='process.flow',
        string='Process Flow', ondelete='cascade', index=True, copy=False, readonly=True,
    )
    sequence_num = fields.Integer(string="Sequence")
    step = fields.Char("Step/Operation Number", store=True)
    stage = fields.Many2one('process.flow.stages', 'Stage')
    class_f = fields.Many2one('process.flow.class', "Class")
    desc_of_operation = fields.Char("Operation Description",translate=True)
    process_op_lines_ids = fields.One2many(
        comodel_name='process.flow.operations.line',
        inverse_name='operation_element_id',
        string="Operations Lines")
    cumulative_cy_time = fields.Integer("Cumulative Cycle Time(sec)", compute='_compute_total_amount')

    def _recalculate_element_numbers(self):
        """Recalculate element_no for all operations in a process step."""
        for process_line in self:
            elements = process_line.process_op_lines_ids.sorted(lambda e: e.sq_handle)

            seq = 1  # Start element numbering from 1
            updates = []
            for element in elements:
                if process_line.step:
                    # Extract numeric part from step (e.g., "10" from "Op-10")
                    match = re.findall(r'\d+', process_line.step)
                    base_number = match[0] if match else process_line.step  # Default to step if no number

                    new_element_no = f"{base_number}.{seq}"  # Format as "10.1", "10.2"
                    if element.sq_handle != seq or element.element_no != new_element_no:
                        updates.append((element.id, seq, new_element_no))

                seq += 1

            # Batch update
            if updates:
                for element_id, new_seq, new_element_no in updates:
                    self.env['process.flow.operations.line'].browse(element_id).sudo().write({
                        'sq_handle': new_seq,
                        'element_no': new_element_no
                    })

    @api.depends('process_op_lines_ids.cycle_time_in_min')
    def _compute_total_amount(self):
        for record in self:
            record.cumulative_cy_time = sum(record.process_op_lines_ids.mapped('cycle_time_in_min'))

    # @api.depends('sequence_num', 'process_flow_id')
    # def _compute_sequence_number(self):
    #     for line in self.mapped('process_flow_id'):
    #         step = 10
    #         for lines in line.process_flow_line_ids:
    #             lines.step = step
    #             step += 10


class ProcessMatrixOperation(models.Model):
    _name = 'process.flow.operations.line'
    _description = "Process Flow Operations line"
    _inherit = "translation.mixin"

    operation_element_id = fields.Many2one('process.flow.line', 'Operation Element Line')
    op = fields.Char(related='operation_element_id.step', string="Operation", store=True)
    sq_handle = fields.Integer("Sequence no", default=10)
    element_no = fields.Char("Element No.", readonly=True)
    element_desc = fields.Char("Element Description",translate=True)
    product_characteristics = fields.Many2one('product.characteristics', 'Product Characteristics')
    process_characteristics = fields.Many2one('process.characteristics', 'Process Characteristics')
    control_method = fields.Many2one('control.method', 'Control Method')
    part_no = fields.Char(string='In Process Part No.')
    work_center_id = fields.Many2one('mrp.workcenter', string='Work center')
    machine_id = fields.Many2one('maintenance.equipment', string="Machine No.")
    cycle_time_in_min = fields.Integer("Cycle Time(sec)")
    remarks = fields.Char('Remarks',translate=True)
    equip_fixture_ids = fields.Many2many(
        'maintenance.equipment',
        string="Equipment & Fixture",
        domain="[('category_id', 'in', ['Equipment', 'Fixture'])]",
        relation="process_flow_operations_line_equip_fixture_rel"
    )
    tools_ids = fields.Many2many(
        'maintenance.equipment',
        string="Tools",
        domain="[('category_id', 'in', ['Tool'])]",
        relation="process_flow_operations_line_tools_rel"
    )
    pokayoke_ids = fields.Many2many('poka.yoka.line', string="Pokayoke", domain="[('operation', '=', op)]")

    @api.model
    def create(self, vals):
        """Ensure sequence and element_no are correct when adding a new element"""
        record = super(ProcessMatrixOperation, self).create(vals)

        # Update element numbers based on manual step
        record.operation_element_id._recalculate_element_numbers()
        return record

    def write(self, vals):
        """Ensure sequence updates when modified"""
        result = super(ProcessMatrixOperation, self).write(vals)
        for record in self:
            record.operation_element_id._recalculate_element_numbers()
        return result

    @api.constrains('cycle_time_in_min')
    def _check_cycle_time(self):
        for rec in self:
            if rec.cycle_time_in_min:
                if rec.cycle_time_in_min < 0:
                    raise ValidationError(_("Cycle time cannot be less than 0"))
                # if not isinstance(rec.cycle_time_in_min, float):
                #     raise ValidationError(_("Cycle time must be a number!"))

    # @api.depends('sq_handle', 'operation_element_id')
    # def _compute_sequence_number(self):
    #     for line in self.mapped('operation_element_id'):
    #         element_no = 1
    #         for lines in line.process_op_lines_ids:
    #             lines.element_no = element_no
    #             element_no += 1


class ProcessFlowClass(models.Model):
    _name = 'process.flow.class'
    _description = 'Process Flow Class'
    # _inherit = "translation.mixin"

    name = fields.Char('Special characteristic', required=True,translate=True)
    symbol = fields.Char('Symbol', required=True)


class ProcessFlowStages(models.Model):
    _name = 'process.flow.stages'
    _description = 'Process Flow Stages'
    # _inherit = "translation.mixin"

    name = fields.Char('Stage', required=True,translate=True)
    symbol = fields.Char('Symbol', required=True)


class MachineMasterList(models.Model):
    _name = 'machine.master.list'

    id = fields.Char("Machine Name")


class ProductCharacteristics(models.Model):
    _name = 'product.characteristics'
    _inherit = "translation.mixin"

    name = fields.Char("Product Characteristics",translate=True)


class ProcessCharacteristics(models.Model):
    _name = 'process.characteristics'
    _inherit = "translation.mixin"

    name = fields.Char("Process Characteristics",translate=True)


class ControlMethod(models.Model):
    _name = 'control.method'
    _inherit = "translation.mixin"

    name = fields.Char("Control Method",translate=True)
