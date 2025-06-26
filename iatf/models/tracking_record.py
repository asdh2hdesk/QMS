from odoo import fields, models
import openpyxl
from io import BytesIO
import io
from openpyxl import Workbook
import base64
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.drawing.image import Image
import copy
from PIL import ImageOps
from PIL import Image as PILImage
from bs4 import BeautifulSoup


class TrackingRecord(models.Model):
    _name = 'tracking.record'
    _inherit = ['iatf.sign.off.members', 'translation.mixin']

    client_id = fields.Many2one('res.partner', 'Client')
    design = fields.Char('Design No.')
    product_id = fields.Many2one('product.template', 'Description')
    default_code = fields.Char('Product Code', related='product_id.default_code')
    comment = fields.Char('Comment',translate=True)
    op_no = fields.Char('OP. NO.')
    fdm = fields.Char('F.D.M')
    operationg_mc = fields.Char('OPERATING M/C',translate=True)
    tracking_line_ids = fields.One2many(
        comodel_name='tracking.record.line',
        inverse_name='tracking_id',
        string="Tracking Line"
    )
    # state = fields.Selection([
    #     ('draft', 'Draft'),
    #     ('hr_approve', 'HR Approval'),
    #     ('design', 'Design'),
    #     ('engineering', 'Engineering'),
    #     ('manufacturing', 'Manufacturing'),
    #     ('quality', 'Quality'),
    #     ('top', 'Top Management'),
    #     ('final_approved', 'Final Approved')
    # ], string='Status', default='draft')
    hr = fields.Many2one('res.users', 'HR')  # HR APPROVED
    design_eng = fields.Many2one('res.users', 'Design Engineering')  # DESIGN
    manf_eng = fields.Many2one('res.users', 'Manufacturing Engineering')  # Engineering
    production = fields.Many2one('res.users', 'Production')  # Manufacturing
    quality = fields.Many2one('res.users', 'Quality')  # Qulity
    top_management_id = fields.Many2one('res.users', 'Top Management')  # Final Approved
    part_development_id = fields.Many2one("part.development.process")
    generate_xls_file = fields.Binary(string="Generated file")

    # state = fields.Selection([
    #                         ('draft', 'Draft'),
    #                         ('confirm', 'Confirmed'),
    #                         ], string='Status', default='draft')
    document_name = fields.Char(string='Document #')

    # approve_department_id = fields.Many2one('hr.department', string='Departments Approvals')
    # # document_pro_id = fields.Many2one("xf.doc.approval.document.package", string="Document #")
    # approve_department_ids = fields.Many2many('hr.department', 'ts3', string='Departments Approvals')
    # approve_by_department_ids = fields.Many2many('hr.department', string='Departments Approved By')
    # manager_id = fields.Many2one('hr.employee', string='Approval Manager')
    # manager_ids = fields.Many2many('hr.employee', 'ts4', string='Approval Managers')
    # approvaed_manager_ids = fields.Many2many('hr.employee', string='Managers Approved By')

    # is_hr_approved = fields.Boolean()
    # is_design_approved = fields.Boolean()
    # is_eng_approved = fields.Boolean()
    # is_manu_approved = fields.Boolean()
    # is_sales_approved = fields.Boolean()
    # is_qc_approved = fields.Boolean()
    # is_maintenance_approved = fields.Boolean()
    # is_marketing_approved = fields.Boolean()
    # is_pm_approved = fields.Boolean()
    # is_management_approved = fields.Boolean()
    # is_approved_approved = fields.Boolean()
    #
    # hr_check_department = fields.Boolean(compute='_check_department')
    # des_check_department = fields.Boolean(compute='_check_department')
    # eng_check_department = fields.Boolean(compute='_check_department')
    # manu_check_department = fields.Boolean(compute='_check_department')
    # sale_check_department = fields.Boolean(compute='_check_department')
    # qc_check_department = fields.Boolean(compute='_check_department')
    # mainte_check_department = fields.Boolean(compute='_check_department')
    # marketing_check_department = fields.Boolean(compute='_check_department')
    # pm_check_department = fields.Boolean(compute='_check_department')
    # managment_check_department = fields.Boolean(compute='_check_department')
    #
    # link = fields.Char()
    #
    # def actipn_confirmed(self):
    #     self.sent_for_approval()
    #     self.state = 'confirm'
    #
    # def sent_for_approval(self):
    #     mail_template = self.env.ref('iatf.mom_document_approval_mail_template')
    #     model_id = self.env['ir.model'].sudo().search([('model', '=', self._name)])
    #     mail_template.write({'model_id' : model_id.id})
    #
    #     web_base_url = self.env['ir.config_parameter'].sudo().get_param('web.base.url')
    #     action_id = self.env.ref('iatf.action_tracking_record_view', raise_if_not_found=False)
    #     link = """{}/web#id={}&view_type=form&model=self._name&action={}""".format(web_base_url,self.id,action_id.id)
    #     self.link = link
    #
    #     user_ids = self.env['hr.employee'].sudo().search([('department_id', 'in', self.approve_department_ids.ids)]).mapped('user_id')
    #     all_admin = ''
    #     for user in user_ids:
    #         if all_admin:
    #             all_admin += ',' + str(user.login)
    #         else:
    #             all_admin = str(user.login)
    #     mail_template.write({
    #             'email_to': all_admin,
    #     })
    #     mail_template.send_mail(self.id, force_send=True)
    #
    # def _check_department(self):
    #     for rec in self:
    #         if self.env.user.employee_id.department_id.id in rec.approve_department_ids.ids and self.env.user.employee_id.department_id.department_type == 'hr':
    #             if rec.state != 'draft':
    #                 rec.hr_check_department = True
    #             else:
    #                 rec.hr_check_department = False
    #         else:
    #             rec.hr_check_department = False
    #         if self.env.user.employee_id.department_id.id in rec.approve_department_ids.ids and self.env.user.employee_id.department_id.department_type == 'design':
    #             if rec.state != 'draft':
    #                 rec.des_check_department = True
    #             else:
    #                 rec.des_check_department = False
    #         else:
    #             rec.des_check_department = False
    #         if self.env.user.employee_id.department_id.id in rec.approve_department_ids.ids and self.env.user.employee_id.department_id.department_type == 'eng':
    #             if rec.state != 'draft':
    #                 rec.eng_check_department = True
    #             else:
    #                 rec.eng_check_department = False
    #         else:
    #             rec.eng_check_department = False
    #         if self.env.user.employee_id.department_id.id in rec.approve_department_ids.ids and self.env.user.employee_id.department_id.department_type == 'manu':
    #             if rec.state != 'draft':
    #                 rec.manu_check_department = True
    #             else:
    #                 rec.manu_check_department = False
    #         else:
    #             rec.manu_check_department = False
    #         if self.env.user.employee_id.department_id.id in rec.approve_department_ids.ids and self.env.user.employee_id.department_id.department_type == 'sales':
    #             if rec.state != 'draft':
    #                 rec.sale_check_department = True
    #             else:
    #                 rec.sale_check_department = False
    #         else:
    #             rec.sale_check_department = False
    #         if self.env.user.employee_id.department_id.id in rec.approve_department_ids.ids and self.env.user.employee_id.department_id.department_type == 'qc':
    #             if rec.state != 'draft':
    #                 rec.qc_check_department = True
    #             else:
    #                 rec.qc_check_department = False
    #         else:
    #             rec.qc_check_department = False
    #         if self.env.user.employee_id.department_id.id in rec.approve_department_ids.ids and self.env.user.employee_id.department_id.department_type == 'maintenance':
    #             if rec.state != 'draft':
    #                 rec.mainte_check_department = True
    #             else:
    #                 rec.mainte_check_department = False
    #         else:
    #             rec.mainte_check_department = False
    #         if self.env.user.employee_id.department_id.id in rec.approve_department_ids.ids and self.env.user.employee_id.department_id.department_type == 'marketing':
    #             if rec.state != 'draft':
    #                 rec.marketing_check_department = True
    #             else:
    #                 rec.marketing_check_department = False
    #         else:
    #             rec.marketing_check_department = False
    #         if self.env.user.employee_id.department_id.id in rec.approve_department_ids.ids and self.env.user.employee_id.department_id.department_type == 'pm':
    #             if rec.state != 'draft':
    #                 rec.pm_check_department = True
    #             else:
    #                 rec.pm_check_department = False
    #         else:
    #             rec.pm_check_department = False
    #         if self.env.user.employee_id.department_id.id in rec.approve_department_ids.ids and self.env.user.employee_id.department_id.department_type == 'management':
    #             if rec.state != 'draft':
    #                 rec.managment_check_department = True
    #             else:
    #                 rec.managment_check_department = False
    #         else:
    #             rec.managment_check_department = False
    #
    #
    #
    # hr_approval_date = fields.Datetime('Hr Approval Date')
    # design_approval_date = fields.Datetime('Design Approval Date')
    # eng_approval_date = fields.Datetime('Eng. Approval Date')
    # manu_approval_date = fields.Datetime('Production Approval Date')
    # sale_approval_date = fields.Datetime('Sales Approval Date')
    # qc_approval_date = fields.Datetime('QC Approval Date')
    # main_approval_date = fields.Datetime('Maintenance Approval Date')
    # mark_approval_date = fields.Datetime('Marketing Approval Date')
    # pm_approval_date = fields.Datetime('Program Management Approval Date')
    # maneg_approval_date = fields.Datetime('Top Management Approval Date')
    #
    # def hr_approval(self):
    #     self.hr_approval_date = datetime.now()
    #     self.is_hr_approved = True
    #     self.approve_by_department_ids = [(4, self.env.user.employee_id.department_id.id)]
    #     self.approvaed_manager_ids = [(4, self.env.user.employee_id.department_id.manager_id.id)]
    # def design_approval(self):
    #     self.design_approval_date = datetime.now()
    #     self.is_design_approved = True
    #     self.approve_by_department_ids = [(4, self.env.user.employee_id.department_id.id)]
    #     self.approvaed_manager_ids = [(4, self.env.user.employee_id.department_id.manager_id.id)]
    # def eng_approval(self):
    #     self.eng_approval_date = datetime.now()
    #     self.is_eng_approved = True
    #     self.approve_by_department_ids = [(4, self.env.user.employee_id.department_id.id)]
    #     self.approvaed_manager_ids = [(4, self.env.user.employee_id.department_id.manager_id.id)]
    # def manu_approval(self):
    #     self.manu_approval_date = datetime.now()
    #     self.is_manu_approved = True
    #     self.approve_by_department_ids = [(4, self.env.user.employee_id.department_id.id)]
    #     self.approvaed_manager_ids = [(4, self.env.user.employee_id.department_id.manager_id.id)]
    # def sales_approval(self):
    #     self.sale_approval_date = datetime.now()
    #     self.is_sales_approved = True
    #     self.approve_by_department_ids = [(4, self.env.user.employee_id.department_id.id)]
    #     self.approvaed_manager_ids = [(4, self.env.user.employee_id.department_id.manager_id.id)]
    # def qc_approval(self):
    #     self.qc_approval_date = datetime.now()
    #     self.is_qc_approved = True
    #     self.approve_by_department_ids = [(4, self.env.user.employee_id.department_id.id)]
    #     self.approvaed_manager_ids = [(4, self.env.user.employee_id.department_id.manager_id.id)]
    # def mainte_approval(self):
    #     self.main_approval_date = datetime.now()
    #     self.is_maintenance_approved = True
    #     self.approve_by_department_ids = [(4, self.env.user.employee_id.department_id.id)]
    #     self.approvaed_manager_ids = [(4, self.env.user.employee_id.department_id.manager_id.id)]
    # def marketing_approval(self):
    #     self.mark_approval_date = datetime.now()
    #     self.is_marketing_approved = True
    #     self.approve_by_department_ids = [(4, self.env.user.employee_id.department_id.id)]
    #     self.approvaed_manager_ids = [(4, self.env.user.employee_id.department_id.manager_id.id)]
    # def pm_approval(self):
    #     self.pm_approval_date = datetime.now()
    #     self.is_pm_approved = True
    #     self.approve_by_department_ids = [(4, self.env.user.employee_id.department_id.id)]
    #     self.approvaed_manager_ids = [(4, self.env.user.employee_id.department_id.manager_id.id)]
    # def managment_approval(self):
    #     self.maneg_approval_date = datetime.now()
    #     self.is_management_approved = True
    #     self.approve_by_department_ids = [(4, self.env.user.employee_id.department_id.id)]
    #     self.approvaed_manager_ids = [(4, self.env.user.employee_id.department_id.manager_id.id)]
    #
    #
    # def button_send_approval(self):
    #     for rec in self:
    #         if rec.hr:
    #             user_email = rec.hr.login
    #             mail_temp = self.env.ref('iatf.sending_mail_template_risk_assessment')
    #             if mail_temp:
    #                 format_name = rec.format_id.name
    #                 cust_name = rec.cust_id.name
    #                 subject = f" Request for approving Risk Assessment {format_name} for {cust_name}"
    #                 if user_email:
    #                     body = """
    #                     <div>
    #                         <p>Dear Recipient  """ + str(user_email) + """,
    #                             <br/><br/>
    #                             Please, kindly Request To accept and approved Sending Approval for the Tracking Record.
    #                         <br></br>
    #                         Thank you.
    #                     <br/>
    #                     <br/>
    #                     <div>"""
    #                     mail_temp.send_mail(self.id, email_values={
    #                         'email_from': self.env.user.login,
    #                         'email_to': user_email,
    #                         'subject': subject,
    #                         'body_html': body,
    #                     }, force_send=True)
    #
    #         rec.write({'state': 'hr_approve'})
    #         rec.part_development_id.cft_id = rec.id
    #
    # def button_hr_approval(self):
    #     for rec in self:
    #         if rec.hr:
    #             user_email = rec.hr.login
    #             mail_temp = self.env.ref('iatf.sending_mail_template_risk_assessment')
    #             if mail_temp:
    #                 format_name = rec.format_id.name
    #                 cust_name = rec.cust_id.name
    #                 subject = f" Request for approving Risk Assessment {format_name} for {cust_name}"
    #                 if user_email:
    #                     body = """
    #                     <div>
    #                         <p>Dear Recipient  """ + str(user_email) + """,
    #                             <br/><br/>
    #                             Please, kindly Request To accept and approved Sending Approval for the Tracking Record.
    #                         <br></br>
    #                         Thank you.
    #                     <br/>
    #                     <br/>
    #                     <div>"""
    #                     mail_temp.send_mail(self.id, email_values={
    #                         'email_from': self.env.user.login,
    #                         'email_to': user_email,
    #                         'subject': subject,
    #                         'body_html': body,
    #                     }, force_send=True)
    #         rec.write({'state': 'design'})
    #
    # def button_design(self):
    #     for rec in self:
    #         if rec.design_eng:
    #             user_email = rec.design_eng.login
    #             mail_temp = self.env.ref('iatf.sending_mail_template_risk_assessment')
    #             if mail_temp:
    #                 format_name = rec.format_id.name
    #                 cust_name = rec.cust_id.name
    #                 subject = f" Request for approving Risk Assessment {format_name} for {cust_name}"
    #                 if user_email:
    #                     body = """
    #                     <div>
    #                         <p>Dear Recipient  """ + str(user_email) + """,
    #                             <br/><br/>
    #                             Please, kindly Request To accept and approved Sending Approval for the Tracking Record.
    #                         <br></br>
    #                         Thank you.
    #                     <br/>
    #                     <br/>
    #                     <div>"""
    #                     mail_temp.send_mail(self.id, email_values={
    #                         'email_from': self.env.user.login,
    #                         'email_to': user_email,
    #                         'subject': subject,
    #                         'body_html': body,
    #                     }, force_send=True)
    #         rec.write({'state': 'engineering'})
    #
    # def button_engineering(self):
    #     for rec in self:
    #         if rec.manf_eng:
    #             user_email = rec.manf_eng.login
    #             mail_temp = self.env.ref('iatf.sending_mail_template_risk_assessment')
    #             if mail_temp:
    #                 format_name = rec.format_id.name
    #                 cust_name = rec.cust_id.name
    #                 subject = f" Request for approving Risk Assessment {format_name} for {cust_name}"
    #                 if user_email:
    #                     body = """
    #                     <div>
    #                         <p>Dear Recipient  """ + str(user_email) + """,
    #                             <br/><br/>
    #                             Please, kindly Request To accept and approved Sending Approval for the Tracking Record.
    #                         <br></br>
    #                         Thank you.
    #                     <br/>
    #                     <br/>
    #                     <div>"""
    #                     mail_temp.send_mail(self.id, email_values={
    #                         'email_from': self.env.user.login,
    #                         'email_to': user_email,
    #                         'subject': subject,
    #                         'body_html': body,
    #                     }, force_send=True)
    #         rec.write({'state': 'manufacturing'})
    #
    # def button_manufacturing(self):
    #     for rec in self:
    #         if rec.production:
    #             user_email = rec.production.login
    #             mail_temp = self.env.ref('iatf.sending_mail_template_risk_assessment')
    #             if mail_temp:
    #                 format_name = rec.format_id.name
    #                 cust_name = rec.cust_id.name
    #                 subject = f" Request for approving Risk Assessment {format_name} for {cust_name}"
    #                 if user_email:
    #                     body = """
    #                     <div>
    #                         <p>Dear Recipient  """ + str(user_email) + """,
    #                             <br/><br/>
    #                             Please, kindly Request To accept and approved Sending Approval for the Tracking Record.
    #                         <br></br>
    #                         Thank you.
    #                     <br/>
    #                     <br/>
    #                     <div>"""
    #                     mail_temp.send_mail(self.id, email_values={
    #                         'email_from': self.env.user.login,
    #                         'email_to': user_email,
    #                         'subject': subject,
    #                         'body_html': body,
    #                     }, force_send=True)
    #         rec.write({'state': 'quality'})
    #
    # def button_quality_test_done(self):
    #     for rec in self:
    #         if rec.quality:
    #             user_email = rec.quality.login
    #             mail_temp = self.env.ref('iatf.sending_mail_template_risk_assessment')
    #             if mail_temp:
    #                 format_name = rec.format_id.name
    #                 cust_name = rec.cust_id.name
    #                 subject = f" Request for approving Risk Assessment {format_name} for {cust_name}"
    #                 if user_email:
    #                     body = """
    #                     <div>
    #                         <p>Dear Recipient  """ + str(user_email) + """,
    #                             <br/><br/>
    #                             Please, kindly Request To accept and approved Sending Approval for the Tracking Record.
    #                         <br></br>
    #                         Thank you.
    #                     <br/>
    #                     <br/>
    #                     <div>"""
    #                     mail_temp.send_mail(self.id, email_values={
    #                         'email_from': self.env.user.login,
    #                         'email_to': user_email,
    #                         'subject': subject,
    #                         'body_html': body,
    #                     }, force_send=True)
    #         rec.write({'state': 'top'})
    #
    # def button_top_managment_final_approved(self):
    #     for rec in self:
    #         if rec.top_management_id:
    #             user_email = rec.top_management_id.login
    #             mail_temp = self.env.ref('iatf.sending_mail_template_risk_assessment')
    #             if mail_temp:
    #                 format_name = rec.format_id.name
    #                 cust_name = rec.cust_id.name
    #                 subject = f" Request for approving Risk Assessment {format_name} for {cust_name}"
    #                 if user_email:
    #                     body = """
    #                     <div>
    #                         <p>Dear Recipient  """ + str(user_email) + """,
    #                             <br/><br/>
    #                             Please, kindly Request To accept and approved Sending Approval for the Tracking Record.
    #                         <br></br>
    #                         Thank you.
    #                     <br/>
    #                     <br/>
    #                     <div>"""
    #                     mail_temp.send_mail(self.id, email_values={
    #                         'email_from': self.env.user.login,
    #                         'email_to': user_email,
    #                         'subject': subject,
    #                         'body_html': body,
    #                     }, force_send=True)
    #         rec.write({'state': 'final_approved'})

    def generate_excel_sheet(self):
        output = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "AQP"

        max_row = 24
        max_col = 14
        start_row = 5

        thin = Side(border_style='thin', color='000000')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        align_center = Alignment(vertical='center', horizontal='center', wrapText=True)
        font_all = Font(name='Arial', size=12, bold=False)
        font_main_header = Font(name='Arial', size=22, bold=True)
        font_header = Font(name='Arial', size=12, bold=True)

        data = {
            'A1': 'TOOL TRACKING RECORD',
            'B2': 'CLIENT',
            'C2': 'DESIGN NO.',
            'E2': 'PART DESCRIPTION  ',
            'F2': 'OP No.',
            'G2': 'F.D.M',
            'H2': 'Operating Machine',
            'I2': 'COMMENT',
            'A4': 'FIXTURE/GAUGE NO.',
            'B4': 'DESCRIPTION',
            'E4': 'ITEM CODE',
            'F4': 'ITEM DESCRIPTION',
            'G4': 'REQUEST NO.',
            'H4': 'ISSUE DATE',
            'I4': 'QTY',
            'J4': 'EXPEXTED DATE',
            'K4': 'COST',
            'L4': 'OPERATION NUMBER',
            'M4': 'OPERATING MACHINE',
            'N4': 'NOTES/ATTACHMENTS',
        }

        for cell, value in data.items():
            ws[cell] = value
        for row in ws.iter_rows(min_row=1, max_row=4, min_col=1, max_col=max_col):
            for cell in row:
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                cell.alignment = align_center
                cell.font = font_header

        cell_ranges_to_merge = [
            'A1:N1',
            'A2:A3', 'C2:D2', 'I2:N2',
            'C3:D3', 'I3:N3',
            'B4:D4',
        ]

        for cell_range in cell_ranges_to_merge:
            ws.merge_cells(cell_range)

        for i in range(start_row, max_row + 1):
            ws.merge_cells(f"B{i}:D{i}")

        ws['A1'].font = font_main_header
        ws['A1'].alignment = align_center

        ws.row_dimensions[1].height = 35
        ws.row_dimensions[2].height = 48
        ws.row_dimensions[3].height = 40
        ws.row_dimensions[4].height = 60

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 22
        ws.column_dimensions['F'].width = 24
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 14
        ws.column_dimensions['J'].width = 15
        ws.column_dimensions['K'].width = 10
        ws.column_dimensions['L'].width = 15
        ws.column_dimensions['M'].width = 15
        ws.column_dimensions['N'].width = 28

        curr_row = start_row

        # region Adding Logo
        if self.env.user.company_id.logo:
            max_width = 250  # Set your desired maximum width
            max_height = 100  # Set your desired maximum height
            image_data = base64.b64decode(self.env.user.company_id.logo)

            # Open the image using PIL
            image = PILImage.open(io.BytesIO(image_data))
            width, height = image.size
            aspect_ratio = width / height

            if width > max_width:
                width = max_width
                height = int(width / aspect_ratio)

            if height > max_height:
                height = max_height
                width = int(height * aspect_ratio)

            # Resize the image using PIL
            # Add space on the top and left side of the image
            padding_top = 10  # Adjust as needed
            padding_left = 10  # Adjust as needed

            resized_image = image.resize((width, height), PILImage.Resampling.LANCZOS)
            ImageOps.expand(resized_image, border=(padding_left, padding_top, 0, 0), fill='rgba(0,0,0,0)')
            img_bytes = io.BytesIO()
            resized_image.save(img_bytes, format='PNG')
            img_bytes.seek(0)
            logo_image = Image(img_bytes)
            # logo_image = Image(self.env.user.company_id.logo)
            ws.add_image(logo_image, 'A2')
        # endregion

        # region Adding data
        for rec in self:
            ws['B3'] = rec.client_id.name if rec.client_id else ''
            ws['C3'] = rec.design if rec.design else ''
            ws['E3'] = rec.product_id.name if rec.product_id else ''
            ws['F3'] = rec.op_no if rec.op_no else ''
            ws['G3'] = rec.fdm if rec.fdm else ''
            ws['H3'] = rec.operationg_mc if rec.operationg_mc else ''
            ws['I3'] = rec.comment if rec.comment else ''
            for trl in rec.tracking_line_ids:
                ws[f'A{curr_row}'] = trl.gauge_id.name if trl.gauge_id else ''
                ws.merge_cells(f'B{curr_row}:D{curr_row}')
                ws[f'B{curr_row}'] = trl.description if trl.description else ''
                ws[f'E{curr_row}'] = trl.item_code if trl.item_code is not False else ''
                ws[f'F{curr_row}'] = trl.item_description if trl.item_description is not False else ''
                ws[f'G{curr_row}'] = trl.request_no if trl.request_no else ''
                ws[f'H{curr_row}'] = trl.issue_date.strftime("%d-%m-%y") if trl.issue_date else ''
                ws[f'I{curr_row}'] = trl.qty if trl.qty is not False else ''
                ws[f'J{curr_row}'] = trl.expected_date.strftime("%d-%m-%y") if trl.expected_date else ''
                ws[f'K{curr_row}'] = trl.cost if trl.cost is not False else ''
                ws[f'L{curr_row}'] = trl.op_no if trl.op_no is not False else ''
                ws[f'M{curr_row}'] = trl.operating_machine if trl.operating_machine is not False else ''
                if trl.note:
                    soup = BeautifulSoup(trl.note, 'html.parser')
                    text_content = soup.find('p').get_text()
                    if text_content:
                        ws[f'N{curr_row}'] = text_content
                curr_row += 1

        if curr_row > max_row:
            max_row = curr_row

        # endregion

        # region formatting the dynamic cells
        for row in ws.iter_rows(min_row=5, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.border = border
                cell.alignment = align_center
                cell.font = font_all
        # endregion

        # region SignOff Members Footer
        cur_row = max_row + 1
        ws.merge_cells(f'A{cur_row}:N{cur_row}')
        cur_row += 1
        sign_row = cur_row

        ws[f'A{cur_row}'] = 'Prepared By'
        ws[f'A{cur_row}'].font = font_header
        ws.merge_cells(f'B{cur_row}:E{cur_row}')

        ws[f'F{cur_row}'] = 'Prepared Date'
        ws[f'F{cur_row}'].font = font_header
        ws.merge_cells(f'G{cur_row}:I{cur_row}')

        for rec in self:
            ws[f'B{cur_row}'] = rec.create_uid.name if rec.create_uid else ''
            ws[f'G{cur_row}'] = rec.create_date if rec.create_date else ''

        ws.row_dimensions[cur_row].height = 18
        cur_row += 1
        ws.merge_cells(f'A{cur_row}:I{cur_row}')
        cur_row += 1
        ws.merge_cells(f'A{cur_row}:I{cur_row}')
        ws[f'A{cur_row}'] = 'Sign OFF'
        ws[f'A{cur_row}'].font = Font(size=18, bold=True)
        ws.row_dimensions[cur_row].height = 25
        cur_row += 1
        ws.merge_cells(f'A{cur_row}:I{cur_row}')
        cur_row += 1
        for cell in ws[cur_row]:
            ws[f'A{cur_row}'] = 'Member'
            ws[f'C{cur_row}'] = 'Department'
            ws[f'E{cur_row}'] = 'Status'
            ws[f'F{cur_row}'] = 'Date'
            ws[f'G{cur_row}'] = 'Comments'
            ws.row_dimensions[cur_row].height = 25
            ws.merge_cells(f'A{cur_row}:B{cur_row}')
            ws.merge_cells(f'C{cur_row}:D{cur_row}')
            ws.merge_cells(f'G{cur_row}:I{cur_row}')
            cell.font = Font(size=12, bold=True, color='ffffff')
            cell.fill = PatternFill(start_color='0070C0', end_color='0070C0', fill_type="solid")

        for row_no in ws.iter_rows(min_row=1, max_row=cur_row + 1, min_col=1, max_col=14):
            for cell in row_no:
                cell.border = border
                cell.alignment = align_center

        # Listing Managers Id
        for rec in self:
            for record in rec.iatf_members_ids:
                name_rec = record.approver_id.name
                dept_rec = record.department_id.name if record.department_id.name else ''
                status_rec = record.approval_status.capitalize()
                date_rec = record.date_approved_rejected if record.date_approved_rejected else ''
                comment_rec = record.comment if record.comment else ''

                ws[f'A{cur_row + 1}'] = name_rec
                ws[f'C{cur_row + 1}'] = dept_rec
                ws[f'E{cur_row + 1}'] = status_rec
                ws[f'F{cur_row + 1}'] = date_rec
                ws[f'G{cur_row + 1}'] = comment_rec
                ws.merge_cells(f'A{cur_row+1}:B{cur_row+1}')
                ws.merge_cells(f'C{cur_row+1}:D{cur_row+1}')
                ws.merge_cells(f'G{cur_row+1}:I{cur_row+1}')

                status_dict = {'Approved': ['CFFFC3', '000000'], 'Rejected': ['FFCDCD', '000000'],
                               'Revision': ['AFEFFF', '000000'], 'Pending': ['FDFFCD', '000000']}

                for state, color in status_dict.items():
                    if state == status_rec:
                        ws[f'E{cur_row + 1}'].fill = PatternFill(start_color=color[0], end_color=color[0],
                                                                 fill_type="solid")
                        ws[f'E{cur_row + 1}'].font = Font(size=12, bold=False, color=color[1])

                cur_row += 1

            ws.merge_cells(f'A{cur_row + 1}:N{cur_row + 1}')
            ws.merge_cells(f'J{sign_row}:N{cur_row}')

            for row_no in ws.iter_rows(min_row=sign_row + 1, max_row=cur_row + 1, min_col=1, max_col=14):
                for cell in row_no:
                    cell.border = border
                    cell.alignment = align_center
            # endregion

        # Save the workbook
        wb.save(output)
        output.seek(0)
        self.generate_xls_file = base64.b64encode(output.getvalue()).decode('utf-8')

        return {
            "type": "ir.actions.act_url",
            "target": "self",
            "url": "/web/content?model=tracking.record&download=true&field=generate_xls_file&filename={filename}.xlsx&id={pid}".format(
                filename="Tool Tracking Record", pid=self[0].id),
        }


class TrackingRecordLines(models.Model):
    _name = "tracking.record.line"
    _description = "Tracking Record Line"
    _inherit = "translation.mixin"

    tracking_id = fields.Many2one(comodel_name='tracking.record', string='FIXTURE/GAUGE', ondelete='cascade',
                                  index=True, copy=False, readonly=True )
    description = fields.Char('Description',translate=True)
    gauge_id = fields.Many2one('gauge.name', string="Fixture/Gauge No")
    qty = fields.Integer('QTY')
    issue_date = fields.Date('Issue Date')
    request_no = fields.Char('Request No.',translate=True)
    expected_date = fields.Date('Expected Date')
    cost = fields.Float('Cost')
    note = fields.Html('Note',translate=True)
    item_code = fields.Char('Item Code',translate=True)
    item_description = fields.Char('Item Description',translate=True)
    op_no = fields.Char('Operation No.',translate=True)
    operating_machine = fields.Char('Operation Machine.',translate=True)


class Gauge(models.Model):
    _name = "gauge.name"
    _description = "Gauge Name"
    _inherit = "translation.mixin"

    name = fields.Char('Name',translate=True)
