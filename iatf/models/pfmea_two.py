from odoo import models, fields, api, _
from odoo.exceptions import ValidationError
import base64
from openpyxl import Workbook
from PIL import Image as PILImage, ImageOps
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
import io
from io import BytesIO
import logging
_logger = logging.getLogger(__name__)


class PfmeaTwo(models.Model):
    _name= 'asd.pfmea'
    _description= 'PFMEA-VDA_AIAG'
    _inherit = 'iatf.sign.off.members'

    company_name = fields.Many2one('res.company', string='Company Name', ondelete='cascade',
                                   default=lambda self: self.env.company.id)
    plant_location = fields.Char(string='Plant Location', related='company_name.partner_id.street', store=True)
    # customer_name = fields.Many2one('res.partner', string='Customer Name',ondelete='cascade')
    myp = fields.Char(string='Model/Year/Platform')
    # spd = fields.Char(string='Subject/Part No/Description')
    pfmea_start_date = fields.Date(string='PFMEA Start Date', default=fields.Date.today())
    pfmea_revision_date = fields.Date(string='PFMEA Revision Date')
    # drg_rev_no = fields.Char(string='Drawing Revision No.', related='part_id.drawing_rev_no', store=True)
    pfmea_aid_no = fields.Char(string='PFMEA AID Number')
    process_responsibility = fields.Char(string='Process Responsibility')
    confidentiality_level = fields.Selection([
        ('business', 'Business'),
        ('confidential', 'Confidential'),
        ('proprietary', 'Proprietary'),

    ], string='Confidentiality Level')
    report_ids= fields.One2many('pfmea.operations', 'operations_id', string='PFMEA Report')
    generate_xls_file = fields.Binary(string='Generated File')

    @api.model_create_multi
    def create(self, vals_list):
        records = super(PfmeaTwo, self).create(vals_list)

        for record in records:
            pfd_records = self.env["process.flow"].search(
                [('project_id', '=', record.project_id.id),
                 ('final_status', '=', 'approved')
                 ], limit=1)

            # if not pfd_records:
            #     raise ValidationError(
            #         _("No Process Flow Diagram is filled in this project first fill PFD before continuing..."))
            for operation in pfd_records.process_flow_line_ids.sorted(key="step"):
                for i in operation.process_op_lines_ids:
                    suggestion = self.env['pfmea.suggestion'].search(
                        [('process_step', '=', operation.desc_of_operation)], limit=1)

                    if not suggestion:
                        suggestion = self.env['pfmea.suggestion'].create({'process_step': operation.desc_of_operation})

                    self.env['pfmea.operations'].create({
                        'operations_id': record.id,
                        'issue': i.element_no,
                        'station_no': operation.step,
                        'process_step': suggestion.id,
                    })

        return records
    def generate_xls_report(self):
        # Create workbook
        output = BytesIO()
        wb = Workbook()

        # Define styles
        border = Border(top=Side(style='thin'), left=Side(style='thin'),
                        right=Side(style='thin'), bottom=Side(style='thin'))
        font_header = Font(name='Times New Roman', bold=True)
        symbol_font = Font(name='Symbol', size=40, bold=True)
        align_center = Alignment(vertical='center', horizontal='center', wrapText=True)

        # Define fill colors
        fill = PatternFill(start_color="e7e7e7", end_color="e7e7e7", fill_type="solid")  # Grey
        skyblue_fill = PatternFill(start_color="06bcf9", end_color="06bcf9", fill_type="solid")  # Sky Blue
        lightgreen_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light Green
        orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

        # Headers dictionary
        headers = {
            'C1': 'PROCESS FAILURE MODE AND EFFECT ANALYSIS',
            'C2': 'Planning & Preparation (STEP1)',
            'C3': 'Company Name:', 'C4': 'Plant location:', 'C5': 'Customer Name:',
            'C6': 'Model/Year/Platform:', 'F3': 'Subject/Part No/Description:',
            'F4': 'PFMEA Start Date:', 'F5': 'PFMEA Completion Date:', 'K3': 'Drg. Rev. No.',
            'O3': 'PFMEA AID Number:', 'O4': 'Process Responsibility:',
            'O5': 'Confidentiality Level:',
            'B8': 'CONTINUOUS IMPROVEMENT',
            'C8': 'STRUCTURE ANALYSIS (STEP 2)', 'G8': 'FUNCTION ANALYSIS (STEP 3)',
            'L8': 'FAILURE ANALYSIS (STEP 4)', 'Q8': 'RISK ANALYSIS (STEP 5)',
            'X8': 'OPTIMIZATION (STEP 6)',
        }

        # Column headers
        column_headers = {
            'A9': 'ISSUE #',
            'B9': 'History/ Change Authorization (Optional)',
            'C9': '1. Process Item System, Subsystem, Part Element or Name of Process',
            'D9': '2. Process Step Station No. and Name of Process Step',
            'E9': '3. Process Work Element 4M Type',
            'G9': '1. Function of the Process Item',
            'I9': '2. Function of the Process Step and Product Characteristics',
            'J9': '3. Function of the Process Work Element and Process Characteristics',
            'L9': '1. Failure Effects (FE)',
            'N9': 'Severity (S) of FE',
            'O9': '2. Failure Mode (FM) of the Process Step',
            'P9': '3. Failure Causes (FC) of the Work Element',
            'Q9': 'Current Prevention Control (PC) of FC',
            'R9': 'occurence (O) of FC',
            'S9': 'Current Detection Controls (DC) of FC or FM',
            'T9': 'Detection (D) of FC/FM',
            'U9': 'FMEA AP (Action Priority)',
            'V9': 'Special Product Characteristics',
            'W9': 'Filter Code (Optional)',
            'X9': 'Prevention Action',
            'Y9': 'Detection Action',
            'Z9': "Responsible Person's Name",
            'AA9': 'Target Completion Date',
            'AB9': 'Status',
            'AC9': 'Action Taken (with pointer to evidence)',
            'AD9': 'Completion Date',
            'AE9': 'Severity (S)',
            'AF9': 'occurence (O)',
            'AG9': 'Detection (D)',
            'AH9': 'Special Characteristics',
            'AI9': 'PFMEA AP',
            'AJ9': 'Remarks',
        }

        # Merge cells
        merge_cells = [
            'A1:B7', 'C1:AJ1',
            'C2:E2', 'D3:E3', 'D4:E4', 'D5:E5', 'D6:E6', 'L3:N3', 'L4:N4', 'L5:N5', 'L6:N6', 'E9:F9', 'G9:H9', 'J9:K9',
            'L9:M9',
            'C7:E7', 'F2:N2', 'F3:G3', 'F4:G4', 'F5:G5', 'F6:G6', 'H3:J3',
            'H4:J4', 'H5:J5', 'H6:J6', 'F7:N7', 'P2:R2', 'S2:AJ2', 'P3:R3', 'S3:AJ3', 'P4:R4', 'S4:AJ4',
            'P5:R5', 'S5:AJ5', 'P6:R6','S6:AJ7', 'C8:F8', 'G8:K8', 'L8:P8', 'Q8:W8', 'X8:AI8',
        ]

        # Set column widths
        col_widths = {
            'A': 3, 'B': 18, 'C': 18, 'D': 15, 'E': 15,
            'F': 15, 'G': 15, 'H': 15, 'I': 15, 'J': 15,
            'K': 15, 'L': 15, 'M': 15, 'N': 5, 'O': 15, 'P': 15, 'Q': 20, 'R': 5, 'S': 20, 'T': 5, 'U': 5, 'V': 8,
            'W': 5,
            'X': 5, 'Y': 5, 'Z': 5, 'AA': 8, 'AB': 5, 'AC': 5, 'AD': 5, 'AE': 5, 'AF': 5, 'AG': 5, 'AH': 5, 'AI': 5,
            'AJ': 18,
        }

        # Apply colors to specific cells
        cells_skyblue = ['A9', 'D9', 'I9', 'O9']
        cells_lightgreen = ['B9']
        cells_orange = ['E9', 'J9', 'P9']

        for rec in self:
            process_sheets = {}

            for process_step in rec.report_ids:
                process_step_name = process_step.process_step_name[:31]  # Limit sheet name to 31 characters


                if process_step_name not in process_sheets:
                    process_sheets[process_step_name] = []
                process_sheets[process_step_name].append(process_step)
            for process_step_name, process_steps in process_sheets.items():
                # Create a safe sheet name (Excel limits sheet names to 31 chars and no special chars)
                safe_name = process_step_name[:31].replace('/', '_').replace('\\', '_').replace('?', '_').replace('*',
                                                                                                                  '_')
                safe_name = safe_name.replace('[', '_').replace(']', '_').replace(':', '_')
                ws = wb.create_sheet(title=safe_name)
                if self.env.user.company_id and self.env.user.company_id.logo:
                    max_width, max_height = 100, 200  # Set max dimensions
                    image_data = base64.b64decode(self.env.user.company_id.logo)

                    # Open image with PIL
                    image = PILImage.open(io.BytesIO(image_data))
                    width, height = image.size
                    aspect_ratio = width / height

                    # Resize with aspect ratio
                    if width > max_width:
                        width = max_width
                        height = int(width / aspect_ratio)
                    if height > max_height:
                        height = max_height
                        width = int(height * aspect_ratio)

                    # Resize and add padding
                    resized_image = image.resize((width, height), PILImage.LANCZOS)
                    padding_top, padding_left = 5, 5
                    resized_image = ImageOps.expand(resized_image, border=(padding_left, padding_top, 0, 0),
                                                    fill='white')

                    img_bytes = io.BytesIO()
                    resized_image.save(img_bytes, format='PNG')
                    img_bytes.seek(0)
                    logo_image = Image(img_bytes)

                    ws.add_image(logo_image, 'A1')

                # Apply headers and column headers
                for cell, value in headers.items():
                    ws[cell] = value
                    ws[cell].font = font_header
                    ws[cell].alignment = align_center
                    ws[cell].fill = fill

                for cell, value in column_headers.items():
                    ws[cell] = value
                    ws[cell].alignment = align_center
                    ws[cell].fill = fill

                # Merge cells
                for cell_range in merge_cells:
                    ws.merge_cells(cell_range)

                # Set row heights
                ws.row_dimensions[1].height = 30
                ws.row_dimensions[3].height = 30
                ws.row_dimensions[4].height = 30
                ws.row_dimensions[5].height = 30
                ws.row_dimensions[6].height = 30
                ws.row_dimensions[8].height = 40
                ws.row_dimensions[9].height = 100

                # Apply the fill to the range F2:AG2
                for row in ws['F2:AG2']:
                    for cell in row:
                        cell.fill = fill

                # Set column widths
                for col, width in col_widths.items():
                    ws.column_dimensions[col].width = width

                # Assign colors to specified cells
                for cell in cells_skyblue:
                    ws[cell].fill = skyblue_fill

                for cell in cells_lightgreen:
                    ws[cell].fill = lightgreen_fill

                for cell in cells_orange:
                    ws[cell].fill = orange_fill

                # Fill in header information
                ws['D3'] = rec.company_name.name if rec.company_name else ''
                ws['D4'] = rec.plant_location if rec.plant_location else ''
                ws['D5'] = rec.partner_id.name if rec.partner_id else ''
                ws['D6'] = rec.myp if rec.myp else ''
                ws['H3'] = rec.part_id.name if rec.part_id else ''
                ws['H4'] = rec.pfmea_start_date if rec.pfmea_start_date else ''
                ws['H5'] = rec.pfmea_revision_date if rec.pfmea_revision_date else ''
                ws['L3'] = rec.drawing_rev_no if rec.drawing_rev_no else ''
                ws['P3'] = rec.pfmea_aid_no if rec.pfmea_aid_no else ''
                ws['P4'] = rec.process_responsibility if rec.process_responsibility else ''
                ws['P5'] = rec.confidentiality_level if rec.confidentiality_level else ''

                # Inside generate_xls_report function, replace the data population part (starting from the row variable)
                # This is the section that writes data to the Excel file

                row = 10
                for process_step in process_steps:
                    start_row = row

                    # Write operation-level data
                    ws[f'A{row}'] = process_step.issue if process_step.issue else ''
                    ws[f'B{row}'] = process_step.hca if process_step.hca else ''
                    ws[f'C{row}'] = process_step.process_item if process_step.process_item else ''
                    ws[f'D{row}'] = ', '.join(
                        filter(None, [process_step.station_no or '', process_step.process_step_name or '']))

                    # Write function process item data
                    fpi_data = [
                        ('Process Item', process_step.fpi_process_item),
                        ('In Plant', process_step.fpi_in_plant),
                        ('Ship to Plant', process_step.fpi_ship_to_plant),
                        ('End User', process_step.fpi_end_user)
                    ]

                    for i, (label, value) in enumerate(fpi_data):
                        ws[f'G{row + i}'] = label
                        ws[f'H{row + i}'] = value if value else ''
                        ws[f'G{row + i}'].font = font_header

                    ws[f'I{row}'] = process_step.fun_of_process_step if process_step.fun_of_process_step else ''

                    # Write failure effects data
                    fe_data = [
                        ('In Plant', process_step.fe_in_plant),
                        ('Customer End', process_step.fe_customer_end),
                        ('End User', process_step.fe_end_user)
                    ]

                    for i, (label, value) in enumerate(fe_data):
                        ws[f'L{row + i}'] = label
                        ws[f'M{row + i}'] = value if value else ''
                        ws[f'L{row + i}'].font = font_header

                    ws[f'O{row}'] = process_step.failure_mode if process_step.failure_mode else ''
                    if process_step.special_product_characteristics:
                        ws[f'V{row}'] = process_step.special_product_characteristics.symbol
                        ws[f'V{row}'].font = symbol_font
                    ws[f'N{row}'] = process_step.severity if process_step.severity else ''

                    # Get all line categories
                    all_line_categories = [
                        ('man', process_step.man_line_ids),
                        ('machine', process_step.machine_line_ids),
                        ('material', process_step.material_line_ids),
                        ('environment', process_step.environment_line_ids),
                        ('method', process_step.method_line_ids)
                    ]

                    # Track rows for each 4M type for merging later
                    type_start_rows = {}

                    # Process each line category
                    for work_type, lines in all_line_categories:
                        if not lines:
                            # Print an empty row if there are no records for this type
                            ws[f'E{row}'] = work_type.capitalize()
                            ws[f'E{row}'].font = font_header
                            ws[f'J{row}'] = work_type.capitalize()
                            ws[f'J{row}'].font = font_header
                            row += 1
                            continue

                        # Record starting row for this work type
                        type_start_row = row
                        type_start_rows[work_type] = type_start_row

                        for line in lines:
                            ws[f'E{row}'] = work_type.capitalize()
                            ws[f'E{row}'].font = font_header
                            ws[f'J{row}'] = work_type.capitalize()
                            ws[f'J{row}'].font = font_header

                            ws[f'F{row}'] = line.process_work_element if line.process_work_element else ''
                            ws[
                                f'K{row}'] = line.function_of_process_work_element if line.function_of_process_work_element else ''
                            ws[f'P{row}'] = line.failure_causes if line.failure_causes else ''
                            ws[f'Q{row}'] = line.current_prevention_control if line.current_prevention_control else ''
                            ws[f'R{row}'] = line.occurrence if line.occurrence else ''
                            ws[f'S{row}'] = line.current_detection_control if line.current_detection_control else ''
                            ws[f'T{row}'] = line.detection if line.detection else ''
                            ws[f'U{row}'] = line.fmea_ap if line.fmea_ap else ''
                            ws[f'W{row}'] = line.filter_code if line.filter_code else ''
                            ws[f'X{row}'] = line.prevention_action if line.prevention_action else ''
                            ws[f'Y{row}'] = line.detection_action if line.detection_action else ''
                            ws[f'Z{row}'] = line.responsible_person_name.name if line.responsible_person_name else ''
                            ws[f'AA{row}'] = line.target_completion_date if line.target_completion_date else ''
                            ws[f'AB{row}'] = line.status if line.status else ''
                            ws[f'AC{row}'] = line.action_taken if line.action_taken else ''
                            ws[f'AD{row}'] = line.completion_date if line.completion_date else ''
                            ws[f'AE{row}'] = line.severity_level if line.severity_level else ''
                            ws[f'AF{row}'] = line.occurrence_level if line.occurrence_level else ''
                            ws[f'AG{row}'] = line.detection_level if line.detection_level else ''
                            ws[f'AH{row}'] = line.special_characteristics if line.special_characteristics else ''
                            ws[f'AI{row}'] = line.pfmea_ap if line.pfmea_ap else ''
                            ws[f'AJ{row}'] = line.remarks if line.remarks else ''

                            row += 1

                        # Merge cells for this work type if multiple lines exist
                        if row > type_start_row + 1:
                            for col in ['E', 'J']:
                                ws.merge_cells(f'{col}{type_start_row}:{col}{row - 1}')

                    # Merge operation-level cells across all rows
                    if row > start_row + 1:
                        for col in ['A', 'B', 'C', 'D', 'I', 'N','O', 'V']:
                            ws.merge_cells(f'{col}{start_row}:{col}{row - 1}')


                cur_row = 30
                if cur_row < row:
                    cur_row = row
                for rows in ws.iter_rows(min_row=8, max_row=cur_row, min_col=1, max_col=36):
                    for cell in rows:
                        cell.alignment = align_center
                        cell.border = border

                # Set vertical text orientation for narrow columns
                vertical_cols = ['A', 'N', 'R', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z',
                                 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ']
                for col in vertical_cols:
                    ws[f'{col}9'].alignment = Alignment(textRotation=90, vertical='center', horizontal='center')

                # region SignOff Members Footer
                sign_row = cur_row
                ws.merge_cells(f'A{cur_row}:S{cur_row}')

                cur_row += 1
                ws[f'A{cur_row}'] = 'Prepared By'
                ws[f'A{cur_row}'].font = font_header
                ws.merge_cells(f'A{cur_row}:B{cur_row}')
                ws.merge_cells(f'C{cur_row}:E{cur_row}')

                ws[f'F{cur_row}'] = 'Prepared Date'
                ws[f'F{cur_row}'].font = font_header
                ws.merge_cells(f'F{cur_row}:G{cur_row}')
                ws.merge_cells(f'H{cur_row}:L{cur_row}')

                for rec in self:
                    ws[f'C{cur_row}'] = rec.create_uid.name if rec.create_uid else ''
                    ws[f'H{cur_row}'] = rec.create_date if rec.create_date else ''

                ws.row_dimensions[cur_row].height = 18
                cur_row += 1
                ws.merge_cells(f'A{cur_row}:L{cur_row}')
                cur_row += 1
                ws.merge_cells(f'A{cur_row}:L{cur_row}')
                ws[f'A{cur_row}'] = 'Sign OFF'
                ws[f'A{cur_row}'].font = Font(size=18, bold=True)
                ws.row_dimensions[cur_row].height = 25
                cur_row += 1
                ws.merge_cells(f'A{cur_row}:L{cur_row}')
                cur_row += 1
                for cell in ws[cur_row]:
                    ws[f'A{cur_row}'] = 'Member'
                    ws[f'D{cur_row}'] = 'Department'
                    ws[f'F{cur_row}'] = 'Status'
                    ws[f'H{cur_row}'] = 'Date'
                    ws[f'J{cur_row}'] = 'Comments'
                    ws.row_dimensions[cur_row].height = 25
                    ws.merge_cells(f'A{cur_row}:C{cur_row}')
                    ws.merge_cells(f'D{cur_row}:E{cur_row}')
                    ws.merge_cells(f'F{cur_row}:G{cur_row}')
                    ws.merge_cells(f'H{cur_row}:I{cur_row}')
                    ws.merge_cells(f'J{cur_row}:L{cur_row}')
                    cell.font = Font(size=12, bold=True, color='ffffff')
                    cell.fill = PatternFill(start_color='0070C0', end_color='0070C0', fill_type="solid")

                # Listing Managers Id
                for rec in self:
                    for record in rec.iatf_members_ids:
                        name_rec = record.approver_id.name
                        dept_rec = record.department_id.name if record.department_id.name else ''
                        status_rec = record.approval_status.capitalize()
                        date_rec = record.date_approved_rejected if record.date_approved_rejected else ''
                        comment_rec = record.comment if record.comment else ''

                        ws[f'A{cur_row + 1}'] = name_rec
                        ws[f'D{cur_row + 1}'] = dept_rec
                        ws[f'F{cur_row + 1}'] = status_rec
                        ws[f'H{cur_row + 1}'] = date_rec
                        ws[f'J{cur_row + 1}'] = comment_rec
                        ws.merge_cells(f'A{cur_row + 1}:C{cur_row + 1}')
                        ws.merge_cells(f'D{cur_row + 1}:E{cur_row + 1}')
                        ws.merge_cells(f'F{cur_row + 1}:G{cur_row + 1}')
                        ws.merge_cells(f'H{cur_row + 1}:I{cur_row + 1}')
                        ws.merge_cells(f'J{cur_row + 1}:L{cur_row + 1}')

                        status_dict = {'Approved': ['CFFFC3', '000000'], 'Rejected': ['FFCDCD', '000000'],
                                       'Revision': ['AFEFFF', '000000'], 'Pending': ['FDFFCD', '000000']}

                        for state, color in status_dict.items():
                            if state == status_rec:
                                ws[f'F{cur_row + 1}'].fill = PatternFill(start_color=color[0], end_color=color[0],
                                                                         fill_type="solid")
                                ws[f'F{cur_row + 1}'].font = Font(size=12, bold=False, color=color[1])

                        cur_row += 1

                    # ws.merge_cells(f'C{cur_row + 1}:D{cur_row + 1}')
                    # ws.merge_cells(f'E{cur_row + 1}:F{cur_row + 1}')
                    ws.merge_cells(f'A{cur_row + 1}:S{cur_row + 1}')
                    ws.merge_cells(f'M{sign_row + 1}:S{cur_row}')

                    for row_no in ws.iter_rows(min_row=sign_row, max_row=cur_row + 1, min_col=1, max_col=17):
                        for cell in row_no:
                            cell.border = border
                            cell.alignment = align_center
                # endregion

        # Remove the default sheet created by openpyxl
        wb.remove(wb['Sheet'])


        # Save workbook to BytesIO
        wb.save(output)
        output.seek(0)

        # Create attachment
        attachment = self.env['ir.attachment'].create({
            'name': 'pfmea.xlsx',
            'type': 'binary',
            'datas': base64.b64encode(output.getvalue()),
            'res_model': 'asd.pfmea',
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        })

        # Return download link
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }

class PfmeaTwoOperations(models.Model):
    _name = 'pfmea.operations'
    _description = 'PFMEA Operations'

    operations_id = fields.Many2one('asd.pfmea', string='Process Report', ondelete='cascade')
    issue = fields.Char(string='Issue #')
    hca = fields.Char(string='History/Change Authorization')
    process_item = fields.Char(string='Name of Process')

    # Function of the Process Item
    fpi_process_item = fields.Char(string='Process Item')
    fpi_in_plant = fields.Char(string='In Plant')
    fpi_ship_to_plant = fields.Char(string='Ship to Plant')
    fpi_end_user = fields.Char(string='End User')

    # Failure Effects (FE)
    fe_in_plant = fields.Char(string='In Plant')
    fe_customer_end = fields.Char(string='Customer End')
    fe_end_user = fields.Char(string='End User')

    station_no = fields.Char(string='Station No.')
    process_step = fields.Many2one('pfmea.suggestion', string='Process Step')
    process_step_name = fields.Char(related='process_step.process_step', string='Process Step')
    # Function of the Process Step and Product Characteristics
    fun_of_process_step = fields.Char(string='Function of Process Step and Product Characteristics')

    # Failure Mode (FM)
    failure_mode = fields.Char(string='Failure Mode (FM)')

    # Special Product Characteristics
    special_product_characteristics = fields.Many2one('process.flow.class', string='Special Product Characteristics')

    severity = fields.Integer(string='Severity (S) of FE', default=1)

    man_line_ids = fields.One2many('pfmea.operations.line', 'man_pro_id',
                                   string='Man Lines',

                                   )
    machine_line_ids = fields.One2many('pfmea.operations.line', 'machine_pro_id',
                                       string='Machine Lines',
                                       )
    material_line_ids = fields.One2many('pfmea.operations.line', 'material_pro_id',
                                        string='Material Lines',

                                        )
    environment_line_ids = fields.One2many('pfmea.operations.line', 'environment_pro_id',
                                           string='Environment Lines',

                                           )
    method_line_ids = fields.One2many('pfmea.operations.line', 'method_pro_id',
                                      string='Method Lines',

                                      )
    _sql_constraints = [
        ('unique_issue_process_step', 'UNIQUE(project_id,process_step, issue)',
         'Issue number must be unique for a given Process Step!')
    ]

    @api.onchange('process_step')
    def _onchange_process_step(self):
        if self.process_step:
            # Use the process_step as the suggestion
            suggestion = self.process_step
            self._update_fields_and_lines_from_suggestion(suggestion)

    @api.onchange('issue')
    def _onchange_issue(self):
        if self.process_step and self.issue:
            # Try to find a specific suggestion with matching process step and issue
            suggestion = self.env['pfmea.suggestion'].search([
                ('process_step', '=', self.process_step.process_step),
                ('issue', '=', self.issue)
            ], limit=1)

            if suggestion:
                self._update_fields_and_lines_from_suggestion(suggestion)

    def _update_fields_and_lines_from_suggestion(self, suggestion):
        """
        Helper method to update fields and lines from a suggestion, ensuring all work types are properly handled
        and avoiding duplicates.
        """
        # Update scalar fields from suggestion
        fields_to_update = [
            'hca', 'process_item', 'fpi_process_item', 'fpi_in_plant',
            'fpi_ship_to_plant', 'fpi_end_user', 'fe_in_plant', 'fe_customer_end',
            'fe_end_user', 'severity', 'station_no', 'fun_of_process_step', 'failure_mode'
        ]

        for field in fields_to_update:
            if hasattr(suggestion, field):
                setattr(self, field, getattr(suggestion, field))

        self.special_product_characteristics = suggestion.special_product_characteristics

        # Map suggestion lines to their respective work types
        line_mapping = {
            'man': (suggestion.man_line_ids, 'man_line_ids', 'man_pro_id'),
            'machine': (suggestion.machine_line_ids, 'machine_line_ids', 'machine_pro_id'),
            'material': (suggestion.material_line_ids, 'material_line_ids', 'material_pro_id'),
            'environment': (suggestion.environment_line_ids, 'environment_line_ids', 'environment_pro_id'),
            'method': (suggestion.method_line_ids, 'method_line_ids', 'method_pro_id'),
        }

        # Track existing work types to avoid duplicates
        existing_work_types = set()
        for field in ['man_line_ids', 'machine_line_ids', 'material_line_ids', 'environment_line_ids',
                      'method_line_ids']:
            if self[field]:  # Check if there are existing lines for this work type
                existing_work_types.add(field.replace('_line_ids', ''))

        # Process each work type and skip if it already exists
        for work_type, (source_lines, target_field, ref_field) in line_mapping.items():
            # Skip this work type if it already has lines
            if work_type in existing_work_types:
                _logger.info(f"Skipping work type '{work_type}' as it already exists for this process operation.")
                continue

            new_lines = []
            for line in source_lines:
                line_vals = {
                    'work_type': work_type,
                    'process_work_element': line.process_work_element,
                    'function_of_process_work_element': line.function_of_process_work_element,
                    'failure_causes': line.failure_causes,
                    'current_prevention_control': line.current_prevention_control,
                    'occurrence': line.occurrence,
                    'current_detection_control': line.current_detection_control,
                    'detection': line.detection,
                    'filter_code': line.filter_code,
                    'prevention_action': line.prevention_action,
                    'detection_action': line.detection_action,
                    'responsible_person_name': line.responsible_person_name,
                    'target_completion_date': line.target_completion_date,
                    'status': line.status,
                    'action_taken': line.action_taken,
                    'completion_date': line.completion_date,
                    'severity_level': line.severity_level,
                    'occurrence_level': line.occurrence_level,
                    'detection_level': line.detection_level,
                    'special_characteristics': line.special_characteristics,
                    'pfmea_ap': line.pfmea_ap,
                    'remarks': line.remarks,
                }
                new_lines.append((0, 0, line_vals))

            if new_lines:
                self[target_field] = new_lines

    @api.constrains('issue')
    def _check_issue(self):
        for record in self:
            if not record.issue:
                raise ValidationError(_("Please enter an Issue number. Issue field cannot be empty."))

    def _sync_with_suggestion(self):
        for record in self:
            if not record.process_step:
                continue

            # Prepare suggestion values
            suggestion_vals = {
                'issue': record.issue,
                'hca': record.hca,
                'process_item': record.process_item,
                'fpi_process_item': record.fpi_process_item,
                'fpi_in_plant': record.fpi_in_plant,
                'fpi_ship_to_plant': record.fpi_ship_to_plant,
                'fpi_end_user': record.fpi_end_user,
                'fe_in_plant': record.fe_in_plant,
                'fe_customer_end': record.fe_customer_end,
                'fe_end_user': record.fe_end_user,
                'station_no': record.station_no,
                'fun_of_process_step': record.fun_of_process_step,
                'failure_mode': record.failure_mode,
                'special_product_characteristics': record.special_product_characteristics.id if record.special_product_characteristics else False,
                'severity': record.severity,
            }

            # Find or create suggestion
            suggestion = self.env['pfmea.suggestion'].search([
                ('process_step', '=', record.process_step.process_step),
                ('issue', '=', record.issue or '')
            ], limit=1)

            if suggestion:
                suggestion.write(suggestion_vals)
            else:
                suggestion_vals['process_step'] = record.process_step.process_step
                suggestion = self.env['pfmea.suggestion'].create(suggestion_vals)
                record.process_step = suggestion

            # Clear existing suggestion lines
            suggestion.man_line_ids = [(5, 0, 0)]
            suggestion.machine_line_ids = [(5, 0, 0)]
            suggestion.material_line_ids = [(5, 0, 0)]
            suggestion.environment_line_ids = [(5, 0, 0)]
            suggestion.method_line_ids = [(5, 0, 0)]

            # Sync all line types
            line_mapping = {
                'man': (record.man_line_ids, 'man_line_ids', 'man_pro_id'),
                'machine': (record.machine_line_ids, 'machine_line_ids', 'machine_pro_id'),
                'material': (record.material_line_ids, 'material_line_ids', 'material_pro_id'),
                'environment': (record.environment_line_ids, 'environment_line_ids', 'environment_pro_id'),
                'method': (record.method_line_ids, 'method_line_ids', 'method_pro_id'),
            }

            for work_type, (source_lines, target_field, ref_field) in line_mapping.items():
                new_lines = []
                for line in source_lines:
                    line_vals = {
                        ref_field: suggestion.id,
                        'work_type': work_type,
                        'process_work_element': line.process_work_element,
                        'function_of_process_work_element': line.function_of_process_work_element,
                        'failure_causes': line.failure_causes,
                        'current_prevention_control': line.current_prevention_control,
                        'occurrence': line.occurrence,
                        'current_detection_control': line.current_detection_control,
                        'detection': line.detection,
                        'filter_code': line.filter_code,
                        'prevention_action': line.prevention_action,
                        'detection_action': line.detection_action,
                        'responsible_person_name': line.responsible_person_name,
                        'target_completion_date': line.target_completion_date,
                        'status': line.status,
                        'action_taken': line.action_taken,
                        'completion_date': line.completion_date,
                        'severity_level': line.severity_level,
                        'occurrence_level': line.occurrence_level,
                        'detection_level': line.detection_level,
                        'special_characteristics': line.special_characteristics,
                        'pfmea_ap': line.pfmea_ap,
                        'remarks': line.remarks,
                    }
                    new_lines.append((0, 0, line_vals))

                if new_lines:
                    suggestion[target_field] = new_lines

    def write(self, vals):
        result = super(PfmeaTwoOperations, self).write(vals)
        self._sync_with_suggestion()
        return result

    @api.model
    def create(self, vals):
        record = super(PfmeaTwoOperations, self).create(vals)
        record._sync_with_suggestion()
        return record

class PfmeaOperationsLine(models.Model):
    _name = 'pfmea.operations.line'
    _description = 'PFMEA Operations Line'

   
    man_pro_id = fields.Many2one('pfmea.operations', string='Man Process Reference')
    machine_pro_id = fields.Many2one('pfmea.operations', string='Machine Process Reference')
    material_pro_id = fields.Many2one('pfmea.operations', string='Material Process Reference')
    environment_pro_id = fields.Many2one('pfmea.operations', string='Environment Process Reference')
    method_pro_id = fields.Many2one('pfmea.operations', string='Method Process Reference')

    work_type = fields.Selection([
        ('man', 'Man'),
        ('machine', 'Machine'),
        ('material', 'Material'),
        ('environment', 'Environment'),
        ('method', 'Method')
    ], string="4M Type", readonly=True)
    process_work_element = fields.Text(string="Process Work Element")
    function_of_process_work_element = fields.Text(
        string="Function of Process Work Element and Process Characteristics")
    severity = fields.Integer(string='Severity (S) of FE', compute='_compute_severity', store=True)
    failure_causes = fields.Char(string='Failure Causes (FC)')
    current_prevention_control = fields.Char(string='Current Prevention Control of FC')
    occurrence = fields.Integer(string='Occurrence (O) of FC', help="Occurrence rating from 1-10", default=1)
    current_detection_control = fields.Char(string='Current Detection Control of FC or FM')
    detection = fields.Integer(string='Detection (D) of FC/FM', help="Detection rating from 1-10", default=1)
    fmea_ap = fields.Char(string="FMEA AP", compute='_compute_fmea_ap', store=True)

    filter_code = fields.Text(string='Filter Code (Optional)')
    prevention_action = fields.Text(string='Prevention Action')
    detection_action = fields.Text(string='Detection Action')
    responsible_person_name = fields.Many2one('res.users', string='Responsible Person Name')
    target_completion_date = fields.Date(string='Target Completion Date')
    status = fields.Selection([
        ('untouched', 'Untouched'),
        ('under_consideration', 'Under Consideration'),
        ('in_progress', 'In Progress'),
        ('completed', 'Completed'),
        ('discarded', 'Discarded'),
    ], string='Status')
    action_taken = fields.Text(string='Action Taken (with pointer to evidence)')
    completion_date = fields.Date(string='Completion Date')

    severity_level = fields.Integer(string='Severity Level', help="Severity rating from 1-10", default=1)
    occurrence_level = fields.Integer(string='Occurrence Level', help="Occurrence rating from 1-10", default=1)
    detection_level = fields.Integer(string='Detection Level', help="Detection rating from 1-10", default=1)
    special_characteristics = fields.Text(string='Special Characteristics')
    pfmea_ap = fields.Char(string='PFMEA AP', compute='_compute_fmea_ap_level', store=True)
    remarks = fields.Text(string='Remarks')

    @api.depends('man_pro_id.severity', 'machine_pro_id.severity', 'material_pro_id.severity',
                 'environment_pro_id.severity', 'method_pro_id.severity')
    def _compute_severity(self):
        """
        Compute the severity based on the parent pfmea.operations record,
        regardless of which reference field is set.
        """
        for record in self:
            # Check each reference field and use the severity from the linked pfmea.operations
            if record.man_pro_id:
                record.severity = record.man_pro_id.severity
            elif record.machine_pro_id:
                record.severity = record.machine_pro_id.severity
            elif record.material_pro_id:
                record.severity = record.material_pro_id.severity
            elif record.environment_pro_id:
                record.severity = record.environment_pro_id.severity
            elif record.method_pro_id:
                record.severity = record.method_pro_id.severity
            else:
                record.severity = 1  # Default value if no parent is linked

    @api.depends('severity', 'occurrence', 'detection')
    def _compute_fmea_ap(self):
        for record in self:
            if record.occurrence and record.occurrence > 10:
                raise ValidationError("OCC Value should not exceed 10")
            if record.detection and record.detection > 10:
                raise ValidationError("Detection Value should not exceed 10")
            if record.occurrence and record.occurrence < 1:
                raise ValidationError("OCC Value should not be less than 1")
            if record.detection and record.detection < 1:
                raise ValidationError("Detection Value should not be less than 1")

            else:
                # Define lookup table for action priority (AP)
                ap_lookup = {
                    (9, 10): {
                        (8, 10): {7: 'H', 5: 'H', 2: 'H', 1: 'H'},
                        (6, 7): {7: 'H', 5: 'H', 2: 'H', 1: 'H'},
                        (4, 5): {7: 'H', 5: 'H', 2: 'H', 1: 'M'},
                        (2, 3): {7: 'H', 5: 'M', 2: 'L', 1: 'L'},
                        (0, 1): 'L',
                    },
                    (7, 8): {
                        (8, 10): {7: 'H', 5: 'H', 2: 'H', 1: 'H'},
                        (6, 7): {7: 'H', 5: 'H', 2: 'H', 1: 'M'},
                        (4, 5): {7: 'H', 5: 'M', 2: 'M', 1: 'M'},
                        (2, 3): {7: 'M', 5: 'M', 2: 'L', 1: 'L'},
                        (0, 1): 'L',
                    },

                    (4, 6): {
                        (8, 10): {7: 'H', 5: 'H', 2: 'M', 1: 'M'},
                        (6, 7): {7: 'M', 5: 'M', 2: 'M', 1: 'L'},
                        (4, 5): {7: 'M', 5: 'L', 2: 'L', 1: 'L'},
                        (2, 3): 'L',
                        (0, 1): 'L',
                    },
                    (2, 3): {
                        (8, 10): {6: 'M', 4: 'M', 1: 'L', 0: 'L'},
                        (6, 7): 'L',
                        (4, 5): 'L',
                        (2, 3): 'L',
                        (0, 1): 'L',
                    },
                    (0, 1): {
                        (1, 10): 'L'
                    }
                }

                # Default value if no match is found
                record.fmea_ap = 'L'

                # Find correct severity range
                for sev_range, occ_dict in ap_lookup.items():
                    if sev_range[0] <= record.severity <= sev_range[1]:
                        # Find correct occurrence range
                        for occ_range, det_dict in occ_dict.items():
                            if occ_range[0] <= record.occurrence <= occ_range[1]:
                                # Assign action priority based on detection
                                if isinstance(det_dict, dict):  # If detection mapping exists
                                    for det_threshold, ap_value in det_dict.items():
                                        if record.detection >= det_threshold:
                                            record.fmea_ap = ap_value
                                            break
                                else:  # If no detection mapping, assign the value directly
                                    record.fmea_ap = det_dict
                                break
                        break

    @api.depends('severity_level', 'occurrence_level', 'detection_level')
    def _compute_fmea_ap_level(self):
        for record in self:
            # Validation for severity, occurrence, and detection

            if record.severity_level > 10:
                raise ValidationError("Severity Level should not exceed 10")
            if record.occurrence_level > 10:
                raise ValidationError("Occurrence Level should not exceed 10")
            if record.detection_level > 10:
                raise ValidationError("Detection Level should not exceed 10")
            if record.severity_level < 1:
                raise ValidationError("Severity Level should not be less than 1")
            if record.occurrence_level < 1:
                raise ValidationError("Occurrence Level should not be less than 1")
            if record.detection_level < 1:
                raise ValidationError("Detection Level should not be less than 1")
            else:
                # Define lookup table for action priority (AP) based on severity_level, occurrence_level, and detection_level

                ap_lookup = {
                    (9, 10): {
                        (8, 10): {7: 'H', 5: 'H', 2: 'H', 1: 'H'},
                        (6, 7): {7: 'H', 5: 'H', 2: 'H', 1: 'H'},
                        (4, 5): {7: 'H', 5: 'H', 2: 'H', 1: 'M'},
                        (2, 3): {7: 'H', 5: 'M', 2: 'L', 1: 'L'},
                        (0, 1): 'L',
                    },
                    (7, 8): {
                        (8, 10): {7: 'H', 5: 'H', 2: 'H', 1: 'H'},
                        (6, 7): {7: 'H', 5: 'H', 2: 'H', 1: 'M'},
                        (4, 5): {7: 'H', 5: 'M', 2: 'M', 1: 'M'},
                        (2, 3): {7: 'M', 5: 'M', 2: 'L', 1: 'L'},
                        (0, 1): 'L',
                    },

                    (4, 6): {
                        (8, 10): {7: 'H', 5: 'H', 2: 'M', 1: 'M'},
                        (6, 7): {7: 'M', 5: 'M', 2: 'M', 1: 'L'},
                        (4, 5): {7: 'M', 5: 'L', 2: 'L', 1: 'L'},
                        (2, 3): 'L',
                        (0, 1): 'L',
                    },
                    (2, 3): {
                        (8, 10): {6: 'M', 4: 'M', 1: 'L', 0: 'L'},
                        (6, 7): 'L',
                        (4, 5): 'L',
                        (2, 3): 'L',
                        (0, 1): 'L',
                    },
                    (0, 1): {
                        (1, 10): 'L'
                    }
                }

            # Default value if no match is found
            record.pfmea_ap = 'L'

            # Find correct severity range for severity_level
            for sev_range, occ_dict in ap_lookup.items():
                if sev_range[0] <= record.severity_level <= sev_range[1]:
                    # Find correct occurrence range for occurrence_level
                    for occ_range, det_dict in occ_dict.items():
                        if occ_range[0] <= record.occurrence_level <= occ_range[1]:
                            # Assign action priority based on detection_level
                            if isinstance(det_dict, dict):  # If detection mapping exists
                                for det_threshold, ap_value in det_dict.items():
                                    if record.detection_level >= det_threshold:
                                        record.pfmea_ap = ap_value
                                        break
                            else:  # If no detection mapping, assign the value directly
                                record.pfmea_ap = det_dict
                            break
                    break

