import base64
import io
from io import BytesIO

from PIL import Image as PILImage
from PIL import ImageOps
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, Border, Side, DEFAULT_FONT, PatternFill

from odoo import api, fields, models


class FeasibilityCommitment(models.Model):
    _name = 'feasibility.commitment'
    _description = 'Feasibility Commitment Report'
    _rec_name = 'part_num'
    _inherit = ['iatf.sign.off.members', 'translation.mixin']

    format_id = fields.Many2one('mom.format', 'Format')

    rfq_no = fields.Char(string='RFQ Number')
    cust_id = fields.Many2one('res.partner', string='Customer')
    date = fields.Date(string='Date')
    company_id = fields.Many2one('res.company', string='Company')
    # signature = fields.Binary(string="Signature")
    title = fields.Char(string='Title',translate=True)
    signature_date = fields.Date(string='Date')
    # Part
    part_id = fields.Many2one('product.template', string='Part Name')
    part_num = fields.Char(related='part_id.default_code', string='Part Number')
    drawing_num = fields.Char(related='part_id.default_code', string='Drawing num')

    # Revision
    revision_date = fields.Date(string='Revision Date')
    revision_number = fields.Char(string='Revision Number')
    hr = fields.Many2one('res.users', 'HR')  # HR APPROVED
    design_eng = fields.Many2one('res.users', 'Design Engineering')  # DESIGN
    manf_eng = fields.Many2one('res.users', 'Manufacturing Engineering')  # Engineering
    marketing = fields.Many2one('res.users', 'Marketing')
    program_management = fields.Many2one('res.users', 'Program Management')
    production = fields.Many2one('res.users', 'Production')  # Manufacturing
    quality = fields.Many2one('res.users', 'Quality')  # Qulity
    top_management_id = fields.Many2one('res.users', 'Top Management')  # Final Approved
    feasibility = fields.Selection([('feasible', 'Feasible'), ('feasible_with_change', 'Feasible With Change'),
                                    ('not_feasible', 'Not Feasible')], string='Feasibility')
    # Feasibility Relational
    feasibility_considerations_ids = fields.One2many(
        comodel_name='feasibility.consideration',
        inverse_name='feasibility_commitment_id',
        string="Considerations")

    feasibility_review_item_ids = fields.One2many(
        comodel_name='feasibility.review',
        inverse_name='review_id',
        string="Review")

    # feasibility_conclusion_ids = fields.One2many(
    #     comodel_name='feasibility.conclusion',
    #     inverse_name='conclusion_id',
    #     string="Conclusion")

    feasibility_sign_off_ids = fields.One2many(
        comodel_name='feasibility.sign.off',
        inverse_name='sign_off_id',
        string="Sign Off")
    # TEXT
    quality_assurance = fields.Text(string="QUALITY ASSURANCE – SUPPLIER",
                                    default='We require that the supplier will adhere to all elements described in Oerlikon Supplier Quality Expectations sheet contained in the RFQ;',translate=True)

    feasibility_review = fields.Text(string="FEASIBILITY REVIEW",
                                     default='Please complete the “Team Feasibility Commitment” form per AIAG APQP and attach along with any comments you have regarding the design of this part/system.  For ex. Spec.s missing/not understood, design alternatives or assumptions.  If no remarks, please state so !!',translate=True)

    part_development_id = fields.Many2one("part.development.process")

    # state = fields.Selection([
    #     ('draft', 'Draft'),
    #     ('hr_approve', 'HR Approval'),
    #     ('design', 'Design'),
    #     ('engineering', 'Engineering'),
    #     ('manufacturing', 'Manufacturing'),
    #     ('quality', 'Quality'),
    #     ('top', 'Top Management'),
    #     ('final_approved', 'Final Approved')
    # ], string='Status', default='draft')
    generate_xls_file = fields.Binary(string="Generated file")

    # state = fields.Selection([
    #     ('draft', 'Draft'),
    #     ('confirm', 'Confirmed'),
    # ], string='Status', default='draft')
    # document_name = fields.Char(string='Document #')
    # approve_department_id = fields.Many2one('hr.department', string='Departments Approvals')
    # # document_pro_id = fields.Many2one("xf.doc.approval.document.package", string="Document #")
    # approve_department_ids = fields.Many2many('hr.department', 'tx9', string='Departments Approvals')
    # approve_by_department_ids = fields.Many2many('hr.department', string='Departments Approved By')
    # manager_id = fields.Many2one('hr.employee', string='Approval Manager')
    # manager_ids = fields.Many2many('hr.employee', 'tx10', string='Approval Managers')
    # approvaed_manager_ids = fields.Many2many('hr.employee', string='Managers Approved By')

    # is_hr_approved = fields.Boolean()
    # is_design_approved = fields.Boolean()
    # is_eng_approved = fields.Boolean()
    # is_manu_approved = fields.Boolean()
    # is_sales_approved = fields.Boolean()
    # is_qc_approved = fields.Boolean()
    # is_maintenance_approved = fields.Boolean()
    # is_marketing_approved = fields.Boolean()
    # is_pm_approved = fields.Boolean()
    # is_management_approved = fields.Boolean()
    # is_approved_approved = fields.Boolean()

    # hr_check_department = fields.Boolean(compute='_check_department')
    # des_check_department = fields.Boolean(compute='_check_department')
    # eng_check_department = fields.Boolean(compute='_check_department')
    # manu_check_department = fields.Boolean(compute='_check_department')
    # sale_check_department = fields.Boolean(compute='_check_department')
    # qc_check_department = fields.Boolean(compute='_check_department')
    # mainte_check_department = fields.Boolean(compute='_check_department')
    # marketing_check_department = fields.Boolean(compute='_check_department')
    # pm_check_department = fields.Boolean(compute='_check_department')
    # managment_check_department = fields.Boolean(compute='_check_department')

    # link = fields.Char()

    # def actipn_confirmed(self):
    #     self.sent_for_approval()
    #     self.state = 'confirm'
    #
    # def sent_for_approval(self):
    #     mail_template = self.env.ref('iatf.mom_document_approval_mail_template')
    #     model_id = self.env['ir.model'].sudo().search([('model', '=', self._name)])
    #     mail_template.write({'model_id' : model_id.id})
    #
    #     web_base_url = self.env['ir.config_parameter'].sudo().get_param('web.base.url')
    #
    #     external_id_data_of_act_window = self.get_form_view_action_external_ids()
    #     action_id = self.env.ref(external_id_data_of_act_window[0]['external_id'], raise_if_not_found=False)
    #     # action_id = self.env.ref('iatf.action_feasibility_commitment_window', raise_if_not_found=False)
    #     link = """{}/web#id={}&view_type=form&model=self._name&action={}""".format(web_base_url,self.id,action_id.id)
    #     self.link = link
    #
    #     user_ids = self.env['hr.employee'].sudo().search([('department_id', 'in', self.approve_department_ids.ids)]).mapped('user_id')
    #     all_admin = ''
    #     for user in user_ids:
    #         if all_admin:
    #             all_admin += ',' + str(user.login)
    #         else:
    #             all_admin = str(user.login)
    #     mail_template.write({
    #             'email_to': all_admin,
    #     })
    #     mail_template.send_mail(self.id, force_send=True)
    #
    # def get_form_view_action_external_ids(self):
    #     # Get the environment
    #     env = self.env  # if used in a controller, or self.env in a model method
    #
    #     # Search for window actions associated with the given model
    #     window_actions = env['ir.actions.act_window'].search(
    #         [('res_model', '=', self._name), ('view_mode', 'like', 'form')])
    #
    #     # Initialize a list to store the details of the form view window actions
    #     form_view_action_details = []
    #
    #     for action in window_actions:
    #         # Search for the external ID in ir.model.data
    #         external_id = env['ir.model.data'].search([
    #             ('model', '=', 'ir.actions.act_window'),
    #             ('res_id', '=', action.id)
    #         ], limit=1).complete_name
    #
    #         # Store the form view action details in the list
    #         form_view_action_details.append({
    #             'name': action.name,
    #             'external_id': external_id,
    #             'view_mode': action.view_mode,
    #         })
    #     return form_view_action_details

    #
    # def _check_department(self):
    #     for rec in self:
    #         if self.env.user.employee_id.department_id.id in rec.approve_department_ids.ids and self.env.user.employee_id.department_id.department_type == 'hr':
    #             if rec.state != 'draft':
    #                 rec.hr_check_department = True
    #             else:
    #                 rec.hr_check_department = False
    #         else:
    #             rec.hr_check_department = False
    #         if self.env.user.employee_id.department_id.id in rec.approve_department_ids.ids and self.env.user.employee_id.department_id.department_type == 'design':
    #             if rec.state != 'draft':
    #                 rec.des_check_department = True
    #             else:
    #                 rec.des_check_department = False
    #         else:
    #             rec.des_check_department = False
    #         if self.env.user.employee_id.department_id.id in rec.approve_department_ids.ids and self.env.user.employee_id.department_id.department_type == 'eng':
    #             if rec.state != 'draft':
    #                 rec.eng_check_department = True
    #             else:
    #                 rec.eng_check_department = False
    #         else:
    #             rec.eng_check_department = False
    #         if self.env.user.employee_id.department_id.id in rec.approve_department_ids.ids and self.env.user.employee_id.department_id.department_type == 'manu':
    #             if rec.state != 'draft':
    #                 rec.manu_check_department = True
    #             else:
    #                 rec.manu_check_department = False
    #         else:
    #             rec.manu_check_department = False
    #         if self.env.user.employee_id.department_id.id in rec.approve_department_ids.ids and self.env.user.employee_id.department_id.department_type == 'sales':
    #             if rec.state != 'draft':
    #                 rec.sale_check_department = True
    #             else:
    #                 rec.sale_check_department = False
    #         else:
    #             rec.sale_check_department = False
    #         if self.env.user.employee_id.department_id.id in rec.approve_department_ids.ids and self.env.user.employee_id.department_id.department_type == 'qc':
    #             if rec.state != 'draft':
    #                 rec.qc_check_department = True
    #             else:
    #                 rec.qc_check_department = False
    #         else:
    #             rec.qc_check_department = False
    #         if self.env.user.employee_id.department_id.id in rec.approve_department_ids.ids and self.env.user.employee_id.department_id.department_type == 'maintenance':
    #             if rec.state != 'draft':
    #                 rec.mainte_check_department = True
    #             else:
    #                 rec.mainte_check_department = False
    #         else:
    #             rec.mainte_check_department = False
    #         if self.env.user.employee_id.department_id.id in rec.approve_department_ids.ids and self.env.user.employee_id.department_id.department_type == 'marketing':
    #             if rec.state != 'draft':
    #                 rec.marketing_check_department = True
    #             else:
    #                 rec.marketing_check_department = False
    #         else:
    #             rec.marketing_check_department = False
    #         if self.env.user.employee_id.department_id.id in rec.approve_department_ids.ids and self.env.user.employee_id.department_id.department_type == 'pm':
    #             if rec.state != 'draft':
    #                 rec.pm_check_department = True
    #             else:
    #                 rec.pm_check_department = False
    #         else:
    #             rec.pm_check_department = False
    #         if self.env.user.employee_id.department_id.id in rec.approve_department_ids.ids and self.env.user.employee_id.department_id.department_type == 'management':
    #             if rec.state != 'draft':
    #                 rec.managment_check_department = True
    #             else:
    #                 rec.managment_check_department = False
    #         else:
    #             rec.managment_check_department = False

    # hr_approval_date = fields.Datetime('Hr Approval Date')
    # design_approval_date = fields.Datetime('Design Approval Date')
    # eng_approval_date = fields.Datetime('Eng. Approval Date')
    # manu_approval_date = fields.Datetime('Production Approval Date')
    # sale_approval_date = fields.Datetime('Sales Approval Date')
    # qc_approval_date = fields.Datetime('QC Approval Date')
    # main_approval_date = fields.Datetime('Maintenance Approval Date')
    # mark_approval_date = fields.Datetime('Marketing Approval Date')
    # pm_approval_date = fields.Datetime('Program Management Approval Date')
    # maneg_approval_date = fields.Datetime('Top Management Approval Date')

    # def hr_approval(self):
    #     self.hr_approval_date = datetime.datetime.now()
    #     self.is_hr_approved = True
    #     self.approve_by_department_ids = [(4, self.env.user.employee_id.department_id.id)]
    #     self.approvaed_manager_ids = [(4, self.env.user.employee_id.department_id.manager_id.id)]
    # def design_approval(self):
    #     self.design_approval_date = datetime.datetime.now()
    #     self.is_design_approved = True
    #     self.approve_by_department_ids = [(4, self.env.user.employee_id.department_id.id)]
    #     self.approvaed_manager_ids = [(4, self.env.user.employee_id.department_id.manager_id.id)]
    # def eng_approval(self):
    #     self.eng_approval_date = datetime.datetime.now()
    #     self.is_eng_approved = True
    #     self.approve_by_department_ids = [(4, self.env.user.employee_id.department_id.id)]
    #     self.approvaed_manager_ids = [(4, self.env.user.employee_id.department_id.manager_id.id)]
    # def manu_approval(self):
    #     self.manu_approval_date = datetime.datetime.now()
    #     self.is_manu_approved = True
    #     self.approve_by_department_ids = [(4, self.env.user.employee_id.department_id.id)]
    #     self.approvaed_manager_ids = [(4, self.env.user.employee_id.department_id.manager_id.id)]
    # def sales_approval(self):
    #     self.sale_approval_date = datetime.datetime.now()
    #     self.is_sales_approved = True
    #     self.approve_by_department_ids = [(4, self.env.user.employee_id.department_id.id)]
    #     self.approvaed_manager_ids = [(4, self.env.user.employee_id.department_id.manager_id.id)]
    # def qc_approval(self):
    #     self.qc_approval_date = datetime.datetime.now()
    #     self.is_qc_approved = True
    #     self.approve_by_department_ids = [(4, self.env.user.employee_id.department_id.id)]
    #     self.approvaed_manager_ids = [(4, self.env.user.employee_id.department_id.manager_id.id)]
    # def mainte_approval(self):
    #     self.main_approval_date = datetime.datetime.now()
    #     self.is_maintenance_approved = True
    #     self.approve_by_department_ids = [(4, self.env.user.employee_id.department_id.id)]
    #     self.approvaed_manager_ids = [(4, self.env.user.employee_id.department_id.manager_id.id)]
    # def marketing_approval(self):
    #     self.mark_approval_date = datetime.datetime.now()
    #     self.is_marketing_approved = True
    #     self.approve_by_department_ids = [(4, self.env.user.employee_id.department_id.id)]
    #     self.approvaed_manager_ids = [(4, self.env.user.employee_id.department_id.manager_id.id)]
    # def pm_approval(self):
    #     self.pm_approval_date = datetime.datetime.now()
    #     self.is_pm_approved = True
    #     self.approve_by_department_ids = [(4, self.env.user.employee_id.department_id.id)]
    #     self.approvaed_manager_ids = [(4, self.env.user.employee_id.department_id.manager_id.id)]
    # def managment_approval(self):
    #     self.maneg_approval_date = datetime.datetime.now()
    #     self.is_management_approved = True
    #     self.approve_by_department_ids = [(4, self.env.user.employee_id.department_id.id)]
    #     self.approvaed_manager_ids = [(4, self.env.user.employee_id.department_id.manager_id.id)]

    @api.model
    def create(self, vals):
        default_considerations = self.env['feasibility.consideration'].search([('default', '=', True)])
        if default_considerations:
            default_consideration_lines = []
            for consideration_point in default_considerations:
                default_consideration_lines.append((0, 0, {'name': consideration_point.name}))
            vals['feasibility_considerations_ids'] = default_consideration_lines
        return super(FeasibilityCommitment, self).create(vals)

    # def button_send_approval(self):
    #     for rec in self:
    #         # Mail Send for approval
    #         if rec.hr:
    #             user_email = rec.hr.login
    #             mail_temp = self.env.ref('iatf.sending_mail_template_risk_assessment')
    #             if mail_temp:
    #                 format_name = rec.format_id.name
    #                 cust_name = rec.cust_id.name
    #                 subject = f" Request for approving Risk Assessment {format_name} for {cust_name}"
    #                 if user_email:
    #                     body = """
    #                     <div>
    #                         <p>Dear Recipient  """ + str(user_email) + """,
    #                             <br/><br/>
    #                             Please, kindly Request To accept and approved Sending Approval for the Risk Analysis Sheet.
    #                         <br></br>
    #                         Thank you.
    #                     <br/>
    #                     <br/>
    #                     <div>"""
    #                     mail_temp.send_mail(self.id, email_values={
    #                         'email_from': self.env.user.login,
    #                         'email_to': user_email,
    #                         'subject': subject,
    #                         'body_html': body,
    #                     }, force_send=True)
    #
    #         rec.write({'state': 'hr_approve'})
    #         rec.part_development_id.feasibility_commitment_id = rec.id
    #
    # def button_hr_approval(self):
    #     for rec in self:
    #         # Mail Send for hr approval
    #         if rec.hr:
    #             user_email = rec.hr.login
    #             mail_temp = self.env.ref('iatf.sending_mail_template_risk_assessment')
    #             if mail_temp:
    #                 format_name = rec.format_id.name
    #                 cust_name = rec.cust_id.name
    #                 subject = f" Request for approving Risk Assessment {format_name} for {cust_name}"
    #                 if user_email:
    #                     body = """
    #                     <div>
    #                         <p>Dear Recipient  """ + str(user_email) + """,
    #                             <br/><br/>
    #                             Please, kindly Request To accept and approved Sending Approval for the Risk Analysis Sheet.
    #                         <br></br>
    #                         Thank you.
    #                     <br/>
    #                     <br/>
    #                     <div>"""
    #                     mail_temp.send_mail(self.id, email_values={
    #                         'email_from': self.env.user.login,
    #                         'email_to': user_email,
    #                         'subject': subject,
    #                         'body_html': body,
    #                     }, force_send=True)
    #         rec.write({'state': 'design'})
    #
    # def button_design(self):
    #     for rec in self:
    #         # Mail Send for Design approval
    #         if rec.design_eng:
    #             user_email = rec.design_eng.login
    #             mail_temp = self.env.ref('iatf.sending_mail_template_risk_assessment')
    #             if mail_temp:
    #                 format_name = rec.format_id.name
    #                 cust_name = rec.cust_id.name
    #                 subject = f" Request for approving Risk Assessment {format_name} for {cust_name}"
    #                 if user_email:
    #                     body = """
    #                     <div>
    #                         <p>Dear Recipient  """ + str(user_email) + """,
    #                             <br/><br/>
    #                             Please, kindly Request To accept and approved Sending Approval for the Risk Analysis Sheet.
    #                         <br></br>
    #                         Thank you.
    #                     <br/>
    #                     <br/>
    #                     <div>"""
    #                     mail_temp.send_mail(self.id, email_values={
    #                         'email_from': self.env.user.login,
    #                         'email_to': user_email,
    #                         'subject': subject,
    #                         'body_html': body,
    #                     }, force_send=True)
    #         rec.write({'state': 'engineering'})
    #
    # def button_engineering(self):
    #     for rec in self:
    #         # Mail Send for Engineering approval
    #         if rec.manf_eng:
    #             user_email = rec.manf_eng.login
    #             mail_temp = self.env.ref('iatf.sending_mail_template_risk_assessment')
    #             if mail_temp:
    #                 format_name = rec.format_id.name
    #                 cust_name = rec.cust_id.name
    #                 subject = f" Request for approving Risk Assessment {format_name} for {cust_name}"
    #                 if user_email:
    #                     body = """
    #                     <div>
    #                         <p>Dear Recipient  """ + str(user_email) + """,
    #                             <br/><br/>
    #                             Please, kindly Request To accept and approved Sending Approval for the Risk Analysis Sheet.
    #                         <br></br>
    #                         Thank you.
    #                     <br/>
    #                     <br/>
    #                     <div>"""
    #                     mail_temp.send_mail(self.id, email_values={
    #                         'email_from': self.env.user.login,
    #                         'email_to': user_email,
    #                         'subject': subject,
    #                         'body_html': body,
    #                     }, force_send=True)
    #         rec.write({'state': 'manufacturing'})
    #
    # def button_manufacturing(self):
    #     for rec in self:
    #         # Mail Send for Manufacturing approval
    #         if rec.production:
    #             user_email = rec.production.login
    #             mail_temp = self.env.ref('iatf.sending_mail_template_risk_assessment')
    #             if mail_temp:
    #                 format_name = rec.format_id.name
    #                 cust_name = rec.cust_id.name
    #                 subject = f" Request for approving Risk Assessment {format_name} for {cust_name}"
    #                 if user_email:
    #                     body = """
    #                     <div>
    #                         <p>Dear Recipient  """ + str(user_email) + """,
    #                             <br/><br/>
    #                             Please, kindly Request To accept and approved Sending Approval for the Risk Analysis Sheet.
    #                         <br></br>
    #                         Thank you.
    #                     <br/>
    #                     <br/>
    #                     <div>"""
    #                     mail_temp.send_mail(self.id, email_values={
    #                         'email_from': self.env.user.login,
    #                         'email_to': user_email,
    #                         'subject': subject,
    #                         'body_html': body,
    #                     }, force_send=True)
    #         rec.write({'state': 'quality'})
    #
    # def button_quality_test_done(self):
    #     for rec in self:
    #         # Mail Send for Quality Test approval
    #         if rec.quality:
    #             user_email = rec.quality.login
    #             mail_temp = self.env.ref('iatf.sending_mail_template_risk_assessment')
    #             if mail_temp:
    #                 format_name = rec.format_id.name
    #                 cust_name = rec.cust_id.name
    #                 subject = f" Request for approving Risk Assessment {format_name} for {cust_name}"
    #                 if user_email:
    #                     body = """
    #                     <div>
    #                         <p>Dear Recipient  """ + str(user_email) + """,
    #                             <br/><br/>
    #                             Please, kindly Request To accept and approved Sending Approval for the Risk Analysis Sheet.
    #                         <br></br>
    #                         Thank you.
    #                     <br/>
    #                     <br/>
    #                     <div>"""
    #                     mail_temp.send_mail(self.id, email_values={
    #                         'email_from': self.env.user.login,
    #                         'email_to': user_email,
    #                         'subject': subject,
    #                         'body_html': body,
    #                     }, force_send=True)
    #         rec.write({'state': 'top'})
    #
    # def button_top_managment_final_approved(self):
    #     for rec in self:
    #         # Mail Send for Final approval
    #         if rec.top_management_id:
    #             user_email = rec.top_management_id.login
    #             mail_temp = self.env.ref('iatf.sending_mail_template_risk_assessment')
    #             if mail_temp:
    #                 format_name = rec.format_id.name
    #                 cust_name = rec.cust_id.name
    #                 subject = f" Request for approving Risk Assessment {format_name} for {cust_name}"
    #                 if user_email:
    #                     body = """
    #                     <div>
    #                         <p>Dear Recipient  """ + str(user_email) + """,
    #                             <br/><br/>
    #                             Please, kindly Request To accept and approved Sending Approval for the Risk Analysis Sheet.
    #                         <br></br>
    #                         Thank you.
    #                     <br/>
    #                     <br/>
    #                     <div>"""
    #                     mail_temp.send_mail(self.id, email_values={
    #                         'email_from': self.env.user.login,
    #                         'email_to': user_email,
    #                         'subject': subject,
    #                         'body_html': body,
    #                     }, force_send=True)
    #         rec.write({'state': 'final_approved'})

    def generate_excel_report(self):
        output = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Feasibility Considerations"
        wb.create_sheet("Quality Assurance")

        # region formatting data
        thin = Side(border_style='thin', color='000000')
        thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
        border = Border(
            top=Side(style='thin'),
            left=Side(style='thin'),
            right=Side(style='thin'),
            bottom=Side(style='thin')
        )

        align_center = Alignment(vertical='center', horizontal='center', wrapText=True)
        align_left = Alignment(vertical='center', horizontal='left', wrapText=True)
        align_right = Alignment(vertical='center', horizontal='right', wrapText=True)

        font_main_header = Font(name='Times New Roman', size=16, bold=True)
        font_header = Font(name='Times New Roman', size=11, bold=True)
        font_all = Font(name='Times New Roman', size=11, bold=False)
        # endregion

        {k: setattr(DEFAULT_FONT, k, v) for k, v in font_all.__dict__.items()}

        # region feasibility Consideration sheet

        ws.row_dimensions[1].height = 75
        ws.row_dimensions[7].height = 75

        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 17
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['H'].width = 20

        ws['A1'] = 'TEAM FEASIBILITY COMMITMENT REPORT'
        ws['A1'].alignment = align_center
        ws['A1'].font = font_main_header

        # region Adding Logo
        if self.env.user.company_id.logo:
            max_width = 200  # Set your desired maximum width
            max_height = 100  # Set your desired maximum height
            image_data = base64.b64decode(self.env.user.company_id.logo)

            # Open the image using PIL
            image = PILImage.open(io.BytesIO(image_data))
            width, height = image.size
            aspect_ratio = width / height

            if width > max_width:
                width = max_width
                height = int(width / aspect_ratio)

            if height > max_height:
                height = max_height
                width = int(height * aspect_ratio)

            # Resize the image using PIL
            # Add space on the top and left side of the image
            padding_top = 10  # Adjust as needed
            padding_left = 10  # Adjust as needed

            resized_image = image.resize((width, height), PILImage.LANCZOS)
            ImageOps.expand(resized_image, border=(padding_left, padding_top, 0, 0), fill='rgba(0,0,0,0)')
            img_bytes = io.BytesIO()
            resized_image.save(img_bytes, format='PNG')
            img_bytes.seek(0)
            logo_image = Image(img_bytes)
            # logo_image = Image(self.env.user.company_id.logo)
            ws.add_image(logo_image, 'A1')
        # endregion

        data = {
            'A2': 'Customer',
            'E2': 'Date',
            'A3': 'Part Number',
            'E3': 'Part Name',
            'A4': 'Revision Date',
            'E4': 'Revision No.',
            'A6': 'Feasibility Considerations',
            'A7': 'Our product quality planning team has considered the following questions, not intended to be all-inclusive in performing a feasibility evaluation.  The drawings and/or specifications provided have been used as a basis for analyzing the ability to meet all specified requirements.  All “no” answers are supported with attached comments identifying our concerns and/or proposed changes to enable us to meet the specified requirements.',
            'A8': 'YES',
            'B8': 'NO',
            'C8': 'CONSIDERATION',
        }
        for cell, value in data.items():
            ws[cell] = value
            ws[cell].font = font_header
            ws[cell].alignment = align_center

        ws['A6'].alignment = align_left
        ws['A7'].alignment = align_left
        ws['A7'].font = font_all

        cells_to_merge = [
            'A1:H1', 'A2:B2', 'C2:D2', 'E2:F2', 'G2:H2',
            'A3:B3', 'C3:D3', 'E3:F3', 'G3:H3',
            'A4:B4', 'C4:D4', 'E4:F4', 'G4:H4',
            'A5:H5', 'A6:H6', 'A7:H7', 'C7:H7',
            'C8:H8',
        ]
        for cell_range in cells_to_merge:
            ws.merge_cells(cell_range)

        start_row = curr_row = 9
        mx_row = 25

        feasibility = None

        sign_off_data = []
        for rec in self:
            ws['C2'] = rec.cust_id.name if rec.cust_id else ''
            ws['G2'] = rec.date.strftime("%d-%m-%y") if rec.date else ''
            ws['C3'] = rec.part_num if rec.part_num else ''
            ws['G3'] = rec.part_id.name if rec.part_id else ''
            ws['C4'] = rec.revision_date.strftime("%d-%m-%y") if rec.revision_date else ''
            ws['G4'] = rec.revision_number if rec.revision_number else ''
            feasibility = rec.feasibility if rec.feasibility else ''

            for fc in rec.feasibility_considerations_ids:
                ws[f'C{curr_row}'] = fc.name if fc.name else ''
                if fc.is_yes:
                    ws[f'A{curr_row}'] = '☑'
                    ws[f'A{curr_row}'].alignment = align_center
                else:
                    ws[f'B{curr_row}'] = '☑'
                    ws[f'B{curr_row}'].alignment = align_center
                curr_row += 1

            for member in rec.feasibility_sign_off_ids:
                sign_off_data.append(
                    {
                        'member': member.sign_off_team_member if member.sign_off_team_member else '',
                        'title': member.sign_off_title if member.sign_off_title else '',
                        'date_': member.sign_off_date if member.sign_off_date else ''
                    }
                )
        if curr_row > mx_row:
            mx_row = curr_row

        for i in range(start_row, mx_row):
            ws.merge_cells(f"C{i}:H{i}")

        ws.merge_cells(f"A{mx_row}:H{mx_row}")
        ws.merge_cells(f"B{mx_row + 2}:C{mx_row + 2}")
        ws.merge_cells(f"B{mx_row + 3}:C{mx_row + 3}")
        ws.merge_cells(f"B{mx_row + 4}:C{mx_row + 4}")
        ws.merge_cells(f"A{mx_row + 1}:H{mx_row + 1}")
        ws.merge_cells(f"D{mx_row + 2}:H{mx_row + 2}")
        ws.merge_cells(f"D{mx_row + 3}:H{mx_row + 3}")
        ws.merge_cells(f"D{mx_row + 4}:H{mx_row + 4}")
        ws.merge_cells(f"A{mx_row + 5}:H{mx_row + 5}")

        ws[f'A{mx_row + 1}'] = 'Conclusion'
        ws[f'B{mx_row + 2}'] = 'Feasible'
        ws[f'D{mx_row + 2}'] = 'Product can be produced as specified with no revisions'
        ws[f'B{mx_row + 3}'] = 'Feasible with Changes'
        ws[f'D{mx_row + 3}'] = 'Changes recommended (see attached)'
        ws[f'B{mx_row + 4}'] = 'Not Feasible'
        ws[f'D{mx_row + 4}'] = 'Design revision required to produce product within the specified requirements'

        ws[f'A{mx_row + 1}'].alignment = align_left
        ws[f'A{mx_row + 1}'].font = font_header

        if feasibility == 'feasible':
            ws[f'A{mx_row + 2}'] = '☑'
            ws[f'A{mx_row + 2}'].alignment = align_center
        elif feasibility == 'feasible_with_change':
            ws[f'A{mx_row + 3}'] = '☑'
            ws[f'A{mx_row + 3}'].alignment = align_center
        elif feasibility == 'not_feasible':
            ws[f'A{mx_row + 4}'] = '☑'
            ws[f'A{mx_row + 4}'].alignment = align_center

        mx_row = mx_row + 5

        for row in ws.iter_rows(min_row=0, max_row=mx_row, min_col=1, max_col=8):
            for cell in row:
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

        cur_row = mx_row + 1
        # region SignOff Members Footer
        ws[f'A{cur_row}'] = 'Prepared By'
        ws[f'A{cur_row}'].font = font_header
        ws.merge_cells(f'A{cur_row}:B{cur_row}')
        ws.merge_cells(f'C{cur_row}:D{cur_row}')

        ws[f'E{cur_row}'] = 'Prepared Date'
        ws[f'E{cur_row}'].font = font_header
        ws.merge_cells(f'E{cur_row}:F{cur_row}')
        ws.merge_cells(f'G{cur_row}:H{cur_row}')

        for rec in self:
            ws[f'C{cur_row}'] = rec.create_uid.name if rec.create_uid else ''
            ws[f'G{cur_row}'] = rec.create_date if rec.create_date else ''

        ws.row_dimensions[cur_row].height = 18
        cur_row += 1
        ws.merge_cells(f'A{cur_row}:H{cur_row}')
        cur_row += 1
        ws.merge_cells(f'A{cur_row}:H{cur_row}')
        ws[f'A{cur_row}'] = 'Sign OFF'
        ws[f'A{cur_row}'].font = Font(size=18, bold=True)
        ws.row_dimensions[cur_row].height = 25
        cur_row += 1
        ws.merge_cells(f'A{cur_row}:H{cur_row}')
        cur_row += 1
        for cell in ws[cur_row]:
            ws[f'A{cur_row}'] = 'Member'
            ws[f'D{cur_row}'] = 'Department'
            ws[f'E{cur_row}'] = 'Status'
            ws[f'F{cur_row}'] = 'Date'
            ws[f'G{cur_row}'] = 'Comments'
            ws.row_dimensions[cur_row].height = 25
            ws.merge_cells(f'A{cur_row}:C{cur_row}')
            ws.merge_cells(f'G{cur_row}:H{cur_row}')
            cell.font = Font(size=12, bold=True, color='ffffff')
            cell.fill = PatternFill(start_color='0070C0', end_color='0070C0', fill_type="solid")

        for row_no in ws.iter_rows(min_row=1, max_row=cur_row + 1, min_col=1, max_col=8):
            for cell in row_no:
                cell.border = border
                cell.alignment = align_center

        sign_row = cur_row

        # Listing Managers Id
        for rec in self:
            for record in rec.iatf_members_ids:
                name_rec = record.approver_id.name
                dept_rec = record.department_id.name if record.department_id.name else ''
                status_rec = record.approval_status.capitalize()
                date_rec = record.date_approved_rejected if record.date_approved_rejected else ''
                comment_rec = record.comment if record.comment else ''

                ws[f'A{cur_row + 1}'] = name_rec
                ws[f'D{cur_row + 1}'] = dept_rec
                ws[f'E{cur_row + 1}'] = status_rec
                ws[f'F{cur_row + 1}'] = date_rec
                ws[f'G{cur_row + 1}'] = comment_rec
                ws.merge_cells(f'A{cur_row + 1}:C{cur_row + 1}')
                ws.merge_cells(f'G{cur_row + 1}:H{cur_row + 1}')

                status_dict = {'Approved': ['CFFFC3', '000000'], 'Rejected': ['FFCDCD', '000000'],
                               'Revision': ['AFEFFF', '000000'], 'Pending': ['FDFFCD', '000000']}

                for state, color in status_dict.items():
                    if state == status_rec:
                        ws[f'E{cur_row + 1}'].fill = PatternFill(start_color=color[0], end_color=color[0],
                                                                 fill_type="solid")
                        ws[f'E{cur_row + 1}'].font = Font(size=12, bold=False, color=color[1])

                cur_row += 1

        ws.merge_cells(f'A{cur_row + 1}:H{cur_row + 1}')

        for row_no in ws.iter_rows(min_row=sign_row + 1, max_row=cur_row + 1, min_col=1, max_col=8):
            for cell in row_no:
                cell.border = border
                cell.alignment = align_center
        # endregion

        # region Quality Assurance sheet
        ws = wb["Quality Assurance"]

        data = {
            'A2': 'RFQ Number',
            'A4': 'Quality Assurance - Supplier',
            'A6': 'Feasibility Review',
            'A8': 'ITEM', 'B8': 'Remarks', 'E8': 'Actions from Customer (mandatory if feasibility review)',
        }
        for cell, value in data.items():
            ws[cell] = value
            ws[cell].font = font_header
            ws[cell].alignment = align_left

        ws['A5'].alignment = align_left
        ws['A5'].font = font_all
        ws['A7'].alignment = align_left
        ws['A7'].font = font_all
        ws['A8'].alignment = align_center
        # ws['A8'].font = font_all

        cells_to_merge = [
            'A1:H1', 'A2:B2', 'C2:H2',
            'A3:H3',
            'A4:H4',
            'A5:H5', 'A6:H6', 'A7:H7',
            'B8:D8', 'E8:H8'
        ]
        for cell_range in cells_to_merge:
            ws.merge_cells(cell_range)

        ws.row_dimensions[1].height = 75
        ws.row_dimensions[5].height = 50
        ws.row_dimensions[7].height = 70

        ws.column_dimensions['A'].width = 13
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['H'].width = 21

        ws['A1'] = 'TEAM FEASIBILITY COMMITMENT REPORT'
        ws['A1'].alignment = align_center
        ws['A1'].font = font_main_header
        # region Adding Logo
        if self.env.user.company_id.logo:
            max_width = 200  # Set your desired maximum width
            max_height = 100  # Set your desired maximum height
            image_data = base64.b64decode(self.env.user.company_id.logo)

            # Open the image using PIL
            image = PILImage.open(io.BytesIO(image_data))
            width, height = image.size
            aspect_ratio = width / height

            if width > max_width:
                width = max_width
                height = int(width / aspect_ratio)

            if height > max_height:
                height = max_height
                width = int(height * aspect_ratio)

            # Resize the image using PIL
            # Add space on the top and left side of the image
            padding_top = 10  # Adjust as needed
            padding_left = 10  # Adjust as needed

            resized_image = image.resize((width, height),PILImage.LANCZOS)
            ImageOps.expand(resized_image, border=(padding_left, padding_top, 0, 0), fill='rgba(0,0,0,0)')
            img_bytes = io.BytesIO()
            resized_image.save(img_bytes, format='PNG')
            img_bytes.seek(0)
            logo_image = Image(img_bytes)
            # logo_image = Image(self.env.user.company_id.logo)
            ws.add_image(logo_image, 'A1')
        # endregion

        start_row = curr_row = 9
        mx_row = 25
        company_name = ''
        for rec in self:
            ws['C2'] = rec.rfq_no if rec.rfq_no else ''
            ws['A5'] = rec.quality_assurance if rec.quality_assurance else ''
            ws['A7'] = rec.feasibility_review if rec.feasibility_review else ''
            company_name = rec.company_id.name if rec.company_id else ''

            for fr in self.feasibility_review_item_ids:
                ws[f'A{curr_row}'] = fr.item if fr.item else ''
                ws[f'B{curr_row}'] = fr.remarks if fr.remarks else ''
                ws[f'E{curr_row}'] = fr.actions_from_customer if fr.actions_from_customer else ''
                ws.merge_cells(f'B{curr_row}:D{curr_row}')
                ws.merge_cells(f'E{curr_row}:H{curr_row}')

                ws[f'A{curr_row}'].font = font_all
                ws[f'A{curr_row}'].alignment = align_center
                ws[f'B{curr_row}'].font = font_all
                ws[f'B{curr_row}'].alignment = align_left
                ws[f'E{curr_row}'].font = font_all
                ws[f'E{curr_row}'].alignment = align_left
                curr_row += 1

        if curr_row > mx_row:
            mx_row = curr_row
        for i in range(start_row, mx_row):
            ws.merge_cells(f"B{i}:D{i}")
            ws.merge_cells(f"E{i}:H{i}")

        ws.merge_cells(f"A{mx_row}:H{mx_row}")
        ws.merge_cells(f"A{mx_row + 1}:C{mx_row + 1}")
        ws.merge_cells(f"E{mx_row + 1}:F{mx_row + 1}")
        ws.merge_cells(f"G{mx_row + 1}:H{mx_row + 1}")
        ws.merge_cells(f"A{mx_row + 2}:H{mx_row + 2}")
        ws.merge_cells(f"A{mx_row + 3}:B{mx_row + 3}")
        ws.merge_cells(f"C{mx_row + 3}:H{mx_row + 3}")
        ws.merge_cells(f"A{mx_row + 4}:H{mx_row + 4}")

        ws[f'A{mx_row + 1}'] = 'Part Specific Quality Requirement?'
        ws[f'A{mx_row + 1}'].font = font_header
        ws[f'E{mx_row + 1}'] = 'Yes, see Enclosed'
        ws[f'E{mx_row + 1}'].font = font_header
        ws[f'A{mx_row + 3}'] = 'Company'
        ws[f'A{mx_row + 3}'].font = font_header
        ws[f'C{mx_row + 3}'] = company_name
        mx_row += 4

        cur_row = mx_row + 1
        # region SignOff Members Footer
        ws[f'A{cur_row}'] = 'Prepared By'
        ws[f'A{cur_row}'].font = font_header
        ws.merge_cells(f'A{cur_row}:B{cur_row}')
        ws.merge_cells(f'C{cur_row}:D{cur_row}')

        ws[f'E{cur_row}'] = 'Prepared Date'
        ws[f'E{cur_row}'].font = font_header
        ws.merge_cells(f'E{cur_row}:F{cur_row}')
        ws.merge_cells(f'G{cur_row}:H{cur_row}')

        for rec in self:
            ws[f'C{cur_row}'] = rec.create_uid.name if rec.create_uid else ''
            ws[f'G{cur_row}'] = rec.create_date if rec.create_date else ''

        ws.row_dimensions[cur_row].height = 18
        cur_row += 1
        ws.merge_cells(f'A{cur_row}:H{cur_row}')
        cur_row += 1
        ws.merge_cells(f'A{cur_row}:H{cur_row}')
        ws[f'A{cur_row}'] = 'Sign OFF'
        ws[f'A{cur_row}'].font = Font(size=18, bold=True)
        ws.row_dimensions[cur_row].height = 25
        cur_row += 1
        ws.merge_cells(f'A{cur_row}:H{cur_row}')
        cur_row += 1
        for cell in ws[cur_row]:
            ws[f'A{cur_row}'] = 'Member'
            ws[f'D{cur_row}'] = 'Department'
            ws[f'E{cur_row}'] = 'Status'
            ws[f'F{cur_row}'] = 'Date'
            ws[f'G{cur_row}'] = 'Comments'
            ws.row_dimensions[cur_row].height = 25
            ws.merge_cells(f'A{cur_row}:C{cur_row}')
            ws.merge_cells(f'G{cur_row}:H{cur_row}')
            cell.font = Font(size=12, bold=True, color='ffffff')
            cell.fill = PatternFill(start_color='0070C0', end_color='0070C0', fill_type="solid")

        for row_no in ws.iter_rows(min_row=1, max_row=cur_row + 1, min_col=1, max_col=8):
            for cell in row_no:
                cell.border = border
                cell.alignment = align_center

        sign_row = cur_row

        # Listing Managers Id
        for rec in self:
            for record in rec.iatf_members_ids:
                name_rec = record.approver_id.name
                dept_rec = record.department_id.name if record.department_id.name else ''
                status_rec = record.approval_status.capitalize()
                date_rec = record.date_approved_rejected if record.date_approved_rejected else ''
                comment_rec = record.comment if record.comment else ''

                ws[f'A{cur_row + 1}'] = name_rec
                ws[f'D{cur_row + 1}'] = dept_rec
                ws[f'E{cur_row + 1}'] = status_rec
                ws[f'F{cur_row + 1}'] = date_rec
                ws[f'G{cur_row + 1}'] = comment_rec
                ws.merge_cells(f'A{cur_row + 1}:C{cur_row + 1}')
                ws.merge_cells(f'G{cur_row + 1}:H{cur_row + 1}')

                status_dict = {'Approved': ['CFFFC3', '000000'], 'Rejected': ['FFCDCD', '000000'],
                               'Revision': ['AFEFFF', '000000'], 'Pending': ['FDFFCD', '000000']}

                for state, color in status_dict.items():
                    if state == status_rec:
                        ws[f'E{cur_row + 1}'].fill = PatternFill(start_color=color[0], end_color=color[0],
                                                                 fill_type="solid")
                        ws[f'E{cur_row + 1}'].font = Font(size=12, bold=False, color=color[1])

                cur_row += 1

        ws.merge_cells(f'A{cur_row + 1}:H{cur_row + 1}')

        for row_no in ws.iter_rows(min_row=sign_row + 1, max_row=cur_row + 1, min_col=1, max_col=8):
            for cell in row_no:
                cell.border = border
                cell.alignment = align_center
        # endregion

        # endregion

        # region Save the workbook
        wb.save(output)
        output.seek(0)
        self.generate_xls_file = base64.b64encode(output.getvalue()).decode('utf-8')
        # endregion
        return {
            "type": "ir.actions.act_url",
            "target": "self",
            "url": "/web/content?model=feasibility.commitment&download=true&field=generate_xls_file&filename={filename}.xlsx&id={pid}".format(
                filename="Feasibility Commitment", pid=self[0].id),
        }

    # def action_print_attached_model(self):
    #     # Get the environment
    #     env = self.env  # if used in a controller, or self.env in a model method
    #
    #     # Search for window actions associated with the given model
    #     window_actions = env['ir.actions.act_window'].search([('res_model', '=', self._name)])
    #
    #     # Initialize a dictionary to store the external IDs
    #     window_action_external_ids = {}
    #
    #     for action in window_actions:
    #         # Search for the external ID in ir.model.data
    #         print(action.name)
    #         print(action.id)
    #         external_id = env['ir.model.data'].search([
    #             ('model', '=', 'ir.actions.act_window'),
    #             ('res_id', '=', action.id)
    #         ], limit=1).complete_name
    #
    #         # Store the external ID in the dictionary
    #         window_action_external_ids[action.name] = external_id
    #
    #     print(window_action_external_ids)


class FeasibilityConsideration(models.Model):
    _name = 'feasibility.consideration'
    _description = 'Feasibility Consideration'
    # _inherit = "translation.mixin"

    feasibility_commitment_id = fields.Many2one('feasibility.commitment', string='Feasibility')
    name = fields.Char(string='Consideration', required=True)
    is_yes = fields.Boolean(string='Yes')
    default = fields.Boolean(readonly=True, default=False)
    description = fields.Text(string='Description')


class FeasibilityConclusion(models.Model):
    _name = 'feasibility.conclusion'
    _description = 'Feasibility Conclusion'
    _inherit = "translation.mixin"

    # conclusion_id = fields.Many2one('feasibility.commitment', string='Feasibility')
    select_conclusion = fields.Selection(
        [('feasible', 'Feasible'), ('feasible_with_changes', 'Feasible With Changes'), ('not_easible', 'Not Feasible')],
        string='Feasibility')
    remarks = fields.Text(string='Remarks',translate=True)


class FeasibilitySignOff(models.Model):
    _name = 'feasibility.sign.off'
    _description = 'Feasibility Sign Off'
    _inherit = "translation.mixin"

    sign_off_id = fields.Many2one('feasibility.commitment', string='Feasibility')
    sign_off_team_member = fields.Char(string='Team Member',translate=True)
    sign_off_title = fields.Char(string='Title',translate=True)
    sign_off_date = fields.Date(string='Date')


class FeasibilityReview(models.Model):
    _name = 'feasibility.review'
    _description = 'Feasibility Review'
    _inherit = "translation.mixin"

    review_id = fields.Many2one('feasibility.commitment', string='Feasibility')
    item = fields.Char(string='Item', required=True,translate=True)
    remarks = fields.Text(string='Remarks',translate=True)
    is_feasibility = fields.Boolean(string='Feasibility')
    actions_from_customer = fields.Text(string='Actions from Customer',translate=True)
