import base64
import io
from io import BytesIO

from PIL import Image as PILImage
from PIL import ImageOps
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill

from odoo import fields, models


class QuaryList(models.Model):
    _name = 'quary.list'
    _description = 'Query List'
    _rec_name = "project_id"
    _inherit = ['iatf.sign.off.members']

    # project_name = fields.Char("Project")
    # partner_id = fields.Many2one('res.partner', string="Customer Details")
    revision = fields.Selection(selection=[('r1', 'R1'), ('r2', 'R2'), ('r3', 'R3'), ('r4', 'R4')], string="Revision")
    first_compilation_date = fields.Date("First Compilation Date")
    first_disposition_date = fields.Date("First Disposition Date")
    second_disposition_date = fields.Date("Second Disposition Date")
    third_disposition_date = fields.Date("Third Disposition Date")
    fourth_disposition_date = fields.Date("Fourth Disposition Date")

    quary_list_line_ids = fields.One2many(
        comodel_name='quary.list.line',
        inverse_name='quart_list_id',
        string="Query List Line"
    )

    # state = fields.Selection([
    #     ('draft', 'Draft'),
    #     ('hr_approve', 'HR Approval'),
    #     ('design', 'Design'),
    #     ('engineering', 'Engineering'),
    #     ('manufacturing', 'Manufacturing'),
    #     ('quality', 'Quality'),
    #     ('top', 'Top Management'),
    #     ('final_approved', 'Final Approved')
    # ], string='Status', default='draft')

    # hr = fields.Many2one('res.users', 'HR')  # HR APPROVED
    # design_eng = fields.Many2one('res.users', 'Design Engineering')  # DESIGN
    # manf_eng = fields.Many2one('res.users', 'Manufacturing Engineering')  # Engineering
    # production = fields.Many2one('res.users', 'Production')  # Manufacturing
    # quality = fields.Many2one('res.users', 'Quality')  # Qulity
    # top_management_id = fields.Many2one('res.users', 'Top Management')  # Final Approved
    # part_development_id = fields.Many2one("part.development.process")
    team_id = fields.Many2one('maintenance.team', string='Internal Team', ondelete="restrict")

    generate_xls_file = fields.Binary(string="Generated file")

    # state = fields.Selection([
    #     ('draft', 'Draft'),
    #     ('confirm', 'Confirmed'),
    # ], string='Status', default='draft')
    document_name = fields.Char(string='Document #')

    # approve_department_id = fields.Many2one('hr.department', string='Departments Approvals')
    # # document_pro_id = fields.Many2one("xf.doc.approval.document.package", string="Document #")
    # approve_department_ids = fields.Many2many('hr.department', 'tb3', string='Departments Approvals')
    # approve_by_department_ids = fields.Many2many('hr.department', string='Departments Approved By')
    # manager_id = fields.Many2one('hr.employee', string='Approval Manager')
    # manager_ids = fields.Many2many('hr.employee', 'tb4', string='Approval Managers')
    # approvaed_manager_ids = fields.Many2many('hr.employee', string='Managers Approved By')

    #
    # is_hr_approved = fields.Boolean()
    # is_design_approved = fields.Boolean()
    # is_eng_approved = fields.Boolean()
    # is_manu_approved = fields.Boolean()
    # is_sales_approved = fields.Boolean()
    # is_qc_approved = fields.Boolean()
    # is_maintenance_approved = fields.Boolean()
    # is_marketing_approved = fields.Boolean()
    # is_pm_approved = fields.Boolean()
    # is_management_approved = fields.Boolean()
    # is_approved_approved = fields.Boolean()
    #
    # hr_check_department = fields.Boolean(compute='_check_department')
    # des_check_department = fields.Boolean(compute='_check_department')
    # eng_check_department = fields.Boolean(compute='_check_department')
    # manu_check_department = fields.Boolean(compute='_check_department')
    # sale_check_department = fields.Boolean(compute='_check_department')
    # qc_check_department = fields.Boolean(compute='_check_department')
    # mainte_check_department = fields.Boolean(compute='_check_department')
    # marketing_check_department = fields.Boolean(compute='_check_department')
    # pm_check_department = fields.Boolean(compute='_check_department')
    # managment_check_department = fields.Boolean(compute='_check_department')
    #
    # link = fields.Char()
    #
    # def actipn_confirmed(self):
    #     self.sent_for_approval()
    #     self.state = 'confirm'

    # def sent_for_approval(self):
    #     mail_template = self.env.ref('iatf.mom_document_approval_mail_template')
    #     model_id = self.env['ir.model'].sudo().search([('model', '=', self._name)])
    #     mail_template.write({'model_id' : model_id.id})
    #
    #     web_base_url = self.env['ir.config_parameter'].sudo().get_param('web.base.url')
    #     action_id = self.env.ref('iatf.action_quary_list_view', raise_if_not_found=False)
    #     link = """{}/web#id={}&view_type=form&model=self._name&action={}""".format(web_base_url,self.id,action_id.id)
    #     self.link = link
    #
    #     user_ids = self.env['hr.employee'].sudo().search([('department_id', 'in', self.approve_department_ids.ids)]).mapped('user_id')
    #     all_admin = ''
    #     for user in user_ids:
    #         if all_admin:
    #             all_admin += ',' + str(user.login)
    #         else:
    #             all_admin = str(user.login)
    #     mail_template.write({
    #             'email_to': all_admin,
    #     })
    #     mail_template.send_mail(self.id, force_send=True)

    # def _check_department(self):
    #     for rec in self:
    #         if self.env.user.employee_id.department_id.id in rec.approve_department_ids.ids and self.env.user.employee_id.department_id.department_type == 'hr':
    #             if rec.state != 'draft':
    #                 rec.hr_check_department = True
    #             else:
    #                 rec.hr_check_department = False
    #         else:
    #             rec.hr_check_department = False
    #         if self.env.user.employee_id.department_id.id in rec.approve_department_ids.ids and self.env.user.employee_id.department_id.department_type == 'design':
    #             if rec.state != 'draft':
    #                 rec.des_check_department = True
    #             else:
    #                 rec.des_check_department = False
    #         else:
    #             rec.des_check_department = False
    #         if self.env.user.employee_id.department_id.id in rec.approve_department_ids.ids and self.env.user.employee_id.department_id.department_type == 'eng':
    #             if rec.state != 'draft':
    #                 rec.eng_check_department = True
    #             else:
    #                 rec.eng_check_department = False
    #         else:
    #             rec.eng_check_department = False
    #         if self.env.user.employee_id.department_id.id in rec.approve_department_ids.ids and self.env.user.employee_id.department_id.department_type == 'manu':
    #             if rec.state != 'draft':
    #                 rec.manu_check_department = True
    #             else:
    #                 rec.manu_check_department = False
    #         else:
    #             rec.manu_check_department = False
    #         if self.env.user.employee_id.department_id.id in rec.approve_department_ids.ids and self.env.user.employee_id.department_id.department_type == 'sales':
    #             if rec.state != 'draft':
    #                 rec.sale_check_department = True
    #             else:
    #                 rec.sale_check_department = False
    #         else:
    #             rec.sale_check_department = False
    #         if self.env.user.employee_id.department_id.id in rec.approve_department_ids.ids and self.env.user.employee_id.department_id.department_type == 'qc':
    #             if rec.state != 'draft':
    #                 rec.qc_check_department = True
    #             else:
    #                 rec.qc_check_department = False
    #         else:
    #             rec.qc_check_department = False
    #         if self.env.user.employee_id.department_id.id in rec.approve_department_ids.ids and self.env.user.employee_id.department_id.department_type == 'maintenance':
    #             if rec.state != 'draft':
    #                 rec.mainte_check_department = True
    #             else:
    #                 rec.mainte_check_department = False
    #         else:
    #             rec.mainte_check_department = False
    #         if self.env.user.employee_id.department_id.id in rec.approve_department_ids.ids and self.env.user.employee_id.department_id.department_type == 'marketing':
    #             if rec.state != 'draft':
    #                 rec.marketing_check_department = True
    #             else:
    #                 rec.marketing_check_department = False
    #         else:
    #             rec.marketing_check_department = False
    #         if self.env.user.employee_id.department_id.id in rec.approve_department_ids.ids and self.env.user.employee_id.department_id.department_type == 'pm':
    #             if rec.state != 'draft':
    #                 rec.pm_check_department = True
    #             else:
    #                 rec.pm_check_department = False
    #         else:
    #             rec.pm_check_department = False
    #         if self.env.user.employee_id.department_id.id in rec.approve_department_ids.ids and self.env.user.employee_id.department_id.department_type == 'management':
    #             if rec.state != 'draft':
    #                 rec.managment_check_department = True
    #             else:
    #                 rec.managment_check_department = False
    #         else:
    #             rec.managment_check_department = False
    #
    #
    # hr_approval_date = fields.Datetime('Hr Approval Date')
    # design_approval_date = fields.Datetime('Design Approval Date')
    # eng_approval_date = fields.Datetime('Eng. Approval Date')
    # manu_approval_date = fields.Datetime('Production Approval Date')
    # sale_approval_date = fields.Datetime('Sales Approval Date')
    # qc_approval_date = fields.Datetime('QC Approval Date')
    # main_approval_date = fields.Datetime('Maintenance Approval Date')
    # mark_approval_date = fields.Datetime('Marketing Approval Date')
    # pm_approval_date = fields.Datetime('Program Management Approval Date')
    # maneg_approval_date = fields.Datetime('Top Management Approval Date')
    #
    # def hr_approval(self):
    #     self.hr_approval_date = datetime.now()
    #     self.is_hr_approved = True
    #     self.approve_by_department_ids = [(4, self.env.user.employee_id.department_id.id)]
    #     self.approvaed_manager_ids = [(4, self.env.user.employee_id.department_id.manager_id.id)]
    # def design_approval(self):
    #     self.design_approval_date = datetime.now()
    #     self.is_design_approved = True
    #     self.approve_by_department_ids = [(4, self.env.user.employee_id.department_id.id)]
    #     self.approvaed_manager_ids = [(4, self.env.user.employee_id.department_id.manager_id.id)]
    # def eng_approval(self):
    #     self.eng_approval_date = datetime.now()
    #     self.is_eng_approved = True
    #     self.approve_by_department_ids = [(4, self.env.user.employee_id.department_id.id)]
    #     self.approvaed_manager_ids = [(4, self.env.user.employee_id.department_id.manager_id.id)]
    # def manu_approval(self):
    #     self.manu_approval_date = datetime.now()
    #     self.is_manu_approved = True
    #     self.approve_by_department_ids = [(4, self.env.user.employee_id.department_id.id)]
    #     self.approvaed_manager_ids = [(4, self.env.user.employee_id.department_id.manager_id.id)]
    # def sales_approval(self):
    #     self.sale_approval_date = datetime.now()
    #     self.is_sales_approved = True
    #     self.approve_by_department_ids = [(4, self.env.user.employee_id.department_id.id)]
    #     self.approvaed_manager_ids = [(4, self.env.user.employee_id.department_id.manager_id.id)]
    # def qc_approval(self):
    #     self.qc_approval_date = datetime.now()
    #     self.is_qc_approved = True
    #     self.approve_by_department_ids = [(4, self.env.user.employee_id.department_id.id)]
    #     self.approvaed_manager_ids = [(4, self.env.user.employee_id.department_id.manager_id.id)]
    # def mainte_approval(self):
    #     self.main_approval_date = datetime.now()
    #     self.is_maintenance_approved = True
    #     self.approve_by_department_ids = [(4, self.env.user.employee_id.department_id.id)]
    #     self.approvaed_manager_ids = [(4, self.env.user.employee_id.department_id.manager_id.id)]
    # def marketing_approval(self):
    #     self.mark_approval_date = datetime.now()
    #     self.is_marketing_approved = True
    #     self.approve_by_department_ids = [(4, self.env.user.employee_id.department_id.id)]
    #     self.approvaed_manager_ids = [(4, self.env.user.employee_id.department_id.manager_id.id)]
    # def pm_approval(self):
    #     self.pm_approval_date = datetime.now()
    #     self.is_pm_approved = True
    #     self.approve_by_department_ids = [(4, self.env.user.employee_id.department_id.id)]
    #     self.approvaed_manager_ids = [(4, self.env.user.employee_id.department_id.manager_id.id)]
    # def managment_approval(self):
    #     self.maneg_approval_date = datetime.now()
    #     self.is_management_approved = True
    #     self.approve_by_department_ids = [(4, self.env.user.employee_id.department_id.id)]
    #     self.approvaed_manager_ids = [(4, self.env.user.employee_id.department_id.manager_id.id)]
    #
    #
    # def button_send_approval(self):
    #     for rec in self:
    #         if rec.hr:
    #             user_email = rec.hr.login
    #             mail_temp = self.env.ref('iatf.sending_mail_template_risk_assessment')
    #             if mail_temp:
    #                 format_name = rec.format_id.name
    #                 cust_name = rec.cust_id.name
    #                 subject = f" Request for approving Risk Assessment {format_name} for {cust_name}"
    #                 if user_email:
    #                     body = """
    #                     <div>
    #                         <p>Dear Recipient  """ + str(user_email) + """,
    #                             <br/><br/>
    #                             Please, kindly Request To accept and approved Sending Approval for the Risk Analysis Sheet.
    #                         <br></br>
    #                         Thank you.
    #                     <br/>
    #                     <br/>
    #                     <div>"""
    #                     mail_temp.send_mail(self.id, email_values={
    #                         'email_from': self.env.user.login,
    #                         'email_to': user_email,
    #                         'subject': subject,
    #                         'body_html': body,
    #                     }, force_send=True)
    #
    #         rec.write({'state': 'hr_approve'})
    #         rec.part_development_id.quart_id = rec.id
    #
    # def button_hr_approval(self):
    #     for rec in self:
    #         if rec.hr:
    #             user_email = rec.hr.login
    #             mail_temp = self.env.ref('iatf.sending_mail_template_risk_assessment')
    #             if mail_temp:
    #                 format_name = rec.format_id.name
    #                 cust_name = rec.cust_id.name
    #                 subject = f" Request for approving Risk Assessment {format_name} for {cust_name}"
    #                 if user_email:
    #                     body = """
    #                     <div>
    #                         <p>Dear Recipient  """ + str(user_email) + """,
    #                             <br/><br/>
    #                             Please, kindly Request To accept and approved Sending Approval for the Risk Analysis Sheet.
    #                         <br></br>
    #                         Thank you.
    #                     <br/>
    #                     <br/>
    #                     <div>"""
    #                     mail_temp.send_mail(self.id, email_values={
    #                         'email_from': self.env.user.login,
    #                         'email_to': user_email,
    #                         'subject': subject,
    #                         'body_html': body,
    #                     }, force_send=True)
    #         rec.write({'state': 'design'})
    #
    # def button_design(self):
    #     for rec in self:
    #         if rec.design_eng:
    #             user_email = rec.design_eng.login
    #             mail_temp = self.env.ref('iatf.sending_mail_template_risk_assessment')
    #             if mail_temp:
    #                 format_name = rec.format_id.name
    #                 cust_name = rec.cust_id.name
    #                 subject = f" Request for approving Risk Assessment {format_name} for {cust_name}"
    #                 if user_email:
    #                     body = """
    #                     <div>
    #                         <p>Dear Recipient  """ + str(user_email) + """,
    #                             <br/><br/>
    #                             Please, kindly Request To accept and approved Sending Approval for the Risk Analysis Sheet.
    #                         <br></br>
    #                         Thank you.
    #                     <br/>
    #                     <br/>
    #                     <div>"""
    #                     mail_temp.send_mail(self.id, email_values={
    #                         'email_from': self.env.user.login,
    #                         'email_to': user_email,
    #                         'subject': subject,
    #                         'body_html': body,
    #                     }, force_send=True)
    #         rec.write({'state': 'engineering'})
    #
    # def button_engineering(self):
    #     for rec in self:
    #         if rec.manf_eng:
    #             user_email = rec.manf_eng.login
    #             mail_temp = self.env.ref('iatf.sending_mail_template_risk_assessment')
    #             if mail_temp:
    #                 format_name = rec.format_id.name
    #                 cust_name = rec.cust_id.name
    #                 subject = f" Request for approving Risk Assessment {format_name} for {cust_name}"
    #                 if user_email:
    #                     body = """
    #                     <div>
    #                         <p>Dear Recipient  """ + str(user_email) + """,
    #                             <br/><br/>
    #                             Please, kindly Request To accept and approved Sending Approval for the Risk Analysis Sheet.
    #                         <br></br>
    #                         Thank you.
    #                     <br/>
    #                     <br/>
    #                     <div>"""
    #                     mail_temp.send_mail(self.id, email_values={
    #                         'email_from': self.env.user.login,
    #                         'email_to': user_email,
    #                         'subject': subject,
    #                         'body_html': body,
    #                     }, force_send=True)
    #         rec.write({'state': 'manufacturing'})
    #
    # def button_manufacturing(self):
    #     for rec in self:
    #         if rec.production:
    #             user_email = rec.production.login
    #             mail_temp = self.env.ref('iatf.sending_mail_template_risk_assessment')
    #             if mail_temp:
    #                 format_name = rec.format_id.name
    #                 cust_name = rec.cust_id.name
    #                 subject = f" Request for approving Risk Assessment {format_name} for {cust_name}"
    #                 if user_email:
    #                     body = """
    #                     <div>
    #                         <p>Dear Recipient  """ + str(user_email) + """,
    #                             <br/><br/>
    #                             Please, kindly Request To accept and approved Sending Approval for the Risk Analysis Sheet.
    #                         <br></br>
    #                         Thank you.
    #                     <br/>
    #                     <br/>
    #                     <div>"""
    #                     mail_temp.send_mail(self.id, email_values={
    #                         'email_from': self.env.user.login,
    #                         'email_to': user_email,
    #                         'subject': subject,
    #                         'body_html': body,
    #                     }, force_send=True)
    #         rec.write({'state': 'quality'})
    #
    # def button_quality_test_done(self):
    #     for rec in self:
    #         if rec.quality:
    #             user_email = rec.quality.login
    #             mail_temp = self.env.ref('iatf.sending_mail_template_risk_assessment')
    #             if mail_temp:
    #                 format_name = rec.format_id.name
    #                 cust_name = rec.cust_id.name
    #                 subject = f" Request for approving Risk Assessment {format_name} for {cust_name}"
    #                 if user_email:
    #                     body = """
    #                     <div>
    #                         <p>Dear Recipient  """ + str(user_email) + """,
    #                             <br/><br/>
    #                             Please, kindly Request To accept and approved Sending Approval for the Risk Analysis Sheet.
    #                         <br></br>
    #                         Thank you.
    #                     <br/>
    #                     <br/>
    #                     <div>"""
    #                     mail_temp.send_mail(self.id, email_values={
    #                         'email_from': self.env.user.login,
    #                         'email_to': user_email,
    #                         'subject': subject,
    #                         'body_html': body,
    #                     }, force_send=True)
    #         rec.write({'state': 'top'})
    #
    # def button_top_managment_final_approved(self):
    #     for rec in self:
    #         if rec.top_management_id:
    #             user_email = rec.top_management_id.login
    #             mail_temp = self.env.ref('iatf.sending_mail_template_risk_assessment')
    #             if mail_temp:
    #                 format_name = rec.format_id.name
    #                 cust_name = rec.cust_id.name
    #                 subject = f" Request for approving Risk Assessment {format_name} for {cust_name}"
    #                 if user_email:
    #                     body = """
    #                     <div>
    #                         <p>Dear Recipient  """ + str(user_email) + """,
    #                             <br/><br/>
    #                             Please, kindly Request To accept and approved Sending Approval for the Risk Analysis Sheet.
    #                         <br></br>
    #                         Thank you.
    #                     <br/>
    #                     <br/>
    #                     <div>"""
    #                     mail_temp.send_mail(self.id, email_values={
    #                         'email_from': self.env.user.login,
    #                         'email_to': user_email,
    #                         'subject': subject,
    #                         'body_html': body,
    #                     }, force_send=True)
    #         rec.write({'state': 'final_approved'})

    def action_query_list(self):
        output = BytesIO()
        wb = Workbook()
        ws = wb.active

        if self.env.user.company_id.logo:
            max_width = 500  # Set your desired maximum width
            max_height = 100  # Set your desired maximum height
            image_data = base64.b64decode(self.env.user.company_id.logo)

            # Open the image using PIL
            image = PILImage.open(io.BytesIO(image_data))
            width, height = image.size
            aspect_ratio = width / height

            if width > max_width:
                width = max_width
                height = int(width / aspect_ratio)

            if height > max_height:
                height = max_height
                width = int(height * aspect_ratio)

            # Resize the image using PIL
            # Add space on the top and left side of the image
            padding_top = 10  # Adjust as needed
            padding_left = 10  # Adjust as needed

            resized_image = image.resize((width, height), PILImage.LANCZOS)
            ImageOps.expand(resized_image, border=(padding_left, padding_top, 0, 0), fill='rgba(0,0,0,0)')
            img_bytes = io.BytesIO()
            resized_image.save(img_bytes, format='PNG')
            img_bytes.seek(0)
            logo_image = Image(img_bytes)
            # logo_image = Image(self.env.user.company_id.logo)
            ws.add_image(logo_image, 'A1')

        border = Border(top=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'),
                        bottom=Side(style='thin'))
        font_header = Font(name='Arial', size=11, bold=True)
        align_left = Alignment(vertical='center', horizontal='left', wrapText=True)
        align_center = Alignment(vertical='center', horizontal='center', wrapText=True)
        font_all = Font(name='Times New Roman', size=11, bold=False)

        data = {
            'A1': 'Technical Queries & Deviation Request',
            'A2': "Project  Name",
            'A3': "Customer Details",
            'F2': "First disposition date : ",
            'G2': "Second disposition date : ",
            'H2': "Third disposition date : ",
            'I2': "Fourth disposition date : ",
            'E3': "First Compilation date : ",
            'F3': "Revision : R1",
            'G3': "Revision : R2",
            'H3': "Revision : R3",
            'I3': "Revision : R4",
            'A4': "Part no.",
            'B4': "Part Name",
            'C4': "Drawing revision number",
            'D4': "Drawing revision date",
            'E4': "Query / Deviation requirement",
            'F4': "Customer disposition",
            'G4': "Internal disposition ",
            'H4': "Customer disposition",
            'I4': "Internal disposition",
        }
        for cell, value in data.items():
            ws[cell] = value

        # region merging and formatting cells
        max_col = 9
        max_row = 4
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.border = border
                cell.alignment = align_left
                cell.font = font_header
        for cell in ws[4]:
            cell.alignment = align_center
            cell.font = font_header

        # Specific Dimensions
        ws['A1'].alignment = align_center
        ws['A1'].font = Font(name='Arial', size=26, bold=True)

        # Row height Dimensions
        ws.row_dimensions[1].height = 75
        ws.row_dimensions[2].height = 23
        ws.row_dimensions[3].height = 23
        ws.row_dimensions[4].height = 45
        for row in range(5, 20 + 1):
            ws.row_dimensions[row].height = 40

        # Column Width Dimensions
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 40
        ws.column_dimensions['F'].width = 40
        ws.column_dimensions['G'].width = 40
        ws.column_dimensions['H'].width = 40
        ws.column_dimensions['I'].width = 40

        ws.merge_cells('A1:I1')
        ws.merge_cells('A2:B2')
        ws.merge_cells('C2:D2')
        ws.merge_cells('A3:B3')
        ws.merge_cells('C3:D3')

        ws.freeze_panes = 'A5'

        # Data Filling
        cur_row = 5
        for rec in self:
            ws['C2'] = rec.project_id.name if rec.project_id else ''
            ws['C3'] = rec.partner_id.name if rec.partner_id.name else ''
            ws['E2'] = f"Doc Type: {rec.doc_type}" if rec.doc_type else 'Doc Type: '
            ws['F3'] = f"Revision: {rec.revision}" if rec.revision else 'Revision: '
            ws['G3'] = f"Revision: {rec.revision}" if rec.revision else 'Revision: '
            ws['H3'] = f"Revision: {rec.revision}" if rec.revision else 'Revision: '
            ws['I3'] = f"Revision: {rec.revision}" if rec.revision else 'Revision: '
            ws[
                'E3'] = f"First Compilation date: {rec.first_compilation_date.strftime('%d-%m-%Y')}" if rec.first_compilation_date else 'First Compilation date: '
            ws[
                'F2'] = f"First disposition date : {rec.first_disposition_date.strftime('%d-%m-%Y')}" if rec.first_disposition_date else 'First disposition date : '
            ws[
                'G2'] = f"Second disposition date: {rec.second_disposition_date.strftime('%d-%m-%Y')}" if rec.second_disposition_date else 'Second disposition date: '
            ws[
                'H2'] = f"Third disposition date: {rec.third_disposition_date.strftime('%d-%m-%Y')}" if rec.third_disposition_date else 'Third disposition date: '
            ws[
                'I2'] = f"Fourth disposition date: {rec.fourth_disposition_date.strftime('%d-%m-%Y')}" if rec.fourth_disposition_date else 'Fourth disposition date: '
            for i in rec.quary_list_line_ids:
                ws[f'A{cur_row}'] = i.part_no if i.part_no else ''
                ws[f'B{cur_row}'] = i.part_name if i.part_name else ''
                ws[f'C{cur_row}'] = i.drawing_revision_number if i.drawing_revision_number else ''
                ws[f'D{cur_row}'] = i.drawing_revision_date.strftime('%d-%m-%Y') if i.drawing_revision_date else ''
                ws[f'E{cur_row}'] = i.query_deviation_requirement if i.query_deviation_requirement else ''
                ws[f'F{cur_row}'] = i.customer_disposition if i.customer_disposition else ''
                ws[f'G{cur_row}'] = i.internal_disposition if i.internal_disposition else ''
                ws.row_dimensions[cur_row].height = 40

                cur_row += 1

            for row_no in ws.iter_rows(min_row=5, max_row=cur_row, min_col=1, max_col=9):
                for cell in row_no:
                    cell.border = border
                    cell.alignment = align_center

        if cur_row > 20:
            cur_row = cur_row
        else:
            cur_row = 21
        ws.merge_cells(f'A{cur_row}:I{cur_row}')
        cur_row += 1

        if cur_row < max_row:
            cur_row = max_row + 1

        # region SignOff Members Footer
        sign_row = cur_row
        ws[f'A{cur_row}'] = 'Prepared By'
        ws[f'A{cur_row}'].font = font_header
        ws.merge_cells(f'A{cur_row}:B{cur_row}')
        ws.merge_cells(f'C{cur_row}:E{cur_row}')

        ws[f'F{cur_row}'] = 'Prepared Date'
        ws[f'F{cur_row}'].font = font_header

        for rec in self:
            ws[f'C{cur_row}'] = rec.create_uid.name if rec.create_uid else ''
            ws[f'G{cur_row}'] = rec.create_date.strftime('%d-%m-%Y') if rec.create_date else ''

        ws.row_dimensions[cur_row].height = 18
        cur_row += 1
        ws.merge_cells(f'A{cur_row}:G{cur_row}')
        cur_row += 1
        ws.merge_cells(f'A{cur_row}:G{cur_row}')
        ws[f'A{cur_row}'] = 'Sign OFF'
        ws[f'A{cur_row}'].font = Font(size=18, bold=True)
        ws.row_dimensions[cur_row].height = 25
        cur_row += 1
        ws.merge_cells(f'A{cur_row}:G{cur_row}')
        cur_row += 1
        for cell in ws[cur_row]:
            ws[f'A{cur_row}'] = 'Member'
            ws[f'C{cur_row}'] = 'Department'
            ws[f'E{cur_row}'] = 'Status'
            ws[f'F{cur_row}'] = 'Date'
            ws[f'G{cur_row}'] = 'Comments'
            ws.row_dimensions[cur_row].height = 25
            ws.merge_cells(f'A{cur_row}:B{cur_row}')
            ws.merge_cells(f'C{cur_row}:D{cur_row}')
            # ws.merge_cells(f'G{cur_row}:H{cur_row}')
            cell.font = Font(size=12, bold=True, color='ffffff')
            cell.fill = PatternFill(start_color='0070C0', end_color='0070C0', fill_type="solid")

        for row_no in ws.iter_rows(min_row=1, max_row=cur_row + 1, min_col=1, max_col=9):
            for cell in row_no:
                cell.border = border
                cell.alignment = align_center

        # Listing Managers Id
        for rec in self:
            for record in rec.iatf_members_ids:
                name_rec = record.approver_id.name
                dept_rec = record.department_id.name if record.department_id.name else ''
                status_rec = record.approval_status.capitalize()
                date_rec = record.date_approved_rejected.strftime('%d-%m-%Y') if record.date_approved_rejected else ''
                comment_rec = record.comment if record.comment else ''

                ws[f'A{cur_row + 1}'] = name_rec
                ws[f'C{cur_row + 1}'] = dept_rec
                ws[f'E{cur_row + 1}'] = status_rec
                ws[f'F{cur_row + 1}'] = date_rec
                ws[f'G{cur_row + 1}'] = comment_rec
                ws.merge_cells(f'A{cur_row + 1}:B{cur_row + 1}')
                ws.merge_cells(f'C{cur_row + 1}:D{cur_row + 1}')
                # ws.merge_cells(f'G{cur_row}:H{cur_row}')

                status_dict = {'Approved': ['CFFFC3', '000000'], 'Rejected': ['FFCDCD', '000000'],
                               'Revision': ['AFEFFF', '000000'], 'Pending': ['FDFFCD', '000000']}

                for state, color in status_dict.items():
                    if state == status_rec:
                        ws[f'E{cur_row + 1}'].fill = PatternFill(start_color=color[0], end_color=color[0],
                                                                 fill_type="solid")
                        ws[f'E{cur_row + 1}'].font = Font(size=12, bold=False, color=color[1])

                cur_row += 1

            ws.merge_cells(f'A{cur_row + 1}:I{cur_row + 1}')
            ws.merge_cells(f'H{sign_row}:I{cur_row}')

            for row_no in ws.iter_rows(min_row=sign_row + 1, max_row=cur_row + 1, min_col=1, max_col=9):
                for cell in row_no:
                    cell.border = border
                    cell.alignment = align_center
            # endregion

        wb.save(output)
        output.seek(0)
        self.generate_xls_file = base64.b64encode(output.getvalue()).decode('utf-8')

        return {
            "type": "ir.actions.act_url",
            "target": "self",
            "url": "/web/content?model=quary.list&download=true&field=generate_xls_file&filename={filename}.xlsx&id={pid}".format(
                filename="Query List", pid=self[0].id),
        }


class QuaryListLine(models.Model):
    _name = 'quary.list.line'
    _description = 'Query List Line'
    _inherit = "translation.mixin"

    quart_list_id = fields.Many2one(
        comodel_name='quary.list',
        string='Query List',
        required=True, ondelete='cascade', index=True, copy=False
    )

    part_no = fields.Integer("Part No")
    part_name = fields.Char("Part Name",translate=True)
    drawing_revision_number = fields.Integer("Drawing Revision Number")
    drawing_revision_date = fields.Date("Drawing Revision Date")
    query_deviation_requirement = fields.Char("Query / Deviation Requirement",translate=True)
    customer_disposition = fields.Char("Customer Disposition",translate=True)
    internal_disposition = fields.Char("Internal Disposition",translate=True)
