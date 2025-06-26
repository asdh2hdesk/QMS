from datetime import date, datetime
from odoo import fields, models, api
import openpyxl
from io import BytesIO
import io
from openpyxl import Workbook
import openpyxl
import base64
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.drawing.image import Image
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
import copy
from PIL import ImageOps
from PIL import Image as PILImage
from bs4 import BeautifulSoup


class PackagingStandard(models.Model):
    _name = 'iatf.packaging.standard'

    part_id = fields.Many2one('product.template', 'Part No.')
    part_description = fields.Char('Product Description')
    supplier_id = fields.Many2one('res.company', 'Supplier Name')
    photo_final_package_with_label = fields.Binary('Photo Final Package with Label')

    primary_polythene_bag = fields.Boolean('Polythene Bag')
    primary_butter_paper = fields.Boolean('Butter Paper')
    primary_bubble_sheet = fields.Boolean('Bubble sheet')
    primary_any_other = fields.Boolean('Any Other (Specify)')
    primary_na = fields.Boolean('Not Applicable')
    primary_other_material = fields.Char('Specify Used Material')
    primary_antirust_oil_used_type = fields.Char('Antirust oil used/type')
    primary_photo_with_primary_packing = fields.Binary('Photo with Primary Packing')

    secondary_corrugated_box = fields.Boolean('Corrugated Box')
    secondary_plastic_box = fields.Boolean('Plastic Box')
    secondary_wooden_box = fields.Boolean('Wooden Box')
    secondary_metallic_box = fields.Boolean('Metallic Box')
    secondary_gunny_bag = fields.Boolean('Gunny Bag')
    secondary_drum = fields.Boolean('Drum')
    secondary_any_other = fields.Boolean('Any Other (Specify)')
    secondary_na = fields.Boolean('Not Applicable')
    secondary_other_material = fields.Char('Specify Used Material')
    secondary_separator_details = fields.Char('Separator Details')
    secondary_photo_with_secondary_packing = fields.Binary('Photo with Secondary Packing')

    final_corrugated_box = fields.Boolean('Corrugated Box')
    final_plastic_box = fields.Boolean('Plastic Box')
    final_wooden_box = fields.Boolean('Wooden Box')
    final_metallic_box = fields.Boolean('Metallic Box')
    final_gunny_bag = fields.Boolean('Gunny Bag')
    final_drum = fields.Boolean('Drum')
    final_dunnages = fields.Boolean('Dunnages')
    final_any_other = fields.Boolean('Any Other (Specify)')
    final_other_material = fields.Char('Specify Used Material')
    final_na = fields.Boolean('Not Applicable')
    final_detail_steel_nylon_other_clamping = fields.Char('Details of Steel/Nylon strip or other clamping')
    final_photo_with_final_packing = fields.Binary('Photo with Final Packing')

    weight_per_packing = fields.Float("Weight per packing")
    box_length = fields.Float("Box length")
    box_width = fields.Float("Box width")
    box_height = fields.Float("Box height")
    num_of_pieces_per_box = fields.Float("Number of pieces per Box")

    prepared_by = fields.Many2one('res.users', 'Prepared By')
    approved_by = fields.Many2one('res.users', 'Approved By')
    approve_date = fields.Date('Approval Date')
    buyer = fields.Many2one('res.partner', 'Buyer')
    customer_approval_date = fields.Date('Customer Approval Date')
    generate_xls_file = fields.Binary('Generate XLSX File')

    def action_generate_report(self):
        output = BytesIO()
        wb = openpyxl.Workbook()
        # Get the active worksheet
        ws = wb.active  # Using 'ws' for abbreviation

        # region formatting
        max_width = 500  # Set your desired maximum width
        max_height = 100  # Set your desired maximum height
        if self.env.user.company_id.logo:
            image_data = base64.b64decode(self.env.user.company_id.logo)

            # Open the image using PIL
            image = PILImage.open(io.BytesIO(image_data))
            width, height = image.size
            aspect_ratio = width / height

            if width > max_width:
                width = max_width
                height = int(width / aspect_ratio)

            if height > max_height:
                height = max_height
                width = int(height * aspect_ratio)

            # Resize the image using PIL
            # Add space on the top and left side of the image
            padding_top = 10  # Adjust as needed
            padding_left = 10  # Adjust as needed

            resized_image = image.resize((width, height), PILImage.ANTIALIAS)
            ImageOps.expand(resized_image, border=(padding_left, padding_top, 0, 0), fill='rgba(0,0,0,0)')
            img_bytes = io.BytesIO()
            resized_image.save(img_bytes, format='PNG')
            img_bytes.seek(0)
            logo_image = Image(img_bytes)
            # logo_image = Image(self.env.user.company_id.logo)
            ws.add_image(logo_image, 'A1')

        min_row = 1
        min_col = 1
        max_row = 25
        max_col = 9

        thin = Side(border_style='thin', color='000000')
        font = Font(name='Arial', size=10, bold=True)
        font_header = Font(name='Arial', size=14, bold=True)
        my_green = openpyxl.styles.colors.Color(rgb='00FF00')
        my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_green)

        alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
        center_align = Alignment(horizontal='center', vertical='center', wrapText=True)

        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                cell.alignment = alignment
                cell.font = font

        ws['A1'].font = font_header

        # Set column widths
        ws.column_dimensions["A"].width = 10.2
        ws.column_dimensions["B"].width = 8.4
        ws.column_dimensions["C"].width = 18.4
        ws.column_dimensions["D"].width = 8.2
        ws.column_dimensions["E"].width = 14.10
        ws.column_dimensions["F"].width = 21.60
        ws.column_dimensions["G"].width = 9
        ws.column_dimensions["H"].width = 11.25
        ws.column_dimensions["I"].width = 18.3

        # Set row heights
        ws.row_dimensions[1].height = 60
        ws.row_dimensions[5].height = 250
        ws.row_dimensions[16].height = 39.5
        ws.row_dimensions[17].height = 102
        ws.row_dimensions[18].height = 36.5
        ws.row_dimensions[19].height = 130
        ws.row_dimensions[20].height = 20
        ws.row_dimensions[22].height = 21

        # Write headers with placeholders for data
        ws["A1"] = "PACKAGING STANDARD"
        ws["A2"] = "Supplier Name"
        ws["A3"] = "Part No."
        ws["D3"] = "Part Description"
        ws["A4"] = "Photograph of final packing with label(including part No./Qty/Date of packing.)"
        ws["A6"] = "Primary packing details"
        ws["D6"] = "Secondary packing details"
        ws["G6"] = "Final packing details"
        ws["B7"] = "Polythene Bag"
        ws["B8"] = "Butter Paper"
        ws["B9"] = "Bubble Sheet"
        ws["B10"] = "Any Other (Specify)"
        ws["B15"] = "Not Applicable"

        ws["E7"] = "Corrugated Box"
        ws["E8"] = "Plastic Box"
        ws["E9"] = "Wooden Box"
        ws["E10"] = "Metallic Box"
        ws["E11"] = "Gunny Bag"
        ws["E12"] = "Drum"
        ws["E13"] = "Any Other (Specify)"
        ws["E15"] = "Not Applicable"

        ws["H7"] = "Corrugated Box"
        ws["H8"] = "Plastic Box"
        ws["H9"] = "Wooden Box"
        ws["H10"] = "Metallic Box"
        ws["H11"] = "Gunny Bag"
        ws["H12"] = "Drum"
        ws["H13"] = "Dunnages"
        ws["H14"] = "Any Other (Specify)"

        ws["A16"] = "Antirust oil used/type"
        ws["D16"] = "Separator details"
        ws["G16"] = "Details of Steel/Nylon strip or other clamping"
        ws["A18"] = "Photograph with Primary Packing"
        ws["D18"] = "Photograph with Secondary Packing"
        ws["G18"] = "Photograph with Final Packing"

        ws["A20"] = "Final Packaging Data"
        ws["A21"] = "Weight Per Packing"  # Replace with actual weight
        ws["D21"] = f"Box size(LxWxH) cm³"  # Replace with actual box dimensions
        ws["G21"] = "No. of PIECES PER BOX"  # Replace with number of pieces
        ws["A23"] = "Prepared by:"
        ws["D23"] = "Approved by:"
        ws["A24"] = "Customer SQA Approval"
        ws["D24"] = "Buyer"
        ws["G23"] = "Date:"
        ws["G24"] = "Date:"
        ws["A25"] = "FM-20018/00 REV NO 00 DATE 12.04.2021"  # Footer text

        # Merge cells with correct notation
        ws.merge_cells("A1:I1")
        ws.merge_cells("A2:B2")
        ws.merge_cells("C2:I2")
        ws.merge_cells("B3:C3")
        ws.merge_cells("D3:E3")
        ws.merge_cells("F3:I3")
        ws.merge_cells("A4:I4")
        ws.merge_cells("A5:I5")
        ws.merge_cells("A6:C6")
        ws.merge_cells("D6:F6")
        ws.merge_cells("G6:I6")
        ws.merge_cells("B7:C7")
        ws.merge_cells("B8:C8")
        ws.merge_cells("B9:C9")
        ws.merge_cells("B10:C10")
        ws.merge_cells("B11:C11")
        ws.merge_cells("B12:C12")
        ws.merge_cells("B13:C13")
        ws.merge_cells("B14:C14")
        ws.merge_cells("B15:C15")
        ws.merge_cells("B11:C14")

        ws.merge_cells("E7:F7")
        ws.merge_cells("E8:F8")
        ws.merge_cells("E9:F9")
        ws.merge_cells("E10:F10")
        ws.merge_cells("E11:F11")
        ws.merge_cells("E12:F12")
        ws.merge_cells("E13:F13")
        ws.merge_cells("E14:F14")
        ws.merge_cells("E15:F15")

        ws.merge_cells("H7:I7")
        ws.merge_cells("H8:I8")
        ws.merge_cells("H9:I9")
        ws.merge_cells("H10:I10")
        ws.merge_cells("H11:I11")
        ws.merge_cells("H12:I12")
        ws.merge_cells("H13:I13")
        ws.merge_cells("H14:I14")
        ws.merge_cells("H15:I15")

        ws.merge_cells("A16:C16")
        ws.merge_cells("D16:F16")
        ws.merge_cells("G16:I16")
        
        ws.merge_cells("A17:C17")
        ws.merge_cells("D17:F17")
        ws.merge_cells("G17:I17")

        ws.merge_cells("A18:C18")
        ws.merge_cells("D18:F18")
        ws.merge_cells("G18:I18")

        ws.merge_cells("A19:C19")
        ws.merge_cells("D19:F19")
        ws.merge_cells("G19:I19")
        ws.merge_cells("A20:I20")

        ws.merge_cells("A21:C21")
        ws.merge_cells("D21:F21")
        ws.merge_cells("G21:I21")
        ws.merge_cells("A22:C22")
        ws.merge_cells("D22:F22")
        ws.merge_cells("G22:I22")

        ws.merge_cells("A23:B23")
        ws.merge_cells("D23:E23")
        ws.merge_cells("H23:I23")

        ws.merge_cells("A24:C24")
        ws.merge_cells("E24:F24")
        ws.merge_cells("H24:I24")

        ws.merge_cells("A25:I25")

        # endregion

        # region Adding Data
        for rec in self:
            ws['C2'] = rec.supplier_id.name if rec.supplier_id.name else ''
            ws['B3'] = rec.part_id.name if rec.part_id.name else ''
            ws['F3'] = rec.part_description if rec.part_description else ''

            if rec.photo_final_package_with_label:
                image_data = base64.b64decode(rec.photo_final_package_with_label)
                max_width = 500
                max_height = 100
                # Open the image using PIL
                image = PILImage.open(io.BytesIO(image_data))
                width, height = image.size
                aspect_ratio = width / height

                if width > max_width:
                    width = max_width
                    height = int(width / aspect_ratio)

                if height > max_height:
                    height = max_height
                    width = int(height * aspect_ratio)

                # Resize the image using PIL
                # Add space on the top and left side of the image
                padding_top = 10  # Adjust as needed
                padding_left = 10  # Adjust as needed

                resized_image = image.resize((width, height), PILImage.ANTIALIAS)
                ImageOps.expand(resized_image, border=(padding_left, padding_top, 0, 0), fill='rgba(0,0,0,0)')
                img_bytes = io.BytesIO()
                resized_image.save(img_bytes, format='PNG')
                img_bytes.seek(0)
                image_to_add = Image(img_bytes)
                # logo_image = Image(self.env.user.company_id.logo)
                ws.add_image(image_to_add, 'A5')
                ws['A5'].alignment = center_align

            ws['A7'] = '☑' if rec.primary_polythene_bag else ''
            ws['A8'] = '☑' if rec.primary_butter_paper else ''
            ws['A9'] = '☑' if rec.primary_bubble_sheet else ''
            ws['A10'] = '☑' if rec.primary_any_other else ''
            ws['B11'] = rec.primary_other_material if rec.primary_other_material else ''
            ws['A15'] = '☑' if rec.primary_na else ''

            ws['D7'] = '☑' if rec.secondary_corrugated_box else ''
            ws['D8'] = '☑' if rec.secondary_plastic_box else ''
            ws['D9'] = '☑' if rec.secondary_wooden_box else ''
            ws['D10'] = '☑' if rec.secondary_metallic_box else ''
            ws['D11'] = '☑' if rec.secondary_gunny_bag else ''
            ws['D12'] = '☑' if rec.secondary_drum else ''
            ws['D13'] = '☑' if rec.secondary_any_other else ''
            ws['E14'] = rec.secondary_other_material if rec.secondary_other_material else ''
            ws['D15'] = '☑' if rec.secondary_na else ''

            ws['G7'] = '☑' if rec.final_corrugated_box else ''
            ws['G8'] = '☑' if rec.final_plastic_box else ''
            ws['G9'] = '☑' if rec.final_wooden_box else ''
            ws['G10'] = '☑' if rec.final_metallic_box else ''
            ws['G11'] = '☑' if rec.final_gunny_bag else ''
            ws['G12'] = '☑' if rec.final_drum else ''
            ws['G13'] = '☑' if rec.final_dunnages else ''
            ws['G14'] = '☑' if rec.final_any_other else ''
            ws['H15'] = rec.final_other_material if rec.final_other_material else ''
            ws['G16'] = '☑' if rec.final_na else ''

            ws['A17'] = rec.primary_antirust_oil_used_type if rec.primary_antirust_oil_used_type else ''
            ws['D17'] = rec.secondary_separator_details if rec.secondary_separator_details else ''
            ws['G17'] = rec.final_detail_steel_nylon_other_clamping if rec.final_detail_steel_nylon_other_clamping else ''

            if rec.primary_photo_with_primary_packing:
                image_data = base64.b64decode(rec.primary_photo_with_primary_packing)
                max_width = 500
                max_height = 100
                # Open the image using PIL
                image = PILImage.open(io.BytesIO(image_data))
                width, height = image.size
                aspect_ratio = width / height

                if width > max_width:
                    width = max_width
                    height = int(width / aspect_ratio)

                if height > max_height:
                    height = max_height
                    width = int(height * aspect_ratio)

                # Resize the image using PIL
                # Add space on the top and left side of the image
                padding_top = 10  # Adjust as needed
                padding_left = 10  # Adjust as needed

                resized_image = image.resize((width, height), PILImage.ANTIALIAS)
                ImageOps.expand(resized_image, border=(padding_left, padding_top, 0, 0), fill='rgba(0,0,0,0)')
                img_bytes = io.BytesIO()
                resized_image.save(img_bytes, format='PNG')
                img_bytes.seek(0)
                image_to_add = Image(img_bytes)
                # logo_image = Image(self.env.user.company_id.logo)
                ws.add_image(image_to_add, 'A19')

            if rec.secondary_photo_with_secondary_packing:
                image_data = base64.b64decode(rec.secondary_photo_with_secondary_packing)
                max_width = 500
                max_height = 100
                # Open the image using PIL
                image = PILImage.open(io.BytesIO(image_data))
                width, height = image.size
                aspect_ratio = width / height

                if width > max_width:
                    width = max_width
                    height = int(width / aspect_ratio)

                if height > max_height:
                    height = max_height
                    width = int(height * aspect_ratio)

                # Resize the image using PIL
                # Add space on the top and left side of the image
                padding_top = 10  # Adjust as needed
                padding_left = 10  # Adjust as needed

                resized_image = image.resize((width, height), PILImage.ANTIALIAS)
                ImageOps.expand(resized_image, border=(padding_left, padding_top, 0, 0), fill='rgba(0,0,0,0)')
                img_bytes = io.BytesIO()
                resized_image.save(img_bytes, format='PNG')
                img_bytes.seek(0)
                image_to_add = Image(img_bytes)
                # logo_image = Image(self.env.user.company_id.logo)
                ws.add_image(image_to_add, 'D19')

            if rec.final_photo_with_final_packing:
                image_data = base64.b64decode(rec.final_photo_with_final_packing)
                max_width = 500
                max_height = 100
                # Open the image using PIL
                image = PILImage.open(io.BytesIO(image_data))
                width, height = image.size
                aspect_ratio = width / height

                if width > max_width:
                    width = max_width
                    height = int(width / aspect_ratio)

                if height > max_height:
                    height = max_height
                    width = int(height * aspect_ratio)

                # Resize the image using PIL
                # Add space on the top and left side of the image
                padding_top = 10  # Adjust as needed
                padding_left = 10  # Adjust as needed

                resized_image = image.resize((width, height), PILImage.ANTIALIAS)
                ImageOps.expand(resized_image, border=(padding_left, padding_top, 0, 0), fill='rgba(0,0,0,0)')
                img_bytes = io.BytesIO()
                resized_image.save(img_bytes, format='PNG')
                img_bytes.seek(0)
                image_to_add = Image(img_bytes)
                # logo_image = Image(self.env.user.company_id.logo)
                ws.add_image(image_to_add, 'G19')

            ws['A22'] = rec.weight_per_packing if rec.weight_per_packing else ''

            a = '='
            if rec.box_length:
                a += f"{rec.box_length}"
            else:
                a += '0'

            if rec.box_width:
                a += f"*{rec.box_width}"
            else:
                a += '*0'

            if rec.box_height:
                a += f"*{rec.box_height}"
            else:
                a += '*0'

            ws['D22'] = a
            ws['G22'] = rec.num_of_pieces_per_box if rec.num_of_pieces_per_box else ''
            ws['C23'] = rec.prepared_by.name if rec.prepared_by else ''
            ws['F23'] = rec.approved_by.name if rec.approved_by else ''
            ws['H23'] = rec.approve_date if rec.approve_date else ''
            ws['E24'] = rec.buyer.name if rec.buyer else ''
            ws['H24'] = rec.customer_approval_date if rec.customer_approval_date else ''

        # endregion
        wb.save(output)
        output.seek(0)
        self.generate_xls_file = base64.b64encode(output.getvalue()).decode('utf-8')
        return {
            "type": "ir.actions.act_url",
            "target": "self",
            "url": "/web/content?model=iatf.packaging.standard&download=true&field=generate_xls_file&filename={filename}.xlsx&id={pid}".format(
                filename="Packaging Report IATF", pid=self.id),
        }