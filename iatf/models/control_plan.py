from odoo import fields, models, api, _
from odoo.exceptions import ValidationError
from datetime import date
import openpyxl
from io import BytesIO
import io
from openpyxl import Workbook
import openpyxl
import base64
from bs4 import BeautifulSoup
from openpyxl.styles import Alignment, Font, Border, Side, DEFAULT_FONT, PatternFill
from openpyxl.drawing.image import Image
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
import copy
import datetime
from PIL import ImageOps
from PIL import Image as PILImage
from bs4 import BeautifulSoup


class ControlPlan(models.Model):
    _name = "control.plan"
    _description = "Control Plan and Inspection-I"
    _inherit = ['iatf.sign.off.members', 'revision.history.mixin']

    doc_type = fields.Selection([('safelaunch', 'Safelaunch'), ('prototype', 'Prototype'), ('prelaunch', 'PreLaunch'),
                             ('production', 'Production')])
    # customer_ids = fields.Many2many(
    #     'res.partner', 'control_plan_customer_rel', 'plan_id', 'customer_id',
    #     string='Customers'
    # )
    supplier_ids = fields.Many2many(
        'res.partner', 'control_plan_supplier_rel', 'plan_id', 'supplier_id',
        string='Suppliers'
    )
    supplier_code = fields.Char("Supplier Code")
    vehicle_model = fields.Char("Vehicle / Model")
    doc_no = fields.Char("Doc. No.")
    # part_name = fields.Char("Part Name")
    # part_no = fields.Char("Part No.")
    # part_rev_no = fields.Char("Part Rev. No.")
    assy_name = fields.Many2one("product.template","Assy. Name")
    assy_no = fields.Char("Assy. No.")
    assy_rev_no = fields.Char("Assy.  Rev. No.")
    date_origin = fields.Date("Date (Orig.)")
    rev_no = fields.Char("Rev. No")
    rev_date=fields.Date("Rev. Date")
    key_contact = fields.Many2many('hr.employee', string="Key Contact")
    telephone = fields.Char(string='Telephones', compute='_compute_telephones')

    process_line_ids = fields.One2many(
        comodel_name='control.plan.process',
        inverse_name='process_id',
        string='Process'
    )
    #
    # def action_duplicate_process_lines(self):
    #     for plan in self:
    #         new_process_lines = []
    #
    #         for process in plan.process_line_ids:
    #             # Duplicate characteristic lines
    #             new_char_lines = []
    #             for char in process.process_char_ids:
    #                 new_char_lines.append((0, 0, {
    #                     'sequence_handle': char.sequence_handle,
    #                     'mc_jig_tool': [(6, 0, char.mc_jig_tool.ids)],
    #                     'char_no': char.char_no,
    #                     'char_product': char.char_product.id,
    #                     'char_process': char.char_process.id,
    #                     'method_description': char.method_description.id,
    #                     'lower_limit': char.lower_limit,
    #                     'upper_limit': char.upper_limit,
    #                     'uom_id': char.uom_id.id,
    #                     'method_evaluation': char.method_evaluation.id,
    #                     'method_sample_size': char.method_sample_size,
    #                     'method_sample_freq': char.method_sample_freq,
    #                     'method_rec_yn': char.method_rec_yn,
    #                     'method_rec_size': char.method_rec_size,
    #                     'method_inspected_by': char.method_inspected_by.id,
    #                     'method_error_proofing_name': [(6, 0, char.method_error_proofing_name.ids)],
    #                     'method_control': char.method_control.id,
    #                     'reaction_plan_action': char.reaction_plan_action,
    #                     'reaction_plan_res': char.reaction_plan_res.id,
    #                 }))
    #
    #             # Add process line with the duplicated characteristic lines
    #             new_process_lines.append((0, 0, {
    #                 'sequence_handle': process.sequence_handle,
    #                 'process_step': process.process_step,
    #                 'process_name': process.process_name,
    #                 'char_class': process.char_class.id,
    #                 'process_char_ids': new_char_lines,
    #             }))
    #
    #         # Write duplicated lines into the current plan
    #         plan.write({'process_line_ids': new_process_lines})
    #
    # def action_duplicate_lines(self):
    #     self.action_duplicate_process_lines()

    def action_open_copy_wizard(self):
        return {
            'type': 'ir.actions.act_window',
            'name': 'Copy Control Plan Data',
            'res_model': 'control.plan.copy.wizard',
            'view_mode': 'form',
            'target': 'new',
            'context': {'default_current_record_id': self.id},
            'views': [[self.env['ir.ui.view'].search([
                ('model', '=', 'control.plan.copy.wizard'),
                ('type', '=', 'form'),
                ('name', '=', 'control.plan.copy.wizard.form')
            ], limit=1).id, 'form']],
        }
    def name_get(self):
        result = []
        for record in self:
            name = record.part_number or 'No Part Name'
            result.append((record.id, name))
        return result

    def _compute_telephones(self):
        for record in self:
            # Extract and join phone numbers from all related employees
            telephones = record.key_contact.mapped('work_phone')
            record.telephone = ', '.join(telephones)

    @api.depends('key_contact')
    def _compute_telephones(self):
        for record in self:
            # Extract and join phone numbers from all related employees
            telephones = record.key_contact.mapped('work_phone')
            record.telephone = ', '.join(telephones)

    generate_xls_file = fields.Binary(string="Generated file")  # do not comment this

    @api.model_create_multi
    def create(self, vals_list):
        records = super(ControlPlan, self).create(vals_list)

        for record in records:
            if not record.project_id:
                continue

            # Fetch approved process reports related to the project
            approved_reports = self.env["process.report"].search(
                [('project_id', '=', record.project_id.id),
                 ('final_status', '=', 'approved')]
            )

            # # Check if there are approved reports
            # if not approved_reports:
            #     raise ValidationError(
            #         _("No approved process reports found for this project.")
            #     )

            # Dictionary to track existing control plan operations by station_no
            existing_operations = {}
            # Dictionary to track item counters for each issue
            issue_item_counters = {}

            for op in approved_reports.report_ids.sorted(key=lambda r: r.station_no):
                # Check if an operation already exists for this station_no
                if op.station_no in existing_operations:
                    cp_operation = existing_operations[op.station_no]
                else:
                    # Create new control plan operation
                    cp_operation = self.env['control.plan.process'].create({
                        'process_id': record.id,
                        'process_step': op.station_no,
                        'process_name': op.process_step_name,
                        'char_class': op.special_product_characteristics.id if op.special_product_characteristics else False,
                    })
                    existing_operations[op.station_no] = cp_operation

                # Initialize item counter for this issue if not already present
                if op.issue not in issue_item_counters:
                    issue_item_counters[op.issue] = 0

                # All 5 line types grouped by work type
                line_types = {
                    'man': op.man_line_ids,
                    'machine': op.machine_line_ids,
                    'material': op.material_line_ids,
                    'environment': op.environment_line_ids,
                    'method': op.method_line_ids,
                }

                for work_type, lines in line_types.items():
                    if not lines:
                        continue

                    # Create character lines for each item in the work type
                    for line in lines:
                        # Increment item counter for this issue
                        issue_item_counters[op.issue] += 1

                        # Generate char_no as issue.item_counter
                        char_no = f"{op.issue}.{issue_item_counters[op.issue]}"

                        # Create method control
                        prevention = line.current_prevention_control or ''
                        detection = line.current_detection_control or ''
                        combined_name = f"{prevention} / {detection}".strip(' /')
                        combined_name = f"{prevention} / {detection}".strip(' /')

                        # Search for existing method control or create a new one
                        method_record = self.env['control.method'].search([('name', '=', combined_name)], limit=1)
                        if not method_record:
                            method_record = self.env['control.method'].create({'name': combined_name})

                        # Create a character line directly linked to the operation
                        self.env['control.chara.line'].create({
                            'chara_id': cp_operation.id,
                            'char_no': char_no,  # Assign computed unique char_no
                            'char_product': op.func_of_product_step.id if op.func_of_product_step else False,
                            'char_process': op.func_of_process_step.id if op.func_of_process_step else False,
                            'method_control': method_record.id,
                        })

        return records

    def action_update_process_flow(self):
        """ Update process flow with modified control plan data """
        for record in self:
            # Find related process flow records based on the project
            process_flow_main = self.env["process.flow"].search([
                ('project_id', '=', record.project_id.id),
                ('final_status', '=', 'approved')  # Changed '!=' to '=' to match the error message
            ])

            if not process_flow_main:
                raise ValidationError(_("No approved process flow found for this project."))

            # Iterate through the control plan process lines
            for cp_operation in record.process_line_ids:
                # Find the corresponding process flow line based on the step
                process_flow_line = process_flow_main.process_flow_line_ids.filtered(
                    lambda x: x.step == cp_operation.process_step
                )

                if process_flow_line:
                    # Update the process flow line with the control plan process data
                    process_flow_line.write({
                        'desc_of_operation': cp_operation.process_name,
                        'class_f': cp_operation.char_class.id if cp_operation.char_class else False,
                    })
                else:
                    # Specify the correct model name when creating a new process flow line
                    process_flow_line = self.env['process.flow.line'].create({
                        'process_flow_id': process_flow_main.id,
                        'step': cp_operation.process_step,  # Assuming you meant 'process_step' here
                        'desc_of_operation': cp_operation.process_name,
                    })

                # Update or create process operation lines
                for cp_char_line in cp_operation.process_char_ids:
                    # Find the corresponding process operation line
                    char_line = process_flow_line.process_op_lines_ids.filtered(
                        lambda x: x.element_no == cp_char_line.char_no
                    )

                    if char_line:
                        # Update the existing process operation line
                        char_line.write({
                            'product_characteristics': cp_char_line.char_product,
                            'process_characteristics': cp_char_line.char_process,
                            'control_method': cp_char_line.method_control,
                            'equip_fixture_ids': [(6, 0, cp_char_line.mc_jig_tool.ids)],
                            'pokayoke_ids': [(6, 0, cp_char_line.method_error_proofing_name.ids)],
                        })
                    else:
                        # Create a new process operation line if it doesn't exist
                        char_line.create({
                            'operation_element_id': process_flow_line.id,
                            'element_no': cp_char_line.char_no,
                            'product_characteristics': cp_char_line.char_product.id,
                            'process_characteristics': cp_char_line.char_process.id,
                            'control_method': cp_char_line.method_control,
                            'equip_fixture_ids': [(6, 0, cp_char_line.mc_jig_tool.ids)],
                            'pokayoke_ids': [(6, 0, cp_char_line.method_error_proofing_name.ids)],
                        })

            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _("Success"),
                    'message': _("Process Flow Diagram updated successfully!"),
                    'sticky': False,  # Notification disappears after a few seconds
                    'type': 'success',  # Green success notification
                }
            }

    def action_update_equipment_fixture_tool(self):
        for record in self:
            # Find approved equipment fixture tracking sheet
            equipment_fixtures = self.env["equipment.trackingsheet"].search([
                ('project_id', '=', record.project_id.id),
                ('final_status', '=', 'approved')
            ])

            # Find approved tool tracking sheet
            tool_tracking = self.env["tools.trackingsheet"].search([
                ('project_id', '=', record.project_id.id),
                ('final_status', '=', 'approved')
            ])

            if not equipment_fixtures and not tool_tracking:
                raise ValidationError(_("No approved Equipment Fixture or Tool Tracking Sheet found for this project."))

            # Update equipment tracking sheets
            for process_line in record.process_line_ids:
                for char_line in process_line.process_char_ids:
                    for equipment in char_line.mc_jig_tool:
                        category_name = equipment.category_id.name if equipment.category_id else False

                        # Handle equipment and fixtures
                        if category_name in ['Equipment', 'Fixture'] and equipment_fixtures:
                            for tracking_sheet in equipment_fixtures:
                                # Search for existing tracking line with same equipment
                                tracking_line = self.env['equipment.trackingsheet.line'].search([
                                    ('tracking_id', '=', tracking_sheet.id),
                                    ('operation_no', '=', process_line.process_step),
                                    ('element_no', '=', char_line.char_no),
                                    ('equipment_id', '=', equipment.id),
                                ], limit=1)

                                if not tracking_line:
                                    # Create new tracking line for this equipment
                                    self.env['equipment.trackingsheet.line'].create({
                                        'tracking_id': tracking_sheet.id,
                                        'element_no': char_line.char_no,
                                        'operation_no': process_line.process_step,
                                        'equipment_id': equipment.id,
                                        'category_id': equipment.category_id.id,
                                    })

                        # Handle tools
                        elif category_name == 'Tool' and tool_tracking:
                            for tracking_sheet in tool_tracking:
                                # Search for existing tracking line with same tool
                                tracking_line = self.env['tools.trackingsheet.line'].search([
                                    ('tracking_id', '=', tracking_sheet.id),
                                    ('operation_no', '=', process_line.process_step),
                                    ('element_no', '=', char_line.char_no),
                                    ('tool_name', '=', equipment.id),  # Tool is stored in equipment variable
                                ], limit=1)

                                if not tracking_line:
                                    # Create new tracking line for this tool
                                    self.env['tools.trackingsheet.line'].create({
                                        'tracking_id': tracking_sheet.id,
                                        'element_no': char_line.char_no,
                                        'operation_no': process_line.process_step,
                                        'tool_name': equipment.id,  # Tool is stored in equipment variable
                                    })

            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _("Success"),
                    'message': _("Equipment, Fixture and Tool Tracking Sheets updated successfully!"),
                    'sticky': False,
                    'type': 'success',
                }
            }

    def action_update_process_matrix(self):
        """ Update process matrix with modified process flow data """
        for record in self:
            # Find related process flow records based on the project
            process_matrix_main = self.env["process.group"].search([
                ('project_id', '=', record.project_id.id),
                ('final_status', '!=', 'approved')
            ])

            if not process_matrix_main:
                raise ValidationError(_("No Inprogress Process Matrix found for this project."))

            # Iterate through the Process Flow Diagram lines
            for cp_operation in record.process_line_ids:
                # Find the corresponding process flow line based on the step
                process_matrix_op = process_matrix_main.process_presentation_ids.filtered(
                    lambda x: x.operation == cp_operation.process_step)
                if process_matrix_op:
                    process_matrix_op.write({
                        'operation_description': cp_operation.process_name,
                    })
                else:
                    process_matrix_op.create({
                        'process_id': process_matrix_main.id,
                        'operation': cp_operation.process_step,
                        'operation_description': cp_operation.process_name,
                    })
                for cp_char_line in cp_operation.process_char_ids:
                    char_line = process_matrix_op.operation_lines_ids.filtered(
                        lambda x: x.element_no == cp_char_line.char_no)
                    if char_line:
                        char_line.write({
                            'product_kpc': cp_char_line.char_product.id,
                            'process_kcc': cp_char_line.char_process.id,
                            'equipment_fixture_ids': [
                                (6, 0, [equipment.id for equipment in cp_char_line.mc_jig_tool])],
                            # [(6, 0, [id1, id2, id3])]
                        })
                    else:
                        char_line.create({
                            'operation_element_id': process_matrix_op.id,
                            'element_no': cp_char_line.char_no,
                            'product_kpc': cp_char_line.char_product.id,
                            'process_kcc': cp_char_line.char_process.id,
                            'equipment_fixture_ids': [(6, 0, [equipment.id for equipment in cp_char_line.mc_jig_tool])],
                        })

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _("Success"),
                'message': _("Process Matrix updated successfully!"),
                'sticky': False,  # False means the notification disappears after a few seconds
                'type': 'success',  # Green success notification
            }
        }

    def action_generate_excel_report(self):
        output = BytesIO()
        wb = Workbook()
        ws = wb.active

        if self.env.user.company_id.logo:
            max_width = 300  # Set your desired maximum width
            max_height = 80  # Set your desired maximum height
            image_data = base64.b64decode(self.env.user.company_id.logo)

            # Open the image using PIL
            image = PILImage.open(io.BytesIO(image_data))
            width, height = image.size
            aspect_ratio = width / height

            if width > max_width:
                width = max_width
                height = int(width / aspect_ratio)

            if height > max_height:
                height = max_height
                width = int(height * aspect_ratio)

            # Resize the image using PIL
            # Add space on the top and left side of the image
            padding_top = 10  # Adjust as needed
            padding_left = 10  # Adjust as needed

            resized_image = image.resize((width, height), PILImage.LANCZOS)
            ImageOps.expand(resized_image, border=(padding_left, padding_top, 0, 0), fill='rgba(0,0,0,0)')
            img_bytes = io.BytesIO()
            resized_image.save(img_bytes, format='PNG')
            img_bytes.seek(0)
            logo_image = Image(img_bytes)
            # logo_image = Image(self.env.user.company_id.logo)
            ws.add_image(logo_image, 'A1')

        border = Border(top=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'),
                        bottom=Side(style='thin'))
        align_center = Alignment(vertical='center', horizontal='center', wrapText=True)
        align_left = Alignment(vertical='center', horizontal='left')
        font_header = Font(name='Arial', size=12, bold=True)
        font_all = Font(name='Times New Roman', size=11, bold=False)
        grey_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        blue_fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')

        data = {
            'A1': 'CONTROL PLAN & INSPECTION-I',
            'A2': 'Safe_Launch/Prototype/Prelaunch/Production',
            'A3': 'Customer',
            'A4': 'Supplier',
            'A5': 'Supplier Code',
            'A6': 'Vehicle/Model',
            'F3': 'Part No. / Rev. No.',
            'F4': 'Part Name',
            'F5': 'Assy. No. / Rev. No.',
            'F6': 'Assy Name',
            'L3': 'Doc. No.',
            'L4': 'Date (ORIG.)',
            'L5': 'Rev. No / Date',
            'L6': 'Telephone No.:',
            # '': 'Key Contact',
            'A8': 'Part/Process Step No.',
            'B8': 'Process Name / Operation Description',
            'C8': 'M/C, JIG, Tools, For MFG. Name',
            'D8': 'M/C, JIG, Tools, For MFG. Number',
            'E8': 'Characteristics (Special & Others)',
            'E9': 'No.',
            'F9': 'Product',
            'G9': 'Process',
            'H9': 'Class',
            'I8': 'Method',
            'I9': 'Product Specs. / Tolerance & Parameters',
            'J9': 'Evaluation Measurement Techniques',
            'K9': 'Sample',
            'K10': 'Size',
            'L10': 'Freq.',
            'M9': 'Recording',
            'M10': 'Y/N',
            'N10': 'Size',
            'O9': 'Inspected By',
            'P9': 'Error Proofing',
            'P10': 'Name',
            'Q10': 'Number',
            'R9': 'Control Method',
            'S8': 'Reaction Plan & Corrective Actions',
            'S10': 'Corrective Actions',
            'T10': 'Owner/Responsibility',
        }

        for cell, value in data.items():
            ws[cell] = value
            ws[cell].font = font_header

        max_row = 30

        cell_ranges_to_merge = ['A1:T1', 'A2:B2', 'C2:T2', 'A7:T7', 'A8:A10', 'B8:B10', 'C8:C10', 'E8:H8',
                                'D8:D10', 'E9:E10', 'F9:F10', 'G9:G10', 'H9:H10', 'I8:R8', 'I9:I10',
                                'J9:J10', 'K9:L9', 'M9:N9', 'O9:O10', 'P9:Q9', 'R9:R10', 'S8:T9', ]
        for cell_mer in cell_ranges_to_merge:
            ws.merge_cells(cell_mer)

        for i in range(3, 7):
            ws.merge_cells(f'B{i}:E{i}')
            ws.merge_cells(f'F{i}:G{i}')
            ws.merge_cells(f'H{i}:K{i}')
            ws.merge_cells(f'L{i}:M{i}')
            ws.merge_cells(f'N{i}:Q{i}')
            for cell in ws[i]:
                cell.alignment = align_left

        for row in ws.iter_rows(min_row=1, max_row=10, min_col=1, max_col=20):
            for cell in row:
                cell.border = border
                cell.alignment = align_center

            for row_no in ws.iter_rows(min_row=11, max_row=29, min_col=1, max_col=20):
                for cell in row_no:
                    cell.border = border
                    cell.alignment = align_center

            ws['A1'].font = Font(name='Arial', size=20, bold=True)
            ws['A1'].fill = PatternFill(start_color='87CEFA', end_color='87CEFA', fill_type='solid')

        ws.row_dimensions[1].height = 60
        ws.row_dimensions[2].height = 30
        ws.row_dimensions[3].height = 25
        ws.row_dimensions[4].height = 25
        ws.row_dimensions[5].height = 25
        ws.row_dimensions[6].height = 25
        ws.row_dimensions[8].height = 30
        ws.row_dimensions[9].height = 30
        ws.row_dimensions[10].height = 30

        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18

        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15

        ws.column_dimensions['I'].width = 18
        ws.column_dimensions['J'].width = 18
        ws.column_dimensions['K'].width = 12
        ws.column_dimensions['L'].width = 12
        ws.column_dimensions['M'].width = 12
        ws.column_dimensions['N'].width = 12
        ws.column_dimensions['O'].width = 15
        ws.column_dimensions['P'].width = 15
        ws.column_dimensions['Q'].width = 15
        ws.column_dimensions['R'].width = 18
        ws.column_dimensions['S'].width = 20
        ws.column_dimensions['T'].width = 20
        grey_cells = ['A2', 'A3', 'F3', 'L3', 'A4', 'F4', 'L4', 'A5', 'F5', 'L5', 'A6', 'F6', 'L6']

        for cell in grey_cells:
            ws[cell].fill = grey_fill

        blue_cells = ['A8', 'B8', 'C8', 'D8','E8', 'I8', 'S8', 'E9', 'F9', 'G9', 'H9','I9', 'J9', 'K9', 'M9', 'O9', 'P9', 'R9', 'K10',
                      'L10', 'M10', 'N10', 'P10', 'Q10', 'S10', 'T10']
        for cell in blue_cells:
            ws[cell].fill = blue_fill

        # Data Filling
        cur_row = 11
        for rec in self:
            ws['C2'] = rec.doc_type if rec.doc_type else ''
            customer_names = ', '.join(rec.partner_id.mapped('name'))
            ws['B3'] = customer_names
            supplier_names = ', '.join(rec.supplier_ids.mapped('name'))
            ws['B4'] = supplier_names
            ws['B5'] = rec.supplier_code if rec.supplier_code else ''
            ws['B6'] = rec.vehicle_model if rec.vehicle_model else ''
            ws['H3'] = rec.part_number if rec.part_number else ''
            ws['H4'] = rec.part_name if rec.part_name else ''
            ws['H6'] = rec.assy_name.name if rec.assy_name else ''
            ws['H5'] = rec.assy_no if rec.assy_no else ''
            ws['N3'] = rec.doc_no if rec.doc_no else ''
            ws['N4'] = rec.date_origin if rec.date_origin else ''
            ws['N5'] = rec.rev_no if rec.rev_no else ''
            # ws['N6'] = rec.telephone if rec.telephone else ''
            # ws['N6'] = rec.key_contact if rec.key_contact else ''
            for line in rec.process_line_ids:
                ws[f'A{cur_row}'] = line.process_step if line.process_step else ''
                ws[f'B{cur_row}'] = line.process_name if line.process_name else ''
                ws[f'H{cur_row}'] = line.char_class.symbol if line.char_class.symbol else ''
                mer_row = cur_row
                for i in line.process_char_ids:
                    ws[f'C{cur_row}'] = ','.join(mc.name if mc.name else '' for mc in i.mc_jig_tool)
                    ws[f'D{cur_row}'] = i.mc_jig_tool_num if i.mc_jig_tool_num else ''

                    ws[f'E{cur_row}'] = i.char_no if i.char_no else ''
                    ws[f'F{cur_row}'] = i.char_product.name if i.char_product else ''
                    ws[f'G{cur_row}'] = i.char_process.name if i.char_process else ''
                    ws[f'I{cur_row}'] = i.method_product_display  if i.method_product_display  else ''
                    ws[f'J{cur_row}'] = (
                        f"{i.method_evaluation.gauge.name or ''} - {i.method_evaluation.gauge_no or ''}"
                        if i.method_evaluation else ''
                    )

                    ws[f'K{cur_row}'] = i.method_sample_size if i.method_sample_size else ''
                    ws[f'L{cur_row}'] = i.method_sample_freq if i.method_sample_freq else ''
                    ws[f'M{cur_row}'] = i.method_rec_yn if i.method_rec_yn else ''
                    ws[f'N{cur_row}'] = i.method_rec_size if i.method_rec_size else ''
                    ws[f'O{cur_row}'] = i.method_inspected_by.name if i.method_inspected_by else ''
                    ws[f'P{cur_row}'] = ','.join(error.poka_yoke_description if error.poka_yoke_description else '' for error in i.method_error_proofing_name)
                    ws[f'Q{cur_row}'] = i.method_error_proofing_num if i.method_error_proofing_num else ''
                    ws[f'R{cur_row}'] = i.method_control.name if i.method_control else ''
                    ws[f'S{cur_row}'] = i.reaction_plan_action if i.reaction_plan_action else ''
                    ws[f'T{cur_row}'] = i.reaction_plan_res.name if i.reaction_plan_res else ''

                    cur_row += 1

        if cur_row < max_row:
            cur_row = max_row
        # region SignOff Members Footer
        sign_row = cur_row
        ws.merge_cells(f'A{cur_row}:T{cur_row}')

        cur_row += 1
        ws[f'A{cur_row}'] = 'Prepared By'
        ws[f'A{cur_row}'].font = font_header
        ws.merge_cells(f'A{cur_row}:B{cur_row}')
        ws.merge_cells(f'C{cur_row}:E{cur_row}')

        ws[f'F{cur_row}'] = 'Prepared Date'
        ws[f'F{cur_row}'].font = font_header
        ws.merge_cells(f'F{cur_row}:G{cur_row}')
        ws.merge_cells(f'H{cur_row}:I{cur_row}')

        for rec in self:
            ws[f'C{cur_row}'] = rec.create_uid.name if rec.create_uid else ''
            ws[f'H{cur_row}'] = rec.create_date if rec.create_date else ''

        ws.row_dimensions[cur_row].height = 18
        cur_row += 1
        ws.merge_cells(f'A{cur_row}:I{cur_row}')
        cur_row += 1
        ws.merge_cells(f'A{cur_row}:I{cur_row}')
        ws[f'A{cur_row}'] = 'Sign OFF'
        ws[f'A{cur_row}'].font = Font(size=18, bold=True)
        ws.row_dimensions[cur_row].height = 25
        cur_row += 1
        ws.merge_cells(f'A{cur_row}:I{cur_row}')
        cur_row += 1
        for cell in ws[cur_row]:
            ws[f'A{cur_row}'] = 'Member'
            ws[f'C{cur_row}'] = 'Department'
            ws[f'D{cur_row}'] = 'Status'
            ws[f'E{cur_row}'] = 'Date'
            ws[f'F{cur_row}'] = 'Comments'
            ws.row_dimensions[cur_row].height = 25
            ws.merge_cells(f'F{cur_row}:I{cur_row}')
            cell.font = Font(size=12, bold=True, color='ffffff')
            cell.fill = PatternFill(start_color='0070C0', end_color='0070C0', fill_type="solid")

        # Listing Managers Id
        for rec in self:
            for record in rec.iatf_members_ids:
                name_rec = record.approver_id.name
                dept_rec = record.department_id.name if record.department_id.name else ''
                status_rec = record.approval_status.capitalize()
                date_rec = record.date_approved_rejected if record.date_approved_rejected else ''
                comment_rec = record.comment if record.comment else ''

                ws[f'A{cur_row + 1}'] = name_rec
                ws[f'C{cur_row + 1}'] = dept_rec
                ws[f'D{cur_row + 1}'] = status_rec
                ws[f'E{cur_row + 1}'] = date_rec
                ws[f'F{cur_row + 1}'] = comment_rec
                ws.merge_cells(f'F{cur_row + 1}:I{cur_row + 1}')

                status_dict = {'Approved': ['CFFFC3', '000000'], 'Rejected': ['FFCDCD', '000000'],
                               'Revision': ['AFEFFF', '000000'], 'Pending': ['FDFFCD', '000000']}

                for state, color in status_dict.items():
                    if state == status_rec:
                        ws[f'D{cur_row + 1}'].fill = PatternFill(start_color=color[0], end_color=color[0],
                                                                 fill_type="solid")
                        ws[f'D{cur_row + 1}'].font = Font(size=12, bold=False, color=color[1])

                cur_row += 1

            ws.merge_cells(f'A{cur_row + 1}:T{cur_row + 1}')
            ws.merge_cells(f'J{sign_row + 1}:T{cur_row}')

            for row_no in ws.iter_rows(min_row=sign_row, max_row=cur_row + 1, min_col=1, max_col=20):
                for cell in row_no:
                    cell.border = border
                    cell.alignment = align_center
            # endregion

            ws['C2'].alignment = align_left

        # Save the workbook
        wb.save(output)
        output.seek(0)
        self.generate_xls_file = base64.b64encode(output.getvalue()).decode('utf-8')
        # endregion

        return {
            "type": "ir.actions.act_url",
            "target": "self",
            "url": "/web/content?model=control.plan&download=true&field=generate_xls_file&filename={filename}.xlsx&id={pid}".format(
                filename="Control Plan", pid=self[0].id),
        }


class ControlProcess(models.Model):
    _name = "control.plan.process"
    _description = "Control Process"
    _rec_name = 'process_step'
    _inherit = "translation.mixin"

    process_id = fields.Many2one('control.plan', 'Control Plan Process')
    sequence_handle = fields.Integer(string="Sequence no", default=1)
    process_step = fields.Char("Process Step")
    process_name = fields.Char("Process Name/ Operation Description",translate=True)
    char_class = fields.Many2one('process.flow.class', "Class")
    process_char_ids = fields.One2many(
        comodel_name='control.chara.line',
        inverse_name='chara_id',
        string='Process characteristics line'
    )
    def _recalculate_sequence_numbers(self):
        """Recalculate sequence_handle and element_no for all elements in the operation."""
        for operation in self:
            # Get elements sorted by existing sequence_handle (ensures correct order)
            elements = operation.operation_lines_ids.sorted(lambda e: e.sequence_handle)

            seq = 1  # Start numbering from 1
            updates = []  # Store updates to batch apply

            for element in elements:
                operation_number = re.findall(r'\d+', operation.operation or "")
                if operation_number:
                    base_number = operation_number[0]  # Extracted integer (e.g., "10" from "OP-10")
                    new_element_no = ".".join([base_number, str(seq)])

                    # Only update if the values have changed
                    if element.sequence_handle != seq or element.element_no != new_element_no:
                        updates.append((element.id, seq, new_element_no))

                seq += 1

            # Batch update to avoid multiple writes
            if updates:
                for element_id, new_seq, new_element_no in updates:
                    self.env['operation.element.line'].browse(element_id).sudo().write({
                        'sequence_handle': new_seq,
                        'element_no': new_element_no
                    })


class ControlCharLine(models.Model):
    _name = 'control.chara.line'
    _description = 'Control Plan Characteristics Line'
    _inherit = "translation.mixin"
    sequence_handle = fields.Integer(string="Sequence no", default=1)
    chara_id = fields.Many2one('control.plan.process', 'Control Plan Process',ondelete='cascade')
    operation = fields.Char(related='chara_id.process_step', string="Operation", store=True,translate=True)
    # equip_fixture_ids = fields.Many2many('maintenance.equipment', string="Equipment & Fixture")
    mc_jig_tool = fields.Many2many('maintenance.equipment', string="MC, Jig, Tool",
                                   domain="[('category_id', 'in', ['Equipment', 'Fixture','Tool'])]")
    mc_jig_tool_num = fields.Char(string="MC, Jig, Tool Number", compute="_compute_mc_jig_tool_num", store=True,translate=True)
    char_no = fields.Char("Characteristics No.",translate=True)
    char_product = fields.Many2one('product.characteristics', 'Product Characteristics')
    char_process = fields.Many2one('process.characteristics', 'Process Characteristics')
    method_description = fields.Many2one('gdt.symbol', string="Method Description",store=True)

    lower_limit = fields.Float(string="Lower Limit")
    upper_limit = fields.Float(string="Upper Limit")
    uom_id = fields.Many2one('uom.uom', string="Unit of Measure")

    method_product_display = fields.Char(
        string="Method Product Spec./ Tolerance",
        compute="_compute_method_product_display",
        translate=True
    )

    @api.depends('lower_limit', 'upper_limit', 'uom_id')
    def _compute_method_product_display(self):
        for rec in self:
            if rec.lower_limit is not None and rec.upper_limit is not None and rec.uom_id:
                rec.method_product_display = f"{rec.lower_limit}-{rec.upper_limit} {rec.uom_id.name}"
            else:
                rec.method_product_display = ''

    method_evaluation = fields.Many2one(
        'gauge.trackingsheet.line',
        string="Evaluation Measurement Techniques",

        store=True
    )

    @api.onchange('lower_limit', 'upper_limit', 'uom_id')
    def _onchange_match_gauge(self):
        for rec in self:
            if rec.lower_limit and rec.upper_limit and rec.uom_id:
                domain = [
                    ('lower_limit', '<=', rec.lower_limit),
                    ('upper_limit', '>=', rec.upper_limit),
                    ('uom_id', '=', rec.uom_id.id),
                ]
                matched_gauge = self.env['gauge.trackingsheet.line'].search(domain, limit=1)
                rec.method_evaluation = matched_gauge or False

    method_sample_size = fields.Char("Sample Size",translate=True)
    method_sample_freq = fields.Char("Sample Freq.",translate=True)
    method_rec_yn = fields.Char("Recording Y/N",translate=True)
    method_rec_size = fields.Char("Recording Size",translate=True)
    method_inspected_by = fields.Many2one('res.users', "Inspected By")


    @api.constrains('method_inspected_by')
    def _check_method_inspected_by(self):
        missing = self.filtered(lambda rec: not rec.method_inspected_by)
        if missing:
            char_nos = ', '.join(filter(None, missing.mapped('char_no')))
            msg = _("The field 'Inspected By' must be filled.")
            if char_nos:
                msg += _(" (Char No: %s)") % char_nos
            raise ValidationError(msg)


    method_error_proofing_name = fields.Many2many('poka.yoka.line', string="Error Proofing Name",
                                                  domain="[('operation', '=', operation)]")
    method_error_proofing_num = fields.Char(
        string="Error Proofing Number",
        compute="_compute_error_proofing_num",
        store=True,
        translate=True
    )

    @api.depends('method_error_proofing_name')
    def _compute_error_proofing_num(self):
        for record in self:
            if record.method_error_proofing_name:
                record.method_error_proofing_num = ', '.join(record.method_error_proofing_name.mapped('poka_yoke_number') if record.method_error_proofing_num else '')
            else:
                record.method_error_proofing_num = ''

    method_control = fields.Many2one('control.method', 'Control Method')
    reaction_plan_action = fields.Char("Reaction Plan Actions",translate=True)
    reaction_plan_res = fields.Many2one('res.users', "Reaction Plan Owner/Resp.")

    @api.onchange('method_evaluation')
    def _onchange_method_description(self):
        for rec in self:
            if rec.method_evaluation:
                rec.method_description = rec.method_evaluation.method_description
            else:
                rec.method_description = False

    @api.depends('mc_jig_tool')
    def _compute_mc_jig_tool_num(self):
        for rec in self:
            # Ensure mc_jig_tool exists and has values
            if rec.mc_jig_tool:
                serial_num = rec.mc_jig_tool.mapped('serial_no')

                # Ensure all values are strings and remove None/False
                serial_num = [str(num) for num in serial_num if num]

                rec.mc_jig_tool_num = ', '.join(serial_num) if serial_num else ''
            else:
                rec.mc_jig_tool_num = '- '  # Assign empty string if mc_jig_tool is empty

    # @api.depends('method_error_proofing_name')
    # def _compute_error_proofing_num(self):
    #     for record in self:
    #         if record.method_error_proofing_name:
    #             poka_nums = record.method_error_proofing_name.mapped('poka_num')
    #
    #             # Ensure all values are strings and filter out empty/None values
    #             poka_nums = [str(num) for num in poka_nums if num]
    #
    #             record.method_error_proofing_num = ', '.join(poka_nums) if poka_nums else ''
    #         else:
    #             record.method_error_proofing_num = ''
class ControlPlanCopyWizard(models.TransientModel):
    _name = 'control.plan.copy.wizard'
    _description = 'Control Plan Copy Wizard'
    _inherit = "translation.mixin"

    current_record_id = fields.Many2one('control.plan', string='Current Record', required=True)
    selected_record_id = fields.Many2one('control.plan', string='Select Record to Copy From', required=True)

    def action_copy_data(self):
        self.ensure_one()
        current_record = self.current_record_id
        selected_record = self.selected_record_id

        if not (current_record and selected_record):
            raise ValidationError(_("Invalid record selection. Please try again."))

        # Update top-level fields
        current_record.write({
            # 'project_id': selected_record.project_id.id if selected_record.project_id else False,
            'doc_type': selected_record.doc_type,
            'partner_id': selected_record.partner_id.id if selected_record.partner_id else False,
            'part_id': selected_record.part_id.id,
            'part_name': selected_record.part_name,
            'part_number': selected_record.part_number,
            'doc_type': selected_record.doc_type,
            'supplier_ids': [(6, 0, selected_record.supplier_ids.ids)] if selected_record.supplier_ids else False,
            'supplier_code': selected_record.supplier_code,
            'vehicle_model': selected_record.vehicle_model,
            'doc_no': selected_record.doc_no,
            'assy_rev_no': selected_record.assy_rev_no,
            'assy_name': selected_record.assy_name.id,
            'assy_no': selected_record.assy_no,
            'assy_rev_no': selected_record.assy_rev_no,
            'date_origin': selected_record.date_origin,
            'rev_no': selected_record.rev_no,
            'rev_date': selected_record.rev_date,
            'key_contact': [(6, 0, selected_record.key_contact.ids)] if selected_record.key_contact else False,
            'process_line_ids': [(5, 0, 0)],  # Clear existing process_line_ids
        })

        # Copy process_line_ids and nested process_char_ids
        new_lines = []
        if not selected_record.process_line_ids:
            raise ValidationError(_("No process lines found in the selected record. Please select a record with process lines."))

        for line in selected_record.process_line_ids:
            # Prepare process_char_ids for the process line
            char_lines = []
            for char_line in line.process_char_ids:
                char_line_data = {
                    'sequence_handle': char_line.sequence_handle,
                    'mc_jig_tool': [(6, 0, char_line.mc_jig_tool.ids)] if char_line.mc_jig_tool else False,
                    'char_no': char_line.char_no,
                    'char_product': char_line.char_product.id if char_line.char_product else False,
                    'char_process': char_line.char_process.id if char_line.char_process else False,
                    'method_description': char_line.method_description.id if char_line.method_description else False,
                    'lower_limit': char_line.lower_limit,
                    'upper_limit': char_line.upper_limit,
                    'uom_id': char_line.uom_id.id if char_line.uom_id else False,
                    'method_evaluation': char_line.method_evaluation.id if char_line.method_evaluation else False,
                    'method_sample_size': char_line.method_sample_size,
                    'method_sample_freq': char_line.method_sample_freq,
                    'method_rec_yn': char_line.method_rec_yn,
                    'method_rec_size': char_line.method_rec_size,
                    'method_inspected_by': char_line.method_inspected_by.id if char_line.method_inspected_by else False,
                    'method_error_proofing_name': [(6, 0, char_line.method_error_proofing_name.ids)] if char_line.method_error_proofing_name else False,
                    'method_control': char_line.method_control.id if char_line.method_control else False,
                    'reaction_plan_action': char_line.reaction_plan_action,
                    'reaction_plan_res': char_line.reaction_plan_res.id if char_line.reaction_plan_res else False,
                }
                char_lines.append((0, 0, char_line_data))

            # Create process line with nested char lines
            line_data = {
                'sequence_handle': line.sequence_handle,
                'process_step': line.process_step,
                'process_name': line.process_name,
                'char_class': line.char_class.id if line.char_class else False,
                'process_char_ids': char_lines,
            }
            new_lines.append((0, 0, line_data))

        current_record.write({'process_line_ids': new_lines})

        return {'type': 'ir.actions.act_window_close'}