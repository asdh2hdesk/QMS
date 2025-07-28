from datetime import date, datetime
from odoo import api, fields, models
from odoo.exceptions import ValidationError
from datetime import date
import openpyxl
from io import BytesIO
import io
from openpyxl import Workbook
import openpyxl
import base64
from bs4 import BeautifulSoup
from openpyxl.styles import Alignment, Font, Border, Side, DEFAULT_FONT, PatternFill
from openpyxl.drawing.image import Image
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
import copy
import datetime
from PIL import ImageOps
from PIL import Image as PILImage
from bs4 import BeautifulSoup



class SpecialCharacterMatrix(models.Model):
    _name = 'special.characteristics.matrix'
    _inherit = ['iatf.sign.off.members']

    partner_id = fields.Many2one('res.partner', 'Customer Name')
    product_id = fields.Many2one('product.template', 'Product Name')
    default_code = fields.Char('Part No.', related='product_id.default_code')
    part_code = fields.Char('Fal Part Code')
    date = fields.Date('REV. Date')
    team_id = fields.Many2many('hr.employee', 'employee_category_rel_team', 'category_id', 'emp_id', string='Team')
    team_id = fields.Many2one('maintenance.team', string='Team', ondelete="restrict")
    special_line_ids = fields.One2many(
        comodel_name='special.characteristics.matrix.line',
        inverse_name='special_id',
        string="NPD Line"
    )

    generate_xls_file = fields.Binary(string="Generated file")  # do not comment this

    def action_generate_excel_report(self):
        output = BytesIO()
        wb = Workbook()
        ws = wb.active

        if self.env.user.company_id.logo:
            max_width = 300  # Set your desired maximum width
            max_height = 80  # Set your desired maximum height
            image_data = base64.b64decode(self.env.user.company_id.logo)

            # Open the image using PIL
            image = PILImage.open(io.BytesIO(image_data))
            width, height = image.size
            aspect_ratio = width / height

            if width > max_width:
                width = max_width
                height = int(width / aspect_ratio)

            if height > max_height:
                height = max_height
                width = int(height * aspect_ratio)

            # Resize the image using PIL
            # Add space on the top and left side of the image
            padding_top = 10  # Adjust as needed
            padding_left = 10  # Adjust as needed

            resized_image = image.resize((width, height), PILImage.LANCZOS)
            ImageOps.expand(resized_image, border=(padding_left, padding_top, 0, 0), fill='rgba(0,0,0,0)')
            img_bytes = io.BytesIO()
            resized_image.save(img_bytes, format='PNG')
            img_bytes.seek(0)
            logo_image = Image(img_bytes)
            # logo_image = Image(self.env.user.company_id.logo)
            ws.add_image(logo_image, 'A1')

        border = Border(top=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'),
                        bottom=Side(style='thin'))
        align_center = Alignment(vertical='center', horizontal='center', wrapText=True)
        align_left = Alignment(vertical='center', horizontal='left')
        font_header = Font(name='Arial', size=12, bold=True)
        font_all = Font(name='Times New Roman', size=11, bold=False)
        grey_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        blue_fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')

        data = {
            'C1': 'SPECIAL CHARACTERISTICS MATRIX',
            'I1': 'EQP-01',
            'H2': 'REV. LEVEL :',
            'H3': 'REV. DATE :',
            'A4': 'CUSTOMER :',
            'A5': 'CUSTOMER PART NO:',
            'A6': 'CUSTOMER PART NAME:',
            'A7': 'FAL PART CODE :',
            'A8': 'FOR CUSTOMER PRINT REVISION LEVEL AND DATE REFER EQP-02',
            'A9': 'Sl. No.',
            'B9': 'PRODUCT   CHARACTERISTICS',
            'B10': 'PARAMETER',
            'D10': 'SYMBOL DESIGNATED',
            'E10': 'SPECIFICATION',
            'F9': 'CLASSIFICATION PRODUCT/PROCESS',
            'G9': 'BASIS (FIT/FUNCTION/SAFETY/ GOVT/REGULATIONS/PERFORMANCE/CUSTOMER DESIGNATED)',
            'H9': 'RELEVANT PROCESS CHARACTERISTICS',
            'A24': 'SIGN OFF',
            'B24': 'TEAM MEMBERS',
            'B25': 'ENGINEERING',
            'B26': 'QUALITY',
            'B27': 'PRDUCTION',
            'B28': 'MATERIALS',
            'B29': 'PPC',
            'A32': 'PREPARED BY :',
            'A33': 'DATE :',
            'E32': 'APPROVED BY :',
            'E33': 'DATE :',
            'F34': 'FM-80717 REV NO.00 DATE: 19.01.2018'

        }
        for cell, value in data.items():
            ws[cell] = value

        for rows in ws.iter_rows(min_row=1, max_row=33, min_col=1, max_col=9):
            for cell in rows:
                cell.alignment = align_center
                cell.border = border
                cell.font = font_header

        ws.merge_cells('A1:B1')
        ws.merge_cells('C1:H1')
        ws.merge_cells('A2:G3')

        ws.merge_cells('A4:B4')
        ws.merge_cells('C4:I4')
        ws.merge_cells('A5:B5')
        ws.merge_cells('C5:I5')
        ws.merge_cells('A6:B6')
        ws.merge_cells('C6:I6')
        ws.merge_cells('A7:B7')
        ws.merge_cells('C7:I7')
        ws.merge_cells('A8:I8')
        ws.merge_cells('B9:E9')
        ws.merge_cells('A9:A10')
        ws.merge_cells('F9:F10')
        ws.merge_cells('G9:G10')
        ws.merge_cells('C32:D32')

        ws.merge_cells('B10:C10')

        ws.merge_cells('H9:I10')
        ws.merge_cells('B24:D24')
        ws.merge_cells('H24:I24')
        ws.merge_cells('A30:I30')
        ws.merge_cells('A31:I31')
        ws.merge_cells('A32:B32')
        ws.merge_cells('A33:B33')
        ws.merge_cells('C33:D33')
        ws.merge_cells('F32:I32')
        ws.merge_cells('F33:I33')
        ws.merge_cells('F34:I34')

        for row in range(25, 30):
            ws.merge_cells(f'C{row}:D{row}')

        for row in range(11, 24):
            ws.merge_cells(f'B{row}:C{row}')

        for row in range(11, 30):
            ws.merge_cells(f'H{row}:I{row}')

        ws['C1'].font = Font(bold=True, size=22, name='Times New Roman')

        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 21
        ws.column_dimensions['D'].width = 17
        ws.column_dimensions['E'].width = 27
        ws.column_dimensions['F'].width = 23
        ws.column_dimensions['G'].width = 30
        ws.column_dimensions['H'].width = 16
        ws.column_dimensions['I'].width = 11

        ws.row_dimensions[1].height = 33
        ws.row_dimensions[2].height = 21
        ws.row_dimensions[3].height = 21
        ws.row_dimensions[4].height = 31
        ws.row_dimensions[5].height = 27
        ws.row_dimensions[6].height = 29
        ws.row_dimensions[7].height = 27
        ws.row_dimensions[8].height = 28
        ws.row_dimensions[9].height = 42
        ws.row_dimensions[10].height = 36
        for i in range(11, 30):
            ws.row_dimensions[i].height = 18
        ws.row_dimensions[30].height = 36
        ws.row_dimensions[31].height = 18
        ws.row_dimensions[32].height = 24
        ws.row_dimensions[33].height = 24
        ws.row_dimensions[34].height = 19

        wb.save(output)
        output.seek(0)
        self.generate_xls_file = base64.b64encode(output.getvalue()).decode('utf-8')
        # endregion

        return {
            "type": "ir.actions.act_url",
            "target": "self",
            "url": "/web/content?model=special.characteristics.matrix&download=true&field=generate_xls_file&filename={filename}.xlsx&id={pid}".format(
                filename="Special Characteristics Matrix", pid=self[0].id),
        }


class SpecialCharacterMatrixLine(models.Model):
    _name = 'special.characteristics.matrix.line'
    _inherit = "translation.mixin"

    special_id = fields.Many2one(comodel_name='special.characteristics.matrix',string='Special Characteristics Matrix',required=True, ondelete='cascade', index=True, copy=False)
    sequence = fields.Integer(string="Sequence", default=10)
    sl_no = fields.Integer("S.No", compute="_compute_sequence_number")
    parameter = fields.Char('Parameter',translate=True)
    symbole_designated = fields.Char("Symbole Designated",translate=True)
    specification = fields.Char('Specification',translate=True)
    process = fields.Char('Classification Of Process',translate=True)
    process_characterhistory = fields.Char('Relevant Process Characterhistory',translate=True)
    characterhistory = fields.Selection([('fit','FIT'),('function','FUNCTION'),('safety','SAFETY'),('govt.','GOVT'),('regulations','REGULATIONS'),('performance','REGULATIONS'),('customer designated','CUSTOMER DESIGNATED')],defaule='fit')


    @api.depends('sequence', 'special_id')
    def _compute_sequence_number(self):
        for line in self.mapped('special_id'):
            sl_no = 1
            for lines in line.special_line_ids:
                lines.sl_no = sl_no
                sl_no += 1
