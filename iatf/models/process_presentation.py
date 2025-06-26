from odoo import api, fields, models
import math
import openpyxl
from io import BytesIO
from odoo.exceptions import ValidationError
import io
from openpyxl import Workbook
import openpyxl
import base64
from bs4 import BeautifulSoup
from openpyxl.styles import Alignment, Font, Border, Side, DEFAULT_FONT, PatternFill
from openpyxl.drawing.image import Image
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
import copy
import pytz
from datetime import date
import datetime
from PIL import ImageOps
from PIL import Image as PILImage
import re
from xlsxwriter.contenttypes import defaults


class ProcessGroup(models.Model):
    _name = 'process.group'
    _inherit = ['iatf.sign.off.members', 'revision.history.mixin', 'translation.mixin']
    _description = 'Process Group'

    project = fields.Char("Project")
    assy_name = fields.Char("Assembly Name")
    org_date = fields.Date("Org. Date")
    rev_date = fields.Date("Rev. Date")
    rev_no = fields.Char("Revision No.")
    process_presentation_ids = fields.One2many(
        comodel_name='process.matrix.operation',
        inverse_name='process_id',
        string='Proces Matrix Operation'
    )
    generate_xls_file = fields.Binary(string="Generated file")
    process_matrix_category = fields.Many2one('presentation.category', string="Process Category")
    process_type=fields.Selection([
        ('machine', 'Machine'),
        ('assembly', 'Assembly'),
        ('machine_and_assembly', 'Machine and Assembly')

    ], string="Process Type")



    # @api.model_create_multi
    # def create(self, vals_list):
    #     records = super(ProcessGroup, self).create(vals_list)
    #
    #     for record in records:
    #         main_records = self.env["xf.doc.approval.document.package"].search([
    #             ('id', '=', record.project_id.id)
    #         ])
    #         record.write({
    #             'doc_type': main_records.doc_type,
    #             'partner_id': main_records.partner_id.id,
    #             'part_id': main_records.part_id.id,
    #         })
    #
    #     return records

    def update_process_flow(self):
        process_flows = self.env["process.flow"].search([('project_id', '=', self.project_id.id)])
        if not process_flows:
            raise ValidationError("Process Flow Diagram Not Created")
        for flow in process_flows:
            for operation in self.process_presentation_ids.sorted(key="operation"):
                # Find existing process flow line
                operation_element = self.env['process.flow.line'].search([
                    ('process_flow_id', '=', flow.id),
                    ('step', '=', operation.operation)
                ], limit=1)

                if operation_element:
                    # Update existing record
                    operation_element.write({
                        'desc_of_operation': operation.operation_description,
                        'cumulative_cy_time': operation.cumulative_cy_time,
                    })
                else:
                    # Create new process flow line
                    operation_element = self.env['process.flow.line'].create({
                        'process_flow_id': flow.id,
                        'step': operation.operation,
                        'desc_of_operation': operation.operation_description,
                        'cumulative_cy_time': operation.cumulative_cy_time,
                    })

                for i in operation.operation_lines_ids:
                    # Find existing process flow operations line
                    operation_line = self.env['process.flow.operations.line'].search([
                        ('operation_element_id', '=', operation_element.id),
                        ('element_no', '=', i.element_no)
                    ], limit=1)

                    if operation_line:
                        # Update existing operation line
                        operation_line.write({
                            'element_desc': i.element_description,
                            'product_characteristics': i.product_kpc.id,
                            'process_characteristics': i.process_kcc.id,
                            'remarks': i.remarks,
                            'cycle_time_in_min': i.cycle_time,
                            'equip_fixture_ids': [(6, 0, i.equipment_fixture_ids.ids)],
                            'tools_ids': [(6, 0, i.tool_ids.ids)],
                        })
                    else:
                        # Create new operation line if not found
                        self.env['process.flow.operations.line'].create({
                            'operation_element_id': operation_element.id,
                            'element_no': i.element_no,
                            'element_desc': i.element_description,
                            'product_characteristics': i.product_kpc.id,
                            'process_characteristics': i.process_kcc.id,
                            'remarks': i.remarks,
                            'cycle_time_in_min': i.cycle_time,
                            'equip_fixture_ids': [(6, 0, i.equipment_fixture_ids.ids)],
                            'tools_ids': [(6, 0, i.tool_ids.ids)],
                            'pokayoke_ids': [(6, 0, i.pokayoke_ids.ids)],
                        })

        return True

    def update_equipment_tracking_sheet(self):
        equip_rec = self.env["equipment.trackingsheet"].search([('project_id', '=', self.project_id.id)])

        for tracking_sheet in equip_rec:
            for operation in self.process_presentation_ids.sorted(key="operation"):
                for operation_line in operation.operation_lines_ids:
                    for equipment in operation_line.equipment_fixture_ids:  # Iterate over each equipment separately
                        # Search for an existing tracking line with the same equipment
                        tracking_line = self.env['equipment.trackingsheet.line'].search([
                            ('tracking_id', '=', tracking_sheet.id),
                            ('operation_no', '=', operation.operation),
                            ('element_no', '=', operation_line.element_no),
                            ('equipment_id', '=', equipment.id),  # Ensure each equipment is stored separately
                        ], limit=1)

                        if not tracking_line:
                            # Create a new tracking line for each equipment
                            self.env['equipment.trackingsheet.line'].create({
                                'tracking_id': tracking_sheet.id,
                                'element_no': operation_line.element_no,
                                'operation_no': operation.operation,
                                'equipment_id': equipment.id,  # Store each equipment as Many2one
                                'category_id': equipment.category_id.id,  # Maintain category info
                            })

        return True

    def update_tool_tracking_sheet(self):
        equip_rec = self.env["tools.trackingsheet"].search([('project_id', '=', self.project_id.id)])

        for tracking_sheet in equip_rec:
            for operation in self.process_presentation_ids.sorted(key="operation"):
                for operation_line in operation.operation_lines_ids:
                    for tool in operation_line.tool_ids:  # Iterate over each tool separately
                        # Search for an existing tracking line with the same tool
                        tracking_line = self.env['tools.trackingsheet.line'].search([
                            ('tracking_id', '=', tracking_sheet.id),
                            ('operation_no', '=', operation.operation),
                            ('element_no', '=', operation_line.element_no),
                            ('tool_name', '=', tool.id),  # Ensure each tool is stored separately
                        ], limit=1)

                        if not tracking_line:
                            # Create a new tracking line for each tool
                            self.env['tools.trackingsheet.line'].create({
                                'tracking_id': tracking_sheet.id,
                                'element_no': operation_line.element_no,
                                'operation_no': operation.operation,
                                'tool_name': tool.id,  # Store each tool as Many2one
                            })

        return True

    def action_process_presentation(self):
        output = BytesIO()
        wb = Workbook()
        ws = wb.active

        if self.env.user.company_id.logo:
            max_width = 500  # Set your desired maximum width
            max_height = 100  # Set your desired maximum height
            image_data = base64.b64decode(self.env.user.company_id.logo)

            # Open the image using PIL
            image = PILImage.open(io.BytesIO(image_data))
            width, height = image.size
            aspect_ratio = width / height

            if width > max_width:
                width = max_width
                height = int(width / aspect_ratio)

            if height > max_height:
                height = max_height
                width = int(height * aspect_ratio)

            # Resize the image using PIL
            # Add space on the top and left side of the image
            padding_top = 10  # Adjust as needed
            padding_left = 10  # Adjust as needed

            resized_image = image.resize((width, height), PILImage.Resampling.LANCZOS)
            ImageOps.expand(resized_image, border=(padding_left, padding_top, 0, 0), fill='rgba(0,0,0,0)')
            img_bytes = io.BytesIO()
            resized_image.save(img_bytes, format='PNG')
            img_bytes.seek(0)
            logo_image = Image(img_bytes)
            # logo_image = Image(self.env.user.company_id.logo)
            ws.add_image(logo_image, 'A1')

        border = Border(top=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'),
                        bottom=Side(style='thin'))
        font_header = Font(name='Times New Roman', size=11, bold=True)
        align_left = Alignment(vertical='center', horizontal='left', wrapText=True)
        align_center = Alignment(vertical='center', horizontal='center', wrapText=True)

        data = {
            'A1': 'Process Matrix',
            'A2': 'Project',
            'E2': 'Type(SafeLaunch/ Prototype/ PreLaunch/ Production)',
            'I2': 'Part Name',
            'M2': 'Part Number',
            'Q2': 'Customer Name',
            'U2': 'Org. Date',
            'W2': 'Rev. Date',
            'Y2': 'Rev. No.',
            'A3': 'Op No.',
            'B3': 'Operation description',
            'C3': 'Element No.',
            'D3': 'Element Description',
            'E3': 'Cycle Time',
            'F3': 'BOM-Item',
            'G3': 'BOM Part No.',
            'H3': 'Revision No.',
            'I3': 'BOM Qty',
            'J3': 'L-mm',
            'K3': 'W/OD-mm',
            'L3': 'T-mm',
            'M3': 'Component Weight - kg',
            'N3': 'manpower',
            'O3': 'Product - KPC',
            'P3': 'Process - KCC',
            'Q3': 'Equipment & Fixture',
            'R3': 'Tools',
            'S3': 'Pokayoke',
            'T3': 'Traceability System',
            'U3': 'Child Part in',
            'V3': 'Child Part / Sub Assy Out',
            'W3': 'Crane',
            'X3': 'Customize MHE',
            'Y3': 'Utility',
            'Z3': 'Miscl. Item',
            'AA3': 'Remarks',

        }
        # Fill data into specific cells using key-value pairs
        for cell, value in data.items():
            ws[cell] = value

        thin = Side(border_style='thin', color='000000')
        font_header = Font(name='Arial', size=13, bold=True)
        font_all = Font(name='Arial', size=11)
        align_center = Alignment(vertical='center', horizontal='center', wrapText=True)

        # region merging and formatting cells
        max_col = 27
        max_row = 68
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                alignment = copy.copy(cell.alignment)
                alignment.wrapText = True
                cell.alignment = align_center
                cell.font = font_all

        # Specific Dimension
        ws['A1'].alignment = align_center
        ws['A1'].font = Font(name='Times New Roman', size=24, bold=True)

        # Merging the cells as per standard sheet
        ws.merge_cells('A1:AA1')
        ws.merge_cells('A2:B2')
        ws.merge_cells('C2:D2')
        ws.merge_cells('E2:F2')
        ws.merge_cells('G2:H2')
        ws.merge_cells('I2:J2')
        ws.merge_cells('K2:L2')
        ws.merge_cells('M2:N2')
        ws.merge_cells('O2:P2')
        ws.merge_cells('Q2:R2')
        ws.merge_cells('S2:T2')
        ws.merge_cells('Z2:AA2')

        # Dimension of Columns
        ws.column_dimensions['A'].width = 8.0
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 35
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 19
        ws.column_dimensions['G'].width = 19
        ws.column_dimensions['H'].width = 19
        ws.column_dimensions['I'].width = 17
        ws.column_dimensions['J'].width = 17
        ws.column_dimensions['K'].width = 17
        ws.column_dimensions['L'].width = 22
        ws.column_dimensions['M'].width = 19
        ws.column_dimensions['N'].width = 19
        ws.column_dimensions['O'].width = 19
        ws.column_dimensions['P'].width = 21
        ws.column_dimensions['Q'].width = 16
        ws.column_dimensions['R'].width = 19
        ws.column_dimensions['S'].width = 19
        ws.column_dimensions['T'].width = 19
        ws.column_dimensions['U'].width = 21
        ws.column_dimensions['V'].width = 17
        ws.column_dimensions['W'].width = 18
        ws.column_dimensions['X'].width = 17
        ws.column_dimensions['Y'].width = 17
        ws.column_dimensions['Z'].width = 17
        ws.column_dimensions['AA'].width = 17

        # Dimension of Rows
        ws.row_dimensions[1].height = 75
        ws.row_dimensions[2].height = 35
        # for row_cell in range(3, 28 + 1):
        #     ws.row_dimensions[row_cell].height = 120
        font = ['A2', 'A3', 'B3', 'C3', 'D3', 'E3', 'F3', 'G3', 'H3', 'I3', 'J3', 'K3', 'L3', 'M3', 'N3', 'O3', 'P3',
                'Q3', 'R3',
                'S3', 'T3', 'U3', 'V3', 'W3', 'X3', 'Y3', 'Z3', 'AA3', 'E2', 'I2', 'M2', 'Q2', 'U2', 'W2', 'Y2']
        for font_col in font:
            ws[font_col].font = Font(name='Arial', bold=True, size=14)
            ws['A30'].alignment = align_center

        # Data Filling
        cur_row = 2
        for rec in self:
            ws[f'C{cur_row}'] = rec.project_id.name if rec.project_id else ""
            ws[f'G{cur_row}'] = rec.doc_type if rec.doc_type else ""
            ws[f'K{cur_row}'] = rec.part_name if rec.part_name else ""
            ws[f'O{cur_row}'] = rec.part_number if rec.part_number else ""
            ws[f'S{cur_row}'] = rec.partner_id.name if rec.partner_id.name else ""
            ws[f'V{cur_row}'] = rec.org_date if rec.org_date else ""
            ws[f'X{cur_row}'] = rec.rev_date if rec.rev_date else ""
            ws[f'Z{cur_row}'] = rec.rev_no if rec.rev_no else ""

            cur_row = 4
            for i in rec.process_presentation_ids:
                op_row = cur_row
                ws[f'A{cur_row}'] = i.operation if i.operation else ''
                ws[f'B{cur_row}'] = i.operation_description if i.operation_description else ''

                for line in i.operation_lines_ids:
                    ws[f'C{cur_row}'] = line.element_no if line.element_no else ''
                    ws[f'D{cur_row}'] = line.element_description if line.element_description else ''
                    ws[f'E{cur_row}'] = line.cycle_time if line.cycle_time else ''
                    ws[f'F{cur_row}'] = line.boq.product_id.name if line.boq else ''
                    ws[f'G{cur_row}'] = line.bom_part_no if line.bom_part_no else ''
                    ws[f'H{cur_row}'] = line.rev_no if line.rev_no else ''
                    ws[f'I{cur_row}'] = line.bom_qty if line.bom_qty else ''
                    ws[f'J{cur_row}'] = line.l_mm if line.l_mm else ''
                    ws[f'K{cur_row}'] = line.w_mm if line.w_mm else ''
                    ws[f'L{cur_row}'] = line.t_mm if line.t_mm else ''
                    ws[f'M{cur_row}'] = line.component_weight if line.component_weight else ''
                    ws[f'N{cur_row}'] = line.manpower if line.manpower else ''
                    ws[f'O{cur_row}'] = line.product_kpc.name if line.product_kpc else ''
                    ws[f'P{cur_row}'] = line.process_kcc.name if line.process_kcc else ''
                    ws[f'Q{cur_row}'] = ','.join(
                        [equipment_fixture_ids.name for equipment_fixture_ids in line.equipment_fixture_ids if
                         equipment_fixture_ids])
                    ws[f'R{cur_row}'] = ','.join([tool.name for tool in line.tool_ids if tool])
                    ws[f'S{cur_row}'] = ','.join(
                        [pokayoke_ids.poka_yoke_description for pokayoke_ids in line.pokayoke_ids if pokayoke_ids])
                    ws[f'T{cur_row}'] = ','.join(
                        [traceability_system_ids.name for traceability_system_ids in line.traceability_system_ids if
                         traceability_system_ids])
                    ws[f'U{cur_row}'] = line.Child_part if line.Child_part else ''
                    ws[f'V{cur_row}'] = line.Child_sub if line.Child_sub else ''
                    ws[f'W{cur_row}'] = line.crane if line.crane else ''
                    ws[f'X{cur_row}'] = line.customize if line.customize else ''
                    ws[f'Y{cur_row}'] = line.utility if line.utility else ''
                    ws[f'Z{cur_row}'] = line.miscl if line.miscl else ''
                    ws[f'AA{cur_row}'] = line.remarks if line.remarks else ''
                    ws.merge_cells(f'A{op_row}:A{cur_row}')
                    ws.merge_cells(f'B{op_row}:B{cur_row}')
                    cur_row += 1
                # cur_row += 1
        if cur_row < max_row:
            cur_row = max_row + 1
        # region SignOff Members Footer
        sign_row = cur_row
        ws.merge_cells(f'A{cur_row}:AA{cur_row}')
        cur_row += 1
        mer_end = cur_row
        ws[f'A{cur_row}'] = 'Prepared By'
        ws[f'A{cur_row}'].font = font_header
        ws.merge_cells(f'A{cur_row}:B{cur_row}')
        ws.merge_cells(f'C{cur_row}:F{cur_row}')

        ws[f'G{cur_row}'] = 'Prepared Date'
        ws[f'G{cur_row}'].font = font_header
        ws.merge_cells(f'G{cur_row}:H{cur_row}')
        ws.merge_cells(f'I{cur_row}:L{cur_row}')

        for rec in self:
            ws[f'C{cur_row}'] = rec.create_uid.name if rec.create_uid else ''
            ws[f'I{cur_row}'] = rec.create_date.strftime('%d-%m-%Y') if rec.create_date else ''

        ws.row_dimensions[cur_row].height = 18
        cur_row += 1
        ws.merge_cells(f'A{cur_row}:L{cur_row}')
        cur_row += 1
        ws.merge_cells(f'A{cur_row}:L{cur_row}')
        ws[f'A{cur_row}'] = 'Sign OFF'
        ws[f'A{cur_row}'].font = Font(size=18, bold=True)
        ws.row_dimensions[cur_row].height = 25
        cur_row += 1
        ws.merge_cells(f'A{cur_row}:L{cur_row}')
        cur_row += 1
        for cell in ws[cur_row]:
            ws[f'A{cur_row}'] = 'Member'
            ws[f'C{cur_row}'] = 'Department'
            ws[f'E{cur_row}'] = 'Status'
            ws[f'G{cur_row}'] = 'Date'
            ws[f'I{cur_row}'] = 'Comments'
            ws.row_dimensions[cur_row].height = 25
            ws.merge_cells(f'A{cur_row}:B{cur_row}')
            ws.merge_cells(f'C{cur_row}:D{cur_row}')
            ws.merge_cells(f'E{cur_row}:F{cur_row}')
            ws.merge_cells(f'G{cur_row}:H{cur_row}')
            ws.merge_cells(f'I{cur_row}:L{cur_row}')
            cell.font = Font(size=12, bold=True, color='ffffff')
            cell.fill = PatternFill(start_color='0070C0', end_color='0070C0', fill_type="solid")

        # Listing Managers Id
        for rec in self:
            for record in rec.iatf_members_ids:
                name_rec = record.approver_id.name
                dept_rec = record.department_id.name if record.department_id.name else ''
                status_rec = record.approval_status.capitalize()
                date_rec = record.date_approved_rejected.strftime(
                    '%d-%m-%Y') if record.date_approved_rejected else ''
                comment_rec = record.comment if record.comment else ''

                ws[f'A{cur_row + 1}'] = name_rec
                ws[f'C{cur_row + 1}'] = dept_rec
                ws[f'E{cur_row + 1}'] = status_rec
                ws[f'G{cur_row + 1}'] = date_rec
                ws[f'I{cur_row + 1}'] = comment_rec
                ws.merge_cells(f'A{cur_row + 1}:B{cur_row + 1}')
                ws.merge_cells(f'C{cur_row + 1}:D{cur_row + 1}')
                ws.merge_cells(f'E{cur_row + 1}:F{cur_row + 1}')
                ws.merge_cells(f'G{cur_row + 1}:H{cur_row + 1}')
                ws.merge_cells(f'I{cur_row + 1}:L{cur_row + 1}')

                status_dict = {'Approved': ['CFFFC3', '000000'], 'Rejected': ['FFCDCD', '000000'],
                               'Revision': ['AFEFFF', '000000'], 'Pending': ['FDFFCD', '000000']}

                for state, color in status_dict.items():
                    if state == status_rec:
                        ws[f'E{cur_row + 1}'].fill = PatternFill(start_color=color[0], end_color=color[0],
                                                                 fill_type="solid")
                        ws[f'E{cur_row + 1}'].font = Font(size=12, bold=False, color=color[1])

                cur_row += 1

            ws.merge_cells(f'A{cur_row + 1}:AA{cur_row + 1}')
            ws.merge_cells(f'M{mer_end}:AA{cur_row}')

            for row_no in ws.iter_rows(min_row=sign_row + 1, max_row=cur_row + 1, min_col=1, max_col=27):
                for cell in row_no:
                    cell.border = border
                    cell.alignment = align_center

        wb.save(output)
        output.seek(0)
        self.generate_xls_file = base64.b64encode(output.getvalue()).decode('utf-8')

        return {
            "type": "ir.actions.act_url",
            "target": "self",
            "url": "/web/content?model=process.group&download=true&field=generate_xls_file&filename={filename}.xlsx&id={pid}".format(
                filename="LIST OF PROCESS PRESENTATION", pid=self[0].id),
        }


class ProcessCategory(models.Model):
    _name = 'presentation.category'
    _inherit = "translation.mixin"
    # _rec_name= 'name'

    name = fields.Char('Category',translate=True)

# class PresentationPokeYoke(models.Model):
#     _name = 'presentation.poke.yoke'
#     name = fields.Char('Poke Yoke')
#     poka_num = fields.Char('Poke Number')


class PresentationTraceabilitySystem(models.Model):
    _name = 'presentation.traceability.system'
    _inherit = "translation.mixin"
    name = fields.Char('Traceability System',translate=True)


class UsedTool(models.Model):
    _name = 'presentation.tool'
    _inherit = "translation.mixin"

    name = fields.Char('Tool',translate=True)


class ProcessMatrixOperation(models.Model):
    _name = 'process.matrix.operation'
    _description = "Process Operation"
    _inherit = "translation.mixin"

    process_id = fields.Many2one('process.group', 'Process Matrix')

    sq_handle = fields.Integer("Sequence no")
    operation = fields.Char("Operation", store=True,translate=True)
    operation_description = fields.Char(string="Operation Description",translate=True)
    operation_lines_ids = fields.One2many(
        comodel_name='operation.element.line',
        inverse_name='operation_element_id',
        string="Operations Lines")
    cumulative_cy_time = fields.Integer("Cumulative Cycle Time(sec)", compute='_compute_total_amount')

    def _recalculate_sequence_numbers(self):
        """Recalculate sequence_handle and element_no for all elements in the operation."""
        for operation in self:
            # Get elements sorted by existing sequence_handle (ensures correct order)
            elements = operation.operation_lines_ids.sorted(lambda e: e.sequence_handle)

            seq = 1  # Start numbering from 1
            updates = []  # Store updates to batch apply

            for element in elements:
                operation_number = re.findall(r'\d+', operation.operation or "")
                if operation_number:
                    base_number = operation_number[0]  # Extracted integer (e.g., "10" from "OP-10")
                    new_element_no = ".".join([base_number, str(seq)])

                    # Only update if the values have changed
                    if element.sequence_handle != seq or element.element_no != new_element_no:
                        updates.append((element.id, seq, new_element_no))

                seq += 1

            # Batch update to avoid multiple writes
            if updates:
                for element_id, new_seq, new_element_no in updates:
                    self.env['operation.element.line'].browse(element_id).sudo().write({
                        'sequence_handle': new_seq,
                        'element_no': new_element_no
                    })

    # def create(self, vals):
    #     records = super(ProcessMatrixOperation, self).create(vals)
    #     sn_o = 1
    #     for rec in records:
    #         rec.sq_handle = sn_o
    #         sn_o += 1
    #     return records
    #
    # @api.depends('sq_handle', 'process_id')
    # def _compute_sequence_number(self):
    #     for rec in self:
    #         rec.operation = rec.sq_handle * 10

    @api.depends('operation_lines_ids.cycle_time')
    def _compute_total_amount(self):
        for record in self:
            record.cumulative_cy_time = sum(record.operation_lines_ids.mapped('cycle_time'))


class ProcessMatrixOperationElement(models.Model):
    _name = 'operation.element.line'
    _inherit = "translation.mixin"

    operation_element_id = fields.Many2one('process.matrix.operation', 'Process Operation Line')

    sequence_handle = fields.Integer(string="Sequence no", default=1)
    element_no = fields.Char(string="Element No", readonly=True)
    # station = fields.Many2one("mrp.workcenter", string="Station/Work Center")
    element_description = fields.Char(string="Element Description",translate=True)
    cycle_time = fields.Integer("Cycle Time (sec)")
    boq = fields.Many2one("mrp.bom.line", string="BOM-Item")
    bom_part_no = fields.Char(related="boq.product_id.default_code", string="BOM Part Number", store=True)
    rev_no = fields.Char("Revision No.")
    bom_qty = fields.Float("BOM Qty", related='boq.product_qty')
    l_mm = fields.Char("L-mm")
    w_mm = fields.Char("W/OD-mm")
    t_mm = fields.Char("T-mm")
    component_weight = fields.Char("Component Weight - kg")
    manpower = fields.Char("manpower")
    product_kpc = fields.Many2one('product.characteristics', 'Product Characteristics')
    process_kcc = fields.Many2one('process.characteristics', 'Process Characteristics')
    special_characteristics = fields.Many2one('process.flow.class', 'Class')
    equipment_fixture_ids = fields.Many2many(
        'maintenance.equipment',
        'operation_element_equipment_rel',  # Unique relation table name
        'operation_id',  # First column
        'equipment_id',  # Second column
        string="Equipment & Fixture",
        domain="[('category_id', 'in', ['Equipment', 'Fixture'])]")
    tool_ids = fields.Many2many(
        'maintenance.equipment',
        'operation_element_tool_rel',  # Unique relation table name
        'operation_id',
        'tool_id',
        string="Tools",
        domain="[('category_id', 'in', ['Tool'])]"
    )
    gauge_id = fields.Many2one(
        'maintenance.equipment',
        string="Gauge",
        domain="[('category_id', 'in', ['Gauge'])]",
        context="{'default_category_id': 'Gauge'}"
    )
    op = fields.Char(related='operation_element_id.operation', string="Operation", store=True)

    pokayoke_ids = fields.Many2many('poka.yoka.line', string="Pokayoke", domain="[('operation', '=', op)]")
    traceability_system_ids = fields.Many2many('presentation.traceability.system', string="Traceability System")
    Child_part = fields.Char(string="Child Part in")
    Child_sub = fields.Char(string="Child Part / Sub Assy Out")
    crane = fields.Char("Crane")
    customize = fields.Char("Customize MHE ")
    utility = fields.Char("Utility")
    miscl = fields.Char("Miscl. Item")
    remarks = fields.Char("Remarks")

    @api.model
    def create(self, vals):
        """Ensure sequence and element_no are correct and ordered when adding a new element"""
        record = super(ProcessMatrixOperationElement, self).create(vals)
        record.operation_element_id._recalculate_sequence_numbers()  # Update all numbers
        return record

    def write(self, vals):
        """Ensure sequence updates when modified"""
        result = super(ProcessMatrixOperationElement, self).write(vals)
        for record in self:
            record.operation_element_id._recalculate_sequence_numbers()  # Update all numbers
        return result
