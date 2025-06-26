import base64
import io
from io import BytesIO

from PIL import Image as PILImage
from PIL import ImageOps
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

from odoo import fields, models, api

from odoo.fields import first

class ControlPlan(models.Model):
    _name = "control.plan.assy"
    _description = "Control Plan and Inspection-II"
    _inherit = ['iatf.sign.off.members', 'translation.mixin']

    type = fields.Selection([('safelaunch', 'Safe Launch'), ('prototype', 'Prototype'), ('prelaunch', 'PreLaunch'),
                             ('production', 'Production')])
    assy_no = fields.Char('Assy. No.')
    operation_no = fields.Char('Operation Number')
    effective_date = fields.Date('Effective Date')
    document = fields.Char('Document',translate=True)
    rev_level = fields.Char('Rev. Level',translate=True)
    rev_date = fields.Date('Rev. Date')
    rev_no = fields.Char('Rev. No.')
    rev_details = fields.Char('Rev. Details',translate=True)
    revised_by = fields.Many2one('res.users','Revised By')
    page_no = fields.Char('Page No.')
    # prepared_by = fields.Many2one('res.users', 'Prepared By')
    # checked_by = fields.Many2one('res.users', 'Checked By')
    # production = fields.Char('Production')
    # quality = fields.Char('Quality')

    procedure_line_ids = fields.One2many(
        comodel_name='control.procedure.line',
        inverse_name='procedure_id',
        string='Control Process Assy Line'
    )
    extra_care = fields.Many2many('extra.care.alerts.line', 'extra_care_relation', 'extra_id', 'alerts_id')
    jigs_fixture_tools = fields.Many2many('jigs.fixture.tools.line', 'jigs_fixture_tools_relation', 'fixture_id',
                                          'tools_id')
    consumable_required = fields.Many2many('consumable.required.line', 'consumable_required_relation', 'consumable_id',
                                           'required_id')
    product_char_ids = fields.One2many(
        comodel_name='control.product.char.line',
        inverse_name='product_char_id',
        string='Control Product Characteristics'
    )

    generate_xls_file = fields.Binary(string="Generated file")  # do not comment this

    @api.model_create_multi
    def create(self, vals_list):
        records = super(ControlPlan, self).create(vals_list)

        for record in records:
            cp_2_records = self.env["process.flow"].search(
                [('project_id', '=', record.project_id.id),
                 ('final_status', '=', 'approved')
                 ])
            # if not pm_records:
            #     raise ValidationError(
            #         _("No Process Matrix is filled in this project first fill Process Matrix before continuing..."))

            for op in cp_2_records.process_flow_line_ids.sorted(key="step"):

                for i in op.process_op_lines_ids:
                    self.env['control.procedure.line'].create({
                        'procedure_id': record.id,
                        'procedure_step': i.element_no,
                        'procedure': i.element_desc,
                    })

        return records

    def action_generate_excel_report(self):
        output = BytesIO()
        wb = Workbook()
        ws = wb.active

        if self.env.user.company_id.logo:
            max_width = 160  # Set your desired maximum width
            max_height = 100  # Set your desired maximum height
            image_data = base64.b64decode(self.env.user.company_id.logo)

            # Open the image using PIL
            image = PILImage.open(io.BytesIO(image_data))
            width, height = image.size
            aspect_ratio = width / height

            if width > max_width:
                width = max_width
                height = int(width / aspect_ratio)

            if height > max_height:
                height = max_height
                width = int(height * aspect_ratio)

            # Resize the image using PIL.lo
            # Add space on the top and left side of the image
            padding_top = 10  # Adjust as needed
            padding_left = 10  # Adjust as needed

            resized_image = image.resize((width, height), PILImage.LANCZOS)
            ImageOps.expand(resized_image, border=(padding_left, padding_top, 0, 0), fill='rgba(0,0,0,0)')
            img_bytes = io.BytesIO()
            resized_image.save(img_bytes, format='PNG')
            img_bytes.seek(0)
            logo_image = Image(img_bytes)
            # logo_image = Image(self.env.user.company_id.logo)
            ws.add_image(logo_image, 'A1')

        border = Border(top=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'),
                        bottom=Side(style='thin'))
        align_center = Alignment(vertical='center', horizontal='center', wrapText=True)
        align_left = Alignment(vertical='center', horizontal='left')
        font_header = Font(name='Arial', size=12, bold=True)
        font_all = Font(name='Times New Roman', size=11, bold=False)

        data = {
            'C1': 'Control Plan and Inspection-II',
            'D3': 'Safe Launch',
            'F3': 'Prototype',
            'I3': 'PreLaunch',
            'L3': 'Production',
            'A4': 'Assembly Number',
            'G4': 'Work Station',
            'N1': 'Effective Date',
            'N2': 'Document',
            'N3': 'Rev. level',
            'N4': 'Page No',
            'Q1': 'Rev. Date',
            'Q2': 'Rev. No',
            'Q3': 'Rev. Details',
            'Q4': 'Revised By',
            'A5': 'Steps',
            'B5': 'Procedure',
            'J5': 'Materials',
            'M5': 'Reference Diagram',
            'J6': 'BOM',
            'K6': 'Qty.',
            'L6': 'Part',
        }

        for cell, value in data.items():
            ws[cell] = value

        max_row = 25

        cell_ranges_to_merge = ['A1:B3', 'C1:M2', 'F3:G3', 'I3:J3', 'L3:M3', 'A4:B4', 'C4:F4',
                                'G4:I4', 'J4:M4', 'A5:A6', 'B5:I6', 'J5:L5', 'M5:S6', ]
        for cell_mer in cell_ranges_to_merge:
            ws.merge_cells(cell_mer)

        for row in ws.iter_rows(min_row=1, max_row=50, min_col=1, max_col=19):
            for cell in row:
                cell.border = border
                cell.alignment = align_center

        for i in range(1, 5):
            ws.merge_cells(f'N{i}:O{i}')
            ws.merge_cells(f'Q{i}:R{i}')
            ws.row_dimensions[i].height = 20
            ws[f'N{i}'].fill = PatternFill(start_color='B1F0F7', end_color='B1F0F7', fill_type="solid")
            ws[f'Q{i}'].fill = PatternFill(start_color='B1F0F7', end_color='B1F0F7', fill_type="solid")
        for col in ws.iter_cols(min_row=5, max_row=6):
            for cell in col:
                cell.fill = PatternFill(start_color='B1F0F7', end_color='B1F0F7', fill_type="solid")

        ws['C1'].font = Font(name='Arial', size=20, bold=True)
        ws[f'D{3}'].fill = PatternFill(start_color='B1F0F7', end_color='B1F0F7', fill_type="solid")
        ws[f'F{3}'].fill = PatternFill(start_color='B1F0F7', end_color='B1F0F7', fill_type="solid")
        ws[f'I{3}'].fill = PatternFill(start_color='B1F0F7', end_color='B1F0F7', fill_type="solid")
        ws[f'L{3}'].fill = PatternFill(start_color='B1F0F7', end_color='B1F0F7', fill_type="solid")
        ws[f'A{4}'].fill = PatternFill(start_color='B1F0F7', end_color='B1F0F7', fill_type="solid")
        ws[f'G{4}'].fill = PatternFill(start_color='B1F0F7', end_color='B1F0F7', fill_type="solid")

        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 8
        ws.column_dimensions['D'].width = 24
        ws.column_dimensions['E'].width = 8
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 8
        ws.column_dimensions['I'].width = 18
        ws.column_dimensions['J'].width = 15
        ws.column_dimensions['K'].width = 9
        ws.column_dimensions['L'].width = 18
        ws.column_dimensions['M'].width = 8.5
        ws.column_dimensions['N'].width = 19
        ws.column_dimensions['O'].width = 19
        ws.column_dimensions['P'].width = 33
        ws.column_dimensions['Q'].width = 13
        ws.column_dimensions['R'].width = 18
        ws.column_dimensions['S'].width = 29

        # Data Filling
        cur_row = 7
        for i in ws[cur_row - 2]:
            i.font = font_header
        for j in ws[cur_row - 1]:
            j.font = font_header
        for rec in self:
            ws['C4'] = rec.assy_no if rec.assy_no else ''
            ws['J4'] = rec.operation_no if rec.operation_no else ''
            ws['P1'] = rec.effective_date if rec.effective_date else ''
            ws['P2'] = rec.document if rec.document else ''
            ws['P3'] = rec.rev_level if rec.rev_level else ''
            ws['P4'] = rec.page_no if rec.page_no else ''
            ws['S1'] = rec.rev_date if rec.rev_date else ''
            ws['S2'] = rec.rev_no if rec.rev_no else ''
            ws['S3'] = rec.rev_details if rec.rev_details else ''
            ws['S4'] = rec.revised_by.name if rec.revised_by.name else ''
            ws['P4'] = "01"

            ws['C3'] = '☑' if rec.type == 'safelaunch' else '☐'
            ws['E3'] = '☑' if rec.type == 'prototype' else '☐'
            ws['H3'] = '☑' if rec.type == 'prelaunch' else '☐'
            ws['K3'] = '☑' if rec.type == 'production' else '☐'
            ws['C3'].font = font_header
            ws['E3'].font = font_header
            ws['H3'].font = font_header
            ws['K3'].font = font_header

            count = 0
            for i in rec.procedure_line_ids:
                ws[f'A{cur_row}'] = i.procedure_step if i.procedure_step else ''
                ws[f'B{cur_row}'] = i.procedure if i.procedure else ''
                material_row = cur_row
                for j in i.material_line_ids:
                    ws[f'J{material_row}'] = j.bom if j.bom else ''
                    ws[f'K{material_row}'] = j.qty if j.qty else ''
                    ws[f'L{material_row}'] = j.part if j.part else ''
                    material_row+=1
                if i.procedure_image:
                    image_data = base64.b64decode(i.procedure_image)
                    # Open the image using PIL
                    image = PILImage.open(io.BytesIO(image_data))
                    width, height = image.size

                    width = 325
                    height = 217

                    resized_image = image.resize((width, height), PILImage.LANCZOS)
                    bordered_image = ImageOps.expand(resized_image, border=(2,2,2,2), fill='black')
                    img_bytes = io.BytesIO()
                    bordered_image.save(img_bytes, format='PNG')
                    img_bytes.seek(0)
                    image_to_add = Image(img_bytes)
                    # logo_image = Image(self.env.user.company_id.logo)
                    print(count)
                    if count==0:
                        ws.add_image(image_to_add, 'M7')
                        ws['M18'] = '01'
                    elif count==1:
                        ws.add_image(image_to_add, 'P7')
                        ws['P18'] = '02'
                    elif count==2:
                        ws.add_image(image_to_add, 'R7')
                        ws['R18'] = '03'
                    elif count==3:
                        ws.add_image(image_to_add, 'M19')
                        ws['M30'] = '04'
                    elif count==4:
                        ws.add_image(image_to_add, 'P19')
                        ws['P30'] = '05'
                    elif count==5:
                        ws.add_image(image_to_add, 'R19')
                        ws['R30'] = '06'
                    elif count==6:
                        ws.add_image(image_to_add, 'M31')
                        ws['M42'] = '07'
                    elif count==7:
                        ws.add_image(image_to_add, 'P31')
                        ws['P42'] = '08'
                    elif count==8:
                        ws.add_image(image_to_add, 'R31')
                        ws['R42'] = '09'
                    count+=1

                ws[f'B{cur_row}'].alignment = align_left
                ws.merge_cells(f'A{cur_row}:A{cur_row + 2}')
                ws.merge_cells(f'B{cur_row}:I{cur_row + 2}')

                cur_row += 3

        # print('uygyhv',cur_row)
        # print('......',max_row)
        if cur_row < max_row:
            for i in range(cur_row, max_row, 2):
                ws.merge_cells(f'A{cur_row}:A{cur_row + 2}')
                ws.merge_cells(f'B{cur_row}:I{cur_row + 2}')
                cur_row += 3
                if cur_row > max_row:
                    break
                # print(cur_row)
                # print(max_row)

        # Extra Care Alerts
        # print("ex1_cur", cur_row)
        # print("ex1_max", max_row)
        max_row = cur_row + 6
        for rec in self:
            ws[f'A{cur_row}'] = "Extra Care Alerts"
            ws[f'A{cur_row}'].alignment = align_left
            ws[f'A{cur_row}'].font = font_header
            ws[f'A{cur_row}'].fill = PatternFill(start_color='B1F0F7', end_color='B1F0F7', fill_type="solid")

            ws.merge_cells(f'A{cur_row}:I{cur_row}')
            # ws.row_dimensions[cur_row].height = 22

            sl_no = 1
            for i in rec.extra_care:
                ws[f'A{cur_row + 1}'] = sl_no
                ws[f'B{cur_row + 1}'] = i.name if i.name else ''
                ws[f'B{cur_row + 1}'].alignment = align_left
                ws.merge_cells(f'B{cur_row + 1}:I{cur_row + 1}')
                cur_row += 1
                sl_no += 1

        # print("ex2_cur", cur_row)
        # print("ex2_max", max_row)
        if cur_row < max_row:
            for i in range(cur_row, max_row):
                ws.merge_cells(f'B{cur_row}:I{cur_row}')
                cur_row += 1
                if cur_row > max_row:
                    break

        # Jig, Fixtures and Tools Required
        max_row = cur_row + 6
        for rec in self:
            ws[f'A{cur_row}'] = "Jig, Fixtures and Tools Required"
            ws[f'A{cur_row}'].alignment = align_left
            ws[f'A{cur_row}'].font = font_header
            ws.merge_cells(f'A{cur_row}:I{cur_row}')
            ws[f'A{cur_row}'].fill = PatternFill(start_color='B1F0F7', end_color='B1F0F7', fill_type="solid")
            # ws.row_dimensions[cur_row].height = 22

            sl_no = 1
            for i in rec.jigs_fixture_tools:
                ws[f'A{cur_row + 1}'] = sl_no
                ws[f'B{cur_row + 1}'] = i.name if i.name else ''
                ws[f'B{cur_row + 1}'].alignment = align_left
                ws.merge_cells(f'B{cur_row + 1}:I{cur_row + 1}')
                cur_row += 1
                sl_no += 1

        # print("jig_cur", cur_row)
        # print("jig_max", max_row)
        if cur_row < max_row:
            for i in range(cur_row, max_row):
                ws.merge_cells(f'B{cur_row}:I{cur_row}')
                cur_row += 1
                if cur_row > max_row:
                    break

        # Consumables Required
        max_row = cur_row + 6
        for rec in self:
            ws[f'A{cur_row}'] = "Consumables Required"
            ws[f'A{cur_row}'].alignment = align_left
            ws[f'A{cur_row}'].font = font_header
            ws.merge_cells(f'A{cur_row}:I{cur_row}')
            ws[f'A{cur_row}'].fill = PatternFill(start_color='B1F0F7', end_color='B1F0F7', fill_type="solid")
            # ws.row_dimensions[cur_row].height = 22

            sl_no = 1
            for i in rec.consumable_required:
                ws[f'A{cur_row + 1}'] = sl_no
                ws[f'B{cur_row + 1}'] = i.name if i.name else ''
                ws[f'B{cur_row + 1}'].alignment = align_left
                ws.merge_cells(f'B{cur_row + 1}:I{cur_row + 1}')
                cur_row += 1
                sl_no += 1

        # print("con_cur", cur_row)
        # print("con_max", max_row)
        if cur_row < max_row:
            for i in range(cur_row, max_row):
                ws.merge_cells(f'B{cur_row}:I{cur_row}')
                cur_row += 1
                if cur_row > max_row:
                    break

        # Control Product Characteristics
        pic_end_row = cur_row - 1
        max_row = cur_row + 8
        for rec in self:
            ws[f'A{cur_row}'] = "Control Product Characteristics"
            ws[f'A{cur_row}'].alignment = align_left
            ws[f'A{cur_row}'].font = font_header
            ws.merge_cells(f'A{cur_row}:S{cur_row}')
            ws.row_dimensions[cur_row].height = 22
            ws[f'A{cur_row}'].fill = PatternFill(start_color='B1F0F7', end_color='B1F0F7', fill_type="solid")

            cur_row += 1

            ws[f'A{cur_row}'] = "Sl.No."
            ws[f'B{cur_row}'] = "Vital Assembly Parameters"
            ws[f'F{cur_row}'] = "Poka-Yoke"
            ws[f'H{cur_row}'] = "Measurements"
            ws[f'N{cur_row}'] = "Inspection Frequency"
            ws[f'P{cur_row}'] = "Control Method"
            ws[f'Q{cur_row}'] = "Reaction Plan"
            ws[f'Q{cur_row + 1}'] = "Action"
            ws[f'S{cur_row + 1}'] = "Owner/ Responsible"
            ws[f'H{cur_row + 1}'] = "Product Specification"
            ws[f'J{cur_row + 1}'] = "Process Specification"
            ws[f'L{cur_row + 1}'] = "Technique"
            ws[f'N{cur_row + 1}'] = "Assembler"
            ws[f'O{cur_row + 1}'] = "Q.A Inspector"

            for i in ws[cur_row]:
                i.font = font_header
            for j in ws[cur_row + 1]:
                j.font = font_header
            for col in ws.iter_cols(min_row=cur_row, max_row=cur_row+1):
                for cell in col:
                    cell.fill = PatternFill(start_color='B1F0F7', end_color='B1F0F7', fill_type="solid")

            ws.merge_cells(f'A{cur_row}:A{cur_row + 1}')
            ws.merge_cells(f'B{cur_row}:E{cur_row + 1}')
            ws.merge_cells(f'F{cur_row}:G{cur_row + 1}')
            ws.merge_cells(f'H{cur_row}:M{cur_row}')
            ws.merge_cells(f'N{cur_row}:O{cur_row}')
            ws.merge_cells(f'P{cur_row}:P{cur_row + 1}')
            ws.merge_cells(f'Q{cur_row}:S{cur_row}')
            ws.merge_cells(f'Q{cur_row + 1}:R{cur_row + 1}')
            ws.merge_cells(f'H{cur_row + 1}:I{cur_row + 1}')
            ws.merge_cells(f'J{cur_row + 1}:K{cur_row + 1}')
            ws.merge_cells(f'L{cur_row + 1}:M{cur_row + 1}')

            cur_row += 2
            for i in rec.product_char_ids:
                ws[f'A{cur_row}'] = i.sl_no if i.sl_no else ''
                ws[f'B{cur_row}'] = i.vital_assy_parameters if i.vital_assy_parameters else ''
                ws[f'F{cur_row}'] = i.poka_yoke if i.poka_yoke else ''
                ws[f'H{cur_row}'] = i.m_product_specification if i.m_product_specification else ''
                ws[f'J{cur_row}'] = i.m_process_specification if i.m_process_specification else ''
                ws[f'L{cur_row}'] = i.m_technique if i.m_technique else ''
                ws[f'N{cur_row}'] = i.i_assembler if i.i_assembler else ''
                ws[f'O{cur_row}'] = i.i_quality_inspector if i.i_quality_inspector else ''
                ws[f'P{cur_row}'] = i.control_method if i.control_method else ''
                ws[f'Q{cur_row}'] = i.reaction_action if i.reaction_action else ''
                ws[f'S{cur_row}'] = i.reaction_owner_responsible if i.reaction_owner_responsible else ''
                ws[f'B{cur_row}'].alignment = align_left

                ws.merge_cells(f'B{cur_row}:E{cur_row}')
                ws.merge_cells(f'F{cur_row}:G{cur_row}')
                ws.merge_cells(f'H{cur_row}:I{cur_row}')
                ws.merge_cells(f'J{cur_row}:K{cur_row}')
                ws.merge_cells(f'L{cur_row}:M{cur_row}')
                ws.merge_cells(f'Q{cur_row}:R{cur_row}')

                cur_row += 1

        # print("cc_cur", cur_row)
        # print("cc_max", max_row)
        if cur_row < max_row:
            for i in range(cur_row, max_row):
                ws.merge_cells(f'B{cur_row}:E{cur_row}')
                ws.merge_cells(f'F{cur_row}:G{cur_row}')
                ws.merge_cells(f'H{cur_row}:I{cur_row}')
                ws.merge_cells(f'J{cur_row}:K{cur_row}')
                ws.merge_cells(f'L{cur_row}:M{cur_row}')
                ws.merge_cells(f'Q{cur_row}:R{cur_row}')
                for j in ws[cur_row]:
                    j.border = border
                cur_row += 1
                if cur_row > max_row:
                    break

        # Reference Diagram Merging
        s_row = 7
        counter = 0
        for i in range(s_row,pic_end_row, 12):
            ws.merge_cells(f'M{i}:O{i + 10}')
            ws.merge_cells(f'P{i}:Q{i + 10}')
            ws.merge_cells(f'R{i}:S{i + 10}')
            ws.merge_cells(f'M{i+11}:O{i + 11}')
            ws.merge_cells(f'P{i+11}:Q{i + 11}')
            ws.merge_cells(f'R{i+11}:S{i + 11}')
            counter+=1
            # print("s_row",s_row)
            # print("pic_row",pic_end_row)
            if counter==3:
                break

        # region SignOff Members Footer
        sign_row = cur_row
        ws.merge_cells(f'A{cur_row}:S{cur_row}')

        cur_row += 1
        ws[f'A{cur_row}'] = 'Prepared By'
        ws[f'A{cur_row}'].font = font_header
        ws.merge_cells(f'A{cur_row}:B{cur_row}')
        ws.merge_cells(f'C{cur_row}:E{cur_row}')

        ws[f'F{cur_row}'] = 'Prepared Date'
        ws[f'F{cur_row}'].font = font_header
        ws.merge_cells(f'F{cur_row}:G{cur_row}')
        ws.merge_cells(f'H{cur_row}:L{cur_row}')

        for rec in self:
            ws[f'C{cur_row}'] = rec.create_uid.name if rec.create_uid else ''
            ws[f'H{cur_row}'] = rec.create_date if rec.create_date else ''

        ws.row_dimensions[cur_row].height = 18
        cur_row += 1
        ws.merge_cells(f'A{cur_row}:L{cur_row}')
        cur_row += 1
        ws.merge_cells(f'A{cur_row}:L{cur_row}')
        ws[f'A{cur_row}'] = 'Sign OFF'
        ws[f'A{cur_row}'].font = Font(size=18, bold=True)
        ws.row_dimensions[cur_row].height = 25
        cur_row += 1
        ws.merge_cells(f'A{cur_row}:L{cur_row}')
        cur_row += 1
        for cell in ws[cur_row]:
            ws[f'A{cur_row}'] = 'Member'
            ws[f'D{cur_row}'] = 'Department'
            ws[f'F{cur_row}'] = 'Status'
            ws[f'H{cur_row}'] = 'Date'
            ws[f'J{cur_row}'] = 'Comments'
            ws.row_dimensions[cur_row].height = 25
            ws.merge_cells(f'A{cur_row}:C{cur_row}')
            ws.merge_cells(f'D{cur_row}:E{cur_row}')
            ws.merge_cells(f'F{cur_row}:G{cur_row}')
            ws.merge_cells(f'H{cur_row}:I{cur_row}')
            ws.merge_cells(f'J{cur_row}:L{cur_row}')
            cell.font = Font(size=12, bold=True, color='ffffff')
            cell.fill = PatternFill(start_color='0070C0', end_color='0070C0', fill_type="solid")

        # Listing Managers Id
        for rec in self:
            for record in rec.iatf_members_ids:
                name_rec = record.approver_id.name
                dept_rec = record.department_id.name if record.department_id.name else ''
                status_rec = record.approval_status.capitalize()
                date_rec = record.date_approved_rejected if record.date_approved_rejected else ''
                comment_rec = record.comment if record.comment else ''

                ws[f'A{cur_row + 1}'] = name_rec
                ws[f'C{cur_row + 1}'] = dept_rec
                ws[f'D{cur_row + 1}'] = status_rec
                ws[f'E{cur_row + 1}'] = date_rec
                ws[f'F{cur_row + 1}'] = comment_rec
                ws.merge_cells(f'A{cur_row + 1}:C{cur_row + 1}')
                ws.merge_cells(f'D{cur_row + 1}:E{cur_row + 1}')
                ws.merge_cells(f'F{cur_row + 1}:G{cur_row + 1}')
                ws.merge_cells(f'H{cur_row + 1}:I{cur_row + 1}')
                ws.merge_cells(f'J{cur_row + 1}:L{cur_row + 1}')

                status_dict = {'Approved': ['CFFFC3', '000000'], 'Rejected': ['FFCDCD', '000000'],
                               'Revision': ['AFEFFF', '000000'], 'Pending': ['FDFFCD', '000000']}

                for state, color in status_dict.items():
                    if state == status_rec:
                        ws[f'F{cur_row + 1}'].fill = PatternFill(start_color=color[0], end_color=color[0],
                                                                 fill_type="solid")
                        ws[f'F{cur_row + 1}'].font = Font(size=12, bold=False, color=color[1])

                cur_row += 1

            # ws.merge_cells(f'C{cur_row + 1}:D{cur_row + 1}')
            # ws.merge_cells(f'E{cur_row + 1}:F{cur_row + 1}')
            ws.merge_cells(f'A{cur_row + 1}:S{cur_row + 1}')
            ws.merge_cells(f'M{sign_row + 1}:S{cur_row}')

            for row_no in ws.iter_rows(min_row=sign_row, max_row=cur_row + 1, min_col=1, max_col=19):
                for cell in row_no:
                    cell.border = border
                    cell.alignment = align_center
            # endregion


        # Save the workbook
        wb.save(output)
        output.seek(0)
        self.generate_xls_file = base64.b64encode(output.getvalue()).decode('utf-8')
        # endregion

        return {
            "type": "ir.actions.act_url",
            "target": "self",
            "url": "/web/content?model=control.plan.assy&download=true&field=generate_xls_file&filename={filename}.xlsx&id={pid}".format(
                filename="Control Plan Assy", pid=self[0].id),
        }


class ControlProcedureAssyLine(models.Model):
    _name = "control.procedure.line"
    _description = "Control Procedure line"
    _inherit = "translation.mixin"

    procedure_id = fields.Many2one('control.plan.assy', 'Control Procedure Line')

    sequence = fields.Integer(string="Sequence", default=10)
    procedure_step = fields.Char("Step",translate=True)
    # procedure_step = fields.Char("Steps")
    procedure = fields.Char("Procedure",translate=True)
    procedure_image = fields.Image("Procedure Image")
    material_line_ids = fields.One2many(
        comodel_name='procedure.material.line',
        inverse_name='material_id',
        string='Materials Line'
    )

    @api.depends('sequence', 'procedure_id')
    def _compute_sequence_num(self):
        for line in self.mapped('procedure_id'):
            procedure_step = 1
            for lines in line.procedure_line_ids:
                lines.procedure_step = procedure_step
                procedure_step += 1


class ControlProcedureMaterialLine(models.Model):
    _name = "procedure.material.line"
    _description = "Control Procedure Material line"
    _inherit = "translation.mixin"

    material_id = fields.Many2one('control.procedure.line', 'Control Procedure Material line')

    bom = fields.Char('BOM',translate=True)
    qty = fields.Char('Quantity',translate=True)
    part = fields.Char('Part',translate=True)


class ExtraCareAlerts(models.Model):
    _name = "extra.care.alerts.line"
    _description = "Extra Care Alerts"
    _inherit = "translation.mixin"

    name = fields.Char('Extra Care Alerts',translate=True)


class JigsFixtureTools(models.Model):
    _name = "jigs.fixture.tools.line"
    _description = "Jigs Fixtures & Tools"
    _inherit = "translation.mixin"

    name = fields.Char('Jigs Fixtures & Tools',translate=True)


class ConsumableRequired(models.Model):
    _name = "consumable.required.line"
    _description = "Consumables Required"
    _inherit = "translation.mixin"

    name = fields.Char('Consumables Required',translate=True)


class ControlProductCharacteristics(models.Model):
    _name = "control.product.char.line"
    _description = "Control Product Characteristics"
    _inherit = "translation.mixin"

    product_char_id = fields.Many2one('control.plan.assy', 'Control Procedure Line')

    sequence = fields.Integer(string="Sequence", default=10)
    sl_no = fields.Integer("S.No", compute="_compute_sequence_num")
    vital_assy_parameters = fields.Char('Vital Assy. Parameters',translate=True)
    poka_yoke = fields.Char('Poka-Yoke',translate=True)
    m_product_specification = fields.Char('Product Specification',translate=True)
    m_process_specification = fields.Char('Process Specification',translate=True)
    m_technique = fields.Char('Technique',translate=True)
    i_assembler = fields.Char('Assembler',translate=True)
    i_quality_inspector = fields.Char('Q.A Inspector',translate=True)
    control_method = fields.Char('Control Method',translate=True)
    reaction_action = fields.Char('Reaction Action',translate=True)
    reaction_owner_responsible = fields.Char('Reaction Owner/Responsible',translate=True)

    @api.depends('sequence', 'product_char_id')
    def _compute_sequence_num(self):
        for line in self.mapped('product_char_id'):
            sl_no = 1
            for lines in line.product_char_ids:
                lines.sl_no = sl_no
                sl_no += 1
