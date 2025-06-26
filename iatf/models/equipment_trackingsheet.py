from odoo import models, fields, api, _
from datetime import date, datetime
from odoo.exceptions import ValidationError
import base64
from openpyxl import Workbook
from PIL import Image as PILImage, ImageOps
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
import io
from io import BytesIO


class EquipmentTrackingSheet(models.Model):
    _name = 'equipment.trackingsheet' 
    _description = 'Equipment Tracking Sheet'
    _inherit = ['iatf.sign.off.members', 'revision.history.mixin']

    # customer_id = fields.Many2one('res.partner', string='Customer/Supplier Name')
    # part_name = fields.Many2one('product.template', string='Part/Assembly Name')
    # part_no = fields.Char("Part No.", related="part_name.default_code", store=True)
    date_of_creation = fields.Date("Date of Creation", default=lambda self: fields.Date.today())
    rev_date = fields.Date("Revision Date")
    rev_no = fields.Char("Revision Details")
    total_planned_cost = fields.Float(
        "Total Planned Project Cost for Fixture and Gauges",
        digits=(12, 2),
        compute="_compute_total_planned_cost",
        store=True
    )
    total_actual_planned_cost = fields.Float(
        "Total Actual Planned Project Cost",
        digits=(12, 2)
    )
    diff = fields.Float(
        "Difference",
        digits=(12, 2),
        compute="_compute_difference",
        store=True
    )
    line_ids = fields.One2many('equipment.trackingsheet.line', 'tracking_id', string="Inspection Lines")

    @api.depends('line_ids.total_qty_cost')
    def _compute_total_planned_cost(self):
        for record in self:
            record.total_planned_cost = sum(record.line_ids.mapped('total_qty_cost'))

    @api.depends('total_planned_cost', 'total_actual_planned_cost')
    def _compute_difference(self):
        for record in self:
            record.diff = record.total_actual_planned_cost - record.total_planned_cost

    generate_xls_file = fields.Binary(string="Generated File", readonly=True)

    @api.model_create_multi
    def create(self, vals_list):
        records = super(EquipmentTrackingSheet, self).create(vals_list)

        for record in records:
            pm_records = self.env["process.group"].search(
                [('project_id', '=', record.project_id.id),
                 ('final_status', '=', 'approved')
                 ])

            for operation in pm_records.process_presentation_ids.sorted(key="operation"):
                for equip_element in operation.operation_lines_ids:
                    for i in equip_element.equipment_fixture_ids:
                        category_id = i.category_id and i.category_id.id or False

                        self.env['equipment.trackingsheet.line'].create({
                            'tracking_id': record.id,
                            'element_no': equip_element.element_no,
                            'operation_no': operation.operation,
                            'equipment_id': i.id,
                            'category_id': category_id,
                        })

        return records

    # For EquipmentTrackingSheet class
    def action_update_control_plan(self):
        self.ensure_one()

        # Fetch related Control Plan records based on the project ID
        control_plan_recs = self.env['control.plan'].search([('project_id', '=', self.project_id.id)])
        if not control_plan_recs:
            raise ValidationError(_("No Control Plan found for the current project."))

        for control_plan in control_plan_recs:
            # Retrieve the approved Equipment Tracking Sheet for the project
            equip_sheet = self.env['equipment.trackingsheet'].search([
                ('project_id', '=', self.project_id.id),
                ('final_status', '=', 'approved')
            ])
            if not equip_sheet:
                raise ValidationError(_("No approved Equipment Tracking Sheet found for this project."))

            # Group Equipment lines by operation number
            operation_groups = {}
            for equip_line in equip_sheet.line_ids:
                if equip_line.operation_no not in operation_groups:
                    operation_groups[equip_line.operation_no] = {
                        'description': equip_line.description,
                        'equipment_lines': []
                    }
                operation_groups[equip_line.operation_no]['equipment_lines'].append(equip_line)

            # Iterate through grouped operations and create/update process lines
            for operation_no, operation_data in sorted(operation_groups.items()):
                # First check if a process line already exists for this operation
                existing_process_line = self.env['control.plan.process'].search([
                    ('process_id', '=', control_plan.id),
                    ('process_step', '=', operation_no)
                ], limit=1)

                if existing_process_line:
                    # Process line exists, use it
                    process_line = existing_process_line
                else:
                    # Create a new process line in Control Plan
                    process_line = self.env['control.plan.process'].create({
                        'process_id': control_plan.id,
                        'process_step': operation_no,  # Operation number
                        'process_name': operation_data['description'],  # Operation description
                    })

                # Create characteristic lines for each piece of equipment
                for equip_line in operation_data['equipment_lines']:
                    # Check if this equipment is already added to this process
                    existing_char_line = self.env['control.chara.line'].search([
                        ('chara_id', '=', process_line.id),
                        ('char_no', '=', equip_line.element_no)
                    ], limit=1)

                    if existing_char_line:
                        # Append new equipment(s) to the existing characteristic line
                        existing_char_line.mc_jig_tool = [
                            (6, 0, list(set(existing_char_line.mc_jig_tool.ids + [equip_line.equipment_id.id])))]
                        existing_char_line.mc_jig_tool_num = ', '.join(
                            filter(None, [existing_char_line.mc_jig_tool_num, equip_line.equipment_no])
                        )
                    else:
                        self.env['control.chara.line'].create({
                            'chara_id': process_line.id,
                            'char_no': equip_line.element_no,
                            'mc_jig_tool': [(6, 0, [equip_line.equipment_id.id])],
                            'mc_jig_tool_num': equip_line.equipment_no,
                        })

        # Notify user upon successful completion
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _("Success"),
                'message': _("Control Plan updated successfully with equipment data!"),
                'sticky': False,
                'type': 'success',
            }
        }
    def generate_xls_report(self):
        # Create workbook and worksheet
        output = BytesIO()
        wb = Workbook()
        ws = wb.active

        # Define styles
        border = Border(top=Side(style='thin'), left=Side(style='thin'),
                        right=Side(style='thin'), bottom=Side(style='thin'))
        font_header = Font(name='Times New Roman', bold=True)
        title_font = Font(size=20, bold=True)
        align_center = Alignment(vertical='center', horizontal='center', wrapText=True)

        # Define fill colors
        fill = PatternFill(start_color="e7e7e7", end_color="e7e7e7", fill_type="solid")  # Grey
        orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

        if self.env.user.company_id and self.env.user.company_id.logo:

            max_width, max_height = 80, 160  # Set max dimensions
            image_data = base64.b64decode(self.env.user.company_id.logo)

            # Open image with PIL
            image = PILImage.open(io.BytesIO(image_data))
            width, height = image.size
            aspect_ratio = width / height

            # Resize with aspect ratio
            if width > max_width:
                width = max_width
                height = int(width / aspect_ratio)
            if height > max_height:
                height = max_height
                width = int(height * aspect_ratio)

            # Resize and add padding
            resized_image = image.resize((width, height), PILImage.LANCZOS)
            padding_top, padding_left = 0, 25
            resized_image = ImageOps.expand(resized_image, border=(padding_left, padding_top, 0, 0), fill='white')

            img_bytes = io.BytesIO()
            resized_image.save(img_bytes, format='PNG')
            img_bytes.seek(0)
            logo_image = Image(img_bytes)

            ws.add_image(logo_image, 'A2')

        # Headers dictionary
        headers = {
            'A1': 'Equipment / Fixture Tracking sheet',
            'B2': 'Type(SafeLaunch/ Prototype/ PreLaunch/ Production)',
            'B3': 'Customer/Supplier Name',
            'E2': 'Part / Assembly Name',
            'E3': 'Part / Assembly No.',
            'H2': 'Revision Date',
            'H3': 'Revision Details',
            'K2': 'Total Planned Project Cost for Fixture and Gauges',
            'K3': 'Total Actual Planned Project Cost',
            'O2': 'Difference',
            'O3': 'Date of Creation',

            'A4': 'Category of Equipment / Fixture',
            'B4': 'Equipment / Fixture Name',
            'C4': 'Equipment / Fixture No.',
            'D4': 'Operation Number',
            'E4': 'Description',
            'F4': 'Availability (Yes or No)',
            'G4': 'Standard Delivery Time',
            'H4': 'To be Ordered by',
            'I4': 'Requirement Raise Date',
            'J4': 'QTY',
            'K4': 'Cost Per Unit',
            'L4': 'Total Quantity Unit',
            'M4': 'Expected Date',
            'N4': 'Indent Date',
            'O4': 'PO Date',
            'P4': 'Actual Received date',
            'Q4': 'Drawing Attachments',
        }
        ws.row_dimensions[1].height = 30
        ws.row_dimensions[2].height = 30
        ws.row_dimensions[3].height = 30
        ws['A1'].font = title_font
        # Apply font and border styles
        for cell, value in headers.items():
            ws[cell] = value
            ws[cell].font = font_header
            ws[cell].alignment = align_center
            ws[cell].fill = fill

        # Merge cells
        merge_cells = [
            'A2:A3', 'A1:Q1', 'C2:D2', 'C3:D3', 'F2:G2', 'F3:G3', 'I2:J2', 'I3:J3', 'K2:L2', 'K3:L3', 'M2:N2', 'M3:N3',
        ]
        for cell_range in merge_cells:
            ws.merge_cells(cell_range)

        # Set column widths
        col_widths = {
            'A': 20, 'B': 30, 'C': 20, 'D': 15, 'E': 20,
            'F': 15, 'G': 15, 'H': 15, 'I': 15, 'J': 15,
            'K': 15, 'L': 15, 'M': 15, 'N': 15, 'O': 15, 'P': 18,'Q':20
        }
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width

        # Populate data for the inspection record
        for rec in self:
            ws['C2'] = rec.doc_type if rec.doc_type else ''
            ws['C3'] = rec.partner_id.name if rec.partner_id else ''
            ws['F2'] = rec.part_name if rec.part_name else ''
            ws['F3'] = rec.part_number if rec.part_number else ''
            ws['I2'] = rec.rev_date if rec.rev_date else ''
            ws['I3'] = rec.rev_no if rec.rev_no else ''
            ws['M2'] = rec.total_planned_cost if rec.total_planned_cost else ''
            ws['M3'] = rec.total_actual_planned_cost if rec.total_actual_planned_cost else ''
            ws['P2'] = rec.diff if rec.diff else ''
            ws['P3'] = rec.date_of_creation if rec.date_of_creation else ''

            row = 5
            for ele in rec.line_ids:
                ws[f'A{row}'] = ele.category_id.name if ele.category_id.name else ''
                ws[f'B{row}'] = ele.equipment_id.name if ele.equipment_id.name else ''
                ws[f'C{row}'] = ele.equipment_no if ele.equipment_no else ''
                ws[f'D{row}'] = ele.operation_no if ele.operation_no else ''
                ws[f'E{row}'] = ele.description if ele.description else ''
                ws[f'F{row}'] = ele.availability if ele.availability else ''

                ws[f'G{row}'] = ele.standard_delivery_time if ele.standard_delivery_time else ''
                ws[f'H{row}'] = ele.to_be_ordered_by if ele.to_be_ordered_by else ''
                ws[f'I{row}'] = ele.requirement_raise_date if ele.requirement_raise_date else ''
                ws[f'J{row}'] = ele.qty if ele.qty else ''
                ws[f'K{row}'] = ele.cost_per_unit if ele.cost_per_unit else ''
                ws[f'L{row}'] = ele.total_qty_cost if ele.total_qty_cost else ''

                ws[f'M{row}'] = ele.expected_date if ele.expected_date else ''
                ws[f'N{row}'] = ele.indent_date if ele.indent_date else ''
                ws[f'O{row}'] = ele.po_date if ele.po_date else ''
                ws[f'P{row}'] = ele.actual_received_date if ele.actual_received_date else ''
                ws[f'Q{row}'] = ele.drawing_attachments if ele.drawing_attachments else ''

                row += 1

        cur_row = 30
        if cur_row < row:
            cur_row = row
        for rows in ws.iter_rows(min_row=1, max_row=cur_row, min_col=1, max_col=17):
            for cell in rows:
                cell.alignment = align_center
                cell.border = border

        # region SignOff Members Footer
        sign_row = cur_row
        ws.merge_cells(f'A{cur_row}:Q{cur_row}')

        cur_row += 1
        ws[f'A{cur_row}'] = 'Prepared By'
        ws[f'A{cur_row}'].font = font_header
        ws.merge_cells(f'A{cur_row}:B{cur_row}')
        ws.merge_cells(f'C{cur_row}:E{cur_row}')

        ws[f'F{cur_row}'] = 'Prepared Date'
        ws[f'F{cur_row}'].font = font_header
        ws.merge_cells(f'F{cur_row}:G{cur_row}')
        ws.merge_cells(f'H{cur_row}:L{cur_row}')

        for rec in self:
            ws[f'C{cur_row}'] = rec.create_uid.name if rec.create_uid else ''
            ws[f'H{cur_row}'] = rec.create_date if rec.create_date else ''

        ws.row_dimensions[cur_row].height = 18
        cur_row += 1
        ws.merge_cells(f'A{cur_row}:L{cur_row}')
        cur_row += 1
        ws.merge_cells(f'A{cur_row}:L{cur_row}')
        ws[f'A{cur_row}'] = 'Sign OFF'
        ws[f'A{cur_row}'].font = Font(size=18, bold=True)
        ws.row_dimensions[cur_row].height = 25
        cur_row += 1
        ws.merge_cells(f'A{cur_row}:L{cur_row}')
        cur_row += 1
        for cell in ws[cur_row]:
            ws[f'A{cur_row}'] = 'Member'
            ws[f'D{cur_row}'] = 'Department'
            ws[f'F{cur_row}'] = 'Status'
            ws[f'H{cur_row}'] = 'Date'
            ws[f'J{cur_row}'] = 'Comments'
            ws.row_dimensions[cur_row].height = 25
            ws.merge_cells(f'A{cur_row}:C{cur_row}')
            ws.merge_cells(f'D{cur_row}:E{cur_row}')
            ws.merge_cells(f'F{cur_row}:G{cur_row}')
            ws.merge_cells(f'H{cur_row}:I{cur_row}')
            ws.merge_cells(f'J{cur_row}:L{cur_row}')
            cell.font = Font(size=12, bold=True, color='ffffff')
            cell.fill = PatternFill(start_color='0070C0', end_color='0070C0', fill_type="solid")

        # Listing Managers Id
        for rec in self:
            for record in rec.iatf_members_ids:
                name_rec = record.approver_id.name
                dept_rec = record.department_id.name if record.department_id.name else ''
                status_rec = record.approval_status.capitalize()
                date_rec = record.date_approved_rejected if record.date_approved_rejected else ''
                comment_rec = record.comment if record.comment else ''

                ws[f'A{cur_row + 1}'] = name_rec
                ws[f'D{cur_row + 1}'] = dept_rec
                ws[f'F{cur_row + 1}'] = status_rec
                ws[f'H{cur_row + 1}'] = date_rec
                ws[f'J{cur_row + 1}'] = comment_rec
                ws.merge_cells(f'A{cur_row + 1}:C{cur_row + 1}')
                ws.merge_cells(f'D{cur_row + 1}:E{cur_row + 1}')
                ws.merge_cells(f'F{cur_row + 1}:G{cur_row + 1}')
                ws.merge_cells(f'H{cur_row + 1}:I{cur_row + 1}')
                ws.merge_cells(f'J{cur_row + 1}:L{cur_row + 1}')

                status_dict = {'Approved': ['CFFFC3', '000000'], 'Rejected': ['FFCDCD', '000000'],
                               'Revision': ['AFEFFF', '000000'], 'Pending': ['FDFFCD', '000000']}

                for state, color in status_dict.items():
                    if state == status_rec:
                        ws[f'F{cur_row + 1}'].fill = PatternFill(start_color=color[0], end_color=color[0],
                                                                 fill_type="solid")
                        ws[f'F{cur_row + 1}'].font = Font(size=12, bold=False, color=color[1])

                cur_row += 1

            # ws.merge_cells(f'C{cur_row + 1}:D{cur_row + 1}')
            # ws.merge_cells(f'E{cur_row + 1}:F{cur_row + 1}')
            ws.merge_cells(f'A{cur_row + 1}:Q{cur_row + 1}')
            ws.merge_cells(f'M{sign_row + 1}:Q{cur_row}')

            for row_no in ws.iter_rows(min_row=sign_row, max_row=cur_row + 1, min_col=1, max_col=17):
                for cell in row_no:
                    cell.border = border
                    cell.alignment = align_center
    # endregion

        # Save workbook to BytesIO
        wb.save(output)
        output.seek(0)

        # Create attachment
        attachment = self.env['ir.attachment'].create({
            'name': 'Equipment_Tracking_Sheet.xlsx',
            'type': 'binary',
            'datas': base64.b64encode(output.getvalue()),
            'res_model': 'in.process.inspection.report',
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        })

            # Return download link
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }


class EquipmentTrackingSheetLine(models.Model):
    _name = 'equipment.trackingsheet.line'
    _description = 'Equipment Tracking Sheet Line'
    # _inherit = "translation.mixin"

    tracking_id = fields.Many2one('equipment.trackingsheet', string='Equipment Tracking Sheet',  index=True, copy=False, readonly=True,
                                  ondelete="cascade")
    element_no = fields.Char(string="Element No")
    category_id = fields.Many2one('maintenance.equipment.category', string="Category of Equipment")
    equipment_id = fields.Many2one('maintenance.equipment',string="Equipment Name")
    # category_of_equipment = fields.Char(string="Category of Equipment")
    # equipment_name = fields.Char(string="Equipment Name")
    equipment_no = fields.Char(related='equipment_id.serial_no', string="Equipment No.")
    description = fields.Text(string="Description")
    availability = fields.Selection([
        ('yes', 'Yes'),
        ('no', 'No')
    ], string="Availability (Yes or No)", default='yes')
    operation_no=fields.Char(string="Operation No.",store=True)

    standard_delivery_time = fields.Date(string="Standard Delivery Time")
    to_be_ordered_by = fields.Date(string="To be Ordered by")
    requirement_raise_date = fields.Date(string="Requirement Raise Date")
    qty = fields.Integer(string="QTY", default=1)
    cost_per_unit = fields.Float(string="Cost per Unit", digits=(12, 2))
    total_qty_cost = fields.Float(string="Total Quantity Cost", digits=(12, 2), compute="_compute_total_qty_cost",
                                  store=True)

    expected_date = fields.Date(string="Expected Date")
    indent_date = fields.Date(string="Indent Date")
    po_date = fields.Date(string="PO Date")
    actual_received_date = fields.Date(string="Actual Received Date")
    drawing_attachments = fields.Binary(string="Drawing Attachments")

    @api.onchange('category_id')
    def _onchange_category_id(self):
        self.equipment_id = False

    @api.depends('qty', 'cost_per_unit')
    def _compute_total_qty_cost(self):
        for record in self:
            record.total_qty_cost = record.qty * record.cost_per_unit

# class EquipmentCategory(models.Model):
#     _name = 'equipment.category.sheet'
#     _description = 'Equipment Category'
#
#     name = fields.Char(string="Category Name", required=True)
#
#
# class EquipmentName(models.Model):
#     _name = 'equipment.name.sheet'
#     _description = 'Equipment Name'
#
#     name = fields.Char(string="Equipment Name", required=True)
#     category_id = fields.Many2one('equipment.category.sheet', string="Category")


