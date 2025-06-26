import base64
import io
from io import BytesIO

from PIL import Image as PILImage
from PIL import ImageOps
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill

from odoo import api, fields, models


class PpapChecklist(models.Model):
    _name = 'ppap.checklist'
    _inherit = ['iatf.sign.off.members']

    partner_id = fields.Many2one('res.partner', 'Customer')
    product_id = fields.Many2one('product.template', 'Part Name')
    default_code = fields.Char('Part Number', related='product_id.default_code')
    rev_no = fields.Integer('Revision No.')
    date = fields.Date('Date')
    ppap_checklist_line_ids = fields.One2many(
        comodel_name='ppap.checklist.line',
        inverse_name='ppap_id',
        string="PPAP Checklist Line"
    )

    generate_xls_file = fields.Binary(string="Generated file")

    def action_ppap_checklist(self):
        output = BytesIO()
        wb = Workbook()
        ws = wb.active

        if self.env.user.company_id.logo:
            max_width = 500  # Set your desired maximum width
            max_height = 100  # Set your desired maximum height
            image_data = base64.b64decode(self.env.user.company_id.logo)

            # Open the image using PIL
            image = PILImage.open(io.BytesIO(image_data))
            width, height = image.size
            aspect_ratio = width / height

            if width > max_width:
                width = max_width
                height = int(width / aspect_ratio)

            if height > max_height:
                height = max_height
                width = int(height * aspect_ratio)

            # Resize the image using PIL
            # Add space on the top and left side of the image
            padding_top = 10  # Adjust as needed
            padding_left = 10  # Adjust as needed

            resized_image = image.resize((width, height), PILImage.Resampling.LANCZOS)
            ImageOps.expand(resized_image, border=(padding_left, padding_top, 0, 0), fill='rgba(0,0,0,0)')
            img_bytes = io.BytesIO()
            resized_image.save(img_bytes, format='PNG')
            img_bytes.seek(0)
            logo_image = Image(img_bytes)
            # logo_image = Image(self.env.user.company_id.logo)
            ws.add_image(logo_image, 'A1')

        data = {
            'A1': 'PPAP CHECKLIST',
            'A2': 'Customer:',
            'A3': 'Part Name:',
            'D2': 'Part Number:',
            'D3': 'Revision No.:',
            'G3': 'Date:',
            'A4': 'Steps',
            'B4': 'Description',
            'C4': 'Requirement',
            'D4': 'Status',
            'E4': 'Comments',
            # #             'A5': 1,
            # 'B5': 'Design Records',
            # 'C5': 'S',
            # #             'A6': 2,
            # 'B6': 'Authorized Engineering Changes (Tech. Review)',
            # 'C6': 'S',
            # #             'A7': 3,
            # 'B7': 'Customer Engineering Approval',
            # 'C7': 'S',
            # #             'A8': 4,
            # 'B8': 'Design F.M.E.A.',
            # 'C8': 'S',
            # #             'A9': 5,
            # 'B9': 'Process Flow Charts',
            # 'C9': 'S',
            # #             'A10': 6,
            # 'B10': 'Process F.M.E.A.',
            # 'C10': 'S',
            # #             'A11': 7,
            # 'B11': 'Control Plans',
            # 'C11': 'S',
            # #             'A12': 8,
            # 'B12': 'Measurement System Analysis Studies',
            # 'C12': 'S',
            # #             'A13': 9,
            # 'B13': 'Dimensional Results',
            # 'C13': 'S',
            # #             'A14': 10,
            # 'B14': 'Record of Material / Performance Test Result',
            # 'C14': 'S',
            # #             'A15': 11,
            # 'B15': 'Initial Process Studies',
            # 'C15': 'S',
            # #             'A16': 12,
            # 'B16': 'Qualified Laboratory Documentation',
            # 'C16': 'S',
            # #             'A17': 13,
            # 'B17': 'Appearance Report',
            # 'C17': 'S',
            # #             'A18': 14,
            # 'B18': 'Sample Production Parts',
            # 'C18': 'S',
            # #             'A19': 15,
            # 'B19': 'Master Samples',
            # 'C19': 'R',
            # #             'A20': 16,
            # 'B20': 'Checking Aids',
            # 'C20': 'R',
            # #             'A22': 17,
            # 'B21': 'List of gauges, instrument & fixture',
            # 'C21': 'R',
            # #             'A23': 18,
            # 'B22': 'Customer - Specific - Requirements.',
            # 'C22': 'S',
            # #             'A24': 19,
            # 'B23': 'Part Submission Warrant',
            # 'B24': 'Key Characteristic Matrix',
            # 'C23': 'S',
            # 'C24': 'R',
            # 'A25': 'Prepared By:',
            # 'C25': 'Approved By:',
            # 'A26': 'FM-80029 / 0 - 0 Rev. No.:  00 DATE : 12.04.2017',
            # 'A27': 'Legends :',
            # 'B27': 'R- Retain',
            # 'B28': 'S- Submit to Customer',
            # 'B29': 'N/A- Not Applicable',

        }

        for cell, value in data.items():
            ws[cell] = value

        thin = Side(border_style='thin', color='000000')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        font_header = Font(name='Arial', size=11, bold=True)
        font_normal = Font(name='Arial', size=10, bold=False)
        align_left = Alignment(vertical='center', horizontal='left', wrapText=True)
        align_right = Alignment(vertical='center', horizontal='right', wrapText=True)
        align_center = Alignment(vertical='center', horizontal='center', wrapText=True)

        max_col = 8
        max_row = 25
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.border = border
                cell.alignment = align_center
                cell.font = font_header

        ws['A1'].alignment = Alignment(vertical='center', horizontal='center')
        ws['A1'].font = Font(name='Arial', size=22, bold=True)

        ws.merge_cells('A1:H1')
        ws.merge_cells('B2:C2')
        ws.merge_cells('E2:H2')
        ws.merge_cells('B3:C3')
        ws.merge_cells('E3:F3')

        for row in range(4, 25):
            ws.merge_cells(f'E{row}:H{row}')

        ws.column_dimensions['A'].width = 28
        ws.column_dimensions['B'].width = 44
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 15

        ws.row_dimensions[1].height = 75
        ws.row_dimensions[2].height = 21
        ws.row_dimensions[3].height = 21
        ws.row_dimensions[4].height = 21

        for i in range(5, 25):
            ws.row_dimensions[i].height = 25

        # Data Filling
        cur_row = 5
        max_row = 25
        for rec in self:
            ws['B2'] = rec.partner_id.name if rec.partner_id.name else ''
            ws['B3'] = rec.product_id.name if rec.product_id.name else ''
            ws['E2'] = rec.default_code if rec.default_code else ''
            ws['E3'] = rec.rev_no if rec.rev_no else ''
            ws['H3'] = rec.date if rec.date else ''

            for i in rec.ppap_checklist_line_ids:
                ws[f'A{cur_row}'] = i.sl_no if i.sl_no else ''
                ws[f'B{cur_row}'] = i.desc.name if i.desc.name else ''
                ws[f'C{cur_row}'] = i.requirement if i.requirement else ''
                ws[f'D{cur_row}'] = i.status if i.status else ''
                ws[f'E{cur_row}'] = i.comment if i.comment else ''
                ws.row_dimensions[cur_row].height = 25
                for row_no in ws.iter_rows(min_row=cur_row, max_row=cur_row, min_col=1, max_col=8):
                    for cell in row_no:
                        cell.border = border
                        cell.font = font_normal
                        cell.alignment = align_center
                        ws.merge_cells(f'E{cur_row}:H{cur_row}')


                cur_row += 1

        # Legends section
        if cur_row < max_row:
            cur_row = max_row
        ws.merge_cells(f'A{cur_row}:H{cur_row}')
        cur_row += 1
        ws[f'A{cur_row}'] = "Legends :"
        ws[f'B{cur_row}'] = "R- Retain"
        ws[f'C{cur_row}'] = "S- Submit to Customer"
        ws[f'E{cur_row}'] = "N/A- Not Applicable"
        ws.merge_cells(f'C{cur_row}:D{cur_row}')
        ws.merge_cells(f'E{cur_row}:H{cur_row}')
        ws.row_dimensions[cur_row].height = 25
        for cell in ws[cur_row]:
            cell.alignment = align_center
            cell.font = font_header
        cur_row += 1

        # region SignOff Members Footer
        ws.merge_cells(f'A{cur_row}:H{cur_row}')
        cur_row += 1
        sign_row = cur_row
        ws[f'A{cur_row}'] = 'Prepared By'
        ws[f'A{cur_row}'].font = font_header

        ws[f'C{cur_row}'] = 'Prepared Date'
        ws[f'C{cur_row}'].font = font_header
        ws.merge_cells(f'C{cur_row}:D{cur_row}')
        ws.merge_cells(f'E{cur_row}:H{cur_row}')

        for rec in self:
            ws[f'B{cur_row}'] = rec.create_uid.name if rec.create_uid else ''
            ws[f'E{cur_row}'] = rec.create_date if rec.create_date else ''

        ws.row_dimensions[cur_row].height = 18
        cur_row += 1
        ws.merge_cells(f'A{cur_row}:H{cur_row}')
        cur_row += 1
        ws.merge_cells(f'A{cur_row}:H{cur_row}')
        ws[f'A{cur_row}'] = 'Sign OFF'
        ws[f'A{cur_row}'].font = Font(size=18, bold=True)
        ws.row_dimensions[cur_row].height = 25
        cur_row += 1
        ws.merge_cells(f'A{cur_row}:H{cur_row}')
        cur_row += 1
        for cell in ws[cur_row]:
            ws[f'A{cur_row}'] = 'Member'
            ws[f'B{cur_row}'] = 'Department'
            ws[f'C{cur_row}'] = 'Status'
            ws[f'D{cur_row}'] = 'Date'
            ws[f'E{cur_row}'] = 'Comments'
            ws.row_dimensions[cur_row].height = 25
            ws.merge_cells(f'E{cur_row}:H{cur_row}')
            cell.font = Font(size=12, bold=True, color='ffffff')
            cell.fill = PatternFill(start_color='0070C0', end_color='0070C0', fill_type="solid")

        for row_no in ws.iter_rows(min_row=1, max_row=cur_row + 1, min_col=1, max_col=8):
            for cell in row_no:
                cell.border = border
                cell.alignment = align_center

        # Listing Managers Id
        for rec in self:
            for record in rec.iatf_members_ids:
                name_rec = record.approver_id.name
                dept_rec = record.department_id.name if record.department_id.name else ''
                status_rec = record.approval_status.capitalize()
                date_rec = record.date_approved_rejected if record.date_approved_rejected else ''
                comment_rec = record.comment if record.comment else ''

                ws[f'A{cur_row + 1}'] = name_rec
                ws[f'B{cur_row + 1}'] = dept_rec
                ws[f'C{cur_row + 1}'] = status_rec
                ws[f'D{cur_row + 1}'] = date_rec
                ws[f'E{cur_row + 1}'] = comment_rec
                ws.merge_cells(f'E{cur_row + 1}:H{cur_row + 1}')

                status_dict = {'Approved': ['CFFFC3', '000000'], 'Rejected': ['FFCDCD', '000000'],
                               'Revision': ['AFEFFF', '000000'], 'Pending': ['FDFFCD', '000000']}

                for state, color in status_dict.items():
                    if state == status_rec:
                        ws[f'C{cur_row + 1}'].fill = PatternFill(start_color=color[0], end_color=color[0],
                                                                 fill_type="solid")
                        ws[f'C{cur_row + 1}'].font = Font(size=12, bold=False, color=color[1])

                cur_row += 1

            ws.merge_cells(f'A{cur_row + 1}:H{cur_row + 1}')

            for row_no in ws.iter_rows(min_row=sign_row + 1, max_row=cur_row + 1, min_col=1, max_col=8):
                for cell in row_no:
                    cell.border = border
                    cell.alignment = align_center
            # endregion

        wb.save(output)
        output.seek(0)
        self.generate_xls_file = base64.b64encode(output.getvalue()).decode('utf-8')

        return {
            "type": "ir.actions.act_url",
            "target": "self",
            "url": "/web/content?model=ppap.checklist&download=true&field=generate_xls_file&filename={filename}.xlsx&id={pid}".format(
                filename="PPAP Checklist", pid=self[0].id),
        }


class PpapChecklistLine(models.Model):
    _name = 'ppap.checklist.line'
    _description = 'PPAP Checklist Line'
    _inherit = "translation.mixin"

    ppap_id = fields.Many2one(
        comodel_name='ppap.checklist',
        string='PPAP Checklist',
        required=True, ondelete='cascade', index=True, copy=False
    )
    sequence = fields.Integer(string="Sequence", default=10)
    sl_no = fields.Integer("S.No", compute="_compute_sequence_number")
    desc = fields.Many2one('description.name', 'Description')
    requirement = fields.Selection([('R', 'retain'), ('S', 'submit for customer'), ('N/A', 'not applicable')],
                                   default='R')
    status = fields.Char('Status',translate=True)
    comment = fields.Char('Comments',translate=True)

    @api.depends('sequence', 'ppap_id')
    def _compute_sequence_number(self):
        for line in self.mapped('ppap_id'):
            sl_no = 1
            for lines in line.ppap_checklist_line_ids:
                lines.sl_no = sl_no
                sl_no += 1


class DescriptionName(models.Model):
    _name = 'description.name'
    _inherit = "translation.mixin"

    name = fields.Char('Name',translate=True)
