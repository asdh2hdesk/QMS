import base64
import io
from io import BytesIO

from PIL import Image as PILImage
from PIL import ImageOps
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

from odoo import fields, models, api


class ChangeRequest(models.Model):
    _name = 'change.request'
    _inherit = 'iatf.sign.off.members'

    date = fields.Datetime('ECR Date', default=lambda self: fields.Datetime.now())
    emp_id = fields.Many2one('hr.employee', 'Proposer Name')
    product_id = fields.Many2one('product.template', 'Part Name')
    part_no = fields.Char('Part No.', related='product_id.default_code')
    customer_id = fields.Many2one('res.partner', 'Customer Name')
    dept_id = fields.Many2one('hr.department', 'Department', related='emp_id.department_id')
    processer_name = fields.Many2one('hr.employee', 'Process Name')
    drg_no = fields.Char('Drg. No.', related='product_id.drg_no')
    control_paln = fields.Char('Control Plan No.')
    description = fields.Html('Change Description')
    material_descrp = fields.Char('Existing Material Disposition')
    Proposed = fields.Char('Proposed')
    scraped = fields.Boolean('Scraped')
    userd_in = fields.Boolean('To be used as it is')
    specification = fields.Boolean('To be converted as per new specification:')
    specific_change = fields.Boolean('Design / Specification Change')
    quality_imp = fields.Boolean('Improvement In Performance / Quality')
    cost_reduce = fields.Boolean('Cost Reduction')
    standardisation = fields.Boolean('Standardisation')
    sub_suppiler = fields.Boolean('Sub- Supplier Requirement')
    improve_machine = fields.Boolean('Improvement In Machining')
    customer_requirement = fields.Boolean('Customer Requirement')
    remark = fields.Html('Remarks')
    first_impact = fields.Boolean('1.Can engineering performance specifications be met as written ?')
    secound_impact = fields.Boolean('2.Is there adequate capacity to produce product ?')
    third_impact = fields.Boolean("3.Can product be manufactured with Cpk's that meet requirements ?")
    forth_impact = fields.Boolean('4.Is statistical process control required on product ?')
    fifth_impact = fields.Boolean('5.Does the design allow the use of efficient material handling techniques ?')
    sixth_impact = fields.Boolean('6.Costs for tooling ? ')
    seven_impact = fields.Boolean('7.Costs for capital equipment ?')
    eight_impact = fields.Boolean('8.Alternative manufacturing methods ?')

    feasible = fields.Boolean('Feasible(Some changes recommended)')
    not_feasible = fields.Boolean(
        'Not Feasible (Design revision required to produce product within the specified requirements.)')
    ppap_approval_required = fields.Boolean('PPAP Approval Required(Yes/No)')
    date_of_ppap_submission = fields.Date('Date of PPAP Submission')
    date_of_approval_received = fields.Date('Date of Approval Submission')
    date_of_pilot_lot_submission = fields.Date('Date of Pilot Lot Submission')
    date_of_regular_submission = fields.Date('Date of Regular Production')

    change_required_ids = fields.One2many(
        comodel_name='change.required.line',
        inverse_name='change_required_id',
        string='Change Required'
    )

    generate_xls_file = fields.Binary('Generate XLSX File')

    # state = fields.Selection([
    #                         ('draft', 'Draft'),
    #                         ('confirm', 'Confirmed'),
    #                         ], string='Status', default='draft')
    document_name = fields.Char(string='Document #')

    # approve_department_id = fields.Many2one('hr.department', string='Departments Approvals')
    # # document_pro_id = fields.Many2one("xf.doc.approval.document.package", string="Document #")
    # approve_department_ids = fields.Many2many('hr.department', 'table9', string='Departments Approvals')
    # approve_by_department_ids = fields.Many2many('hr.department', string='Departments Approved By')
    # manager_id = fields.Many2one('hr.employee', string='Approval Manager')
    # manager_ids = fields.Many2many('hr.employee', 'table10', string='Approval Managers')
    # approvaed_manager_ids = fields.Many2many('hr.employee', string='Managers Approved By')
    #
    # is_hr_approved = fields.Boolean()
    # is_design_approved = fields.Boolean()
    # is_eng_approved = fields.Boolean()
    # is_manu_approved = fields.Boolean()
    # is_sales_approved = fields.Boolean()
    # is_qc_approved = fields.Boolean()
    # is_maintenance_approved = fields.Boolean()
    # is_marketing_approved = fields.Boolean()
    # is_pm_approved = fields.Boolean()
    # is_management_approved = fields.Boolean()
    # is_approved_approved = fields.Boolean()
    #
    # hr_check_department = fields.Boolean(compute='_check_department')
    # des_check_department = fields.Boolean(compute='_check_department')
    # eng_check_department = fields.Boolean(compute='_check_department')
    # manu_check_department = fields.Boolean(compute='_check_department')
    # sale_check_department = fields.Boolean(compute='_check_department')
    # qc_check_department = fields.Boolean(compute='_check_department')
    # mainte_check_department = fields.Boolean(compute='_check_department')
    # marketing_check_department = fields.Boolean(compute='_check_department')
    # pm_check_department = fields.Boolean(compute='_check_department')
    # managment_check_department = fields.Boolean(compute='_check_department')
    #
    # link = fields.Char()
    #
    # def actipn_confirmed(self):
    #     self.sent_for_approval()
    #     self.state = 'confirm'

    # def sent_for_approval(self):
    #     mail_template = self.env.ref('iatf.mom_document_approval_mail_template')
    #     model_id = self.env['ir.model'].sudo().search([('model', '=', self._name)])
    #     mail_template.write({'model_id' : model_id.id})
    #
    #     web_base_url = self.env['ir.config_parameter'].sudo().get_param('web.base.url')
    #     action_id = self.env.ref('iatf.action_change_request_view', raise_if_not_found=False)
    #     link = """{}/web#id={}&view_type=form&model=self._name&action={}""".format(web_base_url,self.id,action_id.id)
    #     self.link = link
    #
    #     user_ids = self.env['hr.employee'].sudo().search([('department_id', 'in', self.approve_department_ids.ids)]).mapped('user_id')
    #     all_admin = ''
    #     for user in user_ids:
    #         if all_admin:
    #             all_admin += ',' + str(user.login)
    #         else:
    #             all_admin = str(user.login)
    #     mail_template.write({
    #             'email_to': all_admin,
    #     })
    #     mail_template.send_mail(self.id, force_send=True)
    #
    # def _check_department(self):
    #     for rec in self:
    #         if self.env.user.employee_id.department_id.id in rec.approve_department_ids.ids and self.env.user.employee_id.department_id.department_type == 'hr':
    #             if rec.state != 'draft':
    #                 rec.hr_check_department = True
    #             else:
    #                 rec.hr_check_department = False
    #         else:
    #             rec.hr_check_department = False
    #         if self.env.user.employee_id.department_id.id in rec.approve_department_ids.ids and self.env.user.employee_id.department_id.department_type == 'design':
    #             if rec.state != 'draft':
    #                 rec.des_check_department = True
    #             else:
    #                 rec.des_check_department = False
    #         else:
    #             rec.des_check_department = False
    #         if self.env.user.employee_id.department_id.id in rec.approve_department_ids.ids and self.env.user.employee_id.department_id.department_type == 'eng':
    #             if rec.state != 'draft':
    #                 rec.eng_check_department = True
    #             else:
    #                 rec.eng_check_department = False
    #         else:
    #             rec.eng_check_department = False
    #         if self.env.user.employee_id.department_id.id in rec.approve_department_ids.ids and self.env.user.employee_id.department_id.department_type == 'manu':
    #             if rec.state != 'draft':
    #                 rec.manu_check_department = True
    #             else:
    #                 rec.manu_check_department = False
    #         else:
    #             rec.manu_check_department = False
    #         if self.env.user.employee_id.department_id.id in rec.approve_department_ids.ids and self.env.user.employee_id.department_id.department_type == 'sales':
    #             if rec.state != 'draft':
    #                 rec.sale_check_department = True
    #             else:
    #                 rec.sale_check_department = False
    #         else:
    #             rec.sale_check_department = False
    #         if self.env.user.employee_id.department_id.id in rec.approve_department_ids.ids and self.env.user.employee_id.department_id.department_type == 'qc':
    #             if rec.state != 'draft':
    #                 rec.qc_check_department = True
    #             else:
    #                 rec.qc_check_department = False
    #         else:
    #             rec.qc_check_department = False
    #         if self.env.user.employee_id.department_id.id in rec.approve_department_ids.ids and self.env.user.employee_id.department_id.department_type == 'maintenance':
    #             if rec.state != 'draft':
    #                 rec.mainte_check_department = True
    #             else:
    #                 rec.mainte_check_department = False
    #         else:
    #             rec.mainte_check_department = False
    #         if self.env.user.employee_id.department_id.id in rec.approve_department_ids.ids and self.env.user.employee_id.department_id.department_type == 'marketing':
    #             if rec.state != 'draft':
    #                 rec.marketing_check_department = True
    #             else:
    #                 rec.marketing_check_department = False
    #         else:
    #             rec.marketing_check_department = False
    #         if self.env.user.employee_id.department_id.id in rec.approve_department_ids.ids and self.env.user.employee_id.department_id.department_type == 'pm':
    #             if rec.state != 'draft':
    #                 rec.pm_check_department = True
    #             else:
    #                 rec.pm_check_department = False
    #         else:
    #             rec.pm_check_department = False
    #         if self.env.user.employee_id.department_id.id in rec.approve_department_ids.ids and self.env.user.employee_id.department_id.department_type == 'management':
    #             if rec.state != 'draft':
    #                 rec.managment_check_department = True
    #             else:
    #                 rec.managment_check_department = False
    #         else:
    #             rec.managment_check_department = False
    #
    #
    # hr_approval_date = fields.Datetime('Hr Approval Date')
    # design_approval_date = fields.Datetime('Design Approval Date')
    # eng_approval_date = fields.Datetime('Eng. Approval Date')
    # manu_approval_date = fields.Datetime('Production Approval Date')
    # sale_approval_date = fields.Datetime('Sales Approval Date')
    # qc_approval_date = fields.Datetime('QC Approval Date')
    # main_approval_date = fields.Datetime('Maintenance Approval Date')
    # mark_approval_date = fields.Datetime('Marketing Approval Date')
    # pm_approval_date = fields.Datetime('Program Management Approval Date')
    # maneg_approval_date = fields.Datetime('Top Management Approval Date')
    #
    # def hr_approval(self):
    #     self.hr_approval_date = datetime.now()
    #     self.is_hr_approved = True
    #     self.approve_by_department_ids = [(4, self.env.user.employee_id.department_id.id)]
    #     self.approvaed_manager_ids = [(4, self.env.user.employee_id.department_id.manager_id.id)]
    # def design_approval(self):
    #     self.design_approval_date = datetime.now()
    #     self.is_design_approved = True
    #     self.approve_by_department_ids = [(4, self.env.user.employee_id.department_id.id)]
    #     self.approvaed_manager_ids = [(4, self.env.user.employee_id.department_id.manager_id.id)]
    # def eng_approval(self):
    #     self.eng_approval_date = datetime.now()
    #     self.is_eng_approved = True
    #     self.approve_by_department_ids = [(4, self.env.user.employee_id.department_id.id)]
    #     self.approvaed_manager_ids = [(4, self.env.user.employee_id.department_id.manager_id.id)]
    # def manu_approval(self):
    #     self.manu_approval_date = datetime.now()
    #     self.is_manu_approved = True
    #     self.approve_by_department_ids = [(4, self.env.user.employee_id.department_id.id)]
    #     self.approvaed_manager_ids = [(4, self.env.user.employee_id.department_id.manager_id.id)]
    # def sales_approval(self):
    #     self.sale_approval_date = datetime.now()
    #     self.is_sales_approved = True
    #     self.approve_by_department_ids = [(4, self.env.user.employee_id.department_id.id)]
    #     self.approvaed_manager_ids = [(4, self.env.user.employee_id.department_id.manager_id.id)]
    # def qc_approval(self):
    #     self.qc_approval_date = datetime.now()
    #     self.is_qc_approved = True
    #     self.approve_by_department_ids = [(4, self.env.user.employee_id.department_id.id)]
    #     self.approvaed_manager_ids = [(4, self.env.user.employee_id.department_id.manager_id.id)]
    # def mainte_approval(self):
    #     self.main_approval_date = datetime.now()
    #     self.is_maintenance_approved = True
    #     self.approve_by_department_ids = [(4, self.env.user.employee_id.department_id.id)]
    #     self.approvaed_manager_ids = [(4, self.env.user.employee_id.department_id.manager_id.id)]
    # def marketing_approval(self):
    #     self.mark_approval_date = datetime.now()
    #     self.is_marketing_approved = True
    #     self.approve_by_department_ids = [(4, self.env.user.employee_id.department_id.id)]
    #     self.approvaed_manager_ids = [(4, self.env.user.employee_id.department_id.manager_id.id)]
    # def pm_approval(self):
    #     self.pm_approval_date = datetime.now()
    #     self.is_pm_approved = True
    #     self.approve_by_department_ids = [(4, self.env.user.employee_id.department_id.id)]
    #     self.approvaed_manager_ids = [(4, self.env.user.employee_id.department_id.manager_id.id)]
    # def managment_approval(self):
    #     self.maneg_approval_date = datetime.now()
    #     self.is_management_approved = True
    #     self.approve_by_department_ids = [(4, self.env.user.employee_id.department_id.id)]
    #     self.approvaed_manager_ids = [(4, self.env.user.employee_id.department_id.manager_id.id)]

    def action_generate_excel_report(self):
        # Create a new workbook
        output = BytesIO()
        wb = Workbook()
        ws = wb.active

        # region Adding static data
        data = {
            'B1': 'Engineering change request(ECR)/Engineering change note(ECN)',
            'I1': 'ECR No.:',
            'I2': 'ECR Date:',
            'A3': 'CHANGE RECEIVED FROM CUSTOMER',
            'F3': 'INTERNAL',
            'A4': 'Proposer Name:',
            'A5': 'Part Name:',
            'A6': 'Part No.:',
            'A7': 'Customer Name:',
            'F4': 'Department.:',
            'F5': 'Process  Name',
            'F6': 'Drg. No.:',
            'F7': 'Control Plan No.:',
            'A8': 'Change Description',
            'A9': 'Existing',
            'F9': 'Proposed',
            'A10': 'Existing Material Disposition :',
            'A12': 'Scraped',
            'A13': 'To be used as it is',
            'A14': 'To be converted as per new specification:',
            'A15': 'Purpose of Change:  (Tick wherever applicable)',
            'A17': 'Design / Specification Change',
            'D17': 'Standardisation',
            'H17': 'Improvement In Machining',
            'A18': 'Improvement In Performance / Quality',
            'D18': 'Sub- Supplier  Requirement',
            'H18': 'Customer  Requirement (ECN No.)',
            'A19': 'Cost Reduction',
            'D19': 'Other (Specify)',
            'A20': 'Remarks:',
            'A22': 'Impact of Changes on :  (Tick wherever applicable)',
            'E22': 'B. Owner: CFT',
            'A23': '1.Can engineering performance specifications be met as written ?',
            'A24': '2.Is there adequate capacity to produce product ?',
            'A25': '3.Can product be manufactured with Cpks that meet requirements ?',
            'A26': '4.Is statistical process control required on product ?',
            'F23': '5.Does the design allow the use of efficient material handling techniques ?',
            'F24': '6.Costs for tooling ?',
            'F25': '7.Costs for capital equipment ?',
            'F26': '8.Alternative manufacturing methods ?',
            'A27': 'Feasible / Not Feasible  (Tick where applicable)',
            'B28': 'Feasible',
            'C28': 'Some changes recommended',
            'B29': 'Not Feasible',
            'C29': 'Design revision required to produce product within the specified requirements.',
            'A31': 'Signature of CFT',
            'B31': 'Marketing',
            'C31': 'Production',
            'E31': 'Engg.',
            'F31': 'Maintenance',
            'G31': 'Logistic',
            'I31': 'Others',
            'A32': 'Formats',
            'B32': 'Change Required (Y/N)',
            'C32': 'Responsibility',
            'G32': 'Target date',
        }
        # Fill data into specific cells using key-value pairs
        for cell, value in data.items():
            ws[cell] = value
        # endregion

        # region Formatting the Sheet
        thin = Side(border_style='thin', color='000000')
        thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        font_header = Font(name='Arial', size=10, bold=False)
        align_center = Alignment(vertical='center', horizontal='center', wrapText=True)
        align_left = Alignment(vertical='center', horizontal='left', wrapText=True)
        align_right = Alignment(vertical='center', horizontal='right')

        # merging and formatting cells
        max_col = 10
        max_row = 46
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.alignment = align_left
                cell.font = font_header

        # Specific Dimension
        ws['B1'].alignment = align_center
        ws['B1'].font = Font(name='Arial', size=18, bold=True)
        ws['F23'].alignment = Alignment(vertical='center', horizontal='left', wrapText=True)
        ws['G31'].alignment = align_right
        ws['A50'].alignment = align_right

        # Bold Cells and underline
        bold_list = ['I1', 'I2', 'A3', 'F3', 'A8', 'A9', 'F9', 'A10', 'A15',
                     'A20', 'A22', 'E22', 'A27', 'A32', 'B32', 'C32', 'G32']
        for bold_cell in bold_list:
            ws[bold_cell].font = Font(name='Arial', size=11, bold=True)
        center_list = ['A3', 'F3', 'A8', 'B32', 'C32', 'G32', 'A49', 'A32']
        for center_cell in center_list:
            ws[center_cell].alignment = align_center

        # Border Cells
        for row in ws.iter_rows(min_row=1, max_row=46, min_col=1, max_col=10):
            for border_cell in row:
                border_cell.border = thin_border

        # Merging the cells as per standard sheet
        merge_list = ['A1:A2', 'B1:H2', 'A3:B3', 'C3:E3', 'F3:G3', 'H3:J3',
                      'A8:J8', 'B9:E9', 'G9:J9', 'A10:A11', 'B10:J11', 'C12:J14',
                      'A15:J16', 'A20:A21', 'B20:J21', 'A27:J27', 'C28:J28', 'C29:J29',
                      'A30:J31', 'E22:J22']
        for merge_cell in merge_list:
            ws.merge_cells(merge_cell)
        for i in range(4, 7 + 1):
            ws.merge_cells(f"B{i}:E{i}")
            ws.merge_cells(f"F{i}:G{i}")
            ws.merge_cells(f"H{i}:J{i}")
        for i in range(17, 19 + 1):
            ws.merge_cells(f"B{i}:C{i}")
            ws.merge_cells(f"D{i}:E{i}")
            ws.merge_cells(f"F{i}:G{i}")
            ws.merge_cells(f"H{i}:I{i}")
        for i in range(22, 26 + 1):
            ws.merge_cells(f"A{i}:D{i}")
            ws.merge_cells(f"F{i}:I{i}")
        for i in range(32, 39 + 1):
            ws.merge_cells(f"C{i}:F{i}")
            ws.merge_cells(f"G{i}:J{i}")
        # for i in range(41, 46 + 1):
        #     ws.merge_cells(f"A{i}:B{i}")
        #     ws.merge_cells(f"C{i}:J{i}")

        # Dimension of Columns
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 8
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 10.5
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 12.5
        ws.column_dimensions['H'].width = 9
        ws.column_dimensions['I'].width = 17
        ws.column_dimensions['J'].width = 20

        # Dimensions of Row
        for rows in range(1, 29 + 1):
            ws.row_dimensions[rows].height = 28
        ws.row_dimensions[1].height = 45
        ws.row_dimensions[2].height = 45
        ws.row_dimensions[9].height = 50
        ws.row_dimensions[16].height = 15

        # endregion

        # region Attaching LOGO here
        max_width = 300  # Set your desired maximum width
        max_height = 100  # Set your desired maximum height
        if self.env.user.company_id.logo:
            image_data = base64.b64decode(self.env.user.company_id.logo)

            # Open the image using PIL
            image = PILImage.open(io.BytesIO(image_data))
            width, height = image.size
            aspect_ratio = width / height

            if width > max_width:
                width = max_width
                height = int(width / aspect_ratio)

            if height > max_height:
                height = max_height
                width = int(height * aspect_ratio)

            # Resize the image using PIL
            # Add space on the top and left side of the image
            padding_top = 10  # Adjust as needed
            padding_left = 10  # Adjust as needed

            resized_image = image.resize((width, height), PILImage.Resampling.LANCZOS)
            ImageOps.expand(resized_image, border=(padding_left, padding_top, 0, 0), fill='rgba(0,0,0,0)')
            img_bytes = io.BytesIO()
            resized_image.save(img_bytes, format='PNG')
            img_bytes.seek(0)
            logo_image = Image(img_bytes)
            # logo_image = Image(self.env.user.company_id.logo)
            ws.add_image(logo_image, 'A1')

        # endregion

        # region Adding dynamic data to sheet
        for rec in self:
            ws['B4'] = rec.emp_id.name if rec.emp_id else ''
            ws['B5'] = rec.product_id.name if rec.product_id else ''
            ws['B6'] = rec.part_no if rec.part_no else ''
            ws['B7'] = rec.customer_id.name if rec.customer_id else ''

            ws['H4'] = rec.dept_id.id if rec.dept_id else ''
            ws['H5'] = rec.processer_name.name if rec.processer_name else ''
            ws['H6'] = rec.drg_no if rec.drg_no else ''
            ws['H7'] = rec.control_paln if rec.control_paln else ''
            # ☑☐
            ws['B12'] = '☑' if rec.scraped else '☐'
            ws['B13'] = '☑' if rec.userd_in else '☐'
            ws['B14'] = '☑' if rec.specification else '☐'

            ws['B17'] = '☑' if rec.specific_change else '☐'
            ws['B18'] = '☑' if rec.quality_imp else '☐'
            ws['B19'] = '☑' if rec.cost_reduce else '☐'

            ws['F17'] = '☑' if rec.standardisation else '☐'
            ws['F18'] = '☑' if rec.sub_suppiler else '☐'

            ws['J17'] = '☑' if rec.improve_machine else '☐'
            ws['J18'] = '☑' if rec.customer_requirement else '☐'

            ws['E23'] = '☑' if rec.first_impact else '☐'
            ws['E24'] = '☑' if rec.secound_impact else '☐'
            ws['E25'] = '☑' if rec.third_impact else '☐'
            ws['E26'] = '☑' if rec.forth_impact else '☐'
            ws['J23'] = '☑' if rec.fifth_impact else '☐'
            ws['J24'] = '☑' if rec.sixth_impact else '☐'
            ws['J25'] = '☑' if rec.seven_impact else '☐'
            ws['J26'] = '☑' if rec.eight_impact else '☐'

            ws['A28'] = '☑' if rec.feasible else '☐'
            ws['A29'] = '☑' if rec.not_feasible else '☐'

        # Checkbox cell formatting
        checkbox_list = ['B12', 'B13', 'B14', 'B17', 'B18', 'B19', 'F17', 'F18',
                         'J17', 'J18', 'E23', 'E24', 'E25', 'E26', 'J23', 'J24',
                         'J25', 'J26', 'A28', 'A29']
        for check_cell in checkbox_list:
            ws[check_cell].alignment = align_center
            if ws[check_cell].value == '☑':
                ws[check_cell].font = Font(size=15, color='18b004', bold=True)
            else:
                ws[check_cell].font = Font(size=15, bold=True)

        # For Changes Required section
        cur_row = 33
        max_row_line = 39
        row_cell = cur_row-1
        for rec in self:
            for i in rec.change_required_ids:
                ws[f'A{cur_row}'] = i.format.name if i.format.name else ''
                ws[f'B{cur_row}'] = 'Yes' if i.change_required else 'No'
                ws[f'C{cur_row}'] = i.responsibility_id.name if i.responsibility_id.name else ''
                ws[f'G{cur_row}'] = i.target_date if i.target_date else ''

                ws.merge_cells(f"C{cur_row}:F{cur_row}")
                ws.merge_cells(f"G{cur_row}:J{cur_row}")

                for cell in ws[cur_row]:
                    cell.alignment = align_center

                cur_row += 1
            if cur_row < max_row_line:
                cur_row = max_row_line
        ws.merge_cells(f'A{cur_row}:J{cur_row}')
        cur_row += 1
        ws[f'A{cur_row}'] = 'PPAP Approval Reqd. (Yes/No)'
        ws[f'A{cur_row}'].font = Font(name='Arial', size=11, bold=True)
        ws[f'A{cur_row+1}'].font = Font(name='Arial', size=11, bold=True)
        ws.merge_cells(f'A{cur_row}:B{cur_row}')
        ws.merge_cells(f'C{cur_row}:J{cur_row}')
        ws[f'A{cur_row + 1}'] = 'PRODUCTION PART APPROVAL PROCESS DETAILS'
        ws[f'A{cur_row + 2}'] = 'Date of PPAP Submission '
        ws[f'A{cur_row + 3}'] = 'Date of Approval Received '
        ws[f'A{cur_row + 4}'] = 'Date of Pilot Lot Submission'
        ws[f'A{cur_row + 5}'] = 'Date of Regular Production'
        for rec in self:
            ws[f'C{cur_row}'] = 'Yes' if rec.ppap_approval_required else 'No'
            ws[f'C{cur_row + 2}'] = rec.date_of_ppap_submission if rec.date_of_ppap_submission else ''
            ws[f'C{cur_row + 3}'] = rec.date_of_approval_received if rec.date_of_approval_received else ''
            ws[f'C{cur_row + 4}'] = rec.date_of_pilot_lot_submission if rec.date_of_pilot_lot_submission else ''
            ws[f'C{cur_row + 5}'] = rec.date_of_regular_submission if rec.date_of_regular_submission else ''
        for i in range(1, 6):
            ws[f'A{cur_row + i}'].alignment = align_center
            ws[f'C{cur_row + i}'].alignment = align_center
            ws.merge_cells(f"A{cur_row + i}:B{cur_row + i}")
            ws.merge_cells(f"C{cur_row + i}:J{cur_row + i}")
        for r_cell in range(row_cell, cur_row+6):
            ws.row_dimensions[r_cell].height = 25
        cur_row += 6
        # endregion

        # region SignOff Members Footer
        sign_row = cur_row
        ws.merge_cells(f'A{cur_row}:J{cur_row}')
        cur_row += 1
        ws[f'A{cur_row}'] = 'Prepared By'
        ws[f'A{cur_row}'].font = font_header
        ws.merge_cells(f'B{cur_row}:E{cur_row}')

        ws[f'F{cur_row}'] = 'Prepared Date'
        ws[f'F{cur_row}'].font = font_header
        ws.merge_cells(f'F{cur_row}:H{cur_row}')
        ws.merge_cells(f'I{cur_row}:J{cur_row}')

        for rec in self:
            ws[f'B{cur_row}'] = rec.create_uid.name if rec.create_uid else ''
            ws[f'I{cur_row}'] = rec.create_date if rec.create_date else ''

        ws.row_dimensions[cur_row].height = 18
        cur_row += 1
        ws.merge_cells(f'A{cur_row}:J{cur_row}')
        cur_row += 1
        ws.merge_cells(f'A{cur_row}:J{cur_row}')
        ws[f'A{cur_row}'] = 'Sign OFF'
        ws[f'A{cur_row}'].font = Font(size=18, bold=True)
        ws.row_dimensions[cur_row].height = 25
        cur_row += 1
        ws.merge_cells(f'A{cur_row}:J{cur_row}')
        cur_row += 1
        for cell in ws[cur_row]:
            ws[f'A{cur_row}'] = 'Member'
            ws[f'B{cur_row}'] = 'Department'
            ws[f'C{cur_row}'] = 'Status'
            ws[f'F{cur_row}'] = 'Date'
            ws[f'H{cur_row}'] = 'Comments'
            ws.row_dimensions[cur_row].height = 25
            ws.merge_cells(f'C{cur_row}:E{cur_row}')
            ws.merge_cells(f'F{cur_row}:G{cur_row}')
            ws.merge_cells(f'H{cur_row}:J{cur_row}')
            cell.font = Font(size=12, bold=True, color='ffffff')
            cell.fill = PatternFill(start_color='0070C0', end_color='0070C0', fill_type="solid")

        for row_no in ws.iter_rows(min_row=47, max_row=cur_row + 1, min_col=1, max_col=10):
            for cell in row_no:
                cell.border = border
                cell.alignment = align_center

        # Listing Managers Id
        for rec in self:
            for record in rec.iatf_members_ids:
                name_rec = record.approver_id.name
                dept_rec = record.department_id.name if record.department_id.name else ''
                status_rec = record.approval_status.capitalize()
                date_rec = record.date_approved_rejected if record.date_approved_rejected else ''
                comment_rec = record.comment if record.comment else ''

                ws[f'A{cur_row + 1}'] = name_rec
                ws[f'B{cur_row + 1}'] = dept_rec
                ws[f'C{cur_row + 1}'] = status_rec
                ws[f'F{cur_row + 1}'] = date_rec
                ws[f'H{cur_row + 1}'] = comment_rec
                ws.merge_cells(f'C{cur_row + 1}:E{cur_row + 1}')
                ws.merge_cells(f'F{cur_row + 1}:G{cur_row + 1}')
                ws.merge_cells(f'H{cur_row + 1}:J{cur_row + 1}')

                status_dict = {'Approved': ['CFFFC3', '000000'], 'Rejected': ['FFCDCD', '000000'],
                               'Revision': ['AFEFFF', '000000'], 'Pending': ['FDFFCD', '000000']}

                for state, color in status_dict.items():
                    if state == status_rec:
                        ws[f'C{cur_row + 1}'].fill = PatternFill(start_color=color[0], end_color=color[0],
                                                                 fill_type="solid")
                        ws[f'C{cur_row + 1}'].font = Font(size=12, bold=False, color=color[1])

                cur_row += 1

            ws.merge_cells(f'A{cur_row + 1}:J{cur_row + 1}')
            ws[
                f'A{cur_row + 1}'] = 'Note 1.If Customer has raised any ECN or request, HOD – QA / Incharge – Mktg raises this ECR.'

            ws.merge_cells(f'A{cur_row + 2}:J{cur_row + 2}')

            for row_no in ws.iter_rows(min_row=sign_row + 1, max_row=cur_row + 2, min_col=1, max_col=10):
                for cell in row_no:
                    cell.border = border
                    cell.alignment = align_center
        # endregion
        ws.sheet_view.showGridLines = False
        ws.row_dimensions[30].height = 10
        ws.row_dimensions[31].height = 10

        wb.save(output)
        output.seek(0)
        self.generate_xls_file = base64.b64encode(output.getvalue()).decode('utf-8')
        return {
            "type": "ir.actions.act_url",
            "target": "self",
            "url": "/web/content?model=change.request&download=true&field=generate_xls_file&filename={filename}.xlsx&id={pid}".format(
                filename="Engineering change request(ECR) Engineering change note(ECN)", pid=self.id),
        }
    # request_ids = fields.One2many(comodel_name='change.request.line',
    #     inverse_name='change_request_id',
    #     string="Change Request Line"
    # )


class ChangeRequired(models.Model):
    _name = 'change.required.line'
    _description = 'Change Required Line'

    change_required_id = fields.Many2one(
        comodel_name='change.request',
        string='Change Required Line ',
        required=True, ondelete='cascade', index=True, copy=False
    )
    sl_no = fields.Integer('S.No', compute='_compute_sequence_number')
    format = fields.Many2one('format.document', 'Format')
    change_required = fields.Boolean("Change Required (Y/N)")
    responsibility_id = fields.Many2one("res.users", 'Responsibility')
    target_date = fields.Date('Target Date')

    @api.depends('change_required_id')
    def _compute_sequence_number(self):
        for line in self.mapped('change_required_id'):
            sl_no = 1
            for lines in line.change_required_ids:
                lines.sl_no = sl_no
                sl_no += 1


class ChangeRequestFormats(models.Model):
    _name = 'format.document'
    _description = 'Format Document'

    name = fields.Char('Format', required=True)


class ProductTemplate(models.Model):
    _inherit = "product.template"

    drg_no = fields.Char('Drawing Number')
