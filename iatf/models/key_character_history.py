from datetime import date, datetime
from odoo import api, fields, models
from odoo.exceptions import ValidationError
from datetime import date
import openpyxl
from io import BytesIO
import io
from openpyxl import Workbook
import openpyxl
import base64
from bs4 import BeautifulSoup
from openpyxl.styles import Alignment, Font, Border, Side, DEFAULT_FONT, PatternFill
from openpyxl.drawing.image import Image
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
import copy
import datetime
from PIL import ImageOps
from PIL import Image as PILImage
from bs4 import BeautifulSoup



class KeyCharacterHistory(models.Model):
    _name = 'key.character.history'
    _inherit = ['iatf.sign.off.members']

    # partner_id = fields.Many2one('res.partner', 'Customer')
    address = fields.Char("Customer Address", related="partner_id.contact_address")
    mobile = fields.Char("Phone Number", related="partner_id.phone")
    email = fields.Char('Email', related="partner_id.email")
    # product_id = fields.Many2one('product.template','Part Name')
    # default_code = fields.Char('Part Number', related='product_id.default_code')
    key_character_line_ids = fields.One2many(
        comodel_name='key.character.history.line',
        inverse_name='key_character_id',
        string="Key Character History Line"
    )

    part_development_id = fields.Many2one("part.development.process")

    document_name = fields.Char(string='Document #')
    generate_xls_file = fields.Binary(string="Generated file")  # do not comment this

    def action_generate_excel_report(self):
        output = BytesIO()
        wb = Workbook()
        ws = wb.active

        if self.env.user.company_id.logo:
            max_width = 300  # Set your desired maximum width
            max_height = 80  # Set your desired maximum height
            image_data = base64.b64decode(self.env.user.company_id.logo)

            # Open the image using PIL
            image = PILImage.open(io.BytesIO(image_data))
            width, height = image.size
            aspect_ratio = width / height

            if width > max_width:
                width = max_width
                height = int(width / aspect_ratio)

            if height > max_height:
                height = max_height
                width = int(height * aspect_ratio)

            # Resize the image using PIL
            # Add space on the top and left side of the image
            padding_top = 10  # Adjust as needed
            padding_left = 10  # Adjust as needed

            resized_image = image.resize((width, height), PILImage.LANCZOS)
            ImageOps.expand(resized_image, border=(padding_left, padding_top, 0, 0), fill='rgba(0,0,0,0)')
            img_bytes = io.BytesIO()
            resized_image.save(img_bytes, format='PNG')
            img_bytes.seek(0)
            logo_image = Image(img_bytes)
            # logo_image = Image(self.env.user.company_id.logo)
            ws.add_image(logo_image, 'A1')

        border = Border(top=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'),
                        bottom=Side(style='thin'))
        align_center = Alignment(vertical='center', horizontal='center', wrapText=True)
        align_left = Alignment(vertical='center', horizontal='left')
        font_header = Font(name='Arial', size=12, bold=True)
        font_all = Font(name='Times New Roman', size=11, bold=False)
        grey_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        blue_fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')

        data = {
            'A1': 'KEY CHARACTERISTIC MATRIX',
            'G2': 'Part Name& No.:',
            'A2': 'Customer :',
            'D3': 'ROUTING',
            'C5': 'Specification',
            'B5': 'Characterstic   Description',
            'A5': 'S.No.',
            'B16': 'Special Characteristics as per drawing',
            'B18': 'Special Characteristics for Process',
            'H16': 'Special Characteristics for    Safety',
            'H19': 1,
            'A19': 'Key characterstic matrix',
            'G19': 'Page',

        }

        for cell, value in data.items():
            ws[cell] = value

        cell_ranges_to_merge = [
            'A1:L1', 'A2:F2', 'G2:L2', 'A3:C4', 'D3:L3', 'A19:F19', 'B16:F17', 'B18:F18', 'H16:L18', 'D4:D4', 'E4:E4',
            'F4:F4',
            'G4:G4', 'H4:H4', 'I4:I4', 'J4:J4', 'K4:K4', 'L4:L4', 'G16:G18', 'A20:L20', 'G19:G19', 'H19:L19', 'A16:A17',
            'A18:A18'
        ]
        for cell_range in cell_ranges_to_merge:
            ws.merge_cells(cell_range)
            for row in ws[cell_range]:
                for cell in row:
                    cell.border = border

        for cell_reference in ['A1']:
            ws[cell_reference].font = Font(name='Arial', size=14, bold=True)

        ws['A5'].font = Font(name='Arial', size=10, bold=True)
        ws['B5'].font = Font(name='Arial', size=10, bold=True)
        ws['C5'].font = Font(name='Arial', size=10, bold=True)
        ws['A2'].font = Font(name='Arial', size=12, bold=True)
        ws['G2'].font = Font(name='Arial', size=12, bold=True)
        ws['D3'].font = Font(name='Arial', size=12, bold=True)

        ws['B5'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws['G19'].alignment = Alignment(horizontal='left', vertical='center')
        ws['H19'].alignment = Alignment(horizontal='left', vertical='center')
        ws['A20'].alignment = Alignment(horizontal='right', vertical='center')
        cell_addresses = ['A1', 'A2', 'G2', 'D3', 'A5', 'A16', 'A18', 'C5']
        for cell_address in cell_addresses:
            ws[cell_address].alignment = Alignment(horizontal='center', vertical='center')

        ws.row_dimensions[16].height = 15
        ws.row_dimensions[17].height = 15
        ws.row_dimensions[18].height = 30
        ws.row_dimensions[1].height = 30
        ws.row_dimensions[2].height = 30
        ws.row_dimensions[3].height = 30
        ws.row_dimensions[4].height = 30
        ws.row_dimensions[5].height = 55
        rows_to_set_height = [6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 19]
        for row_num in rows_to_set_height:
            ws.row_dimensions[row_num].height = 18

        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        cell_range_to_apply_width = ['A', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']
        for cell_range in cell_range_to_apply_width:
            ws.column_dimensions[cell_range].width = 9

        rows_to_apply_border = range(5, 16)
        for row in rows_to_apply_border:
            for col in range(1, ws.max_column + 1):
                cell_reference = ws.cell(row=row, column=col)
                cell_reference.border = border

        wb.save(output)
        output.seek(0)
        self.generate_xls_file = base64.b64encode(output.getvalue()).decode('utf-8')
        # endregion

        return {
            "type": "ir.actions.act_url",
            "target": "self",
            "url": "/web/content?model=key.character.history&download=true&field=generate_xls_file&filename={filename}.xlsx&id={pid}".format(
                filename="Key Character History", pid=self[0].id),
        }

class KeyChatacterHistoryLine(models.Model):
    _name = 'key.character.history.line'
    _inherit = "translation.mixin"

    key_character_id = fields.Many2one(
        comodel_name='key.character.history',
        string='Key Character History',
        required=True, ondelete='cascade', index=True, copy=False
    )
    sequence = fields.Integer(string="Sequence", default=10)
    sl_no = fields.Integer("S.No", compute="_compute_sequence_number")
    description = fields.Char('Characterstic Description',translate=True)
    specification = fields.Char('Specification',translate=True)

    @api.depends('sequence', 'key_character_id')
    def _compute_sequence_number(self):
        for line in self.mapped('key_character_id'):
            sl_no = 1
            for lines in line.key_character_line_ids:
                lines.sl_no = sl_no
                sl_no += 1
