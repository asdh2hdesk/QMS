from odoo import fields, models, api, _
from odoo.exceptions import ValidationError
from datetime import date
import openpyxl
from io import BytesIO
import io
from openpyxl import Workbook
import openpyxl
import base64
from openpyxl.styles import Alignment, Font, Border, Side, DEFAULT_FONT, PatternFill
from openpyxl.drawing.image import Image
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
import copy
import datetime
from PIL import ImageOps
from PIL import Image as PILImage


class PfmeaOne(models.Model):
    _name = "asd.pfmea.one"
    _description = "PFMEA"
    _inherit = ['iatf.sign.off.members', 'translation.mixin']

    # part_name = fields.Char("Part Name")
    fmea_no = fields.Char("FMEA No")
    # core_team = fields.Char("Core Team")
    process_responsibility = fields.Char("Process Responsibility",translate=True)
    prepared_by = fields.Char("Prepared By",translate=True)
    mode_years = fields.Char("Mode Years")
    fmea_date = fields.Date("FMEA Date")
    key_date = fields.Date("Key Date")
    fmea_rev_date = fields.Date("FMEA Rev. Date")
    operation_ids = fields.One2many(
        comodel_name='asd.pfmea.one.operations',
        inverse_name='pfmea_id',
        string="Operations")

    generate_xls_file = fields.Binary(string="Generated file")  # do not comment this

    @api.model_create_multi
    def create(self, vals_list):
        records = super(PfmeaOne, self).create(vals_list)

        for record in records:
            pfd_records = self.env["process.flow"].search(
                [('project_id', '=', record.project_id.id),
                 ('final_status', '=', 'approved')
                 ],limit=1)
            # if not pfd_records:
            #     raise ValidationError(
            #         _("No Process Flow Diagram is filled in this project first fill PFD before continuing..."))

            for operation in pfd_records.process_flow_line_ids.sorted(key="step"):
                pfmea_operation = self.env['asd.pfmea.one.operations'].create({
                    'pfmea_id': record.id,
                    'operation': operation.step,
                    'stage_pfmea': operation.stage.id,
                    'desc_of_operation_pfmea': operation.desc_of_operation,
                    # 'operation': f"{operation.step} - {operation.stage.name} - {operation.desc_of_operation}",
                })
                for i in operation.process_op_lines_ids:
                    self.env['asd.pfmea.one.operations.line'].create({
                        'pfmea_operation_id': pfmea_operation.id,
                        'process_step': i.element_no,
                        'process_desc': i.element_desc,
                    })

        return records

    def action_update_process_flow(self):
        """ Update Process Flow data from PFMEA """
        for record in self:
            # Find related process flow records based on the project
            process_flow_main = self.env["process.flow"].search([
                ('project_id', '=', record.project_id.id),
                ('final_status', '!=', 'approved')
            ])
            if not process_flow_main:
                raise ValidationError(_("No Inprogress Process Flow found for this project."))

            for pfmea_operation in record.operation_ids:
                # Find existing Process Flow operation
                process_flow_operation = process_flow_main.process_flow_line_ids.filtered(lambda x: x.step == pfmea_operation.operation)

                if process_flow_operation:
                    # Update existing Process Flow operation
                    process_flow_operation.write({
                        'stage': pfmea_operation.stage_pfmea.id if pfmea_operation.stage_pfmea else False,
                        'desc_of_operation': pfmea_operation.desc_of_operation_pfmea,
                    })
                else:
                    process_flow_operation.create({
                        'process_flow_id':process_flow_main.id,
                        'step':pfmea_operation.operation,
                        'stage':pfmea_operation.stage_pfmea.id,
                        'desc_of_operation':pfmea_operation.desc_of_operation_pfmea
                    })

                for pfmea_element in pfmea_operation.operation_lines_ids:
                    pf_elements = process_flow_operation.process_op_lines_ids.filtered(
                        lambda x: x.element_no == pfmea_element.process_step)
                    if pf_elements:
                        pf_elements.write({
                            'element_desc': pfmea_element.process_desc
                        })
                    else:
                        pf_elements.create({
                            'operation_element_id': process_flow_operation.id,
                            'element_no': pfmea_element.process_step,
                            'element_desc': pfmea_element.process_desc,
                        })

        return True
    def action_update_process_matrix(self):
        """ Update Process Flow data from PFMEA """
        for record in self:
            # Find related process flow records based on the project
            process_matrix_main = self.env["process.group"].search([
                ('project_id', '=', record.project_id.id),
                ('final_status', '!=', 'approved')
            ])
            if not process_matrix_main:
                raise ValidationError(_("No Inprogress Process Flow found for this project."))

            for pfmea_operation in record.operation_ids:
                # Find existing Process Flow operation
                process_matrix_op = process_matrix_main.process_presentation_ids.filtered(lambda x: x.operation == pfmea_operation.operation)

                if process_matrix_op:
                    # Update existing Process Flow operation
                    process_matrix_op.write({
                        'operation_description': pfmea_operation.desc_of_operation_pfmea,
                    })
                else:
                    process_matrix_op.create({
                        'process_id':process_matrix_main.id,
                        'operation':pfmea_operation.operation,
                        'operation_description':pfmea_operation.desc_of_operation_pfmea
                    })

                for pfmea_element in pfmea_operation.operation_lines_ids:
                    pm_elements = process_matrix_op.operation_lines_ids.filtered(
                        lambda x: x.element_no == pfmea_element.process_step)
                    if pm_elements:
                        pm_elements.write({
                            'element_description': pfmea_element.process_desc
                        })
                    else:
                        pm_elements.create({
                            'operation_element_id': process_matrix_op.id,
                            'element_no': pfmea_element.process_step,
                            'element_description': pfmea_element.process_desc,
                        })

        return True
        pass


    def action_generate_excel_report(self):
        output = BytesIO()
        wb = Workbook()
        ws = wb.active

        if self.env.user.company_id.logo:
            max_width = 300  # Set your desired maximum width
            max_height = 100  # Set your desired maximum height
            image_data = base64.b64decode(self.env.user.company_id.logo)

            # Open the image using PIL
            image = PILImage.open(io.BytesIO(image_data))
            width, height = image.size
            aspect_ratio = width / height

            if width > max_width:
                width = max_width
                height = int(width / aspect_ratio)

            if height > max_height:
                height = max_height
                width = int(height * aspect_ratio)

            # Resize the image using PIL
            # Add space on the top and left side of the image
            padding_top = 10  # Adjust as needed
            padding_left = 10  # Adjust as needed

            resized_image = image.resize((width, height), PILImage.LANCZOS)
            ImageOps.expand(resized_image, border=(padding_left, padding_top, 0, 0), fill='rgba(0,0,0,0)')
            img_bytes = io.BytesIO()
            resized_image.save(img_bytes, format='PNG')
            img_bytes.seek(0)
            logo_image = Image(img_bytes)
            # logo_image = Image(self.env.user.company_id.logo)
            ws.add_image(logo_image, 'A1')

        border = Border(top=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'),
                        bottom=Side(style='thin'))
        align_center = Alignment(vertical='center', horizontal='center', wrapText=True)
        align_left = Alignment(vertical='center', horizontal='left')
        font_header = Font(name='Arial', size=12, bold=True)
        font_all = Font(name='Times New Roman', size=11, bold=False)

        data = {
            'B1': 'POTENTIAL FAILURE MODE AND EFFECT ANALYSIS',
            'B5': 'Part name/Part number:',
            # 'E5': 'Process Responsibility :',
            'G5': "Model Year's) / Vehicle's) : ",
            'L5': 'KEY DATE:-',
            'B6': 'FMEA Number :',
            # 'E6': 'Prepared By :',
            'G6': 'FMEA Date: ',
            'L6': 'FMEA (REV): Rev. Date: - ',
            'B7': 'Customer :',
            'B8': 'Process Step /Function ',
            'C8': 'Requirement',
            'D8': 'Potential Failure Mode',
            'E8': 'Potential Effect(s) of Failure',
            'F8': 'Severity',
            'G8': 'Class',
            'H8': 'Potential Cause(s) / Mechanism(s) of Failure',
            'I8': 'Prevention ',
            'J8': 'O C C.',
            'K8': 'Detection ',
            'L8': 'Detection',
            'M8': 'R P N.',
            'N8': 'Recommended Action(s)',
            'O8': 'Responsibility & Target Completion Date',
            'P8': 'Action Results',
            'P9': 'Action Taken',
            'Q9': 'S E V.',
            'R9': 'O C C.',
            'S9': 'D E T.',
            'T9': 'R P N.',

        }

        for cell, value in data.items():
            ws[cell] = value

        max_row = 30

        cell_ranges_to_merge = ['B1:T4', 'C5:F5', 'I5:K5', 'G5:H5', 'G6:H6', 'I6:K6', 'L5:N5', 'O5:T5',
                                'C6:F6', 'L6:N6', 'P6:T6', 'O6:T6', 'C7:F7', 'G7:H7','B8:B9','B8:B9', 'C8:C9', 'D8:D9',
                                'E8:E9', 'F8:F9', 'G8:G9', 'H8:H9', 'I8:I9', 'J8:J9', 'K8:K9', 'L8:L9', 'M8:M9',
                                'N8:N9', 'O8:O9', 'P8:T8', 'P9:P9', 'Q9:Q9', 'R9:R9', 'S9:S9', 'T9:T9',
                                ]
        for cell_range in cell_ranges_to_merge:
            ws.merge_cells(cell_range)

        for i in range(5,8):
            ws.merge_cells(f'C{i}:F{i}')
            ws.merge_cells(f'G{i}:H{i}')
            ws.merge_cells(f'I{i}:K{i}')
            ws.merge_cells(f'L{i}:N{i}')
            ws.merge_cells(f'O{i}:T{i}')

        for row in ws.iter_rows(min_row=1, max_row=9, min_col=2, max_col=20):
            for cell in row:
                cell.border = border
                cell.alignment = align_center
                cell.font = Font(size=12, bold=True, color='ffffff')
                cell.fill = PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type="solid")

        for row_no in ws.iter_rows(min_row=10, max_row=29, min_col=2, max_col=20):
            for cell in row_no:
                cell.border = border

        columns_to_resize = ['L', 'M', 'J', 'F', 'G', 'Q', 'R', 'S', 'T']
        for column in columns_to_resize:
            ws.column_dimensions[column].width = 10

        columns_to_resize = ['B', 'C', 'D', 'E', 'H', 'I', 'K', 'N']
        for column in columns_to_resize:
            ws.column_dimensions[column].width = 30

        ws['P6'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws['B1'].font = Font(name='Arial', size=20, bold=True)
        ws['L1'].font = Font(name='Arial', size=20, bold=True)

        ws.row_dimensions[1].height = 20
        ws.row_dimensions[2].height = 20
        ws.row_dimensions[3].height = 20
        ws.row_dimensions[4].height = 20
        ws.row_dimensions[5].height = 25
        ws.row_dimensions[6].height = 35
        ws.row_dimensions[7].height = 25
        ws.row_dimensions[8].height = 35
        ws.row_dimensions[9].height = 35

        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['O'].width = 20
        ws.column_dimensions['P'].width = 20

        # Data Filling
        cur_row = 10
        for rec in self:
            ws['C5'] = f'{rec.part_name}/{rec.part_number}' if rec.part_id else ''
            ws['C6'] = rec.fmea_no if rec.fmea_no else ''
            ws['C7'] = rec.partner_id.name if rec.partner_id else ''
            ws['I5'] = rec.mode_years if rec.mode_years else ''
            ws['I6'] = rec.fmea_date.strftime("%d-%m-%Y") if rec.fmea_date else ''
            ws['O5'] = rec.key_date.strftime("%d-%m-%Y") if rec.key_date else ''
            ws['O6'] = rec.fmea_rev_date.strftime("%d-%m-%Y") if rec.fmea_rev_date else ''
            for i in rec.operation_ids:
                ws[f'B{cur_row}'] = f' {i.operation}-{i.desc_of_operation_pfmea}-{i.stage_pfmea.name}' if i.operation else ''
                ws[f'B{cur_row}'].fill = PatternFill(start_color='e3e3e3', end_color='e3e3e3', fill_type="solid")
                ws[f'B{cur_row}'].font = Font(size=12, bold=True, color='000000')
                ws[f'B{cur_row}'].border = border
                ws.merge_cells(f'B{cur_row}:T{cur_row}')
                cur_row += 1
                for j in i.operation_lines_ids:
                    ws[f'B{cur_row}'] = j.process_step_desc if j.process_step_desc else ''
                    ws[f'C{cur_row}'] = j.requirement if j.requirement else ''
                    ws[f'D{cur_row}'] = j.potential_failure if j.potential_failure else ''
                    ws[f'E{cur_row}'] = j.potential_effect_of_failure if j.potential_effect_of_failure else ''
                    ws[f'F{cur_row}'] = j.severity if j.severity else '0'
                    ws[f'G{cur_row}'] = j.class_name if j.class_name else ''
                    ws[
                        f'H{cur_row}'] = j.potential_cause_mechanism_of_failure if j.potential_cause_mechanism_of_failure else ''
                    ws[f'I{cur_row}'] = j.prevention if j.prevention else ''
                    ws[f'J{cur_row}'] = j.occ if j.occ else '0'
                    ws[f'K{cur_row}'] = j.detection_desc if j.detection_desc else ''
                    ws[f'L{cur_row}'] = j.detection_no if j.detection_no else '0'
                    ws[f'M{cur_row}'] = j.rpn if j.rpn else '0'
                    ws[f'N{cur_row}'] = j.recommended_action if j.recommended_action else ''
                    ws[f'O{cur_row}'] = j.responsibility_target_date if j.responsibility_target_date else ''
                    ws[f'P{cur_row}'] = j.action_taken if j.action_taken else ''
                    ws[f'Q{cur_row}'] = j.result_sev if j.result_sev else ''
                    ws[f'R{cur_row}'] = j.result_occ if j.result_occ else ''
                    ws[f'S{cur_row}'] = j.result_det if j.result_det else ''
                    ws[f'T{cur_row}'] = j.result_rpn if j.result_rpn else ''
                    for row_no in ws.iter_rows(min_row=cur_row, max_row=cur_row, min_col=2, max_col=20):
                        for cell in row_no:
                            cell.border = border
                            cell.alignment = align_center

                    cur_row += 1

        if cur_row < max_row:
            cur_row = max_row
        # region SignOff Members Footer
        sign_row = cur_row
        ws.merge_cells(f'B{cur_row}:T{cur_row}')

        cur_row += 1
        ws[f'B{cur_row}'] = 'Prepared By'
        ws[f'B{cur_row}'].font = font_header
        # ws.merge_cells(f'A{cur_row}:B{cur_row}')
        ws.merge_cells(f'C{cur_row}:F{cur_row}')

        ws[f'G{cur_row}'] = 'Prepared Date'
        ws[f'G{cur_row}'].font = font_header
        ws.merge_cells(f'G{cur_row}:H{cur_row}')

        for rec in self:
            ws[f'C{cur_row}'] = rec.create_uid.name if rec.create_uid else ''
            ws[f'I{cur_row}'] = rec.create_date if rec.create_date else ''

        ws.row_dimensions[cur_row].height = 18
        cur_row += 1
        ws.merge_cells(f'B{cur_row}:I{cur_row}')
        cur_row += 1
        ws.merge_cells(f'B{cur_row}:I{cur_row}')
        ws[f'B{cur_row}'] = 'Sign OFF'
        ws[f'B{cur_row}'].font = Font(size=18, bold=True)
        ws.row_dimensions[cur_row].height = 25
        cur_row += 1
        ws.merge_cells(f'B{cur_row}:I{cur_row}')
        cur_row += 1
        for cell in ws[cur_row]:
            ws[f'B{cur_row}'] = 'Member'
            ws[f'C{cur_row}'] = 'Department'
            ws[f'D{cur_row}'] = 'Status'
            ws[f'E{cur_row}'] = 'Date'
            ws[f'F{cur_row}'] = 'Comments'
            ws.row_dimensions[cur_row].height = 25
            ws.merge_cells(f'F{cur_row}:I{cur_row}')
            cell.font = Font(size=12, bold=True, color='ffffff')
            cell.fill = PatternFill(start_color='0070C0', end_color='0070C0', fill_type="solid")

        # Listing Managers Id
        for rec in self:
            for record in rec.iatf_members_ids:
                name_rec = record.approver_id.name
                dept_rec = record.department_id.name if record.department_id.name else ''
                status_rec = record.approval_status.capitalize()
                date_rec = record.date_approved_rejected if record.date_approved_rejected else ''
                comment_rec = record.comment if record.comment else ''

                ws[f'B{cur_row + 1}'] = name_rec
                ws[f'C{cur_row + 1}'] = dept_rec
                ws[f'D{cur_row + 1}'] = status_rec
                ws[f'E{cur_row + 1}'] = date_rec
                ws[f'F{cur_row + 1}'] = comment_rec
                ws.merge_cells(f'F{cur_row + 1}:I{cur_row + 1}')

                status_dict = {'Approved': ['CFFFC3', '000000'], 'Rejected': ['FFCDCD', '000000'],
                               'Revision': ['AFEFFF', '000000'], 'Pending': ['FDFFCD', '000000']}

                for state, color in status_dict.items():
                    if state == status_rec:
                        ws[f'D{cur_row + 1}'].fill = PatternFill(start_color=color[0], end_color=color[0],
                                                                 fill_type="solid")
                        ws[f'D{cur_row + 1}'].font = Font(size=12, bold=False, color=color[1])

                cur_row += 1

            # ws.merge_cells(f'C{cur_row + 1}:D{cur_row + 1}')
            # ws.merge_cells(f'E{cur_row + 1}:F{cur_row + 1}')
            ws.merge_cells(f'B{cur_row + 1}:T{cur_row + 1}')
            ws.merge_cells(f'J{sign_row}:T{cur_row}')
            ws.column_dimensions['A'].hidden = True

            for row_no in ws.iter_rows(min_row=sign_row, max_row=cur_row + 1, min_col=2, max_col=20):
                for cell in row_no:
                    cell.border = border
                    cell.alignment = align_center
            # endregion

        cell_orient_list = ['F8', 'G8', 'J8', 'L8', 'M8']
        for cell_rotate in cell_orient_list:
            ws[f'{cell_rotate}'].alignment = Alignment(vertical='center', horizontal='center', textRotation=90)

        # Save the workbook
        wb.save(output)
        output.seek(0)
        self.generate_xls_file = base64.b64encode(output.getvalue()).decode('utf-8')
        # endregion

        return {
            "type": "ir.actions.act_url",
            "target": "self",
            "url": "/web/content?model=asd.pfmea.one&download=true&field=generate_xls_file&filename={filename}.xlsx&id={pid}".format(
                filename="PFMEA", pid=self[0].id),
        }


class PfmeaOneOperation(models.Model):
    _name = "asd.pfmea.one.operations"
    _description = "PFMEA Operation"
    _rec_name = "operation"
    _inherit = "translation.mixin"

    operation = fields.Char("Operation",translate=True)
    stage_pfmea = fields.Many2one('process.flow.stages', 'Stage')
    desc_of_operation_pfmea = fields.Char("Operation Description",translate=True)
    pfmea_operation_details = fields.Char("Operation-Stage-Description", compute="_compute_pfmea_operation_details", store=True,translate=True)
    # operations = fields.Many2one('process.flow.stages', 'Operation')
    # operation_desc = fields.Char('process.flow.stages', 'Operation Description')
    pfmea_id = fields.Many2one('asd.pfmea.one', 'PFMEA')
    operation_lines_ids = fields.One2many(
        comodel_name='asd.pfmea.one.operations.line',
        inverse_name='pfmea_operation_id',
        string="Operations Lines")

    @api.depends('operation', 'stage_pfmea', 'desc_of_operation_pfmea')
    def _compute_pfmea_operation_details(self):
        for record in self:
            operation = record.operation or ''
            stage = record.stage_pfmea.name if record.stage_pfmea else ''
            description = record.desc_of_operation_pfmea or ''
            record.pfmea_operation_details = f"{operation} - {stage} - {description}"

class PfmeaOneOperationLine(models.Model):
    _name = "asd.pfmea.one.operations.line"
    _description = "PFMEA Operation Line"
    # _inherit = "translation.mixin"

    pfmea_operation_id = fields.Many2one('asd.pfmea.one.operations', 'PFMEA Operation Line')

    process_step = fields.Char("Process Step",translate=True)
    process_desc = fields.Char("Process  Description",translate=True)
    process_step_desc = fields.Char("Process Step Description", compute="_compute_process_step_desc", store=True,translate=True)
    requirement = fields.Char("Requirement",translate=True)
    potential_failure = fields.Char("Potential Failure Mode",translate=True)
    potential_effect_of_failure = fields.Char("Potential Effect of Failure",translate=True)
    severity = fields.Integer("Severity")
    class_name = fields.Char("Class",translate=True)
    potential_cause_mechanism_of_failure = fields.Char("Potential Causes",translate=True)
    prevention = fields.Char("Prevention",translate=True)
    occ = fields.Integer("OCC")
    detection_desc = fields.Char("Detection",translate=True)
    detection_no = fields.Integer("Detection_No")
    rpn = fields.Integer("RPN", compute="_compute_rpn")
    recommended_action = fields.Char("Recommended Action",translate=True)
    responsibility_target_date = fields.Date("Responsibility Target Completion Date")
    action_taken = fields.Char("Action Taken",translate=True)
    result_sev = fields.Char("S.E.V",translate=True)
    result_occ = fields.Char("O.C.C",translate=True)
    result_det = fields.Char("D.E.T",translate=True)
    result_rpn = fields.Char("R.P.N",translate=True)

    @api.depends('severity', 'occ', 'detection_no')
    def _compute_rpn(self):
        for rec in self:
            if rec.severity > 10:
                raise ValidationError("Severity Value should not exceed 10")
            if rec.occ > 10:
                raise ValidationError("OCC Value should not exceed 10")
            if rec.detection_no > 10:
                raise ValidationError("Detection Value should not exceed 10")
            else:
                rec.rpn = rec.severity * rec.occ * rec.detection_no


    @api.depends('process_step', 'process_desc')
    def _compute_process_step_desc(self):
        for record in self:
            step = record.process_step or ''
            desc = record.process_desc or ''
            record.process_step_desc = f"{step} - {desc}"
