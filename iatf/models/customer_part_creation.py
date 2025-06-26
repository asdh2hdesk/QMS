from datetime import date
from odoo import api, fields, models, _
from odoo.exceptions import ValidationError
import openpyxl
from io import BytesIO
import io
from openpyxl import Workbook
import openpyxl
import base64
from bs4 import BeautifulSoup
from openpyxl.styles import Alignment, Font, Border, Side, DEFAULT_FONT, PatternFill
from openpyxl.drawing.image import Image
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
import copy
import datetime
from PIL import ImageOps
from PIL import Image as PILImage
from bs4 import BeautifulSoup


class ProductTemplateInherited(models.Model):
    _inherit = "product.template"

    customer_part_creation_id = fields.Many2one('customer.part.creation', 'Customer part id')
    drg_no = fields.Char("Drawing Nooo.")
    drg_revision_no = fields.Char("Drawing Revision No.")
    drg_revision_date = fields.Date("Drawing Revision Date")

    customer_id = fields.Many2one('res.partner','Customer Name')
    customer_part_name = fields.Char('Customer Part Name')
    customer_part_no = fields.Char('Customer Part No.')


# this code is for integration to how to interlink the models with each other
class HisProductRevisions(models.Model):
    _name = "his.product.revisions"
    _description = "Product Revisions"
    # _rec_name = 'ref'

    customer_id = fields.Many2one(comodel_name='customer.part.creation', string='Customer Part Id')
    revision_no = fields.Char("Revision No", required=True)
    revision_date = fields.Datetime("Revision Date")
    drawing_ids = fields.One2many(
        comodel_name='his.product.revision.drawing',
        inverse_name='revision_id',
        string="Revision Drawing"
    )

    status = fields.Selection(
        [('pending', 'Pending'), ('approved', 'Approved'), ('revision', 'Revision Required'), ('rejected', 'Rejected')],
        string='Status', default='pending'
    )
    approved_manager_ids = fields.Many2many('hr.employee', string='Managers Approved By')
    prepared_manager_ids = fields.Many2many('res.users', string='Prepared Approved By')
    # approved_by = fields.Many2one('hr.employee', 'Approved_By')
    # prepared_by = fields.Many2one('hr.employee', 'prepared_By')


class ProductRevisionsDrawings(models.Model):
    _name = "his.product.revision.drawing"

    revision_id = fields.Many2one('his.product.revisions', string='Revision id')
    drawing = fields.Binary('Drawing')


class CustomerPartCreation(models.Model):
    _name = 'customer.part.creation'
    _description = 'Customer Part Creation'
    _rec_name = 'revision_no'
    _inherit = ['iatf.sign.off.members', 'translation.mixin']
    # cust_part_creation_lines = fields.One2many('customer.part.creation.line', 'cust_part_cre_id', "Customer Part Creation Line")

    # state = fields.Selection([
    #     ('draft', 'Draft'),
    #     ('hr_approve', 'HR Approval'),
    #     ('design', 'Design'),
    #     ('engineering', 'Engineering'),
    #     ('manufacturing', 'Manufacturing'),
    #     ('quality', 'Quality'),
    #     ('top', 'Top Management'),
    #     ('final_approved', 'Final Approved')
    # ], string='Status', default='draft')

    sequence = fields.Integer(string="Sequence", default=10)
    # sl_no = fields.Integer("S.No")

    revision_ids = fields.One2many(comodel_name="his.product.revisions", inverse_name='customer_id', string="Revisions")
    part_no = fields.Char("Part Number")
    part_desc = fields.Char("Part Description",translate=True)
    no_of_cust = fields.Char("No. of Customer")
    partner_id = fields.Many2one('res.partner', "Customer Name")
    cust_part_no = fields.Char("Customer Part Number")
    cust_part_desc = fields.Char("Customer Part Description",translate=True)
    customer_drawing_upload = fields.Char("Customer Drawing upload",translate=True)
    category = fields.Char('Category',translate=True)
    date_creation = fields.Datetime("Date of Creation")
    drawing_upload_internal = fields.Html("Drawing Upload internal")
    date_of_modification = fields.Datetime("Date of Modification")
    revision_no = fields.Char("Revision No")
    hsn_code = fields.Char("HSN Code")
    revision_desc = fields.Char("Revision Description")
    multiple_customer = fields.Char("Multiple Customer")
    raw_part = fields.Char("Supplier Part weight")
    material_grade = fields.Char("Material Grade")
    material_hardness = fields.Char("Material Hardness")
    modification_by = fields.Char("Modification by")
    type_of_purchase = fields.Char("Type of Purchase")
    planned_customer = fields.Char("Type Of Package Planned with Customer")
    cost_of_part = fields.Char("Cost of Part as per Customer")
    to_be_Assigned = fields.Char("To Be assigned as per Customer")
    modification_approved_by = fields.Char("Modification Approved By")
    date_of_modification_approved = fields.Datetime("Date of Modification approved")
    previous_revision = fields.Char("Previous Revision")
    previous_revision_drawings = fields.Char("Previous Revision Drawings ")
    hr = fields.Many2one('res.users', 'HR')  # HR APPROVED
    design_eng = fields.Many2one('res.users', 'Design Engineering')  # DESIGN
    manf_eng = fields.Many2one('res.users', 'Manufacturing Engineering')  # Engineering
    production = fields.Many2one('res.users', 'Production')  # Manufacturing
    quality = fields.Many2one('res.users', 'Quality')  # Qulity
    finish_part_weight = fields.Float("Finished Part weight")
    assigned_part_number = fields.Char("Assigned to Part Number")
    top_management_id = fields.Many2one('res.users', 'Top Management')  # Final Approved
    product_id = fields.Many2one('product.template', 'Product id')
    part_development_id = fields.Many2one("part.development.process")
    generate_xls_file = fields.Binary(string="Generated file")

    state = fields.Selection([
        ('draft', 'Draft'),
        ('confirm', 'Confirmed'),
    ], string='Status', default='draft')
    link = fields.Char()

    @api.model
    def create(self, vals_list):
        if not isinstance(vals_list, list):
            vals_list = [vals_list]

        created_records = []
        for vals in vals_list:
            record = super(CustomerPartCreation, self).create(vals)

            # Prepare product data
            product_data = {
                'customer_part_creation_id': record.id,
                #for adding all part no
                'name': record.part_no or f"New Product {record.id}",
                'l10n_in_hsn_code': record.hsn_code,
                # Add other required fields for product.template here
            }

            product_rec = self.env['product.template'].create(product_data)
            record.product_id = product_rec.id
            created_records.append(record)

        return created_records if len(created_records) > 1 else created_records[0]

    # class CustomerPartCreation(models.Model):
    #     _name = 'your.model.name'  # Replace with your model name
    #
    #     # Define your fields here, including state
    #     state = fields.Selection([
    #         ('draft', 'Draft'),
    #         ('confirmed', 'Confirmed'),
    #         # Add other states as necessary
    #     ], required=True)
    def generate_excel_report(self):
        output = BytesIO()
        wb = Workbook()
        ws = wb.active

        # region Formatting data
        max_col = 19
        thin = Side(border_style='thin', color='000000')
        thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        align_center = Alignment(vertical='center', horizontal='center', wrapText=True)
        align_left = Alignment(vertical='center', horizontal='left', wrapText=True)
        align_right = Alignment(vertical='center', horizontal='right', wrapText=True)

        font_main_header = Font(name='Times New Roman', size=22, bold=True)
        font_header = Font(name='Times New Roman', size=11, bold=True)
        font_all = Font(name='Times New Roman', size=11, bold=False)

        my_clr = openpyxl.styles.colors.Color(rgb='ffa200')
        my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_clr)

        # endregion
        ws.merge_cells('A1:S1')
        ws['A1'] = 'Customer Part Creation'
        ws['A1'].font = font_main_header
        ws['A1'].alignment = align_center

        ws.row_dimensions[1].height = 75

        for i in range(ord('B'), ord('T')):
            ws.column_dimensions[chr(i)].width = 15

        header = ["Sr. No",
                  "Part Number", "Part Description",
                  "No. of Customer", "Customer Name",
                  "Customer Part Number", "Customer Part Description",
                  "Customer Drawing Upload", "Category",
                  "Date of Creation", "Drawing Upload Internal",
                  "Date of Modification", "Revision No.",
                  "Revision Description", "Modification by",
                  "Modification Approved By", "Date of Modification approved",
                  "Previous Revision", "Previous Revision Drawings"]

        # region Adding Logo for image
        if self.env.user.company_id.logo:
            max_width = 500  # Set your desired maximum width
            max_height = 100  # Set your desired maximum height
            image_data = base64.b64decode(self.env.user.company_id.logo)

            # Open the image using PIL
            image = PILImage.open(io.BytesIO(image_data))
            width, height = image.size
            aspect_ratio = width / height

            if width > max_width:
                width = max_width
                height = int(width / aspect_ratio)

            if height > max_height:
                height = max_height
                width = int(height * aspect_ratio)

            # Resize the image using PIL
            # Add space on the top and left side of the image
            padding_top = 10  # Adjust as needed
            padding_left = 10  # Adjust as needed

            resized_image = image.resize((width, height), PILImage.Resampling.LANCZOS)
            ImageOps.expand(resized_image, border=(padding_left, padding_top, 0, 0), fill='rgba(0,0,0,0)')
            img_bytes = io.BytesIO()
            resized_image.save(img_bytes, format='PNG')
            img_bytes.seek(0)
            logo_image = Image(img_bytes)
            # logo_image = Image(self.env.user.company_id.logo)
            ws.add_image(logo_image, 'A1')
        # endregion

        for index, value in enumerate(header):
            ws.cell(row=2, column=index + 1).value = value
            ws.cell(row=2, column=index + 1).font = font_header
            ws.cell(row=2, column=index + 1).alignment = align_center
            ws.cell(row=2, column=index + 1).fill = my_fill

        start_row = curr_row = 3
        mx_row = 25
        sno = 0
  # it is used for html field
        product_records = self.env['product.template'].search([])
        for rec in product_records:
            drawing_upload_internal = ''
            if rec.customer_part_creation_id.drawing_upload_internal:
                drawing_upload_internal = BeautifulSoup(rec.customer_part_creation_id.drawing_upload_internal,
                                                        'html.parser').find('p').get_text()
            ws.append(
                [
                    sno,
                    rec.name if rec.name else '',
                    rec.customer_part_creation_id.part_desc if rec.customer_part_creation_id.part_desc else '',
                    rec.customer_part_creation_id.no_of_cust if rec.customer_part_creation_id.no_of_cust else '',
                    rec.customer_part_creation_id.partner_id.name if rec.customer_part_creation_id.partner_id else '',
                    rec.customer_part_creation_id.cust_part_no if rec.customer_part_creation_id.cust_part_no else '',
                    rec.customer_part_creation_id.cust_part_desc if rec.customer_part_creation_id.cust_part_desc else '',
                    rec.customer_part_creation_id.customer_drawing_upload if rec.customer_part_creation_id.customer_drawing_upload else '',
                    rec.customer_part_creation_id.category if rec.customer_part_creation_id.category else '',
                    rec.customer_part_creation_id.date_creation.strftime("%Y-%m-%d") if rec.customer_part_creation_id.date_creation else '',
                    drawing_upload_internal if drawing_upload_internal else '',
                    rec.customer_part_creation_id.date_of_modification.strftime("%Y-%m-%d") if rec.customer_part_creation_id.date_of_modification else '',
                    rec.customer_part_creation_id.revision_no if rec.customer_part_creation_id.revision_no else '',
                    rec.customer_part_creation_id.revision_desc if rec.customer_part_creation_id.revision_desc else '',
                    rec.customer_part_creation_id.modification_by if rec.customer_part_creation_id.modification_by else '',
                    rec.customer_part_creation_id.modification_approved_by if rec.customer_part_creation_id.modification_approved_by else '',
                    rec.customer_part_creation_id.date_of_modification_approved.strftime("%Y-%m-%d") if rec.customer_part_creation_id.date_of_modification_approved else '',
                    rec.customer_part_creation_id.previous_revision if rec.customer_part_creation_id.previous_revision else '',
                    rec.customer_part_creation_id.previous_revision_drawings if rec.customer_part_creation_id.previous_revision_drawings else '',
                ]
            )
            curr_row += 1
            sno += 1

        if curr_row > mx_row:
            mx_row = curr_row

        for row in ws.iter_rows(min_row=0, max_row=mx_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        # region Save the workbook
        wb.save(output)
        output.seek(0)
        self.generate_xls_file = base64.b64encode(output.getvalue()).decode('utf-8')
        # endregion

        return {
            "type": "ir.actions.act_url",
            "target": "self",
            "url": "/web/content?model=customer.part.creation&download=true&field=generate_xls_file&filename={filename}.xlsx&id={pid}".format(
                filename="Customer Part Creation", pid=self[0].id),
        }

    def action_generate_backpage_excel_report(self):
        # Create a new workbook
        output = BytesIO()
        wb = Workbook()
        ws = wb.active
#it is used for create the fields in excel report according to row and column
        data = {
            'A1': 'PART NUMBER',
            'A2': 'PART NAME',
            'A3': 'CUSTOMER NAME',
            'A4': 'MULTIPLE CUSTOMERS',
            'A5': 'CUSTOMER PART NUMBER',
            'A6': 'CUSTOMER PART DESCRIPTION',
            'A7': 'CATEGORY',
            'A8': 'ROUTING REQUIRED FOR PART',
            'A9': 'FINISHED PART WEIGHT',
            'A10': 'RAW PART WEIGHT',
            'A11': 'SUPPLIER RAW PART WEIGHT',
            'A12': 'MATERIAL GRADE',
            'A13': 'MATERIAL HARDNESS',
            'A14': 'HSN CODE',
            'A15': 'TYPE OF PURCHASE',
            'A16': 'TYPE OF PACKAGING PLANNED WITH CUSTOMER ',
            'A17': 'ASSIGNED TO PART NUMBER',
            'A18': 'COST OF PART AS PER CUSTOMER',
            'A19': 'TO BE ASSIGNED TO THE PROCESS',

        }
        # Fill data into specific cells using key-value pairs
        for cell, value in data.items():
            ws[cell] = value

        thin = Side(border_style='thin', color='000000')
        font_header = Font(name='Arial', size=10, bold=False)
        align_left = Alignment(vertical='center', horizontal='left', wrapText=True)
        align_left = Alignment(vertical='center', horizontal='center', wrapText=True)

        # region merging and formatting cells
        max_col = 2
        max_row = 19
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                alignment = copy.copy(cell.alignment)
                alignment.wrapText = True
                cell.alignment = align_left
                cell.font = font_header

                # NEWLY ADDED
                alignment_style = Alignment(vertical='center', horizontal='center')
                font_style_11_bold = Font(name='Arial', size=10)
                fill_style = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

                # font_style_10_bold = Font(name='Arial', size=10, bold=True)

                # Function to apply styles to a cell
                def apply_styles(ws, cell, alignment, font, fill):
                    ws[cell].alignment = alignment
                    ws[cell].font = font
                    ws[cell].fill = fill

                # Apply styles to cells
                cells_to_style_11_bold = ['A1', 'A2', 'A3', 'A4', 'A5', 'A6', 'A7', 'A8', 'A9', 'A10', 'A11', 'A12',
                                          'A13', 'A14', 'A15', 'A16', 'A17', 'A18', 'A19']

                for cell in cells_to_style_11_bold:
                    apply_styles(ws, cell, alignment_style, font_style_11_bold, fill_style)

        # # Apply the font style to a rectangular range
        # for col in range(4, 35):
        #     cell = ws.cell(row=3, column=col)
        #     cell.font = font_style

        # Merging the cells as per standard sheet

        # Dimension of Columns
        ws.column_dimensions['A'].width = 45
        ws.column_dimensions['B'].width = 40

#it is used for automatically transfer date from form to excel report without confirming
        for rec in self:
            ws['B1'] = rec.part_no if rec.part_no else ''
            ws['B2'] = rec.part_desc if rec.part_desc else ''
            ws['B3'] = rec.partner_id.name if rec.partner_id else ''
            ws['B4'] = rec.multiple_customer if rec.multiple_customer else ''
            ws['B5'] = rec.cust_part_no if rec.cust_part_no else ''
            ws['B6'] = rec.cust_part_desc if rec.cust_part_desc else ''
            ws['B7'] = rec.category if rec.category else ''
            ws['B8'] = rec.customer_drawing_upload if rec.customer_drawing_upload else ''
            ws['B9'] = rec.finish_part_weight if rec.finish_part_weight else '0'
            if rec.drawing_upload_internal:
                soup = BeautifulSoup(rec.drawing_upload_internal, 'html.parser')
                text_content = soup.find('p').get_text()
                if text_content:
                    ws[f'B10'] = text_content

            ws['B11'] = rec.raw_part if rec.raw_part else ''
            ws['B12'] = rec.material_grade if rec.material_grade else ''
            ws['B13'] = rec.material_hardness if rec.material_hardness else ''
            ws['B14'] = rec.hsn_code if rec.hsn_code else ''
            ws['B15'] = rec.type_of_purchase if rec.type_of_purchase else ''
            ws['B16'] = rec.planned_customer if rec.planned_customer else ''
            ws['B17'] = rec.assigned_part_number if rec.assigned_part_number else ''
            ws['B18'] = rec.cost_of_part if rec.cost_of_part else ''
            ws['B19'] = rec.to_be_Assigned if rec.to_be_Assigned else ''



        # Save the workbook
        wb.save(output)
        output.seek(0)
        self.generate_xls_file = base64.b64encode(output.getvalue()).decode('utf-8')
        # endregion

        return {
            "type": "ir.actions.act_url",
            "target": "self",
            "url": "/web/content?model=customer.part.creation&download=true&field=generate_xls_file&filename={filename}.xlsx&id={pid}".format(
                filename="Single Part (customer part creation)", pid=self[0].id),
        }



