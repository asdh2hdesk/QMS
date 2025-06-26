import base64
import io
from fnmatch import translate
from io import BytesIO

from PIL import Image as PILImage
from PIL import ImageOps
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill

from odoo import api, fields, models


class OperationNumber(models.Model):
    _name = 'operation.number'
    _rec_name= 'num'
    _inherit = "translation.mixin"

    num = fields.Char('Operation no.',translate=True)
    op_name = fields.Char('Operation Details',translate=True)


class DevelopmentCycle(models.Model):
    _name = 'development.cycle'
    _inherit = ['iatf.sign.off.members',"translation.mixin"]

    partner_id = fields.Many2one('res.partner', 'Customer')
    product_id = fields.Many2one('product.template', 'Part Name')
    default_code = fields.Char('Part Number', related='product_id.default_code')
    date = fields.Date('Final evaluation date')
    process_of_lot = fields.Char('Qty. of blanks processed for HT trial / Lot :',translate=True)
    reject_of_lot = fields.Char('Qty.of blanks/parts rejected during HT trial / Lot:',translate=True)
    ht_traile = fields.Char('HT Trial',translate=True)
    lot_develop = fields.Char('Lot development',translate=True)
    development_cycle_line_ids = fields.One2many(
        comodel_name='development.cycle.line',
        inverse_name='cycle_id',
        string="Development Cycle Line"
    )
    development_remark_line_ids = fields.One2many(
        comodel_name='development.remark.line',
        inverse_name='remark_id',
        string="Development Remark Line"
    )

    generate_xls_file = fields.Binary(string="Generated file")

    def action_development_cycle(self):
        output = BytesIO()
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1 Page1"
        ws2 = wb.create_sheet(title="Sheet2 Page2")

        if self.env.user.company_id.logo:
            max_width = 500  # Set your desired maximum width
            max_height = 100  # Set your desired maximum height
            image_data = base64.b64decode(self.env.user.company_id.logo)

            # Open the image using PIL
            image = PILImage.open(io.BytesIO(image_data))
            width, height = image.size
            aspect_ratio = width / height

            if width > max_width:
                width = max_width
                height = int(width / aspect_ratio)

            if height > max_height:
                height = max_height
                width = int(height * aspect_ratio)

            # Resize the image using PIL
            # Add space on the top and left side of the image
            padding_top = 10  # Adjust as needed
            padding_left = 10  # Adjust as needed

            resized_image = image.resize((width, height),PILImage.Resampling.LANCZOS)
            ImageOps.expand(resized_image, border=(padding_left, padding_top, 0, 0), fill='rgba(0,0,0,0)')
            img_bytes = io.BytesIO()
            resized_image.save(img_bytes, format='PNG')
            img_bytes.seek(0)
            logo_image = Image(img_bytes)
            # logo_image = Image(self.env.user.company_id.logo)
            ws1.add_image(logo_image, 'A1')

        border = Border(top=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'),
                        bottom=Side(style='thin'))
        font_header = Font(name='Arial', size=11, bold=True)
        align_left = Alignment(vertical='center', horizontal='left', wrapText=True)
        align_right = Alignment(vertical='center', horizontal='right', wrapText=True)
        align_center = Alignment(vertical='center', horizontal='center', wrapText=True)

        data = {
            'C1': 'Development Evaluation Sheet',
            'A2': 'Part Number',
            'A3': 'Part Name .',
            'A4': 'Customer / Project details  ',
            'A5': 'Final Evaluation Date',
            'A6': 'Qty. of blanks processed for HT trial / Lot :',
            'E2': 'Development stage',
            'G2': 'HT Trial ',
            'G4': 'Lot Development',
            'E6': 'Qty.of blanks/parts rejected during HT trial / Lot:',
            'A8': 'Sr.No.',
            'B8': 'Operation Number',
            'C8': 'Operation Details / Description',
            'D8': 'Cycle ',
            'E8': 'Self Control Sheet',
            'F8': 'Gauges',
            'G8': 'Cutting Tools',
            'H8': 'Clamping Tools',
            'I8': 'Quality ',
            'J8': 'Safety',

        }

        for cell, value in data.items():
            ws1[cell] = value

        ws1.row_dimensions[1].height = 75
        ws1.row_dimensions[8].height = 42
        for row in range(2, 6 + 1):
            ws1.row_dimensions[row].height = 25
        ws1.row_dimensions[8].height = 52

        ws1.column_dimensions['B'].width = 15
        ws1.column_dimensions['C'].width = 35
        columns_to_set_width = ['E', 'F', 'G', 'H', 'I', 'J']
        for column in columns_to_set_width:
            ws1.column_dimensions[column].width = 12

        ws1.merge_cells('C1:J1')
        merged_range = ws1['C1:J1']
        for cell in merged_range[0]:
            cell.border = border

        ws1['C1'].font = Font(name='Arial', size=22, bold=True)
        cell_addresses = ['A2', 'A3', 'A4', 'A5', 'A6', 'E2', 'G2', 'G4', 'E6']
        for address in cell_addresses:
            ws1[address].font = Font(name='Arial', size=10, bold=True)

        cell_ranges_to_merge = [
            'A2:B2', 'A3:B3', 'A4:B4', 'A5:B5', 'A6:C6', 'C2:D2', 'C3:D3', 'C4:D4', 'C5:D5', 'D6:D6', 'E2:F5', 'G2:H3',
            'G4:H5', 'A1:B1', 'I2:J3', 'I4:J5', 'E6:I6', 'J6:J6', 'A7:J7'
        ]
        for cell_range in cell_ranges_to_merge:
            ws1.merge_cells(cell_range)
            for row in ws1[cell_range]:
                for cell in row:
                    cell.border = border
                    cell.alignment = align_left

        columns_to_apply_border = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']
        for column in columns_to_apply_border:
            cell_reference = f'{column}8'
            ws1[cell_reference].border = border
            ws1[cell_reference].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws1[cell_reference].font = Font(name='Arial', size=11, bold=True)

        rows_to_apply_border = range(9, 31)
        for row in rows_to_apply_border:
            for col in range(1, ws1.max_column + 1):
                cell_reference = ws1.cell(row=row, column=col)
                cell_reference.border = border

        # Data Filling sheet 1
        cur_row = 9
        max_row = 30
        for rec in self:
            ws1['C4'] = rec.partner_id.name if rec.partner_id else ''
            ws1['C3'] = rec.product_id.id if rec.product_id else ''
            ws1['C2'] = rec.default_code if rec.default_code else ''
            ws1['C5'] = rec.date if rec.date else ''
            ws1['D6'] = rec.process_of_lot if rec.process_of_lot else ''
            ws1['J6'] = rec.reject_of_lot if rec.reject_of_lot else ''
            ws1['I2'] = rec.ht_traile if rec.ht_traile else ''
            ws1['I4'] = rec.lot_develop if rec.lot_develop else ''
            for i in rec.development_cycle_line_ids:
                ws1[f'A{cur_row}'] = i.sl_no if i.sl_no else ''
                ws1[f'B{cur_row}'] = i.operation_num.num if i.operation_num.num else ''
                ws1[f'C{cur_row}'] = i.description if i.description else ''
                ws1[f'D{cur_row}'] = i.cycle if i.cycle else ''
                ws1[f'E{cur_row}'] = i.control_sheet if i.control_sheet else ''
                ws1[f'F{cur_row}'] = i.gauges if i.gauges else ''
                ws1[f'G{cur_row}'] = i.cutting_tool if i.cutting_tool else ''
                ws1[f'H{cur_row}'] = i.clamping_tool if i.clamping_tool else ''
                ws1[f'I{cur_row}'] = i.quality if i.quality else ''
                ws1[f'J{cur_row}'] = i.safety if i.safety else ''

                for row_no in ws1.iter_rows(min_row=cur_row, max_row=cur_row, min_col=1, max_col=10):
                    for cell in row_no:
                        cell.border = border
                        cell.alignment = align_center
                cur_row += 1

        if cur_row < max_row:
            cur_row = max_row

        # region SignOff Members Footer for Sheet Page1
        ws1.merge_cells(f'A{cur_row}:J{cur_row}')
        cur_row += 1
        sign_row = cur_row
        ws1[f'A{cur_row}'] = 'Prepared By'
        ws1[f'A{cur_row}'].font = font_header
        ws1.merge_cells(f'A{cur_row}:B{cur_row}')
        ws1.merge_cells(f'C{cur_row}:E{cur_row}')

        ws1[f'F{cur_row}'] = 'Prepared Date'
        ws1[f'F{cur_row}'].font = font_header
        ws1.merge_cells(f'F{cur_row}:G{cur_row}')
        ws1.merge_cells(f'H{cur_row}:J{cur_row}')

        for rec in self:
            ws1[f'C{cur_row}'] = rec.create_uid.name if rec.create_uid else ''
            ws1[f'H{cur_row}'] = rec.create_date if rec.create_date else ''

        ws1.row_dimensions[cur_row].height = 18
        cur_row += 1
        ws1.merge_cells(f'A{cur_row}:J{cur_row}')
        cur_row += 1
        ws1.merge_cells(f'A{cur_row}:J{cur_row}')
        ws1[f'A{cur_row}'] = 'Sign OFF'
        ws1[f'A{cur_row}'].font = Font(size=18, bold=True)
        ws1.row_dimensions[cur_row].height = 25
        cur_row += 1
        ws1.merge_cells(f'A{cur_row}:J{cur_row}')
        cur_row += 1
        for cell in ws1[cur_row]:
            ws1[f'A{cur_row}'] = 'Member'
            ws1[f'C{cur_row}'] = 'Department'
            ws1[f'D{cur_row}'] = 'Status'
            ws1[f'F{cur_row}'] = 'Date'
            ws1[f'H{cur_row}'] = 'Comments'
            ws1.row_dimensions[cur_row].height = 25
            ws1.merge_cells(f'A{cur_row}:B{cur_row}')
            ws1.merge_cells(f'D{cur_row}:E{cur_row}')
            ws1.merge_cells(f'F{cur_row}:G{cur_row}')
            ws1.merge_cells(f'H{cur_row}:J{cur_row}')
            cell.font = Font(size=12, bold=True, color='ffffff')
            cell.fill = PatternFill(start_color='0070C0', end_color='0070C0', fill_type="solid")

        for row_no in ws1.iter_rows(min_row=1, max_row=cur_row + 1, min_col=1, max_col=10):
            for cell in row_no:
                cell.border = border
                cell.alignment = align_center

        # Listing Managers Id
        for rec in self:
            for record in rec.iatf_members_ids:
                name_rec = record.approver_id.name
                dept_rec = record.department_id.name if record.department_id.name else ''
                status_rec = record.approval_status.capitalize()
                date_rec = record.date_approved_rejected if record.date_approved_rejected else ''
                comment_rec = record.comment if record.comment else ''

                ws1[f'A{cur_row + 1}'] = name_rec
                ws1[f'C{cur_row + 1}'] = dept_rec
                ws1[f'D{cur_row + 1}'] = status_rec
                ws1[f'F{cur_row + 1}'] = date_rec
                ws1[f'H{cur_row + 1}'] = comment_rec
                ws1.merge_cells(f'A{cur_row + 1}:B{cur_row + 1}')
                ws1.merge_cells(f'D{cur_row + 1}:E{cur_row + 1}')
                ws1.merge_cells(f'F{cur_row + 1}:G{cur_row + 1}')
                ws1.merge_cells(f'H{cur_row + 1}:J{cur_row + 1}')

                status_dict = {'Approved': ['CFFFC3', '000000'], 'Rejected': ['FFCDCD', '000000'],
                               'Revision': ['AFEFFF', '000000'], 'Pending': ['FDFFCD', '000000']}

                for state, color in status_dict.items():
                    if state == status_rec:
                        ws1[f'D{cur_row + 1}'].fill = PatternFill(start_color=color[0], end_color=color[0],
                                                                  fill_type="solid")
                        ws1[f'D{cur_row + 1}'].font = Font(size=12, bold=False, color=color[1])

                cur_row += 1

            ws1.merge_cells(f'A{cur_row + 1}:J{cur_row + 1}')

            for row_no in ws1.iter_rows(min_row=sign_row + 1, max_row=cur_row + 1, min_col=1, max_col=10):
                for cell in row_no:
                    cell.border = border
                    cell.alignment = align_center
        # endregion fot fo

        # region for Sheet 2 page2
        sheet_two_data = {
            'C1': 'Development Evaluation Sheet',
            'A3': 'Part Number',
            'A4': 'Part Name .',
            'A5': 'Customer / Project Details  ',
            'A6': 'Final Evaluation Date',
            'A8': 'Sl.No',
            'B8': 'Operation Number',
            'C8': 'Evaluation Remarks',
            'H8': 'Responsibility',
            'I8': 'Target Date',
            'G3': 'Development Stage',
            'H3': 'HT Trial ',
            'H5': 'Lot Development',
        }

        for cell, value in sheet_two_data.items():
            ws2[cell] = value

        for row_no in ws2.iter_rows(min_row=1, max_row=39, min_col=1, max_col=9):
            for cell in row_no:
                cell.border = border
                cell.alignment = align_center

        cell_ranges_to_merge = [
            'A1:B2', 'C1:I2', 'A3:B3', 'A4:B4', 'A5:B5', 'A6:B6', 'C3:F3', 'C4:F4', 'C5:F5', 'C6:F6', 'G3:G6', 'H3:H4',
            'H5:H6', 'I3:I4', 'I5:I6', 'A8:A9', 'C8:G9', 'H8:H9', 'I8:I9','A7:I7','B8:B9',
        ]

        for cell_range in cell_ranges_to_merge:
            ws2.merge_cells(cell_range)
        for i in range(10,40):
            ws2.merge_cells(f'C{i}:G{i}')

        ws2['C3'].value = "='Sheet1 Page1'!C2"
        ws2['C4'].value = "='Sheet1 Page1'!C3"
        ws2['C5'].value = "='Sheet1 Page1'!C4"
        ws2['C6'].value = "='Sheet1 Page1'!C5"
        ws2['I3'].value = "='Sheet1 Page1'!I2"
        ws2['I5'].value = "='Sheet1 Page1'!I4"

        cell_range_apply_font = ['A3', 'A4', 'A5', 'A6', 'G3', 'H3', 'H5', 'I3', 'I5', 'A8', 'B8', 'H8', 'I8', 'C1']
        for cell_range in cell_range_apply_font:
            ws2[cell_range].font = Font(name='Arial', size=10, bold=True)

        cell_range_apply_alignment_center = ['C1', 'G3', 'H3', 'H5', 'I3', 'I5', 'H8', 'I8']
        for cell_range in cell_range_apply_alignment_center:
            ws2[cell_range].alignment = Alignment(horizontal='center', vertical='center')

        ws2['C1'].font = Font(name='Arial', size=22, bold=True)
        ws2['A8'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws2['B8'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws2['A40'].alignment = Alignment(horizontal='right', vertical='center')

        ws2.row_dimensions[1].height = 75
        ws2.row_dimensions[2].height = 20
        ws2.row_dimensions[3].height = 30
        ws2.row_dimensions[4].height = 30
        ws2.row_dimensions[5].height = 30
        ws2.row_dimensions[6].height = 30

        ws2.column_dimensions['A'].width = 15
        ws2.column_dimensions['B'].width = 15
        ws2.column_dimensions['C'].width = 10
        ws2.column_dimensions['D'].width = 10
        ws2.column_dimensions['E'].width = 10
        ws2.column_dimensions['F'].width = 10
        ws2.column_dimensions['G'].width = 20
        ws2.column_dimensions['H'].width = 25
        ws2.column_dimensions['I'].width = 25

        # endregion

        # Data Filling sheet 2
        max_row = 40
        cur_row = 10
        for rec in self:
            for i in rec.development_remark_line_ids:
                ws2[f'A{cur_row}'] = i.sl_no if i.sl_no else ''
                ws2[f'B{cur_row}'] = i.operation_num.num if i.operation_num.num else ''
                ws2[f'C{cur_row}'] = i.evaluation if i.evaluation else ''
                ws2[f'H{cur_row}'] = i.responsibility.name if i.responsibility.name else ''
                ws2[f'I{cur_row}'] = i.target_date if i.target_date else ''

                ws2.merge_cells(f'C{cur_row}:G{cur_row}')
                for row_no in ws2.iter_rows(min_row=cur_row, max_row=cur_row, min_col=1, max_col=9):
                    for cell in row_no:
                        cell.border = border
                        cell.alignment = align_center

                cur_row += 1

        if cur_row < max_row:
            cur_row = max_row

        # region SignOff Members Footer
        ws2.merge_cells(f'A{cur_row}:I{cur_row}')
        cur_row += 1
        sign_row = cur_row
        ws2[f'A{cur_row}'] = 'Prepared By'
        ws2[f'A{cur_row}'].font = font_header
        ws2.merge_cells(f'A{cur_row}:B{cur_row}')
        ws2.merge_cells(f'C{cur_row}:E{cur_row}')

        ws2[f'F{cur_row}'] = 'Prepared Date'
        ws2[f'F{cur_row}'].font = font_header
        ws2.merge_cells(f'F{cur_row}:G{cur_row}')
        ws2.merge_cells(f'H{cur_row}:I{cur_row}')

        for rec in self:
            ws2[f'C{cur_row}'] = rec.create_uid.name if rec.create_uid else ''
            ws2[f'H{cur_row}'] = rec.create_date if rec.create_date else ''

        ws2.row_dimensions[cur_row].height = 18
        cur_row += 1
        ws2.merge_cells(f'A{cur_row}:I{cur_row}')
        cur_row += 1
        ws2.merge_cells(f'A{cur_row}:I{cur_row}')
        ws2[f'A{cur_row}'] = 'Sign OFF'
        ws2[f'A{cur_row}'].font = Font(size=18, bold=True)
        ws2.row_dimensions[cur_row].height = 25
        cur_row += 1
        ws2.merge_cells(f'A{cur_row}:I{cur_row}')
        cur_row += 1
        for cell in ws2[cur_row]:
            ws2[f'A{cur_row}'] = 'Member'
            ws2[f'C{cur_row}'] = 'Department'
            ws2[f'E{cur_row}'] = 'Status'
            ws2[f'G{cur_row}'] = 'Date'
            ws2[f'H{cur_row}'] = 'Comments'
            ws2.row_dimensions[cur_row].height = 25
            ws2.merge_cells(f'A{cur_row}:B{cur_row}')
            ws2.merge_cells(f'C{cur_row}:D{cur_row}')
            ws2.merge_cells(f'E{cur_row}:F{cur_row}')
            ws2.merge_cells(f'H{cur_row}:I{cur_row}')
            cell.font = Font(size=12, bold=True, color='ffffff')
            cell.fill = PatternFill(start_color='0070C0', end_color='0070C0', fill_type="solid")

        for row_no in ws2.iter_rows(min_row=1, max_row=cur_row + 1, min_col=1, max_col=9):
            for cell in row_no:
                cell.border = border
                cell.alignment = align_center
        # endregion

        # Listing Managers Id
        for rec in self:
            for record in rec.iatf_members_ids:
                name_rec = record.approver_id.name
                dept_rec = record.department_id.name if record.department_id.name else ''
                status_rec = record.approval_status.capitalize()
                date_rec = record.date_approved_rejected if record.date_approved_rejected else ''
                comment_rec = record.comment if record.comment else ''

                ws2[f'A{cur_row + 1}'] = name_rec
                ws2[f'C{cur_row + 1}'] = dept_rec
                ws2[f'E{cur_row + 1}'] = status_rec
                ws2[f'G{cur_row + 1}'] = date_rec
                ws2[f'H{cur_row + 1}'] = comment_rec
                ws2.merge_cells(f'A{cur_row + 1}:B{cur_row + 1}')
                ws2.merge_cells(f'C{cur_row + 1}:D{cur_row + 1}')
                ws2.merge_cells(f'E{cur_row + 1}:F{cur_row + 1}')
                ws2.merge_cells(f'H{cur_row + 1}:I{cur_row + 1}')

                status_dict = {'Approved': ['CFFFC3', '000000'], 'Rejected': ['FFCDCD', '000000'],
                               'Revision': ['AFEFFF', '000000'], 'Pending': ['FDFFCD', '000000']}

                for state, color in status_dict.items():
                    if state == status_rec:
                        ws2[f'E{cur_row + 1}'].fill = PatternFill(start_color=color[0], end_color=color[0],
                                                                  fill_type="solid")
                        ws2[f'E{cur_row + 1}'].font = Font(size=12, bold=False, color=color[1])

                cur_row += 1

            ws2.merge_cells(f'A{cur_row + 1}:I{cur_row + 1}')

            for row_no in ws2.iter_rows(min_row=sign_row + 1, max_row=cur_row + 1, min_col=1, max_col=9):
                for cell in row_no:
                    cell.border = border
                    cell.alignment = align_center
            # endregion
        # endregion

        wb.save(output)
        output.seek(0)
        self.generate_xls_file = base64.b64encode(output.getvalue()).decode('utf-8')

        return {
            "type": "ir.actions.act_url",
            "target": "self",
            "url": "/web/content?model=development.cycle&download=true&field=generate_xls_file&filename={filename}.xlsx&id={pid}".format(
                filename="Development Evaluation Cycle", pid=self[0].id),
        }


class DevelopmentCycleLine(models.Model):
    _name = 'development.cycle.line'
    _inherit = "translation.mixin"

    cycle_id = fields.Many2one(
        comodel_name='development.cycle',
        string='Development Cycle',
        required=True, ondelete='cascade', index=True, copy=False
    )
    sequence = fields.Integer(string="Sequence", default=10)
    sl_no = fields.Integer("S.No", compute="_compute_sequence_number")
    operation_num = fields.Many2one('operation.number', 'Operation no.')
    description = fields.Char(related='operation_num.op_name', string='Operation Details/Description',translate=True)
    # description = fields.Many2one('operation.number','Operation details / description',ondelete='cascade', index=True, copy=False)
    cycle = fields.Char('Cycle',translate=True)
    control_sheet = fields.Char('Self control sheet',translate=True)
    gauges = fields.Char('Gauges',translate=True)
    cutting_tool = fields.Char('Cutting tools',translate=True)
    clamping_tool = fields.Char('Clamping tools',translate=True)
    quality = fields.Char('Quality',translate=True)
    safety = fields.Char('Safety',translate=True)


    @api.depends('sequence', 'cycle_id')
    def _compute_sequence_number(self):
        for line in self.mapped('cycle_id'):
            sl_no = 1
            for lines in line.development_cycle_line_ids:
                lines.sl_no = sl_no
                sl_no += 1

class DevelopmentRemarksLine(models.Model):
    _name = 'development.remark.line'
    _inherit = "translation.mixin"

    remark_id = fields.Many2one(
        comodel_name='development.cycle',
        string='Development Remarks',
        required=True, ondelete='cascade', index=True, copy=False
    )
    sequence = fields.Integer(string="Sequence", default=10)
    sl_no = fields.Integer("S.No", compute="_compute_sequence_number")
    operation_num = fields.Many2one('operation.number', 'Operation no.')
    evaluation = fields.Char("Evaluation Remarks",translate=True)
    responsibility = fields.Many2one("hr.employee", "Responsibility")
    target_date = fields.Date("Target Date")

    @api.depends('sequence', 'remark_id')
    def _compute_sequence_number(self):
        for line in self.mapped('remark_id'):
            sl_no = 1
            for lines in line.development_remark_line_ids:
                lines.sl_no = sl_no
                sl_no += 1
