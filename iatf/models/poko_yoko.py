import base64
import io
from odoo.tools.translate import _
from io import BytesIO

from PIL import Image as PILImage
from PIL import ImageOps
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill

from odoo import fields, models
from odoo import api
from odoo.exceptions import ValidationError


class PokaYoke(models.Model):
    _name = 'poka.yoka'
    _inherit = ['iatf.sign.off.members']

    poka_yoke_ids = fields.One2many('poka.yoka.line', 'reference_id', string='Poka Yoka Lines')

    generate_xls_file = fields.Binary(string="Generated file")

    def action_poka_yoke(self):
        output = BytesIO()
        wb = Workbook()
        ws = wb.active

        if self.env.user.company_id.logo:
            max_width = 500  # Set your desired maximum width
            max_height = 100  # Set your desired maximum height
            image_data = base64.b64decode(self.env.user.company_id.logo)

            # Open the image using PIL
            image = PILImage.open(io.BytesIO(image_data))
            width, height = image.size
            aspect_ratio = width / height

            if width > max_width:
                width = max_width
                height = int(width / aspect_ratio)

            if height > max_height:
                height = max_height
                width = int(height * aspect_ratio)

            # Resize the image using PIL
            # Add space on the top and left side of the image
            padding_top = 10  # Adjust as needed
            padding_left = 10  # Adjust as needed

            resized_image = image.resize((width, height), PILImage.Resampling.LANCZOS)
            ImageOps.expand(resized_image, border=(padding_left, padding_top, 0, 0), fill='rgba(0,0,0,0)')
            img_bytes = io.BytesIO()
            resized_image.save(img_bytes, format='PNG')
            img_bytes.seek(0)
            logo_image = Image(img_bytes)
            # logo_image = Image(self.env.user.company_id.logo)
            ws.add_image(logo_image, 'A1')

        border = Border(top=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'),
                        bottom=Side(style='thin'))
        font_header = Font(name='Times New Roman', size=11, bold=True)
        align_left = Alignment(vertical='center', horizontal='left', wrapText=True)
        align_center = Alignment(vertical='center', horizontal='center', wrapText=True)

        data = {
            'A1': 'LIST OF POKA-YOKE AND THEIR EFFECTIVENESS MONITORING PLAN',
            'O1': 'Month & Year : ',
            'A2': 'Sr No.',
            'B2': 'Component',
            'C2': 'Customer',
            'D2': 'Operation',
            'E2': 'Failure Mode',
            'F2': 'Generating work station',
            'G2': 'Detecting work station',
            'H2': 'Poka Yoke description',
            'I2': 'Poka Yoke Number',
            'J2': 'fixture photograph',
            'K2': 'Prevention / Detection',
            'L2': 'Section',
            'M2': 'Check point in Fixture',
            'N2': 'Responsibility',
            'O2': 'Frequency of Validation',
            'P2': 'Alternate Method /Reaction Plan when Mistake Proofing Fails',
            'Q2': '1',
            'R2': '2',
            'S2': '3',
            'T2': '4',
            'U2': '5',
            'V2': '6',
            'W2': '7',
            'X2': '8',
            'Y2': '9',
            'Z2': '10',
            'AA2': '11',
            'AB2': '12',
        }
        # Fill data into specific cells using key-value pairs
        for cell, value in data.items():
            ws[cell] = value

        # region merging and formatting cells
        max_col = 28
        max_row = 28
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.border = border
                cell.alignment = align_center
                cell.font = font_header

        # Specific Dimension
        ws['A1'].alignment = align_center
        ws['A1'].font = Font(name='Times New Roman', size=24, bold=True)

        # Merging the cells as per standard sheet
        ws.merge_cells('A1:N1')
        ws.merge_cells('P1:AB1')

        # Dimension of Columns
        ws.column_dimensions['A'].width = 8.5
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 33
        ws.column_dimensions['D'].width = 16
        ws.column_dimensions['E'].width = 32
        ws.column_dimensions['F'].width = 16
        ws.column_dimensions['G'].width = 16
        ws.column_dimensions['H'].width = 40
        ws.column_dimensions['I'].width = 42
        ws.column_dimensions['J'].width = 40
        ws.column_dimensions['K'].width = 20
        ws.column_dimensions['L'].width = 25
        ws.column_dimensions['M'].width = 19
        ws.column_dimensions['N'].width = 20
        ws.column_dimensions['O'].width = 15
        ws.column_dimensions['P'].width = 37
        ws.column_dimensions['Q'].width = 9
        ws.column_dimensions['R'].width = 9
        ws.column_dimensions['S'].width = 9
        ws.column_dimensions['T'].width = 9
        ws.column_dimensions['U'].width = 9
        ws.column_dimensions['V'].width = 9
        ws.column_dimensions['W'].width = 9
        ws.column_dimensions['X'].width = 9
        ws.column_dimensions['Y'].width = 9
        ws.column_dimensions['Z'].width = 9
        ws.column_dimensions['AA'].width = 9
        ws.column_dimensions['AB'].width = 9

        # Dimension of Rows
        ws.row_dimensions[1].height = 75
        ws.row_dimensions[2].height = 50
        # for row_cell in range(3, 28 + 1):
        #     ws.row_dimensions[row_cell].height = 120

        # Data Filling
        cur_row = 3
        for record in self:
            for rec in record.poka_yoke_ids:
                ws['P1'] = rec.date if rec.date else ''
                ws[f'A{cur_row}'] = str(cur_row - 2)
                ws[f'B{cur_row}'] = rec.product_id.name if rec.product_id.name else ''
                ws[f'C{cur_row}'] = rec.partner_id.name if rec.partner_id.name else ''
                ws[f'D{cur_row}'] = rec.operation if rec.operation else ''
                ws[f'E{cur_row}'] = rec.failer_mode.name if rec.failer_mode.name else ''
                ws[f'F{cur_row}'] = rec.genrationg_work_station if rec.genrationg_work_station else ''
                ws[f'G{cur_row}'] = rec.detecting_work_station if rec.detecting_work_station else ''
                ws[f'H{cur_row}'] = rec.poka_yoke_description if rec.poka_yoke_description else ''
                ws[f'I{cur_row}'] = rec.poka_yoke_number if rec.poka_yoke_number else ''
                ws[f'J{cur_row}'] = rec.fixture_photography if rec.fixture_photography else ''
                ws[f'K{cur_row}'] = rec.detecation if rec.detecation else ''
                ws[f'L{cur_row}'] = rec.section if rec.section else ''
                ws[f'M{cur_row}'] = rec.fixture if rec.fixture else ''
                ws[f'N{cur_row}'] = rec.responsible_id.name if rec.responsible_id.name else ''
                ws[f'O{cur_row}'] = rec.frequency_of_validation if rec.frequency_of_validation else ''
                ws[f'P{cur_row}'] = rec.alternate_method.name if rec.alternate_method.name else ''

                if rec.selection:
                    cell = ord("P")
                    month = cell + int(rec.selection)
                    if month <= 90:
                        ws[f'{str(chr(month))}{cur_row}'] = '☑'
                        ws[f'{str(chr(month))}{cur_row}'].font = Font(size=25, color='05a61d', bold=True)
                        ws[f'{str(chr(month))}{cur_row}'].fill = PatternFill(start_color="d0d4d6", end_color="d0d4d6",
                                                                             fill_type="solid")
                    else:
                        ws[f'AA{cur_row}'] = '☑'
                        ws[f'AA{cur_row}'].font = Font(size=25, color='05a61d', bold=True)
                        ws[f'AA{cur_row}'].fill = PatternFill(start_color="d0d4d6", end_color="d0d4d6",
                                                              fill_type="solid")

                # Set borders for the current row
                for cell in ws.iter_rows(min_row=cur_row, max_row=cur_row, min_col=1, max_col=28):
                    for c in cell:
                        c.border = border
                        c.alignment = align_center

                # Increment row counter for EACH poka yoke line
                cur_row += 1

            if cur_row < max_row:
                cur_row = max_row + 1

            # region SignOff Members Footer
            sign_row = cur_row
            ws.merge_cells(f'A{cur_row}:AA{cur_row}')
            cur_row += 1
            mer_end = cur_row
            ws[f'A{cur_row}'] = 'Prepared By'
            ws[f'A{cur_row}'].font = font_header
            ws.merge_cells(f'A{cur_row}:B{cur_row}')
            ws.merge_cells(f'C{cur_row}:D{cur_row}')

            ws[f'E{cur_row}'] = 'Prepared Date'
            ws[f'E{cur_row}'].font = font_header
            ws.merge_cells(f'E{cur_row}:F{cur_row}')
            ws.merge_cells(f'G{cur_row}:H{cur_row}')

            for rec in self:
                ws[f'C{cur_row}'] = rec.create_uid.name if rec.create_uid else ''
                ws[f'G{cur_row}'] = rec.create_date if rec.create_date else ''

            ws.row_dimensions[cur_row].height = 18
            cur_row += 1
            ws.merge_cells(f'A{cur_row}:H{cur_row}')
            cur_row += 1
            ws.merge_cells(f'A{cur_row}:H{cur_row}')
            ws[f'A{cur_row}'] = 'Sign OFF'
            ws[f'A{cur_row}'].font = Font(size=18, bold=True)
            ws.row_dimensions[cur_row].height = 25
            cur_row += 1
            ws.merge_cells(f'A{cur_row}:H{cur_row}')
            cur_row += 1
            for cell in ws[cur_row]:
                ws[f'A{cur_row}'] = 'Member'
                ws[f'C{cur_row}'] = 'Department'
                ws[f'D{cur_row}'] = 'Status'
                ws[f'E{cur_row}'] = 'Date'
                ws[f'F{cur_row}'] = 'Comments'
                ws.row_dimensions[cur_row].height = 25
                ws.merge_cells(f'A{cur_row}:B{cur_row}')
                ws.merge_cells(f'F{cur_row}:H{cur_row}')
                cell.font = Font(size=12, bold=True, color='ffffff')
                cell.fill = PatternFill(start_color='0070C0', end_color='0070C0', fill_type="solid")

            for row_no in ws.iter_rows(min_row=1, max_row=cur_row + 1, min_col=1, max_col=8):
                for cell in row_no:
                    cell.border = border
                    cell.alignment = align_center

            # Listing Managers Id
            for rec in self:
                for record in rec.iatf_members_ids:
                    name_rec = record.approver_id.name
                    dept_rec = record.department_id.name if record.department_id.name else ''
                    status_rec = record.approval_status.capitalize()
                    date_rec = record.date_approved_rejected if record.date_approved_rejected else ''
                    comment_rec = record.comment if record.comment else ''

                    ws[f'A{cur_row + 1}'] = name_rec
                    ws[f'C{cur_row + 1}'] = dept_rec
                    ws[f'D{cur_row + 1}'] = status_rec
                    ws[f'E{cur_row + 1}'] = date_rec
                    ws[f'F{cur_row + 1}'] = comment_rec
                    ws.merge_cells(f'A{cur_row + 1}:B{cur_row + 1}')
                    ws.merge_cells(f'F{cur_row + 1}:H{cur_row + 1}')

                    status_dict = {'Approved': ['CFFFC3', '000000'], 'Rejected': ['FFCDCD', '000000'],
                                   'Revision': ['AFEFFF', '000000'], 'Pending': ['FDFFCD', '000000']}

                    for state, color in status_dict.items():
                        if state == status_rec:
                            ws[f'D{cur_row + 1}'].fill = PatternFill(start_color=color[0], end_color=color[0],
                                                                     fill_type="solid")
                            ws[f'D{cur_row + 1}'].font = Font(size=12, bold=False, color=color[1])

                    cur_row += 1

                ws.merge_cells(f'A{cur_row + 1}:AA{cur_row + 1}')
                ws.merge_cells(f'I{mer_end}:AA{cur_row}')

                for row_no in ws.iter_rows(min_row=sign_row - 1, max_row=cur_row + 1, min_col=1, max_col=27):
                    for cell in row_no:
                        cell.border = border
                        cell.alignment = align_center
                # endregion

        wb.save(output)
        output.seek(0)
        self.generate_xls_file = base64.b64encode(output.getvalue()).decode('utf-8')

        return {
            "type": "ir.actions.act_url",
            "target": "self",
            "url": "/web/content?model=poka.yoka&download=true&field=generate_xls_file&filename={filename}.xlsx&id={pid}".format(
                filename="LIST OF POKA-YOKE AND THEIR EFFECTIVENESS MONITORING PLAN", pid=self[0].id),
        }

    @api.model_create_multi
    def create(self, vals_list):
        records = super(PokaYoke, self).create(vals_list)
        for record in records:
            pm_records = self.env['process.group'].search(
                [('project_id', '=', record.project_id.id),
                 ('final_status', '=', 'approved')
                 ])

            for operation in pm_records.process_presentation_ids.sorted(key="operation"):
                for poka_yoke_element in operation.operation_lines_ids:
                    for poka_yoke in poka_yoke_element.pokayoke_ids:
                        self.env['poka.yoka.line'].create({
                            'reference_id': record.id,
                            'operation': operation.operation,
                            'poka_yoke_description': poka_yoke.poka_yoke_description,
                            'poka_yoke_number': poka_yoke.poka_yoke_number,
                        })
        return records


class PokaYokeLine(models.Model):
    _name = 'poka.yoka.line'
    _description = 'Poka Yoka Line'
    _inherit = "translation.mixin"

    _rec_name = 'poka_yoke_description'
    reference_id = fields.Many2one('poka.yoka', string='Reference Poka Yoke')
    team_id = fields.Many2one('maintenance.team', string='Team', ondelete="restrict")
    partner_id = fields.Many2one('res.partner', 'Customer')
    product_id = fields.Many2one('product.template', 'Part Name')
    operation = fields.Char('Operation',translate=True)
    date = fields.Date("Date")
    failer_mode = fields.Many2one('failer.mode', 'Failure Mode')
    genrationg_work_station = fields.Char('Generating work station',translate=True)
    detecting_work_station = fields.Char('Detecting work station',translate=True)
    poka_yoke_description = fields.Char('Poka Yoke description',translate=True)
    poka_yoke_number = fields.Char(string='Poka Yoke Number',translate=True)
    fixture_photography = fields.Html('fixture photograph',translate=True)
    detecation = fields.Char("Prevention /Detection",translate=True)
    section = fields.Char('Section',translate=True)
    fixture = fields.Char('Check point in Fixture',translate=True)
    responsible_id = fields.Many2one('hr.employee', 'Responsibility')
    frequency_of_validation = fields.Char('Frequency of Validation', default='default AS per P-80027', readonly=True,translate=True)
    alternate_method = fields.Many2one('alternate.method',
                                       'Alternate Method /Reaction Plan when Mistake Proofing Fails')
    selection = fields.Selection(
        [('1', '1'), ('2', '2'), ('3', '3'), ('4', '4'), ('5', '5'), ('6', '6'), ('7', '7'), ('8', '8'), ('9', '9'),
         ('10', '10'), ('11', '11'), ('12', '12')], string='Selection', default='1')

    #
    # state = fields.Selection([
    #                         ('draft', 'Draft'),
    #                         ('confirm', 'Confirmed'),
    #                         ], string='Status', default='draft')
    # document_name = fields.Char(string='Document #')
    # approve_department_id = fields.Many2one('hr.department', string='Departments Approvals')
    # # document_pro_id = fields.Many2one("xf.doc.approval.document.package", string="Document #")
    # approve_department_ids = fields.Many2many('hr.department', 'ta7', string='Departments Approvals')
    # approve_by_department_ids = fields.Many2many('hr.department', string='Departments Approved By')
    # manager_id = fields.Many2one('hr.employee', string='Approval Manager')
    # manager_ids = fields.Many2many('hr.employee', 'ta8', string='Approval Managers')
    # approvaed_manager_ids = fields.Many2many('hr.employee', string='Managers Approved By')    #  till here

    # is_hr_approved = fields.Boolean()
    # is_design_approved = fields.Boolean()
    # is_eng_approved = fields.Boolean()
    # is_manu_approved = fields.Boolean()
    # is_sales_approved = fields.Boolean()
    # is_qc_approved = fields.Boolean()
    # is_maintenance_approved = fields.Boolean()
    # is_marketing_approved = fields.Boolean()
    # is_pm_approved = fields.Boolean()
    # is_management_approved = fields.Boolean()
    # is_approved_approved = fields.Boolean()
    #
    # hr_check_department = fields.Boolean(compute='_check_department')
    # des_check_department = fields.Boolean(compute='_check_department')
    # eng_check_department = fields.Boolean(compute='_check_department')
    # manu_check_department = fields.Boolean(compute='_check_department')
    # sale_check_department = fields.Boolean(compute='_check_department')
    # qc_check_department = fields.Boolean(compute='_check_department')
    # mainte_check_department = fields.Boolean(compute='_check_department')
    # marketing_check_department = fields.Boolean(compute='_check_department')
    # pm_check_department = fields.Boolean(compute='_check_department')
    # managment_check_department = fields.Boolean(compute='_check_department')
    #
    # link = fields.Char()
    #
    # def actipn_confirmed(self):
    #     self.sent_for_approval()
    #     self.state = 'confirm'

    # def sent_for_approval(self):
    #     mail_template = self.env.ref('iatf.mom_document_approval_mail_template')
    #     model_id = self.env['ir.model'].sudo().search([('model', '=', self._name)])
    #     mail_template.write({'model_id' : model_id.id})
    #
    #     web_base_url = self.env['ir.config_parameter'].sudo().get_param('web.base.url')
    #     action_id = self.env.ref('iatf.action_poka_yoka_view', raise_if_not_found=False)
    #     link = """{}/web#id={}&view_type=form&model=self._name&action={}""".format(web_base_url,self.id,action_id.id)
    #     self.link = link
    #
    #     user_ids = self.env['hr.employee'].sudo().search([('department_id', 'in', self.approve_department_ids.ids)]).mapped('user_id')
    #     all_admin = ''
    #     for user in user_ids:
    #         if all_admin:
    #             all_admin += ',' + str(user.login)
    #         else:
    #             all_admin = str(user.login)
    #     mail_template.write({
    #             'email_to': all_admin,
    #     })
    #     mail_template.send_mail(self.id, force_send=True)

    # def _check_department(self):
    #     for rec in self:
    #         if self.env.user.employee_id.department_id.id in rec.approve_department_ids.ids and self.env.user.employee_id.department_id.department_type == 'hr':
    #             if rec.state != 'draft':
    #                 rec.hr_check_department = True
    #             else:
    #                 rec.hr_check_department = False
    #         else:
    #             rec.hr_check_department = False
    #         if self.env.user.employee_id.department_id.id in rec.approve_department_ids.ids and self.env.user.employee_id.department_id.department_type == 'design':
    #             if rec.state != 'draft':
    #                 rec.des_check_department = True
    #             else:
    #                 rec.des_check_department = False
    #         else:
    #             rec.des_check_department = False
    #         if self.env.user.employee_id.department_id.id in rec.approve_department_ids.ids and self.env.user.employee_id.department_id.department_type == 'eng':
    #             if rec.state != 'draft':
    #                 rec.eng_check_department = True
    #             else:
    #                 rec.eng_check_department = False
    #         else:
    #             rec.eng_check_department = False
    #         if self.env.user.employee_id.department_id.id in rec.approve_department_ids.ids and self.env.user.employee_id.department_id.department_type == 'manu':
    #             if rec.state != 'draft':
    #                 rec.manu_check_department = True
    #             else:
    #                 rec.manu_check_department = False
    #         else:
    #             rec.manu_check_department = False
    #         if self.env.user.employee_id.department_id.id in rec.approve_department_ids.ids and self.env.user.employee_id.department_id.department_type == 'sales':
    #             if rec.state != 'draft':
    #                 rec.sale_check_department = True
    #             else:
    #                 rec.sale_check_department = False
    #         else:
    #             rec.sale_check_department = False
    #         if self.env.user.employee_id.department_id.id in rec.approve_department_ids.ids and self.env.user.employee_id.department_id.department_type == 'qc':
    #             if rec.state != 'draft':
    #                 rec.qc_check_department = True
    #             else:
    #                 rec.qc_check_department = False
    #         else:
    #             rec.qc_check_department = False
    #         if self.env.user.employee_id.department_id.id in rec.approve_department_ids.ids and self.env.user.employee_id.department_id.department_type == 'maintenance':
    #             if rec.state != 'draft':
    #                 rec.mainte_check_department = True
    #             else:
    #                 rec.mainte_check_department = False
    #         else:
    #             rec.mainte_check_department = False
    #         if self.env.user.employee_id.department_id.id in rec.approve_department_ids.ids and self.env.user.employee_id.department_id.department_type == 'marketing':
    #             if rec.state != 'draft':
    #                 rec.marketing_check_department = True
    #             else:
    #                 rec.marketing_check_department = False
    #         else:
    #             rec.marketing_check_department = False
    #         if self.env.user.employee_id.department_id.id in rec.approve_department_ids.ids and self.env.user.employee_id.department_id.department_type == 'pm':
    #             if rec.state != 'draft':
    #                 rec.pm_check_department = True
    #             else:
    #                 rec.pm_check_department = False
    #         else:
    #             rec.pm_check_department = False
    #         if self.env.user.employee_id.department_id.id in rec.approve_department_ids.ids and self.env.user.employee_id.department_id.department_type == 'management':
    #             if rec.state != 'draft':
    #                 rec.managment_check_department = True
    #             else:
    #                 rec.managment_check_department = False
    #         else:
    #             rec.managment_check_department = False
    #
    #
    # hr_approval_date = fields.Datetime('Hr Approval Date')
    # design_approval_date = fields.Datetime('Design Approval Date')
    # eng_approval_date = fields.Datetime('Eng. Approval Date')
    # manu_approval_date = fields.Datetime('Production Approval Date')
    # sale_approval_date = fields.Datetime('Sales Approval Date')
    # qc_approval_date = fields.Datetime('QC Approval Date')
    # main_approval_date = fields.Datetime('Maintenance Approval Date')
    # mark_approval_date = fields.Datetime('Marketing Approval Date')
    # pm_approval_date = fields.Datetime('Program Management Approval Date')
    # maneg_approval_date = fields.Datetime('Top Management Approval Date')
    #
    # def hr_approval(self):
    #     self.hr_approval_date = datetime.now()
    #     self.is_hr_approved = True
    #     self.approve_by_department_ids = [(4, self.env.user.employee_id.department_id.id)]
    #     self.approvaed_manager_ids = [(4, self.env.user.employee_id.department_id.manager_id.id)]
    # def design_approval(self):
    #     self.design_approval_date = datetime.now()
    #     self.is_design_approved = True
    #     self.approve_by_department_ids = [(4, self.env.user.employee_id.department_id.id)]
    #     self.approvaed_manager_ids = [(4, self.env.user.employee_id.department_id.manager_id.id)]
    # def eng_approval(self):
    #     self.eng_approval_date = datetime.now()
    #     self.is_eng_approved = True
    #     self.approve_by_department_ids = [(4, self.env.user.employee_id.department_id.id)]
    #     self.approvaed_manager_ids = [(4, self.env.user.employee_id.department_id.manager_id.id)]
    # def manu_approval(self):
    #     self.manu_approval_date = datetime.now()
    #     self.is_manu_approved = True
    #     self.approve_by_department_ids = [(4, self.env.user.employee_id.department_id.id)]
    #     self.approvaed_manager_ids = [(4, self.env.user.employee_id.department_id.manager_id.id)]
    # def sales_approval(self):
    #     self.sale_approval_date = datetime.now()
    #     self.is_sales_approved = True
    #     self.approve_by_department_ids = [(4, self.env.user.employee_id.department_id.id)]
    #     self.approvaed_manager_ids = [(4, self.env.user.employee_id.department_id.manager_id.id)]
    # def qc_approval(self):
    #     self.qc_approval_date = datetime.now()
    #     self.is_qc_approved = True
    #     self.approve_by_department_ids = [(4, self.env.user.employee_id.department_id.id)]
    #     self.approvaed_manager_ids = [(4, self.env.user.employee_id.department_id.manager_id.id)]
    # def mainte_approval(self):
    #     self.main_approval_date = datetime.now()
    #     self.is_maintenance_approved = True
    #     self.approve_by_department_ids = [(4, self.env.user.employee_id.department_id.id)]
    #     self.approvaed_manager_ids = [(4, self.env.user.employee_id.department_id.manager_id.id)]
    # def marketing_approval(self):
    #     self.mark_approval_date = datetime.now()
    #     self.is_marketing_approved = True
    #     self.approve_by_department_ids = [(4, self.env.user.employee_id.department_id.id)]
    #     self.approvaed_manager_ids = [(4, self.env.user.employee_id.department_id.manager_id.id)]
    # def pm_approval(self):
    #     self.pm_approval_date = datetime.now()
    #     self.is_pm_approved = True
    #     self.approve_by_department_ids = [(4, self.env.user.employee_id.department_id.id)]
    #     self.approvaed_manager_ids = [(4, self.env.user.employee_id.department_id.manager_id.id)]
    # def managment_approval(self):
    #     self.maneg_approval_date = datetime.now()
    #     self.is_management_approved = True
    #     self.approve_by_department_ids = [(4, self.env.user.employee_id.department_id.id)]
    #     self.approvaed_manager_ids = [(4, self.env.user.employee_id.department_id.manager_id.id)]
    #
    #
    # def button_send_approval(self):
    #     for rec in self:
    #         if rec.hr:
    #             user_email = rec.hr.login
    #             mail_temp = self.env.ref('iatf.sending_mail_template_risk_assessment')
    #             if mail_temp:
    #                 format_name = rec.format_id.name
    #                 cust_name = rec.cust_id.name
    #                 subject = f" Request for approving Risk Assessment {format_name} for {cust_name}"
    #                 if user_email:
    #                     body = """
    #                     <div>
    #                         <p>Dear Recipient  """ + str(user_email) + """,
    #                             <br/><br/>
    #                             Please, kindly Request To accept and approved Sending Approval for the List Of Poko Yoko Checklist.
    #                         <br></br>
    #                         Thank you.
    #                     <br/>
    #                     <br/>
    #                     <div>"""
    #                     mail_temp.send_mail(self.id, email_values={
    #                         'email_from': self.env.user.login,
    #                         'email_to': user_email,
    #                         'subject': subject,
    #                         'body_html': body,
    #                     }, force_send=True)
    #
    #         rec.write({'state': 'hr_approve'})
    #         rec.part_development_id.cft_id = rec.id
    #
    # def button_hr_approval(self):
    #     for rec in self:
    #         if rec.hr:
    #             user_email = rec.hr.login
    #             mail_temp = self.env.ref('iatf.sending_mail_template_risk_assessment')
    #             if mail_temp:
    #                 format_name = rec.format_id.name
    #                 cust_name = rec.cust_id.name
    #                 subject = f" Request for approving Risk Assessment {format_name} for {cust_name}"
    #                 if user_email:
    #                     body = """
    #                     <div>
    #                         <p>Dear Recipient  """ + str(user_email) + """,
    #                             <br/><br/>
    #                             Please, kindly Request To accept and approved Sending Approval for the List Of Poko Yoko Checklist.
    #                         <br></br>
    #                         Thank you.
    #                     <br/>
    #                     <br/>
    #                     <div>"""
    #                     mail_temp.send_mail(self.id, email_values={
    #                         'email_from': self.env.user.login,
    #                         'email_to': user_email,
    #                         'subject': subject,
    #                         'body_html': body,
    #                     }, force_send=True)
    #         rec.write({'state': 'design'})
    #
    # def button_design(self):
    #     for rec in self:
    #         if rec.design_eng:
    #             user_email = rec.design_eng.login
    #             mail_temp = self.env.ref('iatf.sending_mail_template_risk_assessment')
    #             if mail_temp:
    #                 format_name = rec.format_id.name
    #                 cust_name = rec.cust_id.name
    #                 subject = f" Request for approving Risk Assessment {format_name} for {cust_name}"
    #                 if user_email:
    #                     body = """
    #                     <div>
    #                         <p>Dear Recipient  """ + str(user_email) + """,
    #                             <br/><br/>
    #                             Please, kindly Request To accept and approved Sending Approval for the List Of Poko Yoko Checklist.
    #                         <br></br>
    #                         Thank you.
    #                     <br/>
    #                     <br/>
    #                     <div>"""
    #                     mail_temp.send_mail(self.id, email_values={
    #                         'email_from': self.env.user.login,
    #                         'email_to': user_email,
    #                         'subject': subject,
    #                         'body_html': body,
    #                     }, force_send=True)
    #         rec.write({'state': 'engineering'})
    #
    # def button_engineering(self):
    #     for rec in self:
    #         if rec.manf_eng:
    #             user_email = rec.manf_eng.login
    #             mail_temp = self.env.ref('iatf.sending_mail_template_risk_assessment')
    #             if mail_temp:
    #                 format_name = rec.format_id.name
    #                 cust_name = rec.cust_id.name
    #                 subject = f" Request for approving Risk Assessment {format_name} for {cust_name}"
    #                 if user_email:
    #                     body = """
    #                     <div>
    #                         <p>Dear Recipient  """ + str(user_email) + """,
    #                             <br/><br/>
    #                             Please, kindly Request To accept and approved Sending Approval for the List Of Poko Yoko Checklist.
    #                         <br></br>
    #                         Thank you.
    #                     <br/>
    #                     <br/>
    #                     <div>"""
    #                     mail_temp.send_mail(self.id, email_values={
    #                         'email_from': self.env.user.login,
    #                         'email_to': user_email,
    #                         'subject': subject,
    #                         'body_html': body,
    #                     }, force_send=True)
    #         rec.write({'state': 'manufacturing'})
    #
    # def button_manufacturing(self):
    #     for rec in self:
    #         if rec.production:
    #             user_email = rec.production.login
    #             mail_temp = self.env.ref('iatf.sending_mail_template_risk_assessment')
    #             if mail_temp:
    #                 format_name = rec.format_id.name
    #                 cust_name = rec.cust_id.name
    #                 subject = f" Request for approving Risk Assessment {format_name} for {cust_name}"
    #                 if user_email:
    #                     body = """
    #                     <div>
    #                         <p>Dear Recipient  """ + str(user_email) + """,
    #                             <br/><br/>
    #                             Please, kindly Request To accept and approved Sending Approval for the List Of Poko Yoko Checklist.
    #                         <br></br>
    #                         Thank you.
    #                     <br/>
    #                     <br/>
    #                     <div>"""
    #                     mail_temp.send_mail(self.id, email_values={
    #                         'email_from': self.env.user.login,
    #                         'email_to': user_email,
    #                         'subject': subject,
    #                         'body_html': body,
    #                     }, force_send=True)
    #         rec.write({'state': 'quality'})
    #
    # def button_quality_test_done(self):
    #     for rec in self:
    #         if rec.quality:
    #             user_email = rec.quality.login
    #             mail_temp = self.env.ref('iatf.sending_mail_template_risk_assessment')
    #             if mail_temp:
    #                 format_name = rec.format_id.name
    #                 cust_name = rec.cust_id.name
    #                 subject = f" Request for approving Risk Assessment {format_name} for {cust_name}"
    #                 if user_email:
    #                     body = """
    #                     <div>
    #                         <p>Dear Recipient  """ + str(user_email) + """,
    #                             <br/><br/>
    #                             Please, kindly Request To accept and approved Sending Approval for the List Of Poko Yoko Checklist.
    #                         <br></br>
    #                         Thank you.
    #                     <br/>
    #                     <br/>
    #                     <div>"""
    #                     mail_temp.send_mail(self.id, email_values={
    #                         'email_from': self.env.user.login,
    #                         'email_to': user_email,
    #                         'subject': subject,
    #                         'body_html': body,
    #                     }, force_send=True)
    #         rec.write({'state': 'top'})
    #
    # def button_top_managment_final_approved(self):
    #     for rec in self:
    #         if rec.top_management_id:
    #             user_email = rec.top_management_id.login
    #             mail_temp = self.env.ref('iatf.sending_mail_template_risk_assessment')
    #             if mail_temp:
    #                 format_name = rec.format_id.name
    #                 cust_name = rec.cust_id.name
    #                 subject = f" Request for approving Risk Assessment {format_name} for {cust_name}"
    #                 if user_email:
    #                     body = """
    #                     <div>
    #                         <p>Dear Recipient  """ + str(user_email) + """,
    #                             <br/><br/>
    #                             Please, kindly Request To accept and approved Sending Approval for the List Of Poko Yoko Checklist.
    #                         <br></br>
    #                         Thank you.
    #                     <br/>
    #                     <br/>
    #                     <div>"""
    #                     mail_temp.send_mail(self.id, email_values={
    #                         'email_from': self.env.user.login,
    #                         'email_to': user_email,
    #                         'subject': subject,
    #                         'body_html': body,
    #                     }, force_send=True)
    #         rec.write({'state': 'final_approved'})




class FailerMode(models.Model):
    _name = 'failer.mode'

    name = fields.Char('Name')


class AlternateMethod(models.Model):
    _name = 'alternate.method'

    name = fields.Char('Name')
