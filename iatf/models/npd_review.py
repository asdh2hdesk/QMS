import base64
import io
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.drawing.image import Image
from PIL import ImageOps
from PIL import Image as PILImage
from bs4 import BeautifulSoup
from datetime import date, datetime
from odoo import api, fields, models


class NPDReview(models.Model):
    _name = 'npd.review'
    _inherit = ['iatf.sign.off.members']

    partner_id = fields.Many2one('res.partner', 'Name of Members')
    employee_ids = fields.Many2many('hr.employee', 'employee_category_rel_participent', 'category_id', 'emp_id',
                                    string='Name of Participants')
    npd_line_ids = fields.One2many(
        comodel_name='npd.review.line',
        inverse_name='npd_id',
        string="NPD Line"
    )
    team_id = fields.Many2one('maintenance.team', string='Team', ondelete="restrict")
    # state = fields.Selection([
    #     ('draft', 'Draft'),
    #     ('hr_approve', 'HR Approval'),
    #     ('design', 'Design'),
    #     ('engineering', 'Engineering'),
    #     ('manufacturing', 'Manufacturing'),
    #     ('quality', 'Quality'),
    #     ('top', 'Top Management'),
    #     ('final_approved', 'Final Approved')
    # ], string='Status', default='draft')
    hr = fields.Many2one('res.users', 'HR')  # HR APPROVED
    design_eng = fields.Many2one('res.users', 'Design Engineering')  # DESIGN
    manf_eng = fields.Many2one('res.users', 'Manufacturing Engineering')  # Engineering
    production = fields.Many2one('res.users', 'Production')  # Manufacturing
    quality = fields.Many2one('res.users', 'Quality')  # Qulity
    top_management_id = fields.Many2one('res.users', 'Top Management')  # Final Approved
    part_development_id = fields.Many2one("part.development.process")

    generate_xls_file = fields.Binary(string="Generated file")

    # state = fields.Selection([
    #                         ('draft', 'Draft'),
    #                         ('confirm', 'Confirmed'),
    #                         ], string='Status', default='draft')
    document_name = fields.Char(string='Document #')
    # approve_department_id = fields.Many2one('hr.department', string='Departments Approvals')
    # document_pro_id = fields.Many2one("xf.doc.approval.document.package", string="Document #")
    # approve_department_ids = fields.Many2many('hr.department', 'ta1', string='Departments Approvals')
    # approve_by_department_ids = fields.Many2many('hr.department', string='Departments Approved By')
    # manager_id = fields.Many2one('hr.employee', string='Approval Manager')
    # manager_ids = fields.Many2many('hr.employee', 'ta2', string='Approval Managers')
    # approvaed_manager_ids = fields.Many2many('hr.employee', string='Managers Approved By')
    #
    # is_hr_approved = fields.Boolean()
    # is_design_approved = fields.Boolean()
    # is_eng_approved = fields.Boolean()
    # is_manu_approved = fields.Boolean()
    # is_sales_approved = fields.Boolean()
    # is_qc_approved = fields.Boolean()
    # is_maintenance_approved = fields.Boolean()
    # is_marketing_approved = fields.Boolean()
    # is_pm_approved = fields.Boolean()
    # is_management_approved = fields.Boolean()
    # is_approved_approved = fields.Boolean()
    #
    # hr_check_department = fields.Boolean(compute='_check_department')
    # des_check_department = fields.Boolean(compute='_check_department')
    # eng_check_department = fields.Boolean(compute='_check_department')
    # manu_check_department = fields.Boolean(compute='_check_department')
    # sale_check_department = fields.Boolean(compute='_check_department')
    # qc_check_department = fields.Boolean(compute='_check_department')
    # mainte_check_department = fields.Boolean(compute='_check_department')
    # marketing_check_department = fields.Boolean(compute='_check_department')
    # pm_check_department = fields.Boolean(compute='_check_department')
    # managment_check_department = fields.Boolean(compute='_check_department')
    #
    # link = fields.Char()
    #
    # def actipn_confirmed(self):
    #     self.sent_for_approval()
    #     self.state = 'confirm'

    # def sent_for_approval(self):
    #     mail_template = self.env.ref('iatf.mom_document_approval_mail_template')
    #     model_id = self.env['ir.model'].sudo().search([('model', '=', self._name)])
    #     mail_template.write({'model_id' : model_id.id})
    #
    #     web_base_url = self.env['ir.config_parameter'].sudo().get_param('web.base.url')
    #     action_id = self.env.ref('iatf.action_npd_review_view', raise_if_not_found=False)
    #     link = """{}/web#id={}&view_type=form&model=self._name&action={}""".format(web_base_url,self.id,action_id.id)
    #     self.link = link
    #
    #     user_ids = self.env['hr.employee'].sudo().search([('department_id', 'in', self.approve_department_ids.ids)]).mapped('user_id')
    #     all_admin = ''
    #     for user in user_ids:
    #         if all_admin:
    #             all_admin += ',' + str(user.login)
    #         else:
    #             all_admin = str(user.login)
    #     mail_template.write({
    #             'email_to': all_admin,
    #     })
    #     mail_template.send_mail(self.id, force_send=True)

    # def _check_department(self):
    #     for rec in self:
    #         if self.env.user.employee_id.department_id.id in rec.approve_department_ids.ids and self.env.user.employee_id.department_id.department_type == 'hr':
    #             if rec.state != 'draft':
    #                 rec.hr_check_department = True
    #             else:
    #                 rec.hr_check_department = False
    #         else:
    #             rec.hr_check_department = False
    #         if self.env.user.employee_id.department_id.id in rec.approve_department_ids.ids and self.env.user.employee_id.department_id.department_type == 'design':
    #             if rec.state != 'draft':
    #                 rec.des_check_department = True
    #             else:
    #                 rec.des_check_department = False
    #         else:
    #             rec.des_check_department = False
    #         if self.env.user.employee_id.department_id.id in rec.approve_department_ids.ids and self.env.user.employee_id.department_id.department_type == 'eng':
    #             if rec.state != 'draft':
    #                 rec.eng_check_department = True
    #             else:
    #                 rec.eng_check_department = False
    #         else:
    #             rec.eng_check_department = False
    #         if self.env.user.employee_id.department_id.id in rec.approve_department_ids.ids and self.env.user.employee_id.department_id.department_type == 'manu':
    #             if rec.state != 'draft':
    #                 rec.manu_check_department = True
    #             else:
    #                 rec.manu_check_department = False
    #         else:
    #             rec.manu_check_department = False
    #         if self.env.user.employee_id.department_id.id in rec.approve_department_ids.ids and self.env.user.employee_id.department_id.department_type == 'sales':
    #             if rec.state != 'draft':
    #                 rec.sale_check_department = True
    #             else:
    #                 rec.sale_check_department = False
    #         else:
    #             rec.sale_check_department = False
    #         if self.env.user.employee_id.department_id.id in rec.approve_department_ids.ids and self.env.user.employee_id.department_id.department_type == 'qc':
    #             if rec.state != 'draft':
    #                 rec.qc_check_department = True
    #             else:
    #                 rec.qc_check_department = False
    #         else:
    #             rec.qc_check_department = False
    #         if self.env.user.employee_id.department_id.id in rec.approve_department_ids.ids and self.env.user.employee_id.department_id.department_type == 'maintenance':
    #             if rec.state != 'draft':
    #                 rec.mainte_check_department = True
    #             else:
    #                 rec.mainte_check_department = False
    #         else:
    #             rec.mainte_check_department = False
    #         if self.env.user.employee_id.department_id.id in rec.approve_department_ids.ids and self.env.user.employee_id.department_id.department_type == 'marketing':
    #             if rec.state != 'draft':
    #                 rec.marketing_check_department = True
    #             else:
    #                 rec.marketing_check_department = False
    #         else:
    #             rec.marketing_check_department = False
    #         if self.env.user.employee_id.department_id.id in rec.approve_department_ids.ids and self.env.user.employee_id.department_id.department_type == 'pm':
    #             if rec.state != 'draft':
    #                 rec.pm_check_department = True
    #             else:
    #                 rec.pm_check_department = False
    #         else:
    #             rec.pm_check_department = False
    #         if self.env.user.employee_id.department_id.id in rec.approve_department_ids.ids and self.env.user.employee_id.department_id.department_type == 'management':
    #             if rec.state != 'draft':
    #                 rec.managment_check_department = True
    #             else:
    #                 rec.managment_check_department = False
    #         else:
    #             rec.managment_check_department = False
    #
    #
    # hr_approval_date = fields.Datetime('Hr Approval Date')
    # design_approval_date = fields.Datetime('Design Approval Date')
    # eng_approval_date = fields.Datetime('Eng. Approval Date')
    # manu_approval_date = fields.Datetime('Production Approval Date')
    # sale_approval_date = fields.Datetime('Sales Approval Date')
    # qc_approval_date = fields.Datetime('QC Approval Date')
    # main_approval_date = fields.Datetime('Maintenance Approval Date')
    # mark_approval_date = fields.Datetime('Marketing Approval Date')
    # pm_approval_date = fields.Datetime('Program Management Approval Date')
    # maneg_approval_date = fields.Datetime('Top Management Approval Date')
    #
    # def hr_approval(self):
    #     self.hr_approval_date = datetime.now()
    #     self.is_hr_approved = True
    #     self.approve_by_department_ids = [(4, self.env.user.employee_id.department_id.id)]
    #     self.approvaed_manager_ids = [(4, self.env.user.employee_id.department_id.manager_id.id)]
    # def design_approval(self):
    #     self.design_approval_date = datetime.now()
    #     self.is_design_approved = True
    #     self.approve_by_department_ids = [(4, self.env.user.employee_id.department_id.id)]
    #     self.approvaed_manager_ids = [(4, self.env.user.employee_id.department_id.manager_id.id)]
    # def eng_approval(self):
    #     self.eng_approval_date = datetime.now()
    #     self.is_eng_approved = True
    #     self.approve_by_department_ids = [(4, self.env.user.employee_id.department_id.id)]
    #     self.approvaed_manager_ids = [(4, self.env.user.employee_id.department_id.manager_id.id)]
    # def manu_approval(self):
    #     self.manu_approval_date = datetime.now()
    #     self.is_manu_approved = True
    #     self.approve_by_department_ids = [(4, self.env.user.employee_id.department_id.id)]
    #     self.approvaed_manager_ids = [(4, self.env.user.employee_id.department_id.manager_id.id)]
    # def sales_approval(self):
    #     self.sale_approval_date = datetime.now()
    #     self.is_sales_approved = True
    #     self.approve_by_department_ids = [(4, self.env.user.employee_id.department_id.id)]
    #     self.approvaed_manager_ids = [(4, self.env.user.employee_id.department_id.manager_id.id)]
    # def qc_approval(self):
    #     self.qc_approval_date = datetime.now()
    #     self.is_qc_approved = True
    #     self.approve_by_department_ids = [(4, self.env.user.employee_id.department_id.id)]
    #     self.approvaed_manager_ids = [(4, self.env.user.employee_id.department_id.manager_id.id)]
    # def mainte_approval(self):
    #     self.main_approval_date = datetime.now()
    #     self.is_maintenance_approved = True
    #     self.approve_by_department_ids = [(4, self.env.user.employee_id.department_id.id)]
    #     self.approvaed_manager_ids = [(4, self.env.user.employee_id.department_id.manager_id.id)]
    # def marketing_approval(self):
    #     self.mark_approval_date = datetime.now()
    #     self.is_marketing_approved = True
    #     self.approve_by_department_ids = [(4, self.env.user.employee_id.department_id.id)]
    #     self.approvaed_manager_ids = [(4, self.env.user.employee_id.department_id.manager_id.id)]
    # def pm_approval(self):
    #     self.pm_approval_date = datetime.now()
    #     self.is_pm_approved = True
    #     self.approve_by_department_ids = [(4, self.env.user.employee_id.department_id.id)]
    #     self.approvaed_manager_ids = [(4, self.env.user.employee_id.department_id.manager_id.id)]
    # def managment_approval(self):
    #     self.maneg_approval_date = datetime.now()
    #     self.is_management_approved = True
    #     self.approve_by_department_ids = [(4, self.env.user.employee_id.department_id.id)]
    #     self.approvaed_manager_ids = [(4, self.env.user.employee_id.department_id.manager_id.id)]
    #
    #
    # def button_send_approval(self):
    #     for rec in self:
    #         if rec.hr:
    #             user_email = rec.hr.login
    #             mail_temp = self.env.ref('iatf.sending_mail_template_risk_assessment')
    #             if mail_temp:
    #                 format_name = rec.format_id.name
    #                 cust_name = rec.cust_id.name
    #                 subject = f" Request for approving Risk Assessment {format_name} for {cust_name}"
    #                 if user_email:
    #                     body = """
    #                     <div>
    #                         <p>Dear Recipient  """ + str(user_email) + """,
    #                             <br/><br/>
    #                             Please, kindly Request To accept and approved Sending Approval for the NPD Review.
    #                         <br></br>
    #                         Thank you.
    #                     <br/>
    #                     <br/>
    #                     <div>"""
    #                     mail_temp.send_mail(self.id, email_values={
    #                         'email_from': self.env.user.login,
    #                         'email_to': user_email,
    #                         'subject': subject,
    #                         'body_html': body,
    #                     }, force_send=True)
    #
    #         rec.write({'state': 'hr_approve'})
    #         rec.part_development_id.cft_id = rec.id
    #
    # def button_hr_approval(self):
    #     for rec in self:
    #         if rec.hr:
    #             user_email = rec.hr.login
    #             mail_temp = self.env.ref('iatf.sending_mail_template_risk_assessment')
    #             if mail_temp:
    #                 format_name = rec.format_id.name
    #                 cust_name = rec.cust_id.name
    #                 subject = f" Request for approving Risk Assessment {format_name} for {cust_name}"
    #                 if user_email:
    #                     body = """
    #                     <div>
    #                         <p>Dear Recipient  """ + str(user_email) + """,
    #                             <br/><br/>
    #                             Please, kindly Request To accept and approved Sending Approval for the NPD Review.
    #                         <br></br>
    #                         Thank you.
    #                     <br/>
    #                     <br/>
    #                     <div>"""
    #                     mail_temp.send_mail(self.id, email_values={
    #                         'email_from': self.env.user.login,
    #                         'email_to': user_email,
    #                         'subject': subject,
    #                         'body_html': body,
    #                     }, force_send=True)
    #         rec.write({'state': 'design'})
    #
    # def button_design(self):
    #     for rec in self:
    #         if rec.design_eng:
    #             user_email = rec.design_eng.login
    #             mail_temp = self.env.ref('iatf.sending_mail_template_risk_assessment')
    #             if mail_temp:
    #                 format_name = rec.format_id.name
    #                 cust_name = rec.cust_id.name
    #                 subject = f" Request for approving Risk Assessment {format_name} for {cust_name}"
    #                 if user_email:
    #                     body = """
    #                     <div>
    #                         <p>Dear Recipient  """ + str(user_email) + """,
    #                             <br/><br/>
    #                             Please, kindly Request To accept and approved Sending Approval for the NPD Review.
    #                         <br></br>
    #                         Thank you.
    #                     <br/>
    #                     <br/>
    #                     <div>"""
    #                     mail_temp.send_mail(self.id, email_values={
    #                         'email_from': self.env.user.login,
    #                         'email_to': user_email,
    #                         'subject': subject,
    #                         'body_html': body,
    #                     }, force_send=True)
    #         rec.write({'state': 'engineering'})
    #
    # def button_engineering(self):
    #     for rec in self:
    #         if rec.manf_eng:
    #             user_email = rec.manf_eng.login
    #             mail_temp = self.env.ref('iatf.sending_mail_template_risk_assessment')
    #             if mail_temp:
    #                 format_name = rec.format_id.name
    #                 cust_name = rec.cust_id.name
    #                 subject = f" Request for approving Risk Assessment {format_name} for {cust_name}"
    #                 if user_email:
    #                     body = """
    #                     <div>
    #                         <p>Dear Recipient  """ + str(user_email) + """,
    #                             <br/><br/>
    #                             Please, kindly Request To accept and approved Sending Approval for the NPD Review.
    #                         <br></br>
    #                         Thank you.
    #                     <br/>
    #                     <br/>
    #                     <div>"""
    #                     mail_temp.send_mail(self.id, email_values={
    #                         'email_from': self.env.user.login,
    #                         'email_to': user_email,
    #                         'subject': subject,
    #                         'body_html': body,
    #                     }, force_send=True)
    #         rec.write({'state': 'manufacturing'})
    #
    # def button_manufacturing(self):
    #     for rec in self:
    #         if rec.production:
    #             user_email = rec.production.login
    #             mail_temp = self.env.ref('iatf.sending_mail_template_risk_assessment')
    #             if mail_temp:
    #                 format_name = rec.format_id.name
    #                 cust_name = rec.cust_id.name
    #                 subject = f" Request for approving Risk Assessment {format_name} for {cust_name}"
    #                 if user_email:
    #                     body = """
    #                     <div>
    #                         <p>Dear Recipient  """ + str(user_email) + """,
    #                             <br/><br/>
    #                             Please, kindly Request To accept and approved Sending Approval for the NPD Review.
    #                         <br></br>
    #                         Thank you.
    #                     <br/>
    #                     <br/>
    #                     <div>"""
    #                     mail_temp.send_mail(self.id, email_values={
    #                         'email_from': self.env.user.login,
    #                         'email_to': user_email,
    #                         'subject': subject,
    #                         'body_html': body,
    #                     }, force_send=True)
    #         rec.write({'state': 'quality'})
    #
    # def button_quality_test_done(self):
    #     for rec in self:
    #         if rec.quality:
    #             user_email = rec.quality.login
    #             mail_temp = self.env.ref('iatf.sending_mail_template_risk_assessment')
    #             if mail_temp:
    #                 format_name = rec.format_id.name
    #                 cust_name = rec.cust_id.name
    #                 subject = f" Request for approving Risk Assessment {format_name} for {cust_name}"
    #                 if user_email:
    #                     body = """
    #                     <div>
    #                         <p>Dear Recipient  """ + str(user_email) + """,
    #                             <br/><br/>
    #                             Please, kindly Request To accept and approved Sending Approval for the NPD Review.
    #                         <br></br>
    #                         Thank you.
    #                     <br/>
    #                     <br/>
    #                     <div>"""
    #                     mail_temp.send_mail(self.id, email_values={
    #                         'email_from': self.env.user.login,
    #                         'email_to': user_email,
    #                         'subject': subject,
    #                         'body_html': body,
    #                     }, force_send=True)
    #         rec.write({'state': 'top'})
    #
    # def button_top_managment_final_approved(self):
    #     for rec in self:
    #         if rec.top_management_id:
    #             user_email = rec.top_management_id.login
    #             mail_temp = self.env.ref('iatf.sending_mail_template_risk_assessment')
    #             if mail_temp:
    #                 format_name = rec.format_id.name
    #                 cust_name = rec.cust_id.name
    #                 subject = f" Request for approving Risk Assessment {format_name} for {cust_name}"
    #                 if user_email:
    #                     body = """
    #                     <div>
    #                         <p>Dear Recipient  """ + str(user_email) + """,
    #                             <br/><br/>
    #                             Please, kindly Request To accept and approved Sending Approval for the NPD Review.
    #                         <br></br>
    #                         Thank you.
    #                     <br/>
    #                     <br/>
    #                     <div>"""
    #                     mail_temp.send_mail(self.id, email_values={
    #                         'email_from': self.env.user.login,
    #                         'email_to': user_email,
    #                         'subject': subject,
    #                         'body_html': body,
    #                     }, force_send=True)
    #         rec.write({'state': 'final_approved'})

    def action_npd_review(self):
        output = BytesIO()
        wb = Workbook()
        ws = wb.active

        if self.env.user.company_id.logo:
            max_width = 500  # Set your desired maximum width
            max_height = 100  # Set your desired maximum height
            image_data = base64.b64decode(self.env.user.company_id.logo)

            # Open the image using PIL
            image = PILImage.open(io.BytesIO(image_data))
            width, height = image.size
            aspect_ratio = width / height

            if width > max_width:
                width = max_width
                height = int(width / aspect_ratio)

            if height > max_height:
                height = max_height
                width = int(height * aspect_ratio)

            # Resize the image using PIL
            # Add space on the top and left side of the image
            padding_top = 10  # Adjust as needed
            padding_left = 10  # Adjust as needed

            resized_image = image.resize((width, height),PILImage.Resampling.LANCZOS)
            ImageOps.expand(resized_image, border=(padding_left, padding_top, 0, 0), fill='rgba(0,0,0,0)')
            img_bytes = io.BytesIO()
            resized_image.save(img_bytes, format='PNG')
            img_bytes.seek(0)
            logo_image = Image(img_bytes)
            # logo_image = Image(self.env.user.company_id.logo)
            ws.add_image(logo_image, 'B1')

        border = Border(top=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'),
                        bottom=Side(style='thin'))
        font_header = Font(name='Arial', size=12, bold=True)
        font_normal = Font(name='Arial', size=11)
        align_left = Alignment(vertical='center', horizontal='left', wrapText=True)
        align_right = Alignment(vertical='center', horizontal='right', wrapText=True)
        align_center = Alignment(vertical='center', horizontal='center', wrapText=True)

        data = {
            'B1': 'NPD review - Minutes Of Meeting',
            'B2': 'Name of Members',
            'F2': 'Name of Participants- ',
            'B4': 'S.NO.',
            'C4': 'Customer',
            'D4': 'Project name',
            'E4': 'Number of parts in project',
            'F4': 'Status as on ',
            'G4': 'Resp.',
        }

        for cell, value in data.items():
            ws[cell] = value

        for rows in ws.iter_rows(min_row=1, max_row=25, min_col=2, max_col=7):
            for cell in rows:
                cell.alignment = align_center
                cell.border = border
                cell.font = font_header

        cell_ranges_to_merge = ['B1:G1', 'B2:E2', 'F2:G2', 'B3:E3', 'F3:G3']
        for cell_range in cell_ranges_to_merge:
            ws.merge_cells(cell_range)

        ws.column_dimensions['A'].hidden = True
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 48
        ws.column_dimensions['G'].width = 20

        ws.row_dimensions[1].height = 75
        ws.row_dimensions[2].height = 30
        ws.row_dimensions[3].height = 100
        for row_number in range(4, 25):
            ws.row_dimensions[row_number].height = 35

        ws['B1'].font = Font(name='Arial', size=22, bold=True)

        # Data Filling
        cur_row = 5
        for rec in self:
            ws['B3'] = rec.partner_id.name if rec.partner_id.name else ''

            if rec.employee_ids:
                ws['F3'] = ', '.join([nop.name for nop in rec.employee_ids])

            for i in rec.npd_line_ids:
                ws[f'B{cur_row}'] = i.sl_no if i.sl_no else ''
                ws[f'C{cur_row}'] = i.customer_id.name if i.customer_id else ''
                ws[f'D{cur_row}'] = i.project_name if i.project_name else ''
                ws[f'E{cur_row}'] = i.part_no if i.part_no else ''
                ws[f'F{cur_row}'] = i.status_on if i.status_on else ''
                ws[f'G{cur_row}'] = i.employee_id.name if i.employee_id.name else ''

                ws.row_dimensions[cur_row].height = 35
                for row_no in ws.iter_rows(min_row=cur_row, max_row=cur_row + 1, min_col=2, max_col=7):
                    for cell in row_no:
                        cell.alignment = align_center
                        cell.font = font_normal
                        cell.border = border

                cur_row += 1
        if cur_row > 24:
            ws.merge_cells(f'B{cur_row}:G{cur_row}')
            ws[f'B{cur_row}'].alignment = align_right
            ws[f'B{cur_row}'] = 'FM-80552 / 0 - 0 Date: 26.08.2017'
        else:
            ws.merge_cells('B25:G25')
            ws['B25'].alignment = align_right
            ws['B25'] = 'FM-80552 / 0 - 0 Date: 26.08.2017'

        wb.save(output)
        output.seek(0)
        self.generate_xls_file = base64.b64encode(output.getvalue()).decode('utf-8')

        return {
            "type": "ir.actions.act_url",
            "target": "self",
            "url": "/web/content?model=npd.review&download=true&field=generate_xls_file&filename={filename}.xlsx&id={pid}".format(
                filename="NPD review - Minutes Of Meeting", pid=self[0].id),
        }


class APDReviewLines(models.Model):
    _name = "npd.review.line"
    _description = "NPD Review Line"
    _inherit = "translation.mixin"

    npd_id = fields.Many2one(comodel_name='npd.review', string='NPD', required=True, ondelete='cascade', index=True,
                             copy=False)
    customer_id = fields.Many2one('res.partner', 'Customer')
    project_name = fields.Char('Project name',translate=True)
    part_no = fields.Char('Number of parts in project')
    status_on = fields.Char('Status as on',translate=True)
    employee_id = fields.Many2one('hr.employee', 'Resp.')
    sequence = fields.Integer(string="Sequence", default=10)
    sl_no = fields.Integer("S.No", compute="_compute_sequence_number")

    @api.depends('sequence', 'npd_id')
    def _compute_sequence_number(self):
        for line in self.mapped('npd_id'):
            sl_no = 1
            for lines in line.npd_line_ids:
                lines.sl_no = sl_no
                sl_no += 1
