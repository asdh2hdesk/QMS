import base64
import copy
import io
from fnmatch import translate
from io import BytesIO

from PIL import Image as PILImage
from PIL import ImageOps
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

from odoo import fields, models, api


# [
#         ('a', 'Project Cost / Return on Investment (ROI)'),
#         ('b', 'Capacity'),
#         ('c', 'Product Liability'),
#         ('d', 'RM Development'),
#         ('e', 'Quality Risk - Internal'),
#         ('f', 'Quality Risk - External'),
#         ('g', 'Delivery'),
#         ('h', 'Transportation & Unforeseen Activities'),
#         ('i', 'Environmental and Political'),
#         ('j', 'Any Other'),
#     ]

class RiskAnalysis(models.Model):
    _name = 'risk.analysis'
    _rec_name = "name"
    # _inherit = "translation.mixin"

    # s_no = fields.Integer('S no.', default=1, required=True)
    name = fields.Char('Risk Analysis', required=True, translate=True)
    criteria_l = fields.Char('Criteria for Low Risk', required=True, translate=True)
    criteria_m = fields.Char('Criteria for Medium Risk', required=True, translate=True)
    criteria_h = fields.Char('Criteria for High Risk', required=True, translate=True)
    default = fields.Binary('is default', readonly=True, default=False)


class RiskAssessment(models.Model):
    _name = 'risk.assessment'
    # _rec_name = "format_id"
    _inherit = ['iatf.sign.off.members','translation.mixin']

    # risk_id = fields.Many2one(
    #     comodel_name='risk.assessment',
    #     string='Risk Assessment',
    #     required=True, ondelete='cascade', index=True, copy=False)

    tree_state = fields.Binary(default=False)

    risk_assessment_line_ids = fields.One2many(
        comodel_name='risk.assessment.lines',
        inverse_name='risk_id',
        string="Risk Assessment Line"
    )
    # format_id = fields.Many2one('mom.format', 'Format')
    # cust_id = fields.Many2one('res.partner', 'Customer Name')
    cust_drawing_num = fields.Char('Customer Drawing Number')
    # product_id = fields.Many2one('Product')   # :todo: link with PLM
    # product = fields.Char('Product')
    # product_num = fields.Char('Product')
    date_ = fields.Date('Date', default=fields.Date.context_today)
    comment = fields.Text('Comment', translate=True)
    viability = fields.Selection([
        ('viable', 'Viable'),
        ('not_viable', 'Not Viable'),
        ('viable_with_change', 'Viable with changes'),
    ], )
    # cft_member_ids = fields.Many2many('res.users', string='CFT Members')
    # top_management_id = fields.Many2one('res.users', 'Top Management')  # Final Approved
    # hr = fields.Many2one('res.users', 'HR')  # HR APPROVED
    # design_eng = fields.Many2one('res.users', 'Design Engineering')  # DESIGN
    # manf_eng = fields.Many2one('res.users', 'Manufacturing Engineering')  # Engineering
    # marketing = fields.Many2one('res.users', 'Marketing')
    # program_management = fields.Many2one('res.users', 'Program Management')
    # production = fields.Many2one('res.users', 'Production')  # Manufacturing
    # quality = fields.Many2one('res.users', 'Quality')  # Quality
    generate_xls_file = fields.Binary(string="Generated file")

    # user_id = fields.Many2one('res.users', string='USer', default=lambda self: self.env.uid)
    # champion = fields.Many2one('res.users', 'Champion')

    # Status
    # state = fields.Selection([
    #     ('draft', 'Draft'),
    #     ('hr_approve', 'HR Approval'),
    #     ('design', 'Design'),
    #     ('engineering', 'Engineering'),
    #     ('manufacturing', 'Manufacturing'),
    #     ('quality', 'Quality'),
    #     ('top', 'Top Management'),
    #     ('final_approved', 'Final Approved')
    # ], string='Status', default='draft')

    # part_development_id = fields.Many2one("part.development.process")
    #
    # state = fields.Selection([
    #                         ('draft', 'Draft'),
    #                         ('confirm', 'Confirmed'),
    #                         ], string='Status', default='draft')
    # document_name = fields.Char(string='Document #')
    # approve_department_id = fields.Many2one('hr.department', string='Departments Approvals')
    # # document_pro_id = fields.Many2one("xf.doc.approval.document.package", string="Document #")
    # approve_department_ids = fields.Many2many('hr.department', 'tx3', string='Departments Approvals')
    # approve_by_department_ids = fields.Many2many('hr.department', string='Departments Approved By')
    # manager_id = fields.Many2one('hr.employee', string='Approval Manager')
    # manager_ids = fields.Many2many('hr.employee', 'tx4', string='Approval Managers')
    # approvaed_manager_ids = fields.Many2many('hr.employee', string='Managers Approved By')

    @api.model
    def create(self, vals):
        default_analysis_line = self.env['risk.analysis'].search([('default', '=', True)])
        if default_analysis_line:
            risk_assessment_lines = []
            for analysis_line in default_analysis_line:
                risk_assessment_lines.append((0, 0, {'risk_analysis_id': analysis_line.id}))
            vals['risk_assessment_line_ids'] = risk_assessment_lines
        return super(RiskAssessment, self).create(vals)

    def _compute_tree_visibility(self):
        if self.risk_assessment_line_ids:
            self.tree_state = True
        else:
            self.tree_state = False

    # def button_send_approval(self):
    #     for rec in self:
    #         # Mail Send for approval
    #         if rec.hr:
    #             user_email = rec.hr.login
    #             mail_temp = self.env.ref('iatf.sending_mail_template_risk_assessment')
    #             if mail_temp:
    #                 format_name = rec.format_id.name
    #                 cust_name = rec.cust_id.name
    #                 subject = f" Request for approving Risk Assessment {format_name} for {cust_name}"
    #                 if user_email:
    #                     body = """
    #                     <div>
    #                         <p>Dear Recipient  """ + str(user_email) + """,
    #                             <br/><br/>
    #                             Please, kindly Request To accept and approved Sending Approval for the Risk Analysis Sheet.
    #                         <br></br>
    #                         Thank you.
    #                     <br/>
    #                     <br/>
    #                     <div>"""
    #                     mail_temp.send_mail(self.id, email_values={
    #                         'email_from': self.env.user.login,
    #                         'email_to': user_email,
    #                         'subject': subject,
    #                         'body_html': body,
    #                     }, force_send=True)
    #
    #         rec.write({'state': 'hr_approve'})
    #         rec.part_development_id.risk_assessment_id = rec.id
    #
    # def button_hr_approval(self):
    #     for rec in self:
    #         # Mail Send for hr approval
    #         if rec.hr:
    #             user_email = rec.hr.login
    #             mail_temp = self.env.ref('iatf.sending_mail_template_risk_assessment')
    #             if mail_temp:
    #                 format_name = rec.format_id.name
    #                 cust_name = rec.cust_id.name
    #                 subject = f" Request for approving Risk Assessment {format_name} for {cust_name}"
    #                 if user_email:
    #                     body = """
    #                     <div>
    #                         <p>Dear Recipient  """ + str(user_email) + """,
    #                             <br/><br/>
    #                             Please, kindly Request To accept and approved Sending Approval for the Risk Analysis Sheet.
    #                         <br></br>
    #                         Thank you.
    #                     <br/>
    #                     <br/>
    #                     <div>"""
    #                     mail_temp.send_mail(self.id, email_values={
    #                         'email_from': self.env.user.login,
    #                         'email_to': user_email,
    #                         'subject': subject,
    #                         'body_html': body,
    #                     }, force_send=True)
    #         rec.write({'state': 'design'})
    #
    # def button_design(self):
    #     for rec in self:
    #         # Mail Send for Design approval
    #         if rec.design_eng:
    #             user_email = rec.design_eng.login
    #             mail_temp = self.env.ref('iatf.sending_mail_template_risk_assessment')
    #             if mail_temp:
    #                 format_name = rec.format_id.name
    #                 cust_name = rec.cust_id.name
    #                 subject = f" Request for approving Risk Assessment {format_name} for {cust_name}"
    #                 if user_email:
    #                     body = """
    #                     <div>
    #                         <p>Dear Recipient  """ + str(user_email) + """,
    #                             <br/><br/>
    #                             Please, kindly Request To accept and approved Sending Approval for the Risk Analysis Sheet.
    #                         <br></br>
    #                         Thank you.
    #                     <br/>
    #                     <br/>
    #                     <div>"""
    #                     mail_temp.send_mail(self.id, email_values={
    #                         'email_from': self.env.user.login,
    #                         'email_to': user_email,
    #                         'subject': subject,
    #                         'body_html': body,
    #                     }, force_send=True)
    #         rec.write({'state': 'engineering'})
    #
    # def button_engineering(self):
    #     for rec in self:
    #         # Mail Send for Engineering approval
    #         if rec.manf_eng:
    #             user_email = rec.manf_eng.login
    #             mail_temp = self.env.ref('iatf.sending_mail_template_risk_assessment')
    #             if mail_temp:
    #                 format_name = rec.format_id.name
    #                 cust_name = rec.cust_id.name
    #                 subject = f" Request for approving Risk Assessment {format_name} for {cust_name}"
    #                 if user_email:
    #                     body = """
    #                     <div>
    #                         <p>Dear Recipient  """ + str(user_email) + """,
    #                             <br/><br/>
    #                             Please, kindly Request To accept and approved Sending Approval for the Risk Analysis Sheet.
    #                         <br></br>
    #                         Thank you.
    #                     <br/>
    #                     <br/>
    #                     <div>"""
    #                     mail_temp.send_mail(self.id, email_values={
    #                         'email_from': self.env.user.login,
    #                         'email_to': user_email,
    #                         'subject': subject,
    #                         'body_html': body,
    #                     }, force_send=True)
    #         rec.write({'state': 'manufacturing'})
    #
    # def button_manufacturing(self):
    #     for rec in self:
    #         # Mail Send for Manufacturing approval
    #         if rec.production:
    #             user_email = rec.production.login
    #             mail_temp = self.env.ref('iatf.sending_mail_template_risk_assessment')
    #             if mail_temp:
    #                 format_name = rec.format_id.name
    #                 cust_name = rec.cust_id.name
    #                 subject = f" Request for approving Risk Assessment {format_name} for {cust_name}"
    #                 if user_email:
    #                     body = """
    #                     <div>
    #                         <p>Dear Recipient  """ + str(user_email) + """,
    #                             <br/><br/>
    #                             Please, kindly Request To accept and approved Sending Approval for the Risk Analysis Sheet.
    #                         <br></br>
    #                         Thank you.
    #                     <br/>
    #                     <br/>
    #                     <div>"""
    #                     mail_temp.send_mail(self.id, email_values={
    #                         'email_from': self.env.user.login,
    #                         'email_to': user_email,
    #                         'subject': subject,
    #                         'body_html': body,
    #                     }, force_send=True)
    #         rec.write({'state': 'quality'})
    #
    # def button_quality_test_done(self):
    #     for rec in self:
    #         # Mail Send for Quality Test approval
    #         if rec.quality:
    #             user_email = rec.quality.login
    #             mail_temp = self.env.ref('iatf.sending_mail_template_risk_assessment')
    #             if mail_temp:
    #                 format_name = rec.format_id.name
    #                 cust_name = rec.cust_id.name
    #                 subject = f" Request for approving Risk Assessment {format_name} for {cust_name}"
    #                 if user_email:
    #                     body = """
    #                     <div>
    #                         <p>Dear Recipient  """ + str(user_email) + """,
    #                             <br/><br/>
    #                             Please, kindly Request To accept and approved Sending Approval for the Risk Analysis Sheet.
    #                         <br></br>
    #                         Thank you.
    #                     <br/>
    #                     <br/>
    #                     <div>"""
    #                     mail_temp.send_mail(self.id, email_values={
    #                         'email_from': self.env.user.login,
    #                         'email_to': user_email,
    #                         'subject': subject,
    #                         'body_html': body,
    #                     }, force_send=True)
    #         rec.write({'state': 'top'})
    #
    # def button_top_managment_final_approved(self):
    #     for rec in self:
    #         # Mail Send for Final approval
    #         if rec.top_management_id:
    #             user_email = rec.top_management_id.login
    #             mail_temp = self.env.ref('iatf.sending_mail_template_risk_assessment')
    #             if mail_temp:
    #                 format_name = rec.format_id.name
    #                 cust_name = rec.cust_id.name
    #                 subject = f" Request for approving Risk Assessment {format_name} for {cust_name}"
    #                 if user_email:
    #                     body = """
    #                     <div>
    #                         <p>Dear Recipient  """ + str(user_email) + """,
    #                             <br/><br/>
    #                             Please, kindly Request To accept and approved Sending Approval for the Risk Analysis Sheet.
    #                         <br></br>
    #                         Thank you.
    #                     <br/>
    #                     <br/>
    #                     <div>"""
    #                     mail_temp.send_mail(self.id, email_values={
    #                         'email_from': self.env.user.login,
    #                         'email_to': user_email,
    #                         'subject': subject,
    #                         'body_html': body,
    #                     }, force_send=True)
    #         rec.write({'state': 'final_approved'})

    def action_print_risk_assessment_sheet(self):
        # Create a new workbook
        output = BytesIO()
        wb = Workbook()
        ws = wb.active

        data = {
            'A1': 'Risk Assessment Sheet',
            'A2': 'Customer Name :',
            'A3': 'Customer Drg. No. :',
            'E2': 'Product Number:',
            'H2': 'Product Name:',
            'E3': 'Date :',
            'B5': 'Risk Analysis',
            'C5': 'Risk Low (Rl)',
            'D5': 'Risk Medium (Rm)',
            'E5': 'Risk High (Rh)',
            'F5': 'Action plan ',
            'G5': 'Target date',
            'H5': 'Responsibility',
            'I5': 'Status',
            'H3': 'Champion :',
            'J5': 'Criteria',
        }
        # Fill data into specific cells using key-value pairs
        for cell, value in data.items():
            ws[cell] = value

        thin = Side(border_style='thin', color='000000')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        font_header = Font(name='Arial', size=10, bold=True)
        font_all = Font(name='Arial', size=10, bold=False)
        align_left = Alignment(vertical='center', horizontal='left', wrapText=True)
        align_center = Alignment(vertical='center', horizontal='center', wrapText=True)

        # region merging and formatting cells
        max_col = 10
        max_row = 15
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                alignment = copy.copy(cell.alignment)
                alignment.wrapText = True
                cell.alignment = align_left
                cell.font = font_all

        # Specific Dimension
        ws['A1'].alignment = Alignment(vertical='center', horizontal='center')
        ws['A1'].font = Font(name='Arial', size=22, bold=True)
        ws['J5'].alignment = Alignment(vertical='center', horizontal='center')
        ws['A20'].alignment = Alignment(vertical='center', horizontal='right')

        # Merging the cells as per standard sheet
        ws.merge_cells('A1:J1')
        ws.merge_cells('A2:B2')
        ws.merge_cells('C2:D2')
        ws.merge_cells('A3:B3')
        ws.merge_cells('C3:D3')
        ws.merge_cells('F2:G2')
        ws.merge_cells('F3:G3')
        ws.merge_cells('I2:J2')
        ws.merge_cells('I3:J3')
        ws.merge_cells('A4:J4')

        # Dimension of Columns
        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 16
        ws.column_dimensions['C'].width = 16
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 27

        # Dimension of Rows
        ws.row_dimensions[1].height = 75
        ws.row_dimensions[5].height = 35

        items_to_bold = ['B5', 'C5', 'D5', 'E5', 'F5', 'G5', 'H5', 'I5', 'J5', 'A2', 'A3', 'E2', 'E3', 'H2', 'H3']

        for item in items_to_bold:
            ws[item].font = font_header

        if self.env.user.company_id.logo:
            max_width = 500  # Set your desired maximum width
            max_height = 100  # Set your desired maximum height
            image_data = base64.b64decode(self.env.user.company_id.logo)

            # Open the image using PIL
            image = PILImage.open(io.BytesIO(image_data))
            width, height = image.size
            aspect_ratio = width / height

            if width > max_width:
                width = max_width
                height = int(width / aspect_ratio)

            if height > max_height:
                height = max_height
                width = int(height * aspect_ratio)

            # Resize the image using PIL
            # Add space on the top and left side of the image
            padding_top = 10  # Adjust as needed
            padding_left = 10  # Adjust as needed

            resized_image = image.resize((width, height),  PILImage.LANCZOS)
            ImageOps.expand(resized_image, border=(padding_left, padding_top, 0, 0), fill='rgba(0,0,0,0)')
            img_bytes = io.BytesIO()
            resized_image.save(img_bytes, format='PNG')
            img_bytes.seek(0)
            logo_image = Image(img_bytes)
            # logo_image = Image(self.env.user.company_id.logo)
            ws.add_image(logo_image, 'A1')

        # region Filling data in sheet

        for rec in self:
            ws['C2'] = rec.partner_id.name if rec.partner_id else ''
            ws['C3'] = rec.cust_drawing_num if rec.cust_drawing_num else ''
            ws['f2'] = rec.part_number if rec.part_number else ''
            ws['f3'] = rec.date_ if rec.date_ else ''
            ws['I2'] = rec.part_name if rec.part_name else ''

            row = 6
            related_ids = []
            for ral in rec.risk_assessment_line_ids:
                ws[f'A{row}'] = row - 5
                ws[f'B{row}'] = ral.risk_analysis_id.name if ral.risk_analysis_id.name else ''

                if ral.risk_level == 'low':
                    ws[f'C{row}'] = '☑'
                elif ral.risk_level == 'medium':
                    ws[f'D{row}'] = '☑'
                elif ral.risk_level == 'high':
                    ws[f'E{row}'] = '☑'
                ws[f'F{row}'] = ral.risk_action_plan if ral.risk_action_plan else ''
                ws[f'G{row}'] = ral.risk_action_plan_date if ral.risk_action_plan_date else ''
                ws[f'H{row}'] = ral.risk_action_plan_responsible.name if ral.risk_action_plan_responsible.name else ''
                # ws[f'I{row}'] = ral.risk_plan_status if ral.risk_plan_status else ''

                if ral.risk_plan_status:
                    soup = BeautifulSoup(ral.risk_plan_status, 'html.parser')
                    text_content = soup.find('p').get_text()
                    if text_content:
                        ws[f'I{row}'] = text_content

                ws[f'J{row}'] = ral.criteria if ral.criteria else ''

                ws[f'B{row}'].alignment = align_center
                ws[f'C{row}'].font = Font(name='Arial', size=28, bold=True)
                ws[f'C{row}'].alignment = align_center
                ws[f'D{row}'].font = Font(name='Arial', size=28, bold=True)
                ws[f'D{row}'].alignment = align_center
                ws[f'E{row}'].font = Font(name='Arial', size=28, bold=True)
                # print(row)
                ws[f'E{row}'].alignment = align_center
                ws[f'F{row}'].alignment = align_center
                ws[f'G{row}'].alignment = align_center
                ws[f'H{row}'].alignment = align_center
                ws[f'I{row}'].alignment = align_center
                ws[f'J{row}'].alignment = align_center
                row += 1

            mx_row = 16
            if row > max_row:
                for row_ in ws.iter_rows(min_row=max_row, max_row=row - 1, min_col=1, max_col=max_col):
                    for cell in row_:
                        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                        alignment = copy.copy(cell.alignment)
                        alignment.wrapText = True
                        cell.alignment = align_center
                        # cell.font = font_all

                mx_row = row
            for i in range(6, mx_row):
                ws.row_dimensions[i].height = 60

            if rec.viability == 'viable':
                ws[f'D{mx_row}'] = '☑'
            elif rec.viability == 'not_viable':
                ws[f'F{mx_row}'] = '☑'
            elif rec.viability == 'viable_with_change':
                ws[f'H{mx_row}'] = '☑'

            ws[f'C{mx_row}'] = 'Viable'
            ws[f'E{mx_row}'] = 'Not Viable'
            ws[f'G{mx_row}'] = 'Viable with changes'

        # endregion

        # region SignOff Members Footer
        cur_row = mx_row + 1
        ws.merge_cells(f'A{cur_row}:J{cur_row}')
        cur_row += 1
        ws[f'A{cur_row}'] = 'Prepared By'
        ws[f'A{cur_row}'].font = font_header
        ws.merge_cells(f'A{cur_row}:B{cur_row}')
        ws.merge_cells(f'C{cur_row}:E{cur_row}')

        ws[f'F{cur_row}'] = 'Prepared Date'
        ws[f'F{cur_row}'].font = font_header
        ws.merge_cells(f'F{cur_row}:H{cur_row}')
        ws.merge_cells(f'I{cur_row}:J{cur_row}')

        for rec in self:
            ws[f'C{cur_row}'] = rec.create_uid.name if rec.create_uid else ''
            ws[f'I{cur_row}'] = rec.create_date if rec.create_date else ''

        ws.row_dimensions[cur_row].height = 18
        cur_row += 1
        ws.merge_cells(f'A{cur_row}:J{cur_row}')
        cur_row += 1
        ws.merge_cells(f'A{cur_row}:J{cur_row}')
        ws[f'A{cur_row}'] = 'Sign OFF'
        ws[f'A{cur_row}'].font = Font(size=16, bold=True)
        ws.row_dimensions[cur_row].height = 25
        cur_row += 1
        ws.merge_cells(f'A{cur_row}:J{cur_row}')
        cur_row += 1

        for cell in ws[cur_row]:
            ws[f'A{cur_row}'] = 'Member'
            ws[f'D{cur_row}'] = 'Department'
            ws[f'E{cur_row}'] = 'Status'
            ws[f'F{cur_row}'] = 'Date'
            ws[f'H{cur_row}'] = 'Comments'
            ws.row_dimensions[cur_row].height = 25
            cell.font = Font(size=12, bold=True, color='ffffff')
            cell.fill = PatternFill(start_color='0070C0', end_color='0070C0', fill_type="solid")
            ws.merge_cells(f'A{cur_row}:C{cur_row}')
            ws.merge_cells(f'F{cur_row}:G{cur_row}')
            ws.merge_cells(f'H{cur_row}:J{cur_row}')

        for row_no in ws.iter_rows(min_row=1, max_row=cur_row + 1, min_col=1, max_col=10):
            for cell in row_no:
                cell.border = border
                cell.alignment = align_center

        sign_row = cur_row

        # Listing Managers Id
        for rec in self:
            for record in rec.iatf_members_ids:
                name_rec = record.approver_id.name
                dept_rec = record.department_id.name if record.department_id.name else ''
                status_rec = record.approval_status.capitalize()
                date_rec = record.date_approved_rejected if record.date_approved_rejected else ''
                comment_rec = record.comment if record.comment else ''

                # print(status_rec)

                ws[f'A{cur_row + 1}'] = name_rec
                ws[f'D{cur_row + 1}'] = dept_rec
                ws[f'E{cur_row + 1}'] = status_rec
                ws[f'F{cur_row + 1}'] = date_rec
                ws[f'H{cur_row + 1}'] = comment_rec
                ws.merge_cells(f'A{cur_row + 1}:C{cur_row + 1}')
                ws.merge_cells(f'F{cur_row + 1}:G{cur_row + 1}')
                ws.merge_cells(f'H{cur_row + 1}:J{cur_row + 1}')

                status_dict = {'Approved': ['CFFFC3', '000000'], 'Rejected': ['FFCDCD', '000000'],
                               'Revision': ['AFEFFF', '000000'], 'Pending': ['FDFFCD', '000000']}

                for state, color in status_dict.items():
                    if state == status_rec:
                        ws[f'E{cur_row + 1}'].fill = PatternFill(start_color=color[0], end_color=color[0],
                                                                 fill_type="solid")
                        ws[f'E{cur_row + 1}'].font = Font(size=12, bold=False, color=color[1])

                cur_row += 1

        ws.merge_cells(f'A{cur_row + 1}:J{cur_row + 1}')

        for row_no in ws.iter_rows(min_row=sign_row + 1, max_row=cur_row + 1, min_col=1, max_col=10):
            for cell in row_no:
                cell.border = border
                cell.alignment = align_center
        # endregion

        # Save the workbook
        wb.save(output)
        output.seek(0)
        self.generate_xls_file = base64.b64encode(output.getvalue()).decode('utf-8')

        return {
            "type": "ir.actions.act_url",
            "target": "self",
            "url": "/web/content?model=risk.assessment&download=true&field=generate_xls_file&filename={filename}.xlsx&id={pid}".format(
                filename="Risk Assessment Sheet", pid=self[0].id),
        }


class RiskAssessmentLines(models.Model):
    _name = 'risk.assessment.lines'
    _inherit = "translation.mixin"

    risk_id = fields.Many2one(
        comodel_name='risk.assessment',
        string='Risk Assessment', ondelete='cascade', index=True, copy=False
    )
    risk_analysis_id = fields.Many2one('risk.analysis', "Risk Analysis type", required=True)
    criteria = fields.Char(string="Criteria", compute="_compute_criteria", store=True,translate=True)
    risk_level = fields.Selection(
        [
            ('low', 'Low'),
            ('medium', 'Medium'),
            ('high', 'High'),
        ],
        required=True,
        default='low'
    )
    risk_action_plan = fields.Text('Risk Action Plan', translate=True)
    risk_action_plan_date = fields.Date('Risk Action Plan Date')  # make it required in view
    risk_action_plan_responsible = fields.Many2one('res.users',
                                                   'Risk Action plan Responsible')  # make it required in view
    risk_plan_status = fields.Html('Risk Plan Status/Attachments')

    @api.depends('risk_level', 'risk_analysis_id.criteria_l', 'risk_analysis_id.criteria_m',
                 'risk_analysis_id.criteria_h')
    def _compute_criteria(self):
        for record in self:
            if record.risk_level == 'low':
                record.criteria = record.risk_analysis_id.criteria_l
            elif record.risk_level == 'medium':
                record.criteria = record.risk_analysis_id.criteria_m
            else:
                record.criteria = record.risk_analysis_id.criteria_h
