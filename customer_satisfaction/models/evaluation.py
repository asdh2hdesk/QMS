from odoo import models, fields, api
import base64
import io
from io import BytesIO
import logging
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, PatternFill
from openpyxl.styles import Font, Border, Side
from PIL import Image as PILImage

class CustomerSatisfactionEvaluation(models.Model):
    _name = 'customer.satisfaction.evaluation'
    _description = 'Customer Satisfaction / Expectation Evaluation Form'

    customer_name = fields.Char(string='Customer')
    period_from = fields.Date(string='From Period')
    period_to = fields.Date(string='To Period')

    perception_lines = fields.One2many(
        'customer.satisfaction.perception.line',
        'evaluation_id',
        string='Perception Criteria'
    )

    consider_other_products = fields.Selection(
        [('yes', 'Yes'), ('no', 'No')],
        string='Would you like to consider NMPL for other products?'
    )
    recommend_nmml = fields.Selection(
        [('yes', 'Yes'), ('no', 'No')],
        string='Would you like to recommend NMPL?'
    )
    requirements_unfulfilled = fields.Selection(
        [('yes', 'Yes'), ('no', 'No')],
        string='Any unfulfilled requirements?'
    )
    unfulfilled_details = fields.Text(string='Details if any')

    initiated_by_name = fields.Char(string='Name')
    initiated_by_designation = fields.Char(string='Designation')
    initiated_by_signature = fields.Char(string='Signature')

    responded_by_name = fields.Char(string='Name')
    responded_by_designation = fields.Char(string='Designation')
    responded_by_signature = fields.Char(string='Signature')
    generate_xlsx_file = fields.Binary(string="Generate XLSX File", attachment=True)


    def generate_xlsx_report(self):
        output = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = 'Customer Feedback'

        border = Border(top=Side(style='thin'),left=Side(style='thin'),right=Side(style='thin'),bottom=Side(style='thin'))
        align_center = Alignment(vertical='center', horizontal='center', wrapText=True)
        align_left = Alignment(vertical='center', horizontal='left')
        font_header = Font(name='Arial', size=12, bold=True)
        font_all = Font(name='Times New Roman', size=11, bold=False)


        if self.env.user.company_id.logo:
            max_width, max_height = 150, 60
            image_data = base64.b64decode(self.env.user.company_id.logo)
            image = PILImage.open(io.BytesIO(image_data))
            image.thumbnail((max_width, max_height), PILImage.LANCZOS)
            img_bytes = io.BytesIO()
            image.save(img_bytes, format='PNG')
            logo_image = Image(img_bytes)
            ws.add_image(logo_image, 'B2')
            ws['B2'].alignment = align_center
        
        data = {
            'C2': 'CUSTOMER SATISFACTION / EXPECTATION EVALUATION FORM',
            'B3' : 'Customer',
            'H3' : 'From Period',

        }

        wb.save(output)
        output.seek(0)

        attachment = self.env["ir.attachment"].create({
            "name": "Customer Feedback.xlsx",
            "type": "binary",
            "datas": base64.b64encode(output.getvalue()),
            "res_model": self._name,
            "res_id": self.id,
            "mimetype": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        })
        return {"type": "ir.actions.act_url", "url": f"/web/content/{attachment.id}?download=true", "target": "self"}

class CustomerSatisfactionPerceptionLine(models.Model):
    _name = 'customer.satisfaction.perception.line'
    _description = 'Perception Line Item'

    evaluation_id = fields.Many2one('customer.satisfaction.evaluation', ondelete='cascade')

    perception_criteria = fields.Char(string='Perception Criteria')
    rating = fields.Selection([
        ('d', 'Delighted'),
        ('hs', 'Highly Satisfied'),
        ('s', 'Satisfied'),
        ('us', 'Unsatisfied'),
        ('hus', 'Highly Unsatisfied'),
        ('nr', 'Not Rated')
    ], string='Rating')

    nmpl_need_to_achieve = fields.Text(string='What NMPL Need to Achieve')
    benchmark = fields.Text(string='Benchmark Opinion')
