from odoo import models, fields, api
from datetime import date, datetime, timedelta
import base64
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.drawing.image import Image
try:
    from PIL import Image as PILImage
except ImportError:
    PILImage = None



class CalibrationSchedule(models.Model):
    _inherit = 'calibration.schedule'

    year_range_id = fields.Many2one('calibration.year.range', string='Year Range')


class CalibrationYearRange(models.Model):
    _name = 'calibration.year.range'
    _description = 'Calibration Year Range'

    name = fields.Char('Name', compute='_compute_name', store=True)
    start = fields.Selection(
        selection='_get_year_selection',
        string='Start Year',
        required=True

    )
    end = fields.Selection(
        selection='_get_year_selection',
        string='End Year',
        compute='_compute_end',
        store=True
    )
    color = fields.Integer('Color Index')
    equipment_id = fields.Many2one('maintenance.equipment', string='Instrument')


    code = fields.Char(related='equipment_id.code', string="Code")
    range = fields.Char(related='equipment_id.range', string="Range")
    make = fields.Char(related='equipment_id.make', string="Make")
    lc = fields.Char(related='equipment_id.lc', string="Least Count")
    location = fields.Char(related='equipment_id.location', string="Location")
    cal_freq = fields.Selection([
        ('daily', 'Daily'),
        ('weekly', 'Weekly'),
        ('monthly', 'Monthly'),
        ('yearly', 'Yearly')
    ], string='Frequency')

    interval = fields.Integer(string='Interval')
    schedule_date = fields.Date(string='Calibration Date')
    approved_by = fields.Many2one('res.users', string="Approved By")

    approval_state = fields.Selection([
        ('draft', 'Draft'),
        ('to_approve', 'To Approve'),
        ('approved', 'Approved'),
        ('rejected', 'Rejected'),
    ], string="Approval Status", default="draft")

    status = fields.Selection([
        ('pending', 'Pending'),
        ('done', 'Done'),
        ('not_done', 'Not Done'),
    ], string='Status')  # No default to allow blank status
    remarks = fields.Char(string='Remarks')
    attachment_ids = fields.Many2many(
        'ir.attachment',
        'calibration_year_range_attachment_rel',  # Unique relation table name
        'line_id',
        'attachment_id',
        string='Attachments'
    )

    # Many2many relationship with calibration lines
    instrument_count = fields.Integer(string="Instrument Count", compute="_compute_instrument_count")
    line_ids = fields.One2many("calibration.schedule", "year_range_id", string="Calibration Schedules")

    @api.model
    def create(self, vals):
        # Create the record and set it to 'to_approve' immediately
        record = super(CalibrationYearRange, self).create(vals)
        record.action_submit_for_approval()
        return record

    def write(self, vals, *args, **kwargs):
        # Call the parent write method with vals and any additional args/kwargs
        res = super(CalibrationYearRange, self).write(vals, *args, **kwargs)

        # Check if line_ids are being modified and approval_state is approved/rejected
        if 'line_ids' in vals and self.approval_state in ['approved', 'rejected']:
            self.approval_state = 'to_approve'
            self.action_submit_for_approval()

        return res

    def action_submit_for_approval(self):
        """Submit the entire year range for approval."""
        self.ensure_one()
        if self.approval_state not in ['draft']:
            raise UserError(_('This year range has already been submitted for approval.'))

        self.approval_state = 'to_approve'

        # Notify admin users
        admin_group = self.env.ref('base.group_system')
        admins = admin_group.users


    def action_approve(self):
        """Approve the year range (admin only)."""
        self.ensure_one()
        if not self.env.user.has_group('base.group_system'):
            raise UserError(_('Only administrators can approve calibration year ranges.'))
        if self.approval_state != 'to_approve':
            raise UserError(_('This year range is not in a state to be approved.'))

        self.approval_state = 'approved'
        self.approved_by = self.env.user


    def action_reject(self):
        """Reject the year range (admin only)."""
        self.ensure_one()
        if not self.env.user.has_group('base.group_system'):
            raise UserError(_('Only administrators can reject calibration year ranges.'))
        if self.approval_state != 'to_approve':
            raise UserError(_('This year range is not in a state to be rejected.'))

        self.approval_state = 'rejected'


    @api.depends('start', 'end')
    def _compute_name(self):
        for record in self:
            if record.start and record.end:
                record.name = f"{record.start}-{record.end}"

    @api.depends('start')
    def _compute_end(self):
        for record in self:
            if record.start:
                record.end = str(int(record.start) + 1)

    @api.depends('line_ids')
    def _compute_instrument_count(self):
        for record in self:
            record.instrument_count = len(record.line_ids.mapped('equipment_id')) if record.line_ids else 0

    @api.model
    def _get_year_selection(self):
        current_year = fields.Date.today().year
        return [(str(i), str(i)) for i in range(current_year, current_year + 11)]

    # Ensure approval state consistency (optional, if you want line_ids to influence parent state)
    @api.depends('line_ids.approval_state')
    def _compute_approval_state(self):
        for record in self:
            if record.approval_state in ['approved', 'rejected']:
                continue  # Once approved/rejected, don't auto-update based on lines
            all_schedules = record.line_ids
            if not all_schedules:
                record.approval_state = 'draft'
            elif any(schedule.approval_state == 'rejected' for schedule in all_schedules):
                record.approval_state = 'rejected'
            elif all(schedule.approval_state == 'approved' for schedule in all_schedules):
                record.approval_state = 'approved'
            elif any(schedule.approval_state == 'to_approve' for schedule in all_schedules):
                record.approval_state = 'to_approve'
            else:
                record.approval_state = 'draft'

    @api.depends('line_ids')
    def _compute_instrument_count(self):
        for record in self:
            record.instrument_count = len(record.line_ids.mapped('equipment_id'))

    # Add this method to the CalibrationYearRange class in your first document
    def action_generate_year_range_excel_report(self):
        """Generate an Excel report for all instruments in this year range"""
        self.ensure_one()

        # Get start and end dates for the year range
        start_date = f"{self.start}-04-01"
        end_date = f"{self.end}-03-31"

        # Find all schedules within this year range
        schedules = self.env['calibration.schedule'].search([
            '|',
            ('year_range_id', '=', self.id),
            '&',
            ('schedule_date', '!=', False),
            '&',
            ('schedule_date', '>=', start_date),
            ('schedule_date', '<=', end_date)
        ])

        if not schedules:
            raise UserError('No calibration schedules found for this year range.')

        # Prepare Excel file
        output = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = f"Year Range {self.name}"

        # Styling
        border = Border(
            top=Side(style='thin'),
            left=Side(style='thin'),
            right=Side(style='thin'),
            bottom=Side(style='thin')
        )
        align_center = Alignment(vertical='center', horizontal='center', wrapText=True)
        font_title = Font(name='Arial', size=12, bold=True)
        font_normal = Font(name='Arial', size=10)
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        light_blue_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
        dark_khaki_fill = PatternFill(start_color='BDB76B', end_color='BDB76B', fill_type='solid')



        for row in ws.iter_rows(min_row=1, max_row=10, min_col=1, max_col=4):
            for cell in row:
                cell.border = border


        # Add company logo if available
        if self.env.user.company_id.logo:
            max_width, max_height = 150, 60
            image_data = base64.b64decode(self.env.user.company_id.logo)
            image = PILImage.open(io.BytesIO(image_data))
            image.thumbnail((max_width, max_height), PILImage.LANCZOS)

            img_bytes = io.BytesIO()
            image.save(img_bytes, format='PNG')
            img_bytes.seek(0)

            logo_image = Image(img_bytes)
            ws.add_image(logo_image, 'A1')
            ws.merge_cells('A1:A2')


        # Header
        ws['B1'] = f'Calibration Schedule Report - Year Range {self.name}'
        ws['B1'].font = font_title
        ws['B1'].alignment = align_center
        ws['B1'].fill = dark_khaki_fill

        ws['D1'] = 'Report Date:'
        ws['D1'].font = font_title
        ws['D1'].alignment = align_center
        ws['D2'] = datetime.now().strftime('%Y-%m-%d')
        ws['D2'].alignment = align_center
        ws['D2'].font = font_title


        # Column headers
        headers = {
            'A3': 'Instrument Name',
            'C3': 'Prepared By',
            'D3': 'Next Calibration Date'
        }

        for cell, value in headers.items():
            ws[cell] = value
            ws[cell].font = font_title
            ws[cell].alignment = align_center
            ws[cell].border = border
            ws[cell].fill = light_blue_fill

        # Set column widths
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 40

        # Set row heights
        ws.row_dimensions[1].height = 40
        ws.row_dimensions[2].height = 30
        ws.row_dimensions[3].height = 30

        # Merge cells
        ws.merge_cells('B1:C2')
        ws.merge_cells('A3:B3')

        # Group schedules by instrument
        current_row = 4
        today = fields.Date.today()

        # Get unique instruments
        instruments = {}
        for schedule in schedules:
            for line in schedule.line_ids:
                if line.equipment_id:
                    if line.equipment_id.id not in instruments:
                        instruments[line.equipment_id.id] = {
                            'name': line.equipment_id.name,
                            'schedules': []
                        }
                    instruments[line.equipment_id.id]['schedules'].append({
                        'prepared_by': schedule.prepared_by.name if schedule.prepared_by else 'N/A',
                        'next_date': line.schedule_date,
                        'status': line.status
                    })

        # Write data
        for instrument_id, data in instruments.items():
            # Sort schedules by date and get the next upcoming one
            future_schedules = [s for s in data['schedules'] if s['next_date'] >= today]
            if future_schedules:
                next_schedule = min(future_schedules, key=lambda x: x['next_date'])

                ws.merge_cells(f'A{current_row}:B{current_row}')
                ws[f'A{current_row}'] = data['name']
                ws[f'C{current_row}'] = next_schedule['prepared_by']
                ws[f'D{current_row}'] = next_schedule['next_date'].strftime('%Y-%m-%d')

                # Apply styling
                for col in ['A', 'C', 'D']:
                    cell = ws[f'{col}{current_row}']
                    cell.font = font_normal
                    cell.alignment = align_center
                    cell.border = border



                    # Check if next calibration is within 10 days
                    days_diff = (next_schedule['next_date'] - today).days
                    if 0 <= days_diff <= 10:
                        cell.fill = yellow_fill

                current_row += 1

        # Apply borders to all cells


        # Save and create attachment
        wb.save(output)
        output.seek(0)

        attachment = self.env["ir.attachment"].create({
            "name": f"Year_Range_Calibration_Report_{self.name}.xlsx",
            "type": "binary",
            "datas": base64.b64encode(output.getvalue()).decode("utf-8"),
            "res_model": self._name,
            "res_id": self.id,
            "mimetype": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        })

        if not attachment:
            raise UserError("Error: The report could not be generated.")

        return {
            "type": "ir.actions.act_url",
            "url": f"/web/content/{attachment.id}?download=true",
            "target": "self"
        }




    @api.depends('line_ids', 'start', 'end')
    def _compute_instrument_count(self):
        for record in self:
            # Get all schedules that fall within this year range
            schedules = record.line_ids.filtered(
                lambda l: l.schedule_date and
                          f"{record.start}-01-01" <= l.schedule_date.strftime(
                    '%Y-%m-%d') <= f"{record.end}-12-31"
            )
            record.instrument_count = len(schedules)

    # Method to find all schedules for this year range
    @api.depends('line_ids', 'start', 'end')
    def _compute_instrument_count(self):
        for record in self:
            # Get start and end dates for the year range
            start_date = f"{record.start}-04-01"
            end_date = f"{record.end}-03-31"

            # Find all schedules that fall within this year range
            schedules = self.env['calibration.schedule'].search([
                '|',
                ('year_range_id', '=', record.id),
                '&',
                ('schedule_date', '!=', False),
                '&',
                ('schedule_date', '>=', start_date),
                ('schedule_date', '<=', end_date)
            ])

            record.instrument_count = len(schedules)

    @api.model
    def create(self, vals):
        # Auto-generate name if not provided
        if not vals.get('name') and vals.get('start') and vals.get('end'):
            vals['name'] = f"{vals['end']}-{vals['end']}"
        return super().create(vals)

    def action_view_yearly_calendar(self):
        self.ensure_one()

        # Fetch all calibration schedule lines linked to this year range
        calibration_lines = self.env['calibration.schedule.line'].search([
            '|',
            ('schedule_date', '>=', fields.Date.today()),
            ('schedule_date', '>=', f"{self.start}-04-01"),

        ])

        return {
            'type': 'ir.actions.act_window',
            'name': f'Yearly Calibration Calendar: {self.name}',
            'res_model': 'calibration.schedule.line',
            'view_mode': 'calendar',
            'views': [(False, 'calendar')],
            'domain': [('id', 'in', calibration_lines.ids)],  # Show all schedules in this year range
            'context': {
                'default_year_range_id': self.id,
                'search_default_group_by_equipment': 1,  # Group by instrument
            },
        }

    def delete_schedule(self):
        for rec in self:
            rec.schedule_line_ids.unlink()  # Delete all related schedules
            rec.unlink()  # Delete the main schedule
        self.env.cr.commit()  # Force database update
        return {'type': 'ir.actions.client', 'tag': 'reload'}  # Refresh calendar

    @api.model
    def _get_year_selection(self):
        current_year = fields.Date.today().year
        return [(str(i), str(i)) for i in range(current_year, current_year + 11)]  # Increase range by 1

    @api.depends('start')
    def _compute_end(self):
        for record in self:
            record.name = f"{record.start}-{record.end}"
            if record.start:
                record.end = str(int(record.start) + 1)

    @api.depends('start', 'end')
    def _compute_name(self):
        for record in self:
            if record.start and record.end:
                record.name = f"{record.start}-{record.end}"

    @api.depends('line_ids')
    def _compute_instrument_count(self):
        for record in self:
            record.instrument_count = len(record.line_ids)

    @api.depends('line_ids')
    def _compute_instrument_count(self):
        for record in self:
            record.instrument_count = len(record.line_ids.mapped('equipment_id')) if record.line_ids else 0

    def action_view_calibrations(self):
        self.ensure_one()
        return {
            'name': f'Calibrations for {self.name}',
            'type': 'ir.actions.act_window',
            'res_model': 'calibration.schedule.line',
            'view_mode': 'tree,form,calendar',
            'domain': [('id', 'in', self.line_ids.ids)],
            'context': {'default_year_range_id': self.id},
        }

    def action_view_instrument_calendar(self):
        self.ensure_one()
        return {
            'type': 'ir.actions.act_window',
            'name': f'Calendar View: {self.equipment_id.name}',
            'res_model': 'calibration.schedule.line',
            'view_mode': 'calendar',
            'views': [(False, 'calendar')],
            'domain': [('equipment_id', '=', self.equipment_id.id)],
            'context': {'default_equipment_id': self.equipment_id.id},
        }

