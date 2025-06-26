from email.policy import default

from odoo import models, fields, api, _
from odoo.exceptions import UserError
from datetime import timedelta
from dateutil.relativedelta import relativedelta
from odoo.tools import format_date
import base64
import logging
from odoo import models, fields, api, _
import calendar
from calendar import monthrange
from odoo.exceptions import UserError

from datetime import timedelta, date
import base64
from openpyxl import Workbook

from PIL import Image as PILImage

from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, Border, Side
import io
from io import BytesIO
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
import logging
from datetime import datetime

_logger = logging.getLogger(__name__)


class MaintenanceEquipment(models.Model):
    _inherit = 'maintenance.equipment'

    code = fields.Char(string="Code")
    range = fields.Char(string="Range")
    make = fields.Char(string="Make")
    lc = fields.Char(string="Least Count")
    location = fields.Char(string="Location")


class CalibrationSchedule(models.Model):
    _name = 'calibration.schedule'
    _description = 'Calibration Schedule'
    _inherit = ['mail.thread', 'mail.activity.mixin']  # Add mail thread for better notification support

    prepared_by = fields.Many2one('res.users', string='Prepared By')
    year_range_id = fields.Many2one("calibration.year.range", string="Year Range")
    schedule_date = fields.Date(string='Creation Date', required=True)
    done_date = fields.Date(string='Calibration Done Date')
    line_ids = fields.One2many('calibration.schedule.line', 'schedule_id', string='Calibration Lines')
    equipment_id = fields.Many2one('maintenance.equipment', string='Instrument', related='line_ids.equipment_id',
                                   store=True)
    future_schedule_created = fields.Boolean(string='Future Schedule Created', default=False)

    generate_xlsx_report = fields.Binary(string="Generated XLSX Report", readonly=True)
    _img_bytes_ref = fields.Binary(string="Temporary Image Bytes", store=False)

    status = fields.Selection([
        ('pending', 'Pending'),
        ('done', 'Done'),
        ('not_done', 'Not Done')
    ], string='Status', default='pending', tracking=True)
    cal_freq = fields.Selection(
        [('daily', 'Daily'), ('weekly', 'Weekly'), ('monthly', 'Monthly'), ('yearly', 'Yearly')],
        string='Frequency',
        related='line_ids.cal_freq',
        store=True
    )
    interval = fields.Integer(
        string='Interval',
        related='line_ids.interval',
        store=True,
        default=''
    )



    # New approval fields
    approved_user_id = fields.Many2one('res.users', string='Approved By')

    approval_state = fields.Selection([
        ('draft', 'Draft'),
        ('to_approve', 'To Approve'),
        ('approved', 'Approved'),
        ('rejected', 'Rejected')
    ], string="Approval Status")

    calendar_data = fields.One2many(
        'calibration.schedule.line',
        'schedule_id',
        string='Calendar Data',
        domain="[('equipment_id', '=', equipment_id)]"
    )

    # name = fields.Char(string="Reference", required=True)
    file_name = fields.Char(string="Filename")
    has_upcoming_schedule = fields.Boolean(
        string="Has Upcoming Schedule",
        compute="_compute_has_upcoming_schedule",
        store=False
    )









    @api.depends('line_ids.schedule_date')
    def _compute_has_upcoming_schedule(self):
        for record in self:
            today = fields.Date.today()
            upcoming = False
            for line in record.line_ids:
                if line.schedule_date:
                    days_until = (line.schedule_date - today).days
                    if 0 <= days_until <= 7:
                        upcoming = True
                        break
            record.has_upcoming_schedule = upcoming


    def action_generate_xlsx_report(self):
        output = io.BytesIO()
        wb = Workbook()

        # Remove default sheet to add custom sheets
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

        # Common styling configuration
        border = Border(
            top=Side(style='thin'),
            left=Side(style='thin'),
            right=Side(style='thin'),
            bottom=Side(style='thin')
        )
        align_center = Alignment(vertical='center', horizontal='center', wrapText=True)
        font_title = Font(name='Arial', size=12, bold=True)
        font_normal = Font(name='Arial', size=10)

        # Helper function to create header row
        def create_header_row(ws, frequency_type):
            # Add logo
            if self.env.user.company_id.logo:
                max_width, max_height = 150, 60
                image_data = base64.b64decode(self.env.user.company_id.logo)
                image = PILImage.open(io.BytesIO(image_data))
                image.thumbnail((max_width, max_height), PILImage.LANCZOS)

                img_bytes = io.BytesIO()
                image.save(img_bytes, format='PNG')
                img_bytes.seek(0)

                logo_image = Image(img_bytes)
                ws.add_image(logo_image, 'A1')
                ws.merge_cells('A1:B1')

            # Add calibration schedule header as shown in image
            ws['C1'] = 'Calibration Schedule'
            ws['C1'].font = font_title
            ws['C1'].alignment = align_center

            # Common header data
            ws['A2'] = 'Creation Date:'
            ws['C2'] = datetime.now().strftime('%Y-%m-%d')
            ws['E2'] = 'Prepared By:'
            ws['F2'] = self.env.user.name

            # Common headers for all types
            headers = {
                'A4': 'S.No',
                'B4': 'Instrument Name',
                'C4': 'Code',
                'D4': 'Range',
                'E4': 'Make',
                'F4': 'Least Count',
                'G4': 'Location',
                'H4': 'Frequency',
                'I4': 'Interval',
                'J4': 'Calibration Date',
                'K4': 'Approved By',  # New column
                'L4': 'Status',  # New column
                'M4': 'Remarks'  # New column
            }

            # Apply styling and populate common headers
            for cell, value in headers.items():
                ws[cell] = value
                ws[cell].font = font_title
                ws[cell].alignment = align_center
                ws[cell].border = border

            # Merge common cells
            merge_ranges = [
                'A1:B1', 'C1:Z1', 'A2:B3', 'C2:D3', 'E2:E3', 'F2:M3'  # Updated to include new columns
            ]
            for merge_range in merge_ranges:
                ws.merge_cells(merge_range)

            # Set common dimensions
            ws.row_dimensions[1].height = 55
            ws.row_dimensions[4].height = 50

            ws.column_dimensions['A'].width = 11
            ws.column_dimensions['B'].width = 50
            ws.column_dimensions['C'].width = 18
            ws.column_dimensions['D'].width = 11
            ws.column_dimensions['E'].width = 17
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 11
            ws.column_dimensions['H'].width = 15
            ws.column_dimensions['I'].width = 11
            ws.column_dimensions['J'].width = 20
            ws.column_dimensions['K'].width = 18  # New column
            ws.column_dimensions['L'].width = 15  # New column
            ws.column_dimensions['M'].width = 20  # New column

        # Process each record
        for rec in self:
            # Group records by frequency
            daily_records = []
            weekly_records = []
            monthly_records = []
            yearly_records = []

            for line in rec.line_ids:
                if line.cal_freq == 'daily':
                    daily_records.append(line)
                elif line.cal_freq == 'weekly':
                    weekly_records.append(line)
                elif line.cal_freq == 'monthly':
                    monthly_records.append(line)
                else:  # yearly or others
                    yearly_records.append(line)

            # Create monthly sheets for daily calibrations
            if daily_records:
                for month_num in range(1, 13):
                    month_name = calendar.month_name[month_num]
                    ws = wb.create_sheet(f"{month_name} - Daily")
                    create_header_row(ws, "daily")

                    # Add day columns for this month
                    last_day = calendar.monthrange(datetime.now().year, month_num)[1]
                    for day in range(1, last_day + 1):
                        col_letter = get_column_letter(13 + day)  # Start from column N (14th column)
                        ws[f'{col_letter}4'] = f"{day}"
                        ws[f'{col_letter}4'].font = font_title
                        ws[f'{col_letter}4'].alignment = align_center
                        ws[f'{col_letter}4'].border = border
                        ws.column_dimensions[col_letter].width = 5

                    # Add records to this sheet
                    cur_row = 5
                    for idx, line in enumerate(daily_records, 1):
                        ws[f'A{cur_row}'] = idx
                        ws[f'B{cur_row}'] = line.equipment_id.name if line.equipment_id else ''
                        ws[f'C{cur_row}'] = line.code if hasattr(line, 'code') else ''
                        ws[f'D{cur_row}'] = line.range if hasattr(line, 'range') else ''
                        ws[f'E{cur_row}'] = line.make if hasattr(line, 'make') else ''
                        ws[f'F{cur_row}'] = line.lc if hasattr(line, 'lc') else ''
                        ws[f'G{cur_row}'] = line.location if hasattr(line, 'location') else ''
                        ws[f'H{cur_row}'] = line.cal_freq if hasattr(line, 'cal_freq') else ''
                        ws[f'I{cur_row}'] = line.interval if hasattr(line, 'interval') else ''
                        ws[f'J{cur_row}'] = line.schedule_date.strftime('%Y-%m-%d') if hasattr(line,
                                                                                               'schedule_date') and line.schedule_date else ''
                        ws[f'K{cur_row}'] = ''  # Approved By - Empty field
                        ws[f'L{cur_row}'] = ''  # Status - Empty field
                        ws[f'M{cur_row}'] = ''  # Remarks - Empty field

                        # Apply daily markers based on interval
                        if hasattr(line, 'interval') and line.interval and hasattr(line,
                                                                                   'schedule_date') and line.schedule_date:
                            schedule_date = line.schedule_date
                            interval_days = int(line.interval)

                            # Now we only create markings for the current month's sheet
                            current_month = month_num
                            current_year = datetime.now().year

                            # Find the first calibration date in this month
                            # Start from the schedule date and find the first occurrence in this month
                            first_date_in_month = None
                            temp_date = schedule_date

                            # If the schedule date is in a future month, go backward
                            while temp_date.month > current_month or temp_date.year > current_year:
                                temp_date = temp_date - timedelta(days=interval_days)

                            # If the schedule date is in a past month, go forward
                            while temp_date.month < current_month or temp_date.year < current_year:
                                temp_date = temp_date + timedelta(days=interval_days)

                            # Now we should be in the correct month or just past it
                            if temp_date.month == current_month and temp_date.year == current_year:
                                # We're in the right month, mark all calibration days
                                day = temp_date.day
                                while day <= last_day:
                                    col_letter = get_column_letter(13 + day)
                                    ws[f'{col_letter}{cur_row}'] = "✓"
                                    ws[f'{col_letter}{cur_row}'].font = Font(name='Arial', size=10, color="00AA00")
                                    ws[f'{col_letter}{cur_row}'].alignment = align_center
                                    day += interval_days

                        cur_row += 1

            # Create monthly sheets for weekly calibrations
            if weekly_records:
                for month_num in range(1, 13):
                    month_name = calendar.month_name[month_num]
                    ws = wb.create_sheet(f"{month_name} - Weekly")
                    create_header_row(ws, "weekly")

                    # Calculate the number of weeks in this month
                    year = datetime.now().year
                    first_day = datetime(year, month_num, 1)
                    last_day = datetime(year, month_num, calendar.monthrange(year, month_num)[1])

                    # Get the first and last week numbers of the month
                    first_week = first_day.isocalendar()[1]
                    last_week = last_day.isocalendar()[1]

                    # Handle year boundary (December -> January)
                    if last_week < first_week:
                        last_week += 52

                    # Calculate number of weeks in this month
                    num_weeks = (last_week - first_week) + 1

                    # Add week columns for this month
                    for week_idx in range(num_weeks):
                        week_num = first_week + week_idx
                        if week_num > 52:  # Handle year wrap-around
                            week_num -= 52

                        col_letter = get_column_letter(13 + week_idx + 1)  # Start from column N (14th column)
                        ws[f'{col_letter}4'] = f"{week_num}"
                        ws[f'{col_letter}4'].font = font_title
                        ws[f'{col_letter}4'].alignment = align_center
                        ws[f'{col_letter}4'].border = border
                        ws.column_dimensions[col_letter].width = 5

                    # Add records to this sheet
                    cur_row = 5
                    for idx, line in enumerate(weekly_records, 1):
                        ws[f'A{cur_row}'] = idx
                        ws[f'B{cur_row}'] = line.equipment_id.name if line.equipment_id else ''
                        ws[f'C{cur_row}'] = line.code if hasattr(line, 'code') else ''
                        ws[f'D{cur_row}'] = line.range if hasattr(line, 'range') else ''
                        ws[f'E{cur_row}'] = line.make if hasattr(line, 'make') else ''
                        ws[f'F{cur_row}'] = line.lc if hasattr(line, 'lc') else ''
                        ws[f'G{cur_row}'] = line.location if hasattr(line, 'location') else ''
                        ws[f'H{cur_row}'] = line.cal_freq if hasattr(line, 'cal_freq') else ''
                        ws[f'I{cur_row}'] = line.interval if hasattr(line, 'interval') else ''
                        ws[f'J{cur_row}'] = line.schedule_date.strftime('%Y-%m-%d') if hasattr(line,
                                                                                               'schedule_date') and line.schedule_date else ''
                        ws[f'K{cur_row}'] = ''  # Approved By - Empty field
                        ws[f'L{cur_row}'] = ''  # Status - Empty field
                        ws[f'M{cur_row}'] = ''  # Remarks - Empty field

                        # Apply weekly markers based on interval
                        if hasattr(line, 'interval') and line.interval and hasattr(line,
                                                                                   'schedule_date') and line.schedule_date:
                            schedule_date = line.schedule_date
                            interval_weeks = int(line.interval)

                            # Calculate the week number for this schedule date
                            schedule_week = schedule_date.isocalendar()[1]

                            # For each week in this month
                            for week_idx in range(num_weeks):
                                week_num = first_week + week_idx
                                if week_num > 52:  # Handle year wrap-around
                                    week_num -= 52

                                # Check if this week needs a calibration
                                if (week_num - schedule_week) % interval_weeks == 0:
                                    col_letter = get_column_letter(13 + week_idx + 1)
                                    ws[f'{col_letter}{cur_row}'] = "✓"
                                    ws[f'{col_letter}{cur_row}'].font = Font(name='Arial', size=10, color="00AA00")
                                    ws[f'{col_letter}{cur_row}'].alignment = align_center

                        cur_row += 1

            # Create a single sheet for monthly calibrations
            if monthly_records:
                ws = wb.create_sheet("Monthly Calibrations")
                create_header_row(ws, "monthly")

                # Add month columns
                months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
                for idx, month in enumerate(months, 1):
                    col_letter = get_column_letter(13 + idx)  # Start from column N (14th column)
                    ws[f'{col_letter}4'] = month
                    ws[f'{col_letter}4'].font = font_title
                    ws[f'{col_letter}4'].alignment = align_center
                    ws[f'{col_letter}4'].border = border
                    ws.column_dimensions[col_letter].width = 8

                # Add records
                cur_row = 5
                for idx, line in enumerate(monthly_records, 1):
                    ws[f'A{cur_row}'] = idx
                    ws[f'B{cur_row}'] = line.equipment_id.name if line.equipment_id else ''
                    ws[f'C{cur_row}'] = line.code if hasattr(line, 'code') else ''
                    ws[f'D{cur_row}'] = line.range if hasattr(line, 'range') else ''
                    ws[f'E{cur_row}'] = line.make if hasattr(line, 'make') else ''
                    ws[f'F{cur_row}'] = line.lc if hasattr(line, 'lc') else ''
                    ws[f'G{cur_row}'] = line.location if hasattr(line, 'location') else ''
                    ws[f'H{cur_row}'] = line.cal_freq if hasattr(line, 'cal_freq') else ''
                    ws[f'I{cur_row}'] = line.interval if hasattr(line, 'interval') else ''
                    ws[f'J{cur_row}'] = line.schedule_date.strftime('%Y-%m-%d') if hasattr(line,
                                                                                           'schedule_date') and line.schedule_date else ''
                    ws[f'K{cur_row}'] = ''  # Approved By - Empty field
                    ws[f'L{cur_row}'] = ''  # Status - Empty field
                    ws[f'M{cur_row}'] = ''  # Remarks - Empty field

                    # Apply monthly markers based on interval
                    if hasattr(line, 'interval') and line.interval and hasattr(line,
                                                                               'schedule_date') and line.schedule_date:
                        schedule_date = line.schedule_date
                        interval_months = int(line.interval)

                        # Calculate start month (0-indexed)
                        start_month = schedule_date.month - 1  # Convert to 0-indexed

                        # Mark months based on interval
                        for i in range(0, 12, interval_months):
                            month_idx = (start_month + i) % 12  # Wrap around to stay within 0-11
                            col_letter = get_column_letter(14 + month_idx)  # +14 because N is the 14th column
                            ws[f'{col_letter}{cur_row}'] = "✓"
                            ws[f'{col_letter}{cur_row}'].font = Font(name='Arial', size=10, color="00AA00")
                            ws[f'{col_letter}{cur_row}'].alignment = align_center

                    cur_row += 1

            # Create a single sheet for yearly calibrations
            if yearly_records:
                ws = wb.create_sheet("Yearly Calibrations")
                create_header_row(ws, "yearly")

                # Add month columns
                months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
                for idx, month in enumerate(months, 1):
                    col_letter = get_column_letter(13 + idx)  # Start from column N (14th column)
                    ws[f'{col_letter}4'] = month
                    ws[f'{col_letter}4'].font = font_title
                    ws[f'{col_letter}4'].alignment = align_center
                    ws[f'{col_letter}4'].border = border
                    ws.column_dimensions[col_letter].width = 8

                # Add records
                cur_row = 5
                for idx, line in enumerate(yearly_records, 1):
                    ws[f'A{cur_row}'] = idx
                    ws[f'B{cur_row}'] = line.equipment_id.name if line.equipment_id else ''
                    ws[f'C{cur_row}'] = line.code if hasattr(line, 'code') else ''
                    ws[f'D{cur_row}'] = line.range if hasattr(line, 'range') else ''
                    ws[f'E{cur_row}'] = line.make if hasattr(line, 'make') else ''
                    ws[f'F{cur_row}'] = line.lc if hasattr(line, 'lc') else ''
                    ws[f'G{cur_row}'] = line.location if hasattr(line, 'location') else ''
                    ws[f'H{cur_row}'] = line.cal_freq if hasattr(line, 'cal_freq') else ''
                    ws[f'I{cur_row}'] = line.interval if hasattr(line, 'interval') else ''
                    ws[f'J{cur_row}'] = line.schedule_date.strftime('%Y-%m-%d') if hasattr(line,
                                                                                           'schedule_date') and line.schedule_date else ''
                    ws[f'K{cur_row}'] = ''  # Approved By - Empty field
                    ws[f'L{cur_row}'] = ''  # Status - Empty field
                    ws[f'M{cur_row}'] = ''  # Remarks - Empty field

                    # Mark the calibration month
                    if hasattr(line, 'schedule_date') and line.schedule_date:
                        schedule_month = line.schedule_date.month
                        col_letter = get_column_letter(13 + schedule_month)
                        ws[f'{col_letter}{cur_row}'] = "✓"
                        ws[f'{col_letter}{cur_row}'].font = Font(name='Arial', size=10, color="00AA00")
                        ws[f'{col_letter}{cur_row}'].alignment = align_center

                    cur_row += 1

        # Apply border to all cells in all sheets
        for ws in wb.worksheets:
            # Add the calibration schedule header row above the number columns
            # This only applies to days/weeks columns, not common headers
            first_day_col = 14  # Column N
            last_col = min(ws.max_column, 45)  # Limit to a reasonable number of columns

            if first_day_col < last_col:
                header_row = "Calibration Schedule"
                merge_range = f"{get_column_letter(first_day_col)}3:{get_column_letter(last_col)}3"
                ws[f"{get_column_letter(first_day_col)}3"] = header_row
                ws.merge_cells(merge_range)
                ws[f"{get_column_letter(first_day_col)}3"].font = font_title
                ws[f"{get_column_letter(first_day_col)}3"].alignment = align_center
                ws[f"{get_column_letter(first_day_col)}3"].border = border

            # Apply borders and styles to cells
            for row in ws.iter_rows(min_row=1, max_row=100, min_col=1, max_col=last_col):
                for cell in row:
                    cell.border = border
                    if not cell.font:
                        cell.font = font_normal
                    if not cell.alignment:
                        cell.alignment = align_center

        # Save workbook
        wb.save(output)
        output.seek(0)

        # Create attachment
        attachment = self.env["ir.attachment"].create({
            "name": "Calibration Schedule Report.xlsx",
            "type": "binary",
            "datas": base64.b64encode(output.getvalue()).decode("utf-8"),
            "res_model": self._name,
            "res_id": self.id,
            "mimetype": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        })

        # Ensure Attachment is Created
        if not attachment:
            raise UserError("Error: The report could not be generated.")

        # Return Download URL
        return {
            "type": "ir.actions.act_url",
            "url": f"/web/content/{attachment.id}?download=true",
            "target": "self"
        }

    def view_instrument_calendar(self):
        self.ensure_one()
        if not self.equipment_id:
            raise UserError(_('Please select an instrument first.'))

        return {
            'type': 'ir.actions.act_window',
            'name': f'Calendar View: {self.equipment_id.name}',
            'res_model': 'calibration.schedule.line',
            'view_mode': 'calendar',
            'views': [(False, 'calendar')],
            'domain': [('equipment_id', '=', self.equipment_id.id)],
            'context': {
                'default_equipment_id': self.equipment_id.id,
                'search_default_group_by_equipment': 1,  # Group by instrument
            },
        }

    # Existing methods from the original code
    def generate_future_calibrations(self):
        for schedule in self:
            for line in schedule.line_ids:
                next_date = max(schedule.schedule_date, fields.Date.today())
                end_date = next_date + timedelta(days=365)

                while next_date <= end_date:
                    next_date = self._get_next_calibration_date(line.cal_freq, line.interval, next_date)
                    if next_date.weekday() == 6:  # Skip Sundays
                        next_date += timedelta(days=1)

                    existing_schedule = self.env['calibration.schedule.line'].search([
                        ('equipment_id', '=', line.equipment_id.id),
                        ('schedule_date', '=', next_date)
                    ])

                    if not existing_schedule:
                        self.env['calibration.schedule.line'].create({
                            'schedule_id': schedule.id,
                            'equipment_id': line.equipment_id.id,
                            'code': line.code,
                            'range': line.range,
                            'lc': line.lc,
                            'make': line.make,
                            'location': line.location,
                            'cal_freq': line.cal_freq,
                            'interval': line.interval,
                            'schedule_date': next_date,
                            'status': 'pending',  # Blank status for manual setting later
                            'done_date': line.done_date,
                            'remarks': line.remarks,
                            'display_name': f"{line.equipment_id.name}",
                        })
                        # Trigger the computation of year ranges

                self.env.cr.commit()

    def _get_next_calibration_date(self, frequency, interval, current_date):
        if frequency == 'daily':
            return current_date + timedelta(days=interval)
        elif frequency == 'weekly':
            return current_date + timedelta(weeks=interval)
        elif frequency == 'monthly':
            return current_date + relativedelta(months=interval)
        elif frequency == 'yearly':
            return current_date + relativedelta(years=interval)
        return current_date

    def check_calibration_status(self):
        for schedule in self:
            for line in schedule.line_ids:
                if line.status == 'done':
                    continue

                if fields.Date.today() > line.schedule_date:
                    line.status = 'not_done'
                    next_date = fields.Date.today() + timedelta(days=1)
                    line.schedule_date = next_date
                    self.rearrange_future_schedules(line)
            self.env.cr.commit()

    def rearrange_future_schedules(self, line):
        future_schedules = self.env['calibration.schedule.line'].search([
            ('equipment_id', '=', line.equipment_id.id),
            ('schedule_date', '>', line.schedule_date)
        ], order='schedule_date asc')

        next_date = line.schedule_date
        for future in future_schedules:
            next_date = self._get_next_calibration_date(future.cal_freq, future.interval, next_date)
            future.schedule_date = next_date
        self.env.cr.commit()

    def generate_report(self):
        for schedule in self:
            report_data = []

            pending_schedules = self.env['calibration.schedule.line'].search([
                ('equipment_id', '=', schedule.equipment_id.id),
                ('schedule_date', '>=', fields.Date.today())
            ])

            for record in pending_schedules:
                report_data.append({
                    'Instrument': record.equipment_id.name,
                    'Schedule Date': format_date(self.env, record.schedule_date),
                    'Location': record.location or "N/A",
                    'Remarks': record.remarks or "N/A",
                })

            if not report_data:
                raise UserError(_('No pending or upcoming schedules found for this instrument.'))

            email_body = self._prepare_report_email(report_data)

            if schedule.equipment_id.technician_user_id and schedule.equipment_id.technician_user_id.email:
                self._send_email(schedule.equipment_id.technician_user_id.email, email_body)
            else:
                raise UserError(_('No responsible person found to send the report.'))

    def _prepare_report_email(self, report_data):
        table_content = "".join(f"""
            <tr>
                <td>{data['Instrument']}</td>
                <td>{data['Schedule Date']}</td>
                <td>{data['Location']}</td>
                <td>{data['Remarks']}</td>
            </tr>
        """ for data in report_data)

        return f"""
            <h3>Calibration Schedule Report</h3>
            <table border="1" cellspacing="0" cellpadding="5">
                <tr>
                    <th>Instrument</th>
                    <th>Schedule Date</th>
                    <th>Location</th>
                    <th>Remarks</th>
                </tr>
                {table_content}
            </table>
        """

    def _send_email(self, email_to, email_body):
        mail_values = {
            'subject': 'Weekly Calibration Schedule Report',
            'body_html': email_body,
            'email_to': email_to,
            'email_from': self.env.user.email,
        }
        self.env['mail.mail'].create(mail_values).send()

    def _generate_calibration_report(self, line):
        """Generate calibration report for a specific line"""

        # Calculate next calibration date
        next_date = self._get_next_calibration_date(line.cal_freq, line.interval, line.schedule_date)

        # Skip Sundays
        if next_date.weekday() == 6:
            next_date += timedelta(days=1)

        # Prepare measurement rows dynamically
        standard_lengths = [25, 50, 75, 100, 150]
        measurement_rows_html = ""
        is_satisfactory = True  # Assume satisfactory unless proven otherwise

        for idx, length in enumerate(standard_lengths, 1):
            min_limit = length - 0.05
            max_limit = length + 0.05

            # Fetch actual readings (Ensure line.measurement_readings is a dictionary containing real readings)
            upper = line.measurement_readings.get(idx, {}).get("upper", 0)
            center = line.measurement_readings.get(idx, {}).get("center", 0)
            lower = line.measurement_readings.get(idx, {}).get("lower", 0)

            # Determine if the values are within acceptable limits
            if min_limit <= upper <= max_limit and min_limit <= center <= max_limit and min_limit <= lower <= max_limit:
                status = "Satisfactory"
            else:
                status = "Not Satisfactory"
                is_satisfactory = False  # If one fails, the entire report is not satisfactory

            # Append to HTML
            measurement_rows_html += f"""
            <tr>
                <td style="border: 1px solid #333; padding: 5px; text-align: center;">{idx}</td>
                <td style="border: 1px solid #333; padding: 5px; text-align: center;">{length}</td>
                <td style="border: 1px solid #333; padding: 5px; text-align: center;">{upper}</td>
                <td style="border: 1px solid #333; padding: 5px; text-align: center;">{center}</td>
                <td style="border: 1px solid #333; padding: 5px; text-align: center;">{lower}</td>
                <td style="border: 1px solid #333; padding: 5px; text-align: center;">{min_limit}</td>
                <td style="border: 1px solid #333; padding: 5px; text-align: center;">{max_limit}</td>
                <td style="border: 1px solid #333; padding: 5px; text-align: center; font-weight: bold;">{status}</td>
            </tr>
            """

        # Determine overall conclusion dynamically
        conclusion = "Satisfactory" if is_satisfactory else "Not Satisfactory"

        # Prepare template variables
        template_vars = {
            'measurement_rows': measurement_rows_html,
            'conclusion': conclusion,  # Dynamically computed
            'equipment_name': line.equipment_id.name or "INSTRUMENT",
            'equipment_code': line.code or "N/A",
            'calibration_date': format_date(self.env, line.schedule_date),
            'next_calibration_date': format_date(self.env, next_date),
            'done_by': self.env.user.name,
            'checked_by': line.equipment_id.technician_user_id.name if line.equipment_id.technician_user_id else "____________",
            'current_date': format_date(self.env, fields.Date.today())
        }

        # Render HTML report
        html_report = self.env['ir.qweb']._render('calibration_schedule.calibration_report_template', template_vars)

        # Create an attachment for the PDF
        filename = f"Calibration_Report_{line.equipment_id.name}_{format_date(self.env, line.schedule_date)}.pdf"
        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'type': 'binary',
            'datas': base64.b64encode(html_report.encode('utf-8')),
            'res_model': self._name,
            'res_id': self.id,
            'mimetype': 'application/pdf',
        })

        # Return the PDF as a download
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }


class CalibrationScheduleLine(models.Model):
    _name = 'calibration.schedule.line'
    _description = 'Calibration Schedule Line'
    _rec_name = 'display_name'

    schedule_id = fields.Many2one('calibration.schedule', string='Schedule Reference', required=True,
                                  ondelete='cascade')

    prepared_by = fields.Many2one('res.users', string='Prepared By')
    schedule_date = fields.Date(string='Creation Date', required=True)
    year_range_id = fields.Many2one("calibration.year.range", string="Year Range")

    approval_status = fields.Selection([
        ('to_approve', 'To Approve'),
        ('approved', 'Approved')
    ], string='Approval Status', default='', tracking=True)


    equipment_id = fields.Many2one('maintenance.equipment', string='Instrument', required=True)
    code = fields.Char(related='equipment_id.code', string="Code")
    range = fields.Char(related='equipment_id.range', string="Range")
    make = fields.Char(related='equipment_id.make', string="Make")
    lc = fields.Char(related='equipment_id.lc', string="Least Count")
    location = fields.Char(related='equipment_id.location', string="Location")

    cal_freq = fields.Selection([
        ('daily', 'Daily'),
        ('weekly', 'Weekly'),
        ('monthly', 'Monthly'),
        ('yearly', 'Yearly')
    ], string='Frequency', required=True)

    interval = fields.Integer(string='Interval', required=True)
    done_date = fields.Date(string='Calibration Done Date')
    display_name = fields.Char(string='Display Name', compute='_compute_display_name', store=True)

    approved_by = fields.Many2one('res.users', string="Approved By")

    approval_state = fields.Selection([
        ('draft', 'Draft'),
        ('to_approve', 'To Approve'),
        ('approved', 'Approved'),
        ('rejected', 'Rejected'),
    ], string="Approval Status", default="")


    def action_open_calibration_report(self):
        """Open existing calibration report or create a new one"""
        self.ensure_one()

        existing_report = self.env['calibration.report.wizard'].search([
            ('line_id', '=', self.id)
        ], limit=1)

        button_label = "Create Report" if not existing_report else "View Report"

        return {
            'type': 'ir.actions.act_window',
            'res_model': 'calibration.report.wizard',
            'view_mode': 'form',
            'res_id': existing_report.id if existing_report else None,
            'target': 'new',
            'context': {
                'default_line_id': self.id,
                'button_label': button_label  # Pass button name dynamically
            },
        }

    status = fields.Selection([
        ('pending', 'Pending'),
        ('done', 'Done'),
        ('not_done', 'Not Done'),
    ], string='Status', default='pending')


    remarks = fields.Char(string='Remarks')
    attachment_ids = fields.Many2many(
        'ir.attachment',
        'calibration_line_attachment_rel',
        'line_id',
        'attachment_id',
        string='Attachments'
    )

    def approve_line(self):
        """Approve an individual calibration line"""
        self.ensure_one()
        if self.status != 'done':
            raise UserError(_('Only lines with status "Done" can be approved.'))

        self.approval_status = 'approved'

        # Check if all lines to approve are now approved
        schedule = self.schedule_id
        pending_lines = schedule.line_ids.filtered(lambda l: l.approval_status == 'to_approve')

        if not pending_lines and schedule.approval_state == 'to_approve':
            schedule.approval_state = 'approved'
            schedule.approved_user_id = self.env.user.id
            schedule.approval_date = fields.Date.today()

            # Notify the preparer
            if schedule.prepared_by:
                schedule.message_post(
                    body=_('Your calibration schedule has been fully approved.'),
                    partner_ids=[schedule.prepared_by.partner_id.id],
                    message_type='notification',
                    subject=_('Calibration Schedule Approved')
                )

        return True

    # Store original values
    original_cal_freq = fields.Selection([
        ('daily', 'Daily'),
        ('weekly', 'Weekly'),
        ('monthly', 'Monthly'),
        ('yearly', 'Yearly')
    ], string='Original Frequency', readonly=True)
    original_interval = fields.Integer(string='Original Interval', readonly=True)
    original_schedule_date = fields.Date(string='Original Schedule Date', readonly=True)

    days_until_schedule = fields.Integer(
        string='Days Until Schedule',
        compute='_compute_days_until_schedule',
        store=False,
    )
    is_schedule_today = fields.Boolean(
        string='Is Schedule Today',
        compute='_compute_is_schedule_today',
        store=False,
    )

    @api.depends('schedule_date')
    def _compute_days_until_schedule(self):
        for record in self:
            if record.schedule_date:
                today = fields.Date.today()
                delta = (record.schedule_date - today).days
                record.days_until_schedule = delta
                _logger.info(
                    f"Computed days_until_schedule for record {record.id}: "
                    f"schedule_date={record.schedule_date}, today={today}, delta={delta}"
                )
            else:
                record.days_until_schedule = 0
                _logger.info(f"No schedule_date for record {record.id}, days_until_schedule set to 0")

    @api.depends('schedule_date')
    def _compute_is_schedule_today(self):
        for record in self:
            if record.schedule_date:
                today = fields.Date.today()
                record.is_schedule_today = (record.schedule_date == today)
                _logger.info(
                    f"Computed is_schedule_today for record {record.id}: "
                    f"schedule_date={record.schedule_date}, today={today}, is_schedule_today={record.is_schedule_today}"
                )
            else:
                record.is_schedule_today = False
                _logger.info(f"No schedule_date for record {record.id}, is_schedule_today set to False")

    # Add a computed field to track whether the parent schedule is approved
    is_approved = fields.Boolean(string='Is Approved', compute='_compute_is_approved', store=True)

    @api.depends('schedule_id.approval_state')
    def _compute_is_approved(self):
        for line in self:
            line.is_approved = line.schedule_id.approval_state == 'approved'

    @api.depends('schedule_id.approval_state')
    def _compute_approval_status_visible(self):
        for line in self:
            line.approval_status_visible = line.schedule_id.approval_state in ['to_approve', 'approved']

    approval_status_visible = fields.Boolean(
        string='Approval Status Visible',
        compute='_compute_approval_status_visible',
        store=False
    )
    approved_user_id = fields.Many2one('res.users', string='Approved By')


    def action_submit_for_approval(self):
        for record in self:
            if not record.approved_user_id:
                raise UserError(_('Please select a user to approve this calibration schedule.'))

            # Find lines with status 'done' that aren't already submitted for approval
            done_lines = record.line_ids.filtered(lambda l: l.status == 'done' and l.approval_status == 'not_approved')

            if not done_lines:
                raise UserError(
                    _('There must be at least one completed calibration (status "Done") before submitting for approval.'))

            # Update the lines for approval
            done_lines.write({'approval_status': 'to_approve'})

            # Update the schedule header
            record.approval_state = 'to_approve'

            # Send notification to the approver
            if record.approved_user_id:
                self.message_post(
                    body=_('A calibration schedule requires your approval.'),
                    partner_ids=[record.approved_user_id.partner_id.id],
                    message_type='notification',
                    subject=_('Calibration Schedule Approval Request')
                )

    @api.model
    def create(self, vals):
        # Save original values when creating records
        record = super(CalibrationScheduleLine, self).create(vals)
        if 'cal_freq' in vals:
            record.original_cal_freq = vals['cal_freq']
        if 'interval' in vals:
            record.original_interval = vals['interval']
        if 'schedule_date' in vals:
            record.original_schedule_date = vals['schedule_date']
        return record

    def write(self, vals):
        result = True
        for record in self:
            # Identify if calibration details are changing
            cal_freq_changed = 'cal_freq' in vals and vals['cal_freq'] != record.cal_freq
            interval_changed = 'interval' in vals and vals['interval'] != record.interval
            date_changed = 'schedule_date' in vals and vals['schedule_date'] != record.schedule_date

            # Save original values if they're being modified for the first time
            update_vals = vals.copy()
            if cal_freq_changed and not record.original_cal_freq:
                update_vals['original_cal_freq'] = record.cal_freq
            if interval_changed and not record.original_interval:
                update_vals['original_interval'] = record.interval
            if date_changed and not record.original_schedule_date:
                update_vals['original_schedule_date'] = record.schedule_date

            # Update the current record - IMPORTANT: Don't modify self here
            # Create a new recordset with the context
            ctx_record = record.with_context(schedule_id=record.schedule_id.id if record.schedule_id else False)
            # Call super on this single record with context
            current_result = super(CalibrationScheduleLine, ctx_record).write(update_vals)
            result = result and current_result

            # If frequency, interval, or schedule date is modified, adjust future schedules
            if cal_freq_changed or interval_changed or date_changed:
                # Get the updated values after write
                current_freq = record.cal_freq if not cal_freq_changed else vals['cal_freq']
                current_interval = record.interval if not interval_changed else vals['interval']
                current_date = record.schedule_date if not date_changed else vals['schedule_date']

                # Find all future schedules
                future_schedules = self.env['calibration.schedule.line'].search([
                    ('equipment_id', '=', record.equipment_id.id),
                    ('schedule_date', '>', current_date),
                    ('schedule_id', '=', record.schedule_id.id if record.schedule_id else False)
                ], order='schedule_date asc')

                # Instead of deleting, update them with new dates
                if future_schedules:
                    # Start with the current date
                    next_date = current_date

                    # Update each future schedule using ORM instead of direct SQL
                    for i, future in enumerate(future_schedules):
                        try:
                            # Calculate next date based on current frequency and interval
                            if record.schedule_id:
                                next_date = record.schedule_id._get_next_calibration_date(
                                    current_freq, current_interval, next_date
                                )
                            else:
                                # Fallback if schedule_id is missing
                                if current_freq == 'monthly':
                                    next_date = next_date + relativedelta(months=current_interval)
                                elif current_freq == 'quarterly':
                                    next_date = next_date + relativedelta(months=3 * current_interval)
                                elif current_freq == 'half_yearly':
                                    next_date = next_date + relativedelta(months=6 * current_interval)
                                elif current_freq == 'yearly':
                                    next_date = next_date + relativedelta(years=current_interval)
                                elif current_freq == 'weekly':
                                    next_date = next_date + relativedelta(weeks=current_interval)
                                elif current_freq == 'daily':
                                    next_date = next_date + relativedelta(days=current_interval)

                            # Skip Sundays
                            if next_date.weekday() == 6:
                                next_date += timedelta(days=1)

                            # Set color status based on existing status
                            if future.status == 'done':
                                color_status = "🔵"
                            elif future.status == 'not_done':
                                color_status = "🔴"
                            elif future.status == 'pending':
                                color_status = "🟡"
                            else:
                                color_status = "⚪"

                            # Use ORM write method instead of direct SQL
                            future.write({
                                'cal_freq': current_freq,
                                'interval': current_interval,
                                'schedule_date': next_date,
                                'display_name': f"{color_status} {future.equipment_id.name} - {format_date(self.env, next_date)}"
                            })
                        except Exception as e:
                            _logger.error(f"Error updating schedule line {future.id}: {str(e)}")
                            continue

        return result

    @api.depends('equipment_id', 'schedule_date', 'status', 'approval_status')
    def _compute_display_name(self):
        for record in self:
            if record.equipment_id and record.schedule_date:
                # Status indicator
                if record.status == 'done':
                    color_status = "🔵"
                elif record.status == 'not_done':
                    color_status = "🔴"
                elif record.status == 'pending':
                    color_status = "🟡"
                else:
                    color_status = "⚪"

                # Approval indicator
                if record.approval_status == 'approved':
                    approval_indicator = "✅"
                elif record.approval_status == 'to_approve':
                    approval_indicator = "⏳"
                else:
                    approval_indicator = ""

                record.display_name = f"{color_status} {approval_indicator} {record.equipment_id.name} - {format_date(self.env, record.schedule_date)}"
            else:
                record.display_name = "New Schedule Line"

    # Method to download report for this specific line
    def download_calibration_report(self):
        """Open wizard to enter standard lengths and download report"""
        self.ensure_one()

        return {
            'name': 'Calibration Report',
            'type': 'ir.actions.act_window',
            'res_model': 'calibration.report.wizard',
            'view_mode': 'form',
            'view_type': 'form',
            'target': 'new',
            'context': {
                'default_line_id': self.id,
            }
        }

    # Override name_get to ensure instrument name appears in calendar view
    def name_get(self):
        result = []
        for record in self:
            name = record.equipment_id.name or "Unnamed"
            result.append((record.id, name))
        return result

    def action_reject(self):
        """Reject an individual calibration line"""
        self.ensure_one()
        if self.approval_state != 'to_approve':
            raise UserError(_('Only lines in "To Approve" state can be rejected.'))

        self.write({
            'approval_state': 'rejected',
            'approved_by': False  # Clear approved_by since it's rejected
        })

        # Notify the preparer if exists
        if self.schedule_id.prepared_by:
            self.message_post(
                body=_('Calibration line for %s has been rejected by %s.') % (
                self.equipment_id.name, self.env.user.name),
                partner_ids=[self.schedule_id.prepared_by.partner_id.id],
                message_type='notification',
                subject=_('Calibration Line Rejected')
            )

        # Check if this affects the parent schedule's state
        schedule = self.schedule_id
        if schedule and schedule.approval_state == 'to_approve':
            pending_lines = schedule.line_ids.filtered(lambda l: l.approval_state in ['to_approve', 'draft'])
            if not pending_lines:
                schedule.approval_state = 'rejected'
                schedule.message_post(
                    body=_('Calibration schedule rejected due to a rejected line.'),
                    message_type='notification',
                    subject=_('Calibration Schedule Rejected')
                )

        return True




class CalibrationReportWizard(models.Model):
    _name = 'calibration.report.wizard'
    _description = 'Calibration Report Wizard'

    approval_state = fields.Selection([

        ('to_approve', 'To Approve'),
        ('approved', 'Approved'),
        ('rejected', 'Rejected')
    ], string='Approval Status', default='draft')
    line_id = fields.Many2one('calibration.schedule.line', string='Calibration Line', required=True)

    def action_submit(self):
        """Submit the line for approval."""
        self.ensure_one()
        self.approval_state = 'to_approve'
        # Notify admins
        admin_group = self.env.ref('base.group_system')
        admins = admin_group.users
        self.message_post(
            body=_('Calibration line requires approval.'),
            partner_ids=[admin.partner_id.id for admin in admins],
            message_type='notification',
            subject=_('Approval Request')
        )

    def action_approve(self):
        """Approve the line (admin only)."""
        self.ensure_one()
        if not self.env.user.has_group('base.group_system'):
            raise UserError(_('Only administrators can approve calibration lines.'))
        self.approval_state = 'approved'
        self.approved_by = self.env.user
        self.message_post(body=_('Calibration line approved by %s.') % self.env.user.name)

    def action_reject(self):
        """Reject the line (admin only)."""
        self.ensure_one()
        if not self.env.user.has_group('base.group_system'):
            raise UserError(_('Only administrators can reject calibration lines.'))
        self.approval_state = 'rejected'
        self.message_post(body=_('Calibration line rejected by %s.') % self.env.user.name)


# class CalibrationApproval(models.Model):
#     _name = 'calibration.approval'
#     _description = 'Calibration Approval'
#
#     line_id = fields.Many2one('calibration.schedule.line', string='Calibration Line', required=True)
#     approved_by = fields.Many2one('res.users', string='Approved By')
#     approval_date = fields.Date(string='Approval Date')
#     approval_state = fields.Selection([
#         ('to_approve', 'To Approve'),
#         ('approved', 'Approved'),
#         ('rejected', 'Rejected')
#     ], string='Approval Status', default='to_approve')
#
#     def action_submit_for_approval(self):
#         """Submit the schedule for approval without any conditions."""
#         self.ensure_one()
#         if self.approval_state != 'draft':
#             raise UserError(_('This schedule has already been submitted for approval.'))
#
#         self.approval_state = 'to_approve'
#
#         # Notify admin users
#         admin_group = self.env.ref('base.group_system')
#         admins = admin_group.users
#         self.message_post(
#             body=_('Calibration schedule for %s requires approval.') % (self.equipment_id.name or "Unnamed Instrument"),
#             partner_ids=[admin.partner_id.id for admin in admins],
#             message_type='notification',
#             subject=_('Approval Request for Calibration Schedule')
#         )
#
#         # Update the parent year range's state
#         self._update_parent_year_range_state()
#
#     def action_approve(self):
#         """Approve the schedule (admin only)."""
#         self.ensure_one()
#         if not self.env.user.has_group('base.group_system'):
#             raise UserError(_('Only administrators can approve calibration schedules.'))
#         if self.approval_state != 'to_approve':
#             raise UserError(_('This schedule is not in a state to be approved.'))
#
#         self.approval_state = 'approved'
#         self.approved_user_id = self.env.user
#         self.message_post(
#             body=_('Calibration schedule for %s approved by %s.') % (self.equipment_id.name or "Unnamed Instrument", self.env.user.name),
#             message_type='notification',
#             subject=_('Calibration Schedule Approved')
#         )
#
#         # Update the parent year range's state
#         self._update_parent_year_range_state()
#
#     def action_reject(self):
#         """Reject the schedule (admin only)."""
#         self.ensure_one()
#         if not self.env.user.has_group('base.group_system'):
#             raise UserError(_('Only administrators can reject calibration schedules.'))
#         if self.approval_state != 'to_approve':
#             raise UserError(_('This schedule is not in a state to be rejected.'))
#
#         self.approval_state = 'rejected'
#         self.message_post(
#             body=_('Calibration schedule for %s rejected by %s.') % (self.equipment_id.name or "Unnamed Instrument", self.env.user.name),
#             message_type='notification',
#             subject=_('Calibration Schedule Rejected')
#         )
#
#         # Update the parent year range's state
#         self._update_parent_year_range_state()
#
#     def _update_parent_year_range_state(self):
#         """Update the state of the parent CalibrationYearRange based on the schedules' states."""
#         year_range = self.year_range_id
#         if not year_range:
#             return
#
#         if hasattr(year_range, 'approval_state'):
#             all_schedules = year_range.line_ids
#             if not all_schedules:
#                 year_range.approval_state = 'draft'
#                 return
#
#             if any(schedule.approval_state == 'rejected' for schedule in all_schedules):
#                 year_range.approval_state = 'rejected'
#                 year_range.message_post(
#                     body=_('Year range schedule rejected due to a rejected schedule.'),
#                     message_type='notification',
#                     subject=_('Year Range Schedule Rejected')
#                 )
#             elif all(schedule.approval_state == 'approved' for schedule in all_schedules):
#                 year_range.approval_state = 'approved'
#                 year_range.message_post(
#                     body=_('Year range schedule fully approved.'),
#                     message_type='notification',
#                     subject=_('Year Range Schedule Approved')
#                 )
#             elif any(schedule.approval_state == 'to_approve' for schedule in all_schedules):
#                 year_range.approval_state = 'to_approve'
#             else:
#                 year_range.approval_state = 'draft'
#
#
#

