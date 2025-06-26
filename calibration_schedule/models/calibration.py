
from odoo import models,fields,api,_
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
import base64
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from PIL import Image as PILImage, ImageOps
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
import io

from odoo.exceptions import ValidationError, UserError
# Import required libraries
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph,Spacer
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.lib.styles import getSampleStyleSheet
from io import BytesIO
from datetime import date

from odoo.tools import format_date


class Calibration(models.Model):
    _name='calibration.sheet'
    _description='Calibration Work Sheet'


    name=fields.Char('Name',compute='_compute_name', store=True)
    start = fields.Selection(
        selection='_get_year_selection',
        string='Start Year',
        required=True
    )
    end = fields.Selection(
        selection='_get_year_selection',
        string='End Year',
        compute='_compute_end',
        store=True
    )

    instrument_count = fields.Integer(string="Instrument Count",compute='_compute_instrument_count',store=True)
    approval_state = fields.Selection([
        ('draft', 'Draft'),
        ('to_approve', 'To Approve'),
        ('approved', 'Approved'),
        ('rejected', 'Rejected'),
    ], string="Approval Status", default="draft",tracking=True)
    cal_sheet_ids=fields.One2many('calibration.sheet.schedule','sched_id',string="calibration")
    approved_by = fields.Many2one('res.users', string='Approved By')
    generate_xlsx_report = fields.Binary(string="Generated XLSX Report", readonly=True)

    def generate_xls_report(self):
        # Create workbook and worksheet
        output = BytesIO()
        wb = Workbook()
        # Remove default sheet
        wb.remove(wb.active)


        # Define styles
        border = Border(
            top=Side(style='thin'), left=Side(style='thin'),
            right=Side(style='thin'), bottom=Side(style='thin')
        )
        font_header = Font(name='Times New Roman', bold=True)
        title_font = Font(size=20, bold=True)
        align_center = Alignment(vertical='center', horizontal='center', wrapText=True)
        fill = PatternFill(start_color="e7e7e7", end_color="e7e7e7", fill_type="solid")

        # Get financial year months (April to March)
        start_year = int(self.start)
        months = [
            ('April', 4, start_year), ('May', 5, start_year), ('June', 6, start_year),
            ('July', 7, start_year), ('August', 8, start_year), ('September', 9, start_year),
            ('October', 10, start_year), ('November', 11, start_year), ('December', 12, start_year),
            ('January', 1, start_year + 1), ('February', 2, start_year + 1), ('March', 3, start_year + 1)
        ]

        # Create a sheet for each month
        for month_name, month_num, year in months:
            ws = wb.create_sheet(title=f"{month_name} {year}")

            # Add company logo if available
            if self.env.user.company_id and self.env.user.company_id.logo:
                max_width, max_height = 100, 200
                image_data = base64.b64decode(self.env.user.company_id.logo)
                image = PILImage.open(BytesIO(image_data))

                # Resize with aspect ratio
                width, height = image.size
                aspect_ratio = width / height
                if width > max_width:
                    width = max_width
                    height = int(width / aspect_ratio)
                if height > max_height:
                    height = max_height
                    width = int(height * aspect_ratio)

                # Resize and add padding
                resized_image = image.resize((width, height), PILImage.LANCZOS)
                resized_image = ImageOps.expand(resized_image, border=(5, 5, 0, 0), fill='white')
                img_bytes = BytesIO()
                resized_image.save(img_bytes, format='PNG')
                img_bytes.seek(0)
                logo_image = Image(img_bytes)
                ws.add_image(logo_image, 'A1')

            # Headers dictionary
            headers = {
                'C1': f'CALIBRATION SCHEDULE FOR {month_name.upper()} {year}',
                'A5': 'Period',
                'E5': 'Start Year',
                'I5': 'End Year',
                'M5': 'Prepared By',
                'Q5': 'Approval Status',
                'A6': 'S.No',
                'B6': 'Instrument Name',
                'C6': 'Instrument Code',
                'D6': 'Instrument Range',
                'E6': 'Instrument Make',
                'F6': 'Instrument Least Count',
                'G6': 'Instrument Location',
                'H6': 'Calibration Frequency',
                'I6': 'Interval',
                'J6': 'Calibration Schedule',
            }


            for i in range(1, 32):
                column_letter = get_column_letter(10 + i - 1)  # J is the 10th column, adjust index
                headers[f'{column_letter}7'] = f'Day {i}'
            # Add additional columns
            headers.update({
                'AO6': 'Calibration Complete Date',
                'AP6': 'Status',
                'AQ6': 'Approved By',
                'AR6': 'Remarks',
            })

            # Set row height
            ws.row_dimensions[1].height = 50

            # Apply font and border styles
            for cell, value in headers.items():
                ws[cell] = value
                ws[cell].font = font_header
                ws[cell].alignment = align_center
                ws[cell].fill = fill
            ws['C1'].font = title_font

            # Merge necessary cells
            merge_cells = [
                'A1:B4', 'C1:AR4', 'A6:A7', 'B6:B7', 'C6:C7', 'D6:D7',
                'E6:E7', 'F6:F7', 'G6:G7', 'H6:H7', 'I6:I7', 'A5:B5','C5:D5','E5:F5','G5:H5','I5:J5','K5:L5','M5:N5','O5:P5','Q5:R5','S5:T5',
                'AP6:AP7', 'AQ6:AQ7', 'AR6:AR7', 'AO6:AO7', 'J6:AN6'
            ]
            for cell_range in merge_cells:
                ws.merge_cells(cell_range)

            # Set column widths
            col_widths = {
                'A': 8, 'B': 25, 'C': 18, 'D': 18, 'E': 18,
                'F': 18, 'G': 18, 'H': 18, 'I': 18, 'AO': 18, 'AP': 18, 'AQ': 18, 'AR': 18,
            }

            # Add columns J through AN with width 15
            for col_idx in range(10, 40):  # 10 is J, 39 is AN (1-based index)
                col_letter = get_column_letter(col_idx)
                col_widths[col_letter] = 15

            # Apply all column widths
            for col, width in col_widths.items():
                ws.column_dimensions[col].width = width

            # Apply formatting for all rows up to row 30
            for rows in ws.iter_rows(min_row=1, max_row=30, min_col=1, max_col=44):
                for cell in rows:
                    cell.alignment = align_center
                    cell.border = border

            for rec in self:
                ws['C5'] = rec.name if rec.name else ''
                ws['G5'] = rec.start if rec.start else ''
                ws['K5'] = rec.end if rec.end else ''
                ws['O5'] = rec.approved_by.name if rec.approved_by else ''
                ws['S5'] = rec.approval_state if rec.approval_state else ''

                # Populate instrument data and schedule dates
                row_num = 8  # Start from row 8 for data

                # Get all unique instruments from calibration sheet schedules
                instruments_done = set()
                today = fields.Date.today()
                start_of_week = today - timedelta(days=today.weekday())  # Monday
                end_of_week = start_of_week + timedelta(days=6)

                for cal_sheet in rec.cal_sheet_ids:
                    # Skip if this instrument has already been processed
                    if cal_sheet.equipment_id.id in instruments_done:
                        continue

                    instruments_done.add(cal_sheet.equipment_id.id)

                    # Fill instrument details
                    ws[f'A{row_num}'] = row_num - 7  # S.No
                    ws[f'B{row_num}'] = cal_sheet.equipment_id.name or ''
                    ws[f'C{row_num}'] = cal_sheet.code or ''
                    ws[f'D{row_num}'] = cal_sheet.equipment_id.range or ''
                    ws[f'E{row_num}'] = cal_sheet.equipment_id.make or ''
                    ws[f'F{row_num}'] = cal_sheet.equipment_id.lc or ''
                    ws[f'G{row_num}'] = cal_sheet.location or ''
                    ws[f'H{row_num}'] = dict(cal_sheet._fields['cal_freq'].selection).get(cal_sheet.cal_freq, '')
                    ws[f'I{row_num}'] = cal_sheet.interval or ''

                    # Check all scheduled dates for this instrument and mark with 'X' if in current month
                    for schedule in cal_sheet.scd_ids:
                        if schedule.scheduled_date:
                            sched_date = schedule.scheduled_date

                            # Check if the date is in the current month and year being processed
                            if sched_date.month == month_num and sched_date.year == year:
                                day_col = get_column_letter(10 + sched_date.day - 1)  # Calculate column for the day
                                ws[f'{day_col}{row_num}'] = 'X'  # Mark with X
                                # Determine fill color based on current week
                                if start_of_week <= sched_date <= end_of_week:
                                    # Orange fill for current week
                                    week_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                                else:
                                    # Blue fill for other dates
                                    week_fill = PatternFill(start_color="4492BD", end_color="0000FF", fill_type="solid")

                                ws[f'{day_col}{row_num}'].fill = week_fill

                                # Add status if available
                                if schedule.status:
                                    status_text = dict(schedule._fields['status'].selection).get(schedule.status, '')
                                    ws[f'AP{row_num}'] = status_text

                                # Add completion date if available
                                if schedule.completion_date:
                                    ws[f'AO{row_num}'] = schedule.completion_date

                                # Add approved by if available
                                if schedule.approved_by:
                                    ws[f'AQ{row_num}'] = schedule.approved_by.name

                                # Add remarks if available
                                if schedule.remarks:
                                    ws[f'AR{row_num}'] = schedule.remarks

                    row_num += 1

        # Save workbook to BytesIO
        wb.save(output)
        output.seek(0)

        # Create attachment
        attachment = self.env['ir.attachment'].create({
            'name': f'Calibration_Schedule_{self.start}_{self.end}.xlsx',
            'type': 'binary',
            'datas': base64.b64encode(output.getvalue()),
            'res_model': 'calibration.sheet',
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        })

        # Return download link
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }

    def action_submit_for_approve(self):
        """
        Submit calibration sheet for approval
        """
        for record in self:
            # Ensure only draft records can be submitted
            if record.approval_state == 'draft':
                record.approval_state = 'to_approve'
            else:
                raise ValidationError(_("Only draft calibration sheets can be submitted for approval."))

    def action_approve(self):
        """
        Approve the calibration sheet
        """
        for record in self:
            # Ensure only 'to_approve' records can be approved
            if record.approval_state == 'to_approve':
                record.approval_state = 'approved'
                record.approved_by = self.env.user
            else:
                raise ValidationError(_("Only calibration sheets in 'To Approve' status can be approved."))

    def action_reject(self):
        """
        Reject the calibration sheet
        """
        for record in self:
            # Ensure only 'to_approve' records can be rejected
            if record.approval_state == 'to_approve':
                record.approval_state = 'rejected'
            else:
                raise ValidationError(_("Only calibration sheets in 'To Approve' status can be rejected."))



    @api.depends('start', 'end')
    def _compute_name(self):
        for record in self:
            if record.start and record.end:
                record.name = f"{record.start}-{record.end}"

    @api.depends('start')
    def _compute_end(self):
        for record in self:
            if record.start:
                record.end = str(int(record.start) + 1)

    @api.model
    def _get_year_selection(self):
        current_year = fields.Date.today().year
        return [(str(i), str(i)) for i in range(current_year, current_year + 11)]

    @api.depends('cal_sheet_ids.equipment_id')  # Track changes in the child model
    def _compute_instrument_count(self):
        for record in self:
            unique_instruments = set(record.cal_sheet_ids.mapped('equipment_id.id'))  # Get unique instrument IDs
            record.instrument_count = len(unique_instruments)



class CalibrationSchedule(models.Model):
    _name='calibration.sheet.schedule'
    _description='Calibration Sheet Schedule'


    sched_id=fields.Many2one('calibration.sheet',string="Calibration Sheet",ondelete="cascade")
    prepared_by = fields.Many2one('res.users', string='Prepared By')
    equipment_id = fields.Many2one('maintenance.equipment', string='Instrument')
    cal_freq = fields.Selection(
        [('daily', 'Daily'), ('weekly', 'Weekly'), ('monthly', 'Monthly'), ('yearly', 'Yearly')],
        string='Frequency',
        store=True
    )
    interval = fields.Integer(
        string='Interval',
        store=True,
        default=''
    )
    current_date = fields.Date(string="Current Date")
    code = fields.Char(string="Code", related="equipment_id.code", store=True)
    scd_ids=fields.One2many('calibration.sheet.generate','gen_id',string="Calibration gen")
    location = fields.Char(string="Location", related="equipment_id.location", store=True)


    _sql_constraints = [
        ('unique_equipment_per_sheet', 'UNIQUE(sched_id, equipment_id)',
         'Each instrument should be unique per calibration sheet!')
    ]

    def generate_future_calibrations(self):
        """
        Generate future calibration schedules based on the calibration frequency
        up to the end of the next financial year (March 31)
        """
        for record in self:
            # Determine the end date (next financial year's March 31)
            today = fields.Date.today()
            current_year = today.year
            financial_year_end = datetime(current_year + 1, 3, 31).date()

            # Check if equipment and frequency are set
            if not record.equipment_id or not record.cal_freq:
                continue

            # Clear existing future schedules
            record.scd_ids.unlink()

            # Determine the starting date
            current_date = record.current_date

            # Generate schedules based on frequency
            while current_date <= financial_year_end:
                # Create a new calibration generation record
                self.env['calibration.sheet.generate'].create({
                    'gen_id': record.id,
                    'scheduled_date': current_date,
                    'status': 'not_started',
                })

                # Calculate next scheduled date based on frequency
                if record.cal_freq == 'daily':
                    current_date += timedelta(days=record.interval or 1)
                elif record.cal_freq == 'weekly':
                    current_date += timedelta(weeks=record.interval or 1)
                elif record.cal_freq == 'monthly':
                    current_date += relativedelta(months=record.interval or 1)
                elif record.cal_freq == 'yearly':
                    current_date += relativedelta(years=record.interval or 1)





class CalibrationGenerate(models.Model):
    _name='calibration.sheet.generate'
    _description='calibration Generation'

    gen_id=fields.Many2one('calibration.sheet.schedule',string="Genarate",ondelete="cascade")
    equipment_id = fields.Many2one('maintenance.equipment',related="gen_id.equipment_id" ,string='Instrument',store=True)
    # Auto-fetch fields from maintenance.equipment
    code = fields.Char(string="Code", related="equipment_id.code", store=True)
    range = fields.Char(string="Range", related="equipment_id.range", store=True)
    make = fields.Char(string="Make", related="equipment_id.make", store=True)
    lc = fields.Char(string="Least Count", related="equipment_id.lc", store=True)
    location = fields.Char(string="Location", related="equipment_id.location", store=True)
    gauge_type = fields.Selection(
        [('attribute', 'Attribute'), ('variable', 'Variable')], string='Gauge Type')
    cal_freq = fields.Selection(
        [('daily', 'Daily'), ('weekly', 'Weekly'), ('monthly', 'Monthly'), ('yearly', 'Yearly')],
        string='Frequency',related="gen_id.cal_freq",
        store=True
    )
    interval = fields.Integer(string='Interval',related="gen_id.interval" ,store=True)
    scheduled_date = fields.Date(string="Scheduled Date")
    completion_date=fields.Date(string="Calibration Completion Date")
    remarks=fields.Text(string="Remarks")
    status = fields.Selection([
        ('not_started', 'Not Started'),
        ('pending', 'Pending'),
        ('done', 'Done'),
        ('not_done', 'Not Done')
    ], string='Status', default='not_started')
    approval_state = fields.Selection([
        ('draft', 'Draft'),
        ('to_approve', 'To Approve'),
        ('approved', 'Approved'),
        ('rejected', 'Rejected'),
    ], string="Approval Status", default="draft",compute="_compute_approval_state",store=True)
    attachment=fields.Binary(string="Attachment")
    approved_by=fields.Many2one('res.users', string='Approved By')
    report_ids = fields.One2many('calibration.report', 'report_id', string='Calibration Reports')
    # Add these fields
    current_date = fields.Date(string="Current Date", compute="_compute_reference_dates")
    next_week_date = fields.Date(string="Next Week Date", compute="_compute_reference_dates")

    # Add this method
    @api.depends('scheduled_date')
    def _compute_reference_dates(self):
        today = fields.Date.today()
        next_week = today + timedelta(days=7)

        for record in self:
            record.current_date = today
            record.next_week_date = next_week

    @api.depends('report_ids.approval_state')
    def _compute_approval_state(self):
        for record in self:
            if record.approval_state in ['approved', 'rejected']:
                continue  # Once approved/rejected, don't auto-update based on lines
            all_schedules = record.report_ids
            if not all_schedules:
                record.approval_state = 'draft'
            elif any(schedule.approval_state == 'rejected' for schedule in all_schedules):
                record.approval_state = 'rejected'
            elif all(schedule.approval_state == 'approved' for schedule in all_schedules):
                record.approval_state = 'approved'
            elif any(schedule.approval_state == 'to_approve' for schedule in all_schedules):
                record.approval_state = 'to_approve'
            else:
                record.approval_state = 'draft'

    def action_open_calibration_report(self):
        """
        Open existing calibration report or create a new one
        """
        self.ensure_one()

        # Search for an existing calibration report for this record
        existing_report = self.env['calibration.report'].search([
            ('report_id', '=', self.id)
        ], limit=1)

        # If no existing report, create a new one
        if not existing_report:
            existing_report = self.env['calibration.report'].create({
                'report_id': self.id,
                'approval_state': 'draft'
            })

        # Return action to open the report
        return {
            'type': 'ir.actions.act_window',
            'name': 'Calibration Report',
            'res_model': 'calibration.report',
            'view_mode': 'form',
            'res_id': existing_report.id,
            'target': 'new',  # or 'new' if you want it in a popup
            'context': {
                'default_report_id': self.id
            }
        }

class CalibrationReport(models.Model):
    _name = 'calibration.report'
    _description = 'Calibration Report'

    equipment_id = fields.Many2one('maintenance.equipment', related="report_id.equipment_id", string='Instrument',
                                   store=True)


    report_id=fields.Many2one('calibration.sheet.generate', string='Calibration Record')
    next_cal_date=fields.Date('Next Calibration Date',compute='_compute_next_calibration_date')
    conclusion=fields.Selection([('satisfactory','Satisfactory'),('not satisfactory','Not Satisfactory')],string="Conclusion")
    approved_by=fields.Many2one('res.users', string='Approved By')
    approval_state = fields.Selection([
        ('draft', 'Draft'),
        ('to_approve', 'To Approve'),
        ('approved', 'Approved'),
        ('rejected', 'Rejected')
    ], string="Approval Status", default="draft")
    cal_freq = fields.Selection(
        [('daily', 'Daily'), ('weekly', 'Weekly'), ('monthly', 'Monthly'), ('yearly', 'Yearly')],
        string='Frequency', related="report_id.cal_freq",
        store=True
    )

    line_ids=fields.One2many('calibration.report.line','line_id',string='report')

    def generate_calibration_report_pdf(self):

        self.ensure_one()

        if not self.report_id:
            raise UserError(_('No calibration line selected.'))



        # Generate measurement rows HTML
        measurement_rows_html = ""

        for length in self.line_ids:
            measurement_rows_html += f"""
            <tr>
                <td style="border: 1px solid #333; padding: 5px; text-align: center;">{length.sn}</td>
                <td style="border: 1px solid #333; padding: 5px; text-align: center;">{length.standard_length}</td>
                <td style="border: 1px solid #333; padding: 5px; text-align: center;">{length.upper_reading or ''}</td>
                <td style="border: 1px solid #333; padding: 5px; text-align: center;">{length.center_reading or ''}</td>
                <td style="border: 1px solid #333; padding: 5px; text-align: center;">{length.lower_reading or ''}</td>
                <td style="border: 1px solid #333; padding: 5px; text-align: center;">{length.avg_reading or ''}</td>
                <td style="border: 1px solid #333; padding: 5px; text-align: center;">{length.min_lim}</td>
                <td style="border: 1px solid #333; padding: 5px; text-align: center;">{length.max_lim}</td>
            </tr>
            """

        # Prepare attachments HTML
        attachments_html = ""
        if self.report_id.attachment:
            attachments_html = "<div style='margin-top: 20px;'><h3>Attachments</h3><ul>"
            for attachment in self.report_id.attachment:
                attachments_html += f"<li>{attachment.name}</li>"
            attachments_html += "</ul></div>"

        # Get frequency text
        freq_text = self.report_id.cal_freq
        interval = self.report_id.interval

        # Handle singular and plural properly
        freq_text = self.report_id.cal_freq
        interval = self.report_id.interval

        if interval > 1:
            if freq_text.endswith("ily"):  # Handle "daily" -> "days"
                freq_text = f"Every {interval} {freq_text[:-3]}ys"
            elif freq_text.endswith("y"):  # Handle "monthly" -> "months", "weekly" -> "weeks"
                freq_text = f"Every {interval} {freq_text[:-1]}ies"
            else:
                freq_text = f"Every {interval} {freq_text}s"  # General case
        else:
            freq_text = f"Every {freq_text}"

        # Format template variables
        template_vars = {
            'equipment_name': self.report_id.equipment_id.name or "INSTRUMENT",
            'equipment_code': self.report_id.code,
            'calibration_date': format_date(self.env, self.report_id.scheduled_date),
            'calibration_frequency': freq_text,
            'instrument_number': self.report_id.equipment_id.serial_no or "N/A",
            'make': self.report_id.make,
            'calibration_instructions': "Place the standard length between measuring jaws of the caliper at three different positions and determine the length.",
            'measurement_rows': measurement_rows_html,
            'next_calibration_date': format_date(self.env, self.next_cal_date),
            'done_by': self.env.user.name,
            'checked_by': self.approved_by.name if self.approved_by else "____________",
            'current_date': format_date(self.env, fields.Date.today()),
            'attachments_html': attachments_html,
            'approval_state':self.approval_state,
            'conclusion': 'satisfactory' if self.conclusion == 'satisfactory' else 'not satisfactory'
        }

        # Render HTML report
        html_report = self.env['ir.qweb']._render('calibration_schedule.calibration_report_template', template_vars)

        try:
            # Try to use wkhtmltopdf first
            pdf_content = self.env['ir.actions.report']._run_wkhtmltopdf(
                [html_report],
                specific_paperformat_args={
                    'data-report-margin-top': 10,
                    'data-report-margin-bottom': 10,
                    'data-report-margin-left': 10,
                    'data-report-margin-right': 10,
                }
            )
        except Exception as e:
            # Create a plain HTML file as fallback if PDF generation fails
            filename = f"Calibration_Report_{self.report_id.equipment_id.name}.html"

            # Create an attachment for the HTML content
            attachment = self.env['ir.attachment'].create({
                'name': filename,
                'type': 'binary',
                'datas': base64.b64encode(html_report.encode('utf-8')),
                'res_model': 'calibration.schedule.line',
                'res_id': self.report_id.id,
                'mimetype': 'text/html',
            })

            # Return the HTML as a download
            return {
                'type': 'ir.actions.act_url',
                'url': f'/web/content/{attachment.id}?download=true',
                'target': 'self',
            }

        # Create an attachment for the PDF
        filename = f"Calibration_Report_{self.report_id.equipment_id.name}.pdf"
        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'type': 'binary',
            'datas': base64.b64encode(pdf_content),
            'res_model': 'calibration.schedule.line',
            'res_id': self.report_id.id,
            'mimetype': 'application/pdf',
        })

        # Return the PDF as a download
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }

    def action_submit_for_approval(self):
        """
        Submit calibration report for approval
        """
        for record in self:
            # Ensure only draft records can be submitted
            if record.approval_state== 'draft':
                record.approval_state= 'to_approve'
            else:
                raise ValidationError(_("Only draft calibration reports can be submitted for approval."))

            return {
                'type': 'ir.actions.act_window',
                'res_model': 'calibration.report',
                'view_mode': 'form',
                'res_id': self.id,
                'views': [(False, 'form')],
                'target': 'new',
                'flags': {'mode': 'readonly'},  # Prevents further edits while in approval
            }

    def action_approve(self):
        """
        Approve the calibration report
        """
        for record in self:
            # Ensure only 'to_approve' records can be approved
            if record.approval_state== 'to_approve':
                record.approval_state= 'approved'
                record.approved_by = self.env.user
            else:
                raise ValidationError(_("Only calibration reports in 'To Approve' status can be approved."))
            return {
                'type': 'ir.actions.act_window',
                'res_model': 'calibration.report',
                'view_mode': 'form',
                'res_id': self.id,
                'views': [(False, 'form')],
                'target': 'new',
                'flags': {'mode': 'readonly'},  # Prevents further edits while in approval
            }

    def action_reject(self):
        """
        Reject the calibration report
        """
        for record in self:
            # Ensure only 'to_approve' records can be rejected
            if record.approval_state== 'to_approve':
                record.approval_state= 'rejected'
            else:
                raise ValidationError(_("Only calibration reports in 'To Approve' status can be rejected."))
            return {
                'type': 'ir.actions.act_window',
                'res_model': 'calibration.report',
                'view_mode': 'form',
                'res_id': self.id,
                'views': [(False, 'form')],
                'target': 'new',
                'flags': {'mode': 'readonly'},  # Prevents further edits while in approval
            }




    @api.depends('report_id', 'report_id.scheduled_date', 'report_id.cal_freq', 'report_id.interval')
    def _compute_next_calibration_date(self):
        for record in self:
            if not record.report_id:
                record.next_cal_date = False
                continue

            # Get the current scheduled date
            current_date = record.report_id.scheduled_date

            # Get the frequency and interval
            freq = record.report_id.cal_freq
            interval = record.report_id.interval or 1  # Default to 1 if not set

            # Calculate the next calibration date based on frequency
            if freq == 'daily':
                record.next_cal_date = current_date + timedelta(days=interval)
            elif freq == 'weekly':
                record.next_cal_date = current_date + timedelta(weeks=interval)
            elif freq == 'monthly':
                record.next_cal_date = current_date + relativedelta(months=interval)
            elif freq == 'yearly':
                record.next_cal_date = current_date + relativedelta(years=interval)
            else:
                record.next_cal_date = False






class CalibrationReportLine(models.Model):
    _name = 'calibration.report.line'
    _description = 'Calibration Report Line'
    line_id=fields.Many2one('calibration.report',string="lines")

    sn=fields.Char(string="S.No")
    standard_length=fields.Float(string="Standard Length (mm)", digits=(10, 2))
    upper_reading=fields.Float(string="Upper Reading (mm)", digits=(10, 2))
    center_reading=fields.Float(string="Center Reading (mm)", digits=(10, 2))
    lower_reading=fields.Float(string="Lower Reading (mm)", digits=(10, 2))
    avg_reading=fields.Float(string="Average Reading (mm)",compute='_compute_average', store=True, digits=(10, 2))
    min_lim=fields.Float(string="Min Limit (mm)", digits=(10, 2))
    max_lim=fields.Float(string="Max Limit (mm)", digits=(10, 2))

    @api.depends('upper_reading', 'center_reading', 'lower_reading')
    def _compute_average(self):
        for record in self:
            readings = [
                reading for reading in [
                    record.upper_reading, record.center_reading, record.lower_reading
                ] if reading
            ]
            record.avg_reading = sum(readings) / len(readings) if readings else 0.0



    @api.onchange('standard_length')
    def _onchange_standard_length(self):
        if self.standard_length:
            self.min_lim = self.standard_length - 0.05
            self.max_lim = self.standard_length + 0.05




class MaintenanceEquipment(models.Model):
    _inherit = 'maintenance.equipment'

    code = fields.Char(string="Code")
    range = fields.Char(string="Range")
    make = fields.Char(string="Make")
    lc = fields.Char(string="Least Count")
    location = fields.Char(string="Location")
    gauge_type = fields.Selection([('variable', 'Variable'), ('attribute', 'Attribute')], string='Gauge Type')











