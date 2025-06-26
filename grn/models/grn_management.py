from fnmatch import translate

from odoo import models, fields, api,_
import base64
import qrcode
from io import BytesIO
from openpyxl import Workbook
from PIL import Image as PILImage, ImageOps
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
import io
import logging

from odoo.exceptions import ValidationError

_logger = logging.getLogger(__name__)



class GRNManagement(models.Model):
    _name = 'grn.management'
    _description = 'GRN Management'
    _inherit = ['mail.thread', 'mail.activity.mixin']
    _rec_name= 'grn_no'
    grn_no = fields.Char(string="GRN No")
    grn_date = fields.Date(string="GRN Date")
    supplier_name = fields.Many2one('res.partner', string="Supplier")
    supplier_code = fields.Char("Supplier Code" , related='supplier_name.supplier_code')
    grn_product = fields.Char("Product", related="part_id.grn_product")
    part_id = fields.Many2one('product.template', string="Reference")
    part_number = fields.Char(string="RIQP Number", related='part_id.default_code', store=True)
    grn_part_number = fields.Char("Part Number",related="part_id.grn_part_number")

    part_name = fields.Char(string="Part Name", related='part_id.name', store=True)
    invoice_no = fields.Char(string="Invoice No")
    no_of_boxes = fields.Integer(string="No. of Boxes/Pallets Received")
    per_box_qty = fields.Integer(string="Per Box/Pallet Qty.")
    total_qty = fields.Integer(string="Total Receipt Qty.", compute="_compute_total_qty", store=True)
    store_location = fields.Char(string="Store Location")
    sample_qty_variable = fields.Integer(string="Sample Quantity / Lot Qty (Nos) For Variable", compute="_compute_sample_quantities", store=True)
    sample_qty_attribute = fields.Integer(string="Sample Quantity / Lot Qty (Nos) For Attribute", compute="_compute_sample_quantities", store=True)
    responsible_person = fields.Many2many('res.users', string="Responsible Persons")
    # Changed from Binary to Many2many fields
    supplier_inspection_report = fields.Many2many('ir.attachment', 'grn_sir_attachment_rel', 'grn_id', 'attachment_id', string="attachment")
    supplier_properties_report = fields.Many2many('ir.attachment', 'grn_spr_attachment_rel', 'grn_id', 'attachment_id', string="attachment")
    internal_verified_sir = fields.Many2many('ir.attachment', 'grn_verified_sir_rel', 'grn_id', 'attachment_id', string="attachment")
    internal_verified_spr = fields.Many2many('ir.attachment', 'grn_verified_spr_rel', 'grn_id', 'attachment_id', string="attachment")
    grn_report_attachment = fields.Many2many('ir.attachment', 'grn_report_attachment_rel', 'grn_id', 'attachment_id', string="Report Attchments")
    product_type = fields.Selection(
        [('regular', 'Regular Product'), ('dol', 'DOL Product')],
        string="Product Type",
        compute="_compute_product_type",
        store=True
    )
    follow_regular_process = fields.Boolean(string="Follow Regular Process", default=False,
                                            help="For DOL products that will follow the regular inspection process")
    @api.depends('part_id')
    def _compute_product_type(self):
        for record in self:
            control_plan = self.env['grn.control.plan'].search([('part_id', '=', record.part_id.id)], limit=1)
            record.product_type = control_plan.product_type if control_plan else False
            _logger.info(f"Computed product_type for GRN {record.grn_no}: {record.product_type}")

    # Add a specific field for camera captures
    camera_attachments = fields.Many2many(
        'ir.attachment',
        'grn_camera_attachment_rel',
        'grn_id',
        'attachment_id',
        string="Camera Captures"
    )
    status = fields.Selection([
        ('draft', 'Draft'),
        ('inspection', 'Under Inspection'),
        ('submit', 'Send Approval'),
        ('accepted', 'Accepted'),
        ('rejected', 'Rejected'),
        ('quarantine', 'Quarantine')
    ], string='Status', default='draft', required=True)

    qr_code = fields.Binary(string="QR Code", readonly=True)
    grn_control_ids = fields.One2many('grn.management.process', 'grn_process_id', string="GRN Processes")
    inspection_report_ids = fields.One2many('grn.inspection.report', 'grn_id', string="Inspection Reports")
    # Add the two new boolean fields
    accepted_before_recheck = fields.Boolean(string="Accepted Before Recheck", default=False,
                                             help="Indicates if the GRN was accepted before any recheck was required")
    accepted_after_recheck = fields.Boolean(string="Accepted After Recheck", default=False,
                                            help="Indicates if the GRN was accepted after recheck was performed")
    # Helper computed fields for graphing
    before_recheck_count = fields.Integer(compute='_compute_recheck_count', store=True)
    after_recheck_count = fields.Integer(compute='_compute_recheck_count', store=True)
    graph_name = fields.Char(compute='_compute_graph_name', store=True)

    # New fields to store rejection wizard data
    rejection_wizard_defect_source = fields.Char(string="Rejection Defect Source")
    rejection_wizard_defective_qty = fields.Integer(string="Rejection Defective Quantity")
    rejection_wizard_suspected_qty = fields.Integer(string="Rejection Suspected Quantity")
    rejection_wizard_defect_description = fields.Text(string="Rejection Defect Description")
    rejection_wizard_defect_snapshot = fields.Binary(string="Rejection Defect Snapshot")
    rejection_wizard_repeat_issue = fields.Char(string="Rejection Repeat Issue")  # Placeholder for Repeat Issue

    @api.depends('accepted_before_recheck', 'accepted_after_recheck')
    def _compute_recheck_count(self):
        for record in self:
            record.before_recheck_count = 1 if record.accepted_before_recheck else 0
            record.after_recheck_count = 1 if record.accepted_after_recheck else 0

    @api.depends('accepted_before_recheck', 'accepted_after_recheck')
    def _compute_graph_name(self):
        for record in self:
            if record.accepted_before_recheck:
                record.graph_name = 'Before Recheck'
            elif record.accepted_after_recheck:
                record.graph_name = 'After Recheck'
            else:
                record.graph_name = 'Unclassified'



    # Add fields needed for the extended approval process
    recheck_required = fields.Boolean(string="Recheck Required", default=False)
    rejection_option = fields.Selection([
        ('scrap', 'Scrap'),
        ('return_to_supplier', 'Return to Supplier')
    ], string="Rejection Option")
    rejection_notes = fields.Text(string="Rejection Notes")
    approval_history_ids = fields.One2many('grn.approval.history', 'grn_approval_id', string="Approval History")
    generate_xls_file = fields.Binary(string="Generated File")

    @api.onchange('part_id')
    def _onchange_part_id(self):
        if self.part_id:
            control_plan = self.env['grn.control.plan'].search([('part_id', '=', self.part_id.id)], limit=1)
            self.grn_control_ids = [(5, 0, 0)]

            if control_plan:
                # Set the product_type field based on the control plan
                self.product_type = control_plan.product_type

                # Create control lines from control plan
                control_lines = [
                    (0, 0, {
                        'process_step': line.process_step,
                        'process_name_id': line.process_name_id.id,
                        'char_class': line.char_class.id if line.char_class else False,
                        'product_spec_tol': line.product_spec_tol,
                        'measure_method_aid': line.measure_method_aid.id if line.measure_method_aid else False,
                        'type_of_inspection': line.type_of_inspection,
                        'control_method': line.control_method.id if line.control_method else False,
                        'reaction_plan': line.reaction_plan.id if line.reaction_plan else False,
                    }) for line in control_plan.grn_process_line_ids
                ]
                self.grn_control_ids = control_lines

                # Add GRN report attachments
                if control_plan.grn_report_attachment:
                    self.grn_report_attachment = [(6, 0, control_plan.grn_report_attachment.ids)]

                # Keep status as draft - will be updated by appropriate actions later
                self.status = 'draft'

                # Show appropriate warning message based on product type
                message = ""
                if control_plan.product_type == 'dol':
                    message = "This is a DOL product. Please save the record and use the 'DOL Product Approval' button to proceed."
                else:
                    message = "This is a Regular product. Please save the record and use the 'Start Inspection' button to proceed."

                return {
                    'warning': {
                        'title': 'Product Type: ' + ('DOL' if control_plan.product_type == 'dol' else 'Regular'),
                        'message': message
                    }
                }
        else:
            self.grn_control_ids = [(5, 0, 0)]
            self.inspection_report_ids = [(5, 0, 0)]
            self.grn_report_attachment = [(5, 0, 0)]
            self.product_type = False

    def action_open_dol_approval_wizard(self):
        """Open the DOL Approval Wizard for DOL products with field validations"""
        self.ensure_one()

        # Validation for required fields
        missing_fields = []

        if not self.supplier_name:
            missing_fields.append("Supplier")

        if not self.responsible_person or len(self.responsible_person) == 0:
            missing_fields.append("Responsible Person")

        if not self.invoice_no:
            missing_fields.append("Invoice No")

        if not self.no_of_boxes or self.no_of_boxes <= 0:
            missing_fields.append("No. of Boxes/Pallets Received")

        if not self.per_box_qty or self.per_box_qty <= 0:
            missing_fields.append("Per Box/Pallet Qty")

        if not self.sample_qty_variable or self.sample_qty_variable <= 0:
            missing_fields.append("Sample Quantity For Variable")

        if not self.sample_qty_attribute or self.sample_qty_attribute <= 0:
            missing_fields.append("Sample Quantity For Attribute")

        # If any required fields are missing, show error message
        if missing_fields:
            missing_fields_str = ", ".join(missing_fields)
            raise ValidationError(_(
                f"Please fill in the following required fields before proceeding: {missing_fields_str}"))

        # If product is DOL type and all validations pass, open the wizard
        if self.product_type == 'dol':
            return {
                'type': 'ir.actions.act_window',
                'res_model': 'grn.dol.approval.wizard',
                'view_mode': 'form',
                'target': 'new',
                'context': {'default_grn_id': self.id}
            }

        return True

    def generate_combined_xls_report(self):
        # Create workbook
        output = BytesIO()
        wb = Workbook()

        # Remove the default sheet created by Workbook()
        wb.remove(wb.active)

        # Define common styles
        border = Border(top=Side(style='thin'), left=Side(style='thin'),
                        right=Side(style='thin'), bottom=Side(style='thin'))
        font_header = Font(name='Times New Roman', bold=True)
        title_font = Font(size=20, bold=True)
        align_center = Alignment(vertical='center', horizontal='center', wrapText=True)
        fill = PatternFill(start_color="e7e7e7", end_color="e7e7e7", fill_type="solid")

        # Helper function to add logo
        def add_logo(ws):
            if self.env.user.company_id and self.env.user.company_id.logo:
                max_width, max_height = 100, 200
                image_data = base64.b64decode(self.env.user.company_id.logo)
                image = PILImage.open(io.BytesIO(image_data))
                width, height = image.size
                aspect_ratio = width / height

                if width > max_width:
                    width = max_width
                    height = int(width / aspect_ratio)
                if height > max_height:
                    height = max_height
                    width = int(height * aspect_ratio)

                resized_image = image.resize((width, height), PILImage.LANCZOS)
                padding_top, padding_left = 5, 5
                resized_image = ImageOps.expand(resized_image, border=(padding_left, padding_top, 0, 0), fill='white')
                img_bytes = io.BytesIO()
                resized_image.save(img_bytes, format='PNG')
                img_bytes.seek(0)
                logo_image = Image(img_bytes)
                ws.add_image(logo_image, 'A1')

        # Worksheet 1: Control Plan
        ws_control = wb.create_sheet(title="Control Plan")
        add_logo(ws_control)

        control_plan = self.env['grn.control.plan'].search([('part_id', '=', self.part_id.id)], limit=1)
        if control_plan:
            ws_control.row_dimensions[1].height = 50
            headers_control = {
                'C1': 'Receipt Inspection Quality Plan (RIQP)', 'C2': 'Product', 'C3': 'RIQP Number',
                'C4': 'Part Name', 'C5': 'Part Number', 'C6': 'Date (Orig.)', 'G2': 'Key Contact',
                'G3': 'Telephone', 'G4': 'CFT Team', 'G5': 'Rev No.', 'G6': 'Rev Date',
                'A7': 'Sr. No.', 'B7': 'Characteristics', 'C7': 'Class', 'D7': 'Product Specification Tolerance',
                'E7': 'Measurement Method / Measuring Aid', 'F7': 'Gauge Type', 'G7': 'Control Method (Records/GRN)',
                'H7': 'Sample Frequency', 'I7': 'Type of Inspection', 'J7': 'Reaction Plan',
            }
            for cell, value in headers_control.items():
                ws_control[cell] = value
                ws_control[cell].font = font_header
                ws_control[cell].alignment = align_center
                ws_control[cell].fill = fill
            ws_control['C1'].font = title_font

            merge_cells_control = [
                'A1:B6', 'C1:J1', 'C2:D2', 'C3:D3', 'C4:D4', 'C5:D5', 'C6:D6',
                'E2:F2', 'E3:F3', 'E4:F4', 'E5:F5', 'E6:F6', 'G2:H2', 'G3:H3',
                'G4:H4', 'G5:H5', 'G6:H6', 'I2:J2', 'I3:J3', 'I4:J4', 'I5:J5', 'I6:J6',
            ]
            for cell_range in merge_cells_control:
                ws_control.merge_cells(cell_range)

            col_widths_control = {'A': 8, 'B': 20, 'C': 18, 'D': 20, 'E': 20, 'F': 18, 'G': 18, 'H': 18, 'I': 18,
                                  'J': 18}
            for col, width in col_widths_control.items():
                ws_control.column_dimensions[col].width = width

            ws_control['E2'] = control_plan.grn_product or ''
            ws_control['E3'] = control_plan.part_number or ''
            ws_control['E4'] = control_plan.part_name or ''
            ws_control['E5'] = control_plan.grn_part_number or ''
            ws_control['E6'] = control_plan.date_origin or ''
            ws_control['I2'] = control_plan.key_contact.name or ''
            ws_control['I3'] = control_plan.telephone or ''
            ws_control['I4'] = ', '.join(control_plan.team_cft.mapped('name')) or ''
            ws_control['I5'] = control_plan.rev_no or ''
            ws_control['I6'] = control_plan.rev_date or ''

            row = 8
            for line in control_plan.grn_process_line_ids:
                ws_control[f'A{row}'] = line.process_step or ''
                ws_control[f'B{row}'] = line.process_name_id.name or ''
                ws_control[f'C{row}'] = line.char_class.name or ''
                ws_control[f'D{row}'] = line.product_spec_tol or ''
                ws_control[f'E{row}'] = line.measure_method_aid.name or ''
                ws_control[f'F{row}'] = line.gauge_type or ''
                ws_control[f'G{row}'] = line.control_method.name or ''
                ws_control[f'H{row}'] = line.sample_freq or ''
                ws_control[f'I{row}'] = line.type_of_inspection or ''
                ws_control[f'J{row}'] = line.reaction_plan.name or ''
                row += 1


            cur_row = 30
            if cur_row < row:
                cur_row = row

            for rows in ws_control.iter_rows(min_row=1, max_row=cur_row, min_col=1, max_col=10):
                for cell in rows:
                    cell.alignment = align_center
                    cell.border = border

        # Worksheet 2: Inspection Report
        ws_inspection = wb.create_sheet(title="Inspection Report")
        add_logo(ws_inspection)

        inspection_report = self.env['grn.inspection.report'].search([('grn_id', '=', self.id)], limit=1)
        if inspection_report:
            ws_inspection.row_dimensions[1].height = 50
            max_observations = inspection_report.max_observations
            headers_inspection = {
                'C1': 'RECEIVING AUDIT CHECK SHEET', 'C2': 'GRN No.', 'C3': 'GRN Date', 'C4': 'Product',
                'C5': 'Part Name', 'F2': 'RIQP Number', 'F3': 'Part Number', 'F4': 'Supplier Name',
                'F5': 'Supplier Code', 'I2': 'Sample Qty / Lot Qty (Nos) For Variable',
                'I3': 'Sample Qty / Lot Qty (Nos) For Attribute', 'I4': 'Date', 'A6': 'Sr. No.',
                'B6': 'Characteristics', 'C6': 'Product Specification Tolerance',
                'D6': 'For Variable characteristics, clearly mention Minimum and Maximum observations within the range of selected sample size, in specified units. For observations recorded from Met Lab or CMM room, ensure filing of back up report'
            }

            def get_column_letter(index):
                result = ""
                while index > 0:
                    index, remainder = divmod(index - 1, 26)
                    result = chr(65 + remainder) + result
                return result

            col_index = 4
            for i in range(1, max_observations + 1):
                obs_col = get_column_letter(col_index)
                date_col = get_column_letter(col_index + 1)
                headers_inspection[f'{obs_col}7'] = 'Part Sample No.'
                headers_inspection[f'{date_col}7'] = f'Observation {i}'
                ws_inspection.merge_cells(f'{obs_col}7:{obs_col}8')
                ws_inspection.merge_cells(f'{date_col}7:{date_col}8')
                ws_inspection.column_dimensions[obs_col].width = 15
                ws_inspection.column_dimensions[date_col].width = 18
                col_index += 2

            for cell, value in headers_inspection.items():
                ws_inspection[cell] = value
                ws_inspection[cell].font = font_header
                ws_inspection[cell].alignment = align_center
                ws_inspection[cell].fill = fill
            ws_inspection['C1'].font = title_font

            max_col_index = 3 + (max_observations * 2)  # Starting from column D (4) plus 2 columns per observation
            max_col_letter = get_column_letter(max_col_index)

            merge_cells_inspection = [
                'A1:B5', 'C1:K1', 'A6:A8', 'B6:B8', 'C6:C8',
                'C2:D2', 'C3:D3', 'C4:D4', 'C5:D5',
                'F2:G2', 'F3:G3', 'F4:G4', 'F5:G5',
                'I2:J2', 'I3:J3', 'I4:J4', 'I5:J5',
            ]
            # Only add L1 merge if we have enough observations
            if max_col_index > 11:  # If we have columns beyond K (10)
                max_col_letter = get_column_letter(max_col_index)
                merge_cells_inspection.append(f'L1:{max_col_letter}5')
                merge_cells_inspection.append(f'D6:{max_col_letter}6')
            else:
                merge_cells_inspection.append(f'D6:K6')
            for cell_range in merge_cells_inspection:
                ws_inspection.merge_cells(cell_range)

            col_widths_inspection = {'A': 8, 'B': 18, 'C': 20}
            for col, width in col_widths_inspection.items():
                ws_inspection.column_dimensions[col].width = width

            ws_inspection['E2'] = inspection_report.grn_no or ''
            ws_inspection['E3'] = inspection_report.grn_date or ''
            ws_inspection['E4'] = inspection_report.grn_product or ''
            ws_inspection['E5'] = inspection_report.part_name or ''
            ws_inspection['H2'] = inspection_report.part_number or ''
            ws_inspection['H3'] = inspection_report.grn_part_number or ''
            ws_inspection['H4'] = inspection_report.supplier_name.name or ''
            ws_inspection['H5'] = inspection_report.supplier_code or ''
            ws_inspection['K2'] = inspection_report.sample_qty_variable or ''
            ws_inspection['K3'] = inspection_report.sample_qty_attribute or ''
            ws_inspection['K4'] = inspection_report.date or ''

            row = 9
            for ele in inspection_report.line_ids:
                ws_inspection[f'A{row}'] = ele.sr_no or ''
                ws_inspection[f'B{row}'] = ele.process_name_id.name or ''
                ws_inspection[f'C{row}'] = ele.product_spec_tol or ''
                col_index = 4
                for obs in ele.observations:
                    ws_inspection.cell(row=row, column=col_index, value=obs.part_sample_no or '')
                    ws_inspection.cell(row=row, column=col_index + 1, value=obs.observation_value or '')
                    col_index += 2
                row += 1



            cur_row = 30
            if cur_row < row:
                cur_row = row



            max_freq = 2 * max_observations
            for rows in ws_inspection.iter_rows(min_row=1, max_row=cur_row, min_col=1, max_col=3 + max_freq):
                for cell in rows:
                    cell.alignment = align_center
                    cell.border = border

        # Worksheet 3: Inspection Monthly Sheet
        ws_monthly = wb.create_sheet(title="Inspection Monthly Sheet")
        add_logo(ws_monthly)

        if inspection_report:
            ws_monthly.row_dimensions[1].height = 50
            headers_monthly = {
                'C1': 'RECEIVING AUDIT CHECK SHEET', 'C2': 'GRN No.', 'C3': 'GRN Date', 'C4': 'Product',
                'C5': 'Part Name', 'G2': 'RIQP Number', 'G3': 'Part Number', 'G4': 'Supplier Name',
                'G5': 'Supplier Code', 'K2': 'Sample Qty / Lot Qty (Nos) For Variable',
                'K3': 'Sample Qty / Lot Qty (Nos) For Attribute', 'K4': 'Date', 'A6': 'Sr. No.',
                'B6': 'Characteristics', 'C6': 'Product Specification Tolerance', 'D6': 'Periodic - Year- ',
                'D7': 'Jan', 'E7': 'Feb', 'F7': 'Mar', 'G7': 'Apr', 'H7': 'May', 'I7': 'Jun',
                'J7': 'Jul', 'K7': 'Aug', 'L7': 'Sep', 'M7': 'Oct', 'N7': 'Nov', 'O7': 'Dec'
            }

            for cell, value in headers_monthly.items():
                ws_monthly[cell] = value
                ws_monthly[cell].font = font_header
                ws_monthly[cell].alignment = align_center
                ws_monthly[cell].fill = fill
            ws_monthly['C1'].font = title_font

            merge_cells_monthly = [
                'A1:B5', 'C1:O1', 'D6:O6', 'A6:A8', 'B6:B8', 'C6:C8', 'D7:D8', 'E7:E8', 'F7:F8', 'G7:G8',
                'H7:H8', 'I7:I8', 'J7:J8', 'K7:K8', 'L7:L8', 'M7:M8', 'N7:N8', 'O7:O8', 'C2:D2', 'C3:D3',
                'C4:D4', 'C5:D5', 'E2:F2', 'E3:F3', 'E4:F4', 'E5:F5', 'G2:H2', 'G3:H3', 'G4:H4', 'G5:H5',
                'I2:J2', 'I3:J3', 'I4:J4', 'I5:J5', 'K2:M2', 'K3:M3', 'K4:M4', 'K5:M5', 'N2:O2', 'N3:O3',
                'N4:O4', 'N5:O5'
            ]
            for cell_range in merge_cells_monthly:
                ws_monthly.merge_cells(cell_range)

            col_widths_monthly = {
                'A': 8, 'B': 18, 'C': 20, 'D': 15, 'E': 15, 'F': 15, 'G': 15, 'H': 15, 'I': 15,
                'J': 15, 'K': 15, 'L': 15, 'M': 15, 'N': 15, 'O': 15
            }
            for col, width in col_widths_monthly.items():
                ws_monthly.column_dimensions[col].width = width

            ws_monthly['E2'] = inspection_report.grn_no or ''
            ws_monthly['E3'] = inspection_report.grn_date or ''
            ws_monthly['E4'] = inspection_report.grn_product or ''
            ws_monthly['E5'] = inspection_report.part_name or ''
            ws_monthly['I2'] = inspection_report.part_number or ''
            ws_monthly['I3'] = inspection_report.grn_part_number or ''
            ws_monthly['I4'] = inspection_report.supplier_name.name or ''
            ws_monthly['I5'] = inspection_report.supplier_code or ''
            ws_monthly['N2'] = inspection_report.sample_qty_variable or ''
            ws_monthly['N3'] = inspection_report.sample_qty_attribute or ''
            ws_monthly['N4'] = inspection_report.date or ''

            row = 9
            month_columns = {
                'January': 'D', 'February': 'E', 'March': 'F', 'April': 'G', 'May': 'H', 'June': 'I',
                'July': 'J', 'August': 'K', 'September': 'L', 'October': 'M', 'November': 'N', 'December': 'O'
            }
            for ele in inspection_report.line_ids:
                ws_monthly[f'A{row}'] = ele.sr_no or ''
                ws_monthly[f'B{row}'] = ele.process_name_id.name or ''
                ws_monthly[f'C{row}'] = ele.product_spec_tol or ''
                for attachment in ele.attchment:
                    if attachment.report_month in month_columns:
                        ws_monthly[f'{month_columns[attachment.report_month]}{row}'] = 'X'
                row += 1

            cur_row = 30
            if cur_row < row:
                cur_row = row

            for rows in ws_monthly.iter_rows(min_row=1, max_row=cur_row, min_col=1, max_col=15):
                for cell in rows:
                    cell.alignment = align_center
                    cell.border = border

        # Save workbook
        wb.save(output)
        output.seek(0)

        # Create attachment
        attachment = self.env['ir.attachment'].create({
            'name': 'Combined_GRN_Report.xlsx',
            'type': 'binary',
            'datas': base64.b64encode(output.getvalue()),
            'res_model': 'grn.management',
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        })

        # Return download link
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }

    # line_ids = fields.One2many('grn.inspection.report.line', compute='_compute_inspection_lines',
    #                            string="Inspection Lines")
    #
    # @api.depends('inspection_report_ids.line_ids')
    # def _compute_inspection_lines(self):
    #     for record in self:
    #         record.line_ids = [(6, 0, record.inspection_report_ids.mapped('line_ids').ids)]

    # ✅ Directly link inspection report lines
    line_ids = fields.One2many('grn.inspection.report.line', 'grn_management_id', string="Inspection Lines")



    @api.depends('no_of_boxes', 'per_box_qty')
    def _compute_total_qty(self):
        for rec in self:
            rec.total_qty = rec.no_of_boxes * rec.per_box_qty if rec.no_of_boxes and rec.per_box_qty else 0

    @api.depends('total_qty')
    def _compute_sample_quantities(self):
        for rec in self:
            total_qty = rec.total_qty or 0
            # Sampling plan based on the provided table
            if 2 <= total_qty <= 150:
                sample_size = 5
            elif 151 <= total_qty <= 500:
                sample_size = 13
            elif 501 <= total_qty <= 35000:
                sample_size = 20
            elif 35001 <= total_qty <= 500000:
                sample_size = 32
            elif total_qty > 500000:
                sample_size = 50
            else:
                sample_size = 0  # For quantities less than 2, or invalid cases

            # Assign the same sample size to both variable and attribute for now
            rec.sample_qty_variable = sample_size
            rec.sample_qty_attribute = sample_size

    @api.model
    def create(self, vals):
        record = super(GRNManagement, self).create(vals)
        record._generate_qr_code()
        return record

    def write(self, vals):
        res = super(GRNManagement, self).write(vals)
        if any(field in vals for field in ['grn_no', 'grn_date', 'supplier_name', 'part_id',
                                           'invoice_no', 'no_of_boxes', 'per_box_qty',
                                           'total_qty', 'store_location']):
            self._generate_qr_code()
        return res

    def _generate_qr_code(self):
        for rec in self:
            qr_data = f"GRN No: {rec.grn_no}\nDate: {rec.grn_date}\nSupplier: {rec.supplier_name.name}\n" \
                      f"Part ID: {rec.part_id.name}\nInvoice No: {rec.invoice_no}\n" \
                      f"Total Quantity: {rec.total_qty}\nStatus: {rec.status}"

            qr = qrcode.make(qr_data)
            temp = BytesIO()
            qr.save(temp, format="PNG")
            qr_code_binary = base64.b64encode(temp.getvalue())

            rec.qr_code = qr_code_binary



    def action_start_inspection(self):
        self.ensure_one()

        # Validation for required fields
        missing_fields = []

        if not self.grn_no:
            missing_fields.append("GRN Number")
        if not self.part_id:
            missing_fields.append("Reference")
        if not self.supplier_name:
            missing_fields.append("Supplier")

        if not self.responsible_person or len(self.responsible_person) == 0:
            missing_fields.append("Responsible Person")

        if not self.invoice_no:
            missing_fields.append("Invoice No")

        if not self.no_of_boxes or self.no_of_boxes <= 0:
            missing_fields.append("No. of Boxes/Pallets Received")

        if not self.per_box_qty or self.per_box_qty <= 0:
            missing_fields.append("Per Box/Pallet Qty")

        if not self.sample_qty_variable or self.sample_qty_variable <= 0:
            missing_fields.append("Sample Quantity For Variable")

        if not self.sample_qty_attribute or self.sample_qty_attribute <= 0:
            missing_fields.append("Sample Quantity For Attribute")

        # If any required fields are missing, show error message
        if missing_fields:
            missing_fields_str = ", ".join(missing_fields)
            raise ValidationError(_(
                f"Please fill in the following required fields before proceeding: {missing_fields_str}"))
        self._create_history_record('start_inspection', 'Started inspection process')
        self.status = 'inspection'

        control_plan = self.env['grn.control.plan'].search([('part_id', '=', self.part_id.id)], limit=1)

        inspection_report = self.env['grn.inspection.report'].search([('grn_no', '=', self.grn_no)], limit=1)

        if not inspection_report:
            inspection_report = self.env['grn.inspection.report'].create({
                'grn_id': self.id,
                'grn_no': self.grn_no,
                'grn_date': self.grn_date,
                'part_id': self.part_id.id,
                'supplier_name': self.supplier_name.id,
                'sample_qty_variable': self.sample_qty_variable,
                'sample_qty_attribute': self.sample_qty_attribute,
                'supplier_inspection_report': [(6, 0, self.supplier_inspection_report.ids)],
                'supplier_properties_report': [(6, 0, self.supplier_properties_report.ids)],
                'internal_verified_sir': [(6, 0, self.internal_verified_sir.ids)],
                'internal_verified_spr': [(6, 0, self.internal_verified_spr.ids)],

                # 'batch_quantity': self.total_qty or 0,

                'line_ids': [
                    (0, 0, {
                        'grn_management_id': self.id,
                        'process_name_id': line.process_name_id.id,
                        'product_spec_tol': line.product_spec_tol,
                        # 'batch_quantity': self.total_qty or 0,
                        # Pass sample frequency from control plan
                        'sample_freq': line.sample_freq or
                                       (self.sample_qty_variable if line.gauge_type == 'variable' else
                                        self.sample_qty_attribute if line.gauge_type == 'attribute' else 0),
                        'measure_method_aid': line.measure_method_aid.id,
                        'observations': [(0, 0, {
                            'observation_number': i + 1,
                            'part_sample_no': f"PS{i + 1}",
                            'observation_value': ''
                        }) for i in range(line.sample_freq)],
                        # ⬇️ Add dummy attachment
                        'attchment': [(0, 0, {
                            'report_filename': 'dummy.pdf',
                            'report_attachment': base64.b64encode(b'Dummy content')
                            # No need to set report_month — it is auto-handled
                        })]


                    }) for line in control_plan.grn_process_line_ids
                ] if control_plan else []
            })

        self.inspection_report_ids = [(4, inspection_report.id)]

        return {
            'type': 'ir.actions.act_window',
            'res_model': 'grn.management',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'current',
            'flags': {'mode': 'readonly'},
            'tag': 'display_notification',
            'params': {
                'title': 'Inspection Started',
                'message': f'GRN {self.grn_no} has been moved to inspection status',
                'type': 'info',
                'sticky': False,
            }
        }

    def _send_notification_email(self, template_xml_id, email_to=None):
        """
        Send notification emails to the appropriate recipients.

        Args:
            template_xml_id: XML ID of the email template to use
            email_to: Optional list of email addresses to send to

        Returns:
            bool: True if email was sent successfully, False otherwise
        """
        self.ensure_one()
        try:
            # Fetch the email template
            template = self.env.ref(template_xml_id, raise_if_not_found=False)
            if not template:
                _logger.error(f"Email template {template_xml_id} not found for GRN {self.grn_no}")
                return False

            # Prepare email values
            email_values = {
                'subject': template.subject,
                'body_html': template.body_html,
                'email_from': template.email_from,
                'model': self._name,
                'res_id': self.id,
            }

            if email_to:
                # Send to specific email addresses
                email_values['email_to'] = ','.join(email_to)
                mail = self.env['mail.mail'].create(email_values)
                mail.send()
            else:
                # Use template's default recipients
                template.with_context(force_send=True).send_mail(
                    self.id,
                    force_send=True,
                    raise_exception=False
                )

            _logger.info(f"Email sent for GRN {self.grn_no} using template {template_xml_id}")
            return True

        except Exception as e:
            _logger.error(f"Failed to send email for GRN {self.grn_no}: {str(e)}")
            return False


    # Submit to Inspection button action
    def action_submit_to_inspection(self):
        self.status = 'submit'
        self._create_history_record('submit_inspection', 'Submitted for inspection approval')
        self._send_notification_email('grn.email_template_grn_submitted')

        return {
            'name': 'Submit GRN',
            'type': 'ir.actions.act_window',
            'res_model': 'grn.submit.wizard',
            'view_mode': 'form',
            'target': 'new',
            'context': {'default_grn_id': self.id}
        }

    # Accept button action
    def action_accept_grn(self):
        self.status = 'accepted'
        self.store_location = 'Moved to Store Location'
        self.accepted_before_recheck = True

        self._create_history_record('accepted', 'GRN accepted - product meets specifications')
        self._send_notification_email('grn.email_template_grn_accepted')

        # Show popup notification
        return {
             'type': 'ir.actions.act_window',
            'res_model': 'grn.management',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'current',
            'flags': {'mode': 'readonly'},
            'tag': 'display_notification',
            'params': {
                'title': 'GRN Accepted',
                'message': f'GRN {self.grn_no} has been moved to stored location and ready to supply',
                'type': 'success',

                'sticky': True,
            }
        }



    # Reject button action
    def action_reject_grn(self):
        self.status = 'rejected'
        self._create_history_record('rejected', 'GRN rejected - product does not meet specifications')
        self._send_notification_email('grn.email_template_grn_rejected')

        return {
            'name': 'Reject GRN',
            'type': 'ir.actions.act_window',
            'res_model': 'grn.rejection.wizard',
            'view_mode': 'form',
            'target': 'new',
            'context': {'default_grn_id': self.id}
        }

    # Move to Quarantine button action
    def action_move_to_quarantine(self):
        self.status = 'quarantine'
        self.store_location = ' Moved to Quarantine'
        self._create_history_record('quarantine', 'GRN moved to quarantine')
        self._send_notification_email('grn.email_template_grn_quarantine')

        return True

    # Send for Recheck button action
    def action_send_for_recheck(self):
        self.recheck_required = True
        self._create_history_record('recheck', 'GRN sent for recheck')
        self._send_notification_email('grn.email_template_grn_recheck')

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': 'Sent for Recheck',
                'message': f'GRN {self.grn_no} has been sent for rechecking',
                'type': 'info',
                'sticky': False,
            }
        }

    # Process after recheck - Accept
    def action_accept_after_recheck(self):
        self.status = 'accepted'
        self.store_location = 'Moved to Store Location'
        self.recheck_required = False
        self.accepted_after_recheck = True
        self._send_notification_email('grn.email_template_grn_accepted_recheck')

        self._create_history_record('accepted_recheck', 'GRN accepted after recheck')
        return {
             'type': 'ir.actions.act_window',
            'tag': 'display_notification',
            'res_model': 'grn.management',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'current',
            'flags': {'mode': 'readonly'},
            'params': {
                'title': 'GRN Accepted After Recheck',
                'message': f'Store location updated for GRN {self.grn_no}',
                'type': 'success',
                'sticky': False,

            }



        }

    # Process after recheck - Reject with options
    def action_reject_after_recheck(self):
        return {
            'name': 'Reject GRN After Recheck',
            'type': 'ir.actions.act_window',
            'res_model': 'grn.final.rejection.wizard',
            'view_mode': 'form',
            'target': 'new',
            'context': {'default_grn_id': self.id}
        }

    # Process final rejection - Scrap
    def action_scrap_rejected_material(self):
        self.status = 'rejected'
        self.store_location = 'Moved to Scrap'
        self.rejection_option = 'scrap'
        self._send_notification_email('grn.email_template_grn_scrap')

        self._create_history_record('scrap', 'Rejected material marked for scrap')
        return {

            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': 'Material Scrapped',
                'message': f'GRN {self.grn_no} material has been marked for scrap',
                'type': 'warning',
                'sticky': False,
            }
        }

    # Process final rejection - Return to Supplier
    def action_return_to_supplier(self):
        self.status = 'quarantine'
        self.store_location = ' Moved to Quarantine'
        self.rejection_option = 'return_to_supplier'
        self._send_notification_email('grn.email_template_grn_return_supplier')

        self._create_history_record('return_supplier', 'Rejected material returned to supplier')
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': 'Returned to Supplier',
                'message': f'GRN {self.grn_no} material has been marked for return to supplier',
                'type': 'warning',
                'sticky': False,
            }
        }

    # Helper method to create history records
    def _create_history_record(self, action_type, description):
        self.env['grn.approval.history'].create({
            'grn_approval_id': self.id,
            'action_type': action_type,
            'description': description,
            'user_id': self.env.user.id,
            'date': fields.Datetime.now(),
        })


    # def action_submit_grn(self):
    #     self.status = 'submit'

    # def action_accept_grn(self):
    #     self.status = 'accepted'
    #
    # def action_reject_grn(self):
    #     self.status = 'rejected'
    #
    # def action_move_to_quarantine(self):
    #     self.status = 'quarantine'



class ControlProcess(models.Model):
    _name = "grn.management.process"
    _description = "Control Process"
    _rec_name = 'process_step'
    _inherit = "translation.mixin"

    grn_process_id = fields.Many2one('grn.management', 'GRN Management Process', ondelete='cascade')

    # process_step = fields.Char("Process Step")
    sequence = fields.Integer(string="Sequence", default=10)
    process_step = fields.Integer("S.No", compute="_compute_sequence_number")


    process_name_id = fields.Many2one('grn.characteristics',string="Characteristics")
    char_class = fields.Many2one('grn.process.class', "Class")
    product_spec_tol = fields.Char("Product Specification / Tolerance",translate=True)
    measure_method_aid  = fields.Many2one("maintenance.equipment", "Measurement Method / Measuring Aid")
    measure_method_aid_details = fields.Char(
        string="Measurement Method Details",
        compute="_compute_measure_method_details",
        store=True,
        help="Combined details of measurement method/aid"
    )
    control_method = fields.Many2one('grn.control.method', "Control Method")
    reaction_plan = fields.Many2one('grn.reaction.plan', "Reaction Plan")
    type_of_inspection= fields.Selection([('visual','Visual'), ('dimensional','Dimensional'),('properties','Properties')], string="Type of Inspection",related='measure_method_aid.type_of_inspection')
    gauge_type = fields.Selection([('variable', 'Variable'), ('attribute', 'Attribute')], string='Gauge Type', related='measure_method_aid.gauge_type')
    sample_freq =fields.Char("Sample Frequency",compute="_compute_sample_frequency",store=True)

    @api.depends('gauge_type', 'grn_process_id.sample_qty_variable', 'grn_process_id.sample_qty_attribute')
    def _compute_sample_frequency(self):
        """
        Dynamically compute sample frequency based on gauge type and GRN management quantities.
        Prioritizes control plan's sample frequency if available.
        """
        for record in self:
            # Check if there's a sample frequency in the control plan line
            if hasattr(record, 'sample_freq_in_control_plan'):
                record.sample_freq = record.sample_freq_in_control_plan
            else:
                # Fallback to GRN management quantities
                if record.gauge_type == 'variable':
                    record.sample_freq = str(record.grn_process_id.sample_qty_variable or 0)
                elif record.gauge_type == 'attribute':
                    record.sample_freq = str(record.grn_process_id.sample_qty_attribute or 0)
                else:
                    record.sample_freq = False
    @api.depends('measure_method_aid')
    def _compute_measure_method_details(self):
        for record in self:
            if record.measure_method_aid:
                # Combine name, code, and least count into a single string
                record.measure_method_aid_details = " | ".join(filter(bool, [
                    record.measure_method_aid.name or '',
                    record.measure_method_aid.serial_no or '',
                    record.measure_method_aid.code or '',
                    record.measure_method_aid.lc or ''
                ]))
            else:
                record.measure_method_aid_details = False



    @api.depends('sequence', 'grn_process_id')
    def _compute_sequence_number(self):
        for order in self.mapped('grn_process_id'):
            process_step = 1
            for lines in order.grn_control_ids:
                lines.process_step = process_step
                process_step += 1

class GRNDOLApprovalWizard(models.TransientModel):
    _name = 'grn.dol.approval.wizard'
    _description = 'GRN DOL Approval Wizard'

    grn_id = fields.Many2one('grn.management', string="GRN", required=True, ondelete='cascade')

    def action_approve_directly(self):
        grn = self.grn_id
        grn.status = 'accepted'
        grn.store_location = 'Moved to Store Location'
        grn.accepted_before_recheck = True
        grn._create_history_record('accepted', 'GRN accepted - product meets specifications')

        return {
            'type': 'ir.actions.act_window',
            'res_model': 'grn.management',
            'res_id': grn.id,
            'view_mode': 'form',
            'target': 'current',
            'flags': {'mode': 'readonly'},
            'tag': 'display_notification',
            'params': {
                'title': 'GRN Accepted',
                'message': f'GRN {grn.grn_no} has been accepted directly.',
                'type': 'success',
                'sticky': False,
            }
        }
    def action_continue_inspection(self):
        """Continue with regular inspection process but keep in draft state"""
        grn = self.grn_id
        # Keep in draft state
        grn.status = 'draft'
        # Set flag to follow regular process
        grn.follow_regular_process = True

        return {
            'type': 'ir.actions.act_window',
            'res_model': 'grn.management',
            'res_id': grn.id,
            'view_mode': 'form',
            'target': 'current',
            'flags': {'mode': 'readonly'},
            'tag': 'display_notification',
            'params': {
                'title': 'Regular Inspection',
                'message': f'GRN {grn.grn_no} will proceed with regular inspection process.',
                'type': 'info',
                'sticky': False,
            }
        }



    def action_cancel(self):
        return {'type': 'ir.actions.act_window_close'}

class SpecialProductCharacteristic(models.Model):
    _name = 'grn.process.class'
    _description = "GRN Process Class"

    name = fields.Char(string='Name', required=True)
    symbol = fields.Char(string='Symbol', required=True)



class GrnControlMethod(models.Model):
    _name = 'grn.control.method'
    _description = "GRN Process Class"

    name = fields.Char(string='Control Method Name')
class Characteristics(models.Model):
    _name = 'grn.characteristics'
    _description = "GRN Characteristics"

    name = fields.Char(string='Characteristics Name')

class SampleFrequency(models.Model):
    _name = 'grn.sample.frequency'
    _description = "GRN Sample Frequency"


    name = fields.Char(string='Sample Frequency Name')

class ReactionPlan(models.Model):
    _name = 'grn.reaction.plan'
    _description = "GRN Reaction Plan"

    name = fields.Char(string='Reaction Plan Name')


class ProductTemplate(models.Model):
    _inherit = 'product.template'
    gauge_type = fields.Selection([('variable', 'Variable'), ('attribute', 'Attribute')], string='Gauge Type')

    grn_product = fields.Char("Product")
    grn_part_number = fields.Char("Part Number")


class HrEmployee(models.Model):
    _inherit = 'hr.employee'

    supplier_code = fields.Char(string='Supplier Code')
class ResPartner(models.Model):
    _inherit = 'res.partner'

    supplier_code = fields.Char(string="Supplier Code")


class MaintenanceEquipment(models.Model):
    _inherit = 'maintenance.equipment'

    code = fields.Char(string="Code")
    range = fields.Char(string="Range")
    make = fields.Char(string="Make")
    lc = fields.Char(string="Least Count")
    location = fields.Char(string="Location")
    gauge_type = fields.Selection([('variable', 'Variable'), ('attribute', 'Attribute')], string='Gauge Type')
    type_of_inspection = fields.Selection(
        [('visual', 'Visual'), ('dimensional', 'Dimensional'), ('properties', 'Properties')],
        string="Type of Inspection")




# History model to track all approval actions
class GRNApprovalHistory(models.Model):
    _name = 'grn.approval.history'
    _description = 'GRN Approval History'
    _order = 'date desc'

    grn_approval_id = fields.Many2one('grn.management', string="GRN", required=True, ondelete='cascade')
    action_type = fields.Selection([
        ('start_inspection', 'Started Inspection'),
        ('submit_inspection', 'Submitted for Inspection'),
        ('accepted', 'Accepted'),
        ('rejected', 'Rejected'),
        ('quarantine', 'Moved to Quarantine'),
        ('recheck', 'Sent for Recheck'),
        ('accepted_recheck', 'Accepted After Recheck'),
        ('scrap', 'Material Scrapped'),
        ('return_supplier', 'Returned to Supplier')
    ], string="Action Type", required=True)
    description = fields.Text(string="Description")
    user_id = fields.Many2one('res.users', string="User", required=True)
    date = fields.Datetime(string="Date", required=True)


# Wizard for capturing rejection details
class GRNRejectionWizard(models.TransientModel):
    _name = 'grn.rejection.wizard'
    _description = 'GRN Rejection Wizard'

    grn_id = fields.Many2one('grn.management', string="GRN", required=True, ondelete='cascade')
    part_number = fields.Char(string="Part Number", related='grn_id.part_number')
    part_description = fields.Char(string="Part Description", related='grn_id.part_name')
    defect_source = fields.Char(string="Defect Source")
    defective_qty = fields.Integer(string="Defective Quantity")
    suspected_qty = fields.Integer(string="Suspected Quantity")
    defect_description = fields.Text(string="Defect Description")
    defect_snapshot = fields.Binary(string="Defect Snapshot")
    rejection_notes = fields.Text(string="Rejection Notes", required=True)
    move_to_quarantine = fields.Boolean(string="Move to Quarantine", default=True)
    send_for_recheck = fields.Boolean(string="Send for Recheck", default=False)

    def action_confirm(self):
        grn = self.grn_id
        grn.write({
            'rejection_notes': self.rejection_notes,
            'rejection_wizard_defect_source': self.defect_source,
            'rejection_wizard_defective_qty': self.defective_qty,
            'rejection_wizard_suspected_qty': self.suspected_qty,
            'rejection_wizard_defect_description': self.defect_description,
            'rejection_wizard_defect_snapshot': self.defect_snapshot,
            'rejection_wizard_repeat_issue': 'N/A',  # Placeholder as per the image; adjust if you have actual data
        })

        if self.move_to_quarantine:
            grn.action_move_to_quarantine()

        if self.send_for_recheck:
            grn.action_send_for_recheck()

        return {'type': 'ir.actions.act_window_close'}


# Wizard for final rejection options after recheck
class GRNFinalRejectionWizard(models.TransientModel):
    _name = 'grn.final.rejection.wizard'
    _description = 'GRN Final Rejection Options'

    grn_id = fields.Many2one('grn.management', string="GRN", required=True, ondelete='cascade')
    rejection_option = fields.Selection([
        ('scrap', 'Scrap Material'),
        ('return_to_supplier', 'Return to Supplier')
    ], string="Rejection Option", required=True)
    part_number = fields.Char(string="Part Number",related='grn_id.part_number')
    part_description = fields.Char(string="Part Description",related='grn_id.part_name')
    defect_source= fields.Char(string="Defect Source")
    defective_qty = fields.Integer(string="Defective Quantity")
    suspected_qty = fields.Integer(string="Suspected Quantity")
    defect_description = fields.Text(string="Defect Description")
    defect_snapshot = fields.Binary(string="Defect Snapshot")

    notes = fields.Text(string="Additional Notes")


    def action_confirm(self):
        grn = self.grn_id
        grn.write({
            'rejection_notes': (grn.rejection_notes or '') + '\n' + (self.notes or ''),
            'rejection_option': self.rejection_option,
            'rejection_wizard_defect_source': self.defect_source,
            'rejection_wizard_defective_qty': self.defective_qty,
            'rejection_wizard_suspected_qty': self.suspected_qty,
            'rejection_wizard_defect_description': self.defect_description,
            'rejection_wizard_defect_snapshot': self.defect_snapshot,
            'rejection_wizard_repeat_issue': 'N/A',  # Placeholder as per the image; adjust if you have actual data
        })

        if self.rejection_option == 'scrap':
            grn.status = 'rejected'
            grn.store_location = 'Scrap'
            grn._create_history_record('scrap', 'Rejected material marked for scrap')
        else:
            grn.status = 'quarantine'
            grn.status = 'quarantine'
            grn._create_history_record('return_supplier',
                                       'Rejected material moved to quarantine for return to supplier')

        # First close the wizard
        result = {'type': 'ir.actions.act_window_close'}

        # Schedule refresh of the parent view
        self.env.context = dict(self.env.context)
        self.env.context.update({'reload_parent': True})

        return result


class GRNSubmitWizard(models.TransientModel):
    _name = 'grn.submit.wizard'
    _description = 'GRN Submit Wizard'

    grn_id = fields.Many2one('grn.management', string="GRN", required=True, ondelete='cascade')
    notes = fields.Text(string="Submit Notes")

    def action_confirm(self):
        grn = self.grn_id

        description = 'Submitted for inspection approval'
        if self.notes:
            description += f": {self.notes}"

        grn._create_history_record('submit_inspection', description)
        grn.write({'status': 'submit'})

        # First close the wizard
        result = {'type': 'ir.actions.act_window_close'}

        # Schedule refresh of the parent view
        self.env.context = dict(self.env.context)
        self.env.context.update({'reload_parent': True})

        return result
