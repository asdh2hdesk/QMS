
from datetime import datetime

from odoo import models, fields, api
from odoo.exceptions import ValidationError
import base64
from io import BytesIO
from openpyxl import Workbook
from PIL import Image as PILImage, ImageOps
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
import io


class IncomingIR(models.Model):
    _name = 'grn.inspection.report'
    _description = 'Report of Incoming Inspection'
    _rec_name = 'grn_no'
    grn_id = fields.Many2one('grn.management', string="GRN Reference")
    grn_no = fields.Char(string="GRN No")
    grn_date = fields.Date(string="GRN Date")
    grn_product = fields.Char("Product", related="part_id.grn_product")

    part_id = fields.Many2one('product.template', string="Reference", required=True)
    part_name = fields.Char("Part Name", related="part_id.name", store=True)
    part_number = fields.Char("RIQP Number", related="part_id.default_code", store=True)
    supplier_name = fields.Many2one('res.partner', string="Supplier")
    supplier_code = fields.Char("Supplier Code", related='supplier_name.supplier_code')

    # batch_quantity = fields.Integer("Batch Quantity", required=True)
    grn_part_number = fields.Char("Part Number", related="part_id.grn_part_number")
    # gauge_type = fields.Selection([('variable', 'Variable'), ('attribute', 'Attribute')], string='Gauge Type')
    sample_qty_variable = fields.Integer(string="Sample Quantity / Lot Qty (Nos) For Variable")
    sample_qty_attribute = fields.Integer(string="Sample Quantity / Lot Qty (Nos) For Attribute")

    # freq = fields.Char("Sample Frequency")
    date = fields.Date("Date", default=fields.Date.context_today)
    # Changed from Binary to Many2many fields
    supplier_inspection_report = fields.Many2many('ir.attachment', 'ins_sir_attachment_rel', 'inspection_id', 'attachment_id', string="attachment")
    supplier_properties_report = fields.Many2many('ir.attachment', 'ins_spr_attachment_rel', 'inspection_id', 'attachment_id', string="attachment")
    internal_verified_sir = fields.Many2many('ir.attachment', 'ins_verified_sir_rel', 'inspection_id', 'attachment_id', string="attachment")
    internal_verified_spr = fields.Many2many('ir.attachment', 'ins_verified_spr_rel', 'inspection_id', 'attachment_id',string="attachment")


    line_ids = fields.One2many('grn.inspection.report.line', 'inspection_id', string="Inspection Lines")
    generate_xls_file = fields.Binary(string="Generated File")

    @api.depends('sample_qty_variable', 'sample_qty_attribute')
    def _compute_max_observations(self):
        for record in self:
            max_val = max(record.sample_qty_variable or 0, record.sample_qty_attribute or 0)
            record.max_observations = max_val if max_val >= 4 else 4

    max_observations = fields.Integer(
        string="Maximum Observations",
        compute="_compute_max_observations",
        store=True
    )

    def generate_xls_report(self):
        # Create workbook and worksheet
        output = BytesIO()
        wb = Workbook()
        ws = wb.active

        # Define styles
        border = Border(top=Side(style='thin'), left=Side(style='thin'),
                        right=Side(style='thin'), bottom=Side(style='thin'))
        font_header = Font(name='Times New Roman', bold=True)
        title_font = Font(size=20, bold=True)
        align_center = Alignment(vertical='center', horizontal='center', wrapText=True)

        # Define fill colors
        fill = PatternFill(start_color="e7e7e7", end_color="e7e7e7", fill_type="solid")  # Grey
        orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

        if self.env.user.company_id and self.env.user.company_id.logo:

            max_width, max_height = 100, 200  # Set max dimensions
            image_data = base64.b64decode(self.env.user.company_id.logo)

            # Open image with PIL
            image = PILImage.open(io.BytesIO(image_data))
            width, height = image.size
            aspect_ratio = width / height

            # Resize with aspect ratio
            if width > max_width:
                width = max_width
                height = int(width / aspect_ratio)
            if height > max_height:
                height = max_height
                width = int(height * aspect_ratio)

            # Resize and add padding
            resized_image = image.resize((width, height), PILImage.LANCZOS)
            padding_top, padding_left = 5, 5
            resized_image = ImageOps.expand(resized_image, border=(padding_left, padding_top, 0, 0), fill='white')

            img_bytes = io.BytesIO()
            resized_image.save(img_bytes, format='PNG')
            img_bytes.seek(0)
            logo_image = Image(img_bytes)

            ws.add_image(logo_image, 'A1')
        # Determine maximum number of observations dynamically
        max_observations = self.max_observations

        # Headers dictionary
        headers = {
            'C1': 'RECEIVING  AUDIT  CHECK SHEET',
            'C2': 'GRN No.',
            'C3': 'GRN Date',
            'C4': 'Product',
            'C5': 'Part Name',
            'F2': 'RIQP Number',
            'F3': 'Part Number',
            'F4': 'Supplier Name',
            'F5': 'Supplier Code',
            # 'I2': 'Total Qty.',
            # 'I3': 'Store Location',
            'I2': 'Sample Qty / Lot Qty (Nos) For Variable',
            'I3': 'Sample Qty / Lot Qty (Nos) For Attribute',
            'I4': 'Date',
            # 'I2': 'CFT Team',
            # 'I3': 'Rev No.',
            # 'I4': 'Rev Date',
            # 'C6': 'Class',
            'A6': 'Sr. No.',
            'B6': 'Characteristics',
            'C6': 'Product Specification Tolerance',
            'D6': 'For Variable characteristics, clearly mention Minimum and Maximum observations within the range of selected sample size, in specified units.  For osbervations recorded from Met Lab or CMM room, ensure filing of back up report'
            # 'E6': 'Measurement Method / Measuring Aid',
            # 'F6': 'Gauge Type',
            # 'G6': 'Control Method (Records/GRN)',
            # 'H6': 'Sample Frequency',
            # 'I6': 'Type of Inspection',
            # 'J6': 'Reaction Plan',

        }

        # Helper function to get Excel column letter for a given index
        def get_column_letter(index):
            result = ""
            while index > 0:
                index, remainder = divmod(index - 1, 26)
                result = chr(65 + remainder) + result
            return result

        col_index = 4
        for i in range(1, max_observations + 1):
            # Get column letters for observation and date columns
            obs_col = get_column_letter(col_index)
            date_col = get_column_letter(col_index + 1)

            # Add headers
            headers[f'{obs_col}7'] = 'Part Sample No.'
            headers[f'{date_col}7'] = f'Observation {i}'

            # Merge cells
            ws.merge_cells(f'{obs_col}7:{obs_col}8')
            ws.merge_cells(f'{date_col}7:{date_col}8')

            # Set column widths
            ws.column_dimensions[obs_col].width = 15
            ws.column_dimensions[date_col].width = 18

            # Increment column index for next pair
            col_index += 2


        ws.row_dimensions[1].height = 50

        # Apply font and border styles
        for cell, value in headers.items():
            ws[cell] = value
            ws[cell].font = font_header
            ws[cell].alignment = align_center
            ws[cell].fill = fill
        ws['C1'].font = title_font

        # Calculate maximum column based on observations
        max_col_index = 3 + (self.max_observations * 2)  # Starting from column D (4) plus 2 columns per observation
        max_col_letter = get_column_letter(max_col_index)

        # Merge cells
        merge_cells = [
            'A1:B5', 'C1:K1','A6:A8', 'B6:B8', 'C6:C8',
            'C2:D2', 'C3:D3', 'C4:D4', 'C5:D5',
            'F2:G2', 'F3:G3', 'F4:G4', 'F5:G5',
            'I2:J2', 'I3:J3', 'I4:J4', 'I5:J5',
        ]
        # Only add L1 merge if we have enough observations
        if max_col_index > 11:  # If we have columns beyond K (10)
            max_col_letter = get_column_letter(max_col_index)
            merge_cells.append(f'L1:{max_col_letter}5')
            merge_cells.append(f'D6:{max_col_letter}6')
        else:
            merge_cells.append(f'D6:K6')
        for cell_range in merge_cells:
            ws.merge_cells(cell_range)

        # Set column widths
        col_widths = {
            'A': 8, 'B': 18, 'C': 20,
        }
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width
        row = 9
        for rec in self:
            ws['E2'] = rec.grn_no if rec.grn_no else ''
            ws['E3'] = rec.grn_date if rec.grn_date else ''
            ws['E4'] = rec.grn_product if rec.grn_product else ''
            ws['E5'] = rec.part_name if rec.part_name else ''
            ws['H2'] = rec.part_number if rec.part_number else ''
            ws['H3'] = rec.grn_part_number if rec.grn_part_number else ''

            ws['H4'] = rec.supplier_name.name if rec.supplier_name else ''
            ws['H5'] = rec.supplier_code if rec.supplier_code else ''

            ws['K2'] = rec.sample_qty_variable if rec.sample_qty_variable else ''
            ws['K3'] = rec.sample_qty_attribute if rec.sample_qty_attribute else ''
            ws['K4'] = rec.date if rec.date else ''


            for ele in rec.line_ids:
                ws[f'A{row}'] = ele.sr_no if ele.sr_no else ''
                ws[f'B{row}'] = ele.process_name_id.name if ele.process_name_id else ''
                ws[f'C{row}'] = ele.product_spec_tol if ele.product_spec_tol else ''
                col_index=4
                for obs in ele.observations:
                    ws.cell(row=row, column=col_index,value =obs.part_sample_no if obs.part_sample_no else '')
                    ws.cell(row=row, column=col_index+1,value =obs.observation_value if obs.observation_value else '')
                    col_index+=2


                row += 1









        cur_row = 30
        if cur_row<row:
            cur_row=row

        max_freq = 2*self.max_observations
        for rows in ws.iter_rows(min_row=1, max_row=cur_row, min_col=1, max_col=3+max_freq):
            for cell in rows:
                cell.alignment = align_center
                cell.border = border

        # Save workbook to BytesIO
        wb.save(output)
        output.seek(0)

        # Create attachment
        attachment = self.env['ir.attachment'].create({
            'name': 'GRN_Report.xlsx',
            'type': 'binary',
            'datas': base64.b64encode(output.getvalue()),
            'res_model': 'grn.inspection.report',
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        })

        # Return download link
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }
    def generate_xls_sheet_report(self):
        # Create workbook and worksheet
        output = BytesIO()
        wb = Workbook()
        ws = wb.active

        # Define styles
        border = Border(top=Side(style='thin'), left=Side(style='thin'),
                        right=Side(style='thin'), bottom=Side(style='thin'))
        font_header = Font(name='Times New Roman', bold=True)
        title_font = Font(size=20, bold=True)
        align_center = Alignment(vertical='center', horizontal='center', wrapText=True)

        # Define fill colors
        fill = PatternFill(start_color="e7e7e7", end_color="e7e7e7", fill_type="solid")  # Grey
        orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

        if self.env.user.company_id and self.env.user.company_id.logo:

            max_width, max_height = 100, 200  # Set max dimensions
            image_data = base64.b64decode(self.env.user.company_id.logo)

            # Open image with PIL
            image = PILImage.open(io.BytesIO(image_data))
            width, height = image.size
            aspect_ratio = width / height

            # Resize with aspect ratio
            if width > max_width:
                width = max_width
                height = int(width / aspect_ratio)
            if height > max_height:
                height = max_height
                width = int(height * aspect_ratio)

            # Resize and add padding
            resized_image = image.resize((width, height), PILImage.LANCZOS)
            padding_top, padding_left = 5, 5
            resized_image = ImageOps.expand(resized_image, border=(padding_left, padding_top, 0, 0), fill='white')

            img_bytes = io.BytesIO()
            resized_image.save(img_bytes, format='PNG')
            img_bytes.seek(0)
            logo_image = Image(img_bytes)

            ws.add_image(logo_image, 'A1')
        # Determine maximum number of observations dynamically
        max_observations = self.max_observations

        # Headers dictionary
        headers = {
            'C1': 'RECEIVING  AUDIT  CHECK SHEET',
            'C2': 'GRN No.',
            'C3': 'GRN Date',
            'C4': 'Product',
            'C5': 'Part Name',
            'G2': 'RIQP Number',
            'G3': 'Part Number',
            'G4': 'Supplier Name',
            'G5': 'Supplier Code',
            # 'I2': 'Total Qty.',
            # 'I3': 'Store Location',
            'K2': 'Sample Qty / Lot Qty (Nos) For Variable',
            'K3': 'Sample Qty / Lot Qty (Nos) For Attribute',
            'K4': 'Date',
            # 'I2': 'CFT Team',
            # 'I3': 'Rev No.',
            # 'I4': 'Rev Date',
            # 'C6': 'Class',
            'A6': 'Sr. No.',
            'B6': 'Characteristics',
            'C6': 'Product Specification Tolerance',
            'D6' : 'Periodic - Year- ',
            'D7' : 'Jan',
            'E7' : 'Feb',
            'F7' : 'Mar',
            'G7' : 'Apr',
            'H7' : 'May',
            'I7' : 'Jun',
            'J7' : 'Jul',
            'K7' : 'Aug',
            'L7' : 'Sep',
            'M7' : 'Oct',
            'N7' : 'Nov',
            'O7' : 'Dec',




        }



        ws.row_dimensions[1].height = 50

        # Apply font and border styles
        for cell, value in headers.items():
            ws[cell] = value
            ws[cell].font = font_header
            ws[cell].alignment = align_center
            ws[cell].fill = fill
        ws['C1'].font = title_font

        # Merge cells
        merge_cells = [
            'A1:B5', 'C1:O1', 'D6:O6','A6:A8', 'B6:B8', 'C6:C8','D7:D8',
            'E7:E8', 'F7:F8', 'G7:G8', 'H7:H8', 'I7:I8', 'J7:J8', 'K7:K8', 'L7:L8', 'M7:M8', 'N7:N8', 'O7:O8',
            'C2:D2', 'C3:D3', 'C4:D4', 'C5:D5',
            'E2:F2', 'E3:F3', 'E4:F4', 'E5:F5',
            'G2:H2', 'G3:H3', 'G4:H4', 'G5:H5',
            'I2:J2', 'I3:J3', 'I4:J4', 'I5:J5',
            'K2:M2', 'K3:M3', 'K4:M4', 'K5:M5',
            'N2:O2', 'N3:O3', 'N4:O4', 'N5:O5',
        ]
        for cell_range in merge_cells:
            ws.merge_cells(cell_range)

        # Set column widths
        col_widths = {
            'A': 8, 'B': 18, 'C': 20,
            'D': 15, 'E': 15, 'F': 15,
            'G': 15, 'H': 15, 'I': 15,
            'J': 15, 'K': 15, 'L': 15,
            'M': 15, 'N': 15, 'O': 15,
        }
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width
        row = 9
        for rec in self:
            ws['E2'] = rec.grn_no if rec.grn_no else ''
            ws['E3'] = rec.grn_date if rec.grn_date else ''
            ws['E4'] = rec.grn_product if rec.grn_product else ''
            ws['E5'] = rec.part_name if rec.part_name else ''
            ws['I2'] = rec.part_number if rec.part_number else ''
            ws['I3'] = rec.grn_part_number if rec.grn_part_number else ''

            ws['I4'] = rec.supplier_name.name if rec.supplier_name else ''
            ws['I5'] = rec.supplier_code if rec.supplier_code else ''

            ws['N2'] = rec.sample_qty_variable if rec.sample_qty_variable else ''
            ws['N3'] = rec.sample_qty_attribute if rec.sample_qty_attribute else ''
            ws['N4'] = rec.date if rec.date else ''

            for ele in rec.line_ids:
                ws[f'A{row}'] = ele.sr_no if ele.sr_no else ''
                ws[f'B{row}'] = ele.process_name_id.name if ele.process_name_id else ''
                ws[f'C{row}'] = ele.product_spec_tol if ele.product_spec_tol else ''

                # Get all attachments for this line
                attachments = ele.attchment

                # Map of month names to column letters
                month_columns = {
                    'January': 'D',
                    'February': 'E',
                    'March': 'F',
                    'April': 'G',
                    'May': 'H',
                    'June': 'I',
                    'July': 'J',
                    'August': 'K',
                    'September': 'L',
                    'October': 'M',
                    'November': 'N',
                    'December': 'O'
                }

                # Check attachments and mark 'X' in the corresponding month column
                for attachment in attachments:
                    if attachment.report_month and attachment.report_month in month_columns:
                        column = month_columns[attachment.report_month]
                        ws[f'{column}{row}'] = 'X'

                row += 1









        cur_row = 30
        if cur_row<row:
            cur_row=row

        for rows in ws.iter_rows(min_row=1, max_row=cur_row, min_col=1, max_col=15):
            for cell in rows:
                cell.alignment = align_center
                cell.border = border

        # Save workbook to BytesIO
        wb.save(output)
        output.seek(0)

        # Create attachment
        attachment = self.env['ir.attachment'].create({
            'name': 'GRN_Report_month.xlsx',
            'type': 'binary',
            'datas': base64.b64encode(output.getvalue()),
            'res_model': 'grn.inspection.report',
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        })

        # Return download link
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }




class IncomingIRLine(models.Model):
    _name = 'grn.inspection.report.line'
    _description = 'Report of Incoming Inspection Line'
    _order = 'sequence'
    _inherit = "translation.mixin"

    inspection_id = fields.Many2one('grn.inspection.report', string='Inspection Report', ondelete='cascade',
                                    required=True)
    grn_management_id = fields.Many2one('grn.management', string="GRN Management", ondelete="cascade")
    sequence = fields.Integer(string="Sequence", index=True, default=10)

    sr_no = fields.Integer(string="Sr. No./Balloon No", compute="_compute_sr_no", store=True)
    process_name_id = fields.Many2one('grn.characteristics', string="Characteristics")
    product_spec_tol = fields.Char(string="Product Specification / Tolerance",translate=True)
   # dimension_des = fields.Many2one('gdt.symbol', string="Dimension Description")
    measure_method_aid = fields.Many2one('maintenance.equipment', string="Inspection Method")
    gauge_type = fields.Selection([('variable', 'Variable'), ('attribute', 'Attribute')], string='Gauge Type',
                                  related='measure_method_aid.gauge_type')

    inspected_by = fields.Many2one('res.users', string="Inspected By", default=lambda self: self.env.user)
    observations = fields.One2many('grn.inspection.observation', 'inspection_line_id', string="Observations")
    attchment = fields.One2many('grn.attachment', 'attch_id', string="Attachments")
    # batch_quantity = fields.Integer("Batch Quantity", required=True,related='inspection_id.batch_quantity')
    sample_freq = fields.Char("Sample Frequency")
    # observation_status = fields.Selection([
    #     ('draft', 'Draft'),
    #     ('auto_filled', 'Auto Filled'),
    #     ('manually_edited', 'Manually Edited')
    # ], string="Observation Status", default='draft')

    # # Attachments
    # attachment_1 = fields.Binary(string="Attachment 1")
    # attachment_filename_1 = fields.Char(string="Filename 1")
    # attachment_2 = fields.Binary(string="Attachment 2")
    # attachment_filename_2 = fields.Char(string="Filename 2")
    # attachment_3 = fields.Binary(string="Attachment 3")
    # attachment_filename_3 = fields.Char(string="Filename 3")

    @api.depends('sequence', 'inspection_id')
    def _compute_sr_no(self):
        for order in self.mapped('inspection_id'):
            sr_no = 1
            # Sort the lines by sequence
            sorted_lines = order.line_ids.sorted(key=lambda l: l.sequence)
            for line in sorted_lines:
                line.sr_no = sr_no
                sr_no += 1

    def _generate_observations(self):
        freq = int(self.sample_freq) if self.sample_freq else 1
        """Dynamically update observation fields based on batch quantity"""
        if freq:
            self.observations = [(5, 0, 0)]  # Clear existing records
            self.observations = [(0, 0, {'observation_number': i + 1}) for i in range(freq)]

    # def _generate_observations(self):
    #     """
    #     Generate observations based on the sample frequency with safe deletion.
    #
    #     This method creates observations for the inspection line based on the specified sample frequency.
    #     It uses a safe deletion approach to handle existing observations.
    #
    #     Raises:
    #         ValidationError: If the sample frequency is not a valid integer
    #     """
    #     for record in self:
    #         # Clean and validate sample frequency
    #         try:
    #             # If sample_freq is empty or None, default to 1
    #             freq = int(record.sample_freq) if record.sample_freq else 1
    #         except (ValueError, TypeError):
    #             freq = 1
    #
    #         # Safe deletion of existing observations
    #         existing_observations = self.env['inspection.observation'].search([
    #             ('inspection_line_id', '=', record.id)
    #         ])
    #
    #         # Instead of unlink, archive the records if possible
    #         if existing_observations:
    #             try:
    #                 existing_observations.write({'active': False})
    #             except Exception:
    #                 # If archiving fails, log the error
    #                 self.env.cr.rollback()
    #                 self.env['ir.logging'].create({
    #                     'name': 'Observation Deletion Error',
    #                     'type': 'server',
    #                     'dbname': self.env.cr.dbname,
    #                     'level': 'ERROR',
    #                     'message': f'Could not archive observations for inspection line {record.id}'
    #                 })
    #
    #         # Generate observations
    #         observations = []
    #         for seq in range(1, freq + 1):
    #             observation_vals = {
    #                 'inspection_line_id': record.id,
    #                 'observation_number': seq,
    #
    #             }
    #             observations.append((0, 0, observation_vals))
    #
    #         # Update the observations
    #         record.write({
    #             'observations': observations,
    #             'observation_status': 'auto_filled'
    #         })
    #




    def action_generate_observations(self):
        """Action triggered by the button to generate observations."""
        for record in self:
            record._generate_observations()




class InspectionObservation(models.Model):
    _name = 'grn.inspection.observation'
    _description = 'Inspection Observations'
    _inherit = "translation.mixin"







    inspection_line_id = fields.Many2one(
        'grn.inspection.report.line',
        string="Inspection Line",
        ondelete="cascade",  # Keep restrict or change to cascade if preferred
        required=True
    )

    observation_number = fields.Integer(string="Sr.No.", required=True)
    part_sample_no = fields.Char(string="Part Sample No.")

    observation_value = fields.Char(string="Observation",translate=True)

    # # time_date_inspection = fields.Datetime(string="Time & Date of Inspection")
    # attachment_1 = fields.Binary(string="Attachment 1")
    # attachment_filename_1 = fields.Char(string="Filename 1")
    # attachment_2 = fields.Binary(string="Attachment 2")
    # attachment_filename_2 = fields.Char(string="Filename 2")
    # attachment_3 = fields.Binary(string="Attachment 3")
    # attachment_filename_3 = fields.Char(string="Filename 3")
#
# class GrnGaugeType(models.Model):
#     _inherit = 'maintenance.equipment'
#
#     gauge_type = fields.Selection([('variable', 'Variable'), ('attribute', 'Attribute')], string='Gauge Type')

class GrnAttachment(models.Model):
    _name = 'grn.attachment'
    _description = 'GRN Attachment'


    attch_id=fields.Many2one('grn.inspection.report.line', string="GRN Reference", ondelete="cascade")
    report_filename = fields.Char(string="Filename")
    report_attachment = fields.Binary(string="Report Attachment")
    report_month = fields.Char(string="Report Month")


    @api.model
    def create(self, vals):
        if vals.get('report_attachment'):
            vals['report_month'] = datetime.today().strftime('%B')
        return super(GrnAttachment, self).create(vals)

    def write(self, vals):
        if 'report_attachment' in vals and vals['report_attachment']:
            vals['report_month'] = datetime.today().strftime('%B')
        return super(GrnAttachment, self).write(vals)
