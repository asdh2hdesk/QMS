from odoo import fields, models, api, _
from odoo.exceptions import ValidationError
from datetime import date
import openpyxl
from io import BytesIO
import io
from openpyxl import Workbook
import openpyxl
import base64
from bs4 import BeautifulSoup
from openpyxl.styles import Alignment, Font, Border, Side, DEFAULT_FONT, PatternFill
from openpyxl.drawing.image import Image
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
import copy
import datetime
from PIL import ImageOps
from PIL import Image as PILImage
from bs4 import BeautifulSoup


class ControlPlan(models.Model):
    _name = "grn.control.plan"
    _description = "GRN Control Plan"
    _rec_name = 'part_id'

    # supplier_ids = fields.Many2one('res.partner',string='Suppliers')
    # supplier_code = fields.Char("Supplier Code",related='supplier_ids.supplier_code')

    # vehicle_model = fields.Char("Vehicle / Model")
    # doc_no = fields.Char("Doc. No.")
    part_id = fields.Many2one("product.template", string="reference")
    part_name = fields.Char("Part Name", related="part_id.name")
    part_number = fields.Char("RIQP Number", related="part_id.default_code")
    grn_part_number = fields.Char("Part Number",related="part_id.grn_part_number")
    grn_product = fields.Char("Product",related="part_id.grn_product")
    date_origin = fields.Date("Date (Orig.)")
    rev_no = fields.Char("Rev. No")
    rev_date= fields.Date("Rev. Date")
    key_contact = fields.Many2one('hr.employee', string="Key Contact")
    telephone = fields.Char(string='Telephones', compute='_compute_telephones')
    team_cft=fields.Many2many('hr.employee', string="CFT Team")
    approved_by = fields.Many2one('res.users', string="Approved By")
    grn_report_attachment = fields.Many2many('ir.attachment', 'ins_report_attachment_rel', 'inspection_id',
                                             'attachment_id',
                                             string="Report Attachments")

    product_type= fields.Selection([('regular','Regular Product'),('dol','DOL Product')],string="Product Type")

    grn_process_line_ids = fields.One2many(
        comodel_name='grn.control.plan.process',
        inverse_name='process_id',
        string='Process'
    )
    # Revision History
    revision_history_ids = fields.One2many(
        comodel_name='grn.revision.history',
        inverse_name='control_plan_id',
        string='Revision History'
    )

    generate_xls_file = fields.Binary(string="Generated File")

    @api.model
    def create(self, vals):
        record = super(ControlPlan, self).create(vals)
        # Create initial revision history entry
        self.env['grn.revision.history'].create({
            'control_plan_id': record.id,
            'serial_no': 1,
            'rev_no': record.rev_no,
            'rev_date': record.rev_date,
            'revised_by': self.env.user.id,
            'approved_by': record.approved_by.id,
        })
        return record

    def write(self, vals):
        # Check if revision-related fields are updated
        if 'rev_no' in vals or 'rev_date' in vals:
            # Get the latest serial number
            latest_revision = self.env['grn.revision.history'].search(
                [('control_plan_id', '=', self.id)],
                order='serial_no desc',
                limit=1
            )
            next_serial = latest_revision.serial_no + 1 if latest_revision else 1

            # Create a new revision history entry
            self.env['grn.revision.history'].create({
                'control_plan_id': self.id,
                'serial_no': next_serial,
                'rev_no': vals.get('rev_no', self.rev_no),
                'rev_date': vals.get('rev_date', self.rev_date), # Default text, user can update later
                'revised_by': self.env.user.id,

            })



        return super(ControlPlan, self).write(vals)

    def generate_xls_report(self):
        # Create workbook and worksheet
        output = BytesIO()
        wb = Workbook()
        ws = wb.active

        # Define styles
        border = Border(top=Side(style='thin'), left=Side(style='thin'),
                        right=Side(style='thin'), bottom=Side(style='thin'))
        font_header = Font(name='Times New Roman', bold=True)
        title_font=Font(size=20, bold=True)
        align_center = Alignment(vertical='center', horizontal='center', wrapText=True)

        # Define fill colors
        fill = PatternFill(start_color="e7e7e7", end_color="e7e7e7", fill_type="solid")  # Grey
        orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")


        if self.env.user.company_id and self.env.user.company_id.logo:

            max_width, max_height = 100, 200  # Set max dimensions
            image_data = base64.b64decode(self.env.user.company_id.logo)

            # Open image with PIL
            image = PILImage.open(io.BytesIO(image_data))
            width, height = image.size
            aspect_ratio = width / height

            # Resize with aspect ratio
            if width > max_width:
                width = max_width
                height = int(width / aspect_ratio)
            if height > max_height:
                height = max_height
                width = int(height * aspect_ratio)

            # Resize and add padding
            resized_image = image.resize((width, height), PILImage.LANCZOS)
            padding_top, padding_left = 5,5
            resized_image = ImageOps.expand(resized_image, border=(padding_left, padding_top, 0, 0), fill='white')

            img_bytes = io.BytesIO()
            resized_image.save(img_bytes, format='PNG')
            img_bytes.seek(0)
            logo_image = Image(img_bytes)

            ws.add_image(logo_image, 'A1')



        # Headers dictionary
        headers = {
            'C1':'Receipt Inspection Quality Plan (RIQP)',

            'C2':'Product',
            'C3': 'RIQP Number',
            'C4':'Part Name',
            'C5':'Part Number',
            'C6': 'Date (Orig.)',
            'G2':'Key Contact',
            'G3':'Telephone',
            'G4':'CFT Team',
            'G5':'Rev No.',
            'G6':'Rev Date',
            'C7':'Class',
            'A7':'Sr. No.',
            'B7' : 'Characteristics',
            'D7' : 'Product Specification Tolerance',
            'E7' : 'Measurement Method / Measuring Aid',
            'F7' : 'Gauge Type',
            'G7' : 'Control Method (Records/GRN)',
            'H7' : 'Sample Frequency',
            'I7' : 'Type of Inspection',
            'J7' : 'Reaction Plan',




        }

        ws.row_dimensions[1].height = 50

        # Apply font and border styles
        for cell, value in headers.items():
            ws[cell] = value
            ws[cell].font = font_header
            ws[cell].alignment = align_center
            ws[cell].fill=fill
        ws['C1'].font = title_font

        # Merge cells
        merge_cells = [
            'A1:B6', 'C1:J1',
            'C2:D2', 'C3:D3', 'C4:D4', 'C5:D5', 'C6:D6',
            'E2:F2', 'E3:F3', 'E4:F4', 'E5:F5', 'E6:F6',
            'G2:H2', 'G3:H3', 'G4:H4', 'G5:H5', 'G6:H6',
            'I2:J2', 'I3:J3', 'I4:J4', 'I5:J5', 'I6:J6',
        ]
        for cell_range in merge_cells:
            ws.merge_cells(cell_range)

        # Set column widths
        col_widths = {
            'A': 8,'B': 20, 'C': 18, 'D': 20, 'E': 20,
            'F': 18,'G':18,'H':18,'I':18,'J':18
        }
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width
        row=8
        for rec in self:
            ws['E2'] = rec.grn_product if rec.grn_product else ''
            ws['E3'] = rec.part_number if rec.part_number else ''
            ws['E4'] = rec.part_name if rec.part_name else ''
            ws['E5'] = rec.grn_part_number if rec.grn_part_number else ''
            ws['E6'] = rec.date_origin if rec.date_origin else ''
            ws['I2'] = rec.key_contact.name if rec.key_contact else ''
            ws['I3'] = rec.telephone if rec.telephone else ''
            # Handle Many2many field team_cft
            if hasattr(rec, 'team_cft') and rec.team_cft:
                team_names = rec.team_cft.mapped('name')  # Get list of names
                ws['I4'] = ', '.join(team_names) if team_names else ''  # Join with comma and space
            else:
                ws['I4'] = ''  # Set empty string if team_cft doesn't exist or is empty
            ws['I5'] = rec.rev_no if rec.rev_no else ''
            ws['I6'] = rec.rev_date if rec.rev_date else ''
            for line in rec.grn_process_line_ids:
                ws[f'A{row}'] = line.process_step if line.process_step else ''
                ws[f'B{row}'] = line.process_name_id.name if line.process_name_id else ''
                ws[f'D{row}'] = line.product_spec_tol if line.product_spec_tol else ''
                ws[f'E{row}'] = line.measure_method_aid.name if line.measure_method_aid else ''
                ws[f'F{row}'] = line.gauge_type if line.gauge_type else ''
                ws[f'G{row}'] = line.control_method.name if line.control_method else ''
                ws[f'H{row}'] = line.control_sample_freq.name if line.control_sample_freq else ''
                ws[f'I{row}'] = line.type_of_inspection if line.type_of_inspection else ''
                ws[f'J{row}'] = line.reaction_plan.name if line.reaction_plan else ''
                row += 1






        cur_row = 30
        if cur_row<row:
            cur_row=row
        # Add spacing between tables
        cur_row += 2

        # Add Revision History Title
        ws.merge_cells(f'A{cur_row}:J{cur_row}')
        ws[f'A{cur_row}'] = "Revision History"
        ws[f'A{cur_row}'].font = title_font
        ws[f'A{cur_row}'].alignment = align_center
        ws[f'A{cur_row}'].fill = fill

        ws.row_dimensions[cur_row].height = 30
        cur_row += 1

        # Add Revision History Headers
        revision_headers = {
            'A': 'Sr. No.',
            'B': 'Rev. No.',
            'C': 'Rev. Date',
            'D': ' Rev Details',

            'G': 'Revised By',
            'I': 'Approved By'
        }

        for col, header in revision_headers.items():
            ws[f'{col}{cur_row}'] = header
            ws[f'{col}{cur_row}'].font = font_header
            ws[f'{col}{cur_row}'].alignment = align_center
            ws[f'{col}{cur_row}'].fill = fill

        # Merge Details and Reason
        ws.merge_cells(f'D{cur_row}:F{cur_row}')
        ws.merge_cells(f'G{cur_row}:H{cur_row}')  # Merge Revised By
        ws.merge_cells(f'I{cur_row}:J{cur_row}')  # Merge Approved By

        # Fill revision history data
        rev_row = cur_row + 1
        for rev in self.revision_history_ids:
            ws[f'A{rev_row}'] = rev.serial_no if rev.serial_no else ''
            ws[f'B{rev_row}'] = rev.rev_no if rev.rev_no else ''
            ws[f'C{rev_row}'] = rev.rev_date if rev.rev_date else ''

            ws[f'D{rev_row}'] = rev.revision_details if rev.revision_details else ''


            ws[f'G{rev_row}'] = rev.revised_by.name if rev.revised_by else ''
            ws[f'I{rev_row}'] = rev.approved_by.name if rev.approved_by else ''


            ws.merge_cells(f'D{rev_row}:F{rev_row}')  # Merge Rev. Date
            ws.merge_cells(f'G{rev_row}:H{rev_row}')  # Merge Revised By
            ws.merge_cells(f'I{rev_row}:J{rev_row}')  # Merge Approved By

            rev_row += 1

        # If no revision history records, add some empty rows
        if rev_row == cur_row + 1:
            for i in range(3):  # Add 3 empty rows

                ws.merge_cells(f'D{rev_row}:F{rev_row}')
                ws.merge_cells(f'G{rev_row}:H{rev_row}')
                ws.merge_cells(f'I{rev_row}:J{rev_row}')
                rev_row += 1

        for rows in ws.iter_rows(min_row=1, max_row=rev_row-1, min_col=1, max_col=10):
            for cell in rows:
                cell.alignment = align_center
                cell.border = border





        # Save workbook to BytesIO
        wb.save(output)
        output.seek(0)

        # Create attachment
        attachment = self.env['ir.attachment'].create({
            'name': 'GRN_Report.xlsx',
            'type': 'binary',
            'datas': base64.b64encode(output.getvalue()),
            'res_model': 'grn.management',
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        })

        # Return download link
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }

    def _compute_telephones(self):
        for record in self:
            # Extract and join phone numbers from all related employees
            telephones = record.key_contact.mapped('work_phone')
            record.telephone = ', '.join(t for t in telephones if isinstance(t, str) and t)


    @api.depends('key_contact')
    def _compute_telephones(self):
        for record in self:
            # Extract and join phone numbers from all related employees
            telephones = record.key_contact.mapped('work_phone')
            record.telephone = ', '.join(t for t in telephones if isinstance(t, str) and t)


class ControlProcess(models.Model):
    _name = "grn.control.plan.process"
    _description = "Control Process"
    _rec_name = 'process_step'
    _inherit = "translation.mixin"

    process_id = fields.Many2one('grn.control.plan', 'Control Plan Process',ondelete='cascade', index=True, copy=False)

    # process_step = fields.Char("Process Step")
    sequence = fields.Integer(string="Sequence", default=10)
    process_step = fields.Integer("S.No", compute="_compute_sequence_number")
    process_name_id = fields.Many2one('grn.characteristics', string="Characteristics")
    control_sample_freq = fields.Many2one('grn.sample.frequency', "Sample Frequency")
    char_class = fields.Many2one('grn.process.class', "Class")
    product_spec_tol = fields.Char("Product Specification / Tolerance",translate=True)
    measure_method_aid = fields.Many2one('maintenance.equipment',"Measurement Method / Measuring Aid")
    sample_freq = fields.Char("Sample Frequency")
    control_method = fields.Many2one('grn.control.method', "Control Method")  # Control
    reaction_plan = fields.Many2one('grn.reaction.plan', "Reaction Plan")

    type_of_inspection= fields.Selection([('visual','Visual'), ('dimensional','Dimensional'),('properties','Properties')], string="Type of Inspection",related='measure_method_aid.type_of_inspection')
    gauge_type = fields.Selection([('variable', 'Variable'), ('attribute', 'Attribute')], string='Gauge Type', related='measure_method_aid.gauge_type')

    measure_method_aid_details = fields.Char(
        string="Measurement Method Details",
        compute="_compute_measure_method_details",
        store=True,
        help="Combined details of measurement method/aid"
    )

    @api.depends('measure_method_aid')
    def _compute_measure_method_details(self):
        for record in self:
            if record.measure_method_aid:
                # Combine name, code, and least count into a single string
                record.measure_method_aid_details = " | ".join(filter(bool, [
                    record.measure_method_aid.name or '',
                    record.measure_method_aid.serial_no or '',
                    record.measure_method_aid.code or '',
                    record.measure_method_aid.lc or ''
                ]))
            else:
                record.measure_method_aid_details = False

    # format_rev = fields.char("Format Revision")
    # cft = fields.char("CFT")  # many to many employeee
    #
    # code, name, leastcount
    #
    # supplier_code = fields.char("Supplier Code")

    @api.depends('sequence', 'process_id')
    def _compute_sequence_number(self):
        for order in self.mapped('process_id'):
            process_step = 1
            for lines in order.grn_process_line_ids:
                lines.process_step = process_step
                process_step += 1

    class GrnRevisionHistory(models.Model):
        _name = "grn.revision.history"
        _description = "GRN Control Plan Revision History"
        _order = "serial_no desc"
        _inherit = "translation.mixin"

        control_plan_id = fields.Many2one('grn.control.plan', string='Control Plan', ondelete='cascade')
        serial_no = fields.Integer(string="Sr. No.", readonly=True)
        rev_no = fields.Char(string="Rev. No.", readonly=True)
        rev_date = fields.Date(string="Rev. Date", readonly=True)
        revision_details = fields.Text(string="Revision Details",translate=True)
        revised_by = fields.Many2one('res.users', string="Revised By", readonly=True)
        approved_by = fields.Many2one('res.users', string="Approved By")

        # For display in tree view
        part_name = fields.Char(related="control_plan_id.part_name", string="Part Name", store=True)
        part_number = fields.Char(related="control_plan_id.part_number", string="RIQP Number", store=True)
