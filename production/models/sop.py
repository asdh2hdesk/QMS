from odoo import fields, models, api, _
import base64
import io
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
from openpyxl.utils import get_column_letter


class StandardOperatingProcedure(models.Model):
    _name = "standard.operating.procedure"
    _description = "Standard Operating Procedure"
    _inherit = 'iatf.sign.off.members'

    sop_no = fields.Char('SOP No.', required=True, copy=False,
                        readonly=True, default=lambda self: _('New'))
    sop_description = fields.Char('SOP Description', required=True)
    prn_number = fields.Char('PRN Number')
    issued_on = fields.Date('Issued On')
    rev_date = fields.Date('Rev. Date')
    rev_no = fields.Char('Rev. No.')
    rev_details = fields.Char('Rev. Details')
    revised_by = fields.Many2one('res.users', 'Revised By')
    model_name = fields.Char('Model Name')
    stage = fields.Char('Stage')
    shop = fields.Char('Shop')
    doc_no = fields.Char('Doc No')
    page_no = fields.Char('Page No.')


    extra_care = fields.Many2many('extra.care.alerts.line', 'extra_care_rel', 'extra_id', 'alerts_id')
    settings_ids = fields.Many2many('sop.settings', 'settings_rel', 'settings_id', 'sop_id')
    safety_ids = fields.Many2many('sop.safety', 'safety_rel', 'safety_id', 'sop_id')
    process_and_tooling_ids = fields.One2many(
        comodel_name='process.and.tooling',
        inverse_name='sop_id',
        string='Process and Tooling'
    )

    @api.model_create_multi
    def create(self, vals_list):
        for vals in vals_list:
            if vals.get('sop_no', _('New')) == _('New'):  # Changed from 'name' to 'sop_no'
                vals['sop_no'] = self.env['ir.sequence'].next_by_code('standard.operating.procedure') or _('New')
        return super().create(vals_list)

    def name_get(self):
        """Display SOP No. and Description"""
        result = []
        for record in self:
            name = f"[{record.sop_no}] {record.sop_description}" if record.sop_description else record.sop_no
            result.append((record.id, name))
        return result
    
    # Write Helper Function for convert bytes to 'PNG' image
    def insert_resized_image(self, ws, base64_img, cell, max_w=230, max_h=100):
        "Insert a resized image into Excel without repeating the same code."

        if not base64_img:
            return

        # Decode base64
        image_data = base64.b64decode(base64_img)
        image = PILImage.open(BytesIO(image_data))

        # Resize logic
        width, height = image.size
        aspect_ratio = width / height

        if width > max_w:
            width = max_w
            height = int(width / aspect_ratio)

        if height > max_h:
            height = max_h
            width = int(aspect_ratio * height)

        # Convert to PNG in memory
        img_bytes = BytesIO()
        image = image.resize((width, height), PILImage.LANCZOS)
        image.save(img_bytes, format="PNG")
        img_bytes.seek(0)
        
        img = Image(img_bytes)

        # Add to Excel
        ws.add_image(img, cell)
        
    
    def generate_excel_report(self):
        """Generate Excel report for Daily Permit Work"""
        output = BytesIO()
        wb = Workbook()
        ws = wb.active
        
        # Add logo in header if exists
        if self.env.user.company_id.logo:
            self.insert_resized_image(ws,self.env.user.company_id.logo,'A1')
            
        # Define Formating
        border = Border(top=Side(style='thin'), left=Side(style='thin'),right=Side(style='thin'), bottom=Side(style='thin'))
        white_side = Side(border_style="thin", color="FFFFFF")  # white color
        white_border = Border(left=white_side, right=white_side, top=white_side, bottom=white_side)
        align_center = Alignment(horizontal='center', vertical='center',wrap_text=True)
        align_left = Alignment(horizontal='left', vertical='center', wrapText=True)
        font_header = Font(name='Arial', size=20, bold=True)
        font_title = Font(name='Arial', size=11, bold=True)
        font_all = Font(name='Arial', size=11, bold=False)
        white_font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
        grey_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        blue_fill = PatternFill(start_color='002060', end_color='002060', fill_type='solid')
        light_blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")


        # Column widths
        for i in range(1,19):
            col = get_column_letter(i)
            ws.column_dimensions[col].width = 15

        
        # ======= ROW 1 : Company Name =======
        ws.merge_cells('A1:R1')
        ws['A1'] = f"{self.env.company.name}\n\nStandard Operating Procedure"
        ws['A1'].font = font_header
        ws['A1'].alignment = align_center
        ws.row_dimensions[1].height = 80  # Logo & company name
        
        # =========================== Hepler Function ===================================
        def write_cell(ws, row, col_start, col_end, value=None, fill=None, font=None, align=None, height=27,border=None):
            # merge cell
            ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
            cell = ws.cell(row=row, column=col_start)

            if value is not None:
                cell.value = value
            if fill:
                cell.fill = fill
            if font:
                cell.font = font
            if align:
                cell.alignment = align
            if height:
                ws.row_dimensions[row].height = height
            if border:
                for col in range(col_start, col_end + 1):
                    ws.cell(row=row, column=col).border = border

        # ============================= End Function ====================================

        # ============================= Data Filling ====================================
        cur_row = 2


        # ------------------------ Basic Information ---------------------------
        # Add heading text 
        write_cell(ws, cur_row, 1, 18,"Basic Information",font=white_font,fill=blue_fill)

        cur_row += 1
        # Add 'SOP No.'
        write_cell(ws, cur_row, 1, 4,"SOP No.",font=font_title,fill=grey_fill)
        write_cell(ws, cur_row, 5, 9,self.sop_no)
        # Add 'Shop'
        write_cell(ws, cur_row, 10, 13,"Shop",font=font_title,fill=grey_fill)
        write_cell(ws, cur_row, 14, 18,self.shop)
        
        cur_row += 1
        # Add 'Doc Type'
        write_cell(ws, cur_row, 1, 4,"Doc Type",font=font_title,fill=grey_fill)
        write_cell(ws, cur_row, 5, 9,self.doc_type)
        # Add 'Doc No'
        write_cell(ws, cur_row, 10, 13,"Doc No",font=font_title,fill=grey_fill)
        write_cell(ws, cur_row, 14, 18,self.doc_no)
        
        cur_row += 1
        # Add 'Part Reference'
        write_cell(ws, cur_row, 1, 4,"Part Reference",font=font_title,fill=grey_fill)
        write_cell(ws, cur_row, 5, 9,f"[{self.part_number}]{self.part_name}")
        # Add 'Page No.'
        write_cell(ws, cur_row, 10, 13,"Page No.",font=font_title,fill=grey_fill)
        write_cell(ws, cur_row, 14, 18,self.page_no)
        
        cur_row += 1
        # Add 'Part Name'
        write_cell(ws, cur_row, 1, 4,"Part Name",font=font_title,fill=grey_fill)
        write_cell(ws, cur_row, 5, 9,self.part_name)
        # Add 'Issued On'
        write_cell(ws, cur_row, 10, 13,"Issued On",font=font_title,fill=grey_fill)
        write_cell(ws, cur_row, 14, 18,self.issued_on)
        
        cur_row += 1
        # Add 'Part Number'
        write_cell(ws, cur_row, 1, 4,"Part Number",font=font_title,fill=grey_fill)
        write_cell(ws, cur_row, 5, 9,self.part_number)
        # Add 'Rev. Date'
        write_cell(ws, cur_row, 10, 13,"Rev. Date",font=font_title,fill=grey_fill)
        write_cell(ws, cur_row, 14, 18,self.rev_date)
        
        cur_row += 1
        # Add 'SOP Description'
        write_cell(ws, cur_row, 1, 4,"SOP Description",font=font_title,fill=grey_fill)
        write_cell(ws, cur_row, 5, 9,self.sop_description)
        # Add 'Rev. No.'
        write_cell(ws, cur_row, 10, 13,"Rev. No.",font=font_title,fill=grey_fill)
        write_cell(ws, cur_row, 14, 18,self.rev_no)
        
        cur_row += 1
        # Add 'PRN Number'
        write_cell(ws, cur_row, 1, 4,"PRN Number",font=font_title,fill=grey_fill)
        write_cell(ws, cur_row, 5, 9,self.prn_number)
        # Add 'Rev. Details'
        write_cell(ws, cur_row, 10, 13,"Rev. Details",font=font_title,fill=grey_fill)
        write_cell(ws, cur_row, 14, 18,self.rev_details)
        
        cur_row += 1
        # Add 'Model Name'
        write_cell(ws, cur_row, 1, 4,"Model Name",font=font_title,fill=grey_fill)
        write_cell(ws, cur_row, 5, 9,self.model_name)
        # Add 'Revised By'
        write_cell(ws, cur_row, 10, 13,"Revised By",font=font_title,fill=grey_fill)
        write_cell(ws, cur_row, 14, 18,self.revised_by.name)
        
        cur_row += 1
        # Add 'Stage'
        write_cell(ws, cur_row, 1, 4,"Stage",font=font_title,fill=grey_fill)
        write_cell(ws, cur_row, 5, 18,self.stage)

        
        
        # ------------------------ Process and Tooling ---------------------------
        # Add heading text 
        cur_row += 1
        write_cell(ws, cur_row, 1, 18,"Process and Tooling",font=white_font,fill=blue_fill)


        cur_row += 1
        # Add 'Sr. No'
        write_cell(ws, cur_row, 1, 1,"Sr. No",font=font_title,fill=grey_fill)
        # Add 'Process Details'
        write_cell(ws, cur_row, 2, 3,"Process Details",font=font_title,fill=grey_fill)
        # Add 'Control Point' 
        write_cell(ws, cur_row, 4, 5,"Control Point",font=font_title,fill=grey_fill)
        # Add 'Spec./Torque'
        write_cell(ws, cur_row, 6, 7,"Spec./Torque",font=font_title,fill=grey_fill)
        # Add 'Tools'
        write_cell(ws, cur_row, 8, 10,"Tools",font=font_title,fill=grey_fill)
        # Add 'Procedure Image'
        write_cell(ws, cur_row, 11, 11,"Procedure Image",font=font_title,fill=grey_fill)
        # Add 'Bill of Materials'
        write_cell(ws, cur_row, 12, 13,"Item Code",font=font_title,fill=grey_fill)
        write_cell(ws, cur_row, 14, 17,"Des.",font=font_title,fill=grey_fill)
        write_cell(ws, cur_row, 18, 18,"Quantity",font=font_title,fill=grey_fill)

        
        # Add dynamic rows
        for rec in self.process_and_tooling_ids:
            bom_count = len(rec.bom_ids)
            if bom_count == 0:
                bom_count = 1
                
            cur_row += 1
            start_row = cur_row
            end_row = cur_row + bom_count - 1
            
            ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
            ws.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=3)
            ws.merge_cells(start_row=start_row, start_column=4, end_row=end_row, end_column=5)
            ws.merge_cells(start_row=start_row, start_column=6, end_row=end_row, end_column=7)
            ws.merge_cells(start_row=start_row, start_column=8, end_row=end_row, end_column=10)
            ws.merge_cells(start_row=start_row, start_column=11, end_row=end_row, end_column=11)
            

            # Sr. No
            write_cell(ws, cur_row, 1, 1, rec.sr_no or "")

            # Process Details
            write_cell(ws, cur_row, 2, 3, rec.process_details or "")

            # Control Point
            write_cell(ws, cur_row, 4, 5, rec.control_point or "")

            # Spec./Torque
            write_cell(ws, cur_row, 6, 7, rec.spec_torque or "")

            # Tools (multiple values in same cell with new lines)
            tools_list = [tool.name for tool in rec.tools if tool.name]
            tools_text = "\n".join(tools_list)   # newline for Excel
            write_cell(ws, cur_row, 8, 10, tools_text)
            
            
            # Procedure Image
            write_cell(ws, cur_row, 11, 11)
            if rec.procedure_image:
                img_data = base64.b64decode(rec.procedure_image)
                pil_img = PILImage.open(BytesIO(img_data))

                # PERFECT BALANCED SIZE
                target_w = 90   # width in pixels
                target_h = 70    # height in pixels

                pil_img = pil_img.resize((target_w, target_h), PILImage.LANCZOS)

                # Convert back to openpyxl image
                buf = BytesIO()
                pil_img.save(buf, format="PNG")
                buf.seek(0)
                xl_img = Image(buf)

                cell = f"K{cur_row}"
                xl_img.anchor = cell

                ws.add_image(xl_img)
                

            row = start_row
            for bom in rec.bom_ids:
                # Bill of Materials
                write_cell(ws, row, 12, 13, bom.item_code or "")
                write_cell(ws, row, 14, 17, bom.description or "")
                if bom_count <= 1:
                    write_cell(ws, row, 18, 18, bom.quantity or "",height=60)
                else:
                    write_cell(ws, row, 18, 18, bom.quantity or "")
                row += 1
            
            cur_row = end_row 

        # ------------------------ Setting & Safety ---------------------------
        # Add heading text 
        cur_row += 1
        write_cell(ws, cur_row, 1, 18,"Setting & Safety",font=white_font,fill=blue_fill)
        
        cur_row += 1
        # Add Settings
        write_cell(ws, cur_row, 1, 4,"Settings",font=font_title,fill=grey_fill)
        write_cell(ws, cur_row, 5, 18,self.settings_ids.name)
        
        cur_row += 1
        # Add Safety
        write_cell(ws, cur_row, 1, 4,"Safety",font=font_title,fill=grey_fill)
        write_cell(ws, cur_row, 5, 18,self.safety_ids.name)
        
        cur_row += 1
        # Add Extra Care
        write_cell(ws, cur_row, 1, 4,"Extra Care",font=font_title,fill=grey_fill)
        write_cell(ws, cur_row, 5, 18,self.extra_care.name)
        
        # ------------------------ Initiator ---------------------------
        # Add heading text 
        cur_row += 1
        write_cell(ws, cur_row, 1, 18,"Initiator",font=white_font,fill=blue_fill)
        
        cur_row += 1
        # Add Created by
        write_cell(ws, cur_row, 1, 4,"Created by",font=font_title,fill=grey_fill)
        write_cell(ws, cur_row, 5, 18,self.create_uid.name)
        
        cur_row += 1
        # Add Created by
        write_cell(ws, cur_row, 1, 4,"Created on",font=font_title,fill=grey_fill)
        write_cell(ws, cur_row, 5, 18,self.create_date)
        
        
        # ------------------------ Approvers ---------------------------
        # Add heading text 
        cur_row += 1
        write_cell(ws, cur_row, 1, 18,"Approvers",font=white_font,fill=blue_fill)
        
        cur_row += 1
        # Add 'Approver'
        write_cell(ws, cur_row, 1, 3,"Approver",font=font_title,fill=grey_fill)
        # Add 'Department'
        write_cell(ws, cur_row, 4, 6,"Department",font=font_title,fill=grey_fill)
        # Add 'Comment' 
        write_cell(ws, cur_row, 7, 10,"Comment",font=font_title,fill=grey_fill)
        # Add 'Approved/Rejected Date'
        write_cell(ws, cur_row, 11, 14,"Approved/Rejected Date",font=font_title,fill=grey_fill)
        # Add 'Status'
        write_cell(ws, cur_row, 15, 18,"Status",font=font_title,fill=grey_fill)


        # Add dynamic rows
        for rec in self.iatf_members_ids:
            cur_row += 1

            # Approver
            write_cell(ws, cur_row, 1, 3, rec.approver_id.name or "")

            # Department
            write_cell(ws, cur_row, 4, 6, rec.department_id.name or "")

            # Comment
            write_cell(ws, cur_row, 7, 10, rec.comment or "")

            # Approved/Rejected Date
            write_cell(ws, cur_row, 11, 14, rec.date_approved_rejected or "")

            # Status
            write_cell(ws, cur_row, 15, 18, rec.approval_status or "")

        # Add blank space
        cur_row += 1
        write_cell(ws, cur_row, 1, 18)
      
        # ======================== End Data Filling ====================================

        # add formating in all rows
        for row in range(1, cur_row + 1):
            for col_idx in range(1, 19):
                col = get_column_letter(col_idx)
                cell = ws[f"{col}{row}"]
                
                if cell.font == Font():
                    cell.font = font_all
                if cell.border == white_border:
                    continue
                else:
                    cell.border = border
                if cell.alignment == Alignment():
                    cell.alignment = align_left


        # Save workbook
        wb.save(output)
        output.seek(0)

        xls_file = base64.b64encode(output.read())
        output.close()

        # Create attachment
        filename = f'sop.xlsx'
        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'type': 'binary',
            'datas': xls_file,
            'res_model': self._name,
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        })

        # Return download action
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }



class ProcessAndTooling(models.Model):
    _name = "process.and.tooling"
    _description = "Process and Tooling"
    _inherit = 'iatf.sign.off.members'

    sop_id = fields.Many2one('standard.operating.procedure', string='SOP')
    sr_no = fields.Char('Sr. No.')
    process_details = fields.Char('Process Details')
    control_point = fields.Char('Control Point')
    spec_torque = fields.Char('Spec./Torque')
    tools = fields.Many2many('maintenance.equipment', string="Tools")
    procedure_image = fields.Image("Procedure Image")
    bom_ids = fields.One2many(
        comodel_name='bill.of.materials',
        inverse_name='material_id',
        string='Bill of Materials'
    )

    def name_get(self):
        """Display Sr. No. and Process Details"""
        result = []
        for record in self:
            name = f"[{record.sr_no}] {record.process_details}" if record.process_details else (record.sr_no or 'Process')
            result.append((record.id, name))
        return result


class BillOfMaterials(models.Model):
    _name = "bill.of.materials"
    _description = "Bill of Materials"
    _inherit = 'iatf.sign.off.members'

    material_id = fields.Many2one('process.and.tooling', string='Material')
    item_code = fields.Char('Item Code')
    description = fields.Char('Description')
    quantity = fields.Char('Quantity')

    def name_get(self):
        """Display Item Code and Description"""
        result = []
        for record in self:
            name = f"[{record.item_code}] {record.description}" if record.item_code and record.description else (record.item_code or record.description or 'BOM Item')
            result.append((record.id, name))
        return result


class SOPSettings(models.Model):
    _name = "sop.settings"
    _description = "SOP Settings"

    name = fields.Char('Settings', required=True)


class SOPSafety(models.Model):
    _name = "sop.safety"
    _description = "SOP Safety"

    name = fields.Char('Safety Information', required=True)