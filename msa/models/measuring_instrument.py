from odoo import models, fields, api
import logging
import math
from odoo import http
from odoo.http import request
import base64
import io
from io import BytesIO
import logging
from openpyxl import Workbook
from odoo.modules.module import get_module_resource
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, PatternFill
from openpyxl.styles import Font, Border, Side
from PIL import Image as PILImage
from matplotlib import pyplot as plt
import numpy as np
_logger = logging.getLogger(__name__)


class GaugeRNR(models.Model):
    _name = 'gauge.rnr'
    _description = 'Gauge R&R Setup'
    _inherit = "translation.mixin"

    # Fields for the Setup Form (First Image)
    date = fields.Date(string='Date', default=fields.Date.today)
    units = fields.Char(string='Units',translate=True)
    prepared_by = fields.Char(string='Prepared by',translate=True)
    gage_name = fields.Char(string='Gage Name',translate=True)
    characteristic = fields.Char(string='Characteristic',translate=True)
    appraiser_a_name = fields.Char(string='Appraiser A ',translate=True)

    gage_type = fields.Selection(
        [('attribute', 'Attribute'), ('variable', 'Variable')], string='Gauge Type')
    specification = fields.Char(string='Specification',translate=True)
    appraiser_b_name = fields.Char(string='Appraiser B ',translate=True)
    gage_number = fields.Char(string='Gage Number')
    test_number = fields.Char(string='Test Number')
    appraiser_c_name = fields.Char(string='Appraiser C ',translate=True)
    sample_size = fields.Selection(
        [(str(i), str(i)) for i in range(1, 11)],
        string='Sample Size (p)',
        default='10'
    )
    generate_xlsx_file = fields.Binary(string="Generate XLSX File", attachment=True)
    appraisers = fields.Integer(string='Appraisers (o)', default=3)
    trials = fields.Integer(string='Trials (n)', default=3)
    upper_spec_limit = fields.Float(string='Upper Specification Limit', digits=(12, 2))
    lower_spec_limit = fields.Float(string='Lower Specification Limit', digits=(12, 2))
    tolerance = fields.Float(string='Tolerance (Upper Spec Limit - Lower Spec Limit)', compute='_compute_tolerance', digits=(12, 2))

    # One2many field to link to appraiser data
    report_id = fields.Many2one('msa.sheet.generate', string="Calibration Record")
    appraiser_ids = fields.One2many('gauge.rnr.appraiser', 'gauge_id', string='Appraiser Data')
    appraiser_a_ids = fields.One2many(
        'gauge.rnr.appraiser', 'gauge_id',
        domain=[('appraiser', '=', 'Appraiser A')]
    ) 

    appraiser_b_ids = fields.One2many(
        'gauge.rnr.appraiser', 'gauge_id',
        domain=[('appraiser', '=', 'Appraiser B')]
    )

    appraiser_c_ids = fields.One2many(
        'gauge.rnr.appraiser', 'gauge_id',
        domain=[('appraiser', '=', 'Appraiser C')]
    )
    started = fields.Boolean(string="Started", default=False)
    approval_state = fields.Selection([
        ('draft', 'Draft'),
        ('to_approve', 'To Approve'),
        ('approved', 'Approved'),
        ('rejected', 'Rejected')
    ], string="Approval State", default='draft', tracking=True)

    appraiser_a_sample_stats_ids = fields.One2many('gauge.rnr.appraiser.sample', 'gauge_id', string='Appraiser A Sample Statistics', domain=[('appraiser', '=', 'Appraiser A')])
    
    appraiser_b_sample_stats_ids = fields.One2many('gauge.rnr.appraiser.sample', 'gauge_id', string='Appraiser B Sample Statistics', domain=[('appraiser', '=', 'Appraiser B')])
    
    appraiser_c_sample_stats_ids = fields.One2many('gauge.rnr.appraiser.sample', 'gauge_id', string='Appraiser C Sample Statistics',
                                                 domain=[('appraiser', '=', 'Appraiser C')])
    
    appraiser_a_trial_stats_ids = fields.One2many('gauge.rnr.appraiser.trial', 'gauge_id',
                                               string='Appraiser A Trial Statistics',
                                               domain=[('appraiser', '=', 'Appraiser A')])
    
    appraiser_b_trial_stats_ids = fields.One2many('gauge.rnr.appraiser.trial', 'gauge_id',
                                               string='Appraiser B Trial Statistics',
                                               domain=[('appraiser', '=', 'Appraiser B')])
    
    appraiser_c_trial_stats_ids = fields.One2many('gauge.rnr.appraiser.trial', 'gauge_id',
                                               string='Appraiser C Trial Statistics',
                                               domain=[('appraiser', '=', 'Appraiser C')])
    
    # Fields for GRR calculations 
    xbar_a = fields.Float(string='Xbar(A)', digits=(12, 2), readonly=True)
    xbar_b = fields.Float(string='Xbar(B)', digits=(12, 2), readonly=True)
    xbar_c = fields.Float(string='Xbar(C)', digits=(12, 2), readonly=True)
    rbar_a = fields.Float(string='Rbar(A)', digits=(12, 2), readonly=True)
    rbar_b = fields.Float(string='Rbar(B)', digits=(12, 2), readonly=True)
    rbar_c = fields.Float(string='Rbar(C)', digits=(12, 2), readonly=True)
    
    # Overall statistics
    xbar_doublebar = fields.Float(string='Xdoublebar', digits=(12, 2), readonly=True)
    rbar_doublebar = fields.Float(string='Rdoublebar', digits=(12, 2), readonly=True)
    rp = fields.Float(string='Rp', digits=(12, 2), readonly=True)
    ro = fields.Float(string='Ro', digits=(12, 2), readonly=True)
    
    # GRR components
    ev = fields.Float(string='EV', digits=(12, 2), readonly=True)
    av = fields.Float(string='AV', digits=(12, 2), readonly=True)
    grr = fields.Float(string='GRR', digits=(12, 2), readonly=True)
    pv = fields.Float(string='PV', digits=(12, 2), readonly=True)
    tv = fields.Float(string='TV', digits=(12, 2), readonly=True)
    
    # Component Variance Method Percentages
    ev_percent_tv = fields.Float(string='%EV', digits=(12, 2), readonly=True)
    av_percent_tv = fields.Float(string='%AV', digits=(12, 2), readonly=True)
    grr_percent_tv = fields.Float(string='%GRR', digits=(12, 2), readonly=True)
    pv_percent_tv = fields.Float(string='%PV', digits=(12, 2), readonly=True)
    
    # AIAG Method Percentages
    ev_percent_tolerance = fields.Float(string='%EV of Tolerance', digits=(12, 2), readonly=True)
    av_percent_tolerance = fields.Float(string='%AV of Tolerance', digits=(12, 2), readonly=True)
    grr_percent_tolerance = fields.Float(string='%GRR of Tolerance', digits=(12, 2), readonly=True)
    pv_percent_tolerance = fields.Float(string='%PV of Tolerance', digits=(12, 2), readonly=True)
    
    # Control limit calculations
    ucl_bias = fields.Float(string='UCL Bias', digits=(12, 4), readonly=True)
    lcl_bias = fields.Float(string='LCL Bias', digits=(12, 4), readonly=True)
    ucl_consistency = fields.Float(string='UCL Consistency', digits=(12, 4), readonly=True)
    lcl_consistency = fields.Float(string='LCL Consistency', digits=(12, 4), readonly=True)

    # charts

    bias_chart = fields.Binary("Appraisal Bias Chart", attachment=True)
    consistency_chart = fields.Binary("Appraisal Consistency Chart", attachment=True)
    range_chart = fields.Binary("Range Chart", attachment=True)
    average_chart = fields.Binary("Average Chart", attachment=True) 
    
    # Results interpretation fields
    utility_result = fields.Selection([
        ('good', 'Good (>= 80%)'),
        ('marginal', 'Marginal (50% - 80%)'),
        ('poor', 'Poor (< 50%)')
    ], string='Utility Result', readonly=True)
    
    bias_result = fields.Selection([
        ('no_bias', 'No appraiser bias present'),
        ('bias', 'Appraiser bias present')
    ], string='Bias Result', readonly=True)
    
    consistency_result = fields.Selection([
        ('consistent', 'Results are consistent across appraisers'),
        ('inconsistent', 'Measurement system shows inconsistency')
    ], string='Consistency Result', readonly=True)
    
    aiag_utility_result = fields.Selection([
        ('good', 'Good (< 10%)'),
        ('marginal', 'Marginal (10% - 30%)'),
        ('poor', 'Poor (> 30%)')
    ], string='AIAG Utility Result', readonly=True)
    
    resolution_result = fields.Selection([
        ('adequate', 'Measurement system has adequate resolution'),
        ('inadequate', 'Measurement system may not have adequate resolution')
    ], string='Resolution Result', readonly=True)

    def action_view_constants_table(self):
        """Open the statistical constants table view"""
        self.ensure_one()
        return {
            'name': 'Statistical Constants for Gauge R&R',
            'type': 'ir.actions.act_window',
            'res_model': 'gauge.rnr.constants',  # Changed to new model
            'view_mode': 'form',
            'view_id': self.env.ref('gauge_rnr.view_gauge_rnr_constants_table').id,
            'target': 'new',
            # No need for context with default_id since it's a separate model
        }
    
    def action_start(self):
        for record in self:
            record.started = True
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'current',
        }

    def action_submit_for_approval(self):
        for record in self:
            if record.approval_state == 'draft':
                record.approval_state = 'to_approve'
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'current',
        }

    def action_approve(self):
        for record in self:
            if record.approval_state == 'to_approve':
                record.approval_state = 'approved'
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'current',
        }

    def action_reject(self):
        for record in self:
            if record.approval_state == 'to_approve':
                record.approval_state = 'rejected'
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'current',
        }
    
    # Create a button to calculate statistics instead of automatic computation
    def calculate_statistics(self):
        """Calculate all statistics for the gauge R&R study"""
        self.ensure_one()
        
        # Clear existing statistics
        self.env['gauge.rnr.appraiser.sample'].search([('gauge_id', '=', self.id)]).unlink()
        self.env['gauge.rnr.appraiser.trial'].search([('gauge_id', '=', self.id)]).unlink()
        # Generate charts after statistics calculation
        self._generate_bias_chart()
        self._generate_consistency_chart()
        
        
        # Calculate for each appraiser
        appraisers = ['Appraiser A', 'Appraiser B', 'Appraiser C']
        appraiser_fields = ['appraiser_a_ids', 'appraiser_b_ids', 'appraiser_c_ids']
        
        # Reset statistics
        self.xbar_a = 0.0
        self.xbar_b = 0.0
        self.xbar_c = 0.0
        self.rbar_a = 0.0
        self.rbar_b = 0.0
        self.rbar_c = 0.0
        
        sample_count = int(self.sample_size)
        
        # Process data for each appraiser
        active_appraiser_count = 0
        xbar_values = []
        rbar_values = []
        
        for i, appraiser_name in enumerate(appraisers[:self.appraisers]):
            appraiser_records = getattr(self, appraiser_fields[i])
            if appraiser_records:
                active_appraiser_count += 1
                appraiser_stats = self._calculate_appraiser_stats(appraiser_name, appraiser_records)
                if appraiser_stats:
                    xbar = appraiser_stats['xbar']
                    rbar = appraiser_stats['rbar']
                    xbar_values.append(xbar)
                    rbar_values.append(rbar)
                    
                    # Set individual appraiser statistics
                    if i == 0:
                        self.xbar_a = xbar
                        self.rbar_a = rbar
                    elif i == 1:
                        self.xbar_b = xbar
                        self.rbar_b = rbar
                    elif i == 2:
                        self.xbar_c = xbar
                        self.rbar_c = rbar
                
        # Calculate overall statistics
        if xbar_values:
            self.xbar_doublebar = sum(xbar_values) / len(xbar_values)
            self.rbar_doublebar = sum(rbar_values) / len(rbar_values)
            self.rp = max(xbar_values) - min(xbar_values)
            self.ro = self.rp
            
            # Calculate GRR components
            
            # For d2 constants, we use standard values based on sample size
            # These are the standard d2 values for different sample sizes
            d2_values = {1: 1.41, 2: 1.28, 3: 1.69, 4: 2.06, 5: 2.33, 6: 2.53, 7: 2.70, 8: 2.85, 9: 2.97, 10: 3.08}
            
            # Get d2 values
            d2_ev = d2_values.get(int(self.trials), 1.41)  # Default to 1.41 if not found
            d2_av = d2_values.get(active_appraiser_count, 1.41)
            d2_pv = d2_values.get(int(self.sample_size), 1.41)
            
            # Calculate equipment variation (EV)
            self.ev = self.rbar_doublebar / d2_ev
            
            # Calculate appraiser variation (AV)
            n = int(self.trials)
            p = int(self.sample_size)
            # Note: In practice, there might be slight variations in this formula based on specific GRR methodologies
            av_squared = ((self.rp / d2_av)**2 - (self.ev**2 / (n * p)))
            self.av = math.sqrt(max(0, av_squared))  # Ensure we don't take square root of negative number
            
            # Calculate GRR
            self.grr = math.sqrt(self.ev**2 + self.av**2)
            
            # Calculate part variation (PV)
            self.pv = self.rp / d2_pv
            
            # Calculate total variation (TV)
            self.tv = math.sqrt(self.grr**2 + self.pv**2)
            
            # Calculate percentages for Component Variance Method
            if self.tv > 0:
                self.ev_percent_tv = (self.ev / self.tv) * 100
                self.av_percent_tv = (self.av / self.tv) * 100
                self.grr_percent_tv = (self.grr / self.tv) * 100
                self.pv_percent_tv = (self.pv / self.tv) * 100
            
            # Calculate percentages for AIAG Method
            if self.tolerance > 0:
                self.ev_percent_tolerance = (self.ev / self.tolerance) * 100
                self.av_percent_tolerance = (self.av / self.tolerance) * 100
                self.grr_percent_tolerance = (self.grr / self.tolerance) * 100
                self.pv_percent_tolerance = (self.pv / self.tolerance) * 100
            
            # Calculate control limits for charts
            self.ucl_bias = self.xbar_doublebar + (0.75 * self.rbar_doublebar)
            self.lcl_bias = self.xbar_doublebar - (0.75 * self.rbar_doublebar)
            self.ucl_consistency = self.rbar_doublebar * 1.315  # Standard factor for UCL
            self.lcl_consistency = self.rbar_doublebar * 0.669  # Standard factor for LCL
            
            # Set interpretation results
            # Utility based on %PV
            if self.pv_percent_tv >= 80:
                self.utility_result = 'good'
            elif 50 <= self.pv_percent_tv < 80:
                self.utility_result = 'marginal'
            else:
                self.utility_result = 'poor'
            
            # Bias result (simplified for this implementation)
            all_within_limits = all(self.lcl_bias <= x <= self.ucl_bias for x in xbar_values)
            self.bias_result = 'no_bias' if all_within_limits else 'bias'
            
            # Consistency result (simplified)
            all_consistent = all(r <= self.ucl_consistency for r in rbar_values)
            self.consistency_result = 'consistent' if all_consistent else 'inconsistent'
            
            # AIAG utility result
            if self.grr_percent_tolerance < 10:
                self.aiag_utility_result = 'good'
            elif 10 <= self.grr_percent_tolerance <= 30:
                self.aiag_utility_result = 'marginal'
            else:
                self.aiag_utility_result = 'poor'
            
            # Resolution result (simplified)
            # This is typically based on how many distinct categories can be reliably measured
            # For simplicity, we use a threshold on GRR%
            self.resolution_result = 'adequate' if self.grr_percent_tv < 30 else 'inadequate'
    
    def _calculate_appraiser_stats(self, appraiser_name, appraiser_records):
        """Calculate statistics for a specific appraiser"""
        if not appraiser_records:
            return None
            
        # Get number of samples and trials
        sample_count = int(self.sample_size)
        
        # Initialize values for calculations
        sample_averages = []
        sample_ranges = []
        trial_averages = []
        trial_ranges = []
        
        # Calculate sample-wise statistics (each sample across all trials)
        for sample_num in range(1, sample_count + 1):
            sample_field = f'sample_{sample_num}'
            values = []
            
            for trial_record in appraiser_records:
                sample_value = getattr(trial_record, sample_field, 0.0)
                if sample_value != 0.0:  # Skip empty values
                    values.append(sample_value)
            
            # Calculate average and range
            if values:
                avg = sum(values) / len(values)
                sample_range = max(values) - min(values)
                sample_averages.append(avg)
                sample_ranges.append(sample_range)
                
                # Create sample statistics record
                self.env['gauge.rnr.appraiser.sample'].create({
                    'gauge_id': self.id,
                    'appraiser': appraiser_name,
                    'sample_number': sample_num,
                    'average': avg,
                    'range': sample_range,
                })
        
        # Calculate trial-wise statistics (each trial across all samples)
        trial_numbers = appraiser_records.mapped('trial')
        for trial_num in trial_numbers:
            trial_record = appraiser_records.filtered(lambda r: r.trial == trial_num)
            if not trial_record:
                continue
                
            # Single record expected per trial
            trial_record = trial_record[0]
            
            # Get all sample values for this trial
            samples = []
            for i in range(1, sample_count + 1):
                sample_value = getattr(trial_record, f'sample_{i}', 0.0)
                if sample_value != 0.0:
                    samples.append(sample_value)
            
            if samples:
                # Calculate average and range for this trial
                trial_avg = sum(samples) / len(samples)
                trial_range = max(samples) - min(samples)
                trial_averages.append(trial_avg)
                trial_ranges.append(trial_range)
                
                # Create trial statistics record
                self.env['gauge.rnr.appraiser.trial'].create({
                    'gauge_id': self.id,
                    'appraiser': appraiser_name,
                    'trial_number': trial_num,
                    'average': trial_avg,
                    'range': trial_range,
                })
                
        # Calculate overall statistics for this appraiser
        xbar = sum(sample_averages) / len(sample_averages) if sample_averages else 0.0
        rbar = sum(sample_ranges) / len(sample_ranges) if sample_ranges else 0.0
        
        return {
            'xbar': xbar,
            'rbar': rbar,
            'sample_averages': sample_averages,
            'sample_ranges': sample_ranges,
            'trial_averages': trial_averages,
            'trial_ranges': trial_ranges
        }

    @api.depends('upper_spec_limit', 'lower_spec_limit')

    def _compute_tolerance(self):
        for record in self:
            record.tolerance = record.upper_spec_limit - record.lower_spec_limit

    @api.model
    def create(self, vals):
        record = super(GaugeRNR, self).create(vals)
        appraisers = ['A', 'B', 'C']
        for appraiser in appraisers[:record.appraisers]:
            for trial in range(1, record.trials + 1):
                appraiser_record = self.env['gauge.rnr.appraiser'].create({
                    'gauge_id': record.id,
                    'appraiser': f'Appraiser {appraiser}',
                    'trial': trial,
                })
                _logger.info(
                    f"Created appraiser record: {appraiser_record.appraiser} with trial {trial} for gauge_id {record.id}")
        return record
        
    # Add this write method here
    def write(self, vals):
        result = super(GaugeRNR, self).write(vals)
        _logger.info(f"Writing values to gauge record: {vals}")
        return result
    
    
        
    def generate_xlsx_report(self):
        output = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws1 = wb.active
        ws.title = 'MSA Report'

        border = Border(top=Side(style='thin'),left=Side(style='thin'),right=Side(style='thin'),bottom=Side(style='thin'))
        align_center = Alignment(vertical='center', horizontal='center', wrapText=True)
        align_left = Alignment(vertical='center', horizontal='left')
        font_header = Font(name='Arial', size=12, bold=True)
        font_all = Font(name='Times New Roman', size=11, bold=False)

        data = {
            'A1': 'MEASUREMENT SYSTEM ANALYSIS (Repeatability and Reproducibility)',
            'A2': 'Date',
            'A3': 'Gauge Name',
            'A4': 'Gauge Type',
            'A5': 'Gauge Number',
            'C2': 'Units',
            'C3': 'Characteristic',
            'C4': 'Specification',
            'C5': 'Test Number',
            'E2': 'Pepared By',
            'E3': 'Appraiser A',
            'E4': 'Appraiser B',
            'E5': 'Appraiser C',
            'A6': 'Specification Limits',
            'A7': 'Sample Size (p)',
            'C7': 'Appraisers (o)',
            'E7': 'Trials (n)',
            'A8': 'Upper Specification Limit',
            'C8': 'Lower Specification Limit',
            'E8': 'Tolerance (Upper Spec Limit - Lower Spec Limit)',

        }

        for cell, value in data.items():
            ws[cell] = value
            ws[cell].font = font_header
            ws[cell].alignment = align_center
            ws[cell].border = border

        ws.sheet_view.showGridLines = False

        for row in ws.iter_rows(min_row=1, max_row=8, min_col=1, max_col=6):
            for cell in row:
                cell.border = border
                cell.alignment = align_center
        
        merge_ranges = [
            'A1:F1', 'A6:F6' 
        ]

        for merge_range in merge_ranges:
            ws.merge_cells(merge_range)

        ws.row_dimensions[1].height = 25
        ws.row_dimensions[5].height = 25
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 25

        for i in self:
            ws['B2'] = i.date if i.date else ''
            ws['B3'] = i.gage_name if i.gage_name else ''
            ws['B4'] = i.gage_type if i.gage_type else ''
            ws['B5'] = i.gage_number if i.gage_number else ''
            ws['D2'] = i.units if i.units else ''
            ws['D3'] = i.characteristic if i.characteristic else ''
            ws['D4'] = i.specification if i.specification else ''
            ws['D5'] = i.test_number if i.test_number else ''
            ws['F2'] = i.prepared_by if i.prepared_by else ''
            ws['F3'] = i.appraiser_a_name if i.appraiser_a_name else ''
            ws['F4'] = i.appraiser_b_name if i.appraiser_b_name else ''
            ws['F5'] = i.appraiser_c_name if i.appraiser_c_name else ''
            ws['B7'] = i.sample_size if i.sample_size else ''
            ws['D7'] = i.appraisers if i.appraisers else ''
            ws['F7'] = i.trials if i.trials else ''
            ws['B8'] = i.upper_spec_limit if i.upper_spec_limit else ''
            ws['D8'] = i.lower_spec_limit if i.lower_spec_limit else ''
            ws['F8'] = i.tolerance if i.tolerance else ''

        ws['F11'].value = 'Go to Appraisal A Data'
        ws['F11'].hyperlink = "#'Appraisal Data'!F11"  # Local sheet link
        ws['F11'].font = Font(color="0000FF", underline="single") 
        
        #Appraisal Data

        wb.create_sheet(title='Appraisal Data')
        ws1 = wb['Appraisal Data']


        border = Border(top=Side(style='thin'),left=Side(style='thin'),right=Side(style='thin'),bottom=Side(style='thin'))
        align_center = Alignment(vertical='center', horizontal='center', wrapText=True)
        align_left = Alignment(vertical='center', horizontal='left')
        font_header = Font(name='Arial', size=12, bold=True)
        font_all = Font(name='Times New Roman', size=11, bold=False)

        data = {
            'A1': 'Appraisal A Data',
            'A2': 'Trail/Sample',
            'B2': 'Sample 1',
            'C2': 'Sample 2',
            'D2': 'Sample 3',
            'E2': 'Sample 4',
            'F2': 'Sample 5',
            'G2': 'Sample 6',
            'H2': 'Sample 7',
            'I2': 'Sample 8',
            'J2': 'Sample 9',
            'K2': 'Sample 10',
            'L2': 'Average',

        }

        for cell, value in data.items():
            ws1[cell] = value
            ws1[cell].font = font_header
            ws1[cell].alignment = align_center
            ws1[cell].border = border
        
        ws1.sheet_view.showGridLines = False

        for row in ws1.iter_rows(min_row=1, max_row=7, min_col=1, max_col=12):
            for cell in row:
                cell.border = border
                cell.alignment = align_center
        
        merge_ranges = [
            'A1:L1'
        ]

        for merge_range in merge_ranges:
            ws1.merge_cells(merge_range)
        
        ws1.row_dimensions[1].height = 25

        column_widths = { 
            'A': 15, 'B': 15, 'C': 15, 'D': 15, 'E': 15, 'F': 15, 'G': 15, 'H': 15, 'I': 15, 'J': 15, 'K': 15, 'L': 15}
        
        for col, width in column_widths.items():
            ws1.column_dimensions[col].width = width
        
        current_row = 3
        for i in self:
            appraiser_rows = len(i.appraiser_a_ids)
            for rec in i.appraiser_a_ids:
                ws1[f'A{current_row}'] = rec.trial
                ws1[f'B{current_row}'] = rec.sample_1
                ws1[f'C{current_row}'] = rec.sample_2
                ws1[f'D{current_row}'] = rec.sample_3
                ws1[f'E{current_row}'] = rec.sample_4
                ws1[f'F{current_row}'] = rec.sample_5
                ws1[f'G{current_row}'] = rec.sample_6
                ws1[f'H{current_row}'] = rec.sample_7
                ws1[f'I{current_row}'] = rec.sample_8
                ws1[f'J{current_row}'] = rec.sample_9
                ws1[f'K{current_row}'] = rec.sample_10
                current_row += 1
            stat_row = current_row - appraiser_rows  # Go back to first row of the trials
            for rec in i.appraiser_a_trial_stats_ids:
                ws1[f'L{stat_row}'] = rec.average
                stat_row += 1
            
            
       
        ws1[f'A{current_row}'] = 'Average'
        ws1[f'A{current_row + 1}'] = 'Range'
        ws1[f'L{current_row}'] = 'Xbar(A)'
        ws1[f'L{current_row+1}'] = 'Rbar(A)'   
        row = current_row
        for i in self:
            col_index = 2  # Start from column B (i.e., after 'A' label column)
            for stat in i.appraiser_a_sample_stats_ids:
                avg_cell = ws1.cell(row=row, column=col_index)
                rng_cell = ws1.cell(row=row + 1, column=col_index)

                avg_cell.value = stat.average
                rng_cell.value = stat.range

                avg_cell.border = border
                rng_cell.border = border

                col_index += 1
        for i in self:
            ws1[f'M{current_row}'] = i.xbar_a if i.xbar_a else ''
            ws1[f'M{current_row}'].border = border
            ws1[f'M{current_row}'].alignment = align_center
            ws1[f'M{current_row+1}'] = i.rbar_a if i.rbar_a else ''
            ws1[f'M{current_row+1}'].border = border
            ws1[f'M{current_row+1}'].alignment = align_center

        
        current_row += 2

        # Set "Appraisal B Data" title
        ws1.merge_cells(f'A{current_row}:L{current_row}')
        ws1[f'A{current_row}'] = 'Appraisal B Data'
        ws1[f'A{current_row}'].border = border
        ws1[f'A{current_row}'].alignment = align_center
        ws1[f'A{current_row}'].font = font_header

    
        # Define header labels
        headers = ['Trial/Sample', 'Sample 1', 'Sample 2', 'Sample 3', 'Sample 4',
                'Sample 5', 'Sample 6', 'Sample 7', 'Sample 8', 'Sample 9', 'Sample 10', 'Average']

        # Apply headers starting from column A
        header_row = current_row + 1
        for col_index, header in enumerate(headers, start=1):
            col_letter = chr(64 + col_index)
            cell = ws1[f'{col_letter}{header_row}']
            cell.value = header
            cell.border = border
            cell.alignment = align_center
            cell.font = font_header
        
        current_row += 2
        for i in self:
            appraiser_rows = len(i.appraiser_b_ids)
            for rec in i.appraiser_b_ids:
                values = [
                    rec.trial, rec.sample_1, rec.sample_2, rec.sample_3, rec.sample_4,
                    rec.sample_5, rec.sample_6, rec.sample_7, rec.sample_8, rec.sample_9,
                    rec.sample_10
                ]
                for col_index, value in enumerate(values, start=1):
                    col_letter = chr(64 + col_index)  # 1 -> A, 2 -> B, etc.
                    cell = ws1[f'{col_letter}{current_row}']
                    cell.value = value
                    cell.border = border  # Apply border
                current_row += 1
            stat_row = current_row - appraiser_rows  # Go back to first row of the trials
            for rec in i.appraiser_b_trial_stats_ids:
                ws1[f'L{stat_row}'] = rec.average
                ws1[f'L{stat_row}'].border = border 
                stat_row += 1
            
            
        ws1[f'A{current_row}'] = 'Average'
        ws1[f'A{current_row}'].border = border
        ws1[f'A{current_row}'].alignment = align_center
        ws1[f'A{current_row}'].font = font_header
        ws1[f'A{current_row+1}'] = 'Range'
        ws1[f'A{current_row+1}'].border = border
        ws1[f'A{current_row+1}'].alignment = align_center
        ws1[f'A{current_row+1}'].font = font_header 
        ws1[f'L{current_row}'] = 'Xbar(B)'
        ws1[f'L{current_row}'].border = border
        ws1[f'L{current_row}'].alignment = align_center
        ws1[f'L{current_row}'].font = font_header
        ws1[f'L{current_row+1}'] = 'Rbar(B)'
        ws1[f'L{current_row+1}'].border = border
        ws1[f'L{current_row+1}'].alignment = align_center
        ws1[f'L{current_row+1}'].font = font_header

        row = current_row
        for i in self:
            col_index = 2  # Start from column B (i.e., after 'A' label column)
            for stat in i.appraiser_b_sample_stats_ids:
                avg_cell = ws1.cell(row=row, column=col_index)
                rng_cell = ws1.cell(row=row + 1, column=col_index)

                avg_cell.value = stat.average
                rng_cell.value = stat.range

                avg_cell.border = border
                rng_cell.border = border

                col_index += 1
        for i in self:
            ws1[f'M{current_row}'] = i.xbar_b if i.xbar_b else ''
            ws1[f'M{current_row}'].border = border
            ws1[f'M{current_row}'].alignment = align_center
            ws1[f'M{current_row+1}'] = i.rbar_b if i.rbar_b else ''
            ws1[f'M{current_row+1}'].border = border
            ws1[f'M{current_row+1}'].alignment = align_center


        current_row += 2

        # Set "Appraisal C Data" title
        ws1.merge_cells(f'A{current_row}:L{current_row}')
        ws1[f'A{current_row}'] = 'Appraisal C Data'
        ws1[f'A{current_row}'].border = border
        ws1[f'A{current_row}'].alignment = align_center
        ws1[f'A{current_row}'].font = font_header

    
        # Define header labels
        headers = ['Trial/Sample', 'Sample 1', 'Sample 2', 'Sample 3', 'Sample 4',
                'Sample 5', 'Sample 6', 'Sample 7', 'Sample 8', 'Sample 9', 'Sample 10', 'Average']

        # Apply headers starting from column A
        header_row = current_row + 1
        for col_index, header in enumerate(headers, start=1):
            col_letter = chr(64 + col_index)
            cell = ws1[f'{col_letter}{header_row}']
            cell.value = header
            cell.border = border
            cell.alignment = align_center
            cell.font = font_header
        
        current_row += 2
        for i in self:
            appraiser_rows = len(i.appraiser_c_ids)
            for rec in i.appraiser_c_ids:
                values = [
                    rec.trial, rec.sample_1, rec.sample_2, rec.sample_3, rec.sample_4,
                    rec.sample_5, rec.sample_6, rec.sample_7, rec.sample_8, rec.sample_9,
                    rec.sample_10
                ]
                for col_index, value in enumerate(values, start=1):
                    col_letter = chr(64 + col_index)  # 1 -> A, 2 -> B, etc.
                    cell = ws1[f'{col_letter}{current_row}']
                    cell.value = value
                    cell.border = border  # Apply border
                current_row += 1
            stat_row = current_row - appraiser_rows  # Go back to first row of the trials
            for rec in i.appraiser_c_trial_stats_ids:
                ws1[f'L{stat_row}'] = rec.average
                ws1[f'L{stat_row}'].border = border
                stat_row += 1
            
        ws1[f'A{current_row}'] = 'Average'
        ws1[f'A{current_row}'].border = border
        ws1[f'A{current_row}'].alignment = align_center
        ws1[f'A{current_row}'].font = font_header
        ws1[f'A{current_row+1}'] = 'Range'
        ws1[f'A{current_row+1}'].border = border
        ws1[f'A{current_row+1}'].alignment = align_center
        ws1[f'A{current_row+1}'].font = font_header
        ws1[f'L{current_row}'] = 'Xbar(C)'
        ws1[f'L{current_row}'].border = border
        ws1[f'L{current_row}'].alignment = align_center
        ws1[f'L{current_row}'].font = font_header        
        ws1[f'L{current_row+1}'] = 'Rbar(C)'
        ws1[f'L{current_row+1}'].border = border
        ws1[f'L{current_row+1}'].alignment = align_center
        ws1[f'L{current_row+1}'].font = font_header
        row = current_row
        for i in self:
            col_index = 2  # Start from column B (i.e., after 'A' label column)
            for stat in i.appraiser_c_sample_stats_ids:
                avg_cell = ws1.cell(row=row, column=col_index)
                rng_cell = ws1.cell(row=row + 1, column=col_index)

                avg_cell.value = stat.average
                rng_cell.value = stat.range

                avg_cell.border = border
                rng_cell.border = border

                col_index += 1
        for i in self:
            ws1[f'M{current_row}'] = i.xbar_c if i.xbar_c else ''
            ws1[f'M{current_row}'].border = border
            ws1[f'M{current_row}'].alignment = align_center
            ws1[f'M{current_row+1}'] = i.rbar_c if i.rbar_c else ''
            ws1[f'M{current_row+1}'].border = border
            ws1[f'M{current_row+1}'].alignment = align_center
        
        current_row += 4
        ws1[f'A{current_row}'].value = 'Go to MSA Report'
        ws1[f'A{current_row}'].hyperlink = "#'MSA Report'!A1"  # Local sheet link
        ws1[f'A{current_row}'].font = Font(color="0000FF", underline="single") 
        
        ws1.merge_cells(f'G{current_row}:H{current_row}')
        ws1[f'G{current_row}'] = 'Xbar Doublebar'
        ws1[f'G{current_row}'].border = border
        ws1[f'G{current_row}'].alignment = align_center
        ws1[f'G{current_row}'].font = font_header
        ws1.merge_cells(f'G{current_row+1}:H{current_row+1}')
        ws1[f'G{current_row+1}'] = 'Rbar Doublebar'
        ws1[f'G{current_row+1}'].border = border
        ws1[f'G{current_row+1}'].alignment = align_center
        ws1[f'G{current_row+1}'].font = font_header
        ws1[f'K{current_row}'] = 'Rp'
        ws1[f'K{current_row}'].border = border
        ws1[f'K{current_row}'].alignment = align_center
        ws1[f'K{current_row}'].font = font_header
        ws1[f'K{current_row+1}'] = 'Ro'
        ws1[f'K{current_row+1}'].border = border
        ws1[f'K{current_row+1}'].alignment = align_center
        ws1[f'K{current_row+1}'].font = font_header

        for i in self:
            ws1[f'I{current_row}'] = i.xbar_doublebar if i.xbar_doublebar else ''
            ws1[f'I{current_row}'].border = border
            ws1[f'I{current_row}'].alignment = align_center
            ws1[f'I{current_row+1}'] = i.rbar_doublebar if i.rbar_doublebar else ''
            ws1[f'I{current_row+1}'].border = border
            ws1[f'I{current_row+1}'].alignment = align_center
            ws1[f'L{current_row}'] = i.rp if i.rp else ''
            ws1[f'L{current_row}'].border = border
            ws1[f'L{current_row}'].alignment = align_center
            ws1[f'L{current_row+1}'] = i.ro if i.ro else ''
            ws1[f'L{current_row+1}'].border = border
            ws1[f'L{current_row+1}'].alignment = align_center

        #GRR RESULTS

        wb.create_sheet(title='GRR RESULTS')
        ws2 = wb['GRR RESULTS']
        ws2.sheet_view.showGridLines = False

        border = Border(top=Side(style='thin'),left=Side(style='thin'),right=Side(style='thin'),bottom=Side(style='thin'))
        align_center = Alignment(vertical='center', horizontal='center', wrapText=True)
        align_left = Alignment(vertical='center', horizontal='left')
        font_header = Font(name='Arial', size=12, bold=True)
        font_all = Font(name='Times New Roman', size=11, bold=False)

        data = {
            'A1': 'GRR RESULTS',
            
        }



        
        


        
        
        
        
        








        wb.save(output)
        output.seek(0)

        attachment = self.env["ir.attachment"].create({
            "name": "MSA Report.xlsx",
            "type": "binary",
            "datas": base64.b64encode(output.getvalue()),
            "res_model": self._name,
            "res_id": self.id,
            "mimetype": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        })










        return {"type": "ir.actions.act_url", "url": f"/web/content/{attachment.id}?download=true", "target": "self"}

        
    
    def _generate_bias_chart(self):
        """Generate and save the Appraiser Bias Chart"""
        self.ensure_one()
        
        # Get appraiser names and their averages
        active_appraisers = []
        xbar_values = []
        
        # Use appraiser names for the chart
        if self.appraisers >= 1 and self.xbar_a > 0:
            active_appraisers.append(self.appraiser_a_name or 'A')
            xbar_values.append(self.xbar_a)
        if self.appraisers >= 2 and self.xbar_b > 0:
            active_appraisers.append(self.appraiser_b_name or 'B')
            xbar_values.append(self.xbar_b)
        if self.appraisers >= 3 and self.xbar_c > 0:
            active_appraisers.append(self.appraiser_c_name or 'C')
            xbar_values.append(self.xbar_c)
        
        # If we have data to plot
        if active_appraisers and xbar_values:
            plt.figure(figsize=(10, 6))
            
            # Plot the appraiser averages
            plt.plot(range(len(active_appraisers)), xbar_values, 'bo-', color='red', label='Appraiser Averages')
            
            # Plot the control limits
            plt.axhline(y=self.ucl_bias, color='blue', linestyle='-', label='UCL')
            plt.axhline(y=self.lcl_bias, color='orange', linestyle='-', label='LCL')
            
            # Plot the centerline (xbar_doublebar)
            plt.axhline(y=self.xbar_doublebar, color='green', linestyle='--', label='Centerline')
            
            # Set labels and title
            plt.xlabel('Appraiser')
            plt.ylabel('Average')
            plt.title('Appraiser Bias')
            
            # Set x-axis ticks to appraiser names
            plt.xticks(range(len(active_appraisers)), active_appraisers)
            
            # Add legend
            plt.legend()
            
            # Add grid for better readability
            plt.grid(True, linestyle='--', alpha=0.7)
            
            # Set y-axis limits with some padding
            y_range = max(self.ucl_bias - self.lcl_bias, 1.0)  # Ensure at least some range
            plt.ylim([self.lcl_bias - 0.1 * y_range, self.ucl_bias + 0.1 * y_range])
            
            # Save the chart to binary field
            buffer = io.BytesIO()
            plt.savefig(buffer, format='png')
            buffer.seek(0)
            self.bias_chart = base64.b64encode(buffer.getvalue())
            plt.close()
    
    def _generate_consistency_chart(self):
        """Generate and save the Appraiser Consistency Chart"""
        self.ensure_one()
        
        # Get appraiser names and their ranges
        active_appraisers = []
        rbar_values = []
        
        # Use appraiser names for the chart
        if self.appraisers >= 1 and self.rbar_a > 0:
            active_appraisers.append(self.appraiser_a_name or 'A')
            rbar_values.append(self.rbar_a)
        if self.appraisers >= 2 and self.rbar_b > 0:
            active_appraisers.append(self.appraiser_b_name or 'B')
            rbar_values.append(self.rbar_b)
        if self.appraisers >= 3 and self.rbar_c > 0:
            active_appraisers.append(self.appraiser_c_name or 'C')
            rbar_values.append(self.rbar_c)
        
        # If we have data to plot
        if active_appraisers and rbar_values:
            plt.figure(figsize=(10, 6))
            
            # Plot the appraiser ranges
            plt.plot(range(len(active_appraisers)), rbar_values, 'bo-', color='red', label='Appraiser Ranges')
            
            # Plot the UCL for consistency
            plt.axhline(y=self.ucl_consistency, color='blue', linestyle='-', label='UCL')
            
            # Plot the LCL for consistency
            plt.axhline(y=self.lcl_consistency, color='orange', linestyle='-', label='LCL')
            
            # Plot the centerline (rbar_doublebar)
            plt.axhline(y=self.rbar_doublebar, color='green', linestyle='--', label='Centerline')
            
            # Set labels and title
            plt.xlabel('Appraiser')
            plt.ylabel('Range')


            
            plt.title('Appraiser Consistency')
            
            # Set x-axis ticks to appraiser names
            plt.xticks(range(len(active_appraisers)), active_appraisers)
            
            # Add legend
            plt.legend()
            
            # Add grid for better readability
            plt.grid(True, linestyle='--', alpha=0.7)
            
            # Set y-axis limits with some padding
            y_max = max(self.ucl_consistency, max(rbar_values)) * 1.1  # Add 10% padding
            plt.ylim([0, y_max])
            
            # Save the chart to binary field
            buffer = io.BytesIO()
            plt.savefig(buffer, format='png')
            buffer.seek(0)
            self.consistency_chart = base64.b64encode(buffer.getvalue())
            plt.close()

    def generate_control_charts(self):
        """Generate Range and Average control charts for all appraisers and samples"""
        self.ensure_one()
        
        # Generate the charts
        self._generate_range_chart()
        self._generate_average_chart()
        
        return {
            'type': 'ir.actions.client',
            'tag': 'reload',
        }
    
    def _generate_range_chart(self):
        """Generate and save the Range Chart"""
        self.ensure_one()
        
        # Get sample count
        sample_count = int(self.sample_size)
        
        # Prepare data structures
        appraiser_data = {}
        
        # Collect data for each appraiser
        appraisers = ['Appraiser A', 'Appraiser B', 'Appraiser C']
        appraiser_fields = ['appraiser_a_ids', 'appraiser_b_ids', 'appraiser_c_ids']
        appraiser_colors = ['blue', 'red', 'green']
        
        # Collect sample range data for each appraiser
        for i, appraiser_name in enumerate(appraisers[:self.appraisers]):
            sample_stats = self.env['gauge.rnr.appraiser.sample'].search([
                ('gauge_id', '=', self.id),
                ('appraiser', '=', appraiser_name)
            ], order='sample_number')
            
            # Extract range values
            range_values = sample_stats.mapped('range')
            
            # If we have range values, store them
            if range_values:
                appraiser_data[appraiser_name] = {
                    'ranges': range_values,
                    'color': appraiser_colors[i],
                    'display_name': self[f'appraiser_{chr(97+i)}_name'] or appraiser_name[-1]  # Use the name or just "A", "B", "C"
                }
        
        # Calculate control limits for range chart
        r_doublebar = self.rbar_doublebar
        
        # For UCL of range, the factor depends on the number of trials
        # These are standard D4 values based on the number of trials
        d4_values = {1: 3.267, 2: 3.267, 3: 2.575, 4: 2.282, 5: 2.115, 6: 2.004, 7: 1.924, 8: 1.864, 9: 1.816, 10: 1.777}
        d4 = d4_values.get(self.trials, 3.267)
        
        ucl_r = r_doublebar * d4
        
        # For trials < 7, LCL_R is typically set to 0
        lcl_r = 0 if self.trials < 7 else r_doublebar * (1 - 3 * 0.3 / d4)
        
        # Create the plot
        plt.figure(figsize=(12, 8))
        
        # X coordinates for plotting - we'll offset each appraiser's data points slightly
        total_appraisers = len(appraiser_data)
        x_positions = {}
        
        # Spacing for x-axis to accommodate all samples and appraisers
        sample_spacing = 3
        appraiser_width = 0.8 / total_appraisers if total_appraisers > 0 else 0.8
        
        # Plot data for each appraiser
        legend_handles = []
        
        max_value = 0  # To track max y value for scaling
        x_labels = []
        
        # First plot the sample points for each appraiser
        for idx, (appraiser, data) in enumerate(appraiser_data.items()):
            ranges = data['ranges']
            color = data['color']
            display_name = data['display_name']
            
            # Calculate x positions for this appraiser
            x_pos = np.arange(1, sample_count + 1) * sample_spacing
            x_positions[appraiser] = x_pos
            
            # Plot range values with connecting lines
            line, = plt.plot(x_pos, ranges, 'o-', color=color, label=display_name, linewidth=1.5, markersize=6)
            legend_handles.append(line)
            
            # Track max value for y-axis scaling
            max_value = max(max_value, max(ranges) if ranges else 0)
        
        # Plot control limits
        plt.axhline(y=ucl_r, color='blue', linestyle='-', label='UCL')
        plt.axhline(y=r_doublebar, color='purple', linestyle='--', label='Mean')
        if lcl_r > 0:
            plt.axhline(y=lcl_r, color='blue', linestyle='-', label='LCL')
        
        # Set axis labels and title
        plt.xlabel('Sample Number')
        plt.ylabel('Range')
        plt.title('Range Chart')
        
        # Set x-axis ticks at sample positions
        sample_positions = np.arange(1, sample_count + 1) * sample_spacing
        plt.xticks(sample_positions, [str(i) for i in range(1, sample_count + 1)])
        
        # Add more detailed x-axis labels showing appraiser groups
        appraiser_label_pos = []
        appraiser_label_names = []
        
        for i, appraiser_name in enumerate(appraisers[:self.appraisers]):
            if appraiser_name in appraiser_data:
                mid_point = (sample_positions[0] + sample_positions[-1]) / 2 + (i - self.appraisers/2 + 0.5) * sample_spacing * 1.2
                appraiser_label_pos.append(mid_point)
                display_name = appraiser_data[appraiser_name]['display_name']
                appraiser_label_names.append(display_name)
        
        # Add a second x-axis for appraiser labels
        ax2 = plt.twiny()
        ax2.set_xticks(appraiser_label_pos)
        ax2.set_xticklabels(appraiser_label_names)
        ax2.set_xlim(plt.gca().get_xlim())
        
        # Set y-axis limits with some padding
        plt.ylim(0, max(max_value * 1.1, ucl_r * 1.1))
        
        # Add grid for better readability
        plt.grid(True, linestyle='--', alpha=0.7)
        
        # Add legend
        plt.legend(handles=legend_handles + [
            plt.Line2D([0], [0], color='blue', lw=2, linestyle='-', label='UCL'),
            plt.Line2D([0], [0], color='purple', lw=2, linestyle='--', label='Mean')
        ])
        
        # Add Plot Area text in the bottom right
        plt.figtext(0.93, 0.05, 'Plot Area', ha='right')
        
        # Save the chart to binary field
        buffer = io.BytesIO()
        plt.savefig(buffer, format='png', dpi=100)
        buffer.seek(0)
        self.range_chart = base64.b64encode(buffer.getvalue())
        plt.close()

    def _generate_average_chart(self):
        """Generate and save the Average Chart"""
        self.ensure_one()
        
        # Get sample count
        sample_count = int(self.sample_size)
        
        # Prepare data structures
        appraiser_data = {}
        
        # Collect data for each appraiser
        appraisers = ['Appraiser A', 'Appraiser B', 'Appraiser C']
        appraiser_fields = ['appraiser_a_ids', 'appraiser_b_ids', 'appraiser_c_ids']
        appraiser_colors = ['blue', 'red', 'green']
        
        # Collect sample average data for each appraiser
        for i, appraiser_name in enumerate(appraisers[:self.appraisers]):
            sample_stats = self.env['gauge.rnr.appraiser.sample'].search([
                ('gauge_id', '=', self.id),
                ('appraiser', '=', appraiser_name)
            ], order='sample_number')
            
            # Extract average values
            avg_values = sample_stats.mapped('average')
            
            # If we have average values, store them
            if avg_values:
                appraiser_data[appraiser_name] = {
                    'averages': avg_values,
                    'color': appraiser_colors[i],
                    'display_name': self[f'appraiser_{chr(97+i)}_name'] or appraiser_name[-1]  # Use the name or just "A", "B", "C"
                }
        
        # Calculate control limits for averages chart
        x_doublebar = self.xbar_doublebar
        r_doublebar = self.rbar_doublebar
        
        # A2 factor depends on the number of trials
        a2_values = {1: 2.660, 2: 1.880, 3: 1.023, 4: 0.729, 5: 0.577, 6: 0.483, 7: 0.419, 8: 0.373, 9: 0.337, 10: 0.308}
        a2 = a2_values.get(self.trials, 1.880)
        
        ucl_x = x_doublebar + (a2 * r_doublebar)
        lcl_x = x_doublebar - (a2 * r_doublebar)
        
        # Create the plot
        plt.figure(figsize=(12, 8))
        
        # X coordinates for plotting - we'll offset each appraiser's data points slightly
        total_appraisers = len(appraiser_data)
        x_positions = {}
        
        # Spacing for x-axis to accommodate all samples and appraisers
        sample_spacing = 3
        appraiser_width = 0.8 / total_appraisers if total_appraisers > 0 else 0.8
        
        # Plot data for each appraiser
        legend_handles = []
        
        max_value = 0  # To track max y value for scaling
        min_value = float('inf')  # To track min y value for scaling
        
        # First plot the sample points for each appraiser
        for idx, (appraiser, data) in enumerate(appraiser_data.items()):
            averages = data['averages']
            color = data['color']
            display_name = data['display_name']
            
            # Calculate x positions for this appraiser
            x_pos = np.arange(1, sample_count + 1) * sample_spacing
            x_positions[appraiser] = x_pos
            
            # Plot average values with connecting lines
            line, = plt.plot(x_pos, averages, 'o-', color=color, label=display_name, linewidth=1.5, markersize=6)
            legend_handles.append(line)
            
            # Track max and min values for y-axis scaling
            max_value = max(max_value, max(averages) if averages else 0)
            min_value = min(min_value, min(averages) if averages else float('inf'))
        
        # Plot control limits
        plt.axhline(y=ucl_x, color='blue', linestyle='-', label='UCL')
        plt.axhline(y=x_doublebar, color='purple', linestyle='--', label='Mean')
        plt.axhline(y=lcl_x, color='orange', linestyle='-', label='LCL')
        
        # Set axis labels and title
        plt.xlabel('Trial Number')
        plt.ylabel('Average')
        plt.title('Averages Chart')
        
        # Set x-axis ticks at sample positions
        sample_positions = np.arange(1, sample_count + 1) * sample_spacing
        plt.xticks(sample_positions, [str(i) for i in range(1, sample_count + 1)])
        
        # Add more detailed x-axis labels showing appraiser groups
        appraiser_label_pos = []
        appraiser_label_names = []
        
        for i, appraiser_name in enumerate(appraisers[:self.appraisers]):
            if appraiser_name in appraiser_data:
                mid_point = (sample_positions[0] + sample_positions[-1]) / 2 + (i - self.appraisers/2 + 0.5) * sample_spacing * 1.2
                appraiser_label_pos.append(mid_point)
                display_name = appraiser_data[appraiser_name]['display_name']
                appraiser_label_names.append(display_name)
        
        # Add a second x-axis for appraiser labels
        ax2 = plt.twiny()
        ax2.set_xticks(appraiser_label_pos)
        ax2.set_xticklabels(appraiser_label_names)
        ax2.set_xlim(plt.gca().get_xlim())
        
        # Set y-axis limits with some padding
        yrange = max_value - min_value
        plt.ylim(min(min_value - 0.1 * yrange, lcl_x - 0.1 * yrange),
                 max(max_value + 0.1 * yrange, ucl_x + 0.1 * yrange))
        
        # Add grid for better readability
        plt.grid(True, linestyle='--', alpha=0.7)
        
        # Add legend
        plt.legend(handles=legend_handles + [
            plt.Line2D([0], [0], color='blue', lw=2, linestyle='-', label='UCL'),
            plt.Line2D([0], [0], color='purple', lw=2, linestyle='--', label='Mean'),
            plt.Line2D([0], [0], color='orange', lw=2, linestyle='-', label='LCL')
        ])
        
        # Add Plot Area text in the bottom right
        plt.figtext(0.93, 0.05, 'Plot Area', ha='right')
        
        # Save the chart to binary field
        buffer = io.BytesIO()
        plt.savefig(buffer, format='png', dpi=100)
        buffer.seek(0)
        self.average_chart = base64.b64encode(buffer.getvalue())
        plt.close()
    


class GaugeRNRAppraiser(models.Model):
    _name = 'gauge.rnr.appraiser'
    _description = 'Gauge R&R Appraiser Data'
    _inherit = "translation.mixin"


    gauge_id = fields.Many2one('gauge.rnr', string='Gauge R&R', required=True, ondelete='cascade')
    appraiser = fields.Char(string='Appraiser',translate=True)
    trial = fields.Integer(string='Trial')
    sample_1 = fields.Float(string='Sample 1', digits=(12, 2), default=0.00)
    sample_2 = fields.Float(string='Sample 2', digits=(12, 2), default=0.00)
    sample_3 = fields.Float(string='Sample 3', digits=(12, 2), default=0.00)
    sample_4 = fields.Float(string='Sample 4', digits=(12, 2), default=0.00)
    sample_5 = fields.Float(string='Sample 5', digits=(12, 2), default=0.00)
    sample_6 = fields.Float(string='Sample 6', digits=(12, 2), default=0.00)
    sample_7 = fields.Float(string='Sample 7', digits=(12, 2), default=0.00)
    sample_8 = fields.Float(string='Sample 8', digits=(12, 2), default=0.00)
    sample_9 = fields.Float(string='Sample 9', digits=(12, 2), default=0.00)
    sample_10 = fields.Float(string='Sample 10', digits=(12, 2), default=0.00)
    average = fields.Float(string='Average', compute='_compute_average', digits=(12, 2))
    range = fields.Float(string='Range', compute='_compute_range', digits=(12, 2))

    @api.depends('sample_1', 'sample_2', 'sample_3', 'sample_4', 'sample_5', 'sample_6', 'sample_7', 'sample_8',
                 'sample_9', 'sample_10')
    def _compute_average(self):
        for record in self:
            samples = [
                record.sample_1, record.sample_2, record.sample_3, record.sample_4, record.sample_5,
                record.sample_6, record.sample_7, record.sample_8, record.sample_9, record.sample_10
            ]
            valid_samples = [s for s in samples if s != 0.00]
            record.average = sum(valid_samples) / len(valid_samples) if valid_samples else 0.00

    @api.depends('sample_1', 'sample_2', 'sample_3', 'sample_4', 'sample_5', 'sample_6', 'sample_7', 'sample_8',
                 'sample_9', 'sample_10')
    def _compute_range(self):
        for record in self:
            samples = [
                record.sample_1, record.sample_2, record.sample_3, record.sample_4, record.sample_5,
                record.sample_6, record.sample_7, record.sample_8, record.sample_9, record.sample_10
            ]
            valid_samples = [s for s in samples if s != 0.00]
            record.range = max(valid_samples) - min(valid_samples) if valid_samples else 0.00
            
    # Add this write method here
    def write(self, vals):
        try:
            _logger.info(f"Writing values to appraiser record: {vals}")
            result = super(GaugeRNRAppraiser, self).write(vals)
            # After successful write, update related statistics
            if self.gauge_id and any(key.startswith('sample_') for key in vals):
                self._compute_average()
                self._compute_range()
            return result
        except Exception as e:
            _logger.error(f"Error saving appraiser data: {e}")
            raise

class GaugeRNRAppraiserSample(models.Model):
    """Stores the average and range for each sample across all trials for an appraiser"""
    _name = 'gauge.rnr.appraiser.sample'
    _description = 'Gauge R&R Appraiser Sample Statistics'
    _inherit = "translation.mixin"

    gauge_id = fields.Many2one('gauge.rnr', string='Gauge R&R', required=True, ondelete='cascade')
    appraiser = fields.Char(string='Appraiser', required=True,translate=True)
    sample_number = fields.Integer(string='Sample Number', required=True)
    average = fields.Float(string='Average', digits=(12, 2))
    range = fields.Float(string='Range', digits=(12, 2))


class GaugeRNRAppraiserTrial(models.Model):
    """Stores the average and range for each trial across all samples for an appraiser"""
    _name = 'gauge.rnr.appraiser.trial'
    _description = 'Gauge R&R Appraiser Trial Statistics'
    _inherit = "translation.mixin"

    gauge_id = fields.Many2one('gauge.rnr', string='Gauge R&R', required=True, ondelete='cascade')
    appraiser = fields.Char(string='Appraiser', required=True,translate=True)
    trial_number = fields.Integer(string='Trial Number', required=True)
    average = fields.Float(string='Average', digits=(12, 2))
    range = fields.Float(string='Range', digits=(12, 2))



class GaugeRNRController(http.Controller):
    @http.route('/gauge_rnr/chart/<string:chart_type>/<int:gauge_id>', type='http', auth='user')
    def display_chart(self, chart_type, gauge_id, **kwargs):
        gauge = request.env['gauge.rnr'].browse(gauge_id)
        
        if chart_type == 'bias' and gauge.bias_chart:
            return request.make_response(
                base64.b64decode(gauge.bias_chart),
                headers=[('Content-Type', 'image/png')]
            )
        elif chart_type == 'consistency' and gauge.consistency_chart:
            return request.make_response(
                base64.b64decode(gauge.consistency_chart),
                headers=[('Content-Type', 'image/png')]
            )
        elif chart_type == 'average' and gauge.average_chart:
            return request.make_response(
                base64.b64decode(gauge.average_chart),
                headers=[('Content-Type', 'image/png')]
            )
        elif chart_type == 'range' and gauge.range_chart:
            return request.make_response(
                base64.b64decode(gauge.range_chart),
                headers=[('Content-Type', 'image/png')]
            )
        
        return request.not_found()
class GaugeRnRConstants(models.TransientModel):
    _name = 'gauge.rnr.constants'
    _description = 'Gauge R&R Statistical Constants'

    
